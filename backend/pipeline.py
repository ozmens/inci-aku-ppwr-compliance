"""Bilingual + photo annex pipeline helpers (preview / catalog).

Writes only into candidates/ when exporting drafts. Never touches frozen deliveries.
"""

from __future__ import annotations

import csv
import sys
from functools import lru_cache
from pathlib import Path

from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook

from masters import get_bom

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from bilingual_format import translate_component, translate_product  # noqa: E402
from photo_annex import PHOTO_ROOT as DEFAULT_PHOTO_ROOT  # noqa: E402
from photo_annex import clear_mapping_cache, resolve_photos  # noqa: E402

CANDIDATES = ROOT / "candidates"
FROZEN = (
    "01_STARTER_INDIVIDUAL_DELIVERY_REV00",
    "02_INDUSTRIAL_DELIVERY_REV00",
    "03_CONTAINER_DELIVERY_REV00",
    "04_COMPONENT_SPARE_DELIVERY_REV00",
)

SCOPE_MAP = {
    "starter": "STARTER",
    "industrial": "INDUSTRIAL",
    "container": "CONTAINER",
    "component": "COMPONENT",
}


def photo_root() -> Path:
    # Prefer local assets; fall back to PIMS if empty
    local = ROOT / "assets" / "representative_component_photos"
    if local.exists() and any(local.rglob("*.jpg")):
        return local
    pims = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS\assets\representative_component_photos")
    if pims.exists():
        return pims
    return local if local.exists() else DEFAULT_PHOTO_ROOT


def _assert_not_frozen(path: Path) -> None:
    s = str(path.resolve()).replace("/", "\\").upper()
    for m in FROZEN:
        if m.upper() in s:
            raise HTTPException(403, f"Refusing write into frozen delivery: {m}")


@lru_cache(maxsize=1)
def _mapping_rows() -> tuple[dict, ...]:
    root = photo_root()
    mapping = root / "PHOTO_MAPPING_INDEX.csv"
    if not mapping.exists():
        return tuple()
    with mapping.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(2048)
        f.seek(0)
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delim)
        rows = []
        for row in reader:
            rows.append({(k or "").strip(): (v or "").strip() for k, v in row.items()})
    return tuple(rows)


def catalog(scope: str | None = None) -> dict:
    root = photo_root()
    rows = []
    for row in _mapping_rows():
        sc = (row.get("Scope") or "").strip().upper()
        if scope:
            want = SCOPE_MAP.get(scope.lower(), scope.upper())
            if sc and sc not in (want, "COMMON", "ALL"):
                continue
        rel = (row.get("Photo File") or row.get("Photo") or "").replace("\\", "/")
        path = root / rel.replace("/", "\\") if rel else None
        rows.append(
            {
                "scope": sc,
                "source": row.get("Source Component Description")
                or row.get("Source Component")
                or row.get("Component Description")
                or "",
                "photo_file": rel,
                "note": row.get("Note") or "",
                "exists": bool(path and path.exists()),
            }
        )
    images = sorted(
        {
            str(p.relative_to(root)).replace("\\", "/")
            for p in root.rglob("*")
            if p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
        }
    )
    return {
        "photo_root": str(root),
        "mapping_count": len(rows),
        "image_count": len(images),
        "mappings": rows[:200],
        "images_sample": images[:40],
    }


def bilingual_preview_texts(texts: list[str], kind: str = "component") -> dict:
    out = []
    for t in texts:
        tr = (t or "").strip()
        if kind == "product":
            en = translate_product(tr)
        else:
            en = translate_component(tr)
        out.append({"tr": tr, "en": en, "display": f"{tr}\n{en}" if en else tr})
    return {"kind": kind, "items": out}


def bilingual_set_preview(set_code: str, scope: str = "starter") -> dict:
    bom = get_bom("starter" if scope == "starter" else "starter", set_code)
    if scope != "starter":
        # BOM currently starter-backed; still allow photo scope switch
        pass
    lines = []
    for line in bom.get("lines") or []:
        tr = str(line.get("description") or "")
        # if already bilingual take TR half
        if "\n" in tr:
            tr = tr.split("\n", 1)[0].strip()
        en = translate_component(tr)
        lines.append(
            {
                "component_code": line.get("component_code"),
                "qty": line.get("qty"),
                "uom": line.get("uom"),
                "tr": tr,
                "en": en,
                "unit_weight": line.get("unit_weight"),
                "line_weight": line.get("line_weight"),
            }
        )
    return {
        "set_code": set_code,
        "scope": scope,
        "meta": bom.get("meta") or {},
        "lines": lines,
        "rule": "Türkçe düz · İngilizce italik",
    }


def resolve_set_photos(set_code: str, scope: str = "starter") -> dict:
    # Ensure photo_annex uses our root
    import photo_annex as pa

    pa.PHOTO_ROOT = photo_root()
    pa.MAPPING = pa.PHOTO_ROOT / "PHOTO_MAPPING_INDEX.csv"
    clear_mapping_cache()

    bom = get_bom("starter", set_code)
    bom_lines = []
    for line in bom.get("lines") or []:
        tr = str(line.get("description") or "")
        if "\n" in tr:
            tr = tr.split("\n", 1)[0].strip()
        bom_lines.append(
            {
                "component_code": line.get("component_code"),
                "description": tr,
                "name_tr": tr,
                "name_en": translate_component(tr),
            }
        )
    scope_key = SCOPE_MAP.get(scope.lower(), scope.upper())
    hits = resolve_photos(scope=scope_key, bom_lines=bom_lines)
    root = photo_root()
    photos = []
    for h in hits:
        try:
            rel = str(h.path.resolve().relative_to(root.resolve())).replace("\\", "/")
        except ValueError:
            rel = h.path.name
        photos.append(
            {
                "component_code": h.component_code,
                "name_tr": h.name_tr,
                "name_en": h.name_en,
                "note": h.note,
                "rel": rel,
                "exists": h.path.exists(),
            }
        )
    matched_codes = {p["component_code"] for p in photos}
    missing = [
        {
            "component_code": bl["component_code"],
            "description": bl["description"],
        }
        for bl in bom_lines
        if bl["component_code"] not in matched_codes
    ]
    return {
        "set_code": set_code,
        "scope": scope,
        "photo_root": str(root),
        "matched": len(photos),
        "missing": missing,
        "photos": photos,
        "bom_line_count": len(bom_lines),
    }


def serve_photo(rel: str) -> FileResponse:
    root = photo_root().resolve()
    clean = (rel or "").replace("\\", "/").lstrip("/")
    if ".." in clean.split("/"):
        raise HTTPException(400, "Invalid path")
    path = (root / clean.replace("/", "\\")).resolve()
    try:
        path.relative_to(root)
    except ValueError as e:
        raise HTTPException(400, "Path outside photo root") from e
    if not path.exists() or path.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise HTTPException(404, "Photo missing")
    return FileResponse(path, filename=path.name)


def export_preview_xlsx(set_code: str, scope: str = "starter") -> dict:
    CANDIDATES.mkdir(parents=True, exist_ok=True)
    bi = bilingual_set_preview(set_code, scope)
    ph = resolve_set_photos(set_code, scope)
    path = CANDIDATES / f"PIPELINE_PREVIEW_{set_code.replace('/', '_')}.xlsx"
    _assert_not_frozen(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "BILINGUAL_BOM"
    ws.append(["Component Code", "TR", "EN", "Qty", "UOM", "Line Weight"])
    for line in bi["lines"]:
        ws.append(
            [
                line.get("component_code"),
                line.get("tr"),
                line.get("en"),
                line.get("qty"),
                line.get("uom"),
                line.get("line_weight"),
            ]
        )
    ws2 = wb.create_sheet("PHOTO_HITS")
    ws2.append(["Component Code", "Name TR", "Name EN", "Photo Rel", "Note"])
    for p in ph["photos"]:
        ws2.append([p["component_code"], p["name_tr"], p["name_en"], p["rel"], p["note"]])
    ws3 = wb.create_sheet("PHOTO_MISSING")
    ws3.append(["Component Code", "Description"])
    for m in ph["missing"]:
        ws3.append([m["component_code"], m["description"]])
    wb.save(path)
    return {
        "path": str(path),
        "set_code": set_code,
        "bilingual_lines": len(bi["lines"]),
        "photos_matched": ph["matched"],
        "photos_missing": len(ph["missing"]),
    }


def status() -> dict:
    root = photo_root()
    imgs = list(root.rglob("*.jpg")) + list(root.rglob("*.png")) if root.exists() else []
    return {
        "photo_root": str(root),
        "photo_root_exists": root.exists(),
        "image_count": len(imgs),
        "mapping_count": len(_mapping_rows()),
        "bilingual_rule": "Türkçe düz · İngilizce italik",
        "write_policy": "candidates_only",
        "candidates_root": str(CANDIDATES),
    }
