"""Candidate Document Engine — SEARCH / DOCUMENT_CENTER under candidates only.

Layout:
  candidates/PPWR_CANDIDATE_DELIVERY_REV00/
    00_CONTROL/<engine>.xlsx
    00_AC_DOCUMENT_ENGINE.cmd
    01_PRODUCTS/<key>/   (junction → ../packs/<key>)

Links from 00_CONTROL are relative: ../01_PRODUCTS/<key>/<file>
Labels: OPEN WORD / OPEN PDF only. Never writes frozen Rev.00 deliveries.
"""

from __future__ import annotations

import json
import os
import subprocess
from datetime import datetime, timezone
from pathlib import Path

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from packs import CANDIDATES, PACKS_ROOT, STEMS, _assert_candidates_only, _file_inventory

DELIVERY = CANDIDATES / "PPWR_CANDIDATE_DELIVERY_REV00"
CONTROL = DELIVERY / "00_CONTROL"
PRODUCTS = DELIVERY / "01_PRODUCTS"
ENGINE_NAME = "INCI_PPWR_CANDIDATE_ENGINE_Rev00.xlsx"
LAUNCHER = DELIVERY / "00_AC_DOCUMENT_ENGINE.cmd"

NAVY = "0E2A47"
GREEN = "1F7A4C"
SOFT = "F3F6F9"
LINE = "D0D7DE"
WHITE = "FFFFFF"
LINK_BLUE = "0563C1"
SLATE = "1C2430"
FONT = "Tahoma"

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_SOFT = PatternFill("solid", fgColor=SOFT)
FONT_TITLE = Font(name=FONT, size=18, bold=True, color=NAVY)
FONT_SUB = Font(name=FONT, size=10, color="1F4E79")
FONT_BODY = Font(name=FONT, size=10, color=SLATE)
FONT_HDR = Font(name=FONT, size=10, bold=True, color=WHITE)
FONT_LINK = Font(name=FONT, size=10, color=LINK_BLUE, underline="single")
FONT_PASS = Font(name=FONT, size=10, bold=True, color=WHITE)
THIN = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STEM_COLS = [
    ("TF WORD", "01_Technical_File.docx"),
    ("TF PDF", "01_Technical_File.pdf"),
    ("DoC WORD", "02_EU_DoC.docx"),
    ("DoC PDF", "02_EU_DoC.pdf"),
    ("Label WORD", "03_Label.docx"),
    ("Label PDF", "03_Label.pdf"),
    ("STM WORD", "04_Shipment_Statement.docx"),
    ("STM PDF", "04_Shipment_Statement.pdf"),
]


def _rel_link(cell, rel_path: str, label: str) -> None:
    """Relative hyperlink from 00_CONTROL (Excel opens relative to workbook)."""
    safe = rel_path.replace('"', '""').replace("/", "\\")
    cell.value = f'=HYPERLINK("{safe}","{label}")'
    cell.font = FONT_LINK
    cell.alignment = CENTER


def _ensure_junction(key: str) -> Path:
    """Point 01_PRODUCTS/<key> at candidates/packs/<key> via directory junction."""
    src = (PACKS_ROOT / key).resolve()
    if not src.is_dir():
        raise HTTPException(404, f"Pack missing: {key}")
    PRODUCTS.mkdir(parents=True, exist_ok=True)
    dest = PRODUCTS / key
    if dest.exists():
        # already correct junction/link?
        try:
            if dest.resolve() == src:
                return dest
        except OSError:
            pass
        # remove existing junction (rmdir works on junctions without deleting target)
        r = subprocess.run(["cmd", "/c", "rmdir", str(dest)], capture_output=True, text=True)
        if dest.exists():
            raise HTTPException(500, f"Cannot refresh junction for {key}: {r.stderr or r.stdout}")

    r = subprocess.run(
        ["cmd", "/c", "mklink", "/J", str(dest), str(src)],
        capture_output=True,
        text=True,
    )
    if not dest.exists():
        raise HTTPException(500, f"Junction failed for {key}: {(r.stderr or r.stdout or '').strip()}")
    return dest


def _collect_records() -> list[dict]:
    records = []
    if not PACKS_ROOT.exists():
        return records
    for folder in sorted(p for p in PACKS_ROOT.iterdir() if p.is_dir()):
        meta = {}
        mp = folder / "PACK_META.json"
        if mp.exists():
            meta = json.loads(mp.read_text(encoding="utf-8"))
        inv = _file_inventory(folder)
        complete = all(f["exists"] for f in inv)
        records.append(
            {
                "key": folder.name,
                "label": str(meta.get("description") or folder.name),
                "set_code": str(meta.get("set_code") or ""),
                "complete": complete,
                "files": inv,
            }
        )
    return records


def verify_candidate_links(keys: list[str]) -> dict:
    missing = []
    checked = 0
    for key in keys:
        base = PRODUCTS / key
        for _, fname in STEM_COLS:
            checked += 1
            target = base / fname
            if not target.exists() or target.stat().st_size < 200:
                missing.append(f"{key}/{fname}")
    return {
        "checked": checked,
        "missing": missing,
        "pass": len(missing) == 0 and checked > 0,
    }


def engine_status() -> dict:
    engine = CONTROL / ENGINE_NAME
    records = _collect_records()
    return {
        "delivery_root": str(DELIVERY),
        "control": str(CONTROL),
        "engine": str(engine) if engine.exists() else None,
        "engine_exists": engine.exists(),
        "launcher": str(LAUNCHER) if LAUNCHER.exists() else None,
        "launcher_exists": LAUNCHER.exists(),
        "packs": len(records),
        "complete_packs": sum(1 for r in records if r["complete"]),
        "write_policy": "candidates_only",
        "link_rule": "relative from 00_CONTROL -> ../01_PRODUCTS/<key>/",
    }


def rebuild_candidate_engine(*, require_complete: bool = True) -> dict:
    """Build/refresh candidate engine. Verify link targets before PASS."""
    _assert_candidates_only(DELIVERY)
    DELIVERY.mkdir(parents=True, exist_ok=True)
    CONTROL.mkdir(parents=True, exist_ok=True)
    PRODUCTS.mkdir(parents=True, exist_ok=True)

    records = _collect_records()
    if require_complete:
        incomplete = [r["key"] for r in records if not r["complete"]]
        if incomplete:
            raise HTTPException(
                400,
                f"Incomplete packs (need 4 DOCX+4 PDF): {', '.join(incomplete[:8])}",
            )
    if not records:
        raise HTTPException(400, "No candidate packs to index — build a pack first")

    # junctions
    for rec in records:
        _ensure_junction(rec["key"])

    keys = [r["key"] for r in records]
    verify = verify_candidate_links(keys)
    if not verify["pass"]:
        raise HTTPException(
            400,
            f"Link verify FAIL — missing {len(verify['missing'])}: "
            + ", ".join(verify["missing"][:6]),
        )

    engine_path = CONTROL / ENGINE_NAME
    _assert_candidates_only(engine_path)
    _write_engine_xlsx(engine_path, records, verify)
    _write_launcher()

    return {
        "engine": str(engine_path),
        "launcher": str(LAUNCHER),
        "delivery_root": str(DELIVERY),
        "records": len(records),
        "verify": verify,
        "link_rule": "relative ../01_PRODUCTS/<key>/",
        "qa": "PASS",
        "built_at": datetime.now(timezone.utc).isoformat(),
    }


def _write_engine_xlsx(path: Path, records: list[dict], verify: dict) -> None:
    n = len(records)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb = Workbook()

    home = wb.active
    home.title = "00_HOME"
    home.sheet_view.showGridLines = False
    home["B2"] = "PPWR CANDIDATE DOCUMENT ENGINE"
    home["B2"].font = FONT_TITLE
    home["B3"] = f"Candidates only  •  Rev.00 draft  •  Publish 11.08.2026  •  {stamp}"
    home["B3"].font = FONT_SUB
    home["B5"] = "QA STATUS"
    home["C5"] = "PASS" if verify["pass"] else "FAIL"
    home["C5"].fill = FILL_GREEN if verify["pass"] else PatternFill("solid", fgColor="A12622")
    home["C5"].font = FONT_PASS
    home["C5"].alignment = CENTER
    home["B6"] = "RECORDS"
    home["C6"] = n
    home["B7"] = "DOCX+PDF"
    home["C7"] = n * 8
    home["B8"] = "LINK RULE"
    home["C8"] = "RELATIVE from 00_CONTROL"
    home["B9"] = "OPEN RULE"
    home["C9"] = "00_AC_DOCUMENT_ENGINE.cmd ONLY"
    home["B11"] = (
        "• Frozen Rev.00 deliveries are NOT modified\n"
        "• SEARCH → Key → OPEN WORD / OPEN PDF\n"
        "• DOCUMENT CENTER → full register\n"
        "• Links: ../01_PRODUCTS/<key>/<file> (do not move engine)\n"
        "• Every pack: DOCX + PDF (LibreOffice PDF)"
    )
    home["B11"].alignment = Alignment(wrap_text=True, vertical="top")
    home.merge_cells("B11:G16")
    home["B11"].fill = FILL_SOFT
    home.column_dimensions["B"].width = 22
    home.column_dimensions["C"].width = 36

    # SEARCH
    search = wb.create_sheet("SEARCH")
    search["A1"] = "GLOBAL KEY SEARCH"
    search["A1"].font = FONT_TITLE
    search["A3"] = "Key"
    search["B3"] = ""
    search["B3"].fill = FILL_SOFT
    search["B3"].border = THIN
    search["A5"] = "Result Key"
    search["A6"] = "Description"
    search["A7"] = "Set"
    search["A8"] = "Status"
    search["B5"] = '=IF($B$3="","",IFERROR(INDEX(SEARCH_DATA!A:A,MATCH($B$3,SEARCH_DATA!A:A,0)),"NOT FOUND"))'
    search["B6"] = '=IF(OR($B$5="",$B$5="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!B:B,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    search["B7"] = '=IF(OR($B$5="",$B$5="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!C:C,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    search["B8"] = '=IF(OR($B$5="",$B$5="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!D:D,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    search["A10"] = "Document"
    search["B10"] = "OPEN WORD"
    search["C10"] = "OPEN PDF"
    for cell in (search["A10"], search["B10"], search["C10"]):
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
    # Relative open formulas — path built from key in B5
    for i, (label, stem) in enumerate(
        [
            ("Technical File", "01_Technical_File"),
            ("EU DoC", "02_EU_DoC"),
            ("Label", "03_Label"),
            ("Shipment Statement", "04_Shipment_Statement"),
        ]
    ):
        r = 11 + i
        search.cell(r, 1, label).font = FONT_BODY
        # HYPERLINK relative
        search.cell(
            r,
            2,
            f'=IF(OR($B$5="",$B$5="NOT FOUND"),"",'
            f'HYPERLINK("../01_PRODUCTS/"&$B$5&"\\{stem}.docx","OPEN WORD"))',
        )
        search.cell(r, 2).font = FONT_LINK
        search.cell(r, 2).alignment = CENTER
        search.cell(
            r,
            3,
            f'=IF(OR($B$5="",$B$5="NOT FOUND"),"",'
            f'HYPERLINK("../01_PRODUCTS/"&$B$5&"\\{stem}.pdf","OPEN PDF"))',
        )
        search.cell(r, 3).font = FONT_LINK
        search.cell(r, 3).alignment = CENTER
    search.column_dimensions["A"].width = 22
    search.column_dimensions["B"].width = 48
    search.column_dimensions["C"].width = 14

    # SEARCH_DATA
    sdata = wb.create_sheet("SEARCH_DATA")
    for i, h in enumerate(["Key", "Description", "Set Code", "Status", "Revision"], 1):
        c = sdata.cell(1, i, h)
        c.fill = FILL_NAVY
        c.font = FONT_HDR
        c.alignment = CENTER
        c.border = THIN

    # DOCUMENT_CENTER
    dc = wb.create_sheet("DOCUMENT_CENTER")
    dc["A1"] = "DOCUMENT CENTER — CANDIDATE REGISTER"
    dc["A1"].font = FONT_TITLE
    dc.merge_cells("A1:L1")
    dc["A2"] = (
        f"Relative links from 00_CONTROL -> ../01_PRODUCTS/<key>/  ·  "
        f"Records: {n}  ·  Open via 00_AC_DOCUMENT_ENGINE.cmd"
    )
    dc["A2"].font = FONT_BODY
    dc.merge_cells("A2:L2")
    headers = ["Key", "Description", "Set Code", "Status", "Revision"] + [h for h, _ in STEM_COLS]
    for i, h in enumerate(headers, 1):
        cell = dc.cell(4, i, h)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    for i, rec in enumerate(records, 1):
        key = rec["key"]
        r = 4 + i
        dc.cell(r, 1, key).font = FONT_BODY
        dc.cell(r, 2, rec["label"]).font = FONT_BODY
        dc.cell(r, 3, rec.get("set_code") or "").font = FONT_BODY
        st = dc.cell(r, 4, "ISSUED")
        st.fill = FILL_GREEN
        st.font = FONT_PASS
        st.alignment = CENTER
        dc.cell(r, 5, "R00").alignment = CENTER
        for c in range(1, 6):
            dc.cell(r, c).border = THIN
        for j, (_, fname) in enumerate(STEM_COLS):
            label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
            cell = dc.cell(r, 6 + j)
            _rel_link(cell, f"../01_PRODUCTS/{key}/{fname}", label)
            cell.border = THIN
        # SEARCH_DATA row
        for c, val in enumerate(
            [key, rec["label"], rec.get("set_code") or "", "ISSUED Rev.00", "R00"], 1
        ):
            cell = sdata.cell(i + 1, c, val)
            cell.font = FONT_BODY
            cell.border = THIN

    dc.column_dimensions["A"].width = 16
    dc.column_dimensions["B"].width = 40
    dc.column_dimensions["C"].width = 16
    for c in range(4, 14):
        dc.column_dimensions[get_column_letter(c)].width = 12
    dc.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + n}"
    dc.freeze_panes = "A5"
    sdata.column_dimensions["A"].width = 16
    sdata.column_dimensions["B"].width = 40
    sdata.auto_filter.ref = f"A1:E{n + 1}"

    # QA
    qa = wb.create_sheet("QA")
    qa["A1"] = "QA — CANDIDATE ENGINE"
    qa["A1"].font = FONT_TITLE
    qa["A3"] = "Link targets checked"
    qa["B3"] = verify["checked"]
    qa["A4"] = "Missing"
    qa["B4"] = len(verify["missing"])
    qa["A5"] = "Result"
    qa["B5"] = "PASS" if verify["pass"] else "FAIL"
    qa["B5"].fill = FILL_GREEN if verify["pass"] else PatternFill("solid", fgColor="A12622")
    qa["B5"].font = FONT_PASS

    wb.save(path)


def _write_launcher() -> None:
    _assert_candidates_only(LAUNCHER)
    engine = CONTROL / ENGINE_NAME
    LAUNCHER.write_text(
        "@echo off\r\n"
        "cd /d \"%~dp0\"\r\n"
        "echo ============================================================\r\n"
        "echo  INCI AKU PPWR — CANDIDATE Document Engine\r\n"
        "echo  Frozen Rev.00 deliveries are NOT modified.\r\n"
        "echo  Open ONLY via this CMD (engine lives in 00_CONTROL).\r\n"
        "echo ============================================================\r\n"
        f'if not exist "00_CONTROL\\{ENGINE_NAME}" (\r\n'
        f"  echo ERROR: engine missing: 00_CONTROL\\{ENGINE_NAME}\r\n"
        "  pause\r\n"
        "  exit /b 1\r\n"
        ")\r\n"
        f'start \"\" \"00_CONTROL\\{ENGINE_NAME}\"\r\n',
        encoding="utf-8",
    )


def open_candidate_engine() -> dict:
    if not LAUNCHER.exists():
        raise HTTPException(404, "Launcher missing — rebuild engine first")
    subprocess.Popen(["cmd", "/c", str(LAUNCHER)], cwd=str(DELIVERY))
    return {"launched": str(LAUNCHER)}


def open_candidate_delivery_folder() -> dict:
    DELIVERY.mkdir(parents=True, exist_ok=True)
    os.startfile(str(DELIVERY))
    return {"opened": str(DELIVERY)}
