"""Optional Workspace Document Engine export.

Source of truth remains workspace/products.
Export layout (software-only, never frozen deliveries):
  workspace/PPWR_WORKSPACE_ENGINE/
    00_CONTROL/INCI_PPWR_WORKSPACE_ENGINE.xlsx
    00_AC_DOCUMENT_ENGINE.cmd
    01_PRODUCTS/<key>/  → junction to products/<key>/revisions/<current>/

Relative links from 00_CONTROL: ../01_PRODUCTS/<key>/<file>
Labels: OPEN WORD / OPEN PDF only.
"""

from __future__ import annotations

import os
import subprocess
from datetime import datetime, timezone
from pathlib import Path

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from workspace_store import (
    WORKSPACE,
    _assert_workspace,
    _rev_dir,
    list_products,
    log_activity,
)

DELIVERY = WORKSPACE / "PPWR_WORKSPACE_ENGINE"
CONTROL = DELIVERY / "00_CONTROL"
PRODUCTS_LINK = DELIVERY / "01_PRODUCTS"
ENGINE_NAME = "INCI_PPWR_WORKSPACE_ENGINE.xlsx"
LAUNCHER = DELIVERY / "00_AC_DOCUMENT_ENGINE.cmd"

NAVY = "0E2A47"
GREEN = "1F7A4C"
SOFT = "F3F6F9"
LINE = "D0D7DE"
WHITE = "FFFFFF"
LINK_BLUE = "0563C1"
SLATE = "1C2430"
FONT = "Tahoma"

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_SOFT = PatternFill("solid", fgColor=SOFT)
FONT_TITLE = Font(name=FONT, size=18, bold=True, color=NAVY)
FONT_SUB = Font(name=FONT, size=10, color="1F4E79")
FONT_BODY = Font(name=FONT, size=10, color=SLATE)
FONT_HDR = Font(name=FONT, size=10, bold=True, color=WHITE)
FONT_LINK = Font(name=FONT, size=10, color=LINK_BLUE, underline="single")
FONT_PASS = Font(name=FONT, size=10, bold=True, color=WHITE)
THIN = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STEM_COLS = [
    ("TF WORD", "01_Technical_File.docx"),
    ("TF PDF", "01_Technical_File.pdf"),
    ("DoC WORD", "02_EU_DoC.docx"),
    ("DoC PDF", "02_EU_DoC.pdf"),
    ("Label WORD", "03_Label.docx"),
    ("Label PDF", "03_Label.pdf"),
    ("STM WORD", "04_Shipment_Statement.docx"),
    ("STM PDF", "04_Shipment_Statement.pdf"),
]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _rel_link(cell, rel_path: str, label: str) -> None:
    safe = rel_path.replace('"', '""').replace("/", "\\")
    cell.value = f'=HYPERLINK("{safe}","{label}")'
    cell.font = FONT_LINK
    cell.alignment = CENTER


def _ensure_junction(key: str, rev: str) -> Path:
    src = _rev_dir(key, rev).resolve()
    if not src.is_dir():
        raise HTTPException(404, f"Revision folder missing: {key} {rev}")
    PRODUCTS_LINK.mkdir(parents=True, exist_ok=True)
    dest = PRODUCTS_LINK / key
    if dest.exists():
        try:
            if dest.resolve() == src:
                return dest
        except OSError:
            pass
        subprocess.run(["cmd", "/c", "rmdir", str(dest)], capture_output=True, text=True)
        if dest.exists():
            raise HTTPException(500, f"Cannot refresh junction for {key}")
    r = subprocess.run(
        ["cmd", "/c", "mklink", "/J", str(dest), str(src)],
        capture_output=True,
        text=True,
    )
    if not dest.exists():
        raise HTTPException(500, f"Junction failed for {key}: {(r.stderr or r.stdout or '').strip()}")
    return dest


def verify_links(keys: list[tuple[str, str]]) -> dict:
    missing = []
    checked = 0
    for key, _rev in keys:
        base = PRODUCTS_LINK / key
        for _, fname in STEM_COLS:
            checked += 1
            target = base / fname
            if not target.exists() or target.stat().st_size < 200:
                missing.append(f"{key}/{fname}")
    return {"checked": checked, "missing": missing, "pass": len(missing) == 0 and checked > 0}


def workspace_engine_status() -> dict:
    engine = CONTROL / ENGINE_NAME
    products = [p for p in list_products() if p.get("complete")]
    return {
        "delivery_root": str(DELIVERY),
        "engine": str(engine) if engine.exists() else None,
        "engine_exists": engine.exists(),
        "launcher": str(LAUNCHER) if LAUNCHER.exists() else None,
        "launcher_exists": LAUNCHER.exists(),
        "complete_products": len(products),
        "link_rule": "relative from 00_CONTROL -> ../01_PRODUCTS/<key>/",
        "role": "optional_export",
    }


def rebuild_workspace_engine(*, issued_only: bool = False) -> dict:
    """Build optional Excel engine from complete workspace packs."""
    _assert_workspace(WORKSPACE)
    DELIVERY.mkdir(parents=True, exist_ok=True)
    CONTROL.mkdir(parents=True, exist_ok=True)
    PRODUCTS_LINK.mkdir(parents=True, exist_ok=True)

    records = []
    for p in list_products():
        if not p.get("complete"):
            continue
        if issued_only and p.get("status") != "ISSUED":
            continue
        rev = p.get("current_revision")
        if not rev:
            continue
        records.append(
            {
                "key": p["product_code"],
                "label": p.get("description") or p["product_code"],
                "set_code": p.get("set_code") or "",
                "revision": rev,
                "status": p.get("status") or "",
            }
        )
    if not records:
        raise HTTPException(400, "No complete workspace packs to export — create/complete packs first")

    for rec in records:
        _ensure_junction(rec["key"], rec["revision"])

    verify = verify_links([(r["key"], r["revision"]) for r in records])
    if not verify["pass"]:
        raise HTTPException(
            400,
            f"Link verify FAIL — missing {len(verify['missing'])}: "
            + ", ".join(verify["missing"][:6]),
        )

    engine_path = CONTROL / ENGINE_NAME
    _assert_workspace(engine_path.parent)
    _write_xlsx(engine_path, records, verify)
    _write_launcher()

    log_activity("workspace_engine_export", records=len(records))
    return {
        "engine": str(engine_path),
        "launcher": str(LAUNCHER),
        "delivery_root": str(DELIVERY),
        "records": len(records),
        "verify": verify,
        "qa": "PASS",
        "built_at": _now(),
        "link_rule": "relative ../01_PRODUCTS/<key>/",
    }


def _write_xlsx(path: Path, records: list[dict], verify: dict) -> None:
    n = len(records)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb = Workbook()

    home = wb.active
    home.title = "00_HOME"
    home.sheet_view.showGridLines = False
    home["B2"] = "PPWR WORKSPACE DOCUMENT ENGINE"
    home["B2"].font = FONT_TITLE
    home["B3"] = f"Optional export from software workspace  ·  {stamp}"
    home["B3"].font = FONT_SUB
    home["B5"] = "QA"
    home["C5"] = "PASS" if verify["pass"] else "FAIL"
    home["C5"].fill = FILL_GREEN if verify["pass"] else PatternFill("solid", fgColor="A12622")
    home["C5"].font = FONT_PASS
    home["C5"].alignment = CENTER
    home["B6"] = "RECORDS"
    home["C6"] = n
    home["B7"] = "OPEN RULE"
    home["C7"] = "00_AC_DOCUMENT_ENGINE.cmd ONLY"
    home["B8"] = "LINK RULE"
    home["C8"] = "RELATIVE from 00_CONTROL"
    home["B10"] = (
        "• Source of truth is Workspace (not this Excel)\n"
        "• SEARCH / DOCUMENT CENTER → OPEN WORD / OPEN PDF\n"
        "• Links: ../01_PRODUCTS/<key>/<file>\n"
        "• Frozen Rev.00 deliveries are never modified"
    )
    home["B10"].alignment = Alignment(wrap_text=True, vertical="top")
    home.merge_cells("B10:G14")
    home["B10"].fill = FILL_SOFT
    home.column_dimensions["B"].width = 22
    home.column_dimensions["C"].width = 40

    search = wb.create_sheet("SEARCH")
    search["A1"] = "GLOBAL KEY SEARCH"
    search["A1"].font = FONT_TITLE
    search["A3"] = "Key"
    search["B3"] = ""
    search["B3"].fill = FILL_SOFT
    search["B3"].border = THIN
    search["A5"] = "Result Key"
    search["A6"] = "Description"
    search["A7"] = "Revision"
    search["A8"] = "Status"
    search["B5"] = '=IF($B$3="","",IFERROR(INDEX(SEARCH_DATA!A:A,MATCH($B$3,SEARCH_DATA!A:A,0)),"NOT FOUND"))'
    search["B6"] = '=IF(OR($B$5="",$B$5="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!B:B,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    search["B7"] = '=IF(OR($B$5="",$B$5="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!C:C,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    search["B8"] = '=IF(OR($B$5="",$B$5="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!D:D,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    search["A10"] = "Document"
    search["B10"] = "OPEN WORD"
    search["C10"] = "OPEN PDF"
    for cell in (search["A10"], search["B10"], search["C10"]):
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
    for i, (label, stem) in enumerate(
        [
            ("Technical File", "01_Technical_File"),
            ("EU DoC", "02_EU_DoC"),
            ("Label", "03_Label"),
            ("Shipment Statement", "04_Shipment_Statement"),
        ]
    ):
        r = 11 + i
        search.cell(r, 1, label).font = FONT_BODY
        search.cell(
            r,
            2,
            f'=IF(OR($B$5="",$B$5="NOT FOUND"),"",'
            f'HYPERLINK("../01_PRODUCTS/"&$B$5&"\\{stem}.docx","OPEN WORD"))',
        )
        search.cell(r, 2).font = FONT_LINK
        search.cell(r, 2).alignment = CENTER
        search.cell(
            r,
            3,
            f'=IF(OR($B$5="",$B$5="NOT FOUND"),"",'
            f'HYPERLINK("../01_PRODUCTS/"&$B$5&"\\{stem}.pdf","OPEN PDF"))',
        )
        search.cell(r, 3).font = FONT_LINK
        search.cell(r, 3).alignment = CENTER
    search.column_dimensions["A"].width = 22
    search.column_dimensions["B"].width = 48
    search.column_dimensions["C"].width = 14

    sdata = wb.create_sheet("SEARCH_DATA")
    for i, h in enumerate(["Key", "Description", "Revision", "Status", "Set Code"], 1):
        c = sdata.cell(1, i, h)
        c.fill = FILL_NAVY
        c.font = FONT_HDR
        c.alignment = CENTER
        c.border = THIN

    dc = wb.create_sheet("DOCUMENT_CENTER")
    dc["A1"] = "DOCUMENT CENTER — WORKSPACE EXPORT"
    dc["A1"].font = FONT_TITLE
    dc.merge_cells("A1:L1")
    dc["A2"] = (
        f"Relative links 00_CONTROL -> ../01_PRODUCTS/<key>/  ·  Records: {n}  ·  "
        "Open via 00_AC_DOCUMENT_ENGINE.cmd"
    )
    dc["A2"].font = FONT_BODY
    dc.merge_cells("A2:L2")
    headers = ["Key", "Description", "Revision", "Status", "Set"] + [h for h, _ in STEM_COLS]
    for i, h in enumerate(headers, 1):
        cell = dc.cell(4, i, h)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    for i, rec in enumerate(records, 1):
        key = rec["key"]
        r = 4 + i
        dc.cell(r, 1, key).font = FONT_BODY
        dc.cell(r, 2, rec["label"]).font = FONT_BODY
        dc.cell(r, 3, rec["revision"]).font = FONT_BODY
        st = dc.cell(r, 4, rec.get("status") or "ISSUED")
        st.fill = FILL_GREEN
        st.font = FONT_PASS
        st.alignment = CENTER
        dc.cell(r, 5, rec.get("set_code") or "").font = FONT_BODY
        for c in range(1, 6):
            dc.cell(r, c).border = THIN
        for j, (_, fname) in enumerate(STEM_COLS):
            label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
            cell = dc.cell(r, 6 + j)
            _rel_link(cell, f"../01_PRODUCTS/{key}/{fname}", label)
            cell.border = THIN
        for c, val in enumerate(
            [key, rec["label"], rec["revision"], rec.get("status") or "ISSUED", rec.get("set_code") or ""],
            1,
        ):
            cell = sdata.cell(i + 1, c, val)
            cell.font = FONT_BODY
            cell.border = THIN

    dc.column_dimensions["A"].width = 16
    dc.column_dimensions["B"].width = 40
    for c in range(3, 14):
        dc.column_dimensions[get_column_letter(c)].width = 12
    dc.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + n}"
    dc.freeze_panes = "A5"
    sdata.column_dimensions["A"].width = 16
    sdata.column_dimensions["B"].width = 40

    qa = wb.create_sheet("QA")
    qa["A1"] = "QA — WORKSPACE ENGINE EXPORT"
    qa["A1"].font = FONT_TITLE
    qa["A3"] = "Checked"
    qa["B3"] = verify["checked"]
    qa["A4"] = "Missing"
    qa["B4"] = len(verify["missing"])
    qa["A5"] = "Result"
    qa["B5"] = "PASS" if verify["pass"] else "FAIL"
    qa["B5"].fill = FILL_GREEN if verify["pass"] else PatternFill("solid", fgColor="A12622")
    qa["B5"].font = FONT_PASS

    wb.save(path)


def _write_launcher() -> None:
    _assert_workspace(LAUNCHER)
    LAUNCHER.write_text(
        "@echo off\r\n"
        "cd /d \"%~dp0\"\r\n"
        "echo ============================================================\r\n"
        "echo  INCI AKU PPWR — WORKSPACE Document Engine (optional export)\r\n"
        "echo  Source of truth is the software Workspace.\r\n"
        "echo  Open ONLY via this CMD.\r\n"
        "echo ============================================================\r\n"
        f'if not exist "00_CONTROL\\{ENGINE_NAME}" (\r\n'
        f"  echo ERROR: engine missing: 00_CONTROL\\{ENGINE_NAME}\r\n"
        "  pause\r\n"
        "  exit /b 1\r\n"
        ")\r\n"
        f'start \"\" \"00_CONTROL\\{ENGINE_NAME}\"\r\n',
        encoding="utf-8",
    )


def open_workspace_engine() -> dict:
    if not LAUNCHER.exists():
        raise HTTPException(404, "Launcher missing — export engine first")
    subprocess.Popen(["cmd", "/c", str(LAUNCHER)], cwd=str(DELIVERY))
    return {"launched": str(LAUNCHER)}


def open_workspace_engine_folder() -> dict:
    DELIVERY.mkdir(parents=True, exist_ok=True)
    os.startfile(str(DELIVERY))
    return {"opened": str(DELIVERY)}
