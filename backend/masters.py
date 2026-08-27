"""Read-only master workbook helpers (Starter / Industrial)."""

from __future__ import annotations

import os
import re
from collections import Counter
from functools import lru_cache
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
_REF = ROOT / "data_reference"
_PIMS = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS\output")


def _master_roots() -> list[Path]:
    roots: list[Path] = [_REF, ROOT / "delivery"]
    env = os.environ.get("INCI_PPWR_MASTERS_ROOT") or os.environ.get("INCI_PPWR_DELIVERY_ROOT")
    if env:
        roots.insert(0, Path(env))
    if _PIMS.exists():
        roots.append(_PIMS)
    seen: set[str] = set()
    out: list[Path] = []
    for r in roots:
        key = str(r)
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out

STARTER_NAMES = ("INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx",)
INDUSTRIAL_NAMES = (
    "INCI_AKU_PPWR_INDUSTRIAL_MASTER_FROM_EXCEL_Rev00.xlsx",
    "INCI_AKU_PPWR_INDUSTRIAL_MASTER_Rev00.xlsx",
)

_STATUS_TR = {
    "CONTROLLED PACKAGING SET": "Kontrollü ambalaj seti",
    "BOM DATA REQUIRED": "Ambalaj verisi eksik",
    "DATA REQUIRED": "Veri eksik",
}


def _public_status(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    return _STATUS_TR.get(raw.upper(), raw)


_MEASURE_TR = {
    "MASS-BASED / N/A": "Ağırlık bazlı / Yok",
    "N/A / MASS-BASED": "Yok / Ağırlık bazlı",
    "N/A / MASS BASED": "Yok / Ağırlık bazlı",
    "MASS BASED / N/A": "Ağırlık bazlı / Yok",
    "MASS-BASED": "Ağırlık bazlı",
    "MASS BASED": "Ağırlık bazlı",
    "N/A": "Yok",
}


def _public_measure(value: Any) -> Any:
    if value is None or isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    raw = str(value).strip()
    if not raw:
        return value
    key = re.sub(r"\s+", " ", raw).upper().replace("–", "/").replace("—", "/")
    if key in _MEASURE_TR:
        return _MEASURE_TR[key]
    out = re.sub(r"MASS[-\s]?BASED", "Ağırlık bazlı", raw, flags=re.I)
    out = re.sub(r"\bN/A\b", "Yok", out, flags=re.I)
    return out


def master_path(kind: str) -> Path:
    names = STARTER_NAMES if kind == "starter" else INDUSTRIAL_NAMES if kind == "industrial" else ()
    if not names:
        raise HTTPException(404, f"Unknown master: {kind}")
    roots = _master_roots()
    for root in roots:
        for name in names:
            p = root / name
            if p.exists():
                return p
    raise HTTPException(404, f"Master file missing for {kind}")


@lru_cache(maxsize=4)
def _sheet_rows(kind: str, sheet: str) -> tuple[tuple[str, ...], tuple[tuple, ...]]:
    path = master_path(kind)
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise HTTPException(404, f"Sheet missing: {sheet}")
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    headers = tuple(str(h or "").strip() for h in next(rows_iter))
    data = tuple(tuple(r) for r in rows_iter)
    wb.close()
    return headers, data


def clear_master_cache() -> None:
    _sheet_rows.cache_clear()


def _hi(headers: tuple[str, ...]) -> dict[str, int]:
    return {h: i for i, h in enumerate(headers) if h}


def starter_summary() -> dict:
    headers, rows = _sheet_rows("starter", "PRODUCT_MASTER")
    hi = _hi(headers)
    status = Counter()
    sets = Counter()
    for row in rows:
        if not row or not row[hi.get("Product Code", 0)]:
            continue
        st = str(row[hi["Physical Packaging Status"]] or "") if "Physical Packaging Status" in hi else ""
        sc = str(row[hi["Packaging Set Code"]] or "") if "Packaging Set Code" in hi else ""
        status[st or "(blank)"] += 1
        if sc and "NOT" not in sc.upper() and "DATA" not in sc.upper():
            sets[sc] += 1
    return {
        "kind": "starter",
        "path": str(master_path("starter")),
        "products": sum(status.values()),
        "status": dict(status.most_common()),
        "unique_sets": len(sets),
        "top_sets": [{"set": k, "products": v} for k, v in sets.most_common(12)],
    }


def industrial_summary() -> dict:
    # try PRODUCT-like sheet names
    path = master_path("industrial")
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    sheet = "PRODUCT_MASTER" if "PRODUCT_MASTER" in sheets else sheets[0]
    headers, rows = _sheet_rows("industrial", sheet)
    hi = _hi(headers)
    code_col = hi.get("Product Code") or hi.get("PRODUCT_CODE") or 0
    n = sum(1 for r in rows if r and r[code_col])
    return {
        "kind": "industrial",
        "path": str(path),
        "sheet": sheet,
        "products": n,
        "sheets": sheets[:20],
    }


def search_products(kind: str, q: str | None, limit: int = 50) -> dict:
    if kind == "starter":
        sheet = "PRODUCT_MASTER"
    else:
        path = master_path(kind)
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet = "PRODUCT_MASTER" if "PRODUCT_MASTER" in wb.sheetnames else wb.sheetnames[0]
        wb.close()
    headers, rows = _sheet_rows(kind, sheet)
    hi = _hi(headers)
    code_i = hi.get("Product Code") or hi.get("PRODUCT_CODE") or 0
    desc_i = hi.get("Technical Description") or hi.get("Product Description") or hi.get("DESCRIPTION")
    set_i = hi.get("Packaging Set Code")
    status_i = hi.get("Physical Packaging Status") or hi.get("Status")
    tare_i = hi.get("Packaging Tare kg")
    qq = (q or "").strip().upper()
    out = []
    for row in rows:
        if not row or not row[code_i]:
            continue
        pc = str(row[code_i]).strip()
        desc = str(row[desc_i] or "") if desc_i is not None else ""
        if qq and qq not in pc.upper() and qq not in desc.upper():
            continue
        out.append(
            {
                "product_code": pc,
                "description": desc,
                "set_code": str(row[set_i] or "") if set_i is not None else "",
                "status": _public_status(str(row[status_i] or "") if status_i is not None else ""),
                "tare_kg": row[tare_i] if tare_i is not None else None,
            }
        )
        if len(out) >= limit:
            break
    return {"kind": kind, "count": len(out), "products": out, "query": q or ""}


def get_product(kind: str, code: str) -> dict:
    if kind == "starter":
        sheet = "PRODUCT_MASTER"
    else:
        path = master_path(kind)
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet = "PRODUCT_MASTER" if "PRODUCT_MASTER" in wb.sheetnames else wb.sheetnames[0]
        wb.close()
    headers, rows = _sheet_rows(kind, sheet)
    hi = _hi(headers)
    code_i = hi.get("Product Code") or hi.get("PRODUCT_CODE") or 0
    desc_i = hi.get("Technical Description") or hi.get("Product Description") or hi.get("DESCRIPTION")
    set_i = hi.get("Packaging Set Code")
    status_i = hi.get("Physical Packaging Status") or hi.get("Status")
    tare_i = hi.get("Packaging Tare kg")
    target = code.strip()
    for row in rows:
        if not row or not row[code_i]:
            continue
        pc = str(row[code_i]).strip()
        if pc != target:
            continue
        p = {
            "product_code": pc,
            "description": str(row[desc_i] or "") if desc_i is not None else "",
            "set_code": str(row[set_i] or "") if set_i is not None else "",
            "status": _public_status(str(row[status_i] or "") if status_i is not None else ""),
            "tare_kg": row[tare_i] if tare_i is not None else None,
        }
        bom = []
        set_code = p.get("set_code") or ""
        if kind == "starter" and set_code and "NOT" not in set_code.upper() and "DATA" not in set_code.upper():
            bom = get_bom("starter", set_code).get("lines", [])
        return {**p, "bom": bom}
    raise HTTPException(404, f"Product not found: {code}")


def get_bom(kind: str, set_code: str) -> dict:
    if kind != "starter":
        raise HTTPException(400, "BOM browse currently supported for starter master")
    headers, rows = _sheet_rows("starter", "BOM_MASTER")
    hi = _hi(headers)
    sc_i = hi["Packaging Set Code"]
    lines = []
    for row in rows:
        if str(row[sc_i] or "").strip() != set_code:
            continue
        lines.append(
            {
                "component_code": str(row[hi.get("Component Code", 3)] or ""),
                "description": str(row[hi.get("Component Description", 4)] or ""),
                "qty": _public_measure(row[hi.get("Quantity", 5)]),
                "uom": _public_measure(str(row[hi.get("UOM", 6)] or "")),
                "unit_weight": _public_measure(row[hi.get("Unit Weight", 7)]) if "Unit Weight" in hi else None,
                "line_weight": _public_measure(row[hi.get("Line Weight", 8)]) if "Line Weight" in hi else None,
            }
        )
    # config meta
    ch, crows = _sheet_rows("starter", "CONFIG_MASTER")
    chi = _hi(ch)
    meta = {}
    for row in crows:
        if str(row[chi["Packaging Set Code"]] or "").strip() == set_code:
            meta = {
                "final_id": str(row[chi.get("Final Configuration ID", 0)] or ""),
                "tare_kg": row[chi.get("Packaging Tare kg")] if "Packaging Tare kg" in chi else None,
                "description": str(row[chi.get("Packaging Description", 0)] or ""),
                "product_count": row[chi.get("Product Count")] if "Product Count" in chi else None,
            }
            break
    return {"set_code": set_code, "lines": lines, "meta": meta}


def search_components(kind: str = "starter", q: str = "", limit: int = 80) -> dict:
    """Unique packaging components from BOM_MASTER (read-only master)."""
    if kind not in ("starter", "industrial"):
        raise HTTPException(404, f"Unknown master: {kind}")
    sheet = "BOM_MASTER"
    try:
        headers, rows = _sheet_rows(kind, sheet)
    except HTTPException:
        if kind == "industrial":
            headers, rows = _sheet_rows("starter", sheet)
            kind = "starter"
        else:
            raise
    hi = _hi(headers)
    code_i = hi.get("Component Code")
    desc_i = hi.get("Component Description")
    set_i = hi.get("Packaging Set Code")
    if code_i is None:
        raise HTTPException(500, "BOM_MASTER missing Component Code")

    needle = (q or "").strip().lower()
    by_code: dict[str, dict] = {}
    for row in rows:
        code = str(row[code_i] or "").strip()
        if not code:
            continue
        desc = str(row[desc_i] or "").strip() if desc_i is not None else ""
        set_code = str(row[set_i] or "").strip() if set_i is not None else ""
        if needle and needle not in code.lower() and needle not in desc.lower() and needle not in set_code.lower():
            continue
        if code not in by_code:
            by_code[code] = {
                "component_code": code,
                "description": desc,
                "set_codes": [],
                "set_count": 0,
            }
        entry = by_code[code]
        if desc and not entry["description"]:
            entry["description"] = desc
        if set_code and set_code not in entry["set_codes"]:
            entry["set_codes"].append(set_code)

    items = list(by_code.values())
    for it in items:
        it["set_count"] = len(it["set_codes"])
        it["set_codes"] = it["set_codes"][:12]
    items.sort(key=lambda x: x["component_code"])
    total = len(items)
    items = items[: max(1, min(limit, 300))]
    return {"kind": kind, "q": q, "total": total, "returned": len(items), "components": items}
