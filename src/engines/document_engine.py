"""
Document Engine — exposes merge variables for future Word templates.

Does NOT generate Word files. Builds a variable catalog (JSON + workbook sheet)
for Technical File, Declaration of Conformity, and Shipment Statement.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Variable catalogs: token → description / source path
TECHNICAL_FILE_VARS: dict[str, str] = {
    "TF.CODE": "TECHNICAL_FILE.TECHNICAL_FILE_CODE",
    "TF.TITLE": "TECHNICAL_FILE.TITLE",
    "TF.REVISION": "TECHNICAL_FILE.REVISION_NO",
    "TF.ASSESSMENT_DATE": "TECHNICAL_FILE.ASSESSMENT_DATE",
    "TF.RECYCLABILITY_SUMMARY": "TECHNICAL_FILE.RECYCLABILITY_SUMMARY",
    "TF.SUBSTANCE_NOTES": "TECHNICAL_FILE.SUBSTANCE_OF_CONCERN_NOTES",
    "TF.DFR_NOTES": "TECHNICAL_FILE.DESIGN_FOR_RECYCLING_NOTES",
    "TF.EFFECTIVE_FROM": "TECHNICAL_FILE.EFFECTIVE_FROM",
    "TF.EFFECTIVE_TO": "TECHNICAL_FILE.EFFECTIVE_TO",
    "TF.OWNER_PERSON_ID": "TECHNICAL_FILE.OWNER_PERSON_ID",
    "PC.ID": "PACKAGING_CONFIGURATION.PACKAGING_CONFIGURATION_ID",
    "PC.GROUP_CODE": "PACKAGING_CONFIGURATION.CONFIG_GROUP_CODE",
    "PC.NAME": "PACKAGING_CONFIGURATION.PACKAGING_CONFIGURATION_NAME",
    "PC.DESCRIPTION": "PACKAGING_CONFIGURATION.DESCRIPTION",
    "PC.REVISION": "PACKAGING_CONFIGURATION.REVISION_NO",
    "PC.COMPONENT_LIST": "Derived from PACKAGING_CONFIGURATION_LINE + COMPONENT",
    "PC.MATERIAL_COMPOSITION": "Derived from COMPONENT_MATERIAL + ENG_MATERIAL_WEIGHT",
    "PC.MATERIAL_SUMMARY": "ENG_PLASTIC/PAPER/WOOD_SUMMARY",
    "PC.WEIGHT_TOTAL_G": "ENG_PACKAGING_WEIGHT totals",
    "PC.SUPPLIER_MATRIX": "COMPONENT.SUPPLIER_ID + DOCUMENT_LINK",
    "PC.EVIDENCE_MATRIX": "DOCUMENT_LIBRARY via DOCUMENT_LINK (TF/Component)",
    "PC.DRAWINGS": "DOCUMENT_LINK filtered by document type DRAWING",
    "PC.PHOTOGRAPHS": "DOCUMENT_LINK filtered by document type PHOTO",
    "PC.MINIMISATION_ASSESSMENT": "Linked assessment documents + NOTES",
    "PC.EMPTY_SPACE_ASSESSMENT": "Linked assessment documents + NOTES",
    "PC.LABELLING_ASSESSMENT": "Linked assessment documents + NOTES",
    "PC.REVISION_HISTORY": "CONFIG revisions + TF.REVISION_NO timeline",
}

DECLARATION_VARS: dict[str, str] = {
    "DOC.NUMBER": "DECLARATION_OF_CONFORMITY.DOC_NUMBER",
    "DOC.TITLE": "DECLARATION_OF_CONFORMITY.TITLE",
    "DOC.REVISION": "DECLARATION_OF_CONFORMITY.REVISION_NO",
    "DOC.ISSUE_DATE": "DECLARATION_OF_CONFORMITY.ISSUE_DATE",
    "DOC.VALID_UNTIL": "DECLARATION_OF_CONFORMITY.VALID_UNTIL",
    "DOC.REGULATION_REFERENCE": "DECLARATION_OF_CONFORMITY.REGULATION_REFERENCE",
    "DOC.CONFORMITY_STATEMENT": "DECLARATION_OF_CONFORMITY.CONFORMITY_STATEMENT",
    "DOC.LEGAL_ENTITY_ID": "DECLARATION_OF_CONFORMITY.LEGAL_ENTITY_ID",
    "DOC.RESPONSIBLE_PERSON_ID": "DECLARATION_OF_CONFORMITY.RESPONSIBLE_PERSON_ID",
    "DOC.PACKAGING_CONFIGURATION_ID": "DECLARATION_OF_CONFORMITY.PACKAGING_CONFIGURATION_ID",
    "DOC.TECHNICAL_FILE_ID": "DECLARATION_OF_CONFORMITY.TECHNICAL_FILE_ID",
    "SCN.CODE": "COMMERCIAL_SCENARIO.COMMERCIAL_SCENARIO_CODE",
    "SCN.NAME": "COMMERCIAL_SCENARIO.COMMERCIAL_SCENARIO_NAME",
    "SCN.INCOTERM_ID": "COMMERCIAL_SCENARIO.INCOTERM_ID",
    "SCN.CUSTOMER_ID": "COMMERCIAL_SCENARIO.CUSTOMER_ID",
    "SCN.COUNTRY_ID": "COMMERCIAL_SCENARIO.DESTINATION_COUNTRY_ID",
    "TF.CODE": "TECHNICAL_FILE.TECHNICAL_FILE_CODE (from Packaging Configuration)",
    "PC.NAME": "PACKAGING_CONFIGURATION.PACKAGING_CONFIGURATION_NAME",
}

STATEMENT_VARS: dict[str, str] = {
    "STM.NUMBER": "STATEMENT.STATEMENT_NUMBER",
    "STM.TITLE": "STATEMENT.TITLE",
    "STM.REVISION": "STATEMENT.REVISION_NO",
    "STM.PERIOD_FROM": "STATEMENT.PERIOD_FROM",
    "STM.PERIOD_TO": "STATEMENT.PERIOD_TO",
    "STM.STATUS_ID": "STATEMENT.STATUS_ID",
    "SHP.NUMBER": "SHIPMENT.SHIPMENT_NUMBER",
    "SHP.DATE": "SHIPMENT.SHIP_DATE",
    "SHP.QTY": "SHIPMENT.QTY_PRODUCT_UNITS",
    "SHP.LOT": "SHIPMENT.EXTERNAL_REF (Lot Number)",
    "SHP.PACKAGING_CONFIGURATION_ID": "SHIPMENT.PACKAGING_CONFIGURATION_ID",
    "SHP.COMMERCIAL_SCENARIO_ID": "SHIPMENT.COMMERCIAL_SCENARIO_ID",
    "SCN.INCOTERM_ID": "COMMERCIAL_SCENARIO.INCOTERM_ID",
    "SCN.CUSTOMER_ID": "COMMERCIAL_SCENARIO.CUSTOMER_ID",
    "SCN.COUNTRY_ID": "COMMERCIAL_SCENARIO.DESTINATION_COUNTRY_ID",
    "PC.NAME": "PACKAGING_CONFIGURATION.PACKAGING_CONFIGURATION_NAME",
    "TF.CODE": "TECHNICAL_FILE for Packaging Configuration",
    "DOC.NUMBER": "DoC variant for Commercial Scenario",
    "STM.COMPOSITION": "STATEMENT_LINE / SHIPMENT_LINE freeze",
    "STM.MATERIAL_BREAKDOWN": "ENG_STATEMENT",
    "STM.TOTAL_WEIGHT_KG": "ENG_SHIPMENT_WEIGHT / ENG_STATEMENT",
    "STM.COMPONENT_LIST": "SHIPMENT_LINE components",
    "STM.SUPPLIER_TRACEABILITY": "COMPONENT.SUPPLIER_ID snapshots",
    "STM.DOCUMENT_REFS": "DOCUMENT_LINK for statement/shipment",
}

CATALOGS = {
    "TECHNICAL_FILE": TECHNICAL_FILE_VARS,
    "DECLARATION_OF_CONFORMITY": DECLARATION_VARS,
    "SHIPMENT_STATEMENT": STATEMENT_VARS,
}


class DocumentEngine:
    """Resolve merge-variable payloads for a future Word template renderer."""

    def catalog(self) -> dict[str, dict[str, str]]:
        return {k: dict(v) for k, v in CATALOGS.items()}

    def tokens(self, document_type: str) -> list[str]:
        return list(CATALOGS[document_type].keys())

    def build_context_stub(
        self,
        document_type: str,
        *,
        packaging_configuration_id: int | None = None,
        commercial_scenario_id: int | None = None,
        shipment_number: str | None = None,
    ) -> dict[str, Any]:
        """Return placeholder context (values filled by Excel/ERP at render time)."""
        tokens = self.tokens(document_type)
        ctx = {t: None for t in tokens}
        ctx["_meta"] = {
            "document_type": document_type,
            "packaging_configuration_id": packaging_configuration_id,
            "commercial_scenario_id": commercial_scenario_id,
            "shipment_number": shipment_number,
            "word_generation": False,
            "note": "Populate values from workbook tables; merge into .docx templates later",
        }
        return ctx

    def write_json_catalog(self, path: Path) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(self.catalog(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return path


def write_document_engine_catalog(output_dir: Path) -> tuple[Path, Path]:
    """Write JSON catalog + Excel variable workbook for admins/template authors."""
    engine = DocumentEngine()
    json_path = engine.write_json_catalog(output_dir / "document_variables.json")

    xlsx_path = output_dir / "Document_Engine_Variables.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    cover = wb.create_sheet("README", 0)
    cover["A1"] = "İNCI AKÜ — Document Engine Variable Catalog"
    cover["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    cover["A3"] = "Tokens use {{TOKEN}} syntax in future Word templates."
    cover["A4"] = "This engine does NOT generate Word files — it only exposes variables."
    cover["A5"] = "Primary object: Packaging Configuration. TF never belongs to Product."

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for doc_type, variables in CATALOGS.items():
        ws = wb.create_sheet(doc_type[:31])
        ws["A1"] = "TOKEN"
        ws["B1"] = "SOURCE"
        ws["C1"] = "WORD_PLACEHOLDER"
        for col in (1, 2, 3):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font
        for i, (token, source) in enumerate(variables.items(), start=2):
            ws.cell(row=i, column=1, value=token)
            ws.cell(row=i, column=2, value=source)
            ws.cell(row=i, column=3, value="{{" + token + "}}")
        end = get_column_letter(3)
        last = 1 + len(variables)
        table = Table(displayName=f"DOC_{doc_type[:20]}", ref=f"A1:{end}{last}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 40
        ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    wb.close()
    return json_path, xlsx_path


def add_document_engine_sheet(wb) -> None:
    """Add hidden DOC_ENGINE_VARS sheet to production workbook."""
    name = "DOC_ENGINE_VARS"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "DOCUMENT_TYPE"
    ws["B1"] = "TOKEN"
    ws["C1"] = "SOURCE"
    ws["D1"] = "WORD_PLACEHOLDER"
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = fill
        ws.cell(row=1, column=col).font = font

    row = 2
    for doc_type, variables in CATALOGS.items():
        for token, source in variables.items():
            ws.cell(row=row, column=1, value=doc_type)
            ws.cell(row=row, column=2, value=token)
            ws.cell(row=row, column=3, value=source)
            ws.cell(row=row, column=4, value="{{" + token + "}}")
            row += 1

    end_row = row - 1
    table = Table(displayName="T_DOC_ENGINE_VARS", ref=f"A1:D{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    ws.sheet_state = "hidden"
    ws.protection.sheet = True
    ws.protection.password = "PIMS_TECH"
