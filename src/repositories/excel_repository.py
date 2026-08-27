"""
Excel repository — Wave 1 operational store access (Phase E).

Table-oriented read helpers for the PIMS workbook.
Does not invent schema; uses frozen sheet/table names.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType

from utils.excel import ExcelGenerationNotEnabledError, assert_excel_allowed, ensure_parent


class ExcelRepository:
    """openpyxl-backed access to the PIMS workbook."""

    def __init__(
        self,
        workbook_path: Path | None = None,
        *,
        enabled: bool = True,
    ) -> None:
        self.workbook_path = workbook_path
        self.enabled = enabled
        self._wb: WorkbookType | None = None

    def open(self, data_only: bool = False) -> WorkbookType:
        assert_excel_allowed(self.enabled)
        if self.workbook_path is None or not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")
        self._wb = load_workbook(self.workbook_path, data_only=data_only)
        return self._wb

    def save(self, path: Path | None = None) -> Path:
        assert_excel_allowed(self.enabled)
        if self._wb is None:
            raise RuntimeError("No workbook open")
        target = path or self.workbook_path
        if target is None:
            raise ValueError("No save path")
        ensure_parent(target)
        self._wb.save(target)
        return target

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    @property
    def workbook(self) -> WorkbookType:
        if self._wb is None:
            raise RuntimeError("Workbook not open — call open() first")
        return self._wb

    def sheet_names(self) -> list[str]:
        return list(self.workbook.sheetnames)

    def headers(self, table_name: str) -> list[str]:
        ws = self.workbook[table_name]
        return [c.value for c in ws[1] if c.value]

    def excel_table_names(self) -> list[str]:
        names: list[str] = []
        for ws in self.workbook.worksheets:
            names.extend(ws.tables.keys())
        return names

    def iter_data_rows(self, table_name: str) -> list[dict[str, Any]]:
        """Return non-empty body rows as dicts keyed by header."""
        ws = self.workbook[table_name]
        headers = self.headers(table_name)
        rows: list[dict[str, Any]] = []
        for r in range(2, ws.max_row + 1):
            values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            if all(v is None or v == "" for v in values):
                continue
            rows.append(dict(zip(headers, values)))
        return rows

    def count_data_rows(self, table_name: str) -> int:
        return len(self.iter_data_rows(table_name))
