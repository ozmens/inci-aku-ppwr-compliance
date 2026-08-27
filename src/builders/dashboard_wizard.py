"""
Phase 11 — PPWR daily workflow wizard on Dashboard.

No schema changes. Lot Number → SHIPMENT.EXTERNAL_REF.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from .ui_theme import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    FILL_DASH,
    FILL_INPUT,
    FILL_LOCKED,
    FILL_OUTPUT,
    FILL_PRIMARY,
    FILL_PRIMARY_MID,
    FONT_BODY,
    FONT_HEADER,
    FONT_KPI,
    FONT_MUTED,
    FONT_SUBTITLE,
    FONT_TITLE,
    THIN,
    style_nav_button_cell,
)


def _link_btn(ws, addr: str, label: str, link: str) -> None:
    cell = ws[addr]
    cell.value = label
    style_nav_button_cell(cell)
    cell.hyperlink = link


def rebuild_ppwr_dashboard(wb, settings) -> None:
    """Replace Dashboard with PPWR 7-step command center."""
    for name in ("DASHBOARD", "Dashboard"):
        if name in wb.sheetnames:
            wb.remove(wb[name])
            break
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    for r in range(1, 48):
        for c in range(1, 14):
            ws.cell(row=r, column=c).fill = FILL_DASH

    ws.merge_cells("B2:J2")
    ws["B2"] = "İNCI AKÜ  ·  PPWR COMMAND CENTER"
    ws["B2"].font = FONT_TITLE
    ws.merge_cells("B3:J3")
    ws["B3"] = (
        f"{settings.RELEASE_NAME}  |  Primary object: Packaging Configuration  ·  "
        "Yellow = select/enter  ·  Green = automatic"
    )
    ws["B3"].font = FONT_MUTED

    # NEXT banner — 7-step PPWR model
    ws.merge_cells("B4:J4")
    ws["B4"] = (
        '=IF(D8="","▶ NEXT: STEP 1 — Select Packaging Configuration",'
        'IF(D9="","▶ NEXT: STEP 2 — Select Commercial Scenario (Incoterms)",'
        'IF(OR(D10="",D10<=0),"▶ NEXT: STEP 3 — Enter Quantity",'
        'IF(D11="","▶ NEXT: STEP 3 — Enter Lot Number",'
        'IF(D12="","▶ NEXT: STEP 3 — Create Shipment (GO), then paste Shipment Number",'
        'IF(D18="MISSING","▶ NEXT: STEP 4 — Generate Shipment Statement",'
        'IF(D20="MISSING","▶ NEXT: STEP 5 — Generate Declaration of Conformity",'
        'IF(D19="MISSING","▶ NEXT: STEP 6 — Reference Technical File for this Packaging Configuration",'
        'IF(D21<>"READY","▶ NEXT: STEP 7 — Finish package, then Archive Shipment",'
        '"✅ READY — Archive Shipment (STEP 7) / export package")))))))))'
    )
    ws["B4"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["B4"].fill = FILL_PRIMARY
    ws["B4"].alignment = ALIGN_LEFT
    ws.row_dimensions[4].height = 28

    ws["B5"] = "QUICK OPEN"
    ws["B5"].font = FONT_MUTED
    for addr, label, link in [
        ("C5", "Packaging", "#'Packaging Configuration'!A1"),
        ("D5", "Shipment", "#Shipment!A1"),
        ("E5", "Statement", "#Statements!A1"),
        ("F5", "DoC", "#'Declaration of Conformity'!A1"),
        ("G5", "Tech File", "#'Technical File'!A1"),
        ("H5", "Components", "#'Component Master'!A1"),
        ("I5", "Products", "#'Product Master'!A1"),
    ]:
        _link_btn(ws, addr, label, link)

    ws.merge_cells("B6:G6")
    ws["B6"] = "PPWR DAILY WORKFLOW  —  Packaging Configuration → Scenario → Shipment → Docs"
    ws["B6"].font = FONT_HEADER
    ws["B6"].fill = FILL_PRIMARY_MID

    for col, text in enumerate(
        ("STEP", "WHAT TO DO", "YOUR INPUT", "AUTO RESULT", "STATUS", "GO"), start=2
    ):
        cell = ws.cell(row=7, column=col, value=text)
        cell.fill = FILL_LOCKED
        cell.font = FONT_BODY
        cell.border = THIN
        cell.alignment = ALIGN_CENTER

    steps = [
        (8, "1", "Select Packaging Configuration", "#'Packaging Configuration'!A1", "Pick"),
        (9, "2", "Select Commercial Scenario (Incoterms)", "#Shipment!A1", "Pick"),
        (10, "3a", "Enter Quantity (units)", "#Shipment!A1", "Type"),
        (11, "3b", "Enter Lot Number", "#Shipment!A1", "Type"),
        (12, "3c", "Create Shipment → paste Number", "#Shipment!A1", "Create"),
        (13, "4", "Generate Shipment Statement", "#Statements!A1", "Open"),
        (14, "5", "Generate Declaration of Conformity", "#'Declaration of Conformity'!A1", "Open"),
        (15, "6", "Reference Technical File", "#'Technical File'!A1", "Open"),
    ]
    for row, num, action, link, go_label in steps:
        ws.cell(row=row, column=2, value=num).fill = FILL_LOCKED
        ws.cell(row=row, column=2).border = THIN
        ws.cell(row=row, column=2).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=action).fill = FILL_LOCKED
        ws.cell(row=row, column=3).border = THIN
        for col in (4, 5, 6):
            ws.cell(row=row, column=col).border = THIN
        _link_btn(ws, f"G{row}", go_label, link)

    for row in (8, 9, 10, 11, 12):
        cell = ws.cell(row=row, column=4, value=None)
        cell.fill = FILL_INPUT
        cell.border = THIN
        cell.protection = Protection(locked=False)

    ws["D10"] = 1
    ws["D10"].fill = FILL_INPUT
    ws["D10"].protection = Protection(locked=False)

    # Step 1 — Packaging Configuration
    ws["E8"] = (
        '=IF(D8="","← select Packaging Configuration ID",'
        'IFERROR(INDEX(T_PACKAGING_CONFIGURATION[CONFIG_GROUP_CODE],MATCH(D8,T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID],0))'
        '&" · "&INDEX(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_NAME],MATCH(D8,T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID],0)),'
        '"⚠ unknown configuration"))'
    )
    ws["F8"] = '=IF(D8="","WAITING","DONE")'

    # Step 2 — Commercial Scenario (+ Incoterms / customer / country)
    ws["E9"] = (
        '=IF(D9="","← select Commercial Scenario ID",'
        'IFERROR(INDEX(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_CODE],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0))'
        '&" · Incoterm "&IFERROR(INDEX(T_COMMERCIAL_SCENARIO[INCOTERM_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),"-")'
        '&" · Country "&IFERROR(INDEX(T_COMMERCIAL_SCENARIO[DESTINATION_COUNTRY_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),"-")'
        '&" · Customer "&IFERROR(INDEX(T_COMMERCIAL_SCENARIO[CUSTOMER_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),"-"),'
        '"⚠ unknown scenario"))'
    )
    ws["F9"] = '=IF(D9="","WAITING","DONE")'

    ws["E10"] = '=IF(OR(D10="",D10<=0),"← enter quantity > 0",D10&" units")'
    ws["F10"] = '=IF(OR(D10="",D10<=0),"WAITING","DONE")'

    ws["E11"] = '=IF(D11="","← enter Lot Number (→ Shipment EXTERNAL_REF)",D11)'
    ws["F11"] = '=IF(D11="","WAITING","DONE")'

    ws["E12"] = (
        '=IF(D12="","← after create, paste SHIPMENT_NUMBER",'
        'IFERROR("OK · "&TEXT(INDEX(T_SHIPMENT[SHIP_DATE],MATCH(D12,T_SHIPMENT[SHIPMENT_NUMBER],0)),"yyyy-mm-dd")'
        '&" · Lot "&IFERROR(INDEX(T_SHIPMENT[EXTERNAL_REF],MATCH(D12,T_SHIPMENT[SHIPMENT_NUMBER],0)),"-"),'
        '"⚠ create shipment / check number"))'
    )
    ws["F12"] = (
        '=IF(D12="","WAITING",'
        'IF(ISNUMBER(MATCH(D12,T_SHIPMENT[SHIPMENT_NUMBER],0)),"DONE","⚠ not found"))'
    )

    # Step 4 — Statement
    ws["D13"] = "auto"
    ws["D13"].fill = FILL_LOCKED
    ws["E13"] = (
        '=IF(D12="","Complete STEP 3 first",'
        'IF(COUNTIF(T_STATEMENT_SHIPMENT[SHIPMENT_ID],'
        'IFERROR(INDEX(T_SHIPMENT[SHIPMENT_ID],MATCH(D12,T_SHIPMENT[SHIPMENT_NUMBER],0)),0))>0,'
        '"Statement linked","Open Statements — link this shipment"))'
    )
    ws["F13"] = '=IF(LEFT(E13,9)="Statement","DONE",IF(D12="","WAITING","NEXT"))'
    ws["D18"] = '=IF(F13="DONE","OK","MISSING")'
    ws["D18"].fill = FILL_OUTPUT

    # Step 5 — DoC (before TF reference in workflow numbering; flags D20)
    ws["D14"] = "auto"
    ws["D14"].fill = FILL_LOCKED
    ws["E14"] = (
        '=IF(D8="","Complete STEP 1 first",'
        'IFERROR(INDEX(T_DECLARATION_OF_CONFORMITY[DOC_NUMBER],MATCH(D8,T_DECLARATION_OF_CONFORMITY[PACKAGING_CONFIGURATION_ID],0)),'
        '"Open DoC — variant from scenario; link TF of Packaging Configuration"))'
    )
    ws["F14"] = '=IF(D8="","WAITING",IF(ISNUMBER(SEARCH("Open DoC",E14)),"NEXT","DONE"))'
    ws["D20"] = '=IF(F14="DONE","OK","MISSING")'
    ws["D20"].fill = FILL_OUTPUT

    # Step 6 — Technical File reference (owned by Packaging Configuration)
    ws["D15"] = "auto"
    ws["D15"].fill = FILL_LOCKED
    ws["E15"] = (
        '=IF(D8="","Complete STEP 1 first",'
        'IFERROR(INDEX(T_TECHNICAL_FILE[TECHNICAL_FILE_CODE],MATCH(D8,T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID],0))'
        '&" · rev "&IFERROR(INDEX(T_TECHNICAL_FILE[REVISION_NO],MATCH(D8,T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID],0)),"?"),'
        '"Open Technical File — PACKAGING_CONFIGURATION_ID only (never for Product)"))'
    )
    ws["F15"] = '=IF(D8="","WAITING",IF(ISNUMBER(SEARCH("Open Technical",E15)),"NEXT","DONE"))'
    ws["D19"] = '=IF(F15="DONE","OK","MISSING")'
    ws["D19"].fill = FILL_OUTPUT

    # Step 7 — Archive
    ws["B16"] = "7"
    ws["B16"].fill = FILL_LOCKED
    ws["B16"].border = THIN
    ws["C16"] = "Archive Shipment / export package"
    ws["C16"].fill = FILL_LOCKED
    ws["C16"].border = THIN
    ws["D16"] = "auto"
    ws["D16"].fill = FILL_LOCKED
    ws["D16"].border = THIN
    ws["E16"] = (
        '=IF(AND(F8="DONE",F9="DONE",F10="DONE",F11="DONE",F12="DONE",F13="DONE",F14="DONE",F15="DONE"),'
        '"READY — archive: Statement + DoC + Tech File ref + shipment freeze (do not edit lines)",'
        '"INCOMPLETE — finish steps showing NEXT / WAITING")'
    )
    ws["E16"].fill = FILL_OUTPUT
    ws["E16"].border = THIN
    ws["F16"] = '=IF(ISNUMBER(SEARCH("READY",E16)),"DONE","NEXT")'
    ws["F16"].fill = FILL_OUTPUT
    ws["F16"].border = THIN
    _link_btn(ws, "G16", "Archive", "#Shipment!A1")
    ws["D21"] = '=IF(F16="DONE","READY","NOT READY")'
    ws["D21"].fill = FILL_OUTPUT

    for row in range(8, 16):
        for col in (5, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_OUTPUT
            cell.protection = Protection(locked=True)

    for formula, cells in (
        ("=NR_PACKAGING_CONFIGURATION_ID", "D8"),
        ("=NR_COMMERCIAL_SCENARIO_ID", "D9"),
    ):
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Select a valid ID"
        dv.errorTitle = "PPWR Workflow"
        ws.add_data_validation(dv)
        dv.add(ws[cells])

    # Draft values for Shipment sheet
    ws.merge_cells("B23:G23")
    ws["B23"] = (
        "STEP 3 DRAFT — paste onto Shipment  |  Customer/Country/Incoterms come from Scenario  |  "
        "Lot Number → EXTERNAL_REF"
    )
    ws["B23"].font = FONT_HEADER
    ws["B23"].fill = FILL_PRIMARY_MID

    draft = [
        (24, "PACKAGING_CONFIGURATION_ID", "=D8"),
        (25, "COMMERCIAL_SCENARIO_ID", "=D9"),
        (26, "QTY_PRODUCT_UNITS", "=D10"),
        (27, "EXTERNAL_REF (Lot Number)", "=D11"),
        (
            28,
            "PRODUCT_ID (auto from scenario)",
            '=IF(D9="","",IFERROR(INDEX(T_COMMERCIAL_SCENARIO[PRODUCT_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),""))',
        ),
        (
            29,
            "TRANSPORT_CONFIGURATION_ID (auto)",
            '=IF(D9="","",IFERROR(INDEX(T_COMMERCIAL_SCENARIO[TRANSPORT_CONFIGURATION_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),""))',
        ),
        (
            30,
            "DESTINATION_COUNTRY_ID (auto)",
            '=IF(D9="","",IFERROR(INDEX(T_COMMERCIAL_SCENARIO[DESTINATION_COUNTRY_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),""))',
        ),
        (
            31,
            "INCOTERM_ID (auto)",
            '=IF(D9="","",IFERROR(INDEX(T_COMMERCIAL_SCENARIO[INCOTERM_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),""))',
        ),
        (
            32,
            "CUSTOMER_ID (auto)",
            '=IF(D9="","",IFERROR(INDEX(T_COMMERCIAL_SCENARIO[CUSTOMER_ID],MATCH(D9,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],0)),""))',
        ),
    ]
    for row, label, formula in draft:
        lab = ws.cell(row=row, column=2, value=label)
        lab.fill = FILL_LOCKED
        lab.border = THIN
        val = ws.cell(row=row, column=3, value=formula)
        val.fill = FILL_OUTPUT
        val.border = THIN
        val.protection = Protection(locked=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)

    ws.merge_cells("B33:G33")
    ws["B33"] = "DOCUMENT PACKAGE (auto) — TF from Packaging Configuration · DoC from Scenario · Statement from Shipment"
    ws["B33"].font = FONT_HEADER
    ws["B33"].fill = FILL_PRIMARY_MID
    for row, label, formula in (
        (34, "Technical File", "=E15"),
        (35, "Declaration of Conformity", "=E14"),
        (36, "Shipment Statement", "=E13"),
        (37, "Archive / export readiness", "=E16"),
    ):
        ws.cell(row=row, column=2, value=label).fill = FILL_LOCKED
        ws.cell(row=row, column=2).border = THIN
        cell = ws.cell(row=row, column=3, value=formula)
        cell.fill = FILL_OUTPUT
        cell.border = THIN
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)

    ws["I7"] = "KPIs"
    ws["I7"].font = FONT_SUBTITLE
    for row, label, formula in (
        (8, "Packaging Configs", "=COUNTA(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID])"),
        (9, "Shipments", "=COUNTA(T_SHIPMENT[SHIPMENT_ID])"),
        (10, "Statements", "=COUNTA(T_STATEMENT[STATEMENT_ID])"),
        (11, "Tech Files", "=COUNTA(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID])"),
        (12, "Validation", '=IF(COUNTIF(T_ENG_VALIDATION[RESULT],"ERROR")>0,"ACTION","OK")'),
        (13, "Missing Lot#", "=COUNTBLANK(T_SHIPMENT[EXTERNAL_REF])"),
        (14, "Latest Ship", '=IFERROR(TEXT(MAX(T_SHIPMENT[SHIP_DATE]),"yyyy-mm-dd"),"-")'),
        (15, "Revision", f'="{settings.WORKBOOK_REVISION}"'),
    ):
        ws.cell(row=row, column=9, value=label).fill = FILL_PRIMARY_MID
        ws.cell(row=row, column=9).font = FONT_HEADER
        ws.cell(row=row, column=9).border = THIN
        v = ws.cell(row=row, column=10, value=formula)
        v.fill = FILL_OUTPUT
        v.font = FONT_KPI
        v.border = THIN
        v.protection = Protection(locked=True)

    ws.merge_cells("B39:G39")
    ws["B39"] = "GLOBAL SEARCH — Packaging Config · Scenario · Shipment/Lot · Component"
    ws["B39"].font = FONT_HEADER
    ws["B39"].fill = FILL_PRIMARY_MID
    ws["B40"] = "Search"
    ws["B40"].fill = FILL_LOCKED
    ws["C40"] = ""
    ws["C40"].fill = FILL_INPUT
    ws["C40"].border = THIN
    ws["C40"].protection = Protection(locked=False)
    ws["D40"] = "Microsoft 365 FILTER"
    ws["D40"].font = FONT_MUTED

    for name in (
        "NR_SEARCH_TERM",
        "NR_WZ_PRODUCT_ID",
        "NR_WZ_PACK_CFG_ID",
        "NR_WZ_SCENARIO_ID",
        "NR_WZ_QTY",
        "NR_WZ_LOT",
        "NR_WZ_SHIPMENT_NO",
    ):
        if name in wb.defined_names:
            del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name="NR_SEARCH_TERM", attr_text="'Dashboard'!$C$40"))
    wb.defined_names.add(DefinedName(name="NR_WZ_PACK_CFG_ID", attr_text="'Dashboard'!$D$8"))
    wb.defined_names.add(DefinedName(name="NR_WZ_SCENARIO_ID", attr_text="'Dashboard'!$D$9"))
    wb.defined_names.add(DefinedName(name="NR_WZ_QTY", attr_text="'Dashboard'!$D$10"))
    wb.defined_names.add(DefinedName(name="NR_WZ_LOT", attr_text="'Dashboard'!$D$11"))
    wb.defined_names.add(DefinedName(name="NR_WZ_SHIPMENT_NO", attr_text="'Dashboard'!$D$12"))

    for row, label, formula in (
        (
            41,
            "Packaging Config",
            '=IF(NR_SEARCH_TERM="","",IFERROR(FILTER(HSTACK(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID],'
            'T_PACKAGING_CONFIGURATION[CONFIG_GROUP_CODE],T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_NAME]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_PACKAGING_CONFIGURATION[CONFIG_GROUP_CODE]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_NAME]&"")))),"(none)"))',
        ),
        (
            42,
            "Scenario",
            '=IF(NR_SEARCH_TERM="","",IFERROR(FILTER(HSTACK(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],'
            'T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_CODE],T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_NAME]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_CODE]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_NAME]&"")))),"(none)"))',
        ),
        (
            43,
            "Shipment / Lot",
            '=IF(NR_SEARCH_TERM="","",IFERROR(FILTER(HSTACK(T_SHIPMENT[SHIPMENT_NUMBER],T_SHIPMENT[EXTERNAL_REF],'
            'T_SHIPMENT[PACKAGING_CONFIGURATION_ID]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_SHIPMENT[SHIPMENT_NUMBER]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_SHIPMENT[EXTERNAL_REF]&"")))),"(none)"))',
        ),
        (
            44,
            "Component",
            '=IF(NR_SEARCH_TERM="","",IFERROR(FILTER(HSTACK(T_COMPONENT[COMPONENT_ID],T_COMPONENT[COMPONENT_CODE],'
            'T_COMPONENT[COMPONENT_NAME]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMPONENT[COMPONENT_CODE]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMPONENT[COMPONENT_NAME]&"")))),"(none)"))',
        ),
    ):
        ws.cell(row=row, column=2, value=label).fill = FILL_LOCKED
        ws.cell(row=row, column=2).border = THIN
        cell = ws.cell(row=row, column=3, value=formula)
        cell.fill = FILL_OUTPUT
        cell.border = THIN
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)

    ws.merge_cells("B46:J46")
    ws["B46"] = (
        "Ship flow: Config → Scenario → Qty+Lot → Statement → DoC → Tech File ref → Archive  ·  "
        "Import: templates/import  ·  Doc engine: DOC_ENGINE_VARS  ·  PIMS_UI / PIMS_TECH"
    )
    ws["B46"].font = FONT_MUTED

    for addr, label in (
        ("C18", "flag_stmt"),
        ("C19", "flag_tf"),
        ("C20", "flag_doc"),
        ("C21", "flag_archive"),
    ):
        ws[addr] = label
        ws[addr].font = FONT_MUTED

    for col, width in (
        ("B", 10),
        ("C", 44),
        ("D", 16),
        ("E", 56),
        ("F", 14),
        ("G", 10),
        ("I", 18),
        ("J", 14),
    ):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 30
