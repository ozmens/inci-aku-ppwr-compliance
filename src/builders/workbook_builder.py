"""
Build normalized PIMS workbook via openpyxl.

Creates:
  - one worksheet per frozen table (no formulas in masters)
  - Excel Tables (ListObjects)
  - PK named ranges + FK data-validation placeholders
  - SYS_RELATIONSHIPS technical sheet (hidden/protected)
  - ENG_* calculation sheets (structured-reference formulas only)

No visual formatting on master tables.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from models.registry import SchemaRegistry

from .engine_builder import all_engines, write_all_engines
from .named_ranges import apply_pk_named_ranges, parent_pk_for_fk
from .relationships import RELATIONSHIP_SHEET, write_relationships_sheet
from .seed_data import apply_seed_data
from .sheet_builder import SheetBuilder, excel_col_letter
from .inci_ops_ux import InciOpsUX
from .production_finalize import protect_all_system_sheets
from .ui_builder import UIBuilder, UI_SHEETS
from engines.document_engine import add_document_engine_sheet

TECHNICAL_SHEETS = frozenset(
    {
        "SYS_WORKBOOK_INFO",
        "SYS_PARAMETER",
        RELATIONSHIP_SHEET,
    }
)


class ExcelGenerationDisabledError(RuntimeError):
    """Raised when workbook generation feature flag is off."""


class WorkbookBuilder:
    """Builds the structural workbook from SchemaRegistry."""

    def __init__(self, registry: SchemaRegistry, settings: Any) -> None:
        self.registry = registry
        self.settings = settings
        self.sheet_builder = SheetBuilder(registry)

    def plan(self) -> list[str]:
        return (
            list(UI_SHEETS)
            + self.registry.names()
            + [RELATIONSHIP_SHEET]
            + [e.sheet_name for e in all_engines()]
        )

    def build(self) -> Path:
        if not getattr(self.settings, "ENABLE_EXCEL_GENERATION", False):
            raise ExcelGenerationDisabledError(
                "Excel generation is disabled. "
                "Set config.ENABLE_EXCEL_GENERATION = True."
            )

        self.settings.ensure_directories()
        output_path: Path = self.settings.output_workbook_path()

        wb = Workbook()
        default = wb.active
        wb.remove(default)

        for table in self.registry.tables:
            ws = wb.create_sheet(title=table.name)
            if table.name == "SYS_WORKBOOK_INFO":
                self._write_sys_workbook_info(ws)
            elif table.name == "SYS_PARAMETER":
                self._write_sys_parameter(ws)
            else:
                self.sheet_builder.write_sheet(ws, table)

        rel_ws = wb.create_sheet(title=RELATIONSHIP_SHEET)
        write_relationships_sheet(rel_ws, self.registry)

        # All calculations — structured references only
        write_all_engines(wb)

        # Production lookup / org seeds (before UI styling)
        if getattr(self.settings, "ENABLE_SEED_DATA", False):
            apply_seed_data(wb)

        apply_pk_named_ranges(wb, self.registry, self.sheet_builder)
        self._add_relationship_named_range(wb)
        self._apply_fk_validations(wb)
        self._protect_technical_sheets(wb)

        # Native Excel UI (theme, dashboard, nav, search, CF, protection)
        if getattr(self.settings, "ENABLE_UI", True):
            UIBuilder(wb, self.registry, self.settings).apply()

        # Phase 9–11 — İnci Akü daily operations UX (no schema changes)
        if getattr(self.settings, "ENABLE_INCI_OPS_UX", True):
            InciOpsUX(wb, self.registry, self.settings).apply()

        # Document engine variable catalog (hidden) + harden protection
        add_document_engine_sheet(wb)
        protect_all_system_sheets(
            wb, getattr(self.settings, "TECH_SHEET_PASSWORD", "PIMS_TECH")
        )

        wb.save(output_path)
        wb.close()
        return output_path

    def _write_keyed_table(
        self,
        ws,
        table_display_name: str,
        headers: list[str],
        rows: list[tuple],
    ) -> Table:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        last_col = excel_col_letter(len(headers))
        last_row = 1 + max(len(rows), 1)
        if not rows:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=2, column=c_idx, value=None)
        excel_table = Table(
            displayName=table_display_name,
            ref=f"A1:{last_col}{last_row}",
        )
        ws.add_table(excel_table)
        return excel_table

    def _write_sys_workbook_info(self, ws) -> None:
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        rows = [
            ("SCHEMA_VERSION", self.settings.SCHEMA_VERSION, generated),
            ("WORKBOOK_NAME", self.settings.DEFAULT_OUTPUT_FILENAME, generated),
            ("PROJECT_NAME", self.settings.PROJECT_NAME, generated),
            (
                "REVISION",
                getattr(self.settings, "WORKBOOK_REVISION", "Rev00"),
                generated,
            ),
            ("GENERATED_AT_UTC", generated, generated),
            ("EXCEL_ENGINE", "openpyxl", generated),
            ("TECH_SHEET_PROTECTION", "ENABLED", generated),
        ]
        self._write_keyed_table(
            ws,
            "T_SYS_WORKBOOK_INFO",
            ["INFO_KEY", "INFO_VALUE", "UPDATED_AT"],
            rows,
        )

    def _write_sys_parameter(self, ws) -> None:
        rows = [
            (
                1,
                "WEIGHT_UOM",
                self.settings.WEIGHT_UOM,
                "Mass unit for COMPONENT.WEIGHT_G",
            ),
            (
                2,
                "TRANSPORT_ALLOCATION_METHOD",
                self.settings.TRANSPORT_ALLOCATION_METHOD,
                "How transport-line weight allocates to product units",
            ),
            (
                3,
                "MATERIAL_SHARE_TOLERANCE_PCT",
                str(self.settings.MATERIAL_SHARE_TOLERANCE_PCT),
                "Allowed deviation from 100% material shares",
            ),
        ]
        self._write_keyed_table(
            ws,
            "T_SYS_PARAMETER",
            ["PARAMETER_ID", "PARAMETER_CODE", "PARAMETER_VALUE", "DESCRIPTION"],
            rows,
        )

    def _add_relationship_named_range(self, wb: Workbook) -> None:
        name = "NR_RELATIONSHIP_ID"
        if name not in wb.defined_names:
            wb.defined_names.add(
                DefinedName(
                    name=name,
                    attr_text=f"'{RELATIONSHIP_SHEET}'!$A$2:$A$1048576",
                )
            )

    def _apply_fk_validations(self, wb: Workbook) -> None:
        for table in self.registry.tables:
            ws = wb[table.name]
            for col in table.columns:
                if not col.is_fk:
                    continue
                nr_name = parent_pk_for_fk(col, self.registry)
                if not nr_name:
                    continue
                col_letter = self.sheet_builder.column_letter(table, col.name)
                dv = DataValidation(
                    type="list",
                    formula1=f"={nr_name}",
                    allow_blank=True,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Invalid FK",
                    error=f"Value must exist in {nr_name}",
                    promptTitle="Foreign Key",
                    prompt=f"Select {col.name} from {nr_name}",
                )
                dv.add(f"{col_letter}2:{col_letter}1048576")
                ws.add_data_validation(dv)

    def _protect_technical_sheets(self, wb: Workbook) -> None:
        password = getattr(self.settings, "TECH_SHEET_PASSWORD", "PIMS_TECH")
        for sheet_name in TECHNICAL_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            ws.sheet_state = "hidden"
            ws.protection.sheet = True
            ws.protection.password = password
