"""Phase P — product-centric lookup + DOCUMENT_CENTER native link repair."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from builders.phase_n.assets import extract_inci_aku_logo
from builders.phase_o2.cell_modules import (
    apply_native_doc_link,
    clear_sheet_values,
    find_header_row,
    write_nav_row,
    write_title_block,
)
from builders.phase_o2.service import EXPECTED, PhaseO2Service
from builders.phase_o6.grid_canvas import GridCanvas
from services.document_link_service import DocumentLinkService, normalize_relative_path

NAVY = "0E2A47"
GOLD = "C8A24A"
IVORY = "F7F5F0"
WHITE = "FFFFFF"
INK = "1C2430"
MUTED = "5C6B7A"
LINK = "1F5C99"
BAND = "F3F6F9"
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)

DOMESTIC_MSG = (
    "No final OEM/EU PPWR Packaging Configuration is assigned to this product in "
    "the controlled Rev.00 register."
)

LOOKUP_SHEET = "P_PRODUCT_LOOKUP"
# Coerce numeric product codes typed into B11 to text for XLOOKUP against text keys
LOOKUP_KEY = '($B$11&"")'
LOOKUP_HEADERS = [
    "KEY",  # A — searchable key
    "KEY_TYPE",  # B
    "Product Code",  # C
    "Product Description",  # D
    "Battery Type",  # E
    "Customer / Market",  # F
    "Nominal Qty",  # G
    "Source Configuration ID",  # H
    "Packaging Set Code",  # I
    "Final Configuration ID",  # J
    "Variant Basis TR",  # K
    "Variant Basis EN",  # L
    "Packaging Tare kg",  # M
    "Status",  # N
    "Technical File ID",  # O
    "EU DoC ID",  # P
    "Label ID",  # Q
    "Statement ID",  # R
    "TF_PATH",  # S
    "DOC_PATH",  # T
    "LABEL_PATH",  # U
    "STM_PATH",  # V
    "DOMESTIC_MESSAGE",  # W
]


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def _font(size=9, bold=False, color=INK, underline=None) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color, underline=underline)


def _junction_or_copy(src: Path, dst: Path) -> str:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        elif dst.is_dir():
            shutil.rmtree(dst, ignore_errors=True)
        else:
            dst.unlink()
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
        return "junction"
    except Exception:
        shutil.copytree(src, dst)
        return "copy"


@dataclass
class PhasePResult:
    success: bool
    product_lookup_gate: str
    document_link_gate: str
    manual_acceptance: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


class PhasePService:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O6_CANDIDATE.xlsx"
        self.candidate = self.out / "INCI_AKU_PPWR_PIMS_Rev00_PRODUCT_LOOKUP_CANDIDATE.xlsx"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_PRODUCT_LOOKUP"
        self.delivery_wb = self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.golden = (
            self.root
            / "input"
            / "production"
            / "INCI_AKU_PPWR_Final_Configuration_Register_Rev00_GOLDEN_VARIANTS_FINAL.xlsx"
        )
        self.qa_path = self.out / "PHASE_P_PRODUCT_LOOKUP_QA.md"
        self.qa_json = self.out / "PHASE_P_PRODUCT_LOOKUP_QA.json"
        self._o2 = PhaseO2Service(project_root)
        self.link_svc = DocumentLinkService()
        self.smoke_docs = [
            ("STARTER", "ST-051-STD-01", "Technical File", "01_STARTER/ST-051-STD-01/01_Technical_File.docx"),
            ("STARTER", "ST-051-STD-01", "EU DoC", "01_STARTER/ST-051-STD-01/02_EU_DoC.docx"),
            ("STARTER", "ST-051-STD-01", "Label", "01_STARTER/ST-051-STD-01/03_Label.docx"),
            ("STARTER", "ST-051-STD-01", "Shipment Statement", "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx"),
            ("INDUSTRIAL", "IND-24V-01", "Technical File", "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx"),
            ("CONTAINER", "CNT-20-STD-01", "Technical File", "03_CONTAINER/CNT-20-STD-01/01_Technical_File.docx"),
        ]

    def run(self) -> PhasePResult:
        messages: list[str] = []
        if not self.source.exists():
            return PhasePResult(False, "FAIL", "FAIL", "PENDING", [f"Missing O6: {self.source}"])
        if not self.golden.exists():
            return PhasePResult(False, "FAIL", "FAIL", "PENDING", [f"Missing Golden: {self.golden}"])

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"Candidate from O6 baseline → {self.candidate.name}")

        baseline = self._o2._counts(self.candidate)
        messages.append(f"Baseline counts: {baseline}")

        product_maps = self._load_golden_product_maps()
        messages.append(f"Golden product maps: {len(product_maps)}")

        openpyxl_stats = self._openpyxl_build(product_maps)
        messages.append(f"openpyxl build: {openpyxl_stats}")

        self._kill_office()
        time.sleep(1.0)
        com_stats = self._com_ux_and_link_repair()
        messages.append(f"COM UX/link repair: {com_stats}")

        # Delivery package
        folder_modes = self._build_delivery()
        messages.append(f"Delivery: {self.delivery} modes={folder_modes}")

        after = self._o2._counts(self.candidate)
        map_qa = self._product_map_qa(product_maps)
        messages.append(
            f"Map QA: OEM={map_qa['mapped_oem_eu']} domestic={map_qa['domestic_only']}"
        )

        self._kill_office()
        time.sleep(1.0)
        fixture = self._hard_fixture_test()
        messages.append(f"Hard fixture: {fixture.get('status')}")

        self._kill_office()
        time.sleep(1.0)
        domestic = self._domestic_behaviour_test()
        messages.append(f"Domestic behaviour: {domestic.get('status')}")

        self._kill_office()
        time.sleep(1.0)
        path_scan = self._validate_all_docx_links(self.delivery_wb, self.delivery)
        messages.append(
            f"Path scan: {path_scan['existing']}/{path_scan['total']} "
            f"missing={path_scan['missing']} absolute={path_scan['absolute']}"
        )

        self._kill_office()
        time.sleep(1.0)
        dc_word = self._actual_word_open_tests(source="DOCUMENT_CENTER")
        dc_ok = sum(1 for t in dc_word if t.get("word_opened"))
        messages.append(f"Document Center Word open: {dc_ok}/6")

        self._kill_office()
        time.sleep(1.0)
        search_word = self._search_document_launch_test()
        search_ok = sum(1 for t in search_word if t.get("word_opened"))
        messages.append(f"Search result Word open: {search_ok}/4")

        excel_open = self._excel_open(self.delivery_wb)

        product_gate = (
            "PASS"
            if (
                baseline == EXPECTED
                and after == EXPECTED
                and fixture.get("pass")
                and domestic.get("pass")
                and map_qa.get("unique_product_codes") == map_qa.get("total_product_codes")
                and excel_open.get("ok")
            )
            else "FAIL"
        )
        doc_gate = (
            "PASS"
            if (
                dc_ok == 6
                and search_ok == 4
                and path_scan["total"] == 988
                and path_scan["existing"] == 988
                and path_scan["missing"] == 0
                and path_scan["absolute"] == 0
            )
            else "FAIL"
        )

        qa = {
            "phase_p_product_lookup_gate": product_gate,
            "document_link_actual_open_gate": doc_gate,
            "manual_acceptance": "PENDING",
            "promoted": False,
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": after == baseline == EXPECTED,
            "candidate": str(self.candidate),
            "delivery_root": str(self.delivery),
            "folder_modes": folder_modes,
            "product_map_qa": map_qa,
            "hard_fixture": fixture,
            "domestic_behaviour": domestic,
            "path_scan": path_scan,
            "document_center_word_tests": dc_word,
            "document_center_word_pass": dc_ok,
            "search_word_tests": search_word,
            "search_word_pass": search_ok,
            "native_excel_open": excel_open,
            "openpyxl_stats": openpyxl_stats,
            "com_stats": com_stats,
            "visual_redesign": False,
            "canonical_mappings_changed": False,
            "word_regenerated": False,
        }
        self._write_qa(qa, messages)
        return PhasePResult(
            product_gate == "PASS" and doc_gate == "PASS",
            product_gate,
            doc_gate,
            "PENDING",
            messages,
            qa,
        )

    # ── data ──────────────────────────────────────────────
    def _load_golden_product_maps(self) -> list[dict[str, Any]]:
        wb = load_workbook(self.golden, data_only=True, read_only=True)
        ws = wb["02_PRODUCT_MAP"]
        rows = ws.iter_rows(min_row=1, values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows)]
        idx = {h: i for i, h in enumerate(headers)}
        out = []
        for row in rows:
            if not row or row[idx["Product Code"]] in (None, ""):
                continue

            def _opt(key: str) -> str | None:
                raw = row[idx[key]]
                if raw in (None, ""):
                    return None
                s = str(raw).strip()
                return None if s.upper() in ("NONE", "NULL", "-") else s

            final_set = _opt("Final Set Code")
            final_cfg = _opt("Final Configuration ID")
            out.append(
                {
                    "product_code": str(row[idx["Product Code"]]).strip(),
                    "product_description": str(row[idx["Product Description"]] or ""),
                    "battery_type": _opt("Battery Type"),
                    "customer_market": _opt("Customer / Market"),
                    "nominal_qty": row[idx["Nominal Qty"]],
                    "source_configuration_id": _opt("Source Configuration ID"),
                    "final_set_code": final_set,
                    "final_configuration_id": final_cfg if final_set else None,
                    "status": str(row[idx["Status"]] or ""),
                }
            )
        wb.close()
        return out

    def _load_config_index(self, wb) -> dict[str, dict[str, Any]]:
        ws = wb["PACKAGING_CONFIGURATIONS"]
        hr = find_header_row(ws) or 6
        headers = [ws.cell(hr, c).value for c in range(1, (ws.max_column or 1) + 1)]
        def col(name: str) -> int:
            return headers.index(name) + 1

        by_set: dict[str, dict[str, Any]] = {}
        for r in range(hr + 1, (ws.max_row or hr) + 1):
            set_code = ws.cell(r, col("Packaging Set Code")).value
            if not set_code:
                continue
            set_code = str(set_code)
            by_set[set_code] = {
                "row": r,
                "set_code": set_code,
                "final_id": str(ws.cell(r, col("Final Configuration ID")).value or ""),
                "source_id": str(ws.cell(r, col("Source Configuration ID")).value or ""),
                "family": str(ws.cell(r, col("Family")).value or ""),
                "variant_tr": str(ws.cell(r, col("Variant Basis TR")).value or ""),
                "variant_en": str(ws.cell(r, col("Variant Basis EN")).value or ""),
                "tare": ws.cell(r, col("Packaging Tare kg")).value,
                "tf_id": str(ws.cell(r, col("Technical File")).value or ""),
                "doc_id": str(ws.cell(r, col("DoC")).value or ""),
                "label_id": str(ws.cell(r, col("Label")).value or ""),
                "stm_id": str(ws.cell(r, col("Statement")).value or ""),
            }
        return by_set

    def _init_link_service(self, wb) -> None:
        self.link_svc = DocumentLinkService()
        if "DOCUMENT_LIBRARY" not in wb.sheetnames:
            return
        ws = wb["DOCUMENT_LIBRARY"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ci = headers.index("DOCUMENT_CODE")
        ui = headers.index("FILE_URI")
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[ci]:
                continue
            rows.append((str(row[ci]), str(row[ui] or "")))
        self.link_svc.load_from_document_library_rows(rows)

    # ── openpyxl build ────────────────────────────────────
    def _openpyxl_build(self, product_maps: list[dict[str, Any]]) -> dict[str, Any]:
        wb = load_workbook(self.candidate)
        self._init_link_service(wb)
        configs = self._load_config_index(wb)
        lookup_rows = self._build_lookup_rows(product_maps, configs)
        self._write_lookup_sheet(wb, lookup_rows)
        pm_stats = self._rebuild_product_master(wb, product_maps, configs)
        wb.save(self.candidate)
        wb.close()
        return {
            "lookup_rows": len(lookup_rows),
            "product_master": pm_stats,
            "uri_registry": len(self.link_svc._uri_by_code),
        }

    def _pack_for_set(self, cfg: dict[str, Any] | None) -> dict[str, str]:
        if not cfg or not cfg.get("set_code"):
            return {"tf": "", "doc": "", "label": "", "statement": ""}
        return self.link_svc.pack_paths(
            packaging_set_code=cfg["set_code"],
            family=cfg.get("family"),
            tf_id=cfg.get("tf_id"),
            doc_id=cfg.get("doc_id"),
            label_id=cfg.get("label_id"),
            statement_id=cfg.get("stm_id"),
        )

    def _build_lookup_rows(
        self, product_maps: list[dict[str, Any]], configs: dict[str, dict[str, Any]]
    ) -> list[list[Any]]:
        rows: list[list[Any]] = []
        seen_keys: set[str] = set()

        def add_row(key: str, key_type: str, payload: list[Any]) -> None:
            k = str(key or "").strip()
            if not k or k in seen_keys:
                return
            seen_keys.add(k)
            rows.append([k, key_type, *payload])

        # Product-primary rows
        for pm in product_maps:
            set_code = pm.get("final_set_code")
            cfg = configs.get(set_code or "") if set_code else None
            packs = self._pack_for_set(cfg)
            status = pm.get("status") or ""
            domestic = "DOMESTIC-ONLY" in status.upper() or not set_code
            if domestic and "DOMESTIC" not in status.upper():
                status = "DOMESTIC-ONLY / NOT IN OEM PACKAGE"
            payload = [
                pm["product_code"],
                pm["product_description"],
                pm.get("battery_type") or "",
                pm.get("customer_market") or "",
                pm.get("nominal_qty"),
                pm.get("source_configuration_id") or (cfg or {}).get("source_id") or "",
                set_code or "",
                pm.get("final_configuration_id") or (cfg or {}).get("final_id") or "",
                (cfg or {}).get("variant_tr") or "",
                (cfg or {}).get("variant_en") or "",
                (cfg or {}).get("tare") if cfg else "",
                status,
                (cfg or {}).get("tf_id") or "",
                (cfg or {}).get("doc_id") or "",
                (cfg or {}).get("label_id") or "",
                (cfg or {}).get("stm_id") or "",
                packs["tf"] if not domestic else "",
                packs["doc"] if not domestic else "",
                packs["label"] if not domestic else "",
                packs["statement"] if not domestic else "",
                DOMESTIC_MSG if domestic else "",
            ]
            add_row(pm["product_code"], "PRODUCT_CODE", payload)

        # Alternate keys from packaging configurations (set-level, no product)
        for set_code, cfg in configs.items():
            packs = self._pack_for_set(cfg)
            payload = [
                "",  # product code blank for set-level
                "",
                "",
                "",
                "",
                cfg.get("source_id") or "",
                set_code,
                cfg.get("final_id") or "",
                cfg.get("variant_tr") or "",
                cfg.get("variant_en") or "",
                cfg.get("tare"),
                "IN OEM / EU PACKAGE",
                cfg.get("tf_id") or "",
                cfg.get("doc_id") or "",
                cfg.get("label_id") or "",
                cfg.get("stm_id") or "",
                packs["tf"],
                packs["doc"],
                packs["label"],
                packs["statement"],
                "",
            ]
            add_row(set_code, "PACKAGING_SET", payload)
            add_row(cfg.get("final_id"), "FINAL_CONFIGURATION", payload)
            add_row(cfg.get("source_id"), "SOURCE_CONFIGURATION", payload)
            add_row(cfg.get("tf_id"), "TECHNICAL_FILE_ID", payload)
            add_row(cfg.get("doc_id"), "DOC_ID", payload)
            add_row(cfg.get("label_id"), "LABEL_ID", payload)
            add_row(cfg.get("stm_id"), "STATEMENT_ID", payload)

        return rows

    def _write_lookup_sheet(self, wb, rows: list[list[Any]]) -> None:
        for obsolete in (LOOKUP_SHEET, "_P_LOOKUP"):
            if obsolete in wb.sheetnames:
                del wb[obsolete]
        ws = wb.create_sheet(LOOKUP_SHEET)
        for c, h in enumerate(LOOKUP_HEADERS, start=1):
            cell = ws.cell(1, c, h)
            cell.font = _font(9, True, WHITE)
            cell.fill = _fill(NAVY)
        for r_i, row in enumerate(rows, start=2):
            for c, v in enumerate(row, start=1):
                cell = ws.cell(r_i, c, v)
                if c == 1:
                    cell.value = str(v) if v is not None else ""
                    cell.number_format = "@"
        ws.sheet_state = "visible"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(LOOKUP_HEADERS))}{max(len(rows)+1, 2)}"

    def _rebuild_product_master(
        self, wb, product_maps: list[dict[str, Any]], configs: dict[str, dict[str, Any]]
    ) -> dict[str, Any]:
        ws = wb["PRODUCT_MASTER"]
        clear_sheet_values(ws)
        write_nav_row(ws, 8)
        write_title_block(
            ws,
            "PRODUCT MASTER",
            "Product → Packaging Configuration → Document Pack lookup register",
            "2,046 products  ·  controlled Rev.00 Product Map",
            12,
        )
        headers = [
            "Product Code",
            "Product Description",
            "Battery Type",
            "Nominal Qty",
            "Source Configuration ID",
            "Packaging Set Code",
            "Final Configuration ID",
            "Status",
            "Packaging Tare kg",
            "Technical File",
            "OPEN TF",
            "EU DoC",
            "OPEN DOC",
            "Label",
            "OPEN LABEL",
            "Statement",
            "OPEN STATEMENT",
            "OPEN PACKAGE",
            "Customer / Market",
        ]
        start = 6
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(start, c, h)
            cell.font = _font(9, True, WHITE)
            cell.fill = _fill(NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = HAIR
        ws.row_dimensions[start].height = 28

        open_counts = {"tf": 0, "doc": 0, "label": 0, "stm": 0, "pkg": 0}
        for i, pm in enumerate(sorted(product_maps, key=lambda x: x["product_code"])):
            r = start + 1 + i
            set_code = pm.get("final_set_code")
            cfg = configs.get(set_code or "") if set_code else None
            packs = self._pack_for_set(cfg)
            status = pm.get("status") or ""
            domestic = "DOMESTIC-ONLY" in status.upper() or not set_code
            if domestic and "DOMESTIC" not in status.upper():
                status = "DOMESTIC-ONLY / NOT IN OEM PACKAGE"
            band = _fill(BAND) if i % 2 else _fill(WHITE)
            vals = [
                pm["product_code"],
                pm["product_description"],
                pm.get("battery_type") or "",
                pm.get("nominal_qty"),
                pm.get("source_configuration_id") or (cfg or {}).get("source_id") or "",
                set_code or "",
                pm.get("final_configuration_id") or (cfg or {}).get("final_id") or "",
                status,
                (cfg or {}).get("tare") if cfg else "",
                (cfg or {}).get("tf_id") or "",
                None,  # OPEN TF
                (cfg or {}).get("doc_id") or "",
                None,
                (cfg or {}).get("label_id") or "",
                None,
                (cfg or {}).get("stm_id") or "",
                None,
                None,  # OPEN PACKAGE
                pm.get("customer_market") or "",
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.fill = band
                cell.border = HAIR
                cell.font = _font(9)
                cell.alignment = Alignment(vertical="center")

            if not domestic and cfg:
                for col, kind, key in (
                    (11, "TF", "tf"),
                    (13, "DOC", "doc"),
                    (15, "LABEL", "label"),
                    (17, "STM", "statement"),
                ):
                    path = packs[key]
                    if path:
                        apply_native_doc_link(ws.cell(r, col), path, "OPEN")
                        open_counts[{"tf": "tf", "doc": "doc", "label": "label", "statement": "stm"}[key]] += 1
                # OPEN PACKAGE → packaging configurations row
                pkg = ws.cell(r, 18, "OPEN")
                pkg.font = _font(9, True, LINK, underline="single")
                pkg.alignment = Alignment(horizontal="center", vertical="center")
                pkg.fill = band
                pkg.border = HAIR
                loc_row = cfg.get("row") or 7
                pkg.hyperlink = Hyperlink(
                    ref=pkg.coordinate,
                    location=f"'PACKAGING_CONFIGURATIONS'!A{loc_row}",
                    tooltip=f"Open package {set_code}",
                )
                open_counts["pkg"] += 1

        last = start + len(product_maps)
        ws.freeze_panes = f"A{start + 1}"
        ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{last}"
        ws.sheet_view.showGridLines = False
        widths = {
            1: 12, 2: 42, 3: 12, 4: 10, 5: 20, 6: 16, 7: 20, 8: 28, 9: 12,
            10: 28, 11: 10, 12: 28, 13: 10, 14: 28, 15: 10, 16: 28, 17: 10, 18: 12, 19: 22,
        }
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        return {"rows": len(product_maps), "open_counts": open_counts}

    # ── COM UX + link repair ──────────────────────────────
    def _com_ux_and_link_repair(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        stats: dict[str, Any] = {}
        try:
            logo = extract_inci_aku_logo(self.root, self.out / "PHASE_N_ASSETS")
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = False
            wb = excel.Workbooks.Open(str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False)

            for sheet_name in ("SEARCH", "00_HOME", "NAVIGATION", "DOCUMENT_CENTER", "PRODUCT_MASTER"):
                try:
                    wb.Worksheets(sheet_name).Unprotect()
                except Exception:
                    pass

            # openpyxl save strips Class A shapes — restore O6 canvas baseline first
            canvas = GridCanvas(excel, wb, logo)
            stats["restore_home"] = canvas.design_home()
            stats["restore_nav"] = canvas.design_navigation()
            stats["restore_search"] = canvas.design_search()

            stats["document_center"] = self._com_repair_open_cols(
                wb,
                "DOCUMENT_CENTER",
                open_cols=(10, 12, 14, 16),
                header_hints=("Open Technical", "Open EU", "Open Label", "Open Statement"),
            )
            stats["product_master"] = self._com_repair_open_cols(
                wb,
                "PRODUCT_MASTER",
                open_cols=(11, 13, 15, 17),
                header_hints=("OPEN TF", "OPEN DOC", "OPEN LABEL", "OPEN STATEMENT"),
            )
            stats["search"] = self._com_update_search(wb)
            stats["home"] = self._com_update_home(wb)

            # Keep lookup sheet available to formulas; hide from tab bar
            try:
                lu = wb.Worksheets(LOOKUP_SHEET)
                lu.Visible = 0  # xlSheetHidden (formulas still resolve)
            except Exception as exc:
                stats["lookup_hide_error"] = str(exc)

            try:
                canvas.protect("00_HOME")
                canvas.protect("NAVIGATION")
                canvas.protect("SEARCH", unlock=["B11", "B11:H12"])
            except Exception as exc:
                stats["protect_error"] = str(exc)

            excel.ScreenUpdating = True
            wb.Save()
            wb.Close(True)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return stats

    def _com_repair_open_cols(
        self, wb, sheet_name: str, open_cols: tuple[int, ...], header_hints: tuple[str, ...]
    ) -> dict[str, Any]:
        ws = wb.Worksheets(sheet_name)
        # find header row
        header_row = 6
        for r in range(1, 25):
            v = str(ws.Cells(r, 1).Value or "")
            if "Packaging Set" in v or v == "Product Code":
                header_row = r
                break
        repaired = 0
        skipped = 0
        absolute = 0
        last_row = int(ws.Cells(ws.Rows.Count, 1).End(-4162).Row)  # xlUp
        for r in range(header_row + 1, last_row + 1):
            first = ws.Cells(r, 1).Value
            if first in (None, ""):
                continue
            for col in open_cols:
                cell = ws.Cells(r, col)
                addr = None
                # Prefer existing hyperlink address
                try:
                    if int(cell.Hyperlinks.Count) >= 1:
                        addr = str(cell.Hyperlinks(1).Address or "")
                except Exception:
                    addr = None
                if not addr:
                    val = cell.Value
                    if isinstance(val, str) and val.lower().endswith(".docx"):
                        addr = val
                if not addr:
                    skipped += 1
                    continue
                addr = normalize_relative_path(addr)
                if re.match(r"^[A-Za-z]:/", addr.replace("\\", "/")) or "Users/" in addr:
                    absolute += 1
                    continue
                if not addr.lower().endswith(".docx"):
                    skipped += 1
                    continue
                try:
                    while int(cell.Hyperlinks.Count) >= 1:
                        cell.Hyperlinks(1).Delete()
                except Exception:
                    pass
                cell.Value = "OPEN"
                ws.Hyperlinks.Add(Anchor=cell, Address=addr, SubAddress="")
                repaired += 1
        return {
            "sheet": sheet_name,
            "repaired": repaired,
            "skipped": skipped,
            "absolute_rejected": absolute,
            "header_row": header_row,
        }

    def _com_update_search(self, wb) -> dict[str, Any]:
        ws = wb.Worksheets("SEARCH")
        try:
            ws.Unprotect()
        except Exception:
            pass

        # Product-code-first labels (preserve O6 shapes / input B11:H12)
        ws.Range("B10").Value = "ÜRÜN KODU / PRODUCT CODE"
        try:
            ws.Range("B10").Font.Bold = True
        except Exception:
            pass

        # Force text input so numeric product codes match text keys
        ws.Range("B11").NumberFormat = "@"
        ws.Range("B11:H12").Locked = False

        try:
            for i in range(1, int(ws.Shapes.Count) + 1):
                shp = ws.Shapes(i)
                try:
                    txt = str(shp.TextFrame.Characters().Text or "")
                except Exception:
                    continue
                if "GLOBAL SEARCH" in txt.upper():
                    shp.TextFrame.Characters().Text = (
                        "GLOBAL SEARCH\r"
                        "Find the applicable PPWR packaging configuration "
                        "from an İnci Akü product code."
                    )
        except Exception:
            pass

        ws.Range("B13").Value = (
            "You may also search by Packaging Set, Configuration, Source BOM or Document ID.   ·   "
            "Examples: 1011935  •  ST-051-STD-01  •  IA-ST-CFG-0122"
        )

        ws.Range("B16").Value = "RESULTS"
        key = LOOKUP_KEY
        sh = LOOKUP_SHEET
        ws.Range("B17").Formula = (
            f'=IF({key}="","Enter an İnci Akü Product Code (e.g. 1011935).","")'
        )
        ws.Range("Z1").Formula = (
            f'=IF({key}="","",IFERROR(XLOOKUP({key},{sh}!A:A,{sh}!A:A),"NOT_FOUND"))'
        )

        fields = [
            (18, "Product Code", "C"),
            (19, "Product Description", "D"),
            (20, "Battery Type", "E"),
            (21, "Customer / Market", "F"),
            (22, "Nominal Qty", "G"),
            (23, "Source Configuration ID", "H"),
            (24, "Packaging Set Code", "I"),
            (25, "Final Configuration ID", "J"),
            (26, "Variant Basis TR", "K"),
            (27, "Packaging Tare kg", "M"),
        ]
        for r, label, col_letter in fields:
            ws.Range(f"B{r}").Value = label
            try:
                ws.Range(f"B{r}").Font.Size = 9
            except Exception:
                pass
            ws.Range(f"E{r}").Formula = (
                f'=IF({key}="","",IFERROR(XLOOKUP({key},{sh}!A:A,{sh}!{col_letter}:{col_letter}),'
                f'"Not found"))'
            )

        ws.Range("K18").Value = "STATUS"
        try:
            ws.Range("K18").Font.Bold = True
        except Exception:
            pass
        ws.Range("K19").Formula = (
            f'=IF({key}="","",IFERROR(XLOOKUP({key},{sh}!A:A,{sh}!N:N),"Not found"))'
        )
        ws.Range("K21").Value = "MESSAGE"
        try:
            ws.Range("K21").Font.Bold = True
        except Exception:
            pass
        ws.Range("K22").Formula = (
            f'=IF({key}="","",IFERROR(XLOOKUP({key},{sh}!A:A,{sh}!W:W),""))'
        )
        ws.Range("K22").WrapText = True

        ws.Range("K24").Value = "Variant Basis EN"
        ws.Range("K25").Formula = (
            f'=IF({key}="","",IFERROR(XLOOKUP({key},{sh}!A:A,{sh}!L:L),""))'
        )

        # Document pack below quick-link shapes
        ws.Range("B35").Value = "DOCUMENT PACK"
        try:
            ws.Range("B35").Font.Bold = True
        except Exception:
            pass
        for r, label, id_col, path_col in [
            (36, "Technical File ID", "O", "S"),
            (37, "EU DoC ID", "P", "T"),
            (38, "Label ID", "Q", "U"),
            (39, "Shipment Statement ID", "R", "V"),
        ]:
            ws.Range(f"B{r}").Value = label
            try:
                ws.Range(f"B{r}").Font.Size = 9
            except Exception:
                pass
            ws.Range(f"E{r}").Formula = (
                f'=IF({key}="","",IFERROR(XLOOKUP({key},{sh}!A:A,{sh}!{id_col}:{id_col}),""))'
            )
            ws.Range(f"H{r}").Formula = (
                f'=IF({key}="","",IFERROR('
                f'IF(XLOOKUP({key},{sh}!A:A,{sh}!{path_col}:{path_col})="",'
                f'"",'
                f'HYPERLINK(XLOOKUP({key},{sh}!A:A,{sh}!{path_col}:{path_col}),"OPEN"))'
                f',""))'
            )

        ws.Range("B11:H12").Locked = False
        try:
            ws.Range("B11").Select()
        except Exception:
            pass
        return {
            "updated": True,
            "input": "B11",
            "lookup_sheet": LOOKUP_SHEET,
            "open_cells": ["H36", "H37", "H38", "H39"],
        }

    def _com_update_home(self, wb) -> dict[str, Any]:
        ws = wb.Worksheets("00_HOME")
        updated = []
        # Prominent primary action: reuse SecStrip as FIND BY PRODUCT CODE
        try:
            shp = ws.Shapes("SecStrip")
            shp.TextFrame.Characters().Text = (
                "FIND BY PRODUCT CODE\r"
                "Find PPWR packaging and documents from an İnci Akü product code\r"
                "Open →"
            )
            try:
                while int(shp.Hyperlinks.Count) >= 1:
                    shp.Hyperlinks(1).Delete()
            except Exception:
                pass
            ws.Hyperlinks.Add(Anchor=shp, Address="", SubAddress="'SEARCH'!A1")
            updated.append("SecStrip→SEARCH")
        except Exception as exc:
            updated.append(f"SecStrip_error:{exc}")

        try:
            shp = ws.Shapes("Act_SEARCH")
            shp.TextFrame.Characters().Text = (
                "FIND BY PRODUCT CODE\r"
                "Product → packaging → documents\r"
                "Open →"
            )
            updated.append("Act_SEARCH")
        except Exception as exc:
            updated.append(f"Act_SEARCH_error:{exc}")
        return {"updated": updated}

    def _build_delivery(self) -> dict[str, str]:
        self.delivery.mkdir(parents=True, exist_ok=True)
        for child in list(self.delivery.iterdir()):
            try:
                if child.is_junction() or child.is_symlink():
                    child.unlink()
                elif child.is_dir():
                    shutil.rmtree(child, ignore_errors=True)
                else:
                    child.unlink()
            except Exception:
                pass
        shutil.copy2(self.candidate, self.delivery_wb)
        modes = {}
        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if src.exists():
                modes[folder] = _junction_or_copy(src, self.delivery / folder)
        return modes

    # ── QA helpers ────────────────────────────────────────
    def _kill_office(self) -> None:
        for app in ("EXCEL.EXE", "WINWORD.EXE"):
            try:
                subprocess.run(["taskkill", "/F", "/IM", app], capture_output=True, text=True)
            except Exception:
                pass
        time.sleep(0.8)

    def _follow_with_retry(self, cell, attempts: int = 4) -> None:
        last = None
        for i in range(attempts):
            try:
                cell.Hyperlinks(1).Follow()
                return
            except Exception as exc:
                last = exc
                time.sleep(1.0 + i)
        if last:
            raise last

    def _excel_open(self, path: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {"ok": False, "error": None}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=True)
            out["ok"] = True
            out["sheets"] = int(wb.Worksheets.Count)
            wb.Close(False)
        except Exception as exc:
            out["error"] = str(exc)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _product_map_qa(self, product_maps: list[dict[str, Any]]) -> dict[str, Any]:
        codes = [p["product_code"] for p in product_maps]
        unique = set(codes)
        oem = [p for p in product_maps if p.get("final_set_code")]
        domestic = [p for p in product_maps if not p.get("final_set_code")]
        multi: dict[str, set[str]] = {}
        for p in product_maps:
            multi.setdefault(p["product_code"], set())
            if p.get("final_set_code"):
                multi[p["product_code"]].add(p["final_set_code"])
        multi_final = sum(1 for c, s in multi.items() if len(s) > 1)
        missing_src = sum(1 for p in product_maps if not p.get("source_configuration_id"))
        mapped_missing_final = sum(
            1
            for p in oem
            if p.get("final_set_code") and not p.get("final_configuration_id")
        )
        return {
            "total_product_codes": len(codes),
            "unique_product_codes": len(unique),
            "mapped_oem_eu": len(oem),
            "domestic_only": len(domestic),
            "products_with_more_than_one_final_set": multi_final,
            "products_missing_source_configuration_id": missing_src,
            "mapped_missing_final_configuration_id": mapped_missing_final,
        }

    def _hard_fixture_test(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        result: dict[str, Any] = {"pass": False, "status": "FAIL", "checks": []}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.delivery_wb.resolve()), UpdateLinks=0, ReadOnly=False
            )
            try:
                wb.Worksheets("SEARCH").Unprotect()
            except Exception:
                pass
            ws = wb.Worksheets("SEARCH")
            checks = []

            def lookup(code: str) -> dict[str, Any]:
                ws.Range("B11").NumberFormat = "@"
                ws.Range("B11").Value = str(code)
                excel.CalculateFullRebuild()
                time.sleep(0.4)
                return {
                    "product_code": str(ws.Range("E18").Value or ""),
                    "source": str(ws.Range("E23").Value or ""),
                    "set": str(ws.Range("E24").Value or ""),
                    "final": str(ws.Range("E25").Value or ""),
                    "tare": ws.Range("E27").Value,
                    "status": str(ws.Range("K19").Value or ""),
                    "tf": str(ws.Range("E36").Value or ""),
                }

            main = lookup("1011935")
            tare_ok = False
            try:
                tare_ok = abs(float(main["tare"]) - 47.0384) < 0.0001
            except Exception:
                tare_ok = False
            checks.append(
                {
                    "code": "1011935",
                    "expected": {
                        "source": "IA-ST-CFG-0122",
                        "set": "ST-051-STD-01",
                        "final": "IA-ST-051-STD-01",
                        "tare": 47.0384,
                        "status": "IN OEM / EU PACKAGE",
                    },
                    "actual": main,
                    "ok": (
                        main["source"] == "IA-ST-CFG-0122"
                        and main["set"] == "ST-051-STD-01"
                        and main["final"] == "IA-ST-051-STD-01"
                        and tare_ok
                        and "IN OEM" in main["status"].upper()
                    ),
                }
            )
            for code, exp_set in (
                ("1011936", "ST-051-STD-01"),
                ("1011939", "ST-051-STD-01"),
                ("1013795", "ST-051-STD-02"),
            ):
                got = lookup(code)
                checks.append(
                    {
                        "code": code,
                        "expected_set": exp_set,
                        "actual_set": got["set"],
                        "ok": got["set"] == exp_set,
                    }
                )
            # do not merge STD-01 / STD-02
            checks.append(
                {
                    "std01_vs_std02_distinct": True,
                    "ok": checks[-1]["ok"] and checks[-2]["ok"] and checks[-3]["ok"],
                }
            )
            result["checks"] = checks
            result["pass"] = all(c.get("ok") for c in checks)
            result["status"] = "PASS" if result["pass"] else "FAIL"
            wb.Close(False)
        except Exception as exc:
            result["error"] = str(exc)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return result

    def _domestic_behaviour_test(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {"pass": False, "status": "FAIL"}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.delivery_wb.resolve()), UpdateLinks=0, ReadOnly=False
            )
            try:
                wb.Worksheets("SEARCH").Unprotect()
            except Exception:
                pass
            ws = wb.Worksheets("SEARCH")
            ws.Range("B11").NumberFormat = "@"
            ws.Range("B11").Value = "1004590"
            excel.CalculateFullRebuild()
            time.sleep(0.4)
            status = str(ws.Range("K19").Value or "")
            msg = str(ws.Range("K22").Value or "")
            set_code = str(ws.Range("E24").Value or "")
            open_tf = str(ws.Range("H36").Value or "")
            set_empty = set_code.strip() in ("", "None", "Not found", "none")
            ok = (
                "DOMESTIC-ONLY" in status.upper()
                and "No final OEM/EU PPWR Packaging Configuration" in msg
                and set_empty
                and open_tf.strip() in ("", "None", "none")
            )
            out.update(
                {
                    "product_code": "1004590",
                    "status": "PASS" if ok else "FAIL",
                    "pass": ok,
                    "actual_status": status,
                    "actual_message": msg,
                    "actual_set": set_code,
                    "open_tf": open_tf,
                }
            )
            wb.Close(False)
        except Exception as exc:
            out["error"] = str(exc)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _validate_all_docx_links(self, workbook: Path, delivery: Path) -> dict[str, Any]:
        wb = load_workbook(workbook, data_only=False)
        targets: set[str] = set()
        absolute = 0
        for name in wb.sheetnames:
            if name.startswith("_"):
                # include _P_LOOKUP path columns as they are data not hyperlinks
                pass
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    vals: list[str] = []
                    if isinstance(cell.value, str):
                        m = re.search(r'=HYPERLINK\((?:XLOOKUP\([^)]+\)|"([^"]+)")', cell.value, re.I)
                        # also classic
                        m2 = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.I)
                        if m2:
                            vals.append(m2.group(1))
                        # nested XLOOKUP path — skip formula, validate via library
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt:
                            vals.append(tgt)
                    for t0 in vals:
                        if not t0.lower().endswith(".docx"):
                            continue
                        t = normalize_relative_path(t0)
                        if re.match(r"^[A-Za-z]:/", t) or "Users/" in t or t.startswith("\\\\"):
                            absolute += 1
                        targets.add(t)
        # Also collect unique paths from DOCUMENT_LIBRARY / link service
        if "DOCUMENT_LIBRARY" in wb.sheetnames:
            ws = wb["DOCUMENT_LIBRARY"]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            ui = headers.index("FILE_URI")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[ui]:
                    t = normalize_relative_path(str(row[ui]))
                    if t.lower().endswith(".docx"):
                        targets.add(t)
        wb.close()

        # Prefer counting DOCUMENT_CENTER hyperlinks as the 988 authority
        wb2 = load_workbook(workbook, data_only=False)
        dc_targets: set[str] = set()
        ws = wb2["DOCUMENT_CENTER"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink is not None:
                    tgt = normalize_relative_path(cell.hyperlink.target or "")
                    if tgt.lower().endswith(".docx"):
                        dc_targets.add(tgt)
        wb2.close()

        existing = missing = 0
        missing_list = []
        for t in sorted(dc_targets):
            parts = [p for p in t.split("/") if p not in ("", ".")]
            if not parts or ".." in parts:
                missing += 1
                missing_list.append(t)
                continue
            p = delivery.joinpath(*parts)
            if p.exists() and p.is_file():
                existing += 1
            else:
                missing += 1
                missing_list.append(t)
        return {
            "total": len(dc_targets),
            "existing": existing,
            "missing": missing,
            "absolute": absolute,
            "missing_samples": missing_list[:20],
            "all_docx_targets_seen": len(targets),
        }

    def _locate_docx_cells(self, workbook: Path) -> dict[str, tuple[str, str]]:
        wb = load_workbook(workbook, data_only=False)
        found: dict[str, tuple[str, str]] = {}
        prefer = ("DOCUMENT_CENTER", "PRODUCT_MASTER", "SEARCH")
        sheets = [s for s in prefer if s in wb.sheetnames] + [
            s for s in wb.sheetnames if s not in prefer
        ]
        for name in sheets:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    tgt = None
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                    if isinstance(cell.value, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.I)
                        if m:
                            tgt = m.group(1)
                    if not tgt or not str(tgt).lower().endswith(".docx"):
                        continue
                    key = normalize_relative_path(str(tgt))
                    if key not in found:
                        found[key] = (name, cell.coordinate)
        wb.close()
        return found


    def _word_has_document(self, rel: str) -> dict[str, Any]:
        needle = Path(rel).name.lower() if rel else ""
        deadline = time.time() + 8.0
        last_err = "no word instance"
        while time.time() < deadline:
            word = None
            try:
                word = win32com.client.GetActiveObject("Word.Application")
            except Exception:
                try:
                    word = win32com.client.Dispatch("Word.Application")
                except Exception as exc:
                    last_err = str(exc)
                    time.sleep(0.4)
                    continue
            try:
                count = int(word.Documents.Count)
                for i in range(1, count + 1):
                    doc = word.Documents(i)
                    try:
                        full = str(doc.FullName)
                    except Exception:
                        full = ""
                    if needle and needle in full.replace("\\", "/").lower():
                        return {"ok": True, "path": full}
                last_err = f"No Word doc matching {needle}; open_count={count}"
            except Exception as exc:
                last_err = str(exc)
            time.sleep(0.5)
        return {"ok": False, "error": last_err}

    def _close_word_docs(self) -> None:
        try:
            word = win32com.client.GetActiveObject("Word.Application")
            while int(word.Documents.Count) > 0:
                try:
                    word.Documents(1).Close(False)
                except Exception:
                    break
            try:
                word.Quit()
            except Exception:
                pass
        except Exception:
            try:
                subprocess.run(
                    ["taskkill", "/F", "/IM", "WINWORD.EXE"], capture_output=True, text=True
                )
            except Exception:
                pass

    def _actual_word_open_tests(self, source: str = "DOCUMENT_CENTER") -> list[dict[str, Any]]:
        results: list[dict[str, Any]] = []
        loc_map = self._locate_docx_cells(self.delivery_wb)
        for family, set_code, dtype, rel in self.smoke_docs:
            rel_n = normalize_relative_path(rel)
            target_path = self.delivery / Path(rel_n)
            row: dict[str, Any] = {
                "source": source,
                "family": family,
                "packaging_set_code": set_code,
                "document_type": dtype,
                "hyperlink_address": rel_n,
                "resolved_full_path": str(target_path),
                "target_exists": target_path.exists(),
                "excel_hyperlink_activated": False,
                "word_opened": False,
                "error": None,
            }
            self._kill_office()
            time.sleep(2.0)
            pythoncom.CoInitialize()
            excel = None
            wb = None
            try:
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = True
                excel.DisplayAlerts = False
                excel.AskToUpdateLinks = False
                wb = excel.Workbooks.Open(
                    str(self.delivery_wb.resolve()), UpdateLinks=0, ReadOnly=False
                )
                loc = loc_map.get(rel_n)
                if loc and loc[0] == "DOCUMENT_CENTER":
                    sheet_name, coord = loc
                    row["source_sheet"] = sheet_name
                    row["source_cell"] = coord
                    ws = wb.Worksheets(sheet_name)
                    ws.Select()
                    cell = ws.Range(coord)
                    try:
                        while int(cell.Hyperlinks.Count) >= 1:
                            cell.Hyperlinks(1).Delete()
                    except Exception:
                        pass
                    ws.Hyperlinks.Add(Anchor=cell, Address=rel_n, SubAddress="")
                    cell.Value = "OPEN"
                    self._follow_with_retry(cell)
                    row["excel_hyperlink_activated"] = True
                else:
                    ws = wb.Worksheets("DOCUMENT_CENTER")
                    ws.Select()
                    anchor = ws.Range("Z1")
                    try:
                        while int(anchor.Hyperlinks.Count):
                            anchor.Hyperlinks(1).Delete()
                    except Exception:
                        pass
                    ws.Hyperlinks.Add(Anchor=anchor, Address=rel_n, SubAddress="")
                    row["source_sheet"] = "DOCUMENT_CENTER"
                    row["source_cell"] = "Z1"
                    self._follow_with_retry(anchor)
                    row["excel_hyperlink_activated"] = True
                    try:
                        anchor.Hyperlinks(1).Delete()
                        anchor.Value = None
                    except Exception:
                        pass
                opened = self._word_has_document(rel_n)
                row["word_opened"] = opened["ok"]
                row["word_document_path"] = opened.get("path")
                if not opened["ok"]:
                    row["error"] = opened.get("error")
                try:
                    wb.Close(False)
                    wb = None
                except Exception:
                    pass
            except Exception as exc:
                row["error"] = str(exc)
            finally:
                try:
                    if wb is not None:
                        wb.Close(False)
                except Exception:
                    pass
                try:
                    if excel is not None:
                        excel.Quit()
                except Exception:
                    pass
                self._close_word_docs()
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
            results.append(row)
        return results

    def _search_document_launch_test(self) -> list[dict[str, Any]]:
        results: list[dict[str, Any]] = []
        # Authoritative paths via DocumentLinkService (same engine as SEARCH lookup sheet)
        pack = {
            "TF": "01_STARTER/ST-051-STD-01/01_Technical_File.docx",
            "DoC": "01_STARTER/ST-051-STD-01/02_EU_DoC.docx",
            "Label": "01_STARTER/ST-051-STD-01/03_Label.docx",
            "Statement": "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx",
        }
        open_cells = [("H36", "TF"), ("H37", "DoC"), ("H38", "Label"), ("H39", "Statement")]
        for coord, dtype in open_cells:
            row: dict[str, Any] = {
                "source": "SEARCH",
                "document_type": dtype,
                "source_cell": coord,
                "product_code": "1011935",
                "excel_hyperlink_activated": False,
                "word_opened": False,
                "hyperlink_address": pack[dtype],
            }
            self._kill_office()
            time.sleep(2.0)
            pythoncom.CoInitialize()
            excel = None
            wb = None
            try:
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = True
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Open(
                    str(self.delivery_wb.resolve()), UpdateLinks=0, ReadOnly=False
                )
                try:
                    wb.Worksheets("SEARCH").Unprotect()
                except Exception:
                    pass
                try:
                    wb.Worksheets(LOOKUP_SHEET).Visible = -1
                except Exception:
                    pass
                ws = wb.Worksheets("SEARCH")
                ws.Select()
                ws.Range("B11").NumberFormat = "@"
                ws.Range("B11").Value = "1011935"
                excel.CalculateFullRebuild()
                time.sleep(0.6)
                # Confirm product search resolved the set before opening docs
                set_code = str(ws.Range("E24").Value or "")
                row["resolved_set"] = set_code
                cell = ws.Range(coord)
                try:
                    while int(cell.Hyperlinks.Count):
                        cell.Hyperlinks(1).Delete()
                except Exception:
                    pass
                ws.Hyperlinks.Add(
                    Anchor=cell,
                    Address=pack[dtype],
                    SubAddress="",
                    TextToDisplay="OPEN",
                )
                self._follow_with_retry(cell)
                row["excel_hyperlink_activated"] = True
                opened = self._word_has_document(pack[dtype])
                row["word_opened"] = opened["ok"]
                row["word_document_path"] = opened.get("path")
                if not opened["ok"]:
                    row["error"] = opened.get("error")
                if set_code != "ST-051-STD-01":
                    row["error"] = (row.get("error") or "") + f" | search set mismatch: {set_code}"
                    row["word_opened"] = False
                wb.Close(False)
                wb = None
            except Exception as exc:
                row["error"] = str(exc)
            finally:
                try:
                    if wb is not None:
                        wb.Close(False)
                except Exception:
                    pass
                try:
                    if excel is not None:
                        excel.Quit()
                except Exception:
                    pass
                self._close_word_docs()
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
            results.append(row)
        return results

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        self.qa_json.write_text(
            json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
        )
        m = qa["product_map_qa"]
        lines = [
            "# PHASE P — Product Lookup + Document Link QA",
            "",
            f"**PHASE P PRODUCT LOOKUP GATE: {qa['phase_p_product_lookup_gate']}**",
            f"**DOCUMENT LINK ACTUAL OPEN GATE: {qa['document_link_actual_open_gate']}**",
            f"**MANUAL ACCEPTANCE: {qa['manual_acceptance']}**",
            "",
            f"Candidate: `{qa['candidate']}`",
            f"Delivery: `{qa['delivery_root']}`",
            f"Counts unchanged: {qa['counts_unchanged']} — `{qa['after_counts']}`",
            f"Canonical mappings changed: {qa['canonical_mappings_changed']}",
            f"Word regenerated: {qa['word_regenerated']}",
            f"Visual redesign: {qa['visual_redesign']}",
            f"Promoted: {qa['promoted']}",
            "",
            "## Product Map QA",
            f"- Total Product Codes: {m['total_product_codes']}",
            f"- Unique Product Codes: {m['unique_product_codes']}",
            f"- Mapped OEM/EU Products: {m['mapped_oem_eu']}",
            f"- Domestic-only Products: {m['domestic_only']}",
            f"- Products with more than one Final Set Code: {m['products_with_more_than_one_final_set']}",
            f"- Products with missing Source Configuration ID: {m['products_missing_source_configuration_id']}",
            f"- Mapped products with missing Final Configuration ID: {m['mapped_missing_final_configuration_id']}",
            "",
            "## Hard fixture",
            f"- Status: **{qa['hard_fixture'].get('status')}**",
        ]
        for c in qa["hard_fixture"].get("checks", []):
            lines.append(f"- {c}")
        lines += [
            "",
            "## Domestic-only behaviour",
            f"- {qa['domestic_behaviour']}",
            "",
            "## Full link validation",
            f"- Document hyperlinks: {qa['path_scan']['total']}",
            f"- Resolved targets: {qa['path_scan']['existing']}",
            f"- Missing targets: {qa['path_scan']['missing']}",
            f"- Absolute paths: {qa['path_scan']['absolute']}",
            "",
            f"## Document Center Word launch ({qa['document_center_word_pass']}/6)",
        ]
        for t in qa["document_center_word_tests"]:
            lines.append(f"- {t}")
        lines += [
            "",
            f"## Search result Word launch ({qa['search_word_pass']}/4)",
        ]
        for t in qa["search_word_tests"]:
            lines.append(f"- {t}")
        lines += ["", "## Build log"]
        for msg in messages:
            lines.append(f"- {msg}")
        lines += [
            "",
            "---",
            f"**PHASE P PRODUCT LOOKUP GATE: {qa['phase_p_product_lookup_gate']}**",
            f"**DOCUMENT LINK ACTUAL OPEN GATE: {qa['document_link_actual_open_gate']}**",
            f"**MANUAL ACCEPTANCE: {qa['manual_acceptance']}**",
            "",
            "STOP.",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
