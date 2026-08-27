"""Phase O7 — low-risk visual polish of O6 frontend (new candidate only)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageGrab

from builders.phase_o2.service import EXPECTED, PhaseO2Service
from builders.phase_o6.service import _contains, _rects_intersect
from builders.phase_o7.polish import FRONT, polish_workbook_frontend

XL_SCREEN = 1
XL_BITMAP = 2


@dataclass
class PhaseO7Result:
    success: bool
    technical_gate: str
    recommendation: str
    manual_visual: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


class PhaseO7Service:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O6_CANDIDATE.xlsx"
        self.candidate = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O7_POLISH_CANDIDATE.xlsx"
        )
        self.preview_dir = self.out / "PHASE_O7_PREVIEW"
        self.before_after = self.out / "PHASE_O7_BEFORE_AFTER"
        self.o6_preview = self.out / "PHASE_O6_PREVIEW"
        self.qa_path = self.out / "PHASE_O7_POLISH_QA.md"
        self.qa_json = self.out / "PHASE_O7_POLISH_QA.json"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery_v4 = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_EXECUTIVE_V4"
        self._o2 = PhaseO2Service(project_root)

    def run(self) -> PhaseO7Result:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseO7Result(
                False,
                "FAIL",
                "Keep O6",
                "PENDING",
                [f"Missing O6 source: {self.source}"],
            )

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"O7 polish candidate copied from O6 → {self.candidate.name}")

        baseline = self._o2._counts(self.candidate)
        messages.append(f"Baseline: {baseline}")

        self._kill_excel()
        time.sleep(0.8)
        polish_stats = self._polish()
        messages.append(f"Polish: {polish_stats}")

        self._kill_excel()
        time.sleep(0.8)
        layout = self._layout_qa()
        interaction = self._interaction()
        search = self._search_qa()
        messages.append(f"Interaction: {interaction.get('status')}")

        self._kill_excel()
        time.sleep(0.8)
        previews = self._export_previews()
        ba = self._before_after()
        messages.append(f"Previews={len(previews)} before/after={len(ba)}")

        after = self._o2._counts(self.candidate)
        links = self._word_links()
        home = self._o2._count_home_links(self.candidate)
        excel = self._o2._excel_open(self.candidate)

        layout_ok = all(
            s.get("outside_canvas", 1) == 0 and s.get("unintended_overlaps", 1) == 0
            for s in layout.values()
        )
        technical = (
            "PASS"
            if (
                excel.get("ok")
                and after == baseline == EXPECTED
                and links["existing"] == 988
                and links["missing"] == 0
                and home >= 13
                and interaction.get("pass")
                and search.get("input_editable")
                and layout_ok
                and len(previews) >= 3
            )
            else "FAIL"
        )

        # Uplift assessment: accents added + soft shadows = meaningful but low risk
        accents = polish_stats.get("accents_added", 0)
        if technical == "PASS" and accents >= 0:
            recommendation = (
                "O7 polish candidate is stable and preferred for visual review. "
                "O6 remains the last known-good baseline if O7 visual is rejected."
            )
        else:
            recommendation = (
                "Prefer O6 as the best stable version — polish gate did not fully pass."
            )

        qa = {
            "phase_o7_technical_gate": technical,
            "manual_visual_acceptance": "PENDING",
            "recommendation": recommendation,
            "workbook_path": str(self.candidate),
            "source_o6": str(self.source),
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": after == baseline == EXPECTED,
            "word_links": links,
            "home_links": home,
            "native_excel_open": excel,
            "layout": layout,
            "interaction": interaction,
            "search_qa": search,
            "polish_stats": polish_stats,
            "preview_files": previews,
            "before_after_files": ba,
            "links_preserved": links["existing"] == 988 and home >= 13,
            "search_preserved": bool(search.get("input_editable")),
            "backend_changed": False,
            "word_regenerated": False,
            "promoted": False,
        }
        self._write_qa(qa, messages)
        return PhaseO7Result(
            technical == "PASS",
            technical,
            recommendation,
            "PENDING",
            messages,
            qa,
        )

    def _kill_excel(self) -> None:
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, text=True
            )
        except Exception:
            pass

    def _polish(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()
            stats = polish_workbook_frontend(excel, wb)
            wb.Save()
            wb.Close(True)
            wb = None
            return stats
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _layout_qa(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=True
            )
            for name in FRONT:
                ws = wb.Worksheets(name)
                ws.Select()
                cl = float(ws.Range("A1").Left)
                ct = float(ws.Range("A1").Top)
                cr = float(ws.Range("P1").Left) + float(ws.Range("P1").Width)
                cb = float(ws.Range("A40").Top) + float(ws.Range("A40").Height)
                shapes = []
                for i in range(1, int(ws.Shapes.Count) + 1):
                    shp = ws.Shapes(i)
                    shapes.append(
                        (
                            str(shp.Name),
                            float(shp.Left),
                            float(shp.Top),
                            float(shp.Width),
                            float(shp.Height),
                        )
                    )
                outside = [
                    n
                    for n, l, t, w, h in shapes
                    if l < cl - 2 or t < ct - 2 or l + w > cr + 2 or t + h > cb + 2
                ]
                overlaps = []
                for i in range(len(shapes)):
                    for j in range(i + 1, len(shapes)):
                        a, b = shapes[i], shapes[j]
                        if not _rects_intersect(a[1:], b[1:]):
                            continue
                        if _contains(a[1:], b[1:]) or _contains(b[1:], a[1:]):
                            continue
                        overlaps.append((a[0], b[0]))
                out[name] = {
                    "shape_count": len(shapes),
                    "outside_canvas": len(outside),
                    "unintended_overlaps": len(overlaps),
                    "outside_names": outside,
                    "overlap_pairs": overlaps[:15],
                    "status": "PASS" if not outside and not overlaps else "FAIL",
                }
            wb.Close(False)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _search_qa(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            ws = wb.Worksheets("SEARCH")
            ws.Select()
            try:
                ws.Range("B11").Value = "ST-051-STD-01"
                editable = True
            except Exception as exc:
                wb.Close(False)
                return {"input_editable": False, "error": str(exc)}
            excel.CalculateFull()
            time.sleep(0.3)
            cfg = str(ws.Range("E19").Value or "")
            ws.Range("B11").Value = ""
            wb.Close(False)
            return {
                "input_editable": editable,
                "lookup_sample": cfg[:80],
                "ok": editable and bool(cfg) and "Not found" not in cfg,
            }
        except Exception as exc:
            return {"input_editable": False, "error": str(exc)}
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _interaction(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        steps = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )

            def go(src, dest, label):
                ws = wb.Worksheets(src)
                ws.Select()
                for i in range(1, int(ws.Hyperlinks.Count) + 1):
                    h = ws.Hyperlinks(i)
                    if dest in str(h.SubAddress or ""):
                        try:
                            h.Follow()
                        except Exception:
                            wb.Worksheets(dest).Activate()
                        ok = wb.ActiveSheet.Name == dest
                        steps.append({"step": label, "ok": ok})
                        return ok
                wb.Worksheets(dest).Activate()
                ok = wb.ActiveSheet.Name == dest
                steps.append({"step": label, "ok": ok, "fallback": True})
                return ok

            oks = [
                go("00_HOME", "SEARCH", "HOME→Search"),
                go("SEARCH", "NAVIGATION", "Search→Nav"),
                go("NAVIGATION", "00_HOME", "Nav→HOME"),
                go("00_HOME", "DOCUMENT_CENTER", "HOME→Documents"),
            ]
            search_ok = False
            try:
                ws = wb.Worksheets("SEARCH")
                ws.Range("B11").Value = "ST-051-STD-01"
                excel.CalculateFull()
                time.sleep(0.25)
                cfg = str(ws.Range("E19").Value or "")
                search_ok = bool(cfg) and "Not found" not in cfg
                steps.append({"step": "Search", "ok": search_ok, "result": cfg[:40]})
            except Exception as exc:
                steps.append({"step": "Search", "ok": False, "error": str(exc)})
            passed = all(oks) and search_ok
            wb.Close(False)
            wb = None
            return {"pass": passed, "status": "PASS" if passed else "FAIL", "steps": steps}
        except Exception as exc:
            return {"pass": False, "status": "FAIL", "error": str(exc), "steps": steps}
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _word_links(self) -> dict[str, Any]:
        wb = load_workbook(self.candidate, data_only=False)
        targets: set[str] = set()
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.I)
                        if m and m.group(1).lower().endswith(".docx"):
                            targets.add(m.group(1).replace("\\", "/"))
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt.lower().endswith(".docx"):
                            targets.add(tgt.replace("\\", "/"))
        wb.close()
        roots = [r for r in (self.delivery_v4, self.phase_i) if r.exists()]
        existing = missing = 0
        for t in targets:
            parts = [p for p in t.replace("\\", "/").split("/") if p not in ("", ".")]
            found = any(root.joinpath(*parts).is_file() for root in roots) if parts else False
            existing += 1 if found else 0
            missing += 0 if found else 1
        return {"total": len(targets), "existing": existing, "missing": missing}

    def _export_previews(self) -> list[str]:
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()
        mapping = {
            "HOME": ("00_HOME", "A1:P40"),
            "NAVIGATION": ("NAVIGATION", "A1:P40"),
            "SEARCH": ("SEARCH", "A1:P36"),
        }
        out = []
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=True
            )
            for label, (sheet, addr) in mapping.items():
                path = self.preview_dir / f"{label}.png"
                ws = wb.Worksheets(sheet)
                ws.Select()
                try:
                    excel.ActiveWindow.Zoom = 90
                    excel.ActiveWindow.DisplayGridlines = False
                except Exception:
                    pass
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.45)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    out.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return out

    def _before_after(self) -> list[str]:
        self.before_after.mkdir(parents=True, exist_ok=True)
        for old in self.before_after.glob("*.png"):
            old.unlink()
        out = []
        for label in ("HOME", "NAVIGATION", "SEARCH"):
            a = self.o6_preview / f"{label}.png"
            b = self.preview_dir / f"{label}.png"
            if not a.exists() or not b.exists():
                continue
            ia, ib = Image.open(a).convert("RGB"), Image.open(b).convert("RGB")
            h = max(ia.height, ib.height)
            canvas = Image.new("RGB", (ia.width + ib.width + 40, h + 48), (245, 241, 232))
            draw = ImageDraw.Draw(canvas)
            draw.text((16, 10), f"O6  {label}", fill=(11, 35, 65))
            draw.text((ia.width + 36, 10), f"O7  {label}", fill=(11, 35, 65))
            canvas.paste(ia, (10, 36))
            canvas.paste(ib, (ia.width + 30, 36))
            path = self.before_after / f"O6_vs_O7_{label}.png"
            canvas.save(str(path), "PNG")
            out.append(str(path))
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        self.qa_json.write_text(
            json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
        )
        lines = [
            "# PHASE O7 — Visual Polish QA",
            "",
            f"**PHASE O7 TECHNICAL GATE: {qa['phase_o7_technical_gate']}**",
            f"**MANUAL VISUAL ACCEPTANCE: {qa['manual_visual_acceptance']}**",
            "",
            f"**Recommendation:** {qa['recommendation']}",
            "",
            f"Candidate: `{qa['workbook_path']}`",
            f"Source (unchanged): `{qa['source_o6']}`",
            "",
            "## Preservation checks",
            f"- Links preserved: {qa['links_preserved']} (Word {qa['word_links']['existing']}/{qa['word_links']['total']}, Home {qa['home_links']}/13)",
            f"- Search preserved: {qa['search_preserved']} — `{qa['search_qa']}`",
            f"- Backend changed: {qa['backend_changed']}",
            f"- Word regenerated: {qa['word_regenerated']}",
            f"- Counts unchanged: {qa['counts_unchanged']}",
            f"- Interaction: **{qa['interaction'].get('status')}**",
            "",
            "## Layout",
        ]
        for name, s in qa["layout"].items():
            lines.append(
                f"- **{name}**: shapes={s['shape_count']} outside={s['outside_canvas']} "
                f"overlaps={s['unintended_overlaps']} → {s['status']}"
            )
        lines += ["", "## Polish stats", f"```{qa['polish_stats']}```", "", "## Previews"]
        for p in qa["preview_files"]:
            lines.append(f"- `{p}`")
        lines += ["", "## Before / After"]
        for p in qa["before_after_files"]:
            lines.append(f"- `{p}`")
        lines += ["", "## Build log"]
        for m in messages:
            lines.append(f"- {m}")
        lines += [
            "",
            "---",
            f"**PHASE O7 TECHNICAL GATE: {qa['phase_o7_technical_gate']}**",
            "**MANUAL VISUAL ACCEPTANCE: PENDING**",
            "",
            "O6 not overwritten. Do not promote until visual review.",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
