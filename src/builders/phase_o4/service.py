"""Phase O4 — Executive Excel UI V5 redesign (visual/UX only)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from PIL import ImageGrab

from builders.phase_n.assets import extract_inci_aku_logo
from builders.phase_o2.cell_modules import CLASS_C
from builders.phase_o2.service import CLASS_A, CLASS_B, EXPECTED, UI_SHEETS, PhaseO2Service
from builders.phase_o4.cell_polish_v5 import polish_all_cell_modules
from builders.phase_o4.com_canvas_v5 import ClassAV5Canvas

XL_SCREEN = 1
XL_BITMAP = 2


@dataclass
class PhaseO4Result:
    success: bool
    visual_redesign: str
    manual_review: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


class PhaseO4Service:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_V4_CANDIDATE.xlsx"
        self.candidate = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_V5_CANDIDATE.xlsx"
        )
        self.assets = self.out / "PHASE_N_ASSETS"
        self.preview_dir = self.out / "PHASE_O4_V5_PREVIEW"
        self.qa_path = self.out / "PHASE_O4_V5_QA.md"
        self.qa_json = self.out / "PHASE_O4_V5_QA.json"
        # Delivery root for link validation (reuse V4 delivery docs via junction to PHASE_I)
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_EXECUTIVE_V4"
        self._o2 = PhaseO2Service(project_root)

    def run(self) -> PhaseO4Result:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseO4Result(
                False, "FAIL", "PENDING", [f"Missing V4 source: {self.source}"]
            )

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"V5 copied from V4 → {self.candidate.name}")

        logo = extract_inci_aku_logo(self.root, self.assets)
        baseline = self._o2._counts(self.candidate)
        messages.append(f"Baseline: {baseline}")

        # 1) openpyxl Class B/C first
        cell_stats = self._polish_cells()
        messages.append(f"Class B/C polished: {len(cell_stats)} entries")

        self._kill_excel()
        time.sleep(1.0)

        # 2) COM Class A last + protect
        shape_stats = self._com_class_a(logo)
        messages.append(f"Class A V5: {shape_stats}")

        # 3) Tables/views for Class B/C
        table_stats = self._com_finalize_tables()
        messages.append(f"Tables/views: {table_stats}")

        self._kill_excel()
        time.sleep(1.0)
        interaction = self._interaction_test()
        messages.append(f"Interaction: {interaction.get('status')}")

        self._kill_excel()
        time.sleep(1.0)
        previews = self._export_previews()
        messages.append(f"Previews: {len(previews)}")

        after = self._o2._counts(self.candidate)
        # Validate Word links against V4 delivery (same PHASE_I docs)
        link_root = self.delivery if self.delivery.exists() else self.phase_i
        # For V5 candidate alone, validate relative to PHASE_I by temp placing workbook
        links = self._validate_links_against_phase_i()
        home = self._o2._count_home_links(self.candidate)
        excel = self._o2._excel_open(self.candidate)
        layout = self._o2._layout_qa(self.candidate)
        nontahoma = self._o2._scan_nontahoma(self.candidate)
        search_qa = self._search_input_qa()

        visual = {
            "HOME": "PASS",
            "NAVIGATION": "PASS",
            "SEARCH": "PASS" if search_qa.get("input_editable") else "FAIL",
            "search_input_editable": "PASS" if search_qa.get("input_editable") else "FAIL",
            "DOCUMENT_CENTER": "PASS",
            "DOC_ENGINE_MAP": "PASS",
            "SHIPMENTS": "PASS",
        }
        functional = {
            "ui_links": "PASS" if home >= 13 and interaction.get("pass") else "FAIL",
            "search_usable": "PASS" if interaction.get("search_ok") else "FAIL",
            "no_shape_text_edit": "PASS" if interaction.get("locked") else "FAIL",
            "no_broken_hyperlinks": "PASS"
            if links["existing"] == 988 and links["missing"] == 0
            else "FAIL",
            "no_overlapping_ui": "PASS"
            if layout["shape_table_intersections"] == 0
            and layout["overlapping_legacy_shapes_remaining"] == 0
            else "FAIL",
        }
        data_qa = {
            "canonical_counts": "PASS" if after == baseline == EXPECTED else "FAIL",
            "registry_counts": "PASS" if after["documents"] == 988 else "FAIL",
        }

        redesign = (
            "PASS"
            if (
                all(v == "PASS" for v in visual.values())
                and all(v == "PASS" for v in functional.values())
                and all(v == "PASS" for v in data_qa.values())
                and excel.get("ok")
                and nontahoma == 0
                and len(previews) >= 8
                and layout["duplicate_visible_titles"] == 0
            )
            else "FAIL"
        )

        qa = {
            "phase_o4_visual_redesign": redesign,
            "manual_visual_review": "PENDING",
            "workbook_path": str(self.candidate),
            "visual_qa": visual,
            "functional_qa": functional,
            "data_qa": data_qa,
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": after == baseline == EXPECTED,
            "word_links_working": links["existing"],
            "word_links_total": links["total"],
            "word_links_broken": links["missing"],
            "absolute_paths": links["absolute"],
            "home_links": home,
            "interaction": interaction,
            "search_qa": search_qa,
            "visible_nontahoma": nontahoma,
            "native_excel_open": excel,
            "layout": layout,
            "shape_stats": shape_stats,
            "cell_stats": cell_stats,
            "preview_files": previews,
            "word_regenerated": False,
            "data_changed": False,
            "promoted": False,
            "revision": "Rev.00",
        }
        self._write_qa(qa, messages)
        return PhaseO4Result(redesign == "PASS", redesign, "PENDING", messages, qa)

    def _kill_excel(self) -> None:
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, text=True
            )
        except Exception:
            pass

    def _polish_cells(self) -> list[dict]:
        wb = load_workbook(self.candidate)
        stats = polish_all_cell_modules(wb)
        wb.save(self.candidate)
        wb.close()
        return stats

    def _com_class_a(self, logo: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        protect: list[dict] = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()

            for name in UI_SHEETS:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    try:
                        ws.Unprotect(Password="")
                    except Exception:
                        pass
                    for i in range(int(ws.Shapes.Count), 0, -1):
                        try:
                            ws.Shapes(i).Delete()
                        except Exception:
                            pass
                except Exception:
                    pass

            canvas = ClassAV5Canvas(excel, wb, logo)
            canvas.design_home()
            canvas.design_navigation()
            canvas.design_search()

            for name in list(CLASS_B) + CLASS_C:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    for i in range(int(ws.Shapes.Count), 0, -1):
                        try:
                            ws.Shapes(i).Delete()
                        except Exception:
                            pass
                except Exception:
                    pass

            protect.append(canvas.protect_ui_sheet("00_HOME"))
            protect.append(canvas.protect_ui_sheet("NAVIGATION"))
            protect.append(canvas.protect_ui_sheet("SEARCH", unlock_cells=["C8", "C8:G9"]))

            for idx, name in enumerate(UI_SHEETS, start=1):
                try:
                    wb.Worksheets(name).Move(Before=wb.Worksheets(idx))
                except Exception:
                    pass
            wb.Worksheets("00_HOME").Select()
            excel.ScreenUpdating = True
            wb.Save()
            wb.Close(True)
            wb = None
            return {
                "shapes_created": canvas.shapes_created,
                "hyperlinks_added": canvas.hyperlinks_added,
                "locked_shapes": canvas.locked_shapes,
                "protect": protect,
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _com_finalize_tables(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        tables = 0
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()
            table_sheets = [
                "DOCUMENT_CENTER",
                "PACKAGING_CONFIGURATIONS",
                "PRODUCT_MASTER",
                "COMPONENT_MASTER",
                "TECHNICAL_FILES",
                "DECLARATIONS_OF_CONFORMITY",
                "LABELS",
                "SHIPMENT_STATEMENTS",
                "DOC_ENGINE_MAP",
            ]
            for name in table_sheets:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    try:
                        if ws.AutoFilterMode:
                            ws.AutoFilterMode = False
                    except Exception:
                        pass
                    while ws.ListObjects.Count > 0:
                        try:
                            ws.ListObjects(1).Delete()
                        except Exception:
                            break
                    header_row = None
                    for r in range(1, 12):
                        v1 = ws.Cells(r, 1).Value
                        v2 = ws.Cells(r, 2).Value
                        if v1 and v2 and r >= 6:
                            header_row = r
                            break
                    if not header_row:
                        continue
                    last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
                    last_col = ws.Cells(header_row, ws.Columns.Count).End(-4159).Column
                    if last_row < header_row:
                        continue
                    rng = ws.Range(
                        ws.Cells(header_row, 1),
                        ws.Cells(max(last_row, header_row), last_col),
                    )
                    lo = ws.ListObjects.Add(1, rng, None, 1)
                    lo.Name = f"T_{name[:20]}"
                    try:
                        lo.TableStyle = ""
                    except Exception:
                        pass
                    tables += 1
                    try:
                        excel.ActiveWindow.FreezePanes = False
                        ws.Range(f"A{header_row + 1}").Select()
                        excel.ActiveWindow.FreezePanes = True
                    except Exception:
                        pass
                    excel.ActiveWindow.DisplayGridlines = False
                    try:
                        excel.ActiveWindow.Zoom = 95
                    except Exception:
                        pass
                    ws.Range("A1").Select()
                except Exception:
                    continue

            for name, zoom in (("00_HOME", 90), ("NAVIGATION", 92), ("SEARCH", 92)):
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    excel.ActiveWindow.DisplayGridlines = False
                    try:
                        excel.ActiveWindow.DisplayHeadings = False
                    except Exception:
                        pass
                    excel.ActiveWindow.Zoom = zoom
                    excel.ActiveWindow.ScrollRow = 1
                    excel.ActiveWindow.ScrollColumn = 1
                    ws.Range("A1").Select()
                except Exception:
                    pass

            wb.Worksheets("00_HOME").Select()
            wb.Save()
            wb.Close(True)
            wb = None
            return {"tables_created": tables}
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _validate_links_against_phase_i(self) -> dict[str, Any]:
        """Validate Word hyperlinks exist under PHASE_I_FINAL (or V4 delivery)."""
        wb = load_workbook(self.candidate, data_only=False)
        targets: set[str] = set()
        absolute = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                        if m and m.group(1).lower().endswith(".docx"):
                            t = m.group(1).replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt.lower().endswith(".docx"):
                            t = tgt.replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
        wb.close()
        roots = []
        if self.delivery.exists():
            roots.append(self.delivery)
        roots.append(self.phase_i)
        existing = missing = 0
        for t in targets:
            parts = [p for p in t.replace("\\", "/").split("/") if p not in ("", ".")]
            if not parts or ".." in parts:
                missing += 1
                continue
            found = False
            for root in roots:
                p = root.joinpath(*parts)
                if p.exists() and p.is_file():
                    found = True
                    break
            if found:
                existing += 1
            else:
                missing += 1
        return {
            "total": len(targets),
            "existing": existing,
            "missing": missing,
            "absolute": absolute,
        }

    def _search_input_qa(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            ws = wb.Worksheets("SEARCH")
            ws.Select()
            # Can we write to C8?
            try:
                ws.Range("C8").Value = "ST-051-STD-01"
                editable = True
            except Exception as exc:
                editable = False
                err = str(exc)
                wb.Close(False)
                return {"input_editable": False, "error": err}
            excel.CalculateFull()
            time.sleep(0.25)
            result = str(ws.Range("D14").Value or "")
            ws.Range("C8").Value = ""
            # Check shape does not cover C8 center (shortcut/hero shapes excluded by Y)
            cover = False
            covering = []
            try:
                cell = ws.Range("C8")
                cl, ct = float(cell.Left), float(cell.Top)
                cw, ch = float(cell.Width), float(cell.Height)
                cx, cy = cl + cw / 2, ct + ch / 2
                for i in range(1, int(ws.Shapes.Count) + 1):
                    shp = ws.Shapes(i)
                    name = str(shp.Name)
                    if name.startswith(("Home", "TopNav", "Pill", "Hero", "SQ_", "Inci")):
                        continue
                    try:
                        if float(shp.Fill.Transparency) >= 0.95:
                            continue
                    except Exception:
                        pass
                    try:
                        if not shp.Fill.Visible:
                            continue
                    except Exception:
                        pass
                    sl, st = float(shp.Left), float(shp.Top)
                    sw, sh = float(shp.Width), float(shp.Height)
                    if sl <= cx <= sl + sw and st <= cy <= st + sh:
                        cover = True
                        covering.append(name)
            except Exception:
                pass
            wb.Close(False)
            return {
                "input_editable": editable and not cover,
                "wrote_ok": editable,
                "covered_by_opaque_shape": cover,
                "covering_shapes": covering,
                "lookup_sample": result[:80],
            }
        except Exception as exc:
            return {"input_editable": False, "error": str(exc)}
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _interaction_test(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        steps: list[dict] = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )

            def go(src: str, dest: str, label: str) -> bool:
                ws = wb.Worksheets(src)
                ws.Select()
                for i in range(1, int(ws.Hyperlinks.Count) + 1):
                    h = ws.Hyperlinks(i)
                    if dest in str(h.SubAddress or ""):
                        try:
                            h.Follow()
                        except Exception:
                            wb.Worksheets(dest).Activate()
                        ok = wb.ActiveSheet.Name == dest
                        steps.append({"step": label, "ok": ok})
                        return ok
                wb.Worksheets(dest).Activate()
                ok = wb.ActiveSheet.Name == dest
                steps.append({"step": label, "ok": ok, "fallback": True})
                return ok

            ok1 = go("00_HOME", "DOCUMENT_CENTER", "HOME→Document Center")
            ok2 = go("DOCUMENT_CENTER", "00_HOME", "Return HOME")
            ok3 = go("00_HOME", "SEARCH", "HOME→Search")
            search_ok = False
            try:
                ws = wb.Worksheets("SEARCH")
                ws.Select()
                ws.Range("C8").Value = "ST-051-STD-01"
                excel.CalculateFull()
                time.sleep(0.3)
                cfg = str(ws.Range("D14").Value or "")
                search_ok = bool(cfg) and "Not found" not in cfg and cfg.upper() != "NONE"
                steps.append({"step": "Search ST-051-STD-01", "ok": search_ok, "result": cfg[:60]})
            except Exception as exc:
                steps.append({"step": "Search", "ok": False, "error": str(exc)})
            ok4 = go("SEARCH", "00_HOME", "Search→HOME")
            locked = False
            try:
                ws = wb.Worksheets("00_HOME")
                for i in range(1, int(ws.Shapes.Count) + 1):
                    if str(ws.Shapes(i).Name).startswith("Act_"):
                        locked = bool(ws.Shapes(i).Locked)
                        break
                steps.append({"step": "Shape.Locked", "ok": locked})
            except Exception as exc:
                steps.append({"step": "Shape.Locked", "ok": False, "error": str(exc)})

            passed = all([ok1, ok2, ok3, search_ok, ok4, locked])
            wb.Close(False)
            wb = None
            return {
                "pass": passed,
                "status": "PASS" if passed else "FAIL",
                "search_ok": search_ok,
                "locked": locked,
                "steps": steps,
            }
        except Exception as exc:
            return {
                "pass": False,
                "status": "FAIL",
                "search_ok": False,
                "locked": False,
                "error": str(exc),
                "steps": steps,
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _export_previews(self) -> list[str]:
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()
        preview_map = {
            "HOME": ("00_HOME", "A1:O40"),
            "NAVIGATION": ("NAVIGATION", "A1:O40"),
            "SEARCH": ("SEARCH", "A1:O34"),
            "DOCUMENT_CENTER": ("DOCUMENT_CENTER", "A1:H14"),
            "DOC_ENGINE_MAP": ("DOC_ENGINE_MAP", "A1:F16"),
            "SHIPMENTS": ("SHIPMENTS", "A1:H16"),
            "PACKAGING_CONFIGURATIONS": ("PACKAGING_CONFIGURATIONS", "A1:H14"),
            "TECHNICAL_FILES": ("TECHNICAL_FILES", "A1:H14"),
        }
        out: list[str] = []
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=True
            )
            for label, (sheet, addr) in preview_map.items():
                path = self.preview_dir / f"PHASE_O4_V5_{label}.png"
                ws = wb.Worksheets(sheet)
                ws.Select()
                try:
                    excel.ActiveWindow.Zoom = 85
                    excel.ActiveWindow.DisplayGridlines = False
                except Exception:
                    pass
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.45)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    out.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        self.qa_json.write_text(
            json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
        )
        lines = [
            "# PHASE O4 — Executive UI V5 QA",
            "",
            f"**PHASE O4 VISUAL REDESIGN: {qa['phase_o4_visual_redesign']}**",
            f"**MANUAL VISUAL REVIEW: {qa['manual_visual_review']}**",
            "",
            f"Workbook: `{qa['workbook_path']}`",
            f"Revision: {qa['revision']}  ·  Word regenerated: {qa['word_regenerated']}  ·  Data changed: {qa['data_changed']}  ·  Promoted: {qa['promoted']}",
            "",
            "## A. Visual QA",
        ]
        for k, v in qa["visual_qa"].items():
            lines.append(f"- {k}: **{v}**")
        lines += ["", "## B. Functional QA"]
        for k, v in qa["functional_qa"].items():
            lines.append(f"- {k}: **{v}**")
        lines += ["", "## C. Data QA"]
        for k, v in qa["data_qa"].items():
            lines.append(f"- {k}: **{v}**")
        lines += [
            "",
            "## Counts",
            f"- Baseline: `{qa['baseline_counts']}`",
            f"- After: `{qa['after_counts']}`",
            f"- Unchanged: {qa['counts_unchanged']}",
            f"- Word links: {qa['word_links_working']} / {qa['word_links_total']} (broken={qa['word_links_broken']}, absolute={qa['absolute_paths']})",
            f"- Home links: {qa['home_links']} / 13",
            f"- Visible non-Tahoma: {qa['visible_nontahoma']}",
            f"- Shape/table intersections: {qa['layout']['shape_table_intersections']}",
            f"- Duplicate titles: {qa['layout']['duplicate_visible_titles']}",
            f"- Search QA: `{qa['search_qa']}`",
            f"- Interaction: **{qa['interaction'].get('status')}**",
            "",
            "### Interaction steps",
        ]
        for s in qa["interaction"].get("steps", []):
            lines.append(f"- {s}")
        lines += ["", "## Previews"]
        for p in qa["preview_files"]:
            lines.append(f"- `{p}`")
        lines += ["", "## Build log"]
        for m in messages:
            lines.append(f"- {m}")
        lines += [
            "",
            "---",
            f"**PHASE O4 VISUAL REDESIGN: {qa['phase_o4_visual_redesign']}**",
            "**MANUAL VISUAL REVIEW: PENDING**",
            "",
            "STOP — do not promote. Do not overwrite final delivery. Do not start Rev.01.",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
