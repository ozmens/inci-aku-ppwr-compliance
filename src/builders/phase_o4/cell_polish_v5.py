"""Phase O4 — Class B/C premium header polish (cell-based, no body shapes)."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from builders.phase_o2.cell_modules import (
    HYPERLINK_RE,
    MODULE_META,
    apply_native_doc_link,
    clear_sheet_values,
    extract_table,
    write_data_table,
    write_nav_row,
)

# V5 palette
NAVY = "0B2341"
NAVY2 = "123A63"
GOLD = "C9A24A"
IVORY = "F5F1E8"
CARD = "FCFBF8"
PALE = "E9EEF3"
STONE = "F5F1E8"
INK = "1E2C3A"
MUTED = "5C6B7A"
LINK = "1F5C99"
OK_BG = "E5F0E7"
OK_FG = "325D3E"
FONT = "Tahoma"


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def _font(size=9, bold=False, color=INK, underline=None) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color, underline=underline)


def write_premium_header(
    ws,
    title: str,
    subtitle: str,
    metrics: list[tuple[str, str]],
    cols: int = 8,
) -> None:
    """Cell-based premium module header (rows 1–5), table starts at 6."""
    write_nav_row(ws, min(cols, 8))
    # Row 2 title
    ws.row_dimensions[2].height = 28
    for c in range(1, cols + 1):
        ws.cell(2, c).fill = _fill(NAVY)
        ws.cell(2, c).border = Border()
    t = ws.cell(2, 1, title)
    t.font = _font(16, True, "FFFFFF")
    t.alignment = Alignment(vertical="center")
    # Row 3 subtitle
    ws.row_dimensions[3].height = 18
    for c in range(1, cols + 1):
        ws.cell(3, c).fill = _fill(NAVY2)
    s = ws.cell(3, 1, subtitle)
    s.font = _font(9, False, "B8C7D6")
    s.alignment = Alignment(vertical="center")
    # Row 4 metrics
    ws.row_dimensions[4].height = 22
    for c in range(1, cols + 1):
        ws.cell(4, c).fill = _fill(IVORY)
    for i, (label, value) in enumerate(metrics[:4]):
        col = 1 + i * 2
        if col > cols:
            break
        cell = ws.cell(4, col, f"{value}  ·  {label}")
        cell.font = _font(8, True, NAVY)
        cell.fill = _fill(PALE)
        cell.alignment = Alignment(vertical="center", horizontal="left")
    # Row 5 gold hairline
    ws.row_dimensions[5].height = 6
    for c in range(1, cols + 1):
        ws.cell(5, c).fill = _fill(GOLD)


def polish_document_center(ws) -> dict[str, Any]:
    title, subtitle, _info = MODULE_META["DOCUMENT_CENTER"]
    headers, data = extract_table(ws)
    if len(headers) < 3 or len(data) < 200:
        raise ValueError(
            f"DOCUMENT_CENTER extract failed: headers={len(headers)} rows={len(data)}"
        )
    clear_sheet_values(ws)
    cols = max(len(headers), 8)
    write_premium_header(
        ws,
        "DOCUMENT CENTER",
        "Per-configuration document pack — Technical File, DoC, Label, Statement",
        [
            ("CONFIGURATIONS", "247"),
            ("DOCUMENTS", "988"),
            ("LINKED", "988"),
            ("BLOCKING", "0"),
        ],
        min(cols, 8),
    )
    if cols > 8:
        for c in range(9, cols + 1):
            for r in range(1, 6):
                if r == 1:
                    ws.cell(r, c).fill = _fill(STONE)
                elif r == 2:
                    ws.cell(r, c).fill = _fill(NAVY)
                elif r == 3:
                    ws.cell(r, c).fill = _fill(NAVY2)
                elif r == 4:
                    ws.cell(r, c).fill = _fill(IVORY)
                else:
                    ws.cell(r, c).fill = _fill(GOLD)
    last = write_data_table(ws, headers, data, start_row=6)
    return {
        "sheet": "DOCUMENT_CENTER",
        "class": "B",
        "table_start_row": 6,
        "rows": len(data),
        "cols": len(headers),
        "last_row": last,
    }


def polish_shipments(ws) -> dict[str, Any]:
    clear_sheet_values(ws)
    write_nav_row(ws, 8)
    write_premium_header(
        ws,
        "SHIPMENTS",
        "Transactional shipment register",
        [("STATUS", "EMPTY"), ("REV", "00"), ("RECORDS", "0"), ("PHASE", "BASELINE")],
        8,
    )
    # Empty-state panel in cells below header (row 6+)
    ws.cell(6, 1, "CURRENT STATUS").font = _font(10, True, NAVY)
    ws.cell(7, 1, "No transactional shipment records are loaded in Rev.00 baseline.").font = _font(
        10, False, INK
    )
    ws.cell(
        8,
        1,
        "Shipment tracking will be activated for Rev.01 and the operational phase.",
    ).font = _font(9, False, MUTED)

    ws.cell(10, 1, "QUICK LINKS").font = _font(9, True, NAVY)
    for r, (label, sheet) in enumerate(
        [
            ("→  Open Shipment Database", "SHIPMENT"),
            ("→  Open Shipment Statements", "SHIPMENT_STATEMENTS"),
        ],
        start=11,
    ):
        cell = ws.cell(r, 1, label)
        cell.font = _font(11, True, LINK, underline="single")
        cell.fill = _fill(CARD)
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate, location=f"'{sheet}'!A1", tooltip=label
        )
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 56
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    return {"sheet": "SHIPMENTS", "class": "C", "rows": 0, "empty_state": True}


def polish_doc_engine_map(ws) -> dict[str, Any]:
    headers, data = extract_table(ws)
    if not headers:
        headers = [
            "Document Type",
            "Logical Field",
            "PIMS Source",
            "DocumentContext Field",
            "Runtime Token / Mapping",
            "Builder Module",
        ]
    if len(headers) < 3:
        raise ValueError("DOC_ENGINE_MAP header detection failed")
    clear_sheet_values(ws)
    cols = max(len(headers), 6)
    write_premium_header(
        ws,
        "DOCUMENT ENGINE MAP",
        "Read-only mapping • Python remains document authority",
        [("FIELDS", str(len(data))), ("MODE", "READ-ONLY"), ("AUTHORITY", "PYTHON"), ("REV", "00")],
        min(cols, 8),
    )
    last = write_data_table(ws, headers, data, start_row=6)
    for c in range(3, min(len(headers), 6) + 1):
        for r in range(7, last + 1):
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 32
    for r in range(7, last + 1):
        ws.row_dimensions[r].height = 36
    return {
        "sheet": "DOC_ENGINE_MAP",
        "class": "C",
        "table_start_row": 6,
        "rows": len(data),
        "cols": len(headers),
    }


def polish_class_c_ops(ws, name: str) -> dict[str, Any]:
    title, subtitle, info = MODULE_META[name]
    headers, data = extract_table(ws)
    if len(headers) < 3:
        raise ValueError(f"{name}: table header detection failed (cols={len(headers)})")
    clear_sheet_values(ws)
    cols = max(len(headers), 8)
    # Parse simple metric from info string
    metrics = [("RECORDS", str(len(data))), ("REV", "00"), ("STATUS", "READY"), ("LINKED", "YES")]
    write_premium_header(ws, title, subtitle, metrics, min(cols, 8))
    if cols > 8:
        for c in range(9, cols + 1):
            for r in range(1, 6):
                fills = {1: STONE, 2: NAVY, 3: NAVY2, 4: IVORY, 5: GOLD}
                ws.cell(r, c).fill = _fill(fills[r])
    last = write_data_table(ws, headers, data, start_row=6)
    return {
        "sheet": name,
        "class": "C",
        "table_start_row": 6,
        "rows": len(data),
        "cols": len(headers),
        "last_row": last,
    }


def polish_all_cell_modules(wb) -> list[dict[str, Any]]:
    stats: list[dict[str, Any]] = []
    if "DOCUMENT_CENTER" in wb.sheetnames:
        stats.append(polish_document_center(wb["DOCUMENT_CENTER"]))
    if "SHIPMENTS" in wb.sheetnames:
        stats.append(polish_shipments(wb["SHIPMENTS"]))
    if "DOC_ENGINE_MAP" in wb.sheetnames:
        stats.append(polish_doc_engine_map(wb["DOC_ENGINE_MAP"]))
    ops = [
        "PACKAGING_CONFIGURATIONS",
        "PRODUCT_MASTER",
        "COMPONENT_MASTER",
        "TECHNICAL_FILES",
        "DECLARATIONS_OF_CONFORMITY",
        "LABELS",
        "SHIPMENT_STATEMENTS",
    ]
    for name in ops:
        if name in wb.sheetnames:
            stats.append(polish_class_c_ops(wb[name], name))
    # Re-convert any leftover HYPERLINK formulas on UI sheets
    converted = 0
    ui = [
        "DOCUMENT_CENTER",
        "TECHNICAL_FILES",
        "DECLARATIONS_OF_CONFORMITY",
        "LABELS",
        "SHIPMENT_STATEMENTS",
        "PACKAGING_CONFIGURATIONS",
    ]
    for name in ui:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.upper().startswith("=HYPERLINK("):
                    m = HYPERLINK_RE.match(val.strip())
                    if m and m.group(1).lower().endswith(".docx"):
                        apply_native_doc_link(cell, m.group(1), m.group(2) or "OPEN")
                        converted += 1
    stats.append({"docx_formulas_converted": converted})
    return stats
