"""
Production seed data for lookup / reference tables.

Master transactional data remains empty for customer entry.
Lookup seeds make FK dropdowns and engines operational.
"""

from __future__ import annotations

from datetime import datetime, timezone

from openpyxl.worksheet.table import Table
from openpyxl.workbook.workbook import Workbook

from .sheet_builder import excel_col_letter

NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


SEED_ROWS: dict[str, list[tuple]] = {
    "LKP_STATUS": [
        (1, "DRAFT", "Draft", "MASTER", True, 1),
        (2, "ACTIVE", "Active", "MASTER", True, 2),
        (3, "OBSOLETE", "Obsolete", "MASTER", False, 3),
        (4, "CONFIRMED", "Confirmed", "SHIPMENT", False, 4),
        (5, "CANCELLED", "Cancelled", "SHIPMENT", False, 5),
        (6, "APPROVED_DOC", "Approved (Document)", "DOCUMENT", False, 6),
        (7, "REVOKED", "Revoked", "DOCUMENT", False, 7),
        (8, "APPROVED_STM", "Approved (Statement)", "STATEMENT", False, 8),
    ],
    "LKP_UOM": [
        (1, "G", "Gram", "MASS"),
        (2, "KG", "Kilogram", "MASS"),
        (3, "MM", "Millimeter", "LENGTH"),
        (4, "PCS", "Pieces", "COUNT"),
        (5, "PAL", "Pallet", "COUNT"),
        (6, "TEU", "TEU", "VOLUME"),
    ],
    "LKP_PACKAGING_LEVEL": [
        (1, "PRIMARY", "Primary", 1),
        (2, "SECONDARY", "Secondary", 2),
        (3, "TERTIARY", "Tertiary", 3),
    ],
    "LKP_PACKAGING_FUNCTION": [
        (1, "SALES", "Sales Packaging"),
        (2, "GROUPED", "Grouped Packaging"),
        (3, "TRANSPORT", "Transport Packaging"),
    ],
    "LKP_COMPONENT_TYPE": [
        (1, "CARTON", "Carton", 2),
        (2, "DIVIDER", "Divider", 2),
        (3, "FILM", "Film", 3),
        (4, "STRAP", "Strap", 3),
        (5, "LABEL", "Label", 1),
        (6, "PALLET", "Pallet", 3),
        (7, "BAG", "Bag", 1),
        (8, "CAP", "Cap", 1),
        (9, "DUNNAGE", "Dunnage", 3),
        (10, "AIRBAG", "Airbag", 3),
        (11, "DESICCANT", "Desiccant", 3),
        (12, "CONTAINER_LINER", "Container Liner", 3),
        (13, "LASHING", "Lashing", 3),
        (14, "OTHER", "Other", None),
    ],
    "LKP_MATERIAL_FAMILY": [
        (1, "PLASTIC", "Plastic"),
        (2, "PAPER", "Paper / Board"),
        (3, "WOOD", "Wood"),
        (4, "METAL", "Metal"),
        (5, "COMPOSITE", "Composite"),
        (6, "OTHER", "Other"),
    ],
    "LKP_PPWR_MATERIAL_CATEGORY": [
        (1, "PLASTIC", "Plastic"),
        (2, "PAPER_CARDBOARD", "Paper and cardboard"),
        (3, "WOOD", "Wood"),
        (4, "FERROUS", "Ferrous metals"),
        (5, "ALUMINIUM", "Aluminium"),
        (6, "GLASS", "Glass"),
        (7, "OTHER", "Other"),
    ],
    "LKP_MATERIAL": [
        (1, "LDPE", "LDPE", 1, 1, False, None),
        (2, "HDPE", "HDPE", 1, 1, False, None),
        (3, "PP", "Polypropylene", 1, 1, False, None),
        (4, "PET", "PET", 1, 1, False, None),
        (5, "PAPER_BOARD", "Paperboard", 2, 2, False, None),
        (6, "CORRUGATED", "Corrugated board", 2, 2, False, None),
        (7, "WOOD", "Wood", 3, 3, False, None),
        (8, "STEEL", "Steel", 4, 4, False, None),
        (9, "ALUMINIUM", "Aluminium", 4, 5, False, None),
        (10, "COMPOSITE", "Composite", 5, 7, True, None),
    ],
    "LKP_RECYCLABILITY_CLASS": [
        (1, "A", "Class A"),
        (2, "B", "Class B"),
        (3, "C", "Class C"),
        (4, "NOT_ASSESSED", "Not assessed"),
    ],
    "LKP_OWNERSHIP_TYPE": [
        (1, "DISPOSABLE", "Disposable"),
        (2, "COMPANY_RETURNABLE", "Company returnable"),
        (3, "CUSTOMER_RETURNABLE", "Customer returnable"),
        (4, "POOL", "Pool packaging"),
    ],
    "LKP_TRANSPORT_UNIT_TYPE": [
        (1, "PIECE_PACK", "Piece pack"),
        (2, "PALLET", "Pallet"),
        (3, "CONTAINER", "Container"),
        (4, "TRUCK", "Truck"),
        (5, "STILLAGE", "Stillage"),
    ],
    "LKP_LINE_ROLE": [
        (1, "BASE", "Base"),
        (2, "PALLET", "Pallet"),
        (3, "WRAP", "Wrap"),
        (4, "CORNER", "Corner"),
        (5, "LAYER_PAD", "Layer pad"),
        (6, "LABEL", "Label"),
        (7, "DUNNAGE", "Dunnage"),
        (8, "AIRBAG", "Airbag"),
        (9, "LINER", "Liner"),
        (10, "LASHING", "Lashing"),
        (11, "OTHER", "Other"),
    ],
    "LKP_PRODUCT_CATEGORY": [
        (1, "STARTER_BATTERY", "Starter Battery"),
        (2, "INDUSTRIAL_BATTERY", "Industrial Battery"),
        (3, "OTHER", "Other"),
    ],
    "LKP_COUNTRY": [
        (1, "TR", "TUR", "Türkiye", False),
        (2, "DE", "DEU", "Germany", True),
        (3, "FR", "FRA", "France", True),
        (4, "IT", "ITA", "Italy", True),
        (5, "ES", "ESP", "Spain", True),
        (6, "NL", "NLD", "Netherlands", True),
        (7, "BE", "BEL", "Belgium", True),
        (8, "AT", "AUT", "Austria", True),
        (9, "PL", "POL", "Poland", True),
        (10, "GB", "GBR", "United Kingdom", False),
    ],
    "LKP_TRANSPORT_MODE": [
        (1, "ROAD", "Road"),
        (2, "SEA", "Sea"),
        (3, "AIR", "Air"),
        (4, "RAIL", "Rail"),
        (5, "MULTI", "Multimodal"),
    ],
    "LKP_INCOTERM": [
        (1, "EXW", "Ex Works"),
        (2, "FCA", "Free Carrier"),
        (3, "FOB", "Free On Board"),
        (4, "CIF", "Cost Insurance Freight"),
        (5, "DAP", "Delivered At Place"),
        (6, "DDP", "Delivered Duty Paid"),
    ],
    "LKP_STATEMENT_TYPE": [
        (1, "ANNUAL_MARKET", "Annual market statement"),
        (2, "QUARTERLY", "Quarterly statement"),
        (3, "INTERNAL_AUDIT", "Internal audit"),
    ],
    "LKP_DOCUMENT_TYPE": [
        (1, "TECH_FILE", "Technical File"),
        (2, "TEST_REPORT", "Test Report"),
        (3, "SPEC", "Specification"),
        (4, "DOC_PDF", "Declaration PDF"),
        (5, "CERTIFICATE", "Certificate"),
        (6, "OTHER", "Other"),
    ],
    "LKP_SCENARIO_TYPE": [
        (1, "ENGINEERING_DEFAULT", "Engineering default"),
        (2, "CUSTOMER", "Customer scenario"),
        (3, "EXPORT", "Export scenario"),
        (4, "INTERNAL", "Internal"),
        (5, "SAMPLE", "Sample"),
    ],
    "LEGAL_ENTITY": [
        (1, "INCI_AKU", "İnci Akü Sanayi ve Ticaret A.Ş.", 1, 2, None),
    ],
    "PERSON": [
        (1, "P-COMPLIANCE", "PPWR Compliance Officer", "compliance@inciaku.example", "Compliance Officer", 2),
    ],
    "PLANT": [
        (1, "MANISA", "Manisa Plant", 1, 1, 2, "PLANT-01"),
    ],
}


def _resize_table(ws, table_name: str, headers: list, rows: list[tuple]) -> None:
    # Remove existing table registration then rewrite
    if table_name in ws.tables:
        del ws.tables[table_name]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            if c_idx <= len(headers):
                ws.cell(row=r_idx, column=c_idx, value=value)
    last_col = excel_col_letter(len(headers))
    last_row = 1 + len(rows)
    ws.add_table(Table(displayName=table_name, ref=f"A1:{last_col}{last_row}"))


def apply_seed_data(workbook: Workbook) -> list[str]:
    """Populate seeded sheets; return list of seeded sheet names."""
    seeded: list[str] = []
    for sheet_name, rows in SEED_ROWS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = [c.value for c in ws[1] if c.value]
        if not headers:
            continue
        # Pad/truncate rows to header width
        normalized: list[tuple] = []
        for row in rows:
            values = list(row)
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))
            normalized.append(tuple(values[: len(headers)]))
        _resize_table(ws, f"T_{sheet_name}", headers, normalized)
        seeded.append(sheet_name)
    return seeded
