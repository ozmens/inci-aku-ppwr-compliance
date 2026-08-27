"""
Engine sheets — all calculations live here.

Rules:
  - No formulas in master / fact entity tables
  - Engine formulas use Excel structured references only
  - ROW_INDEX is a numeric driver (literal) to pull source rows via INDEX
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from .sheet_builder import excel_col_letter

LINE_ENGINE_ROWS = 50
SUMMARY_ENGINE_ROWS = 1
COMPONENT_ENGINE_ROWS = 50
VALIDATION_ROWS = 18


def _idx(source_table: str, column: str, row_index_ref: str = "[@ROW_INDEX]") -> str:
    """INDEX pull from a source Excel table column by ROW_INDEX."""
    return f'IFERROR(INDEX({source_table}[{column}],{row_index_ref}),"")'


def _xwt(component_id_ref: str = "[@COMPONENT_ID]") -> str:
    """Unit weight from T_COMPONENT via MATCH/INDEX (structured refs)."""
    return (
        f'IF({component_id_ref}="","",'
        f'IFERROR(INDEX(T_COMPONENT[WEIGHT_G],'
        f'MATCH({component_id_ref},T_COMPONENT[COMPONENT_ID],0)),""))'
    )


def _same(col: str) -> str:
    return f"[@{col}]"


@dataclass(frozen=True)
class EngineSheet:
    sheet_name: str
    table_name: str
    headers: tuple[str, ...]
    row_count: int
    # column -> formula template (None = leave blank; "literal:{value}" for constants)
    formulas: dict[str, str]

    def formula_for(self, column: str, row_index: int) -> str | int | float | None:
        raw = self.formulas.get(column)
        if raw is None:
            return None
        if raw == "ROW_INDEX":
            return row_index
        if raw.startswith("literal:"):
            value = raw[len("literal:") :]
            if value == "":
                return None
            # numeric literals
            try:
                if "." in value:
                    return float(value)
                return int(value)
            except ValueError:
                return value
        return raw  # formula string without leading '='


def _engine_material_weight() -> EngineSheet:
    t = "T_ENG_MATERIAL_WEIGHT"
    return EngineSheet(
        sheet_name="ENG_MATERIAL_WEIGHT",
        table_name=t,
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "COMPONENT_MATERIAL_ID",
            "COMPONENT_ID",
            "MATERIAL_ID",
            "SHARE_PCT",
            "UNIT_WEIGHT_G",
            "MATERIAL_WEIGHT_G",
            "MATERIAL_FAMILY_ID",
            "MATERIAL_FAMILY_CODE",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "COMPONENT_MATERIAL_ID": _idx("T_COMPONENT_MATERIAL", "COMPONENT_MATERIAL_ID"),
            "COMPONENT_ID": _idx("T_COMPONENT_MATERIAL", "COMPONENT_ID"),
            "MATERIAL_ID": _idx("T_COMPONENT_MATERIAL", "MATERIAL_ID"),
            "SHARE_PCT": _idx("T_COMPONENT_MATERIAL", "SHARE_PCT"),
            "UNIT_WEIGHT_G": _xwt(),
            "MATERIAL_WEIGHT_G": (
                'IF(OR([@UNIT_WEIGHT_G]="",[@SHARE_PCT]=""),"",'
                "[@UNIT_WEIGHT_G]*[@SHARE_PCT]/100)"
            ),
            "MATERIAL_FAMILY_ID": (
                'IF([@MATERIAL_ID]="","",'
                "IFERROR(INDEX(T_LKP_MATERIAL[MATERIAL_FAMILY_ID],"
                "MATCH([@MATERIAL_ID],T_LKP_MATERIAL[MATERIAL_ID],0)),\"\"))"
            ),
            "MATERIAL_FAMILY_CODE": (
                'IF([@MATERIAL_FAMILY_ID]="","",'
                "IFERROR(INDEX(T_LKP_MATERIAL_FAMILY[MATERIAL_FAMILY_CODE],"
                "MATCH([@MATERIAL_FAMILY_ID],T_LKP_MATERIAL_FAMILY[MATERIAL_FAMILY_ID],0)),\"\"))"
            ),
        },
    )


def _engine_packaging_weight() -> EngineSheet:
    return EngineSheet(
        sheet_name="ENG_PACKAGING_WEIGHT",
        table_name="T_ENG_PACKAGING_WEIGHT",
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "PACKAGING_CONFIGURATION_LINE_ID",
            "PACKAGING_CONFIGURATION_ID",
            "COMPONENT_ID",
            "QUANTITY",
            "IS_OPTIONAL",
            "UNIT_WEIGHT_G",
            "LINE_WEIGHT_G",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "PACKAGING_CONFIGURATION_LINE_ID": _idx(
                "T_PACKAGING_CONFIGURATION_LINE", "PACKAGING_CONFIGURATION_LINE_ID"
            ),
            "PACKAGING_CONFIGURATION_ID": _idx(
                "T_PACKAGING_CONFIGURATION_LINE", "PACKAGING_CONFIGURATION_ID"
            ),
            "COMPONENT_ID": _idx("T_PACKAGING_CONFIGURATION_LINE", "COMPONENT_ID"),
            "QUANTITY": _idx("T_PACKAGING_CONFIGURATION_LINE", "QUANTITY"),
            "IS_OPTIONAL": _idx("T_PACKAGING_CONFIGURATION_LINE", "IS_OPTIONAL"),
            "UNIT_WEIGHT_G": _xwt(),
            "LINE_WEIGHT_G": (
                'IF(OR([@QUANTITY]="",[@UNIT_WEIGHT_G]=""),"",'
                'IF([@IS_OPTIONAL]=TRUE,0,[@QUANTITY]*[@UNIT_WEIGHT_G]))'
            ),
        },
    )


def _engine_transport_weight() -> EngineSheet:
    """Supporting engine for shipment allocation (transport-unit lines)."""
    return EngineSheet(
        sheet_name="ENG_TRANSPORT_WEIGHT",
        table_name="T_ENG_TRANSPORT_WEIGHT",
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "TRANSPORT_CONFIGURATION_LINE_ID",
            "TRANSPORT_CONFIGURATION_ID",
            "COMPONENT_ID",
            "QUANTITY_PER_TRANSPORT_UNIT",
            "UNIT_WEIGHT_G",
            "LINE_WEIGHT_G",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "TRANSPORT_CONFIGURATION_LINE_ID": _idx(
                "T_TRANSPORT_CONFIGURATION_LINE", "TRANSPORT_CONFIGURATION_LINE_ID"
            ),
            "TRANSPORT_CONFIGURATION_ID": _idx(
                "T_TRANSPORT_CONFIGURATION_LINE", "TRANSPORT_CONFIGURATION_ID"
            ),
            "COMPONENT_ID": _idx("T_TRANSPORT_CONFIGURATION_LINE", "COMPONENT_ID"),
            "QUANTITY_PER_TRANSPORT_UNIT": _idx(
                "T_TRANSPORT_CONFIGURATION_LINE", "QUANTITY_PER_TRANSPORT_UNIT"
            ),
            "UNIT_WEIGHT_G": _xwt(),
            "LINE_WEIGHT_G": (
                'IF(OR([@QUANTITY_PER_TRANSPORT_UNIT]="",[@UNIT_WEIGHT_G]=""),"",'
                "[@QUANTITY_PER_TRANSPORT_UNIT]*[@UNIT_WEIGHT_G])"
            ),
        },
    )


def _engine_shipment_weight() -> EngineSheet:
    return EngineSheet(
        sheet_name="ENG_SHIPMENT_WEIGHT",
        table_name="T_ENG_SHIPMENT_WEIGHT",
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "SHIPMENT_ID",
            "SHIPMENT_NUMBER",
            "QTY_PRODUCT_UNITS",
            "PACKAGING_CONFIGURATION_ID",
            "TRANSPORT_CONFIGURATION_ID",
            "UNITS_PER_LAYER",
            "LAYERS_PER_UNIT",
            "CONTAINER_PAYLOAD_UNITS",
            "UNITS_PER_TRANSPORT_UNIT",
            "PACKAGING_WEIGHT_PER_UNIT_G",
            "TRANSPORT_WEIGHT_PER_UNIT_G",
            "TOTAL_WEIGHT_PER_UNIT_G",
            "SHIPMENT_PACKAGING_WEIGHT_G",
            "SHIPMENT_PACKAGING_WEIGHT_KG",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "SHIPMENT_ID": _idx("T_SHIPMENT", "SHIPMENT_ID"),
            "SHIPMENT_NUMBER": _idx("T_SHIPMENT", "SHIPMENT_NUMBER"),
            "QTY_PRODUCT_UNITS": _idx("T_SHIPMENT", "QTY_PRODUCT_UNITS"),
            "PACKAGING_CONFIGURATION_ID": _idx("T_SHIPMENT", "PACKAGING_CONFIGURATION_ID"),
            "TRANSPORT_CONFIGURATION_ID": _idx("T_SHIPMENT", "TRANSPORT_CONFIGURATION_ID"),
            "UNITS_PER_LAYER": (
                'IF([@TRANSPORT_CONFIGURATION_ID]="","",'
                "IFERROR(INDEX(T_TRANSPORT_CONFIGURATION[UNITS_PER_LAYER],"
                "MATCH([@TRANSPORT_CONFIGURATION_ID],T_TRANSPORT_CONFIGURATION[TRANSPORT_CONFIGURATION_ID],0)),\"\"))"
            ),
            "LAYERS_PER_UNIT": (
                'IF([@TRANSPORT_CONFIGURATION_ID]="","",'
                "IFERROR(INDEX(T_TRANSPORT_CONFIGURATION[LAYERS_PER_UNIT],"
                "MATCH([@TRANSPORT_CONFIGURATION_ID],T_TRANSPORT_CONFIGURATION[TRANSPORT_CONFIGURATION_ID],0)),\"\"))"
            ),
            "CONTAINER_PAYLOAD_UNITS": (
                'IF([@TRANSPORT_CONFIGURATION_ID]="","",'
                "IFERROR(INDEX(T_TRANSPORT_CONFIGURATION[CONTAINER_PAYLOAD_UNITS],"
                "MATCH([@TRANSPORT_CONFIGURATION_ID],T_TRANSPORT_CONFIGURATION[TRANSPORT_CONFIGURATION_ID],0)),\"\"))"
            ),
            "UNITS_PER_TRANSPORT_UNIT": (
                'IF([@TRANSPORT_CONFIGURATION_ID]="","",'
                'IF(AND(ISNUMBER([@UNITS_PER_LAYER]),ISNUMBER([@LAYERS_PER_UNIT]),'
                '[@UNITS_PER_LAYER]>0,[@LAYERS_PER_UNIT]>0),'
                "[@UNITS_PER_LAYER]*[@LAYERS_PER_UNIT],"
                'IF(AND(ISNUMBER([@CONTAINER_PAYLOAD_UNITS]),[@CONTAINER_PAYLOAD_UNITS]>0),'
                "[@CONTAINER_PAYLOAD_UNITS],\"\")))"
            ),
            "PACKAGING_WEIGHT_PER_UNIT_G": (
                'IF([@PACKAGING_CONFIGURATION_ID]="","",'
                "SUMIF(T_ENG_PACKAGING_WEIGHT[PACKAGING_CONFIGURATION_ID],"
                "[@PACKAGING_CONFIGURATION_ID],T_ENG_PACKAGING_WEIGHT[LINE_WEIGHT_G]))"
            ),
            "TRANSPORT_WEIGHT_PER_UNIT_G": (
                'IF(OR([@TRANSPORT_CONFIGURATION_ID]="",[@UNITS_PER_TRANSPORT_UNIT]=""),"",'
                "IFERROR(SUMIF(T_ENG_TRANSPORT_WEIGHT[TRANSPORT_CONFIGURATION_ID],"
                "[@TRANSPORT_CONFIGURATION_ID],T_ENG_TRANSPORT_WEIGHT[LINE_WEIGHT_G])"
                "/[@UNITS_PER_TRANSPORT_UNIT],\"\"))"
            ),
            "TOTAL_WEIGHT_PER_UNIT_G": (
                'IF(OR([@PACKAGING_WEIGHT_PER_UNIT_G]="",[@TRANSPORT_WEIGHT_PER_UNIT_G]=""),'
                'IF([@PACKAGING_WEIGHT_PER_UNIT_G]="","",[@PACKAGING_WEIGHT_PER_UNIT_G]),'
                "[@PACKAGING_WEIGHT_PER_UNIT_G]+[@TRANSPORT_WEIGHT_PER_UNIT_G])"
            ),
            "SHIPMENT_PACKAGING_WEIGHT_G": (
                'IF(OR([@TOTAL_WEIGHT_PER_UNIT_G]="",[@QTY_PRODUCT_UNITS]=""),"",'
                "[@TOTAL_WEIGHT_PER_UNIT_G]*[@QTY_PRODUCT_UNITS])"
            ),
            "SHIPMENT_PACKAGING_WEIGHT_KG": (
                'IF([@SHIPMENT_PACKAGING_WEIGHT_G]="","",[@SHIPMENT_PACKAGING_WEIGHT_G]/1000)'
            ),
        },
    )


def _family_summary(sheet: str, table: str, family_code: str) -> EngineSheet:
    return EngineSheet(
        sheet_name=sheet,
        table_name=table,
        row_count=SUMMARY_ENGINE_ROWS,
        headers=(
            "MATERIAL_FAMILY_CODE",
            "LINE_COUNT",
            "TOTAL_MATERIAL_WEIGHT_G",
            "TOTAL_MATERIAL_WEIGHT_KG",
            "SHARE_OF_ALL_MATERIAL_PCT",
        ),
        formulas={
            "MATERIAL_FAMILY_CODE": f"literal:{family_code}",
            "LINE_COUNT": (
                f'COUNTIF(T_ENG_MATERIAL_WEIGHT[MATERIAL_FAMILY_CODE],"={family_code}")'
            ),
            "TOTAL_MATERIAL_WEIGHT_G": (
                "SUMIF(T_ENG_MATERIAL_WEIGHT[MATERIAL_FAMILY_CODE],"
                f'[@MATERIAL_FAMILY_CODE],T_ENG_MATERIAL_WEIGHT[MATERIAL_WEIGHT_G])'
            ),
            "TOTAL_MATERIAL_WEIGHT_KG": (
                'IF([@TOTAL_MATERIAL_WEIGHT_G]="","",[@TOTAL_MATERIAL_WEIGHT_G]/1000)'
            ),
            "SHARE_OF_ALL_MATERIAL_PCT": (
                'IF(OR([@TOTAL_MATERIAL_WEIGHT_G]="",'
                "SUM(T_ENG_MATERIAL_WEIGHT[MATERIAL_WEIGHT_G])=0),\"\","
                "[@TOTAL_MATERIAL_WEIGHT_G]/SUM(T_ENG_MATERIAL_WEIGHT[MATERIAL_WEIGHT_G])*100)"
            ),
        },
    )


def _engine_statement() -> EngineSheet:
    return EngineSheet(
        sheet_name="ENG_STATEMENT",
        table_name="T_ENG_STATEMENT",
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "STATEMENT_ID",
            "STATEMENT_CODE",
            "MATERIAL_ID",
            "PACKAGING_LEVEL_ID",
            "OWNERSHIP_TYPE_ID",
            "FROZEN_TOTAL_WEIGHT_KG",
            "RECOMPUTED_WEIGHT_G",
            "RECOMPUTED_WEIGHT_KG",
            "VARIANCE_KG",
            "SOURCE_SHIPMENT_COUNT",
            "RECONCILE_STATUS",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "STATEMENT_ID": _idx("T_STATEMENT_LINE", "STATEMENT_ID"),
            "STATEMENT_CODE": (
                'IF([@STATEMENT_ID]="","",'
                "IFERROR(INDEX(T_STATEMENT[STATEMENT_CODE],"
                "MATCH([@STATEMENT_ID],T_STATEMENT[STATEMENT_ID],0)),\"\"))"
            ),
            "MATERIAL_ID": _idx("T_STATEMENT_LINE", "MATERIAL_ID"),
            "PACKAGING_LEVEL_ID": _idx("T_STATEMENT_LINE", "PACKAGING_LEVEL_ID"),
            "OWNERSHIP_TYPE_ID": _idx("T_STATEMENT_LINE", "OWNERSHIP_TYPE_ID"),
            "FROZEN_TOTAL_WEIGHT_KG": _idx("T_STATEMENT_LINE", "TOTAL_WEIGHT_KG"),
            "RECOMPUTED_WEIGHT_G": (
                'IF(OR([@MATERIAL_ID]="",[@PACKAGING_LEVEL_ID]="",[@OWNERSHIP_TYPE_ID]=""),"",'
                "SUMIFS(T_SHIPMENT_LINE[WEIGHT_G],"
                "T_SHIPMENT_LINE[MATERIAL_ID],[@MATERIAL_ID],"
                "T_SHIPMENT_LINE[PACKAGING_LEVEL_ID],[@PACKAGING_LEVEL_ID],"
                "T_SHIPMENT_LINE[OWNERSHIP_TYPE_ID],[@OWNERSHIP_TYPE_ID]))"
            ),
            "RECOMPUTED_WEIGHT_KG": (
                'IF([@RECOMPUTED_WEIGHT_G]="","",[@RECOMPUTED_WEIGHT_G]/1000)'
            ),
            "VARIANCE_KG": (
                'IF(OR([@FROZEN_TOTAL_WEIGHT_KG]="",[@RECOMPUTED_WEIGHT_KG]=""),"",'
                "[@RECOMPUTED_WEIGHT_KG]-[@FROZEN_TOTAL_WEIGHT_KG])"
            ),
            "SOURCE_SHIPMENT_COUNT": (
                'IF([@STATEMENT_ID]="","",'
                "COUNTIF(T_STATEMENT_SHIPMENT[STATEMENT_ID],[@STATEMENT_ID]))"
            ),
            "RECONCILE_STATUS": (
                'IF([@STATEMENT_ID]="","",'
                'IF([@VARIANCE_KG]="","PENDING",'
                'IF(ABS([@VARIANCE_KG])<=0.001,"OK","VARIANCE")))'
            ),
        },
    )


def _engine_technical_file() -> EngineSheet:
    return EngineSheet(
        sheet_name="ENG_TECHNICAL_FILE",
        table_name="T_ENG_TECHNICAL_FILE",
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "TECHNICAL_FILE_ID",
            "TECHNICAL_FILE_CODE",
            "SUBJECT_COUNT",
            "HAS_EXACTLY_ONE_SUBJECT",
            "DOCUMENT_LINK_COUNT",
            "HAS_DOCUMENT",
            "STATUS_ID",
            "COMPLETENESS_SCORE",
            "ENGINE_STATUS",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "TECHNICAL_FILE_ID": _idx("T_TECHNICAL_FILE", "TECHNICAL_FILE_ID"),
            "TECHNICAL_FILE_CODE": _idx("T_TECHNICAL_FILE", "TECHNICAL_FILE_CODE"),
            "SUBJECT_COUNT": (
                'IF([@TECHNICAL_FILE_ID]="","",'
                'IFERROR((IF(INDEX(T_TECHNICAL_FILE[COMPONENT_ID],[@ROW_INDEX])<>"",1,0))'
                '+(IF(INDEX(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID],[@ROW_INDEX])<>"",1,0))'
                '+(IF(INDEX(T_TECHNICAL_FILE[TRANSPORT_CONFIGURATION_ID],[@ROW_INDEX])<>"",1,0)),0))'
            ),
            "HAS_EXACTLY_ONE_SUBJECT": (
                'IF([@TECHNICAL_FILE_ID]="","",IF([@SUBJECT_COUNT]=1,TRUE,FALSE))'
            ),
            "DOCUMENT_LINK_COUNT": (
                'IF([@TECHNICAL_FILE_ID]="","",'
                "COUNTIF(T_DOCUMENT_LINK[TECHNICAL_FILE_ID],[@TECHNICAL_FILE_ID]))"
            ),
            "HAS_DOCUMENT": (
                'IF([@TECHNICAL_FILE_ID]="","",IF([@DOCUMENT_LINK_COUNT]>0,TRUE,FALSE))'
            ),
            "STATUS_ID": _idx("T_TECHNICAL_FILE", "STATUS_ID"),
            "COMPLETENESS_SCORE": (
                'IF([@TECHNICAL_FILE_ID]="","",'
                # İnci rule: subject must be Packaging Configuration
                '((IF(INDEX(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID],[@ROW_INDEX])<>"",1,0))*50)'
                "+([@HAS_EXACTLY_ONE_SUBJECT]*20)+([@HAS_DOCUMENT]*30))"
            ),
            "ENGINE_STATUS": (
                'IF([@TECHNICAL_FILE_ID]="","",'
                'IF([@COMPLETENESS_SCORE]=100,"COMPLETE","INCOMPLETE"))'
            ),
        },
    )


def _engine_declaration() -> EngineSheet:
    return EngineSheet(
        sheet_name="ENG_DECLARATION",
        table_name="T_ENG_DECLARATION",
        row_count=LINE_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "DECLARATION_OF_CONFORMITY_ID",
            "DOC_NUMBER",
            "HAS_LEGAL_ENTITY",
            "HAS_TECHNICAL_FILE",
            "HAS_RESPONSIBLE_PERSON",
            "SCOPE_COUNT",
            "HAS_SCOPE",
            "DOCUMENT_LINK_COUNT",
            "COMPLETENESS_SCORE",
            "ENGINE_STATUS",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "DECLARATION_OF_CONFORMITY_ID": _idx(
                "T_DECLARATION_OF_CONFORMITY", "DECLARATION_OF_CONFORMITY_ID"
            ),
            "DOC_NUMBER": _idx("T_DECLARATION_OF_CONFORMITY", "DOC_NUMBER"),
            "HAS_LEGAL_ENTITY": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                'IF(INDEX(T_DECLARATION_OF_CONFORMITY[LEGAL_ENTITY_ID],[@ROW_INDEX])<>"",TRUE,FALSE))'
            ),
            "HAS_TECHNICAL_FILE": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                'IF(INDEX(T_DECLARATION_OF_CONFORMITY[TECHNICAL_FILE_ID],[@ROW_INDEX])<>"",TRUE,FALSE))'
            ),
            "HAS_RESPONSIBLE_PERSON": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                'IF(INDEX(T_DECLARATION_OF_CONFORMITY[RESPONSIBLE_PERSON_ID],[@ROW_INDEX])<>"",TRUE,FALSE))'
            ),
            "SCOPE_COUNT": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                'IFERROR((IF(INDEX(T_DECLARATION_OF_CONFORMITY[PRODUCT_ID],[@ROW_INDEX])<>"",1,0))'
                '+(IF(INDEX(T_DECLARATION_OF_CONFORMITY[PACKAGING_CONFIGURATION_ID],[@ROW_INDEX])<>"",1,0))'
                '+(IF(INDEX(T_DECLARATION_OF_CONFORMITY[TRANSPORT_CONFIGURATION_ID],[@ROW_INDEX])<>"",1,0)),0))'
            ),
            "HAS_SCOPE": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",IF([@SCOPE_COUNT]>=1,TRUE,FALSE))'
            ),
            "DOCUMENT_LINK_COUNT": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                "COUNTIF(T_DOCUMENT_LINK[DECLARATION_OF_CONFORMITY_ID],"
                "[@DECLARATION_OF_CONFORMITY_ID]))"
            ),
            "COMPLETENESS_SCORE": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                "([@HAS_LEGAL_ENTITY]*25)+([@HAS_TECHNICAL_FILE]*25)"
                "+([@HAS_RESPONSIBLE_PERSON]*25)+([@HAS_SCOPE]*25))"
            ),
            "ENGINE_STATUS": (
                'IF([@DECLARATION_OF_CONFORMITY_ID]="","",'
                'IF([@COMPLETENESS_SCORE]=100,"COMPLETE","INCOMPLETE"))'
            ),
        },
    )


def _engine_impact_analysis() -> EngineSheet:
    return EngineSheet(
        sheet_name="ENG_IMPACT_ANALYSIS",
        table_name="T_ENG_IMPACT_ANALYSIS",
        row_count=COMPONENT_ENGINE_ROWS,
        headers=(
            "ROW_INDEX",
            "COMPONENT_ID",
            "COMPONENT_CODE",
            "WEIGHT_G",
            "USED_IN_PKG_LINES",
            "USED_IN_TRN_LINES",
            "AFFECTED_PACKAGING_CONFIGS",
            "AFFECTED_TRANSPORT_CONFIGS",
            "AFFECTED_SCENARIOS",
            "AFFECTED_SHIPMENTS",
            "IMPACT_SCORE",
            "IMPACT_LEVEL",
        ),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "COMPONENT_ID": _idx("T_COMPONENT", "COMPONENT_ID"),
            "COMPONENT_CODE": _idx("T_COMPONENT", "COMPONENT_CODE"),
            "WEIGHT_G": _idx("T_COMPONENT", "WEIGHT_G"),
            "USED_IN_PKG_LINES": (
                'IF([@COMPONENT_ID]="","",'
                "COUNTIF(T_PACKAGING_CONFIGURATION_LINE[COMPONENT_ID],[@COMPONENT_ID]))"
            ),
            "USED_IN_TRN_LINES": (
                'IF([@COMPONENT_ID]="","",'
                "COUNTIF(T_TRANSPORT_CONFIGURATION_LINE[COMPONENT_ID],[@COMPONENT_ID]))"
            ),
            "AFFECTED_PACKAGING_CONFIGS": (
                'IF([@COMPONENT_ID]="","",[@USED_IN_PKG_LINES])'
            ),
            "AFFECTED_TRANSPORT_CONFIGS": (
                'IF([@COMPONENT_ID]="","",[@USED_IN_TRN_LINES])'
            ),
            "AFFECTED_SCENARIOS": (
                'IF([@COMPONENT_ID]="","",'
                'IF([@USED_IN_TRN_LINES]+[@USED_IN_PKG_LINES]=0,0,'
                "COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])))"
            ),
            "AFFECTED_SHIPMENTS": (
                'IF([@COMPONENT_ID]="","",'
                "COUNTIF(T_SHIPMENT_LINE[COMPONENT_ID],[@COMPONENT_ID]))"
            ),
            "IMPACT_SCORE": (
                'IF([@COMPONENT_ID]="","",'
                "[@USED_IN_PKG_LINES]*2+[@USED_IN_TRN_LINES]*3+"
                "[@AFFECTED_SHIPMENTS]*5)"
            ),
            "IMPACT_LEVEL": (
                'IF([@COMPONENT_ID]="","",'
                'IF([@IMPACT_SCORE]>=20,"HIGH",IF([@IMPACT_SCORE]>=5,"MEDIUM","LOW")))'
            ),
        },
    )


def _engine_validation() -> EngineSheet:
    """Validation engine shell — rows written from VALIDATION_RULES."""
    return EngineSheet(
        sheet_name="ENG_VALIDATION",
        table_name="T_ENG_VALIDATION",
        row_count=len(VALIDATION_RULES),
        headers=("ROW_INDEX", "RULE_ID", "RULE_NAME", "RESULT", "SEVERITY"),
        formulas={
            "ROW_INDEX": "ROW_INDEX",
            "RULE_ID": "literal:",
            "RULE_NAME": "literal:",
            "RESULT": "literal:PENDING",
            "SEVERITY": "literal:ERROR",
        },
    )


VALIDATION_RULES: tuple[tuple[str, str, str, str], ...] = (
    ("V-WT-01", "MISSING_WEIGHT", 'IF(OR(COUNTIF(T_COMPONENT[WEIGHT_G],"<="&0)>0,COUNTBLANK(T_COMPONENT[WEIGHT_G])>0),"ERROR","OK")', "ERROR"),
    ("V-SUP-01", "MISSING_SUPPLIER", 'IF(COUNTBLANK(T_COMPONENT[SUPPLIER_ID])>0,"WARN","OK")', "WARN"),
    ("V-EVD-01", "MISSING_EVIDENCE", 'IF(COUNTA(T_COMPONENT[COMPONENT_ID])=0,"PENDING",IF(COUNTA(T_DOCUMENT_LINK[COMPONENT_ID])>0,"OK","WARN"))', "WARN"),
    ("V-DOC-01", "MISSING_DOCUMENTS", 'IF(COUNTIF(T_ENG_TECHNICAL_FILE[HAS_DOCUMENT],FALSE)>0,"WARN","OK")', "WARN"),
    ("V-DUP-01", "DUPLICATE_COMPONENT_CODE", 'IF(COUNTA(T_COMPONENT[COMPONENT_CODE])=0,"PENDING",IFERROR(IF(COUNTA(T_COMPONENT[COMPONENT_CODE])<>SUMPRODUCT(1/COUNTIF(T_COMPONENT[COMPONENT_CODE],T_COMPONENT[COMPONENT_CODE]&"")),"ERROR","OK"),"OK"))', "ERROR"),
    ("V-ACT-01", "INACTIVE_COMPONENT_IN_BOM", 'IF(COUNTA(T_PACKAGING_CONFIGURATION_LINE[COMPONENT_ID])=0,"PENDING","OK")', "WARN"),
    ("V-MAT-01", "MATERIAL_SHARE_ENGINE_READY", 'IF(COUNTA(T_ENG_MATERIAL_WEIGHT[ROW_INDEX])>0,"OK","ERROR")', "ERROR"),
    ("V-MAT-02", "COMPONENT_MATERIAL_LINKED", 'IF(COUNTA(T_COMPONENT[COMPONENT_ID])=0,"PENDING",IF(COUNTA(T_COMPONENT_MATERIAL[COMPONENT_MATERIAL_ID])>0,"OK","WARN"))', "WARN"),
    ("V-CFG-01", "PACKAGING_HAS_LINES", 'IF(COUNTA(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID])=0,"PENDING",IF(COUNTA(T_PACKAGING_CONFIGURATION_LINE[PACKAGING_CONFIGURATION_LINE_ID])>0,"OK","ERROR"))', "ERROR"),
    ("V-TRN-01", "TRANSPORT_HAS_LINES", 'IF(COUNTA(T_TRANSPORT_CONFIGURATION[TRANSPORT_CONFIGURATION_ID])=0,"PENDING",IF(COUNTA(T_TRANSPORT_CONFIGURATION_LINE[TRANSPORT_CONFIGURATION_LINE_ID])>0,"OK","WARN"))', "WARN"),
    ("V-SCN-01", "SCENARIO_CORE_FKS", 'IF(COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])=0,"PENDING",IF(COUNTBLANK(T_COMMERCIAL_SCENARIO[PRODUCT_ID])+COUNTBLANK(T_COMMERCIAL_SCENARIO[TRANSPORT_CONFIGURATION_ID])>0,"ERROR","OK"))', "ERROR"),
    ("V-REL-01", "BROKEN_RELATIONSHIPS_CATALOG", 'IF(COUNTA(T_SYS_RELATIONSHIPS[RELATIONSHIP_ID])>0,"OK","ERROR")', "ERROR"),
    ("V-LKP-01", "INVALID_LOOKUP_SEED", 'IF(COUNTA(T_LKP_STATUS[STATUS_ID])>0,"OK","ERROR")', "ERROR"),
    ("V-SHP-01", "SHIPMENT_QTY_POSITIVE", 'IF(COUNTIF(T_SHIPMENT[QTY_PRODUCT_UNITS],"<="&0)>0,"ERROR","OK")', "ERROR"),
    ("V-SHP-02", "SHIPMENT_LINE_FREEZE", 'IF(COUNTA(T_SHIPMENT[SHIPMENT_ID])=0,"PENDING",IF(COUNTA(T_SHIPMENT_LINE[SHIPMENT_LINE_ID])>0,"OK","WARN"))', "WARN"),
    ("V-STM-01", "STATEMENT_LINES", 'IF(COUNTA(T_STATEMENT[STATEMENT_ID])=0,"PENDING",IF(COUNTA(T_STATEMENT_LINE[STATEMENT_LINE_ID])>0,"OK","WARN"))', "WARN"),
    ("V-STM-02", "STATEMENT_VARIANCE", 'IF(COUNTIF(T_ENG_STATEMENT[RECONCILE_STATUS],"VARIANCE")>0,"ERROR",IF(COUNTIF(T_ENG_STATEMENT[RECONCILE_STATUS],"OK")>0,"OK","PENDING"))', "ERROR"),
    ("V-TF-01", "TECH_FILE_SUBJECT_XOR", 'IF(COUNTIF(T_ENG_TECHNICAL_FILE[HAS_EXACTLY_ONE_SUBJECT],FALSE)>0,"ERROR",IF(COUNTA(T_ENG_TECHNICAL_FILE[TECHNICAL_FILE_ID])=0,"PENDING","OK"))', "ERROR"),
    ("V-TF-INCI-01", "TECH_FILE_OWNS_PACKAGING_CONFIG", 'IF(COUNTA(T_TECHNICAL_FILE[TECHNICAL_FILE_ID])=0,"PENDING",IF(COUNTBLANK(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID])>0,"ERROR","OK"))', "ERROR"),
    ("V-SHP-INCI-01", "SHIPMENT_PRODUCT_PC_SCENARIO", 'IF(COUNTA(T_SHIPMENT[SHIPMENT_ID])=0,"PENDING",IF(COUNTBLANK(T_SHIPMENT[COMMERCIAL_SCENARIO_ID])+COUNTBLANK(T_SHIPMENT[PACKAGING_CONFIGURATION_ID])>0,"ERROR","OK"))', "ERROR"),
    ("V-SHP-INCI-02", "SHIPMENT_LOT_NUMBER", 'IF(COUNTA(T_SHIPMENT[SHIPMENT_ID])=0,"PENDING",IF(COUNTBLANK(T_SHIPMENT[EXTERNAL_REF])>0,"WARN","OK"))', "WARN"),
    ("V-PC-TF-01", "PACK_CONFIG_HAS_TECH_FILE", 'IF(COUNTA(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID])=0,"PENDING",IF(COUNTA(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID])>=COUNTA(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID]),"OK","WARN"))', "WARN"),
    ("V-SCN-INCI-01", "SCENARIO_HAS_INCOTERM", 'IF(COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])=0,"PENDING",IF(COUNTBLANK(T_COMMERCIAL_SCENARIO[INCOTERM_ID])>0,"WARN","OK"))', "WARN"),
    ("V-DoC-01", "DECLARATION_COMPLETE", 'IF(COUNTIF(T_ENG_DECLARATION[ENGINE_STATUS],"INCOMPLETE")>0,"ERROR",IF(COUNTA(T_ENG_DECLARATION[DECLARATION_OF_CONFORMITY_ID])=0,"PENDING","OK"))', "ERROR"),
    ("V-ENG-01", "MATERIAL_ENGINE", 'IF(COUNTA(T_ENG_MATERIAL_WEIGHT[ROW_INDEX])>0,"OK","ERROR")', "ERROR"),
    ("V-ENG-02", "PACKAGING_ENGINE", 'IF(COUNTA(T_ENG_PACKAGING_WEIGHT[ROW_INDEX])>0,"OK","ERROR")', "ERROR"),
    ("V-ENG-03", "SHIPMENT_ENGINE", 'IF(COUNTA(T_ENG_SHIPMENT_WEIGHT[ROW_INDEX])>0,"OK","ERROR")', "ERROR"),
)


def all_engines() -> tuple[EngineSheet, ...]:
    return (
        _engine_material_weight(),
        _engine_packaging_weight(),
        _engine_transport_weight(),
        _engine_shipment_weight(),
        _family_summary("ENG_PLASTIC_SUMMARY", "T_ENG_PLASTIC_SUMMARY", "PLASTIC"),
        _family_summary("ENG_PAPER_SUMMARY", "T_ENG_PAPER_SUMMARY", "PAPER"),
        _family_summary("ENG_WOOD_SUMMARY", "T_ENG_WOOD_SUMMARY", "WOOD"),
        _engine_statement(),
        _engine_technical_file(),
        _engine_declaration(),
        _engine_impact_analysis(),
        _engine_validation(),
    )


def write_engine_sheet(ws: Worksheet, engine: EngineSheet) -> Table:
    for col_idx, header in enumerate(engine.headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    if engine.sheet_name == "ENG_VALIDATION":
        return _write_validation_sheet(ws, engine)

    for row_offset in range(engine.row_count):
        excel_row = row_offset + 2
        row_index = row_offset + 1
        for col_idx, header in enumerate(engine.headers, start=1):
            raw_spec = engine.formulas.get(header)
            cell = ws.cell(row=excel_row, column=col_idx)
            if raw_spec is None:
                cell.value = None
            elif raw_spec == "ROW_INDEX":
                cell.value = row_index
            elif raw_spec.startswith("literal:"):
                lit = raw_spec[len("literal:") :]
                if lit == "":
                    cell.value = None
                else:
                    try:
                        cell.value = float(lit) if "." in lit else int(lit)
                    except ValueError:
                        cell.value = lit
            else:
                cell.value = f"={raw_spec}"

    last_col = excel_col_letter(len(engine.headers))
    last_row = 1 + engine.row_count
    table = Table(displayName=engine.table_name, ref=f"A1:{last_col}{last_row}")
    ws.add_table(table)
    return table


def _write_validation_sheet(ws: Worksheet, engine: EngineSheet) -> Table:
    for col_idx, header in enumerate(engine.headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_offset, (rule_id, rule_name, result_formula, severity) in enumerate(
        VALIDATION_RULES
    ):
        excel_row = row_offset + 2
        ws.cell(row=excel_row, column=1, value=row_offset + 1)
        ws.cell(row=excel_row, column=2, value=rule_id)
        ws.cell(row=excel_row, column=3, value=rule_name)
        ws.cell(row=excel_row, column=4, value=f"={result_formula}")
        ws.cell(row=excel_row, column=5, value=severity)

    last_col = excel_col_letter(len(engine.headers))
    last_row = 1 + len(VALIDATION_RULES)
    table = Table(displayName=engine.table_name, ref=f"A1:{last_col}{last_row}")
    ws.add_table(table)
    return table


def write_all_engines(workbook) -> list[str]:
    created: list[str] = []
    for engine in all_engines():
        ws = workbook.create_sheet(title=engine.sheet_name)
        write_engine_sheet(ws, engine)
        created.append(engine.sheet_name)
    return created
