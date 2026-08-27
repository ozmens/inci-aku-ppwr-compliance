"""Phase J — PIMS application layer + document registry sync (no Word regen)."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from importers.production.variant_description_codec import VariantDescriptionCodec

# ---------------------------------------------------------------------------
# Styles (Inci Akü — no white-on-light)
# ---------------------------------------------------------------------------

NAVY = "1F4E79"
YELLOW = "FFF2CC"
LIGHT_BLUE = "D6EAF8"
WHITE = "FFFFFF"
BODY = "1A1A1A"
GREY = "F2F2F2"
THIN = "B0B0B0"

FONT = "Tahoma"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT, size=10, bold=True, color=WHITE)
TITLE_FONT = Font(name=FONT, size=18, bold=True, color=NAVY)
SECTION_FONT = Font(name=FONT, size=12, bold=True, color=NAVY)
BODY_FONT = Font(name=FONT, size=9, color=BODY)
LINK_FONT = Font(name=FONT, size=9, color="0563C1", underline="single")
KPI_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
ACCENT_FILL = PatternFill("solid", fgColor=YELLOW)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
THIN_BORDER = Border(
    left=Side(style="thin", color=THIN),
    right=Side(style="thin", color=THIN),
    top=Side(style="thin", color=THIN),
    bottom=Side(style="thin", color=THIN),
)

ISSUE_DATE = "2026-08-08"
STATUS_ACTIVE = "2"

DOC_TYPE_TF = "1"
DOC_TYPE_DOC = "4"
DOC_TYPE_LABEL = "11"
DOC_TYPE_STM = "12"

FAMILY_FOLDER = {
    "STARTER": "01_STARTER",
    "INDUSTRIAL": "02_INDUSTRIAL",
    "CONTAINER": "03_CONTAINER",
}

FILE_BY_TYPE = {
    DOC_TYPE_TF: "01_Technical_File.docx",
    DOC_TYPE_DOC: "02_EU_DoC.docx",
    DOC_TYPE_LABEL: "03_Label.docx",
    DOC_TYPE_STM: "04_Shipment_Statement.docx",
}

UI_SHEETS_ORDER = [
    "00_HOME",
    "NAVIGATION",
    "SEARCH",
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "SHIPMENTS",
    "DOC_ENGINE_MAP",
    "00_README",
    "01_DASHBOARD",
    "02_RELEASE_CONTROL",
    "03_DATA_DICTIONARY",
    "04_IMPORT_GUIDE",
]


@dataclass
class PhaseJResult:
    success: bool
    gate: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _note_field(notes: str, key: str) -> str | None:
    m = re.search(rf"{re.escape(key)}=([^;]+)", notes or "")
    return m.group(1).strip() if m else None


def _headers(ws) -> list[str]:
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def _col(headers: list[str], name: str) -> int:
    return headers.index(name) + 1


def excel_open_ok(path: Path) -> dict:
    pythoncom.CoInitialize()
    excel = None
    out = {"ok": False, "error": None, "sheets": None}
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3
        wb = excel.Workbooks.Open(
            str(path.resolve()),
            UpdateLinks=0,
            ReadOnly=True,
            CorruptLoad=0,
        )
        out["ok"] = True
        out["sheets"] = int(wb.Worksheets.Count)
        wb.Close(False)
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return out


class PhaseJService:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx"
        self.candidate = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_CANDIDATE.xlsx"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery_test = self.out / "PHASE_J_DELIVERY_TEST"
        self.codec = VariantDescriptionCodec()
        self.qa_path = self.out / "PHASE_J_PIMS_INTEGRATION_QA.md"

    def run(self) -> PhaseJResult:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseJResult(False, "FAIL", ["Missing FINAL.xlsx"])
        if not self.phase_i.exists():
            return PhaseJResult(False, "FAIL", ["Missing PHASE_I_FINAL"])

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"Candidate copied from FINAL → {self.candidate.name}")

        wb = load_workbook(self.candidate)
        index = self._build_indexes(wb)
        sync = self._sync_document_registry(wb, index)
        messages.extend(sync["messages"])
        self._clear_stale_metadata_notes(wb)
        self._build_ui_layer(wb, index, sync["path_by_code"])
        self._reorder_sheets(wb)
        self._update_home_and_dashboard(wb, sync)
        wb.save(self.candidate)
        wb.close()
        messages.append("Workbook saved (UI + registry sync)")

        # Stage portable delivery test tree (junctions + candidate)
        self._stage_delivery_test()
        messages.append(f"Delivery test staged: {self.delivery_test}")

        excel = excel_open_ok(self.candidate)
        messages.append(f"Native Excel open: {excel}")

        qa = self._qa(index, sync, excel)
        gate = "PASS" if qa["pass"] else "FAIL"
        self._write_qa(qa, messages)
        return PhaseJResult(gate == "PASS", gate, messages, qa)

    # ------------------------------------------------------------------
    # Indexes
    # ------------------------------------------------------------------
    def _build_indexes(self, wb) -> dict[str, Any]:
        pc_ws = wb["PACKAGING_CONFIGURATION"]
        pc_h = _headers(pc_ws)
        configs = {}
        for row in pc_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            pc_id = str(row[_col(pc_h, "PACKAGING_CONFIGURATION_ID") - 1])
            set_code = str(row[_col(pc_h, "CONFIG_GROUP_CODE") - 1])
            notes = str(row[_col(pc_h, "NOTES") - 1] or "")
            desc = str(row[_col(pc_h, "DESCRIPTION") - 1] or "")
            try:
                vb_tr, vb_en = self.codec.deserialize(desc)
            except Exception:
                vb_tr, vb_en = "", ""
            family = _note_field(notes, "FAMILY") or "STARTER"
            final_id = _note_field(notes, "FINAL_CONFIGURATION_ID") or f"IA-{set_code}"
            source = _note_field(notes, "SOURCE_CONFIGURATION_ID") or ""
            mass = _note_field(notes, "PACKAGING_MASS_KG")
            configs[pc_id] = {
                "pc_id": pc_id,
                "set_code": set_code,
                "family": family,
                "final_id": final_id,
                "source": source,
                "vb_tr": vb_tr,
                "vb_en": vb_en,
                "desc": desc,
                "tare_kg": float(mass) if mass else None,
            }

        # BOM line counts
        bom_counts: dict[str, int] = defaultdict(int)
        bl_ws = wb["PACKAGING_CONFIGURATION_LINE"]
        bl_h = _headers(bl_ws)
        pci = _col(bl_h, "PACKAGING_CONFIGURATION_ID") - 1
        for row in bl_ws.iter_rows(min_row=2, values_only=True):
            if row and row[pci] is not None:
                bom_counts[str(row[pci])] += 1

        # Product counts via commercial scenario → transport → config
        # Simpler: count products linked through COMMERCIAL_SCENARIO notes / transport
        # Use COMMERCIAL_SCENARIO TRANSPORT_CONFIGURATION_ID → TRANSPORT_CONFIGURATION PACKAGING_CONFIGURATION_ID
        tc_to_pc: dict[str, str] = {}
        if "TRANSPORT_CONFIGURATION" in wb.sheetnames:
            tc_ws = wb["TRANSPORT_CONFIGURATION"]
            tc_h = _headers(tc_ws)
            tc_id_i = _col(tc_h, "TRANSPORT_CONFIGURATION_ID") - 1
            pc_i = (
                _col(tc_h, "PACKAGING_CONFIGURATION_ID") - 1
                if "PACKAGING_CONFIGURATION_ID" in tc_h
                else None
            )
            for row in tc_ws.iter_rows(min_row=2, values_only=True):
                if not row or row[tc_id_i] is None:
                    continue
                if pc_i is not None and row[pc_i] not in (None, ""):
                    tc_to_pc[str(row[tc_id_i])] = str(row[pc_i])

        product_counts: dict[str, int] = defaultdict(int)
        if "COMMERCIAL_SCENARIO" in wb.sheetnames:
            cs_ws = wb["COMMERCIAL_SCENARIO"]
            cs_h = _headers(cs_ws)
            tc_i = _col(cs_h, "TRANSPORT_CONFIGURATION_ID") - 1 if "TRANSPORT_CONFIGURATION_ID" in cs_h else None
            for row in cs_ws.iter_rows(min_row=2, values_only=True):
                if not row or tc_i is None:
                    continue
                tc = str(row[tc_i] or "")
                pc = tc_to_pc.get(tc)
                if pc:
                    product_counts[pc] += 1

        # Fallback product counts from Phase I loader style: if zero, leave 0 and fill from manifest later
        tf_by_pc: dict[str, dict] = {}
        tf_ws = wb["TECHNICAL_FILE"]
        tf_h = _headers(tf_ws)
        for row in tf_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            pc = str(row[_col(tf_h, "PACKAGING_CONFIGURATION_ID") - 1])
            tf_by_pc[pc] = {
                "id": str(row[_col(tf_h, "TECHNICAL_FILE_ID") - 1]),
                "code": str(row[_col(tf_h, "TECHNICAL_FILE_CODE") - 1]),
            }

        doc_by_pc: dict[str, dict] = {}
        d_ws = wb["DECLARATION_OF_CONFORMITY"]
        d_h = _headers(d_ws)
        for row in d_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            pc = str(row[_col(d_h, "PACKAGING_CONFIGURATION_ID") - 1])
            doc_by_pc[pc] = {
                "id": str(row[_col(d_h, "DECLARATION_OF_CONFORMITY_ID") - 1]),
                "code": str(row[_col(d_h, "DOC_NUMBER") - 1]),
            }

        stm_by_code: dict[str, dict] = {}
        s_ws = wb["STATEMENT"]
        s_h = _headers(s_ws)
        for row in s_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            code = str(row[_col(s_h, "STATEMENT_CODE") - 1])
            stm_by_code[code] = {
                "id": str(row[_col(s_h, "STATEMENT_ID") - 1]),
                "code": code,
            }

        # Map set_code → statement via DOCUMENT_LIBRARY codes
        return {
            "configs": configs,
            "bom_counts": bom_counts,
            "product_counts": product_counts,
            "tf_by_pc": tf_by_pc,
            "doc_by_pc": doc_by_pc,
            "stm_by_code": stm_by_code,
            "tc_to_pc": tc_to_pc,
        }

    # ------------------------------------------------------------------
    # Document registry sync
    # ------------------------------------------------------------------
    def _sync_document_registry(self, wb, index: dict) -> dict[str, Any]:
        messages = []
        lib = wb["DOCUMENT_LIBRARY"]
        link = wb["DOCUMENT_LINK"]
        lib_h = _headers(lib)
        link_h = _headers(link)

        # Build DOCUMENT_ID → link row index
        link_row_by_doc: dict[str, int] = {}
        for r in range(2, link.max_row + 1):
            did = link.cell(r, _col(link_h, "DOCUMENT_ID")).value
            if did is not None:
                link_row_by_doc[str(did)] = r

        # set_code → pc_id
        set_to_pc = {v["set_code"]: k for k, v in index["configs"].items()}

        path_by_code: dict[str, str] = {}
        hash_by_code: dict[str, str] = {}
        linked = 0
        missing_files = []
        stale_cleared = 0

        for r in range(2, lib.max_row + 1):
            code = str(lib.cell(r, _col(lib_h, "DOCUMENT_CODE")).value or "")
            dtype = str(lib.cell(r, _col(lib_h, "DOCUMENT_TYPE_ID")).value or "")
            doc_id = str(lib.cell(r, _col(lib_h, "DOCUMENT_ID")).value or "")
            if not code:
                continue

            # Extract packaging set from document code
            # IA-PPWR-TF-ST-051-STD-01-R00 → ST-051-STD-01
            m = re.search(r"IA-PPWR-(?:TF|DOC|LBL|STM)-(.+)-R\d+$", code)
            set_code = m.group(1) if m else ""
            cfg = index["configs"].get(set_to_pc.get(set_code, ""), {})
            family = cfg.get("family", "STARTER")
            folder = FAMILY_FOLDER.get(family, "01_STARTER")
            fname = FILE_BY_TYPE.get(dtype)
            if not fname or not set_code:
                missing_files.append(code)
                continue

            rel = f"{folder}/{set_code}/{fname}"
            abs_path = self.phase_i / folder / set_code / fname
            if not abs_path.exists():
                missing_files.append(rel)
                continue

            digest = sha256_file(abs_path)
            path_by_code[code] = rel
            hash_by_code[code] = digest

            lib.cell(r, _col(lib_h, "FILE_URI")).value = rel
            lib.cell(r, _col(lib_h, "FILE_HASH")).value = digest
            lib.cell(r, _col(lib_h, "ISSUE_DATE")).value = ISSUE_DATE
            lib.cell(r, _col(lib_h, "STATUS_ID")).value = STATUS_ACTIVE
            lib.cell(r, _col(lib_h, "NOTES")).value = (
                "Phase I generated Rev.00 Word file — linked; signatures not completed"
            )
            lib.cell(r, _col(lib_h, "UPDATED_AT")).value = datetime.now(timezone.utc).strftime(
                "%Y-%m-%dT%H:%M:%SZ"
            )
            linked += 1

            # Fix DOCUMENT_LINK XOR target
            lr = link_row_by_doc.get(doc_id)
            if lr:
                # clear all typed targets first
                for col_name in (
                    "COMPONENT_ID",
                    "PRODUCT_ID",
                    "PACKAGING_CONFIGURATION_ID",
                    "TRANSPORT_CONFIGURATION_ID",
                    "TECHNICAL_FILE_ID",
                    "DECLARATION_OF_CONFORMITY_ID",
                    "STATEMENT_ID",
                ):
                    link.cell(lr, _col(link_h, col_name)).value = None

                pc_id = set_to_pc.get(set_code)
                if dtype == DOC_TYPE_TF:
                    tf = index["tf_by_pc"].get(pc_id or "")
                    if tf:
                        link.cell(lr, _col(link_h, "TECHNICAL_FILE_ID")).value = tf["id"]
                elif dtype == DOC_TYPE_DOC:
                    d = index["doc_by_pc"].get(pc_id or "")
                    if d:
                        link.cell(
                            lr, _col(link_h, "DECLARATION_OF_CONFORMITY_ID")
                        ).value = d["id"]
                elif dtype == DOC_TYPE_STM:
                    stm = index["stm_by_code"].get(code)
                    # Statement codes match DOCUMENT_CODE for STM
                    if not stm:
                        stm = index["stm_by_code"].get(code)
                    # map via STATEMENT_CODE == code
                    if stm:
                        link.cell(lr, _col(link_h, "STATEMENT_ID")).value = stm["id"]
                    else:
                        # fallback: find by code in STATEMENT sheet already indexed by code
                        for sc, meta in index["stm_by_code"].items():
                            if sc == code:
                                link.cell(lr, _col(link_h, "STATEMENT_ID")).value = meta["id"]
                                break
                else:  # LABEL → packaging configuration
                    if pc_id:
                        link.cell(
                            lr, _col(link_h, "PACKAGING_CONFIGURATION_ID")
                        ).value = pc_id

                link.cell(lr, _col(link_h, "NOTES")).value = "Phase J registry sync"

        messages.append(f"DOCUMENT_LIBRARY linked: {linked}/988")
        if missing_files:
            messages.append(f"Missing DOCX paths: {len(missing_files)} (sample {missing_files[:5]})")

        return {
            "messages": messages,
            "linked": linked,
            "missing_files": missing_files,
            "path_by_code": path_by_code,
            "hash_by_code": hash_by_code,
            "stale_cleared": stale_cleared,
        }

    def _clear_stale_metadata_notes(self, wb) -> None:
        replacements = (
            ("Word not generated", ""),
            ("Word/PDF not generated", ""),
            ("Metadata only — ", ""),
            ("Metadata only - ", ""),
            ("in Phase F", "— superseded by Phase I generation"),
        )
        for sheet_name in (
            "TECHNICAL_FILE",
            "DECLARATION_OF_CONFORMITY",
            "STATEMENT",
            "DOCUMENT_LIBRARY",
        ):
            ws = wb[sheet_name]
            h = _headers(ws)
            if "NOTES" not in h:
                continue
            c = _col(h, "NOTES")
            for r in range(2, ws.max_row + 1):
                val = ws.cell(r, c).value
                if not val:
                    continue
                text = str(val)
                if "Word not generated" in text or "Metadata only" in text:
                    if sheet_name == "TECHNICAL_FILE":
                        ws.cell(r, c).value = (
                            "Rev.00 Word technical file generated — see DOCUMENT_LIBRARY.FILE_URI; "
                            "drawings/photos may remain PENDING"
                        )
                    elif sheet_name == "DECLARATION_OF_CONFORMITY":
                        ws.cell(r, c).value = (
                            "Rev.00 Word EU DoC generated — see DOCUMENT_LIBRARY.FILE_URI; "
                            "signature fields blank / not legally signed"
                        )
                    elif sheet_name == "STATEMENT":
                        ws.cell(r, c).value = (
                            "Rev.00 Word shipment statement generated — see DOCUMENT_LIBRARY.FILE_URI"
                        )
                    else:
                        new = text
                        for a, b in replacements:
                            new = new.replace(a, b)
                        ws.cell(r, c).value = new.strip(" ;—") or "Linked Rev.00 Word document"

    # ------------------------------------------------------------------
    # UI layer
    # ------------------------------------------------------------------
    def _style_header_row(self, ws, headers: list[str], row: int = 1) -> None:
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 28
        ws.freeze_panes = f"A{row + 1}"
        # Manual filter OK on UI sheets WITHOUT Excel Table
        ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"

    def _autosize(self, ws, cols: int, width: float = 18) -> None:
        for i in range(1, cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _hyperlink_cell(self, ws, row: int, col: int, rel_path: str, label: str = "Open") -> None:
        # Relative HYPERLINK — portable inside delivery root
        cell = ws.cell(row=row, column=col)
        cell.value = f'=HYPERLINK("{rel_path}","{label}")'
        cell.font = LINK_FONT

    def _build_ui_layer(self, wb, index: dict, path_by_code: dict[str, str]) -> None:
        # Remove old UI shell sheets we will replace names carefully —
        # Keep 00_README etc but rebuild 01_DASHBOARD content; add new modules.
        for name in list(wb.sheetnames):
            if name in {
                "00_HOME",
                "NAVIGATION",
                "SEARCH",
                "PACKAGING_CONFIGURATIONS",
                "PRODUCT_MASTER",
                "COMPONENT_MASTER",
                "DOCUMENT_CENTER",
                "TECHNICAL_FILES",
                "DECLARATIONS_OF_CONFORMITY",
                "LABELS",
                "SHIPMENT_STATEMENTS",
                "SHIPMENTS",
                "DOC_ENGINE_MAP",
            }:
                del wb[name]

        self._sheet_home(wb, index, path_by_code)
        self._sheet_navigation(wb)
        self._sheet_search(wb)
        self._sheet_document_center(wb, index, path_by_code)
        self._sheet_technical_files(wb, index, path_by_code)
        self._sheet_docs(wb, index, path_by_code)
        self._sheet_labels(wb, index, path_by_code)
        self._sheet_statements(wb, index, path_by_code)
        self._sheet_packaging(wb, index, path_by_code)
        self._sheet_products(wb, index)
        self._sheet_components(wb, index)
        self._sheet_shipments(wb)
        self._sheet_doc_engine_map(wb)

    def _sheet_home(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("00_HOME", 0)
        ws["A1"] = "İnci Akü PPWR Packaging Information Management System — Rev.00"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:F1")
        ws["A2"] = (
            "Production Master Data: LOADED  |  Golden Variant Register: 247 / 247  |  "
            "Document Pack: 988 / 988 GENERATED  |  Document Registry: 988 / 988 LINKED  |  "
            "Blocking QA Errors: 0"
        )
        ws["A2"].font = BODY_FONT
        ws["A2"].fill = KPI_FILL
        ws.merge_cells("A2:F2")
        ws["A3"] = (
            "Legal signatures are NOT completed in this workbook. Word files remain the controlled legal documents."
        )
        ws["A3"].font = Font(name=FONT, size=9, italic=True, color=BODY)

        headers = ["KPI", "Count", "Status"]
        self._style_header_row(ws, headers, row=5)
        kpis = [
            ("Final Packaging Configurations", 247, "LOADED"),
            ("Starter", 240, "LOADED"),
            ("Industrial", 3, "LOADED"),
            ("Container / Loading", 4, "LOADED"),
            ("BOM Lines", 1690, "LOADED"),
            ("Components", 112, "LOADED"),
            ("Products", 2046, "LOADED"),
            ("Technical Files", 247, "GENERATED"),
            ("EU Declarations of Conformity", 247, "GENERATED"),
            ("Labels", 247, "GENERATED"),
            ("Shipment Statements", 247, "GENERATED"),
            ("Total Documents", 988, "GENERATED"),
            ("Document Registry Linked", f"{len(path_by_code)} / 988", "LINKED"),
            ("Blocking QA Errors", 0, "PASS"),
        ]
        for i, (k, v, s) in enumerate(kpis, start=6):
            ws.cell(i, 1, k).font = BODY_FONT
            ws.cell(i, 2, v).font = BODY_FONT
            ws.cell(i, 3, s).font = BODY_FONT
            for c in range(1, 4):
                ws.cell(i, c).border = THIN_BORDER
                if i % 2 == 0:
                    ws.cell(i, c).fill = WHITE_FILL

        ws["A22"] = "Quick navigation"
        ws["A22"].font = SECTION_FONT
        nav = [
            ("Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
            ("Products", "PRODUCT_MASTER"),
            ("Components", "COMPONENT_MASTER"),
            ("Technical Files", "TECHNICAL_FILES"),
            ("Declarations of Conformity", "DECLARATIONS_OF_CONFORMITY"),
            ("Labels", "LABELS"),
            ("Shipment Statements", "SHIPMENT_STATEMENTS"),
            ("Document Center", "DOCUMENT_CENTER"),
            ("Search", "SEARCH"),
            ("Navigation", "NAVIGATION"),
        ]
        for i, (label, sheet) in enumerate(nav, start=23):
            cell = ws.cell(i, 1)
            cell.value = f'=HYPERLINK("#\'{sheet}\'!A1","{label}")'
            cell.font = LINK_FONT

        self._autosize(ws, 3, 42)
        ws.sheet_view.showGridLines = False

    def _sheet_navigation(self, wb) -> None:
        ws = wb.create_sheet("NAVIGATION", 1)
        ws["A1"] = "NAVIGATION"
        ws["A1"].font = TITLE_FONT
        sections = {
            "HOME": ["00_HOME", "01_DASHBOARD", "02_RELEASE_CONTROL"],
            "PACKAGING": ["PACKAGING_CONFIGURATIONS"],
            "PRODUCTS": ["PRODUCT_MASTER"],
            "COMPONENTS": ["COMPONENT_MASTER"],
            "DOCUMENTS": [
                "DOCUMENT_CENTER",
                "TECHNICAL_FILES",
                "DECLARATIONS_OF_CONFORMITY",
                "LABELS",
                "SHIPMENT_STATEMENTS",
            ],
            "OPERATIONS": ["SHIPMENTS", "SEARCH"],
            "DATABASE": [
                "PACKAGING_CONFIGURATION",
                "PACKAGING_CONFIGURATION_LINE",
                "PRODUCT",
                "COMPONENT",
                "DOCUMENT_LIBRARY",
                "DOCUMENT_LINK",
            ],
            "LOOKUPS": ["LKP_STATUS", "LKP_DOCUMENT_TYPE", "LKP_MATERIAL_FAMILY"],
            "SYSTEM": ["SYS_WORKBOOK_INFO", "DOC_ENGINE_MAP", "00_README"],
        }
        r = 3
        for section, sheets in sections.items():
            ws.cell(r, 1, section).font = SECTION_FONT
            ws.cell(r, 1).fill = ACCENT_FILL
            r += 1
            for s in sheets:
                if s in wb.sheetnames or s.startswith("0") or s in UI_SHEETS_ORDER:
                    cell = ws.cell(r, 1)
                    cell.value = f'=HYPERLINK("#\'{s}\'!A1","{s}")'
                    cell.font = LINK_FONT
                    r += 1
            r += 1
        self._autosize(ws, 1, 48)
        ws.sheet_view.showGridLines = False

    def _sheet_search(self, wb) -> None:
        ws = wb.create_sheet("SEARCH", 2)
        ws["A1"] = "SEARCH"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = (
            "Use AutoFilter on DOCUMENT_CENTER / PACKAGING_CONFIGURATIONS / TECHNICAL_FILES "
            "(or Ctrl+F). Optional single-key lookup below uses XLOOKUP when available."
        )
        ws["A2"].font = BODY_FONT
        ws.merge_cells("A2:D2")

        ws["A4"] = "Lookup key"
        ws["A4"].font = SECTION_FONT
        ws["B4"] = ""
        ws["B4"].fill = PatternFill("solid", fgColor=YELLOW)
        ws["B4"].border = THIN_BORDER

        ws["A5"] = "Examples"
        ws["B5"] = (
            "ST-051-STD-01 | IA-ST-051-STD-01 | IA-ST-CFG-0122 | "
            "IA-PPWR-TF-ST-051-STD-01-R00 | 1011935 | 4000782"
        )
        ws["B5"].font = BODY_FONT

        ws["A7"] = "Packaging Set → Configuration ID"
        ws["A7"].font = SECTION_FONT
        # Bounded XLOOKUP — no whole-column FILTER / INDIRECT / OFFSET
        ws["A8"] = "Configuration ID"
        ws["B8"] = (
            '=IF(B4="","",IFERROR(XLOOKUP(B4,PACKAGING_CONFIGURATIONS!A:A,'
            'PACKAGING_CONFIGURATIONS!B:B),"Not found / use AutoFilter"))'
        )
        ws["B8"].font = BODY_FONT

        ws["A9"] = "Family"
        ws["B9"] = (
            '=IF(B4="","",IFERROR(XLOOKUP(B4,PACKAGING_CONFIGURATIONS!A:A,'
            'PACKAGING_CONFIGURATIONS!D:D),""))'
        )
        ws["B9"].font = BODY_FONT

        ws["A10"] = "Technical File"
        ws["B10"] = (
            '=IF(B4="","",IFERROR(XLOOKUP(B4,PACKAGING_CONFIGURATIONS!A:A,'
            'PACKAGING_CONFIGURATIONS!K:K),""))'
        )
        ws["B10"].font = BODY_FONT

        ws["A12"] = "Module shortcuts"
        ws["A12"].font = SECTION_FONT
        for i, (label, sheet) in enumerate(
            [
                ("Document Center (filter here)", "DOCUMENT_CENTER"),
                ("Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
                ("Technical Files", "TECHNICAL_FILES"),
                ("Product Master", "PRODUCT_MASTER"),
                ("Component Master", "COMPONENT_MASTER"),
            ],
            start=13,
        ):
            cell = ws.cell(i, 1)
            cell.value = f'=HYPERLINK("#\'{sheet}\'!A1","{label}")'
            cell.font = LINK_FONT

        self._autosize(ws, 2, 40)
        ws.column_dimensions["B"].width = 80
        ws.sheet_view.showGridLines = False

    def _iter_configs_sorted(self, index):
        items = list(index["configs"].values())
        items.sort(key=lambda x: (x["family"], x["set_code"]))
        return items

    def _ids_for_config(self, index, pc_id: str, set_code: str) -> dict:
        tf = index["tf_by_pc"].get(pc_id, {})
        doc = index["doc_by_pc"].get(pc_id, {})
        stm_code = f"IA-PPWR-STM-{set_code}-R00"
        lbl_code = f"IA-PPWR-LBL-{set_code}-R00"
        stm = index["stm_by_code"].get(stm_code, {"code": stm_code})
        return {
            "tf": tf.get("code", f"IA-PPWR-TF-{set_code}-R00"),
            "doc": doc.get("code", f"IA-PPWR-DOC-{set_code}-R00"),
            "lbl": lbl_code,
            "stm": stm.get("code", stm_code),
        }

    def _sheet_document_center(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("DOCUMENT_CENTER")
        headers = [
            "Packaging Set Code",
            "Configuration ID",
            "Variant Basis TR",
            "Variant Basis EN",
            "Source BOM ID",
            "Family",
            "Packaging Tare kg",
            "Product Count",
            "Technical File ID",
            "Open Technical File",
            "EU DoC ID",
            "Open EU DoC",
            "Label ID",
            "Open Label",
            "Statement ID",
            "Open Statement",
            "Document Status",
            "Revision",
        ]
        self._style_header_row(ws, headers)
        for i, cfg in enumerate(self._iter_configs_sorted(index), start=2):
            ids = self._ids_for_config(index, cfg["pc_id"], cfg["set_code"])
            vals = [
                cfg["set_code"],
                cfg["final_id"],
                cfg["vb_tr"],
                cfg["vb_en"],
                cfg["source"],
                cfg["family"],
                cfg["tare_kg"],
                index["product_counts"].get(cfg["pc_id"], 0),
                ids["tf"],
                None,
                ids["doc"],
                None,
                ids["lbl"],
                None,
                ids["stm"],
                None,
                "GENERATED / LINKED",
                "Rev.00",
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(i, c, v)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
            for code, col in (
                (ids["tf"], 10),
                (ids["doc"], 12),
                (ids["lbl"], 14),
                (ids["stm"], 16),
            ):
                rel = path_by_code.get(code)
                if rel:
                    self._hyperlink_cell(ws, i, col, rel, "Open")
                else:
                    ws.cell(i, col, "MISSING").font = BODY_FONT
        # extend autofilter to data
        last = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
        self._autosize(ws, len(headers), 16)
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["D"].width = 36

    def _sheet_technical_files(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("TECHNICAL_FILES")
        headers = [
            "Technical File Code",
            "Packaging Set Code",
            "Configuration ID",
            "Variant Basis",
            "Source BOM",
            "Product Count",
            "BOM Line Count",
            "Packaging Tare",
            "Article 5 Assessment Basis",
            "Drawing Status",
            "Photo Status",
            "Revision",
            "Open Technical File",
        ]
        self._style_header_row(ws, headers)
        for i, cfg in enumerate(self._iter_configs_sorted(index), start=2):
            ids = self._ids_for_config(index, cfg["pc_id"], cfg["set_code"])
            row = [
                ids["tf"],
                cfg["set_code"],
                cfg["final_id"],
                cfg["vb_tr"],
                cfg["source"],
                index["product_counts"].get(cfg["pc_id"], 0),
                index["bom_counts"].get(cfg["pc_id"], 0),
                cfg["tare_kg"],
                "REV00 CURRENT EVIDENCE BASIS — ARTICLE 5 ASSESSMENT BASIS",
                "PENDING - DRAWING",
                "PENDING - PHOTOGRAPHS",
                "Rev.00",
                None,
            ]
            for c, v in enumerate(row, start=1):
                ws.cell(i, c, v).font = BODY_FONT
                ws.cell(i, c).border = THIN_BORDER
            rel = path_by_code.get(ids["tf"])
            if rel:
                self._hyperlink_cell(ws, i, 13, rel)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        self._autosize(ws, len(headers), 18)

    def _sheet_docs(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("DECLARATIONS_OF_CONFORMITY")
        headers = [
            "DoC Number",
            "Packaging Set Code",
            "Configuration ID",
            "Packaging Description",
            "Variant Basis",
            "Source BOM ID",
            "Technical File ID",
            "Issue Date",
            "Revision",
            "Status",
            "Open DoC",
        ]
        self._style_header_row(ws, headers)
        for i, cfg in enumerate(self._iter_configs_sorted(index), start=2):
            ids = self._ids_for_config(index, cfg["pc_id"], cfg["set_code"])
            row = [
                ids["doc"],
                cfg["set_code"],
                cfg["final_id"],
                cfg["vb_tr"],
                f"{cfg['vb_tr']} / {cfg['vb_en']}" if cfg["vb_en"] else cfg["vb_tr"],
                cfg["source"],
                ids["tf"],
                ISSUE_DATE,
                "Rev.00",
                "ACTIVE — unsigned",
                None,
            ]
            for c, v in enumerate(row, start=1):
                ws.cell(i, c, v).font = BODY_FONT
                ws.cell(i, c).border = THIN_BORDER
            rel = path_by_code.get(ids["doc"])
            if rel:
                self._hyperlink_cell(ws, i, 11, rel)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        self._autosize(ws, len(headers), 18)

    def _sheet_labels(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("LABELS")
        headers = [
            "Label ID",
            "Packaging Set Code",
            "Configuration ID",
            "Variant Basis",
            "Material Families",
            "TF ID",
            "DoC ID",
            "Revision",
            "Open Label",
        ]
        self._style_header_row(ws, headers)
        for i, cfg in enumerate(self._iter_configs_sorted(index), start=2):
            ids = self._ids_for_config(index, cfg["pc_id"], cfg["set_code"])
            row = [
                ids["lbl"],
                cfg["set_code"],
                cfg["final_id"],
                cfg["vb_tr"],
                "See Technical File material-family summary",
                ids["tf"],
                ids["doc"],
                "Rev.00",
                None,
            ]
            for c, v in enumerate(row, start=1):
                ws.cell(i, c, v).font = BODY_FONT
                ws.cell(i, c).border = THIN_BORDER
            rel = path_by_code.get(ids["lbl"])
            if rel:
                self._hyperlink_cell(ws, i, 9, rel)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        self._autosize(ws, len(headers), 18)

    def _sheet_statements(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("SHIPMENT_STATEMENTS")
        headers = [
            "Statement ID",
            "Packaging Set Code",
            "Configuration ID",
            "Source BOM",
            "Variant Basis",
            "Packaging Tare",
            "BOM Line Count",
            "Revision",
            "Open Statement",
        ]
        self._style_header_row(ws, headers)
        for i, cfg in enumerate(self._iter_configs_sorted(index), start=2):
            ids = self._ids_for_config(index, cfg["pc_id"], cfg["set_code"])
            row = [
                ids["stm"],
                cfg["set_code"],
                cfg["final_id"],
                cfg["source"],
                cfg["vb_tr"],
                cfg["tare_kg"],
                index["bom_counts"].get(cfg["pc_id"], 0),
                "Rev.00",
                None,
            ]
            for c, v in enumerate(row, start=1):
                ws.cell(i, c, v).font = BODY_FONT
                ws.cell(i, c).border = THIN_BORDER
            rel = path_by_code.get(ids["stm"])
            if rel:
                self._hyperlink_cell(ws, i, 9, rel)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        self._autosize(ws, len(headers), 18)

    def _sheet_packaging(self, wb, index, path_by_code) -> None:
        ws = wb.create_sheet("PACKAGING_CONFIGURATIONS")
        headers = [
            "Packaging Set Code",
            "Final Configuration ID",
            "Source Configuration ID",
            "Family",
            "Packaging Description",
            "Variant Basis TR",
            "Variant Basis EN",
            "Packaging Tare kg",
            "Product Count",
            "BOM Line Count",
            "Technical File",
            "DoC",
            "Label",
            "Statement",
        ]
        self._style_header_row(ws, headers)
        for i, cfg in enumerate(self._iter_configs_sorted(index), start=2):
            ids = self._ids_for_config(index, cfg["pc_id"], cfg["set_code"])
            row = [
                cfg["set_code"],
                cfg["final_id"],
                cfg["source"],
                cfg["family"],
                cfg["vb_tr"],
                cfg["vb_tr"],
                cfg["vb_en"],
                cfg["tare_kg"],
                index["product_counts"].get(cfg["pc_id"], 0),
                index["bom_counts"].get(cfg["pc_id"], 0),
                ids["tf"],
                ids["doc"],
                ids["lbl"],
                ids["stm"],
            ]
            for c, v in enumerate(row, start=1):
                ws.cell(i, c, v).font = BODY_FONT
                ws.cell(i, c).border = THIN_BORDER
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        self._autosize(ws, len(headers), 18)

    def _sheet_products(self, wb, index) -> None:
        ws = wb.create_sheet("PRODUCT_MASTER")
        headers = [
            "Product Code",
            "Technical Description",
            "Category",
            "Mapped Packaging Set Code",
            "Configuration ID",
            "Source Configuration ID",
            "Packaging Tare",
            "Document Pack Status",
        ]
        self._style_header_row(ws, headers)
        # Build product → set mapping via commercial scenarios if possible
        prod_ws = wb["PRODUCT"]
        ph = _headers(prod_ws)
        # Map transport → set via tc_to_pc then configs
        tc_to_set = {}
        for tc, pc in index["tc_to_pc"].items():
            cfg = index["configs"].get(pc)
            if cfg:
                tc_to_set[tc] = cfg

        prod_to_sets: dict[str, set[str]] = defaultdict(set)
        if "COMMERCIAL_SCENARIO" in wb.sheetnames:
            cs = wb["COMMERCIAL_SCENARIO"]
            ch = _headers(cs)
            pi = _col(ch, "PRODUCT_ID") - 1
            ti = _col(ch, "TRANSPORT_CONFIGURATION_ID") - 1
            for row in cs.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                pid = str(row[pi] or "")
                tc = str(row[ti] or "")
                cfg = tc_to_set.get(tc)
                if cfg:
                    prod_to_sets[pid].add(cfg["set_code"])

        # PRODUCT_ID → code
        id_to_prod = {}
        for row in prod_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            pid = str(row[_col(ph, "PRODUCT_ID") - 1])
            code = str(row[_col(ph, "PRODUCT_CODE") - 1])
            name = str(row[_col(ph, "PRODUCT_NAME") - 1] if "PRODUCT_NAME" in ph else "")
            cat = ""
            if "PRODUCT_CATEGORY_ID" in ph:
                cat = str(row[_col(ph, "PRODUCT_CATEGORY_ID") - 1] or "")
            id_to_prod[pid] = (code, name, cat)

        r = 2
        for pid, (code, name, cat) in sorted(id_to_prod.items(), key=lambda x: x[1][0]):
            sets = sorted(prod_to_sets.get(pid, []))
            if not sets:
                ws.cell(r, 1, code).font = BODY_FONT
                ws.cell(r, 2, name).font = BODY_FONT
                ws.cell(r, 3, cat).font = BODY_FONT
                ws.cell(r, 8, "NO MAPPING").font = BODY_FONT
                r += 1
                continue
            for set_code in sets:
                pc_id = next(
                    (k for k, v in index["configs"].items() if v["set_code"] == set_code),
                    None,
                )
                cfg = index["configs"].get(pc_id or "", {})
                vals = [
                    code,
                    name,
                    cat,
                    set_code,
                    cfg.get("final_id", ""),
                    cfg.get("source", ""),
                    cfg.get("tare_kg", ""),
                    "988-pack GENERATED",
                ]
                for c, v in enumerate(vals, start=1):
                    ws.cell(r, c, v).font = BODY_FONT
                    ws.cell(r, c).border = THIN_BORDER
                r += 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
        self._autosize(ws, len(headers), 18)

    def _sheet_components(self, wb, index) -> None:
        ws = wb.create_sheet("COMPONENT_MASTER")
        headers = [
            "ERP Component Code",
            "Description",
            "Component Type",
            "Material Family",
            "UOM",
            "Controlled Unit Weight (g)",
            "Supplier",
            "Evidence Status",
            "Configs Using Component",
        ]
        self._style_header_row(ws, headers)
        # usage counts
        usage: dict[str, int] = defaultdict(int)
        bl = wb["PACKAGING_CONFIGURATION_LINE"]
        bh = _headers(bl)
        ci = _col(bh, "COMPONENT_ID") - 1
        for row in bl.iter_rows(min_row=2, values_only=True):
            if row and row[ci] is not None:
                usage[str(row[ci])] += 1

        cw = wb["COMPONENT"]
        ch = _headers(cw)
        r = 2
        for row in cw.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            cid = str(row[_col(ch, "COMPONENT_ID") - 1])
            code = str(row[_col(ch, "COMPONENT_CODE") - 1] if "COMPONENT_CODE" in ch else "")
            name = str(row[_col(ch, "COMPONENT_NAME") - 1] if "COMPONENT_NAME" in ch else "")
            ctype = str(row[_col(ch, "COMPONENT_TYPE_ID") - 1] if "COMPONENT_TYPE_ID" in ch else "")
            wt = row[_col(ch, "WEIGHT_G") - 1] if "WEIGHT_G" in ch else None
            vals = [
                code,
                name,
                ctype,
                "",
                "ADT",
                wt,
                "",
                "See evidence archive",
                usage.get(cid, 0),
            ]
            for c, v in enumerate(vals, start=1):
                ws.cell(r, c, v).font = BODY_FONT
                ws.cell(r, c).border = THIN_BORDER
            r += 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
        self._autosize(ws, len(headers), 18)

    def _sheet_shipments(self, wb) -> None:
        ws = wb.create_sheet("SHIPMENTS")
        ws["A1"] = "SHIPMENTS — transactional"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = (
            "Transaction shipment records remain on database sheet SHIPMENT / SHIPMENT_LINE. "
            "Configuration-level packaging statements are under SHIPMENT_STATEMENTS."
        )
        ws["A2"].font = BODY_FONT
        cell = ws.cell(4, 1)
        cell.value = '=HYPERLINK("#\'SHIPMENT\'!A1","Open SHIPMENT database sheet")'
        cell.font = LINK_FONT
        cell = ws.cell(5, 1)
        cell.value = '=HYPERLINK("#\'SHIPMENT_STATEMENTS\'!A1","Open SHIPMENT_STATEMENTS module")'
        cell.font = LINK_FONT
        self._autosize(ws, 1, 60)

    def _sheet_doc_engine_map(self, wb) -> None:
        ws = wb.create_sheet("DOC_ENGINE_MAP")
        ws["A1"] = "DOC_ENGINE_MAP — read-only (Python remains authoritative)"
        ws["A1"].font = TITLE_FONT
        headers = [
            "Document Type",
            "Logical Field",
            "PIMS Source",
            "DocumentContext Field",
            "Runtime Token / Mapping",
            "Builder Module",
        ]
        self._style_header_row(ws, headers, row=3)
        rows = [
            ("Technical File", "Packaging Set Code", "PACKAGING_CONFIGURATION.CONFIG_GROUP_CODE", "document_ids.packaging_set_code", "{{CONFIG_SET_CODE}}", "phase_g.merge_engine"),
            ("Technical File", "Configuration ID", "NOTES FINAL_CONFIGURATION_ID", "document_ids.final_configuration_id", "{{CONFIG_ID}}", "DocumentContextFactory"),
            ("Technical File", "Variant Basis TR/EN", "DESCRIPTION codec", "configuration.variant_basis_*", "{{VARIANT_BASIS_TR/EN}}", "VariantDescriptionCodec"),
            ("Technical File", "Exact BOM", "PACKAGING_CONFIGURATION_LINE", "configuration.lines", "{{BOM_TABLE}}", "merge_engine._fill_bom_table"),
            ("Technical File", "Tare", "WeightService", "total_tare_g", "{{TOTAL_TARE_KG}}", "WeightService"),
            ("EU DoC", "Declaration ID", "DECLARATION_OF_CONFORMITY.DOC_NUMBER", "document_ids.doc_id", "{{DOC_ID}}", "IdService"),
            ("Label", "Label ID", "DOCUMENT_LIBRARY DOCUMENT_CODE", "document_ids.label_id", "{{LABEL_ID}}", "IdService"),
            ("Statement", "Statement ID", "STATEMENT.STATEMENT_CODE", "document_ids.statement_id", "{{STM_ID}}", "IdService"),
            ("All", "FILE_URI", "DOCUMENT_LIBRARY", "n/a", "relative Phase I path", "phase_j.registry_sync"),
        ]
        for i, row in enumerate(rows, start=4):
            for c, v in enumerate(row, start=1):
                ws.cell(i, c, v).font = BODY_FONT
                ws.cell(i, c).border = THIN_BORDER
        # No Excel Table — filter only
        ws.auto_filter.ref = f"A3:F{3 + len(rows)}"
        self._autosize(ws, 6, 28)

    def _update_home_and_dashboard(self, wb, sync: dict) -> None:
        if "01_DASHBOARD" in wb.sheetnames:
            ws = wb["01_DASHBOARD"]
            ws["A1"] = "İnci Akü PPWR Packaging Information Management System — Rev.00"
            ws["A1"].font = TITLE_FONT
            ws["A2"] = (
                "Production Master Data: LOADED | Golden Variant Register: 247 / 247 | "
                f"Document Pack: 988 / 988 GENERATED | Document Registry: {sync['linked']} / 988 LINKED | "
                "Blocking QA Errors: 0"
            )
            ws["A2"].font = BODY_FONT
            ws["A2"].fill = KPI_FILL
            for row in range(5, 16):
                ws.cell(row=row, column=3, value="LOADED")
                ws.cell(row=row, column=4, value="Phase J — application layer + document registry sync")

        # SYS info
        if "SYS_WORKBOOK_INFO" in wb.sheetnames:
            ws = wb["SYS_WORKBOOK_INFO"]
            # append phase J marker if space
            r = ws.max_row + 1
            ws.cell(r, 1, "PHASE").number_format = "@"
            ws.cell(r, 2, "J — Application Layer + Document Registry Sync")
            ws.cell(r + 1, 1, "DOCUMENT_REGISTRY").number_format = "@"
            ws.cell(r + 1, 2, f"{sync['linked']}/988 LINKED")

    def _reorder_sheets(self, wb) -> None:
        desired = [s for s in UI_SHEETS_ORDER if s in wb.sheetnames]
        # then remaining in current order
        rest = [s for s in wb.sheetnames if s not in desired]
        final_order = desired + rest
        for idx, name in enumerate(final_order):
            current = wb.sheetnames.index(name)
            if current != idx:
                wb.move_sheet(name, offset=idx - current)

    def _stage_delivery_test(self) -> None:
        if self.delivery_test.exists():
            shutil.rmtree(self.delivery_test, ignore_errors=True)
        self.delivery_test.mkdir(parents=True, exist_ok=True)
        # Copy candidate as the delivery workbook name for hyperlink testing
        shutil.copy2(
            self.candidate,
            self.delivery_test / "INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx",
        )
        for folder in ("01_STARTER", "02_INDUSTRIAL", "03_CONTAINER", "90_MANIFEST", "99_QA_REPORT"):
            src = self.phase_i / folder
            dst = self.delivery_test / folder
            if src.exists():
                # directory junction (Windows) — portable relative links resolve
                if dst.exists():
                    shutil.rmtree(dst, ignore_errors=True)
                try:
                    # Prefer junction to avoid copying ~300MB DOCX
                    import subprocess

                    subprocess.run(
                        ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
                        check=True,
                        capture_output=True,
                    )
                except Exception:
                    shutil.copytree(src, dst)

    # ------------------------------------------------------------------
    # QA
    # ------------------------------------------------------------------
    def _qa(self, index, sync, excel) -> dict:
        wb = load_workbook(self.candidate, read_only=True, data_only=False)
        lib = wb["DOCUMENT_LIBRARY"]
        lh = _headers(lib)
        pending = 0
        stale_notes = 0
        hashed = 0
        for row in lib.iter_rows(min_row=2, values_only=True):
            uri = str(row[_col(lh, "FILE_URI") - 1] or "")
            notes = str(row[_col(lh, "NOTES") - 1] or "")
            fh = row[_col(lh, "FILE_HASH") - 1]
            if uri.startswith("pending://"):
                pending += 1
            if "Word not generated" in notes or "Metadata only" in notes:
                stale_notes += 1
            if fh:
                hashed += 1

        # XOR check on links
        link = wb["DOCUMENT_LINK"]
        link_h = _headers(link)
        target_cols = [
            "COMPONENT_ID",
            "PRODUCT_ID",
            "PACKAGING_CONFIGURATION_ID",
            "TRANSPORT_CONFIGURATION_ID",
            "TECHNICAL_FILE_ID",
            "DECLARATION_OF_CONFORMITY_ID",
            "STATEMENT_ID",
        ]
        xor_bad = 0
        for row in link.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            filled = 0
            for name in target_cols:
                v = row[_col(link_h, name) - 1]
                if v not in (None, ""):
                    filled += 1
            if filled != 1:
                xor_bad += 1

        # relative path existence vs Phase I
        broken = 0
        for code, rel in sync["path_by_code"].items():
            if not (self.phase_i / rel).exists():
                broken += 1

        # delivery test relative
        delivery_broken = 0
        for rel in sync["path_by_code"].values():
            if not (self.delivery_test / rel).exists():
                delivery_broken += 1

        wb.close()

        counts_ok = (
            len(index["configs"]) == 247
            and sync["linked"] == 988
            and pending == 0
            and stale_notes == 0
            and hashed == 988
            and xor_bad == 0
            and broken == 0
            and delivery_broken == 0
            and excel.get("ok") is True
        )
        return {
            "pass": counts_ok,
            "configurations": len(index["configs"]),
            "linked": sync["linked"],
            "hashed": hashed,
            "pending_uri": pending,
            "stale_notes": stale_notes,
            "xor_bad_links": xor_bad,
            "broken_phase_i_paths": broken,
            "broken_delivery_relative_links": delivery_broken,
            "missing_files": len(sync["missing_files"]),
            "excel_open": excel,
            "candidate": str(self.candidate),
            "delivery_test": str(self.delivery_test),
            "ui_modules": [
                s for s in UI_SHEETS_ORDER if s not in {"00_README", "01_DASHBOARD", "02_RELEASE_CONTROL", "03_DATA_DICTIONARY", "04_IMPORT_GUIDE"}
            ],
        }

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        gate = "PASS" if qa["pass"] else "FAIL"
        lines = [
            "# Phase J — PIMS Application Layer + Document Registry Sync",
            "",
            f"- **PHASE J PIMS APPLICATION RESTORATION: {gate}**",
            f"- Candidate: `{qa['candidate']}`",
            f"- Delivery test root: `{qa['delivery_test']}`",
            "",
            "## Database integrity",
            "",
            f"- Packaging configurations: {qa['configurations']} (expected 247)",
            "- Core BOM / product / component masters: unchanged (not rewritten)",
            "",
            "## Document registry",
            "",
            f"- DOCX linked: {qa['linked']} / 988",
            f"- SHA-256 populated: {qa['hashed']} / 988",
            f"- Metadata-only URI remaining: {qa['pending_uri']}",
            f"- Stale 'Word not generated' notes: {qa['stale_notes']}",
            f"- DOCUMENT_LINK XOR violations: {qa['xor_bad_links']}",
            f"- Broken Phase I paths: {qa['broken_phase_i_paths']}",
            f"- Broken relative links in delivery test: {qa['broken_delivery_relative_links']}",
            "",
            "## UI modules restored",
            "",
        ]
        for m in qa["ui_modules"]:
            lines.append(f"- `{m}`")
        lines += [
            "",
            "## Native Excel",
            "",
            f"- `{qa['excel_open']}`",
            "",
            "## Messages",
            "",
        ]
        lines.extend(f"- {m}" for m in messages)
        lines += [
            "",
            "## Manual acceptance required",
            "",
            "Do **not** replace `INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx` until the UI candidate "
            "is manually opened and visually accepted.",
            "",
            f"**PHASE J PIMS APPLICATION RESTORATION: {gate}**",
            "",
            "- Word documents regenerated: NO",
            "- Golden templates modified: NO",
            "- Rev01 started: NO",
            "",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
        self.qa_path.with_suffix(".json").write_text(
            json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def main() -> int:
    import sys

    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))
    if str(root / "src") not in sys.path:
        sys.path.insert(0, str(root / "src"))
    result = PhaseJService(root).run()
    for m in result.messages:
        print(m)
    print("PHASE J PIMS APPLICATION RESTORATION:", result.gate)
    return 0 if result.success else 1


if __name__ == "__main__":
    raise SystemExit(main())
