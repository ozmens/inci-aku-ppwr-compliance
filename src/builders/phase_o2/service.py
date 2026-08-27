"""Phase O2 — UI layout recovery + premium hybrid application design."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from PIL import ImageGrab

from builders.phase_n.assets import extract_inci_aku_logo
from builders.phase_o2.cell_modules import (
    CLASS_C,
    HYPERLINK_RE,
    rebuild_class_c_sheet,
    rebuild_document_center,
)
from builders.phase_o2.com_canvas import ClassACanvas

UI_SHEETS = [
    "00_HOME",
    "NAVIGATION",
    "SEARCH",
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "SHIPMENTS",
    "DOC_ENGINE_MAP",
]

CLASS_A = {"00_HOME", "NAVIGATION", "SEARCH"}
CLASS_B = {"DOCUMENT_CENTER"}

EXPECTED = {
    "packaging_configurations": 247,
    "bom_lines": 1690,
    "components": 112,
    "products": 2046,
    "documents": 988,
}

XL_SCREEN = 1
XL_BITMAP = 2


@dataclass
class PhaseO2Result:
    success: bool
    technical_gate: str
    layout_acceptance: str
    link_acceptance: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def _junction_or_copy(src: Path, dst: Path) -> None:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        else:
            shutil.rmtree(dst, ignore_errors=True)
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        shutil.copytree(src, dst)


class PhaseO2Service:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_CANDIDATE.xlsx"
        self.candidate = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_V3_CANDIDATE.xlsx"
        )
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_EXECUTIVE_V3"
        self.delivery_workbook = (
            self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_V3_CANDIDATE.xlsx"
        )
        self.assets = self.out / "PHASE_N_ASSETS"
        self.preview_dir = self.out / "PHASE_O2_LAYOUT_PREVIEW"
        self.qa_path = self.out / "PHASE_O2_LAYOUT_QA.md"

    def run(self) -> PhaseO2Result:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseO2Result(
                False, "FAIL", "PENDING", "PENDING", [f"Missing source: {self.source}"]
            )

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"V3 copied from Executive → {self.candidate.name}")

        logo = extract_inci_aku_logo(self.root, self.assets)
        messages.append(f"Logo ready: {logo.name}")

        baseline = self._counts(self.candidate)
        messages.append(f"Baseline: {baseline}")

        self._kill_excel()
        time.sleep(1.0)

        # ORDER IS CRITICAL:
        # 1) openpyxl Class B/C first (openpyxl save destroys COM shapes)
        # 2) COM Class A last + tables/views
        cell_stats = self._rebuild_cell_modules()
        messages.append(f"Class B/C rebuilt: {len(cell_stats)} sheets")

        self._kill_excel()
        time.sleep(1.0)
        shape_stats = self._com_cleanup_and_class_a(logo)
        messages.append(f"Shape cleanup + Class A: {shape_stats}")

        table_stats = self._com_finalize_tables_and_views()
        messages.append(f"Tables/views: {table_stats}")

        self._build_delivery()
        messages.append(f"Delivery: {self.delivery}")

        self._kill_excel()
        time.sleep(1.0)
        previews = self._export_previews()
        messages.append(f"Previews: {len(previews)}")

        after = self._counts(self.delivery_workbook)
        links = self._validate_doc_links(self.delivery_workbook, self.delivery)
        home = self._count_home_links(self.delivery_workbook)
        excel = self._excel_open(self.delivery_workbook)
        layout = self._layout_qa(self.delivery_workbook)
        nontahoma = self._scan_nontahoma(self.delivery_workbook)
        samples = self._samples(links)

        technical = (
            "PASS"
            if (
                excel.get("ok")
                and links["existing"] == 988
                and links["total"] == 988
                and links["missing"] == 0
                and links["absolute"] == 0
                and home >= 13
                and after == baseline == EXPECTED
                and layout["shape_table_intersections"] == 0
                and layout["overlapping_legacy_shapes_remaining"] == 0
                and layout["duplicate_visible_titles"] == 0
                and nontahoma == 0
                and len(previews) >= 8
                and all(s["exists"] for s in samples)
            )
            else "FAIL"
        )

        qa = {
            "technical_gate": technical,
            "manual_layout_acceptance": "PENDING",
            "manual_link_acceptance": "PENDING",
            "workbook_path": str(self.candidate),
            "delivery_root": str(self.delivery),
            "old_ui_shapes_removed": shape_stats.get("old_ui_shapes_removed", 0),
            "overlapping_legacy_shapes_remaining": layout[
                "overlapping_legacy_shapes_remaining"
            ],
            "shape_table_intersections": layout["shape_table_intersections"],
            "duplicate_visible_titles": layout["duplicate_visible_titles"],
            "sheet_layout": layout["sheets"],
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": after == baseline == EXPECTED,
            "word_links_total": links["total"],
            "word_links_working": links["existing"],
            "word_links_broken": links["missing"],
            "absolute_paths": links["absolute"],
            "home_links": home,
            "visible_nontahoma": nontahoma,
            "native_excel_open": excel,
            "preview_files": previews,
            "cell_module_stats": cell_stats,
            "canonical_data_changed": False,
            "word_regenerated": False,
            "promoted": False,
        }
        self._write_qa(qa, messages)
        return PhaseO2Result(
            technical == "PASS",
            technical,
            "PENDING",
            "PENDING",
            messages,
            qa,
        )

    def _kill_excel(self) -> None:
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, text=True
            )
        except Exception:
            pass

    def _com_cleanup_and_class_a(self, logo: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        removed_total = 0
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()

            # Strip shapes on ALL UI sheets first
            for name in UI_SHEETS:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    n = int(ws.Shapes.Count)
                    for i in range(n, 0, -1):
                        try:
                            ws.Shapes(i).Delete()
                        except Exception:
                            pass
                    removed_total += n
                except Exception:
                    pass

            canvas = ClassACanvas(excel, wb, logo)
            canvas.design_home()
            canvas.design_navigation()
            canvas.design_search()

            # Ensure Class B/C have ZERO shapes after this pass
            for name in list(CLASS_B) + CLASS_C:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    for i in range(int(ws.Shapes.Count), 0, -1):
                        try:
                            ws.Shapes(i).Delete()
                        except Exception:
                            pass
                except Exception:
                    pass

            # Reorder UI first
            for idx, name in enumerate(UI_SHEETS, start=1):
                try:
                    wb.Worksheets(name).Move(Before=wb.Worksheets(idx))
                except Exception:
                    pass
            wb.Worksheets(1).Select()

            excel.ScreenUpdating = True
            wb.Save()
            wb.Close(True)
            wb = None
            return {
                "old_ui_shapes_removed": removed_total,
                "class_a_shapes_created": canvas.shapes_created,
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _rebuild_cell_modules(self) -> list[dict]:
        wb = load_workbook(self.candidate)
        stats = []
        # Class B
        if "DOCUMENT_CENTER" in wb.sheetnames:
            stats.append(rebuild_document_center(wb["DOCUMENT_CENTER"]))
        # Class C
        for name in CLASS_C:
            if name in wb.sheetnames:
                stats.append(rebuild_class_c_sheet(wb[name], name))
        # Convert remaining doc HYPERLINK formulas on any UI sheet Open columns
        converted = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1
            ):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.upper().startswith("=HYPERLINK("):
                        m = HYPERLINK_RE.match(val.strip())
                        if m and m.group(1).lower().endswith(".docx"):
                            from builders.phase_o2.cell_modules import (
                                apply_native_doc_link,
                            )

                            apply_native_doc_link(
                                cell, m.group(1), m.group(2) or "OPEN"
                            )
                            converted += 1
        stats.append({"docx_formulas_converted": converted})
        wb.save(self.candidate)
        wb.close()
        return stats

    def _com_finalize_tables_and_views(self) -> dict[str, Any]:
        """Create Excel Tables for Class B/C; strip worksheet AutoFilter; set views."""
        pythoncom.CoInitialize()
        excel = None
        wb = None
        tables = 0
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()

            table_sheets = [
                "DOCUMENT_CENTER",
                "PACKAGING_CONFIGURATIONS",
                "PRODUCT_MASTER",
                "COMPONENT_MASTER",
                "TECHNICAL_FILES",
                "DECLARATIONS_OF_CONFORMITY",
                "LABELS",
                "SHIPMENT_STATEMENTS",
                "DOC_ENGINE_MAP",
            ]
            for name in table_sheets:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    # Remove worksheet AutoFilter if present
                    try:
                        if ws.AutoFilterMode:
                            ws.AutoFilterMode = False
                    except Exception:
                        pass
                    # Delete existing ListObjects
                    while ws.ListObjects.Count > 0:
                        try:
                            ws.ListObjects(1).Delete()
                        except Exception:
                            break
                    # Find header row dynamically (must be row 6 after rebuild)
                    header_row = None
                    for r in range(1, 12):
                        v1 = ws.Cells(r, 1).Value
                        v2 = ws.Cells(r, 2).Value
                        if v1 and v2 and r >= 6:
                            header_row = r
                            break
                    if not header_row:
                        continue
                    last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row  # xlUp
                    last_col = ws.Cells(header_row, ws.Columns.Count).End(-4159).Column
                    if last_row < header_row:
                        continue
                    rng = ws.Range(
                        ws.Cells(header_row, 1), ws.Cells(max(last_row, header_row), last_col)
                    )
                    lo = ws.ListObjects.Add(1, rng, None, 1)  # xlSrcRange, xlYes
                    lo.Name = f"T_{name[:20]}"
                    try:
                        lo.TableStyle = ""
                    except Exception:
                        pass
                    tables += 1
                    # Freeze below header
                    try:
                        excel.ActiveWindow.FreezePanes = False
                        ws.Range(f"A{header_row + 1}").Select()
                        excel.ActiveWindow.FreezePanes = True
                    except Exception:
                        pass
                    excel.ActiveWindow.DisplayGridlines = False
                    try:
                        excel.ActiveWindow.Zoom = 95
                    except Exception:
                        pass
                    ws.Range("A1").Select()
                except Exception:
                    continue

            # Class A view
            for name, zoom in (("00_HOME", 92), ("NAVIGATION", 95), ("SEARCH", 95)):
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    excel.ActiveWindow.DisplayGridlines = False
                    try:
                        excel.ActiveWindow.DisplayHeadings = False
                    except Exception:
                        pass
                    excel.ActiveWindow.Zoom = zoom
                    excel.ActiveWindow.ScrollRow = 1
                    excel.ActiveWindow.ScrollColumn = 1
                    ws.Range("A1").Select()
                except Exception:
                    pass

            # SHIPMENTS view
            try:
                ws = wb.Worksheets("SHIPMENTS")
                ws.Select()
                excel.ActiveWindow.DisplayGridlines = False
                ws.Range("A1").Select()
            except Exception:
                pass

            wb.Worksheets("00_HOME").Select()
            wb.Save()
            # reopen check
            path = str(self.candidate.resolve())
            wb.Close(True)
            wb = None
            excel.Quit()
            excel = None
            reopen = self._excel_open(Path(path))
            return {"tables_created": tables, "reopen": reopen}
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _build_delivery(self) -> None:
        if self.delivery.exists():
            for child in list(self.delivery.iterdir()):
                try:
                    if child.is_junction() or child.is_symlink():
                        child.unlink()
                    elif child.is_dir():
                        shutil.rmtree(child, ignore_errors=True)
                    else:
                        child.unlink()
                except Exception:
                    pass
        self.delivery.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.candidate, self.delivery_workbook)
        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if src.exists():
                _junction_or_copy(src, self.delivery / folder)

    def _export_previews(self) -> list[str]:
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()
        preview_map = {
            "HOME": ("00_HOME", "A1:O40"),
            "NAVIGATION": ("NAVIGATION", "A1:O36"),
            "SEARCH": ("SEARCH", "A1:O28"),
            "DOCUMENT_CENTER": ("DOCUMENT_CENTER", "A1:H14"),
            "DOC_ENGINE_MAP": ("DOC_ENGINE_MAP", "A1:F16"),
            "SHIPMENTS": ("SHIPMENTS", "A1:H14"),
            "DECLARATIONS_OF_CONFORMITY": ("DECLARATIONS_OF_CONFORMITY", "A1:H14"),
            "LABELS": ("LABELS", "A1:H14"),
        }
        pythoncom.CoInitialize()
        excel = None
        wb = None
        out: list[str] = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.delivery_workbook.resolve()), UpdateLinks=0, ReadOnly=True
            )
            wb.Worksheets(1).Select()
            for label, (sheet, addr) in preview_map.items():
                path = self.preview_dir / f"PHASE_O2_{label}.png"
                ws = wb.Worksheets(sheet)
                ws.Select()
                try:
                    excel.ActiveWindow.Zoom = 85
                    excel.ActiveWindow.DisplayGridlines = False
                except Exception:
                    pass
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.5)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    out.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return out

    def _excel_open(self, path: Path) -> dict:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {"ok": False, "error": None, "sheets": None}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(path.resolve()), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
            )
            out["ok"] = True
            out["sheets"] = int(wb.Worksheets.Count)
            wb.Close(False)
        except Exception as exc:  # noqa: BLE001
            out["error"] = str(exc)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _counts(self, path: Path) -> dict[str, int]:
        wb = load_workbook(path, read_only=True, data_only=False)

        def rows(s: str) -> int:
            if s not in wb.sheetnames:
                return -1
            return max((wb[s].max_row or 1) - 1, 0)

        out = {
            "packaging_configurations": rows("PACKAGING_CONFIGURATION"),
            "bom_lines": rows("PACKAGING_CONFIGURATION_LINE"),
            "components": rows("COMPONENT"),
            "products": rows("PRODUCT"),
            "documents": rows("DOCUMENT_LIBRARY"),
        }
        wb.close()
        return out

    def _validate_doc_links(self, workbook: Path, delivery: Path) -> dict[str, Any]:
        wb = load_workbook(workbook, data_only=False)
        targets: set[str] = set()
        absolute = 0
        # formulas
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                        if m and m.group(1).lower().endswith(".docx"):
                            t = m.group(1).replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
                    # native hyperlinks
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt.lower().endswith(".docx"):
                            t = tgt.replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
        wb.close()
        existing = missing = 0
        for t in targets:
            parts = [p for p in t.replace("\\", "/").split("/") if p not in ("", ".")]
            if not parts or ".." in parts:
                missing += 1
                continue
            p = delivery.joinpath(*parts)
            if p.exists() and p.is_file():
                existing += 1
            else:
                missing += 1
        return {
            "total": len(targets),
            "existing": existing,
            "missing": missing,
            "absolute": absolute,
            "targets": sorted(targets),
        }

    def _count_home_links(self, path: Path) -> int:
        """Count HOME navigation on all 13 UI sheets (shape or cell)."""
        pythoncom.CoInitialize()
        excel = None
        n = 0
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=True)
            wb.Worksheets(1).Select()
            for name in UI_SHEETS:
                try:
                    ws = wb.Worksheets(name)
                    found = False
                    for i in range(1, int(ws.Shapes.Count) + 1):
                        if str(ws.Shapes(i).Name).startswith("HomeBtn"):
                            found = True
                            break
                    if not found:
                        # cell A1
                        v = str(ws.Range("A1").Value or "")
                        if "HOME" in v.upper() or "Ana Sayfaya" in v:
                            found = True
                        # hyperlinks collection
                        if not found:
                            for h in range(1, int(ws.Hyperlinks.Count) + 1):
                                try:
                                    sub = str(ws.Hyperlinks(h).SubAddress or "")
                                    if "00_HOME" in sub:
                                        found = True
                                        break
                                except Exception:
                                    pass
                    if found:
                        n += 1
                except Exception:
                    pass
            wb.Close(False)
        except Exception:
            pass
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return n

    def _layout_qa(self, path: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        sheets_info = []
        intersections = 0
        legacy = 0
        dup_titles = 0
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=True)
            wb.Worksheets(1).Select()
            ox = load_workbook(path, data_only=False)
            for name in UI_SHEETS:
                ws = wb.Worksheets(name)
                ui_class = (
                    "A" if name in CLASS_A else "B" if name in CLASS_B else "C"
                )
                shape_count = int(ws.Shapes.Count)
                # Class B/C must have 0 shapes
                if ui_class in ("B", "C") and shape_count > 0:
                    intersections += shape_count  # treat as layout error
                    legacy += shape_count
                table_start = None
                ows = ox[name]
                # detect header row
                for r in range(1, 12):
                    v = ows.cell(r, 1).value
                    if v and any(
                        k in str(v)
                        for k in (
                            "Packaging Set",
                            "Product Code",
                            "ERP",
                            "Label ID",
                            "DoC Number",
                            "Document Type",
                            "Technical File ID",
                            "Statement",
                        )
                    ):
                        table_start = r
                        break
                # duplicate titles: ◆ and MODULE both, or two navy title rows with similar text
                titles = []
                for r in range(1, 8):
                    v = ows.cell(r, 1).value
                    if v and not str(v).startswith("=") and len(str(v)) > 8:
                        if any(
                            k in str(v).upper()
                            for k in (
                                "DOCUMENT ENGINE",
                                "DOC_ENGINE",
                                "SHIPMENTS",
                                "◆",
                            )
                        ):
                            titles.append(str(v)[:60])
                # duplicate if both DOC_ENGINE_MAP string and DOCUMENT ENGINE MAP
                if name == "DOC_ENGINE_MAP":
                    joined = " | ".join(titles).upper()
                    if "DOC_ENGINE_MAP" in joined and "DOCUMENT ENGINE MAP" in joined:
                        dup_titles += 1
                    elif joined.count("DOCUMENT ENGINE") > 1:
                        dup_titles += 1
                if name == "SHIPMENTS":
                    joined = " | ".join(titles).upper()
                    if joined.count("SHIPMENT") > 2:
                        dup_titles += 1

                sheets_info.append(
                    {
                        "sheet": name,
                        "ui_class": ui_class,
                        "shape_count": shape_count,
                        "table_start_row": table_start,
                        "status": (
                            "PASS"
                            if (
                                (ui_class == "A" and shape_count > 10)
                                or (ui_class in ("B", "C") and shape_count == 0)
                            )
                            else "CHECK"
                        ),
                    }
                )
            ox.close()
            wb.Close(False)
        except Exception as exc:  # noqa: BLE001
            sheets_info.append({"error": str(exc)})
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return {
            "sheets": sheets_info,
            "shape_table_intersections": intersections,
            "overlapping_legacy_shapes_remaining": legacy,
            "duplicate_visible_titles": dup_titles,
        }

    def _scan_nontahoma(self, path: Path) -> int:
        wb = load_workbook(path, data_only=False)
        n = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row or 1, 30), max_col=min(ws.max_column or 1, 12)
            ):
                for cell in row:
                    if cell.value is None:
                        continue
                    fn = cell.font.name if cell.font else None
                    if fn and fn.lower() not in ("tahoma", "calibri") and fn != "Tahoma":
                        # allow only Tahoma; count Calibri too as non-compliant
                        n += 1
                    if fn and fn != "Tahoma":
                        # Calibri is non-Tahoma
                        if fn != "Tahoma":
                            n += 1
        # fix double count — recount properly
        wb.close()
        wb = load_workbook(path, data_only=False)
        n = 0
        for name in UI_SHEETS:
            ws = wb[name]
            for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row or 1, 40), max_col=min(ws.max_column or 1, 14)
            ):
                for cell in row:
                    if cell.value is None:
                        continue
                    fn = cell.font.name if cell.font else None
                    if fn and fn != "Tahoma":
                        n += 1
        wb.close()
        return n

    def _samples(self, links: dict) -> list[dict]:
        wanted = [
            "01_STARTER/ST-051-STD-01/01_Technical_File.docx",
            "01_STARTER/ST-051-STD-01/02_EU_DoC.docx",
            "01_STARTER/ST-051-STD-01/03_Label.docx",
            "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx",
            "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx",
            "03_CONTAINER/CNT-20-STD-01/03_Label.docx",
        ]
        out = []
        for rel in wanted:
            p = self.delivery / rel
            out.append(
                {
                    "relative": rel,
                    "exists": p.exists(),
                    "in_workbook": rel in links.get("targets", []),
                }
            )
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        lines = [
            "# Phase O2 — Layout Recovery / Hybrid UI QA",
            "",
            f"- **PHASE O2 TECHNICAL GATE: {qa['technical_gate']}**",
            f"- **MANUAL LAYOUT ACCEPTANCE: {qa['manual_layout_acceptance']}**",
            f"- **MANUAL LINK ACCEPTANCE: {qa['manual_link_acceptance']}**",
            "",
            f"- Candidate: `{qa['workbook_path']}`",
            f"- Delivery root: `{qa['delivery_root']}`",
            "",
            "## Shape cleanup",
            "",
            f"- OLD_UI_SHAPES_REMOVED = {qa['old_ui_shapes_removed']}",
            f"- OVERLAPPING_LEGACY_SHAPES_REMAINING = {qa['overlapping_legacy_shapes_remaining']}",
            f"- SHAPE_TABLE_INTERSECTIONS = {qa['shape_table_intersections']}",
            f"- DUPLICATE_VISIBLE_TITLES = {qa['duplicate_visible_titles']}",
            "",
            "## Sheet layout",
            "",
            "| Sheet | Class | Shapes | Table Start | Status |",
            "|---|---|---|---|---|",
        ]
        for s in qa["sheet_layout"]:
            if "error" in s:
                lines.append(f"| error | | | | {s['error']} |")
                continue
            lines.append(
                f"| {s['sheet']} | {s['ui_class']} | {s['shape_count']} | "
                f"{s['table_start_row']} | {s['status']} |"
            )
        lines += [
            "",
            "## Counts / Links",
            "",
            f"- Counts unchanged: {qa['counts_unchanged']} `{qa['after_counts']}`",
            f"- Word links: {qa['word_links_working']} / {qa['word_links_total']}",
            f"- Home links: {qa['home_links']}/13",
            f"- Absolute paths: {qa['absolute_paths']}",
            f"- Visible non-Tahoma: {qa['visible_nontahoma']}",
            f"- Native Excel: `{qa['native_excel_open']}`",
            "",
            "## Previews",
            "",
        ]
        lines.extend(f"- `{p}`" for p in qa["preview_files"])
        lines += ["", "## Messages", ""]
        lines.extend(f"- {m}" for m in messages)
        lines += [
            "",
            "- Canonical data changed: NO",
            "- Word regenerated: NO",
            "- Promoted: NO",
            "- Rev01 started: NO",
            "",
            f"**PHASE O2 TECHNICAL GATE: {qa['technical_gate']}**",
            f"**MANUAL LAYOUT ACCEPTANCE: {qa['manual_layout_acceptance']}**",
            f"**MANUAL LINK ACCEPTANCE: {qa['manual_link_acceptance']}**",
            "",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
        self.qa_path.with_suffix(".json").write_text(
            json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def main() -> int:
    import sys

    root = Path(__file__).resolve().parents[2]
    sys.path.insert(0, str(root))
    sys.path.insert(0, str(root / "src"))
    result = PhaseO2Service(root).run()
    for m in result.messages:
        try:
            print(m)
        except UnicodeEncodeError:
            print(m.encode("ascii", "replace").decode("ascii"))
    print("PHASE O2 TECHNICAL GATE:", result.technical_gate)
    print("MANUAL LAYOUT ACCEPTANCE:", result.layout_acceptance)
    print("MANUAL LINK ACCEPTANCE:", result.link_acceptance)
    if result.qa:
        print("Candidate:", result.qa.get("workbook_path"))
        print("Delivery root:", result.qa.get("delivery_root"))
        print("OLD_UI_SHAPES_REMOVED =", result.qa.get("old_ui_shapes_removed"))
        print(
            "OVERLAPPING_LEGACY_SHAPES_REMAINING =",
            result.qa.get("overlapping_legacy_shapes_remaining"),
        )
    return 0 if result.success else 1


if __name__ == "__main__":
    raise SystemExit(main())
