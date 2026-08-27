"""Phase O2 — Class C / B cell-based operational module recovery."""

from __future__ import annotations

import re
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

NAVY = "0E2A47"
STEEL = "315E87"
GOLD = "C8A24A"
IVORY = "F7F5F0"
STONE = "F3F1EB"
WHITE = "FFFFFF"
INK = "1C2430"
MUTED = "5C6B7A"
LINK = "1F5C99"
BAND = "F3F6F9"
FONT = "Tahoma"

NONE = Border()
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)

CLASS_C = [
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "SHIPMENTS",
    "DOC_ENGINE_MAP",
]

MODULE_META = {
    "PACKAGING_CONFIGURATIONS": (
        "PACKAGING CONFIGURATIONS",
        "Final packaging set register — 247 controlled configurations",
        "247 configs  ·  Starter 240  ·  Industrial 3  ·  Container 4",
    ),
    "PRODUCT_MASTER": (
        "PRODUCT MASTER",
        "Products linked to packaging configurations",
        "2,046 products  ·  247 configurations",
    ),
    "COMPONENT_MASTER": (
        "COMPONENT MASTER",
        "Packaging component catalogue",
        "112 components",
    ),
    "TECHNICAL_FILES": (
        "TECHNICAL FILES",
        "PPWR technical file index — one per packaging configuration",
        "247 / 247 linked",
    ),
    "DECLARATIONS_OF_CONFORMITY": (
        "DECLARATIONS OF CONFORMITY",
        "EU Declaration of Conformity index",
        "247 / 247 linked",
    ),
    "LABELS": (
        "LABELS",
        "Packaging identification label index",
        "247 / 247 linked",
    ),
    "SHIPMENT_STATEMENTS": (
        "SHIPMENT STATEMENTS",
        "Shipment statement index — one per packaging configuration",
        "247 / 247 linked",
    ),
    "SHIPMENTS": (
        "SHIPMENTS",
        "Transactional shipment register",
        "Rev.00 baseline — no transactional shipment records loaded",
    ),
    "DOC_ENGINE_MAP": (
        "DOCUMENT ENGINE MAP",
        "Read-only mapping • Python remains document authority",
        "Controlled document-field mapping reference",
    ),
    "DOCUMENT_CENTER": (
        "DOCUMENT CENTER",
        "Per-configuration document pack — Technical File, DoC, Label, Statement",
        "247 configurations  ·  988 documents  ·  988 linked  ·  Rev.00",
    ),
}

HYPERLINK_RE = re.compile(
    r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)', re.IGNORECASE
)


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def _font(size=9, bold=False, color=INK, underline=None) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color, underline=underline)


def find_header_row(ws) -> int | None:
    """Find real table header: requires ≥2 populated header-like cells."""
    keys = (
        "Packaging Set Code",
        "Final Configuration ID",
        "Product Code",
        "ERP Component Code",
        "Label ID",
        "DoC Number",
        "Document Type",
        "Technical File Code",
        "Technical File ID",
        "Statement ID",
        "Configuration ID",
        "Open Technical",
        "Open EU",
        "Open Label",
        "Open Statement",
        "Open DoC",
        "Logical Field",
    )
    for r in range(1, min(25, (ws.max_row or 1) + 1)):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if v1 is None or v2 is None:
            continue
        s1, s2 = str(v1), str(v2)
        if s1.startswith("=") or s1.startswith("◆") or "Ana Sayfaya" in s1:
            continue
        if s1.startswith("←") or s1 in ("NAVIGATION", "SEARCH", "DOCUMENT CENTER"):
            continue
        # reject prose/summary lines
        if " · " in s1 or s1.endswith(".") or len(s1) > 80:
            continue
        if "|" in s1 and "247" in s1:
            continue
        populated = sum(
            1
            for c in range(1, min((ws.max_column or 1), 12) + 1)
            if ws.cell(r, c).value not in (None, "")
        )
        if populated < 3:
            continue
        joined = f"{s1} | {s2}"
        if any(k in joined for k in keys) or any(
            k in s1 or k in s2
            for k in (
                "Packaging Set",
                "Product Code",
                "ERP Component",
                "Label ID",
                "DoC Number",
                "Document Type",
                "Technical File",
                "Statement ID",
                "Logical Field",
            )
        ):
            return r
    return None


def _export_cell(cell) -> Any:
    """Preserve doc links when reading cells (formula or native hyperlink)."""
    if cell.hyperlink is not None:
        tgt = cell.hyperlink.target or ""
        if tgt.lower().endswith(".docx"):
            label = cell.value if isinstance(cell.value, str) and cell.value else "OPEN"
            if isinstance(label, str) and label.startswith("="):
                label = "OPEN"
            return f'=HYPERLINK("{tgt.replace(chr(92), "/")}","{label}")'
    return cell.value


def extract_table(ws) -> tuple[list[Any], list[list[Any]]]:
    """Return (headers, data_rows) from detected header row downward."""
    hr = find_header_row(ws)
    if hr is None:
        return [], []
    max_col = ws.max_column or 1
    headers = [_export_cell(ws.cell(hr, c)) for c in range(1, max_col + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    if len(headers) < 3:
        return [], []
    data = []
    for r in range(hr + 1, (ws.max_row or hr) + 1):
        row = [_export_cell(ws.cell(r, c)) for c in range(1, len(headers) + 1)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        data.append(row)
    return headers, data


def clear_sheet_values(ws) -> None:
    # Unmerge
    for m in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(m))
        except Exception:
            pass
    ws.auto_filter.ref = None
    ws.freeze_panes = None
    if ws.tables:
        for name in list(ws.tables.keys()):
            del ws.tables[name]
    max_r = ws.max_row or 1
    max_c = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name=FONT)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.hyperlink = None


def write_nav_row(ws, max_fill_col: int = 8) -> None:
    """Row 1 cell-based navigation — no shapes."""
    ws.row_dimensions[1].height = 22
    links = [
        (1, "← HOME", "00_HOME"),
        (2, "NAVIGATION", "NAVIGATION"),
        (3, "SEARCH", "SEARCH"),
        (4, "DOCUMENT CENTER", "DOCUMENT_CENTER"),
    ]
    for col, label, sheet in links:
        cell = ws.cell(1, col, label)
        cell.font = _font(9, True, NAVY, underline="single")
        cell.fill = _fill(STONE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate, location=f"'{sheet}'!A1", tooltip=label
        )
        cell.border = NONE
    for c in range(5, max_fill_col):
        ws.cell(1, c).fill = _fill(STONE)
        ws.cell(1, c).border = NONE
    rev = ws.cell(1, max_fill_col, "REV.00")
    rev.font = _font(9, True, NAVY)
    rev.fill = _fill(GOLD)
    rev.alignment = Alignment(horizontal="center", vertical="center")


def write_title_block(ws, title: str, subtitle: str, info: str, cols: int) -> None:
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 8
    # Row 2–3 navy fills across columns (no giant merge)
    for c in range(1, cols + 1):
        ws.cell(2, c).fill = _fill(NAVY)
        ws.cell(2, c).border = NONE
        ws.cell(3, c).fill = _fill(STEEL)
        ws.cell(3, c).border = NONE
        ws.cell(4, c).fill = _fill(IVORY)
        ws.cell(4, c).border = NONE
        ws.cell(5, c).fill = _fill(IVORY)
        ws.cell(5, c).border = NONE
    t = ws.cell(2, 1, title)
    t.font = _font(14, True, WHITE)
    t.alignment = Alignment(vertical="center")
    s = ws.cell(3, 1, subtitle)
    s.font = _font(9, False, "D5DEE8")
    s.alignment = Alignment(vertical="center")
    i = ws.cell(4, 1, info)
    i.font = _font(8, True, NAVY)
    i.alignment = Alignment(vertical="center")
    # gold hairline via row 5 thin gold fill strip on row 5 top - use gold on row5 partial
    for c in range(1, min(cols, 4) + 1):
        ws.cell(5, c).fill = _fill(GOLD)


def apply_native_doc_link(cell, target: str, label: str = "OPEN") -> None:
    """Convert to native hyperlink object (relative path)."""
    target = target.replace("\\", "/").lstrip("./")
    cell.value = label
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=target, tooltip=target)
    cell.font = _font(9, True, LINK, underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def maybe_convert_hyperlink_formula(cell) -> bool:
    val = cell.value
    if not isinstance(val, str):
        return False
    m = HYPERLINK_RE.match(val.strip())
    if not m:
        return False
    target, label = m.group(1), m.group(2) or "OPEN"
    if target.startswith("#"):
        # internal — keep as formula or convert to location hyperlink
        loc = target[1:]
        cell.value = label
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=loc, tooltip=label)
        cell.font = _font(9, True, NAVY, underline="single")
        return True
    if target.lower().endswith(".docx"):
        apply_native_doc_link(cell, target, label if label else "OPEN")
        return True
    return False


def write_data_table(ws, headers: list, data: list[list], start_row: int = 6) -> int:
    cols = max(len(headers), 1)
    # header
    ws.row_dimensions[start_row].height = 26
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(start_row, c, h)
        cell.font = _font(9, True, WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HAIR
    for r_i, row in enumerate(data):
        r = start_row + 1 + r_i
        band = _fill(BAND) if r_i % 2 else _fill(WHITE)
        for c in range(1, cols + 1):
            val = row[c - 1] if c - 1 < len(row) else None
            cell = ws.cell(r, c, val)
            cell.fill = band
            cell.border = HAIR
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if isinstance(val, str) and val.upper().startswith("=HYPERLINK("):
                maybe_convert_hyperlink_formula(cell)
            elif isinstance(val, str) and val.lower().endswith(".docx"):
                apply_native_doc_link(cell, val, "OPEN")
            else:
                cell.font = _font(9, False, INK)
    last = start_row + len(data)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.sheet_view.showGridLines = False
    # Column widths
    for c, h in enumerate(headers, start=1):
        letter = get_column_letter(c)
        hs = str(h or "")
        if "Open" in hs or hs.upper() == "OPEN":
            ws.column_dimensions[letter].width = 10
        elif "Variant" in hs or "Description" in hs or "Mapping" in hs or "Source" in hs:
            ws.column_dimensions[letter].width = 36
            # wrap for text-heavy
            for r in range(start_row + 1, last + 1):
                ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
        elif "ID" in hs or "Code" in hs:
            ws.column_dimensions[letter].width = 26
        else:
            ws.column_dimensions[letter].width = 18
    return last


def rebuild_class_c_sheet(ws, name: str) -> dict[str, Any]:
    title, subtitle, info = MODULE_META[name]
    headers, data = extract_table(ws)

    # Special: SHIPMENTS may have no real table
    if name == "SHIPMENTS":
        clear_sheet_values(ws)
        write_nav_row(ws, 8)
        write_title_block(ws, title, subtitle, info, 8)
        # info panel in cells below
        ws.cell(6, 1, "TRANSACTIONAL SHIPMENT REGISTER").font = _font(11, True, NAVY)
        ws.cell(7, 1, "Current status:").font = _font(9, True, MUTED)
        ws.cell(8, 1, "No transactional shipment records loaded in Rev.00 baseline.").font = _font(
            9, False, INK
        )
        for r, (label, sheet) in enumerate(
            [
                ("→ Open Shipment Database", "SHIPMENT"),
                ("→ Open Shipment Statements", "SHIPMENT_STATEMENTS"),
            ],
            start=10,
        ):
            cell = ws.cell(r, 1, label)
            cell.font = _font(10, True, LINK, underline="single")
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate, location=f"'{sheet}'!A1", tooltip=label
            )
        for c in range(1, 9):
            ws.column_dimensions[get_column_letter(c)].width = 18
        ws.column_dimensions["A"].width = 42
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A6"
        return {
            "sheet": name,
            "class": "C",
            "table_start_row": None,
            "rows": 0,
            "duplicate_titles_cleared": True,
        }

    if name == "DOC_ENGINE_MAP" and not headers:
        # fallback known headers
        headers = [
            "Document Type",
            "Logical Field",
            "PIMS Source",
            "DocumentContext Field",
            "Runtime Token / Mapping",
            "Builder Module",
        ]

    if len(headers) < 3:
        raise ValueError(
            f"{name}: table header detection failed (cols={len(headers)}). "
            "Aborting to protect canonical UI data."
        )

    clear_sheet_values(ws)
    cols = max(len(headers), 8)
    write_nav_row(ws, min(cols, 8))
    write_title_block(ws, title, subtitle, info, min(cols, 8))
    # fill remaining title cols if table wider
    if cols > 8:
        for c in range(9, cols + 1):
            for r in range(1, 6):
                if r == 1:
                    ws.cell(r, c).fill = _fill(STONE)
                elif r == 2:
                    ws.cell(r, c).fill = _fill(NAVY)
                elif r == 3:
                    ws.cell(r, c).fill = _fill(STEEL)
                else:
                    ws.cell(r, c).fill = _fill(IVORY)

    last = write_data_table(ws, headers, data, start_row=6)

    # Wrap specific DOC_ENGINE_MAP columns
    if name == "DOC_ENGINE_MAP":
        for c in range(3, min(len(headers), 6) + 1):
            for r in range(7, last + 1):
                ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = 32
        for r in range(7, last + 1):
            ws.row_dimensions[r].height = 36

    return {
        "sheet": name,
        "class": "C",
        "table_start_row": 6,
        "rows": len(data),
        "cols": len(headers),
        "duplicate_titles_cleared": True,
    }


def rebuild_document_center(ws) -> dict[str, Any]:
    """Class B: cell header + summary chips in cells (no floating body shapes)."""
    title, subtitle, info = MODULE_META["DOCUMENT_CENTER"]
    headers, data = extract_table(ws)
    if len(headers) < 3 or len(data) < 200:
        raise ValueError(
            f"DOCUMENT_CENTER extract failed: headers={len(headers)} rows={len(data)}"
        )
    clear_sheet_values(ws)
    cols = max(len(headers), 8)
    write_nav_row(ws, 8)
    write_title_block(ws, title, subtitle, info, 8)
    # KPI chips as cells in row 4 already has info; enhance row 4
    ws.cell(4, 1, "247 CONFIGURATIONS").font = _font(8, True, NAVY)
    ws.cell(4, 3, "988 DOCUMENTS").font = _font(8, True, NAVY)
    ws.cell(4, 5, "988 LINKED").font = _font(8, True, NAVY)
    ws.cell(4, 7, "REV.00").font = _font(8, True, NAVY)
    for c in (1, 3, 5, 7):
        ws.cell(4, c).fill = _fill("E8EEF4")
    last = write_data_table(ws, headers, data, start_row=6)
    return {
        "sheet": "DOCUMENT_CENTER",
        "class": "B",
        "table_start_row": 6,
        "rows": len(data),
        "cols": len(headers),
        "duplicate_titles_cleared": True,
    }
