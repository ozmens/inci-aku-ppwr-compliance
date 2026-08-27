"""
İnci Akü operational UX (Phases 9–11).

Does NOT change database architecture or entity/table names.

PPWR operational model (Phase 11):
- Primary object: PACKAGING_CONFIGURATION
- Hierarchy: Component → Packaging Config → Technical File →
  Commercial Scenario (Incoterms) → DoC Variant → Shipment Statement
- Lot Number maps to SHIPMENT.EXTERNAL_REF (no new column)

Visible sheets (max 8):
  Dashboard, Packaging Configuration, Component Master, Product Master,
  Shipment, Statements, Technical File, Declaration of Conformity
"""

from __future__ import annotations

from copy import copy

from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from models.registry import SchemaRegistry

from .sheet_builder import excel_col_letter
from .ui_theme import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    FILL_CALC,
    FILL_DASH,
    FILL_INPUT,
    FILL_LOCKED,
    FILL_NAV,
    FILL_OUTPUT,
    FILL_PRIMARY,
    FILL_PRIMARY_MID,
    FONT_BODY,
    FONT_HEADER,
    FONT_KPI,
    FONT_LINK,
    FONT_MUTED,
    FONT_NAV,
    FONT_SUBTITLE,
    FONT_TITLE,
    TABLE_STYLE,
    THIN,
    style_nav_button_cell,
)

# Internal sheet name → visible tab title (entity/table names unchanged)
VISIBLE_SHEET_TITLES: dict[str, str] = {
    "DASHBOARD": "Dashboard",
    "COMPONENT": "Component Master",
    "PRODUCT": "Product Master",
    "PACKAGING_CONFIGURATION": "Packaging Configuration",
    "SHIPMENT": "Shipment",
    "STATEMENT": "Statements",
    "TECHNICAL_FILE": "Technical File",
    "DECLARATION_OF_CONFORMITY": "Declaration of Conformity",
}

VISIBLE_ORDER = [
    "DASHBOARD",
    "PACKAGING_CONFIGURATION",
    "COMPONENT",
    "PRODUCT",
    "SHIPMENT",
    "STATEMENT",
    "TECHNICAL_FILE",
    "DECLARATION_OF_CONFORMITY",
]

CHROME_ROWS = 4  # rows inserted above data tables for nav/instructions


class InciOpsUX:
    """Apply İnci Akü daily-operations UX on top of the frozen workbook."""

    def __init__(self, workbook, registry: SchemaRegistry, settings) -> None:
        self.wb = workbook
        self.registry = registry
        self.settings = settings
        self.password = getattr(settings, "UI_SHEET_PASSWORD", "PIMS_UI")

    def apply(self) -> None:
        self._colocate_packaging_lines()
        self._insert_chrome_on_ops_sheets()
        self._add_business_panels()
        self._rename_visible_sheets()
        self._refresh_named_ranges_after_rename()
        self._rebuild_dashboard()
        self._set_visibility_and_order()
        self._reprotect_visible_sheets()

    # ------------------------------------------------------------------
    # Packaging recipe: keep LINE table on same visible sheet as header
    # ------------------------------------------------------------------

    def _colocate_packaging_lines(self) -> None:
        if (
            "PACKAGING_CONFIGURATION" not in self.wb.sheetnames
            or "PACKAGING_CONFIGURATION_LINE" not in self.wb.sheetnames
        ):
            return

        src = self.wb["PACKAGING_CONFIGURATION_LINE"]
        dst = self.wb["PACKAGING_CONFIGURATION"]

        # Destination header table extent
        hdr_table = dst.tables.get("T_PACKAGING_CONFIGURATION")
        start_row = (dst.max_row + 3) if not hdr_table else int(hdr_table.ref.split(":")[1][1:]) + 3

        # Title for recipe section
        dst.cell(row=start_row - 1, column=1).value = "PACKAGING RECIPE LINES (edit quantities here)"
        dst.cell(row=start_row - 1, column=1).font = FONT_SUBTITLE

        # Read source
        headers = [c.value for c in src[1]]
        rows = []
        for r in range(2, src.max_row + 1):
            rows.append([src.cell(row=r, column=c).value for c in range(1, len(headers) + 1)])

        # Remove old table from source sheet
        if "T_PACKAGING_CONFIGURATION_LINE" in src.tables:
            del src.tables["T_PACKAGING_CONFIGURATION_LINE"]

        # Write to destination
        for c, h in enumerate(headers, start=1):
            cell = dst.cell(row=start_row, column=c, value=h)
            cell.fill = FILL_PRIMARY
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
        body_rows = rows if rows else [[None] * len(headers)]
        for r_idx, row in enumerate(body_rows, start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                cell = dst.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = FILL_INPUT
                cell.border = THIN
                cell.protection = Protection(locked=False)

        last_col = excel_col_letter(len(headers))
        last_row = start_row + len(body_rows)
        if "T_PACKAGING_CONFIGURATION_LINE" in dst.tables:
            del dst.tables["T_PACKAGING_CONFIGURATION_LINE"]
        dst.add_table(
            Table(
                displayName="T_PACKAGING_CONFIGURATION_LINE",
                ref=f"A{start_row}:{last_col}{last_row}",
                tableStyleInfo=TABLE_STYLE,
            )
        )

        # Leave a stub note on original sheet (stays hidden later)
        src.delete_rows(1, src.max_row)
        src["A1"] = "NOTE"
        src["B1"] = "DETAIL"
        src["A2"] = "MOVED"
        src["B2"] = "T_PACKAGING_CONFIGURATION_LINE lives on Packaging Configuration sheet"
        src.add_table(Table(displayName="T_PKG_LINE_STUB", ref="A1:B2"))

    # ------------------------------------------------------------------
    # Chrome: HOME / BACK / Dashboard above each ops table
    # ------------------------------------------------------------------

    def _insert_chrome_on_ops_sheets(self) -> None:
        for internal in VISIBLE_ORDER:
            if internal == "DASHBOARD" or internal not in self.wb.sheetnames:
                continue
            ws = self.wb[internal]
            # Remove legacy UIBuilder "Dashboard" link cells that sit on header rows
            for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), max_col=ws.max_column):
                for cell in row:
                    if cell.value == "Dashboard" and cell.hyperlink:
                        cell.value = None
                        cell.hyperlink = None
            # Shift existing content/tables down
            ws.insert_rows(1, CHROME_ROWS)
            self._shift_tables(ws, CHROME_ROWS)
            self._trim_entity_table_refs(ws, internal)
            self._write_chrome(ws, internal)

    def _trim_entity_table_refs(self, ws, internal_name: str) -> None:
        """Keep Excel Table refs limited to schema columns (exclude stray link cells)."""
        try:
            table_def = self.registry.get(internal_name)
            col_count = len(table_def.columns)
        except KeyError:
            col_count = None
        for name in list(ws.tables):
            table = ws.tables[name]
            if name == "T_PACKAGING_CONFIGURATION_LINE":
                try:
                    line_def = self.registry.get("PACKAGING_CONFIGURATION_LINE")
                    n = len(line_def.columns)
                except KeyError:
                    n = 8
            elif col_count and name == f"T_{internal_name}":
                n = col_count
            else:
                continue
            start = table.ref.split(":")[0]
            end = table.ref.split(":")[1]
            start_row = int("".join(ch for ch in start if ch.isdigit()))
            end_row = int("".join(ch for ch in end if ch.isdigit()))
            table.ref = f"A{start_row}:{excel_col_letter(n)}{end_row}"

    def _shift_tables(self, ws, rows: int) -> None:
        # openpyxl Table.ref must be rewritten after insert_rows
        for name in list(ws.tables.keys()):
            table = ws.tables[name]
            ref = table.ref  # e.g. A1:G10
            start, end = ref.split(":")
            def shift_addr(addr: str) -> str:
                col = "".join(ch for ch in addr if ch.isalpha())
                row = int("".join(ch for ch in addr if ch.isdigit()))
                return f"{col}{row + rows}"
            table.ref = f"{shift_addr(start)}:{shift_addr(end)}"

    def _write_chrome(self, ws, internal_name: str) -> None:
        title = VISIBLE_SHEET_TITLES[internal_name]
        ws.merge_cells("A1:D1")
        ws["A1"] = f"İNCI AKÜ  |  {title}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_DASH

        # Nav buttons
        self._link_btn(ws, "F1", "HOME", "#Dashboard!B4")
        self._link_btn(ws, "G1", "Dashboard", "#Dashboard!B4")
        self._link_btn(ws, "H1", "BACK", "#Dashboard!B4")

        tips = {
            "COMPONENT": "Hierarchy bottom: parts of Packaging Configurations. Weight + supplier + evidence required.",
            "PRODUCT": "ERP link only — NOT the PPWR primary object. Products link to Packaging Configurations via scenarios.",
            "PACKAGING_CONFIGURATION": "PRIMARY OBJECT (Dashboard STEP 1). One physical arrangement. Owns exactly one Technical File. Edit recipe here.",
            "SHIPMENT": "Dashboard STEP 3: paste Packaging Config + Scenario + Qty + Lot Number (EXTERNAL_REF). Customer/Country/Incoterms come from scenario.",
            "STATEMENT": "Dashboard STEP 4: Shipment Statement. Link shipment; composition assembles automatically.",
            "TECHNICAL_FILE": "Dashboard STEP 6: belongs ONLY to Packaging Configuration — never to a Product. Version when components change.",
            "DECLARATION_OF_CONFORMITY": "Dashboard STEP 5: variant from Commercial Scenario (Incoterms). TF stays on Packaging Configuration.",
        }
        ws.merge_cells("A2:H2")
        ws["A2"] = tips.get(internal_name, "Yellow = input · Blue = calculation · Green = output · Gray = protected")
        ws["A2"].font = FONT_MUTED
        ws["A2"].fill = FILL_DASH

        # Legend chips
        ws["A3"] = "INPUT"
        ws["A3"].fill = FILL_INPUT
        ws["B3"] = "CALC"
        ws["B3"].fill = FILL_CALC
        ws["C3"] = "OUTPUT"
        ws["C3"].fill = FILL_OUTPUT
        ws["D3"] = "PROTECTED"
        ws["D3"].fill = FILL_LOCKED
        for col in "ABCD":
            ws[f"{col}3"].font = FONT_BODY
            ws[f"{col}3"].alignment = ALIGN_CENTER
            ws[f"{col}3"].border = THIN

        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 32

    def _link_btn(self, ws, addr: str, label: str, link: str) -> None:
        cell = ws[addr]
        cell.value = label
        style_nav_button_cell(cell)
        cell.hyperlink = link

    # ------------------------------------------------------------------
    # Business logic panels (green outputs) — no schema change
    # ------------------------------------------------------------------

    def _add_business_panels(self) -> None:
        self._panel_component()
        self._panel_product()
        self._panel_packaging()
        self._panel_shipment()
        self._panel_statement()
        self._panel_technical_file()
        self._panel_declaration()

    def _panel_anchor_col(self, ws) -> int:
        """Place panels to the right of the widest table on the sheet."""
        max_col = 1
        for table in ws.tables.values():
            end = table.ref.split(":")[1]
            col = "".join(ch for ch in end if ch.isalpha())
            # column letters to index
            idx = 0
            for ch in col:
                idx = idx * 26 + (ord(ch.upper()) - 64)
            max_col = max(max_col, idx)
        return max_col + 2

    def _write_panel(self, ws, start_col: int, start_row: int, title: str, rows: list[tuple[str, str]]) -> None:
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = FONT_SUBTITLE
        title_cell.fill = FILL_PRIMARY_MID
        title_cell.font = FONT_HEADER
        ws.cell(row=start_row, column=start_col + 1).fill = FILL_PRIMARY_MID

        for i, (label, formula) in enumerate(rows, start=1):
            lab = ws.cell(row=start_row + i, column=start_col, value=label)
            lab.fill = FILL_LOCKED
            lab.font = FONT_BODY
            lab.border = THIN
            val = ws.cell(row=start_row + i, column=start_col + 1, value=formula)
            val.fill = FILL_OUTPUT
            val.font = FONT_KPI if i <= 3 else FONT_BODY
            val.border = THIN
            val.protection = Protection(locked=True)
        ws.column_dimensions[get_column_letter(start_col)].width = 28
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 22

    def _panel_component(self) -> None:
        ws = self.wb["COMPONENT"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "COMPONENT STATUS (auto)",
            [
                ("Total Components", "=COUNTA(T_COMPONENT[COMPONENT_ID])"),
                ("Missing Weight", '=COUNTIF(T_COMPONENT[WEIGHT_G],"<=0")+COUNTBLANK(T_COMPONENT[WEIGHT_G])'),
                ("Missing Supplier", "=COUNTBLANK(T_COMPONENT[SUPPLIER_ID])"),
                ("With Evidence Docs", "=SUMPRODUCT((T_DOCUMENT_LINK[COMPONENT_ID]<>\"\")*1)"),
                ("Returnable / Pool", '=COUNTIF(T_COMPONENT[OWNERSHIP_TYPE_ID],">1")'),
                ("Active Hint", "Set STATUS_ID = ACTIVE (2)"),
            ],
        )
        # Guidance block
        g = c
        ws.cell(row=CHROME_ROWS + 9, column=g, value="INCI RULES").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 9, column=g).font = FONT_HEADER
        rules = [
            "1) Enter WEIGHT_G (grams) — only weight source",
            "2) Fill COMPONENT_MATERIAL shares (~100%)",
            "3) Supplier required for PPWR traceability",
            "4) Attach evidence via DOCUMENT_LINK",
            "5) Container materials = type DUNNAGE/LINER/…",
        ]
        for i, text in enumerate(rules):
            cell = ws.cell(row=CHROME_ROWS + 10 + i, column=g, value=text)
            cell.fill = FILL_CALC
            cell.font = FONT_MUTED

    def _panel_product(self) -> None:
        ws = self.wb["PRODUCT"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "ERP PRODUCT REFERENCE (auto)",
            [
                ("Total Products", "=COUNTA(T_PRODUCT[PRODUCT_ID])"),
                ("Starter Batteries", '=COUNTIF(T_PRODUCT[PRODUCT_CATEGORY_ID],1)'),
                ("Industrial Batteries", '=COUNTIF(T_PRODUCT[PRODUCT_CATEGORY_ID],2)'),
                ("Linked Scenarios", "=COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])"),
                ("Products w/ Scenario", "=SUMPRODUCT((T_COMMERCIAL_SCENARIO[PRODUCT_ID]<>\"\")*1)"),
            ],
        )
        ws.cell(row=CHROME_ROWS + 8, column=c, value="NOT THE PRIMARY OBJECT").font = FONT_HEADER
        ws.cell(row=CHROME_ROWS + 8, column=c).fill = FILL_PRIMARY
        for i, text in enumerate(
            [
                "PPWR primary object = Packaging Configuration",
                "Product only links to configurations (via scenario)",
                "Never create a Technical File for a Product",
                "Daily work starts on Dashboard STEP 1 (Packaging Config)",
                "Scenario chooses DoC variant only — never edits TF",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 9 + i, column=c, value=text)
            cell.fill = FILL_CALC
            cell.font = FONT_MUTED

    def _panel_packaging(self) -> None:
        ws = self.wb["PACKAGING_CONFIGURATION"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "PRIMARY OBJECT · RECIPE TOTALS",
            [
                ("Configurations", "=COUNTA(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID])"),
                ("Recipe Lines", "=COUNTA(T_PACKAGING_CONFIGURATION_LINE[PACKAGING_CONFIGURATION_LINE_ID])"),
                ("Technical Files", "=COUNTA(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID])"),
                ("Total Line Weight g", "=IFERROR(SUM(T_ENG_PACKAGING_WEIGHT[LINE_WEIGHT_G]),0)"),
                ("Plastic kg", "=IFERROR(INDEX(T_ENG_PLASTIC_SUMMARY[TOTAL_MATERIAL_WEIGHT_KG],1),0)"),
                ("Paper kg", "=IFERROR(INDEX(T_ENG_PAPER_SUMMARY[TOTAL_MATERIAL_WEIGHT_KG],1),0)"),
                ("Wood kg", "=IFERROR(INDEX(T_ENG_WOOD_SUMMARY[TOTAL_MATERIAL_WEIGHT_KG],1),0)"),
            ],
        )
        ws.cell(row=CHROME_ROWS + 10, column=c, value="OWNERSHIP RULES").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 10, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "1 Technical File ↔ 1 Packaging Configuration",
                "Commercial Scenario does NOT edit this recipe/TF",
                "Starter: pallet-based recipe · Industrial: 1 batt/pallet",
                "Shipments pin this configuration revision",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 11 + i, column=c, value=text)
            cell.fill = FILL_CALC
            cell.font = FONT_MUTED

    def _panel_shipment(self) -> None:
        ws = self.wb["SHIPMENT"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "SHIPMENT AUTO CALC",
            [
                ("Shipments", "=COUNTA(T_SHIPMENT[SHIPMENT_ID])"),
                ("Latest Ship Date", "=IFERROR(MAX(T_SHIPMENT[SHIP_DATE]),\"-\")"),
                ("Total Pack Weight kg", "=IFERROR(SUM(T_ENG_SHIPMENT_WEIGHT[SHIPMENT_PACKAGING_WEIGHT_KG]),0)"),
                ("Freeze Lines", "=COUNTA(T_SHIPMENT_LINE[SHIPMENT_LINE_ID])"),
                ("Scenarios Ready", "=COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])"),
                ("Tech Files on Configs", "=COUNTA(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID])"),
                ("DoC Variants Ready", "=COUNTA(T_DECLARATION_OF_CONFORMITY[DECLARATION_OF_CONFORMITY_ID])"),
            ],
        )
        ws.cell(row=CHROME_ROWS + 10, column=c, value="MINIMUM INPUT (PPWR)").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 10, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "Customer / Country / Incoterms ← Scenario",
                "Packaging Configuration",
                "Quantity",
                "Lot Number → EXTERNAL_REF",
                "Paste draft from Dashboard STEP 3",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 11 + i, column=c, value=text)
            cell.fill = FILL_INPUT
            cell.font = FONT_MUTED

        ws.cell(row=CHROME_ROWS + 17, column=c, value="AUTO DETERMINED").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 17, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "Technical File ← Packaging Configuration",
                "DoC variant ← Commercial Scenario",
                "Statement ← shipment freeze",
                "Required documents ← package checklist",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 18 + i, column=c, value=text)
            cell.fill = FILL_OUTPUT
            cell.font = FONT_MUTED

    def _panel_statement(self) -> None:
        ws = self.wb["STATEMENT"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "STATEMENT PACK (auto)",
            [
                ("Statements", "=COUNTA(T_STATEMENT[STATEMENT_ID])"),
                ("Statement Lines", "=COUNTA(T_STATEMENT_LINE[STATEMENT_LINE_ID])"),
                ("Linked Shipments", "=COUNTA(T_STATEMENT_SHIPMENT[SHIPMENT_ID])"),
                ("Variance Rows", '=COUNTIF(T_ENG_STATEMENT[RECONCILE_STATUS],"VARIANCE")'),
                ("OK Rows", '=COUNTIF(T_ENG_STATEMENT[RECONCILE_STATUS],"OK")'),
            ],
        )
        ws.cell(row=CHROME_ROWS + 8, column=c, value="AUTO CONTENT").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 8, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "Composition ← SHIPMENT_LINE freeze",
                "Material breakdown ← ENG_STATEMENT",
                "Weights / component list ← engines",
                "Supplier traceability ← COMPONENT.SUPPLIER_ID",
                "Customer/Country/Incoterms ← scenario/shipment",
                "Document refs ← DOCUMENT_LINK",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 9 + i, column=c, value=text)
            cell.fill = FILL_OUTPUT
            cell.font = FONT_MUTED

    def _panel_technical_file(self) -> None:
        ws = self.wb["TECHNICAL_FILE"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "TECH FILE ASSEMBLY (auto)",
            [
                ("Technical Files", "=COUNTA(T_TECHNICAL_FILE[TECHNICAL_FILE_ID])"),
                ("On Packaging Config", "=COUNTA(T_TECHNICAL_FILE[PACKAGING_CONFIGURATION_ID])"),
                ("Complete", '=COUNTIF(T_ENG_TECHNICAL_FILE[ENGINE_STATUS],"COMPLETE")'),
                ("Incomplete", '=COUNTIF(T_ENG_TECHNICAL_FILE[ENGINE_STATUS],"INCOMPLETE")'),
                ("Linked Documents", "=COUNTA(T_DOCUMENT_LINK[DOCUMENT_LINK_ID])"),
            ],
        )
        ws.cell(row=CHROME_ROWS + 8, column=c, value="OWNERSHIP").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 8, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "NEVER create TF for a Product",
                "PACKAGING_CONFIGURATION_ID only",
                "Leave COMPONENT_ID / TRANSPORT blank",
                "Component change → increase REVISION_NO",
                "Commercial Scenario NEVER changes this file",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 9 + i, column=c, value=text)
            cell.fill = FILL_CALC
            cell.font = FONT_MUTED

        ws.cell(row=CHROME_ROWS + 15, column=c, value="DOSSIER CONTENT").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 15, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "Config description · drawings · photos (DOCUMENT_LINK)",
                "Component list · materials · weights (engines)",
                "Supplier + evidence matrices (DOCUMENT_LINK)",
                "Recyclability / minimisation / empty space / PFAS / labelling",
                "Revision history (REVISION_NO + dates)",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 16 + i, column=c, value=text)
            cell.fill = FILL_OUTPUT
            cell.font = FONT_MUTED

    def _panel_declaration(self) -> None:
        ws = self.wb["DECLARATION_OF_CONFORMITY"]
        c = self._panel_anchor_col(ws)
        self._write_panel(
            ws,
            c,
            CHROME_ROWS + 1,
            "DoC EXPORT READY (auto)",
            [
                ("Declarations", "=COUNTA(T_DECLARATION_OF_CONFORMITY[DECLARATION_OF_CONFORMITY_ID])"),
                ("Complete", '=COUNTIF(T_ENG_DECLARATION[ENGINE_STATUS],"COMPLETE")'),
                ("Incomplete", '=COUNTIF(T_ENG_DECLARATION[ENGINE_STATUS],"INCOMPLETE")'),
                ("With Documents", "=COUNTIF(T_DOCUMENT_LINK[DECLARATION_OF_CONFORMITY_ID],\">0\")"),
                ("Scenarios (variants)", "=COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])"),
            ],
        )
        ws.cell(row=CHROME_ROWS + 8, column=c, value="VARIANT RULE").fill = FILL_PRIMARY
        ws.cell(row=CHROME_ROWS + 8, column=c).font = FONT_HEADER
        for i, text in enumerate(
            [
                "DoC variant ← Commercial Scenario",
                "Technical File ← Packaging Configuration",
                "Scenario attaches DoC — does not rewrite TF",
                "Link TECHNICAL_FILE_ID from the config's TF",
                "Legal entity + responsible person required",
                "Attach signed PDF in DOCUMENT_LIBRARY",
            ]
        ):
            cell = ws.cell(row=CHROME_ROWS + 9 + i, column=c, value=text)
            cell.fill = FILL_OUTPUT
            cell.font = FONT_MUTED

    # ------------------------------------------------------------------
    # Dashboard — PPWR command center (Phase 11)
    # ------------------------------------------------------------------

    def _rebuild_dashboard(self) -> None:
        from .dashboard_wizard import rebuild_ppwr_dashboard

        rebuild_ppwr_dashboard(self.wb, self.settings)

    # ------------------------------------------------------------------
    # Rename / hide / order
    # ------------------------------------------------------------------

    def _rename_visible_sheets(self) -> None:
        """Rename internal entity sheets to user-facing titles (tables unchanged).

        Excel sheet titles are case-insensitive, so SHIPMENT→Shipment must use a
        two-step rename via a temporary unique name.
        """
        # Remove accidental empty *1/*2 leftovers
        for name in list(self.wb.sheetnames):
            if name.endswith("1") or name.endswith("2"):
                ws = self.wb[name]
                if not ws.tables:
                    del self.wb[name]

        for internal, title in VISIBLE_SHEET_TITLES.items():
            if internal not in self.wb.sheetnames:
                # Recover Shipment1 / Dashboard1 style leftovers that hold real tables
                for name in list(self.wb.sheetnames):
                    if name.lower().startswith(title.lower()) and name != title:
                        tmp = f"_TMP_{internal}"
                        self.wb[name].title = tmp
                        self.wb[tmp].title = title
                        break
                continue
            if internal == title:
                continue
            tmp = f"_TMP_{internal}"
            # ensure temp name free
            n = 0
            while tmp in self.wb.sheetnames:
                n += 1
                tmp = f"_TMP_{internal}_{n}"
            self.wb[internal].title = tmp
            # if destination exists empty, drop it
            if title in self.wb.sheetnames and not self.wb[title].tables:
                del self.wb[title]
            self.wb[tmp].title = title

    def _refresh_named_ranges_after_rename(self) -> None:
        """Update defined names that pointed at old sheet titles."""
        replacements = {
            f"'{old}'": f"'{new}'"
            for old, new in VISIBLE_SHEET_TITLES.items()
            if old != new
        }
        # Also map DASHBOARD search if still present
        replacements["'DASHBOARD'"] = "'Dashboard'"

        for name in list(self.wb.defined_names.keys()):
            defn = self.wb.defined_names[name]
            text = defn.attr_text
            new_text = text
            for old, new in replacements.items():
                if old in new_text:
                    new_text = new_text.replace(old, new)
            if new_text != text:
                # recreate
                del self.wb.defined_names[name]
                self.wb.defined_names.add(DefinedName(name=name, attr_text=new_text))

        # Search + wizard named ranges are (re)created in _rebuild_dashboard after rename.
        # Keep search pointer valid if dashboard already exists at this stage.
        if "NR_SEARCH_TERM" in self.wb.defined_names:
            del self.wb.defined_names["NR_SEARCH_TERM"]
        self.wb.defined_names.add(
            DefinedName(name="NR_SEARCH_TERM", attr_text="'Dashboard'!$C$40")
        )

        # Packaging line PK range now on Packaging Configuration sheet
        pk_name = "NR_PACKAGING_CONFIGURATION_LINE_ID"
        if pk_name in self.wb.defined_names:
            del self.wb.defined_names[pk_name]
        # Find column of PK on Packaging Configuration sheet
        ws = self.wb["Packaging Configuration"]
        if "T_PACKAGING_CONFIGURATION_LINE" in ws.tables:
            ref = ws.tables["T_PACKAGING_CONFIGURATION_LINE"].ref
            start = ref.split(":")[0]
            start_row = int("".join(ch for ch in start if ch.isdigit()))
            # PK is first column of line table (PACKAGING_CONFIGURATION_LINE_ID)
            self.wb.defined_names.add(
                DefinedName(
                    name=pk_name,
                    attr_text=f"'Packaging Configuration'!$A${start_row + 1}:$A$1048576",
                )
            )

    def _set_visibility_and_order(self) -> None:
        visible_titles = set(VISIBLE_SHEET_TITLES.values())
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            if name in visible_titles:
                ws.sheet_state = "visible"
            else:
                ws.sheet_state = "hidden"

        # Order visible sheets to front
        for title in reversed(list(VISIBLE_SHEET_TITLES.values())):
            if title in self.wb.sheetnames:
                idx = self.wb.sheetnames.index(title)
                if idx != 0:
                    self.wb.move_sheet(title, offset=-idx)

        # Ensure exact order
        for i, internal in enumerate(VISIBLE_ORDER):
            title = VISIBLE_SHEET_TITLES[internal]
            if title not in self.wb.sheetnames:
                continue
            current = self.wb.sheetnames.index(title)
            offset = i - current
            if offset:
                self.wb.move_sheet(title, offset=offset)

    def _reprotect_visible_sheets(self) -> None:
        for title in VISIBLE_SHEET_TITLES.values():
            if title not in self.wb.sheetnames:
                continue
            ws = self.wb[title]
            # Unlock yellow input cells inside tables (body), keep headers locked
            for table in ws.tables.values():
                ref = table.ref
                start, end = ref.split(":")
                start_row = int("".join(ch for ch in start if ch.isdigit()))
                end_row = int("".join(ch for ch in end if ch.isdigit()))
                end_col_letters = "".join(ch for ch in end if ch.isalpha())
                end_col = 0
                for ch in end_col_letters:
                    end_col = end_col * 26 + (ord(ch.upper()) - 64)
                # header locked
                for col in range(1, end_col + 1):
                    ws.cell(row=start_row, column=col).protection = Protection(locked=True)
                # stub table skip unlock
                if table.name == "T_PKG_LINE_STUB":
                    continue
                for row in range(start_row + 1, end_row + 1):
                    for col in range(1, end_col + 1):
                        cell = ws.cell(row=row, column=col)
                        # output/calc panels elsewhere; table body = input
                        if cell.fill.fgColor and getattr(cell.fill.fgColor, "rgb", None) in {
                            "00C8E6C9",
                            "00D6EAF8",
                            "00E0E0E0",
                        }:
                            cell.protection = Protection(locked=True)
                        else:
                            cell.fill = FILL_INPUT
                            cell.protection = Protection(locked=False)

            if title == "Dashboard":
                for addr in ("D8", "D9", "D10", "D11", "D12", "C40"):
                    ws[addr].protection = Protection(locked=False)
                    ws[addr].fill = FILL_INPUT

            ws.protection.sheet = True
            ws.protection.password = self.password
            ws.protection.autoFilter = True
            ws.protection.sort = True
