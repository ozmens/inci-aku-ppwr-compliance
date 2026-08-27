"""Non-database UI / QA sheets for Phase E workbook platform."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from models.registry import SchemaRegistry
from models.table_definition import TableDefinition

from .styles import (
    BODY_TEXT,
    HEADER_FILL,
    HEADER_FONT,
    LIGHT_BLUE_FILL,
    NAVY,
    UI_FONT,
    UI_SECTION_FONT,
    UI_TITLE_FONT,
    WHITE,
)

UI_SHEETS = (
    "00_README",
    "01_DASHBOARD",
    "02_RELEASE_CONTROL",
    "03_DATA_DICTIONARY",
    "04_IMPORT_GUIDE",
)

QA_WEIGHT_SHEET = "ZZ_QA_WEIGHT_FIXTURE"

# Weight fixture matching tests/test_weights.py (500*1 + 50*6 = 800)
WEIGHT_FIXTURE_LINES = (
    (1, "C1", "4000782", "Carton", 1, 500.0),
    (2, "C2", "CORNER", "Corner", 6, 50.0),
)


def create_ui_sheets(wb: Workbook, registry: SchemaRegistry) -> None:
    write_readme(wb.create_sheet("00_README", 0))
    write_dashboard(wb.create_sheet("01_DASHBOARD", 1), registry)
    write_release_control(wb.create_sheet("02_RELEASE_CONTROL", 2))
    write_data_dictionary(wb.create_sheet("03_DATA_DICTIONARY", 3), registry)
    write_import_guide(wb.create_sheet("04_IMPORT_GUIDE", 4))
    write_weight_fixture(wb.create_sheet(QA_WEIGHT_SHEET))


def write_readme(ws: Worksheet) -> None:
    _title(ws, "A1", "İnci Akü PPWR PIMS — Workbook Platform (Rev00)")
    lines = [
        "",
        "TR | EN",
        "Amaç / Purpose",
        "Bu çalışma kitabı, İnci Akü PPWR Ambalaj Bilgi Yönetim Sistemi’nin (PIMS) Wave-1 operasyonel ilişkisel veritabanıdır.",
        "This workbook is the Wave-1 operational relational database for the İnci Akü PPWR Packaging Information Management System (PIMS).",
        "",
        "Kaynak Gerçeği / Source of Truth",
        "Şema otoritesi: FINAL_DATABASE.md Schema 1.0.0 (FROZEN). Her ilişkisel tablo = bir sayfa. UI sayfaları (00–04) kaynak değildir.",
        "Schema authority: FINAL_DATABASE.md Schema 1.0.0 (FROZEN). One relational table = one sheet. UI sheets (00–04) are never source of truth.",
        "",
        "Sayfa Kategorileri / Sheet Categories",
        "00–04: Arayüz / rehber. LKP_*: kontrollü lookup. SYSTEM: meta. Diğerleri: master / yapılandırma / sevkiyat / belge metadata.",
        "00–04: UI / guides. LKP_*: controlled lookups. SYSTEM: metadata. Others: masters / configs / shipments / document metadata.",
        "",
        "Golden Variant",
        "CONFIG_GROUP_CODE = Packaging Set Code (ör. ST-051-STD-01). Variant numarası kod son ekidir (-01). Aile kodu soneksiz kısımdır.",
        "Variant Basis TR/EN → DESCRIPTION alanı (yapı: TR: … | EN: …) — Schema 1.0.0’da ayrı kolon yoktur; normalize tablolar korunur.",
        "PACKAGING_CONFIGURATION_ID = Final Configuration technical PK. Lineage: SUPERSEDES_ID + EXTERNAL_REF alanları (ilgili varlıklarda).",
        "",
        "Exact BOM",
        "PACKAGING_CONFIGURATION → PACKAGING_CONFIGURATION_LINE (COMPONENT + QUANTITY + LINE_ROLE). Tare = Σ(qty × COMPONENT.WEIGHT_G). Python WeightService otoriterdir.",
        "",
        "Ağırlık Sahipliği / Weight Ownership",
        "COMPONENT.WEIGHT_G = master. Configuration tare = derived. SHIPMENT_LINE / STATEMENT_LINE = transactional snapshot (dondurulmuş kopya).",
        "",
        "Ürün → Konfigürasyon / Product Mapping",
        "PRODUCT → COMMERCIAL_SCENARIO → TRANSPORT_CONFIGURATION → PACKAGING_CONFIGURATION. Ayrı Product Map tablosu yoktur (Schema 1.0.0).",
        "",
        "Sevkiyat Anlık Görüntüsü / Shipment Snapshot",
        "SHIPMENT / SHIPMENT_LINE onay anındaki bileşen/malzeme/ağırlık anlığını saklar. Lot → SHIPMENT.EXTERNAL_REF.",
        "",
        "Belge Kontrolü / Document Control",
        "TECHNICAL_FILE, DECLARATION_OF_CONFORMITY, DOCUMENT_LIBRARY, DOCUMENT_LINK, STATEMENT* metadata hazırdır. Word/PDF üretimi Phase E’de KAPALI.",
        "",
        "Revizyon / Revision",
        "REVISION_NO + SUPERSEDES_ID + STATUS_ID. Workbook: Rev00 platform. Production data: Phase F.",
        "",
        "Veri Giriş Kuralları / Data Entry",
        "1) Lookup kodlarını kullanın. 2) PK boş/çift olmasın. 3) CREATED_AT / UPDATED_AT sistem alanlarını elle değiştirmeyin.",
        "4) Derived tare’yi elle yazmayın. 5) ID’leri metin olarak tutun (bilimsel gösterim yok).",
        "",
        "Koruma / Protection",
        "Phase E: parola kilidi yok (istemci kullanımını engellemez). Sistem/hesaplanan alanlar dokunulmaz kabul edilir; Python QA doğrular.",
        "",
        "ERP / SQL / Power BI",
        "Tablo adları ve kolonlar Schema 1.0.0 ile birebir. Gelecek SQL migrasyonu ve Power BI aynı ilişkisel modeli kullanır.",
        "",
        "CIP / CIF",
        "CIF ve CIP ayrı Incoterms 2020 kodlarıdır; ikisi de LKP_INCOTERM’de aktiftir. Alias değildir.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = UI_FONT
        if text in {
            "TR | EN",
            "Amaç / Purpose",
            "Kaynak Gerçeği / Source of Truth",
            "Sayfa Kategorileri / Sheet Categories",
            "Golden Variant",
            "Exact BOM",
            "Ağırlık Sahipliği / Weight Ownership",
            "Ürün → Konfigürasyon / Product Mapping",
            "Sevkiyat Anlık Görüntüsü / Shipment Snapshot",
            "Belge Kontrolü / Document Control",
            "Revizyon / Revision",
            "Veri Giriş Kuralları / Data Entry",
            "Koruma / Protection",
            "ERP / SQL / Power BI",
            "CIP / CIF",
        }:
            cell.font = UI_SECTION_FONT
    ws.column_dimensions["A"].width = 120
    ws.sheet_view.showGridLines = False


def write_dashboard(ws: Worksheet, registry: SchemaRegistry) -> None:
    _title(ws, "A1", "01_DASHBOARD — Operational Shell (NOT source of truth)")
    ws["A2"] = "Phase E: production master/transactional data = NOT LOADED"
    ws["A2"].font = UI_FONT

    headers = ["KPI", "Count", "Status", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    kpis = [
        ("Components", "COMPONENT", "NOT LOADED"),
        ("Products", "PRODUCT", "NOT LOADED"),
        ("Final Packaging Configurations", "PACKAGING_CONFIGURATION", "NOT LOADED"),
        ("Packaging BOM lines", "PACKAGING_CONFIGURATION_LINE", "NOT LOADED"),
        ("Transport Configurations", "TRANSPORT_CONFIGURATION", "NOT LOADED"),
        ("Commercial Scenarios", "COMMERCIAL_SCENARIO", "NOT LOADED"),
        ("Shipments", "SHIPMENT", "NOT LOADED"),
        ("Technical Files", "TECHNICAL_FILE", "NOT LOADED"),
        ("DoCs", "DECLARATION_OF_CONFORMITY", "NOT LOADED"),
        ("Statements", "STATEMENT", "NOT LOADED"),
        ("Evidence / Document Library", "DOCUMENT_LIBRARY", "NOT LOADED"),
        ("Validation errors (platform)", None, "SEE 02_RELEASE_CONTROL"),
    ]

    for i, (label, table, status) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=label).font = UI_FONT
        if table:
            # Count non-blank PK cells minus empty template placeholder noise via COUNTA
            pk = registry.get(table).primary_key
            formula = f'=IF(COUNTA({table}!A:A)<=1,0,COUNTA({table}!A:A)-1)'
            # Safer: use structured table if exists — COUNTA on PK column range
            col_idx = list(registry.get(table).column_names).index(pk) + 1
            col = get_column_letter(col_idx)
            formula = f'=COUNTA(\'{table}\'!{col}2:{col}5000)'
            ws.cell(row=i, column=2, value=formula).font = UI_FONT
        else:
            ws.cell(row=i, column=2, value=0).font = UI_FONT
        ws.cell(row=i, column=3, value=status).font = UI_FONT
        ws.cell(row=i, column=4, value="Platform shell — Phase F loads production").font = UI_FONT

    ws["A18"] = "Category breakdown (future)"
    ws["A18"].font = UI_SECTION_FONT
    for i, cat in enumerate(
        ["Starter", "Industrial", "Container", "Returnable / Pool (schema hooks only)"],
        start=19,
    ):
        ws.cell(row=i, column=1, value=cat).font = UI_FONT
        ws.cell(row=i, column=2, value=0).font = UI_FONT
        ws.cell(row=i, column=3, value="NOT LOADED").font = UI_FONT

    ws["A24"] = "Lookup seed count (controlled vocabulary only)"
    ws["A24"].font = UI_SECTION_FONT
    ws["A25"] = '=COUNTA(\'LKP_STATUS\'!A2:A5000)'
    ws["B25"] = "LKP_STATUS sample (others on LKP_* sheets)"
    ws["A25"].font = UI_FONT
    ws["B25"].font = UI_FONT

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 36
    ws.sheet_view.showGridLines = False


def write_release_control(ws: Worksheet) -> None:
    _title(ws, "A1", "02_RELEASE_CONTROL — Platform readiness (not packaging conformity)")
    ws["A2"] = (
        "Statuses refer to SYSTEM / WORKBOOK readiness. Production conformity is NOT assessed in Phase E."
    )
    ws["A2"].font = UI_FONT

    headers = [
        "Control Area",
        "Rule",
        "Error Count",
        "Warning Count",
        "Status",
        "Blocking?",
        "Responsible Role",
        "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    rows = [
        (
            "Schema",
            "FINAL_DATABASE.md 1.0.0 frozen; 43 table sheets present",
            0,
            0,
            "READY",
            "Yes",
            "Architecture",
            "Workbook schema: READY",
        ),
        (
            "Master data",
            "Production COMPONENT / PRODUCT / CONFIG not loaded",
            0,
            0,
            "NOT LOADED",
            "Yes",
            "Packaging Engineering",
            "Production master data: NOT LOADED",
        ),
        (
            "BOM",
            "PACKAGING_CONFIGURATION_LINE empty until Phase F",
            0,
            0,
            "NOT LOADED",
            "Yes",
            "Packaging Engineering",
            "Production configuration data: NOT LOADED",
        ),
        (
            "Product mapping",
            "COMMERCIAL_SCENARIO links product→transport→packaging",
            0,
            0,
            "NOT LOADED",
            "Yes",
            "Sales / Engineering",
            "Platform ready; data Phase F",
        ),
        (
            "Evidence",
            "DOCUMENT_LIBRARY / LINK architecture present",
            0,
            0,
            "PLATFORM READY",
            "No",
            "Compliance",
            "No production evidence records",
        ),
        (
            "Drawing / photo",
            "Document types DRAWING/PHOTO/ARTWORK seeded",
            0,
            0,
            "PLATFORM READY",
            "No",
            "Engineering",
            "Files not imported in Phase E",
        ),
        (
            "Document generation",
            "Word/PDF builders exist as stubs only",
            0,
            0,
            "DISABLED IN PHASE E",
            "Yes",
            "IT / Compliance",
            "Document generation: DISABLED IN PHASE E",
        ),
        (
            "Lookup vocabulary",
            "LKP_* seeded; CIP and CIF both present as distinct Incoterms",
            0,
            0,
            "READY",
            "Yes",
            "MDM",
            "Controlled seed only",
        ),
        (
            "PK integrity framework",
            "PK highlight + blank/duplicate checks via Python QA / future CF",
            0,
            0,
            "FRAMEWORK READY",
            "Yes",
            "IT",
            "Empty masters intentionally blank",
        ),
        (
            "FK integrity framework",
            "Lookup FK dropdowns; high-volume FK via Python QA",
            0,
            0,
            "FRAMEWORK READY",
            "Yes",
            "IT",
            "Relational FK defined by schema",
        ),
        (
            "Weight derivation",
            "Python WeightService authoritative; ZZ_QA_WEIGHT_FIXTURE visibility",
            0,
            0,
            "READY",
            "No",
            "Engineering",
            "Fixture total must = 800 g",
        ),
        (
            "Article 5",
            "Assessment fields available on TF; no auto ≤100 mg/kg",
            0,
            0,
            "NOT ASSESSED",
            "No",
            "Compliance",
            "No production assessment records",
        ),
    ]
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = UI_FONT
            if c_idx == 5 and val in {"READY", "PLATFORM READY", "FRAMEWORK READY"}:
                cell.fill = LIGHT_BLUE_FILL

    last = 4 + len(rows)
    ws.add_table(
        Table(displayName="T_UI_RELEASE_CONTROL", ref=f"A4:H{last}")
    )
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22 if col > 2 else 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["H"].width = 40
    ws.sheet_view.showGridLines = False


def write_data_dictionary(ws: Worksheet, registry: SchemaRegistry) -> None:
    _title(ws, "A1", "03_DATA_DICTIONARY — Generated from Schema 1.0.0 metadata")
    ws["A2"] = "Generated from Python TableDefinition registry — do not maintain as parallel SoT."
    ws["A2"].font = UI_FONT

    headers = [
        "Table",
        "Column",
        "Data Type",
        "Required",
        "PK",
        "FK",
        "FK Target",
        "Business Meaning",
        "Data Owner",
        "Update Frequency",
        "Validation Rule",
        "System / User Managed",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    r = 5
    for table in registry.tables:
        for col in table.columns:
            managed = (
                "SYSTEM"
                if col.name
                in {
                    "CREATED_AT",
                    "UPDATED_AT",
                    "CONFIRMED_AT",
                    "APPROVED_AT",
                    "GENERATED_AT",
                }
                else "USER"
            )
            rule = _validation_rule(col)
            values = [
                table.name,
                col.name,
                col.data_type,
                "Y" if col.required else "N",
                "Y" if col.is_pk else "N",
                "Y" if col.is_fk else "N",
                col.fk_table or "",
                col.description or table.business_purpose,
                table.data_owner,
                table.update_frequency,
                rule,
                managed,
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = UI_FONT
            r += 1

    last = r - 1
    if last >= 5:
        ws.add_table(Table(displayName="T_UI_DATA_DICTIONARY", ref=f"A4:L{last}"))
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["H"].width = 40
    ws.freeze_panes = "A5"


def write_import_guide(ws: Worksheet) -> None:
    _title(ws, "A1", "04_IMPORT_GUIDE — Phase F migration order (do not import yet)")
    lines = [
        "",
        "Phase E does NOT import production data. Follow this order in Phase F:",
        "",
        "1. Lookup tables (LKP_*) — already seeded in Phase E; extend codes if needed",
        "2. Legal entities / external IDs (LEGAL_ENTITY, PERSON, SUPPLIER, CUSTOMER, PLANT)",
        "3. Components (COMPONENT)",
        "4. Component materials (COMPONENT_MATERIAL) — shares ≈ 100%",
        "5. Products (PRODUCT)",
        "6. Packaging configurations (PACKAGING_CONFIGURATION) — Golden Variant CONFIG_GROUP_CODE",
        "7. Packaging configuration lines (PACKAGING_CONFIGURATION_LINE) — exact BOM",
        "8. Transport configurations (+ lines)",
        "9. Commercial scenarios (product → transport → packaging mapping)",
        "10. Document / evidence metadata (DOCUMENT_LIBRARY, DOCUMENT_LINK, TF, DoC)",
        "11. Shipment data if applicable (SHIPMENT, SHIPMENT_LINE, STATEMENT*)",
        "",
        "Source lineage requirements (preserve on every import row where schema supports):",
        "- Source system",
        "- Source configuration ID / EXTERNAL_REF",
        "- Source file / sheet / row reference (NOTES or EXTERNAL_REF convention)",
        "- Legacy identifier",
        "",
        "Do not invent columns. Use Schema 1.0.0 fields only.",
        "Python ExcelRepository must own table-oriented access — avoid ad-hoc worksheet edits.",
        "",
        "Weight: never store competing editable configuration totals; derive via WeightService.",
        "Article 5: do not auto-populate ≤100 mg/kg without evidence.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = UI_SECTION_FONT if text.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "10.", "11.")) or text.startswith("Source lineage") or text.startswith("Phase E") else UI_FONT
    ws.column_dimensions["A"].width = 110
    ws.sheet_view.showGridLines = False


def write_weight_fixture(ws: Worksheet) -> None:
    """Non-SoT QA sheet: Excel visibility of WeightService fixture (800 g)."""
    _title(ws, "A1", "ZZ_QA_WEIGHT_FIXTURE — visibility only (matches Python WeightService)")
    ws["A2"] = "NOT a relational source table. Line weight = qty × component WEIGHT_G."
    ws["A2"].font = UI_FONT

    headers = [
        "LINE_NO",
        "COMPONENT_ID",
        "COMPONENT_ERP_CODE",
        "COMPONENT_NAME",
        "QUANTITY",
        "WEIGHT_G",
        "LINE_WEIGHT_G",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, row in enumerate(WEIGHT_FIXTURE_LINES, start=5):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = UI_FONT
            if c in (2, 3):
                cell.number_format = "@"
                cell.value = str(v)
        # LINE_WEIGHT_G = QUANTITY * WEIGHT_G  (same as WeightService)
        formula_cell = ws.cell(row=i, column=7, value=f"=E{i}*F{i}")
        formula_cell.font = UI_FONT
        formula_cell.number_format = "0.00"

    ws["A7"] = "TOTAL_TARE_G"
    ws["A7"].font = UI_SECTION_FONT
    ws["G7"] = "=SUM(G5:G6)"
    ws["G7"].font = UI_SECTION_FONT
    ws["G7"].number_format = "0.00"
    ws["H7"] = "Must equal Python WeightService fixture = 800"
    ws["H7"].font = UI_FONT

    ws.add_table(Table(displayName="T_QA_WEIGHT_FIXTURE", ref="A4:G6"))
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _title(ws: Worksheet, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    cell.font = UI_TITLE_FONT


def _validation_rule(col) -> str:
    if col.is_pk:
        return "PK unique; not blank when row used"
    if col.is_fk and col.fk_table and col.fk_table.startswith("LKP_"):
        return f"Lookup list → {col.fk_table}"
    if col.is_fk:
        return f"FK → {col.fk_table} (Python QA; no huge dropdown)"
    if col.data_type == "BOOL":
        return "TRUE/FALSE"
    if col.data_type == "DATE":
        return "Date DD.MM.YYYY"
    if col.data_type == "DECIMAL":
        return "Decimal number"
    if col.data_type == "INT":
        return "Whole number / text-capable ID"
    return "Text"
