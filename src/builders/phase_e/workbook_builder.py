"""
Phase E WorkbookBuilder — production PIMS Excel platform.

Builds output/INCI_AKU_PPWR_PIMS_Rev00.xlsx from frozen Schema 1.0.0.
Does not import production data. Does not generate Word/PDF.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from models.registry import SchemaRegistry

from .db_sheets import (
    restyle_seeded_body,
    write_database_sheet,
    write_sys_parameter,
    write_sys_workbook_info,
)
from .lookup_seed import apply_lookup_seeds, lookup_row_count
from .ui_sheets import QA_WEIGHT_SHEET, UI_SHEETS, create_ui_sheets
from .validations import apply_pk_named_ranges, apply_validations

PHASE_E_OUTPUT_FILENAME = "INCI_AKU_PPWR_PIMS_Rev00.xlsx"


class PhaseEWorkbookBuilder:
    """Orchestrates schema sheets, lookup seeds, UI, validations, QA metadata."""

    def __init__(self, registry: SchemaRegistry, settings: Any) -> None:
        self.registry = registry
        self.settings = settings
        self.stats: dict[str, Any] = {}

    def output_path(self) -> Path:
        return self.settings.OUTPUT_DIR / PHASE_E_OUTPUT_FILENAME

    def build(self) -> Path:
        self.settings.ensure_directories()
        output = self.output_path()

        wb = Workbook()
        default = wb.active
        wb.remove(default)

        # Database sheets first (will be reordered after UI insert)
        for table in self.registry.tables:
            ws = wb.create_sheet(title=table.name)
            if table.name == "SYS_WORKBOOK_INFO":
                write_sys_workbook_info(ws, self._sys_workbook_info_rows())
            elif table.name == "SYS_PARAMETER":
                write_sys_parameter(ws, self._sys_parameter_rows())
            else:
                write_database_sheet(ws, table)

        seeded = apply_lookup_seeds(wb)
        for name in seeded:
            restyle_seeded_body(wb[name], self.registry.get(name))

        create_ui_sheets(wb, self.registry)
        self._reorder_sheets(wb)

        named = apply_pk_named_ranges(wb, self.registry)
        validation_count = apply_validations(wb, self.registry)

        formula_count = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1

        table_count = sum(len(ws.tables) for ws in wb.worksheets)
        defined_name_count = len(list(wb.defined_names.keys()))

        wb.save(output)
        wb.close()

        self.stats = {
            "output_path": str(output),
            "schema_version": self.registry.schema_version,
            "database_table_count": self.registry.table_count,
            "database_sheet_count": self.registry.table_count,
            "ui_sheet_count": len(UI_SHEETS),
            "qa_sheet_count": 1,
            "lookup_row_count": lookup_row_count(),
            "seeded_lookup_sheets": seeded,
            "excel_table_count": table_count,
            "formula_count": formula_count,
            "defined_name_count": defined_name_count,
            "validation_count": validation_count,
            "named_ranges_created": len(named),
            "production_record_count": 0,
            "word_generation": "NOT RUN",
            "cip_cif_decision": (
                "CIF and CIP are distinct Incoterms 2020 codes; both seeded in "
                "LKP_INCOTERM. Neither is an alias. Schema 1.0.0 unchanged."
            ),
        }
        return output

    def _reorder_sheets(self, wb: Workbook) -> None:
        """UI sheets → database tables (registry order) → QA fixture."""
        desired = list(UI_SHEETS) + self.registry.names() + [QA_WEIGHT_SHEET]
        for idx, name in enumerate(desired):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    def _sys_workbook_info_rows(self) -> list[tuple]:
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        return [
            ("SCHEMA_VERSION", self.registry.schema_version, generated),
            ("WORKBOOK_NAME", PHASE_E_OUTPUT_FILENAME, generated),
            ("PROJECT_NAME", "INCI_AKU_PPWR_PIMS", generated),
            ("REVISION", "Rev00", generated),
            ("PHASE", "E — Workbook Platform", generated),
            ("PRODUCTION_DATA", "NOT LOADED", generated),
            ("WORD_GENERATION", "DISABLED", generated),
            ("GENERATED_AT_UTC", generated, generated),
            ("EXCEL_ENGINE", "openpyxl", generated),
            (
                "CIP_CIF_NOTE",
                "CIF and CIP are distinct Incoterms; both active in LKP_INCOTERM",
                generated,
            ),
        ]

    def _sys_parameter_rows(self) -> list[tuple]:
        return [
            (1, "WEIGHT_UOM", self.settings.WEIGHT_UOM, "Mass unit for COMPONENT.WEIGHT_G"),
            (
                2,
                "TRANSPORT_ALLOCATION_METHOD",
                self.settings.TRANSPORT_ALLOCATION_METHOD,
                "Transport-line weight allocation method",
            ),
            (
                3,
                "MATERIAL_SHARE_TOLERANCE_PCT",
                str(self.settings.MATERIAL_SHARE_TOLERANCE_PCT),
                "Allowed deviation from 100% material shares",
            ),
            (
                4,
                "ENABLE_WORD_GENERATION",
                "FALSE",
                "Document generation disabled in Phase E",
            ),
            (
                5,
                "PRODUCTION_DATA_LOADED",
                "FALSE",
                "Phase F loads production masters",
            ),
            (
                6,
                "ARTICLE5_BASIS_LABEL",
                "REV00 CURRENT EVIDENCE BASIS — ARTICLE 5 ASSESSMENT BASIS",
                "Controlled customer language — do not auto-fill numerical compliance",
            ),
        ]
