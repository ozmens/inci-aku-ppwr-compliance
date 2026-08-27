"""Database worksheet writer — one frozen table = one sheet."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from models.table_definition import ColumnDefinition, TableDefinition

from .styles import (
    BODY_ALIGNMENT,
    BODY_FONT,
    FK_FILL,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    PK_FILL,
    SYSTEM_FILL,
    SYSTEM_MANAGED_COLUMNS,
    THIN_BORDER,
    WHITE_FILL,
    is_identifier_column,
)

PLACEHOLDER_ROWS = 1


def excel_table_name(table_name: str) -> str:
    return f"T_{table_name}"


def write_database_sheet(ws: Worksheet, table: TableDefinition) -> Table:
    """Write header + empty body, Excel Table, styles, freeze, widths."""
    headers = list(table.column_names)
    if not headers:
        raise ValueError(f"{table.name} has no columns")

    for col_idx, col in enumerate(table.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx in range(2, 2 + PLACEHOLDER_ROWS):
        for col_idx, col in enumerate(table.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=None)
            _style_body_cell(cell, col)

    last_col = get_column_letter(len(headers))
    last_row = 1 + PLACEHOLDER_ROWS
    display = excel_table_name(table.name)
    excel_table = Table(displayName=display, ref=f"A1:{last_col}{last_row}")
    ws.add_table(excel_table)

    ws.freeze_panes = "A2"
    # Do NOT set ws.auto_filter when an Excel Table is present — duplicate AutoFilter
    # XML makes Microsoft Excel refuse to open the workbook.
    ws.row_dimensions[1].height = 28

    for col_idx, col in enumerate(table.columns, start=1):
        width = _column_width(col)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return excel_table


def restyle_seeded_body(ws: Worksheet, table: TableDefinition) -> None:
    """Re-apply body styles after lookup seed population."""
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col in enumerate(table.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            _style_body_cell(cell, col)
            if cell.value is not None and is_identifier_column(col.name, col.data_type):
                if not isinstance(cell.value, bool):
                    cell.value = str(cell.value)
                cell.number_format = "@"
    last_col = get_column_letter(len(table.columns))
    last_row = max(ws.max_row, 2)
    # Refresh table ref if present (table owns AutoFilter — do not set ws.auto_filter)
    tname = excel_table_name(table.name)
    if tname in ws.tables:
        ws.tables[tname].ref = f"A1:{last_col}{last_row}"


def write_sys_workbook_info(ws: Worksheet, rows: list[tuple]) -> Table:
    from models.tables import SYS_WORKBOOK_INFO

    return _write_populated(ws, SYS_WORKBOOK_INFO, rows)


def write_sys_parameter(ws: Worksheet, rows: list[tuple]) -> Table:
    from models.tables import SYS_PARAMETER

    return _write_populated(ws, SYS_PARAMETER, rows)


def _write_populated(
    ws: Worksheet, table: TableDefinition, rows: list[tuple]
) -> Table:
    for col_idx, col in enumerate(table.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    body_rows = rows if rows else [(None,) * len(table.columns)]
    for r_idx, row in enumerate(body_rows, start=2):
        for c_idx, col in enumerate(table.columns, start=1):
            value = row[c_idx - 1] if c_idx - 1 < len(row) else None
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            _style_body_cell(cell, col)
            if value is not None and is_identifier_column(col.name, col.data_type):
                if not isinstance(value, bool):
                    cell.value = str(value)
                cell.number_format = "@"

    last_col = get_column_letter(len(table.columns))
    last_row = 1 + len(body_rows)
    display = excel_table_name(table.name)
    excel_table = Table(displayName=display, ref=f"A1:{last_col}{last_row}")
    ws.add_table(excel_table)
    ws.freeze_panes = "A2"
    # Do NOT set ws.auto_filter when an Excel Table is present — duplicate AutoFilter
    # XML makes Microsoft Excel refuse to open the workbook.
    ws.row_dimensions[1].height = 28
    for col_idx, col in enumerate(table.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _column_width(col)
    return excel_table


def _style_body_cell(cell, col: ColumnDefinition) -> None:
    cell.font = BODY_FONT
    cell.alignment = BODY_ALIGNMENT
    cell.border = THIN_BORDER

    if col.is_pk:
        cell.fill = PK_FILL
    elif col.is_fk:
        cell.fill = FK_FILL
    elif col.name in SYSTEM_MANAGED_COLUMNS:
        cell.fill = SYSTEM_FILL
    else:
        cell.fill = WHITE_FILL

    if is_identifier_column(col.name, col.data_type):
        cell.number_format = "@"
    elif col.data_type == "DATE":
        cell.number_format = "DD.MM.YYYY"
    elif col.data_type == "DATETIME":
        cell.number_format = "DD.MM.YYYY HH:MM:SS"
    elif col.data_type == "DECIMAL":
        cell.number_format = "0.00"
    elif col.data_type == "INT" and not col.is_pk and not col.is_fk:
        cell.number_format = "0"
    elif col.data_type == "BOOL":
        cell.number_format = "@"


def _column_width(col: ColumnDefinition) -> float:
    name_len = len(col.name)
    if col.is_pk or col.is_fk:
        return max(14, min(22, name_len + 2))
    if col.data_type in {"TEXT"} and name_len > 18:
        return 28
    if col.name in {"NOTES", "DESCRIPTION", "CONFORMITY_STATEMENT"}:
        return 36
    return max(12, min(24, name_len + 2))
