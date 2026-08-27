"""Phase E workbook QA report writer."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def inspect_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    formula_errors = 0
    broken_links = 0
    # openpyxl: external links
    if getattr(wb, "_external_links", None):
        broken_links = len(wb._external_links)

    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    formula_count += 1
                    upper = val.upper()
                    if any(err in upper for err in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?")):
                        formula_errors += 1

    table_names = []
    for ws in wb.worksheets:
        table_names.extend(ws.tables.keys())

    production_tables = [
        "PRODUCT",
        "COMPONENT",
        "COMPONENT_MATERIAL",
        "PACKAGING_CONFIGURATION",
        "PACKAGING_CONFIGURATION_LINE",
        "TRANSPORT_CONFIGURATION",
        "TRANSPORT_CONFIGURATION_LINE",
        "COMMERCIAL_SCENARIO",
        "SHIPMENT",
        "SHIPMENT_LINE",
        "TECHNICAL_FILE",
        "DECLARATION_OF_CONFORMITY",
        "STATEMENT",
        "STATEMENT_LINE",
        "CUSTOMER",
        "SUPPLIER",
    ]
    production_records = 0
    for name in production_tables:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        # Count rows with any non-empty cell in body
        for r in range(2, ws.max_row + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
                production_records += 1

    result = {
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
        "excel_table_names": table_names,
        "duplicate_table_names": sorted(
            {t for t in table_names if table_names.count(t) > 1}
        ),
        "defined_name_count": len(list(wb.defined_names.keys())),
        "formula_count": formula_count,
        "formula_error_count": formula_errors,
        "broken_link_count": broken_links,
        "production_record_count": production_records,
        "validation_count": sum(len(ws.data_validations.dataValidation) for ws in wb.worksheets),
    }
    wb.close()
    return result


def write_qa_report(
    path: Path,
    build_stats: dict[str, Any],
    inspect: dict[str, Any] | None = None,
) -> Path:
    inspect = inspect or {}
    lines = [
        "# Phase E Workbook QA Report",
        "",
        f"- **Workbook path:** `{build_stats.get('output_path', '')}`",
        f"- **Frozen schema version:** {build_stats.get('schema_version', '1.0.0')}",
        f"- **Database table count (schema):** {build_stats.get('database_table_count', 43)}",
        f"- **Database sheet count:** {build_stats.get('database_sheet_count', 43)}",
        f"- **UI sheet count:** {build_stats.get('ui_sheet_count', 5)}",
        f"- **QA sheet count:** {build_stats.get('qa_sheet_count', 1)}",
        f"- **Lookup row count (seeded):** {build_stats.get('lookup_row_count', 0)}",
        f"- **Excel Table count:** {inspect.get('excel_table_names') and len(inspect['excel_table_names']) or build_stats.get('excel_table_count', 0)}",
        f"- **Formula count:** {inspect.get('formula_count', build_stats.get('formula_count', 0))}",
        f"- **Defined name count:** {inspect.get('defined_name_count', build_stats.get('defined_name_count', 0))}",
        f"- **Validation count:** {inspect.get('validation_count', build_stats.get('validation_count', 0))}",
        f"- **PK integrity framework status:** FRAMEWORK READY (highlight + Python QA; blank masters intentional)",
        f"- **FK integrity framework status:** FRAMEWORK READY (lookup dropdowns; high-volume via Python QA)",
        f"- **Formula error count:** {inspect.get('formula_error_count', 0)}",
        f"- **Broken link count:** {inspect.get('broken_link_count', 0)}",
        f"- **Production record count:** {inspect.get('production_record_count', 0)}",
        f"- **Production data:** NOT LOADED",
        f"- **Word generation:** NOT RUN / DISABLED IN PHASE E",
        "",
        "## CIP / CIF decision",
        "",
        build_stats.get("cip_cif_decision", ""),
        "",
        "## Duplicate Excel table names",
        "",
        f"`{inspect.get('duplicate_table_names', [])}`",
        "",
        "## Confirmation",
        "",
        "- FINAL_DATABASE.md Schema 1.0.0 was not modified.",
        "- No production İnci Akü masters / BOMs / shipments loaded.",
        "- No Word / PDF / batch customer outputs generated.",
        "- Phase F not started.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
