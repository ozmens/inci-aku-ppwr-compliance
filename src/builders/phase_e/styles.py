"""İnci Akü corporate-compatible restrained workbook styles (Phase E)."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FONT_NAME = "Tahoma"

NAVY = "1F4E79"
LIGHT_BLUE = "D6EAF8"
PK_YELLOW = "FFF2CC"
SYSTEM_GREY = "F2F2F2"
WHITE = "FFFFFF"
BODY_TEXT = "1A1A1A"
THIN_GREY = "B0B0B0"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
BODY_FONT = Font(name=FONT_NAME, size=9, color=BODY_TEXT)
UI_FONT = Font(name=FONT_NAME, size=10, color=BODY_TEXT)
UI_TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
UI_SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
PK_FILL = PatternFill("solid", fgColor=PK_YELLOW)
FK_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
SYSTEM_FILL = PatternFill("solid", fgColor=SYSTEM_GREY)
LIGHT_BLUE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)

HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin", color=THIN_GREY),
    right=Side(style="thin", color=THIN_GREY),
    top=Side(style="thin", color=THIN_GREY),
    bottom=Side(style="thin", color=THIN_GREY),
)

SYSTEM_MANAGED_COLUMNS = frozenset(
    {
        "CREATED_AT",
        "UPDATED_AT",
        "CONFIRMED_AT",
        "APPROVED_AT",
        "GENERATED_AT",
    }
)

# Columns that must remain text-capable (IDs / codes / external refs)
TEXT_ID_HINTS = ("_ID", "_CODE", "_REF", "EXTERNAL_REF", "INFO_KEY")


def is_identifier_column(name: str, data_type: str) -> bool:
    if name.endswith("_ID") or name.endswith("_CODE") or name == "EXTERNAL_REF":
        return True
    if name in {"INFO_KEY", "ISO2", "ISO3"}:
        return True
    if data_type == "TEXT" and any(h in name for h in ("_CODE", "_REF", "NUMBER")):
        return True
    return False
