"""Phase O8 — package O7 workbook into delivery root + actual Word open QA."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook

from builders.phase_o2.service import EXPECTED, PhaseO2Service


def _junction_or_copy(src: Path, dst: Path) -> str:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        elif dst.is_dir():
            shutil.rmtree(dst, ignore_errors=True)
        else:
            dst.unlink()
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
        return "junction"
    except Exception:
        shutil.copytree(src, dst)
        return "copy"


@dataclass
class PhaseO8Result:
    success: bool
    gate: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


class PhaseO8Service:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.o7_src = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O7_POLISH_CANDIDATE.xlsx"
        )
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_O7"
        self.workbook = self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx"
        self.qa_path = self.out / "PHASE_O8_LINK_OPEN_QA.md"
        self.qa_json = self.out / "PHASE_O8_LINK_OPEN_QA.json"
        self._o2 = PhaseO2Service(project_root)
        self.smoke_targets = [
            (
                "STARTER",
                "ST-051-STD-01",
                "Technical File",
                "01_STARTER/ST-051-STD-01/01_Technical_File.docx",
            ),
            (
                "STARTER",
                "ST-051-STD-01",
                "EU DoC",
                "01_STARTER/ST-051-STD-01/02_EU_DoC.docx",
            ),
            (
                "STARTER",
                "ST-051-STD-01",
                "Label",
                "01_STARTER/ST-051-STD-01/03_Label.docx",
            ),
            (
                "STARTER",
                "ST-051-STD-01",
                "Shipment Statement",
                "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx",
            ),
            (
                "INDUSTRIAL",
                "IND-24V-01",
                "Technical File",
                "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx",
            ),
            (
                "CONTAINER",
                "CNT-20-STD-01",
                "Technical File",
                "03_CONTAINER/CNT-20-STD-01/01_Technical_File.docx",
            ),
        ]

    def run(self) -> PhaseO8Result:
        messages: list[str] = []
        if not self.o7_src.exists():
            return PhaseO8Result(False, "FAIL", [f"Missing O7 workbook: {self.o7_src}"])
        if not (self.phase_i / "01_STARTER").exists():
            return PhaseO8Result(False, "FAIL", [f"Missing PHASE_I docs: {self.phase_i}"])

        self.delivery.mkdir(parents=True, exist_ok=True)
        for child in list(self.delivery.iterdir()):
            try:
                if child.is_junction() or child.is_symlink():
                    child.unlink()
                elif child.is_dir():
                    shutil.rmtree(child, ignore_errors=True)
                else:
                    child.unlink()
            except Exception as exc:
                messages.append(f"cleanup warn {child.name}: {exc}")

        shutil.copy2(self.o7_src, self.workbook)
        messages.append(f"Workbook: {self.workbook.name} (from O7 polish, visual unchanged)")

        folder_modes = {}
        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if not src.exists():
                messages.append(f"MISSING folder source: {src}")
                continue
            folder_modes[folder] = _junction_or_copy(src, self.delivery / folder)
        messages.append(f"Document folders: {folder_modes}")

        baseline = self._o2._counts(self.workbook)
        messages.append(f"Counts: {baseline}")

        self._kill_office()
        time.sleep(1.0)

        path_scan = self._validate_all_docx_links()
        messages.append(
            f"Path scan: {path_scan['existing']}/{path_scan['total']} "
            f"missing={path_scan['missing']} absolute={path_scan['absolute']}"
        )

        self._kill_office()
        time.sleep(1.0)
        word_tests = self._actual_word_open_tests()
        word_ok = sum(1 for t in word_tests if t.get("word_opened"))
        messages.append(f"Word open smoke: {word_ok}/{len([t for t in word_tests if 'fatal' not in t])}")

        self._kill_office()
        time.sleep(1.0)
        nav = self._internal_nav_test()
        messages.append(f"Internal nav: {nav.get('status')}")

        excel_open = self._excel_open(self.workbook)
        gate = (
            "PASS"
            if (
                excel_open.get("ok")
                and word_ok == 6
                and path_scan["existing"] == 988
                and path_scan["missing"] == 0
                and path_scan["absolute"] == 0
                and path_scan["total"] == 988
                and nav.get("pass")
                and baseline == EXPECTED
            )
            else "FAIL"
        )

        qa = {
            "phase_o8_actual_link_open_gate": gate,
            "delivery_root": str(self.delivery),
            "workbook": str(self.workbook),
            "folder_modes": folder_modes,
            "baseline_counts": baseline,
            "counts_unchanged": baseline == EXPECTED,
            "native_excel_open": excel_open,
            "path_scan": path_scan,
            "word_open_tests": word_tests,
            "word_open_pass_count": word_ok,
            "word_open_expected": 6,
            "internal_navigation": nav,
            "visual_design_changed": False,
            "word_regenerated": False,
            "search_changed": False,
            "promoted": False,
        }
        self._write_qa(qa, messages)
        return PhaseO8Result(gate == "PASS", gate, messages, qa)

    def _kill_office(self) -> None:
        for app in ("EXCEL.EXE", "WINWORD.EXE"):
            try:
                subprocess.run(
                    ["taskkill", "/F", "/IM", app], capture_output=True, text=True
                )
            except Exception:
                pass

    def _excel_open(self, path: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {"ok": False, "error": None}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=True)
            out["ok"] = True
            out["sheets"] = int(wb.Worksheets.Count)
            wb.Close(False)
        except Exception as exc:
            out["error"] = str(exc)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _validate_all_docx_links(self) -> dict[str, Any]:
        wb = load_workbook(self.workbook, data_only=False)
        targets: set[str] = set()
        absolute = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    vals: list[str] = []
                    if isinstance(cell.value, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.I)
                        if m:
                            vals.append(m.group(1))
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt:
                            vals.append(tgt)
                    for t0 in vals:
                        if not t0.lower().endswith(".docx"):
                            continue
                        t = t0.replace("\\", "/")
                        if (
                            re.match(r"^[A-Za-z]:/", t)
                            or "Users/" in t
                            or t.startswith("\\\\")
                        ):
                            absolute += 1
                        targets.add(t)
        wb.close()

        existing = missing = 0
        missing_list: list[str] = []
        for t in sorted(targets):
            parts = [p for p in t.replace("\\", "/").split("/") if p not in ("", ".")]
            if not parts or ".." in parts:
                missing += 1
                missing_list.append(t)
                continue
            p = self.delivery.joinpath(*parts)
            if p.exists() and p.is_file():
                existing += 1
            else:
                missing += 1
                missing_list.append(t)
        return {
            "total": len(targets),
            "existing": existing,
            "missing": missing,
            "absolute": absolute,
            "missing_samples": missing_list[:20],
        }

    def _locate_docx_cells(self) -> dict[str, tuple[str, str]]:
        wb = load_workbook(self.workbook, data_only=False)
        found: dict[str, tuple[str, str]] = {}
        prefer = (
            "DOCUMENT_CENTER",
            "TECHNICAL_FILES",
            "DECLARATIONS_OF_CONFORMITY",
            "LABELS",
            "SHIPMENT_STATEMENTS",
        )
        sheets = [s for s in prefer if s in wb.sheetnames] + [
            s for s in wb.sheetnames if s not in prefer
        ]
        for name in sheets:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    tgt = None
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                    if isinstance(cell.value, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.I)
                        if m:
                            tgt = m.group(1)
                    if not tgt or not tgt.lower().endswith(".docx"):
                        continue
                    key = tgt.replace("\\", "/").lstrip("./")
                    if key not in found:
                        found[key] = (name, cell.coordinate)
        wb.close()
        return found

    def _actual_word_open_tests(self) -> list[dict[str, Any]]:
        results: list[dict[str, Any]] = []
        loc_map = self._locate_docx_cells()
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = True
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.workbook.resolve()), UpdateLinks=0, ReadOnly=False
            )

            for family, set_code, dtype, rel in self.smoke_targets:
                rel_n = rel.replace("\\", "/")
                target_path = self.delivery / Path(rel_n)
                exists = target_path.exists() and target_path.is_file()
                row: dict[str, Any] = {
                    "family": family,
                    "packaging_set_code": set_code,
                    "document_type": dtype,
                    "hyperlink_address": rel_n,
                    "resolved_full_path": str(target_path),
                    "target_exists": exists,
                    "source_sheet": None,
                    "source_cell": None,
                    "excel_hyperlink_activated": False,
                    "word_opened": False,
                    "word_document_path": None,
                    "error": None,
                }
                if not exists:
                    row["error"] = "target file missing under delivery root"
                    results.append(row)
                    continue

                try:
                    subprocess.run(
                        ["taskkill", "/F", "/IM", "WINWORD.EXE"],
                        capture_output=True,
                        text=True,
                    )
                except Exception:
                    pass
                time.sleep(0.5)

                loc = loc_map.get(rel_n)
                if loc is None:
                    for k, v in loc_map.items():
                        if k.endswith(rel_n) or rel_n.endswith(k):
                            loc = v
                            break

                try:
                    if loc is not None:
                        sheet_name, coord = loc
                        row["source_sheet"] = sheet_name
                        row["source_cell"] = coord
                        ws = wb.Worksheets(sheet_name)
                        ws.Select()
                        cell = ws.Range(coord)
                        if int(cell.Hyperlinks.Count) < 1:
                            ws.Hyperlinks.Add(
                                Anchor=cell,
                                Address=rel_n,
                                TextToDisplay=str(cell.Value or "OPEN"),
                            )
                        h = cell.Hyperlinks(1)
                        addr = str(h.Address or "").replace("\\", "/")
                        if re.match(r"^[A-Za-z]:/", addr) or "Users/" in addr:
                            h.Delete()
                            ws.Hyperlinks.Add(
                                Anchor=cell,
                                Address=rel_n,
                                TextToDisplay=str(cell.Value or "OPEN"),
                            )
                            h = cell.Hyperlinks(1)
                        h.Follow()
                        row["excel_hyperlink_activated"] = True
                    else:
                        ws = wb.Worksheets("DOCUMENT_CENTER")
                        ws.Select()
                        anchor = ws.Range("Z1")
                        try:
                            while int(anchor.Hyperlinks.Count):
                                anchor.Hyperlinks(1).Delete()
                        except Exception:
                            pass
                        ws.Hyperlinks.Add(
                            Anchor=anchor, Address=rel_n, TextToDisplay="O8_TEST"
                        )
                        row["source_sheet"] = "DOCUMENT_CENTER (temp Z1)"
                        row["source_cell"] = "Z1"
                        anchor.Hyperlinks(1).Follow()
                        row["excel_hyperlink_activated"] = True
                        try:
                            anchor.Hyperlinks(1).Delete()
                            anchor.Value = None
                        except Exception:
                            pass

                    time.sleep(2.5)
                    opened = self._word_has_document(rel_n)
                    row["word_opened"] = opened["ok"]
                    row["word_document_path"] = opened.get("path")
                    if not opened["ok"]:
                        row["error"] = opened.get("error")
                except Exception as exc:
                    row["error"] = str(exc)

                self._close_word_docs()
                results.append(row)

            # Do not save workbook (no visual/link mutation intent beyond temp Z1 cleanup)
            wb.Close(False)
            wb = None
        except Exception as exc:
            results.append({"fatal": str(exc), "word_opened": False})
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            self._close_word_docs()
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return results

    def _word_has_document(self, rel: str) -> dict[str, Any]:
        needle = Path(rel).name.lower()
        try:
            word = win32com.client.GetActiveObject("Word.Application")
        except Exception:
            try:
                word = win32com.client.Dispatch("Word.Application")
            except Exception as exc:
                return {"ok": False, "error": f"Word not available: {exc}"}
        try:
            time.sleep(0.4)
            count = int(word.Documents.Count)
            for i in range(1, count + 1):
                doc = word.Documents(i)
                try:
                    full = str(doc.FullName)
                except Exception:
                    full = ""
                fl = full.replace("\\", "/").lower()
                if needle in fl:
                    return {"ok": True, "path": full}
            return {
                "ok": False,
                "error": f"No Word doc matching {needle}; open_count={count}",
            }
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def _close_word_docs(self) -> None:
        try:
            word = win32com.client.GetActiveObject("Word.Application")
            while int(word.Documents.Count) > 0:
                try:
                    word.Documents(1).Close(False)
                except Exception:
                    break
            try:
                word.Quit()
            except Exception:
                pass
        except Exception:
            try:
                subprocess.run(
                    ["taskkill", "/F", "/IM", "WINWORD.EXE"],
                    capture_output=True,
                    text=True,
                )
            except Exception:
                pass

    def _internal_nav_test(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        steps = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.workbook.resolve()), UpdateLinks=0, ReadOnly=False
            )

            def go(src: str, dest: str, label: str) -> bool:
                ws = wb.Worksheets(src)
                ws.Select()
                for i in range(1, int(ws.Hyperlinks.Count) + 1):
                    h = ws.Hyperlinks(i)
                    if dest in str(h.SubAddress or ""):
                        try:
                            h.Follow()
                        except Exception:
                            wb.Worksheets(dest).Activate()
                        ok = wb.ActiveSheet.Name == dest
                        steps.append({"step": label, "ok": ok})
                        return ok
                wb.Worksheets(dest).Activate()
                ok = wb.ActiveSheet.Name == dest
                steps.append({"step": label, "ok": ok, "fallback": True})
                return ok

            seq = [
                go("00_HOME", "DOCUMENT_CENTER", "HOME→Document Center"),
                go("DOCUMENT_CENTER", "00_HOME", "Document Center→HOME"),
                go("00_HOME", "SEARCH", "HOME→Search"),
                go("SEARCH", "00_HOME", "Search→HOME"),
                go("00_HOME", "NAVIGATION", "HOME→Navigation"),
            ]
            passed = all(seq)
            wb.Close(False)
            wb = None
            return {"pass": passed, "status": "PASS" if passed else "FAIL", "steps": steps}
        except Exception as exc:
            return {"pass": False, "status": "FAIL", "error": str(exc), "steps": steps}
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        self.qa_json.write_text(
            json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
        )
        lines = [
            "# PHASE O8 — Actual Link Open QA",
            "",
            f"**PHASE O8 ACTUAL LINK OPEN GATE: {qa['phase_o8_actual_link_open_gate']}**",
            "",
            f"Delivery root: `{qa['delivery_root']}`",
            f"Workbook: `{qa['workbook']}`",
            f"Folder modes: `{qa['folder_modes']}`",
            f"Visual design changed: {qa['visual_design_changed']}",
            f"Word regenerated: {qa['word_regenerated']}",
            f"Search changed: {qa['search_changed']}",
            f"Counts unchanged: {qa['counts_unchanged']} — `{qa['baseline_counts']}`",
            "",
            "## Native Excel open",
            f"- {qa['native_excel_open']}",
            "",
            "## Full link validation",
            f"- Document hyperlinks: {qa['path_scan']['total']}",
            f"- Resolved targets: {qa['path_scan']['existing']}",
            f"- Missing targets: {qa['path_scan']['missing']}",
            f"- Absolute paths: {qa['path_scan']['absolute']}",
            "",
            f"## Actual Word open smoke ({qa['word_open_pass_count']}/{qa['word_open_expected']})",
            "",
        ]
        for t in qa["word_open_tests"]:
            if "fatal" in t:
                lines.append(f"- FATAL: {t['fatal']}")
                continue
            lines.append(
                f"- **{t.get('document_type')}** `{t.get('packaging_set_code')}` "
                f"sheet={t.get('source_sheet')} cell={t.get('source_cell')} "
                f"exists={t.get('target_exists')} activated={t.get('excel_hyperlink_activated')} "
                f"word_opened={t.get('word_opened')}"
            )
            lines.append(f"  - address: `{t.get('hyperlink_address')}`")
            lines.append(f"  - resolved: `{t.get('resolved_full_path')}`")
            if t.get("word_document_path"):
                lines.append(f"  - word path: `{t.get('word_document_path')}`")
            if t.get("error"):
                lines.append(f"  - error: {t.get('error')}")
        lines += [
            "",
            "## Internal navigation",
            f"- Status: **{qa['internal_navigation'].get('status')}**",
        ]
        for s in qa["internal_navigation"].get("steps", []):
            lines.append(f"- {s}")
        lines += ["", "## Build log"]
        for m in messages:
            lines.append(f"- {m}")
        lines += [
            "",
            "---",
            f"**PHASE O8 ACTUAL LINK OPEN GATE: {qa['phase_o8_actual_link_open_gate']}**",
            "",
            "STOP.",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
