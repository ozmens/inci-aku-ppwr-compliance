"""SYS_RELATIONSHIPS technical sheet — FK relationship catalog."""

from __future__ import annotations

from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from models.registry import SchemaRegistry

from .sheet_builder import PLACEHOLDER_ROWS, excel_col_letter

RELATIONSHIP_SHEET = "SYS_RELATIONSHIPS"

RELATIONSHIP_HEADERS = (
    "RELATIONSHIP_ID",
    "FROM_TABLE",
    "FROM_COLUMN",
    "TO_TABLE",
    "TO_COLUMN",
    "TO_NAMED_RANGE",
    "IS_REQUIRED",
    "RELATIONSHIP_CODE",
)


def build_relationship_rows(registry: SchemaRegistry) -> list[tuple]:
    rows: list[tuple] = []
    rel_id = 1
    for table in registry.tables:
        for col in table.columns:
            if not col.is_fk or not col.fk_table:
                continue
            try:
                parent = registry.get(col.fk_table)
            except KeyError:
                continue
            rows.append(
                (
                    rel_id,
                    table.name,
                    col.name,
                    parent.name,
                    parent.primary_key,
                    f"NR_{parent.primary_key}",
                    "Y" if col.required else "N",
                    f"FK_{table.name}_{col.name}",
                )
            )
            rel_id += 1
    return rows


def write_relationships_sheet(ws: Worksheet, registry: SchemaRegistry) -> Table:
    for col_idx, header in enumerate(RELATIONSHIP_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    rows = build_relationship_rows(registry)
    if not rows:
        # Keep table body present
        for col_idx in range(1, len(RELATIONSHIP_HEADERS) + 1):
            ws.cell(row=2, column=col_idx, value=None)
        last_row = 1 + PLACEHOLDER_ROWS
    else:
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        last_row = 1 + len(rows)

    last_col = excel_col_letter(len(RELATIONSHIP_HEADERS))
    ref = f"A1:{last_col}{last_row}"
    excel_table = Table(displayName="T_SYS_RELATIONSHIPS", ref=ref)
    ws.add_table(excel_table)
    return excel_table
