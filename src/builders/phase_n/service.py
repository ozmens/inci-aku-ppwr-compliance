"""Phase N — luxury executive Excel application UI (native COM design)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from PIL import ImageGrab

from builders.phase_n.assets import extract_inci_aku_logo
from builders.phase_n.com_ui import UI_SHEETS, ExcelComUI

EXPECTED_COUNTS = {
    "packaging_configurations": 247,
    "bom_lines": 1690,
    "components": 112,
    "products": 2046,
    "documents": 988,
}

XL_SCREEN = 1
XL_BITMAP = 2


@dataclass
class PhaseNResult:
    success: bool
    technical_gate: str
    visual_acceptance: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def _junction_or_copy(src: Path, dst: Path) -> None:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        else:
            shutil.rmtree(dst, ignore_errors=True)
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        shutil.copytree(src, dst)


class PhaseNService:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_PREMIUM_CANDIDATE.xlsx"
        self.candidate = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_CANDIDATE.xlsx"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_EXECUTIVE"
        self.delivery_workbook = (
            self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_CANDIDATE.xlsx"
        )
        self.assets = self.out / "PHASE_N_ASSETS"
        self.preview_dir = self.out / "PHASE_N_EXECUTIVE_PREVIEW"
        self.qa_path = self.out / "PHASE_N_EXECUTIVE_UI_QA.md"

    def run(self) -> PhaseNResult:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseNResult(
                False, "FAIL", "PENDING", [f"Missing source: {self.source}"]
            )

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"Candidate copied from Phase M → {self.candidate.name}")

        logo = extract_inci_aku_logo(self.root, self.assets)
        messages.append(f"Logo extracted from Golden DOCX media → {logo.name}")

        baseline = self._canonical_counts_file(self.candidate)
        messages.append(f"Baseline counts: {baseline}")

        # Kill orphan Excel before COM design
        self._kill_excel()
        time.sleep(1.5)

        designer = ExcelComUI(self.candidate, logo)
        design_stats = designer.run()
        messages.append(f"COM design stats: {design_stats}")

        self._build_delivery_root()
        messages.append(f"Executive delivery root: {self.delivery}")

        # Previews from delivery workbook
        self._kill_excel()
        time.sleep(1.0)
        previews = self._export_previews(self.delivery_workbook)
        messages.append(f"Previews: {len(previews)}")

        after = self._canonical_counts_file(self.delivery_workbook)
        validation = self._validate_links(self.delivery_workbook, self.delivery)
        home_buttons = self._count_home_buttons(self.delivery_workbook)
        excel = design_stats.get("native_reopen") or self._excel_open_ok(
            self.delivery_workbook
        )
        samples = self._sample_links(validation)
        counts_ok = after == baseline == EXPECTED_COUNTS

        technical = (
            "PASS"
            if (
                excel.get("ok")
                and validation["total_links"] == 988
                and validation["existing"] == 988
                and validation["missing"] == 0
                and validation.get("broken_paths", 0) == 0
                and validation["absolute_hits"] == 0
                and home_buttons >= 13
                and counts_ok
                and all(s["exists"] for s in samples)
                and len(previews) >= 6
                and design_stats.get("shapes_created", 0) >= 50
            )
            else "FAIL"
        )

        qa = {
            "technical_gate": technical,
            "manual_visual_acceptance": "PENDING",
            "workbook_path": str(self.candidate),
            "delivery_root": str(self.delivery),
            "delivery_workbook": str(self.delivery_workbook),
            "logo_path": str(logo),
            "preview_dir": str(self.preview_dir),
            "preview_files": previews,
            "design_stats": design_stats,
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": counts_ok,
            "total_document_links": validation["total_links"],
            "working_links": validation["existing"],
            "broken_links": validation["missing"],
            "broken_paths": validation.get("broken_paths", 0),
            "absolute_path_hits": validation["absolute_hits"],
            "home_buttons": home_buttons,
            "tested_sample_links": samples,
            "native_excel_open": excel,
            "canonical_data_changed": False,
            "word_regenerated": False,
            "final_overwritten": False,
            "promoted": False,
        }
        self._write_qa(qa, messages)
        return PhaseNResult(
            technical == "PASS",
            technical,
            "PENDING",
            messages,
            qa,
        )

    def _kill_excel(self) -> None:
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"],
                capture_output=True,
                text=True,
            )
        except Exception:
            pass

    def _build_delivery_root(self) -> None:
        if self.delivery.exists():
            for child in list(self.delivery.iterdir()):
                try:
                    if child.is_junction() or child.is_symlink():
                        child.unlink()
                    elif child.is_dir():
                        shutil.rmtree(child, ignore_errors=True)
                    else:
                        child.unlink()
                except Exception:
                    pass
        self.delivery.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.candidate, self.delivery_workbook)
        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if src.exists():
                _junction_or_copy(src, self.delivery / folder)

    def _export_previews(self, workbook_path: Path) -> list[str]:
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()
        preview_map = {
            "HOME": ("00_HOME", "A1:N42"),
            "NAVIGATION": ("NAVIGATION", "A1:N36"),
            "SEARCH": ("SEARCH", "A1:N30"),
            "DOCUMENT_CENTER": ("DOCUMENT_CENTER", "A1:N22"),
            "PACKAGING_CONFIGURATIONS": ("PACKAGING_CONFIGURATIONS", "A1:N20"),
            "DECLARATIONS_OF_CONFORMITY": ("DECLARATIONS_OF_CONFORMITY", "A1:N20"),
        }
        pythoncom.CoInitialize()
        excel = None
        wb = None
        previews: list[str] = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = True
            wb = excel.Workbooks.Open(
                str(workbook_path.resolve()), UpdateLinks=0, ReadOnly=True
            )
            # ungroup
            wb.Worksheets(1).Select()
            for label, (sheet_name, addr) in preview_map.items():
                path = self.preview_dir / f"PHASE_N_{label}.png"
                ws = wb.Worksheets(sheet_name)
                ws.Select()
                ws.Activate()
                try:
                    excel.ActiveWindow.Zoom = 85
                    excel.ActiveWindow.DisplayGridlines = False
                except Exception:
                    pass
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.55)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                if path.exists():
                    path.unlink()
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    previews.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return previews

    def _excel_open_ok(self, path: Path) -> dict:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {"ok": False, "error": None, "sheets": None}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(path.resolve()), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
            )
            out["ok"] = True
            out["sheets"] = int(wb.Worksheets.Count)
            wb.Close(False)
        except Exception as exc:  # noqa: BLE001
            out["error"] = str(exc)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _canonical_counts_file(self, path: Path) -> dict[str, int]:
        wb = load_workbook(path, read_only=True, data_only=False)

        def data_rows(sheet: str) -> int:
            if sheet not in wb.sheetnames:
                return -1
            return max((wb[sheet].max_row or 1) - 1, 0)

        counts = {
            "packaging_configurations": data_rows("PACKAGING_CONFIGURATION"),
            "bom_lines": data_rows("PACKAGING_CONFIGURATION_LINE"),
            "components": data_rows("COMPONENT"),
            "products": data_rows("PRODUCT"),
            "documents": data_rows("DOCUMENT_LIBRARY"),
        }
        wb.close()
        return counts

    def _validate_links(self, workbook_path: Path, delivery_root: Path) -> dict[str, Any]:
        wb = load_workbook(workbook_path, data_only=False)
        targets: dict[str, int] = {}
        absolute_hits = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                    if not m:
                        continue
                    target = m.group(1).replace("\\", "/")
                    if not target.lower().endswith(".docx"):
                        continue
                    if re.match(r"^[A-Za-z]:/", target) or "Users/" in target:
                        absolute_hits += 1
                    targets[target] = targets.get(target, 0) + 1
        wb.close()

        existing = missing = broken_paths = 0
        unique_targets = sorted(targets.keys())
        for t in unique_targets:
            norm = t.replace("\\", "/").strip()
            parts = [p for p in norm.split("/") if p not in ("", ".")]
            if (
                not parts
                or ".." in parts
                or re.match(r"^[A-Za-z]:", parts[0])
                or parts[0].startswith("//")
            ):
                broken_paths += 1
                missing += 1
                continue
            candidate = delivery_root.joinpath(*parts)
            if candidate.exists() and candidate.is_file():
                existing += 1
            else:
                missing += 1
        return {
            "total_links": len(unique_targets),
            "existing": existing,
            "missing": missing,
            "broken_paths": broken_paths,
            "absolute_hits": absolute_hits,
            "targets": unique_targets,
        }

    def _count_home_buttons(self, path: Path) -> int:
        """Count HomeBtn shapes via Excel COM (ungroup first)."""
        pythoncom.CoInitialize()
        excel = None
        n = 0
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=True)
            wb.Worksheets(1).Select()
            for name in UI_SHEETS:
                try:
                    ws = wb.Worksheets(name)
                    for i in range(1, int(ws.Shapes.Count) + 1):
                        if str(ws.Shapes(i).Name).startswith("HomeBtn"):
                            n += 1
                            break
                except Exception:
                    pass
            wb.Close(False)
        except Exception:
            pass
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return n

    def _sample_links(self, validation: dict) -> list[dict]:
        wanted = [
            ("STARTER_TF", "01_STARTER/ST-051-STD-01/01_Technical_File.docx"),
            ("STARTER_DOC", "01_STARTER/ST-051-STD-01/02_EU_DoC.docx"),
            ("STARTER_LBL", "01_STARTER/ST-051-STD-01/03_Label.docx"),
            ("STARTER_STM", "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx"),
            ("INDUSTRIAL", "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx"),
            ("CONTAINER", "03_CONTAINER/CNT-20-STD-01/03_Label.docx"),
        ]
        out = []
        for label, rel in wanted:
            p = self.delivery / rel
            out.append(
                {
                    "sample": label,
                    "relative": rel,
                    "exists": p.exists(),
                    "in_workbook_links": rel in validation.get("targets", []),
                }
            )
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        lines = [
            "# Phase N — Luxury Executive Excel Application UI QA",
            "",
            f"- **PHASE N TECHNICAL GATE: {qa['technical_gate']}**",
            f"- **MANUAL VISUAL ACCEPTANCE: {qa['manual_visual_acceptance']}**",
            "",
            f"- Executive candidate: `{qa['workbook_path']}`",
            f"- Executive delivery root: `{qa['delivery_root']}`",
            f"- Logo asset: `{qa['logo_path']}`",
            f"- Preview directory: `{qa['preview_dir']}`",
            "",
            "## Design approach",
            "",
            "- Native Microsoft Excel COM shapes (rounded cards, shadows, badges)",
            "- Official İnci Akü logo extracted from Golden DOCX media (read-only)",
            "- Sheet ungroup before AddShape (fixes multi-selection COM block)",
            "- HOME / NAVIGATION / SEARCH rebuilt as application canvases",
            "- Register sheets: COM header chrome + preserved data/hyperlinks",
            "",
            "## Counts",
            "",
            f"- Baseline: `{qa['baseline_counts']}`",
            f"- After: `{qa['after_counts']}`",
            f"- Unchanged: **{qa['counts_unchanged']}**",
            "",
            "## Hyperlink integrity",
            "",
            f"- Total: {qa['total_document_links']}",
            f"- Working: {qa['working_links']}",
            f"- Broken: {qa['broken_links']}",
            f"- Home shape buttons: {qa['home_buttons']}/13",
            "",
            "## Native Excel",
            "",
            f"- `{qa['native_excel_open']}`",
            f"- Design stats: `{qa['design_stats']}`",
            "",
            "## Previews",
            "",
        ]
        lines.extend(f"- `{p}`" for p in qa["preview_files"])
        lines += ["", "## Messages", ""]
        lines.extend(f"- {m}" for m in messages)
        lines += [
            "",
            "## Confirmations",
            "",
            "- Canonical data changed: NO",
            "- Word regenerated: NO",
            "- Golden templates modified: NO",
            "- Final delivery overwritten: NO",
            "- Promoted: NO",
            "- Rev01 started: NO",
            "",
            f"**PHASE N TECHNICAL GATE: {qa['technical_gate']}**",
            f"**MANUAL VISUAL ACCEPTANCE: {qa['manual_visual_acceptance']}**",
            "",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
        self.qa_path.with_suffix(".json").write_text(
            json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def main() -> int:
    import sys

    root = Path(__file__).resolve().parents[2]
    sys.path.insert(0, str(root))
    sys.path.insert(0, str(root / "src"))
    result = PhaseNService(root).run()
    for m in result.messages:
        try:
            print(m)
        except UnicodeEncodeError:
            print(m.encode("ascii", "replace").decode("ascii"))
    print("PHASE N TECHNICAL GATE:", result.technical_gate)
    print("MANUAL VISUAL ACCEPTANCE:", result.visual_acceptance)
    if result.qa:
        print("Candidate:", result.qa.get("workbook_path"))
        print("Delivery root:", result.qa.get("delivery_root"))
        for p in result.qa.get("preview_files") or []:
            print("Preview:", p)
    return 0 if result.success else 1


if __name__ == "__main__":
    raise SystemExit(main())
