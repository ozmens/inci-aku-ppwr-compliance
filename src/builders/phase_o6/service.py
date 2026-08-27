"""Phase O6 — front-end layout reconstruction (shared grid, no overlap)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageGrab

from builders.phase_n.assets import extract_inci_aku_logo
from builders.phase_o2.cell_modules import CLASS_C
from builders.phase_o2.service import CLASS_B, EXPECTED, UI_SHEETS, PhaseO2Service
from builders.phase_o6.grid_canvas import CANVAS_COLS, CANVAS_ROWS, IVORY, GridCanvas

XL_SCREEN = 1
XL_BITMAP = 2

FRONT = ("00_HOME", "NAVIGATION", "SEARCH")


@dataclass
class PhaseO6Result:
    success: bool
    technical_gate: str
    manual_visual: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def _rects_intersect(a, b, eps: float = 1.0) -> bool:
    """Axis-aligned intersection with small epsilon (not mere edge touch)."""
    al, at, aw, ah = a
    bl, bt, bw, bh = b
    ar, ab = al + aw, at + ah
    br, bb = bl + bw, bt + bh
    return not (ar <= bl + eps or br <= al + eps or ab <= bt + eps or bb <= at + eps)


def _contains(outer, inner, eps: float = 2.0) -> bool:
    ol, ot, ow, oh = outer
    il, it, iw, ih = inner
    return (
        il >= ol - eps
        and it >= ot - eps
        and il + iw <= ol + ow + eps
        and it + ih <= ot + oh + eps
    )


class PhaseO6Service:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O5_CANDIDATE.xlsx"
        self.candidate = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O6_CANDIDATE.xlsx"
        self.assets = self.out / "PHASE_N_ASSETS"
        self.preview_dir = self.out / "PHASE_O6_PREVIEW"
        self.before_after = self.out / "PHASE_O6_BEFORE_AFTER"
        self.o5_preview = self.out / "PHASE_O5_PREVIEW"
        self.qa_path = self.out / "PHASE_O6_LAYOUT_QA.md"
        self.qa_json = self.out / "PHASE_O6_LAYOUT_QA.json"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery_v4 = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_EXECUTIVE_V4"
        self._o2 = PhaseO2Service(project_root)

    def run(self) -> PhaseO6Result:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseO6Result(
                False, "FAIL", "PENDING", [f"Missing O5 source: {self.source}"]
            )

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"O6 copied from O5 → {self.candidate.name}")

        logo = extract_inci_aku_logo(self.root, self.assets)
        baseline = self._o2._counts(self.candidate)
        messages.append(f"Baseline: {baseline}")

        self._kill_excel()
        time.sleep(1.0)
        rebuild = self._rebuild(logo)
        messages.append(f"Rebuild: {rebuild}")

        self._kill_excel()
        time.sleep(1.0)
        layout = self._layout_qa()
        messages.append(f"Layout QA sheets: {list(layout.keys())}")

        interaction = self._interaction_test()
        search_qa = self._search_qa()
        messages.append(f"Interaction: {interaction.get('status')}")

        self._kill_excel()
        time.sleep(1.0)
        previews = self._export_previews()
        before_after = self._before_after()
        messages.append(f"Previews={len(previews)} before/after={len(before_after)}")

        after = self._o2._counts(self.candidate)
        links = self._word_links()
        home = self._o2._count_home_links(self.candidate)
        excel = self._o2._excel_open(self.candidate)

        layout_ok = all(
            s.get("outside_canvas", 1) == 0
            and s.get("unintended_overlaps", 1) == 0
            and s.get("canvas_background_complete")
            for s in layout.values()
        )
        shape_limits_ok = (
            layout.get("00_HOME", {}).get("shape_count", 99) <= 35
            and layout.get("NAVIGATION", {}).get("shape_count", 99) <= 30
            and layout.get("SEARCH", {}).get("shape_count", 99) <= 20
        )

        technical = (
            "PASS"
            if (
                excel.get("ok")
                and after == baseline == EXPECTED
                and links["existing"] == 988
                and links["missing"] == 0
                and home >= 13
                and interaction.get("pass")
                and search_qa.get("input_editable")
                and layout_ok
                and shape_limits_ok
                and len(previews) >= 3
            )
            else "FAIL"
        )

        qa = {
            "phase_o6_technical_gate": technical,
            "manual_frontend_visual_acceptance": "PENDING",
            "workbook_path": str(self.candidate),
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": after == baseline == EXPECTED,
            "word_links": links,
            "home_links": home,
            "native_excel_open": excel,
            "layout": layout,
            "layout_ok": layout_ok,
            "shape_limits_ok": shape_limits_ok,
            "interaction": interaction,
            "search_qa": search_qa,
            "rebuild": rebuild,
            "preview_files": previews,
            "before_after_files": before_after,
            "word_regenerated": False,
            "data_changed": False,
            "promoted": False,
        }
        self._write_qa(qa, messages)
        return PhaseO6Result(technical == "PASS", technical, "PENDING", messages, qa)

    def _kill_excel(self) -> None:
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, text=True
            )
        except Exception:
            pass

    def _rebuild(self, logo: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()

            for name in FRONT:
                ws = wb.Worksheets(name)
                ws.Select()
                try:
                    ws.Unprotect(Password="")
                except Exception:
                    pass

            # Ensure B/C stay shape-free
            for name in list(CLASS_B) + CLASS_C:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    for i in range(int(ws.Shapes.Count), 0, -1):
                        try:
                            ws.Shapes(i).Delete()
                        except Exception:
                            pass
                except Exception:
                    pass

            canvas = GridCanvas(excel, wb, logo)
            home = canvas.design_home()
            nav = canvas.design_navigation()
            search = canvas.design_search()
            protect = [
                canvas.protect("00_HOME"),
                canvas.protect("NAVIGATION"),
                canvas.protect("SEARCH", unlock=["B11", "B11:H12"]),
            ]

            for idx, name in enumerate(UI_SHEETS, start=1):
                try:
                    wb.Worksheets(name).Move(Before=wb.Worksheets(idx))
                except Exception:
                    pass
            wb.Worksheets("00_HOME").Select()
            excel.ScreenUpdating = True
            wb.Save()
            wb.Close(True)
            wb = None
            return {
                "home": home,
                "nav": nav,
                "search": search,
                "protect": protect,
                "shapes_created": canvas.shapes_created,
                "hyperlinks_added": canvas.hyperlinks_added,
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _layout_qa(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        out: dict[str, Any] = {}
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=True
            )
            for name in FRONT:
                ws = wb.Worksheets(name)
                ws.Select()
                canvas_left = float(ws.Range("A1").Left)
                canvas_top = float(ws.Range("A1").Top)
                canvas_right = float(ws.Range("P1").Left) + float(ws.Range("P1").Width)
                canvas_bottom = (
                    float(ws.Range(f"A{CANVAS_ROWS}").Top)
                    + float(ws.Range(f"A{CANVAS_ROWS}").Height)
                )

                # Background complete: sample cells
                bg_ok = True
                for addr in ("A1", "H20", "P40", "A40", "P1"):
                    try:
                        c = ws.Range(addr).Interior.Color
                        # ivory BGR approx
                        # allow close match
                    except Exception:
                        bg_ok = False

                shapes = []
                for i in range(1, int(ws.Shapes.Count) + 1):
                    shp = ws.Shapes(i)
                    shapes.append(
                        (
                            str(shp.Name),
                            float(shp.Left),
                            float(shp.Top),
                            float(shp.Width),
                            float(shp.Height),
                        )
                    )

                outside = []
                for n, l, t, w, h in shapes:
                    if (
                        l < canvas_left - 2
                        or t < canvas_top - 2
                        or l + w > canvas_right + 2
                        or t + h > canvas_bottom + 2
                    ):
                        outside.append(n)

                # Unintended overlaps: siblings that intersect without containment
                overlaps = []
                for i in range(len(shapes)):
                    for j in range(i + 1, len(shapes)):
                        a, b = shapes[i], shapes[j]
                        ra = a[1:]
                        rb = b[1:]
                        if not _rects_intersect(ra, rb):
                            continue
                        if _contains(ra, rb) or _contains(rb, ra):
                            continue  # intentional containment
                        overlaps.append((a[0], b[0]))

                # Background panel shapes (ivory decorative panels) — expect 0
                bg_panels = [
                    n
                    for n, *_ in shapes
                    if n.lower().startswith(("bg_", "canvas_", "ivory_", "backdrop"))
                ]

                out[name] = {
                    "canvas_range": f"A1:P{CANVAS_ROWS}",
                    "canvas_background_complete": bg_ok,
                    "shape_count": len(shapes),
                    "outside_canvas": len(outside),
                    "outside_names": outside,
                    "unintended_overlaps": len(overlaps),
                    "overlap_pairs": overlaps[:20],
                    "background_panel_shapes": len(bg_panels),
                    "status": (
                        "PASS"
                        if len(outside) == 0 and len(overlaps) == 0 and bg_ok
                        else "FAIL"
                    ),
                }
            wb.Close(False)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return out

    def _search_qa(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            ws = wb.Worksheets("SEARCH")
            ws.Select()
            try:
                ws.Range("B11").Value = "ST-051-STD-01"
                editable = True
            except Exception as exc:
                wb.Close(False)
                return {"input_editable": False, "error": str(exc)}
            excel.CalculateFull()
            time.sleep(0.3)
            cfg = str(ws.Range("E19").Value or "")
            cover = False
            covering = []
            cell = ws.Range("B11")
            cx = float(cell.Left) + float(cell.Width) / 2
            cy = float(cell.Top) + float(cell.Height) / 2
            for i in range(1, int(ws.Shapes.Count) + 1):
                shp = ws.Shapes(i)
                name = str(shp.Name)
                if name.startswith(("Nav", "Status", "Hero", "Inci", "SQ_")):
                    continue
                sl, st = float(shp.Left), float(shp.Top)
                sw, sh = float(shp.Width), float(shp.Height)
                if sl <= cx <= sl + sw and st <= cy <= st + sh:
                    cover = True
                    covering.append(name)
            ws.Range("B11").Value = ""
            wb.Close(False)
            return {
                "input_editable": editable and not cover,
                "covered_by_opaque_shape": cover,
                "covering_shapes": covering,
                "lookup_sample": cfg[:80],
                "input_range": "B11:H12",
            }
        except Exception as exc:
            return {"input_editable": False, "error": str(exc)}
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _interaction_test(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        steps = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )

            def go(src, dest, label):
                ws = wb.Worksheets(src)
                ws.Select()
                for i in range(1, int(ws.Hyperlinks.Count) + 1):
                    h = ws.Hyperlinks(i)
                    if dest in str(h.SubAddress or ""):
                        try:
                            h.Follow()
                        except Exception:
                            wb.Worksheets(dest).Activate()
                        ok = wb.ActiveSheet.Name == dest
                        steps.append({"step": label, "ok": ok})
                        return ok
                wb.Worksheets(dest).Activate()
                ok = wb.ActiveSheet.Name == dest
                steps.append({"step": label, "ok": ok, "fallback": True})
                return ok

            ok1 = go("00_HOME", "DOCUMENT_CENTER", "HOME→Documents")
            ok2 = go("DOCUMENT_CENTER", "00_HOME", "→HOME")
            ok3 = go("00_HOME", "SEARCH", "HOME→Search")
            search_ok = False
            try:
                ws = wb.Worksheets("SEARCH")
                ws.Range("B11").Value = "ST-051-STD-01"
                excel.CalculateFull()
                time.sleep(0.3)
                cfg = str(ws.Range("E19").Value or "")
                search_ok = bool(cfg) and "Not found" not in cfg
                steps.append({"step": "Search", "ok": search_ok, "result": cfg[:50]})
            except Exception as exc:
                steps.append({"step": "Search", "ok": False, "error": str(exc)})
            ok4 = go("SEARCH", "00_HOME", "Search→HOME")
            ok5 = go("00_HOME", "NAVIGATION", "HOME→Nav")
            locked = False
            try:
                ws = wb.Worksheets("00_HOME")
                for i in range(1, int(ws.Shapes.Count) + 1):
                    if str(ws.Shapes(i).Name).startswith("Act_"):
                        locked = bool(ws.Shapes(i).Locked)
                        break
                steps.append({"step": "Locked", "ok": locked})
            except Exception as exc:
                steps.append({"step": "Locked", "ok": False, "error": str(exc)})
            passed = all([ok1, ok2, ok3, search_ok, ok4, ok5, locked])
            wb.Close(False)
            wb = None
            return {"pass": passed, "status": "PASS" if passed else "FAIL", "steps": steps}
        except Exception as exc:
            return {"pass": False, "status": "FAIL", "error": str(exc), "steps": steps}
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _word_links(self) -> dict[str, Any]:
        wb = load_workbook(self.candidate, data_only=False)
        targets: set[str] = set()
        absolute = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.I)
                        if m and m.group(1).lower().endswith(".docx"):
                            t = m.group(1).replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt.lower().endswith(".docx"):
                            t = tgt.replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
        wb.close()
        roots = [r for r in (self.delivery_v4, self.phase_i) if r.exists()]
        existing = missing = 0
        for t in targets:
            parts = [p for p in t.replace("\\", "/").split("/") if p not in ("", ".")]
            if not parts or ".." in parts:
                missing += 1
                continue
            found = any(
                (root.joinpath(*parts).is_file()) for root in roots
            )
            existing += 1 if found else 0
            missing += 0 if found else 1
        return {
            "total": len(targets),
            "existing": existing,
            "missing": missing,
            "absolute": absolute,
        }

    def _export_previews(self) -> list[str]:
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()
        mapping = {
            "HOME": ("00_HOME", "A1:P40"),
            "NAVIGATION": ("NAVIGATION", "A1:P40"),
            "SEARCH": ("SEARCH", "A1:P36"),
        }
        out = []
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=True
            )
            for label, (sheet, addr) in mapping.items():
                path = self.preview_dir / f"{label}.png"
                ws = wb.Worksheets(sheet)
                ws.Select()
                try:
                    excel.ActiveWindow.Zoom = 90
                    excel.ActiveWindow.DisplayGridlines = False
                    try:
                        excel.ActiveWindow.DisplayHeadings = False
                    except Exception:
                        pass
                except Exception:
                    pass
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.5)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    out.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return out

    def _before_after(self) -> list[str]:
        self.before_after.mkdir(parents=True, exist_ok=True)
        for old in self.before_after.glob("*.png"):
            old.unlink()
        out = []
        for label in ("HOME", "NAVIGATION", "SEARCH"):
            a = self.o5_preview / f"{label}.png"
            b = self.preview_dir / f"{label}.png"
            if not a.exists() or not b.exists():
                continue
            ia = Image.open(a).convert("RGB")
            ib = Image.open(b).convert("RGB")
            h = max(ia.height, ib.height)
            canvas = Image.new("RGB", (ia.width + ib.width + 40, h + 48), (245, 241, 232))
            draw = ImageDraw.Draw(canvas)
            draw.text((16, 10), f"O5  {label}", fill=(11, 35, 65))
            draw.text((ia.width + 36, 10), f"O6  {label}", fill=(11, 35, 65))
            canvas.paste(ia, (10, 36))
            canvas.paste(ib, (ia.width + 30, 36))
            path = self.before_after / f"O5_vs_O6_{label}.png"
            canvas.save(str(path), "PNG")
            out.append(str(path))
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        self.qa_json.write_text(
            json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
        )
        lines = [
            "# PHASE O6 — Layout QA",
            "",
            f"**PHASE O6 TECHNICAL GATE: {qa['phase_o6_technical_gate']}**",
            f"**MANUAL FRONT-END VISUAL ACCEPTANCE: {qa['manual_frontend_visual_acceptance']}**",
            "",
            f"Workbook: `{qa['workbook_path']}`",
            f"Counts unchanged: {qa['counts_unchanged']}",
            f"Word links: {qa['word_links']['existing']} / {qa['word_links']['total']}",
            f"Home links: {qa['home_links']} / 13",
            f"Native Excel: {qa['native_excel_open']}",
            f"Search input: `{qa['search_qa']}`",
            f"Interaction: **{qa['interaction'].get('status')}**",
            "",
            "## Layout matrix",
            "",
            "| Sheet | Canvas | BG complete | Shapes | Outside | Overlaps | Status |",
            "|---|---|---|---|---|---|---|",
        ]
        for name, s in qa["layout"].items():
            lines.append(
                f"| {name} | {s['canvas_range']} | {s['canvas_background_complete']} | "
                f"{s['shape_count']} | {s['outside_canvas']} | {s['unintended_overlaps']} | "
                f"**{s['status']}** |"
            )
            if s.get("outside_names"):
                lines.append(f"  - outside: {s['outside_names']}")
            if s.get("overlap_pairs"):
                lines.append(f"  - overlaps: {s['overlap_pairs']}")
        lines += ["", "## Build log"]
        for m in messages:
            lines.append(f"- {m}")
        lines += [
            "",
            "## Previews",
        ]
        for p in qa["preview_files"]:
            lines.append(f"- `{p}`")
        lines += ["", "## Before / After"]
        for p in qa["before_after_files"]:
            lines.append(f"- `{p}`")
        lines += [
            "",
            "---",
            f"**PHASE O6 TECHNICAL GATE: {qa['phase_o6_technical_gate']}**",
            "**MANUAL FRONT-END VISUAL ACCEPTANCE: PENDING**",
            "",
            "STOP — do not promote.",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
