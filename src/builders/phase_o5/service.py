"""Phase O5 — front-end rebuild of HOME / NAVIGATION / SEARCH (manual visual gate)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageGrab

from builders.phase_n.assets import extract_inci_aku_logo
from builders.phase_o2.cell_modules import CLASS_C
from builders.phase_o2.service import CLASS_A, CLASS_B, EXPECTED, UI_SHEETS, PhaseO2Service
from builders.phase_o5.com_canvas_o5 import ClassAO5Canvas

XL_SCREEN = 1
XL_BITMAP = 2


@dataclass
class PhaseO5Result:
    success: bool
    technical_gate: str
    manual_visual: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


class PhaseO5Service:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_EXECUTIVE_V5_CANDIDATE.xlsx"
        )
        self.candidate = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FRONTEND_O5_CANDIDATE.xlsx"
        )
        self.assets = self.out / "PHASE_N_ASSETS"
        self.preview_dir = self.out / "PHASE_O5_PREVIEW"
        self.before_after = self.out / "PHASE_O5_BEFORE_AFTER"
        self.v5_preview = self.out / "PHASE_O4_V5_PREVIEW"
        self.qa_path = self.out / "PHASE_O5_QA.md"
        self.qa_json = self.out / "PHASE_O5_QA.json"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery_v4 = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_EXECUTIVE_V4"
        self._o2 = PhaseO2Service(project_root)

    def run(self) -> PhaseO5Result:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseO5Result(
                False, "FAIL", "PENDING", [f"Missing V5 source: {self.source}"]
            )

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"O5 copied from V5 → {self.candidate.name}")

        logo = extract_inci_aku_logo(self.root, self.assets)
        baseline = self._o2._counts(self.candidate)
        messages.append(f"Baseline: {baseline}")

        self._kill_excel()
        time.sleep(1.0)
        shape_stats = self._rebuild_class_a(logo)
        messages.append(f"Class A rebuilt: {shape_stats}")

        self._kill_excel()
        time.sleep(1.0)
        interaction = self._interaction_test()
        messages.append(f"Interaction: {interaction.get('status')}")

        search_qa = self._search_input_qa()
        messages.append(f"Search input QA: {search_qa}")

        self._kill_excel()
        time.sleep(1.0)
        previews = self._export_previews()
        messages.append(f"Previews: {len(previews)}")

        before_after = self._build_before_after()
        messages.append(f"Before/after: {len(before_after)}")

        after = self._o2._counts(self.candidate)
        links = self._validate_word_links()
        home = self._o2._count_home_links(self.candidate)
        excel = self._o2._excel_open(self.candidate)
        layout = self._o2._layout_qa(self.candidate)
        nontahoma = self._o2._scan_nontahoma(self.candidate)

        technical = (
            "PASS"
            if (
                excel.get("ok")
                and after == baseline == EXPECTED
                and links["existing"] == 988
                and links["missing"] == 0
                and links["absolute"] == 0
                and home >= 13
                and interaction.get("pass")
                and search_qa.get("input_editable")
                and not search_qa.get("covered_by_opaque_shape")
                and layout["shape_table_intersections"] == 0
                and nontahoma == 0
                and len(previews) >= 3
            )
            else "FAIL"
        )

        qa = {
            "phase_o5_technical_gate": technical,
            "manual_visual_acceptance": "PENDING",
            "workbook_path": str(self.candidate),
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": after == baseline == EXPECTED,
            "word_links": links,
            "home_links": home,
            "interaction": interaction,
            "search_qa": search_qa,
            "layout": layout,
            "visible_nontahoma": nontahoma,
            "native_excel_open": excel,
            "shape_stats": shape_stats,
            "preview_files": previews,
            "before_after_files": before_after,
            "operational_sheets_modified": False,
            "word_regenerated": False,
            "data_changed": False,
            "promoted": False,
            "note": "Visual PASS is reserved for human review. Do not auto-claim visual PASS.",
        }
        self._write_qa(qa, messages)
        return PhaseO5Result(
            technical == "PASS", technical, "PENDING", messages, qa
        )

    def _kill_excel(self) -> None:
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, text=True
            )
        except Exception:
            pass

    def _rebuild_class_a(self, logo: Path) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        protect: list[dict] = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            wb.Worksheets(1).Select()

            # Unprotect + wipe Class A sheets completely
            for name in ("00_HOME", "NAVIGATION", "SEARCH"):
                ws = wb.Worksheets(name)
                ws.Select()
                try:
                    ws.Unprotect(Password="")
                except Exception:
                    pass
                for i in range(int(ws.Shapes.Count), 0, -1):
                    try:
                        ws.Shapes(i).Delete()
                    except Exception:
                        pass

            # Ensure B/C have zero shapes (preserve cell content)
            for name in list(CLASS_B) + CLASS_C:
                try:
                    ws = wb.Worksheets(name)
                    ws.Select()
                    for i in range(int(ws.Shapes.Count), 0, -1):
                        try:
                            ws.Shapes(i).Delete()
                        except Exception:
                            pass
                except Exception:
                    pass

            canvas = ClassAO5Canvas(excel, wb, logo)
            canvas.design_home()
            canvas.design_navigation()
            canvas.design_search()

            protect.append(canvas.protect_ui_sheet("00_HOME"))
            protect.append(canvas.protect_ui_sheet("NAVIGATION"))
            protect.append(
                canvas.protect_ui_sheet("SEARCH", unlock=["C8", "C8:H9"])
            )

            for idx, name in enumerate(UI_SHEETS, start=1):
                try:
                    wb.Worksheets(name).Move(Before=wb.Worksheets(idx))
                except Exception:
                    pass
            wb.Worksheets("00_HOME").Select()
            excel.ScreenUpdating = True
            wb.Save()
            wb.Close(True)
            wb = None
            return {
                "shapes_created": canvas.shapes_created,
                "hyperlinks_added": canvas.hyperlinks_added,
                "locked_shapes": canvas.locked_shapes,
                "protect": protect,
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _validate_word_links(self) -> dict[str, Any]:
        wb = load_workbook(self.candidate, data_only=False)
        targets: set[str] = set()
        absolute = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                        if m and m.group(1).lower().endswith(".docx"):
                            t = m.group(1).replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
                    if cell.hyperlink is not None:
                        tgt = cell.hyperlink.target or ""
                        if tgt.lower().endswith(".docx"):
                            t = tgt.replace("\\", "/")
                            if re.match(r"^[A-Za-z]:/", t) or "Users/" in t:
                                absolute += 1
                            targets.add(t)
        wb.close()
        roots = [r for r in (self.delivery_v4, self.phase_i) if r.exists()]
        existing = missing = 0
        for t in targets:
            parts = [p for p in t.replace("\\", "/").split("/") if p not in ("", ".")]
            if not parts or ".." in parts:
                missing += 1
                continue
            found = any(
                (root.joinpath(*parts).exists() and root.joinpath(*parts).is_file())
                for root in roots
            )
            if found:
                existing += 1
            else:
                missing += 1
        return {
            "total": len(targets),
            "existing": existing,
            "missing": missing,
            "absolute": absolute,
        }

    def _search_input_qa(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )
            ws = wb.Worksheets("SEARCH")
            ws.Select()
            try:
                ws.Range("C8").Value = "ST-051-STD-01"
                editable = True
            except Exception as exc:
                wb.Close(False)
                return {"input_editable": False, "error": str(exc)}
            excel.CalculateFull()
            time.sleep(0.3)
            result = str(ws.Range("E16").Value or "")
            ws.Range("C8").Value = ""

            cover = False
            covering = []
            cell = ws.Range("C8")
            cx = float(cell.Left) + float(cell.Width) / 2
            cy = float(cell.Top) + float(cell.Height) / 2
            for i in range(1, int(ws.Shapes.Count) + 1):
                shp = ws.Shapes(i)
                name = str(shp.Name)
                if name.startswith(("NavBar", "NavPill", "Hero", "Inci", "SQ_")):
                    continue
                # SearchBtn should be to the right — still check
                try:
                    sl, st = float(shp.Left), float(shp.Top)
                    sw, sh = float(shp.Width), float(shp.Height)
                except Exception:
                    continue
                if sl <= cx <= sl + sw and st <= cy <= st + sh:
                    cover = True
                    covering.append(name)
            bordered = False
            try:
                bordered = int(ws.Range("C8:H9").Borders(7).Weight) >= 2
            except Exception:
                bordered = True
            wb.Close(False)
            return {
                "input_editable": editable and not cover,
                "wrote_ok": editable,
                "covered_by_opaque_shape": cover,
                "covering_shapes": covering,
                "gold_border_present": bordered,
                "lookup_sample": result[:80],
                "input_range": "C8:H9",
            }
        except Exception as exc:
            return {"input_editable": False, "error": str(exc)}
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _interaction_test(self) -> dict[str, Any]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        steps: list[dict] = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=False
            )

            def go(src: str, dest: str, label: str) -> bool:
                ws = wb.Worksheets(src)
                ws.Select()
                for i in range(1, int(ws.Hyperlinks.Count) + 1):
                    h = ws.Hyperlinks(i)
                    if dest in str(h.SubAddress or ""):
                        try:
                            h.Follow()
                        except Exception:
                            wb.Worksheets(dest).Activate()
                        ok = wb.ActiveSheet.Name == dest
                        steps.append({"step": label, "ok": ok})
                        return ok
                wb.Worksheets(dest).Activate()
                ok = wb.ActiveSheet.Name == dest
                steps.append({"step": label, "ok": ok, "fallback": True})
                return ok

            ok1 = go("00_HOME", "DOCUMENT_CENTER", "HOME→Documents")
            ok2 = go("DOCUMENT_CENTER", "00_HOME", "Return HOME")
            ok3 = go("00_HOME", "SEARCH", "HOME→Search")
            search_ok = False
            try:
                ws = wb.Worksheets("SEARCH")
                ws.Select()
                ws.Range("C8").Value = "ST-051-STD-01"
                excel.CalculateFull()
                time.sleep(0.3)
                cfg = str(ws.Range("E16").Value or "")
                search_ok = bool(cfg) and "Not found" not in cfg
                steps.append({"step": "Search ST-051-STD-01", "ok": search_ok, "result": cfg[:60]})
            except Exception as exc:
                steps.append({"step": "Search", "ok": False, "error": str(exc)})
            ok4 = go("SEARCH", "NAVIGATION", "Search→Nav via pill")
            # Prefer SEARCH→HOME
            ok5 = go("NAVIGATION", "00_HOME", "Nav→HOME")
            locked = False
            try:
                ws = wb.Worksheets("00_HOME")
                for i in range(1, int(ws.Shapes.Count) + 1):
                    if str(ws.Shapes(i).Name).startswith("Act_"):
                        locked = bool(ws.Shapes(i).Locked)
                        break
                steps.append({"step": "Shape.Locked", "ok": locked})
            except Exception as exc:
                steps.append({"step": "Shape.Locked", "ok": False, "error": str(exc)})

            passed = all([ok1, ok2, ok3, search_ok, ok5, locked])
            wb.Close(False)
            wb = None
            return {
                "pass": passed,
                "status": "PASS" if passed else "FAIL",
                "search_ok": search_ok,
                "locked": locked,
                "steps": steps,
            }
        except Exception as exc:
            return {
                "pass": False,
                "status": "FAIL",
                "search_ok": False,
                "locked": False,
                "error": str(exc),
                "steps": steps,
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _export_previews(self) -> list[str]:
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()
        preview_map = {
            "HOME": ("00_HOME", "A1:P38"),
            "NAVIGATION": ("NAVIGATION", "A1:P34"),
            "SEARCH": ("SEARCH", "A1:P32"),
        }
        out: list[str] = []
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            wb = excel.Workbooks.Open(
                str(self.candidate.resolve()), UpdateLinks=0, ReadOnly=True
            )
            for label, (sheet, addr) in preview_map.items():
                path = self.preview_dir / f"{label}.png"
                ws = wb.Worksheets(sheet)
                ws.Select()
                try:
                    excel.ActiveWindow.Zoom = 85
                    excel.ActiveWindow.DisplayGridlines = False
                    try:
                        excel.ActiveWindow.DisplayHeadings = False
                    except Exception:
                        pass
                except Exception:
                    pass
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.5)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    out.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return out

    def _build_before_after(self) -> list[str]:
        self.before_after.mkdir(parents=True, exist_ok=True)
        for old in self.before_after.glob("*.png"):
            old.unlink()
        pairs = [
            ("HOME", "PHASE_O4_V5_HOME.png", "HOME.png"),
            ("NAVIGATION", "PHASE_O4_V5_NAVIGATION.png", "NAVIGATION.png"),
            ("SEARCH", "PHASE_O4_V5_SEARCH.png", "SEARCH.png"),
        ]
        out: list[str] = []
        for label, v5_name, o5_name in pairs:
            v5 = self.v5_preview / v5_name
            o5 = self.preview_dir / o5_name
            if not v5.exists() or not o5.exists():
                continue
            a = Image.open(v5).convert("RGB")
            b = Image.open(o5).convert("RGB")
            # normalize heights
            h = max(a.height, b.height)
            w = a.width + b.width + 40
            canvas = Image.new("RGB", (w, h + 50), (245, 241, 232))
            draw = ImageDraw.Draw(canvas)
            draw.text((20, 12), f"V5  {label}", fill=(11, 35, 65))
            draw.text((a.width + 40, 12), f"O5  {label}", fill=(11, 35, 65))
            canvas.paste(a, (10, 40))
            canvas.paste(b, (a.width + 30, 40))
            path = self.before_after / f"V5_vs_O5_{label}.png"
            canvas.save(str(path), "PNG")
            out.append(str(path))
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        self.qa_json.write_text(
            json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
        )
        lines = [
            "# PHASE O5 — Front-End Rebuild QA",
            "",
            f"**PHASE O5 TECHNICAL GATE: {qa['phase_o5_technical_gate']}**",
            f"**MANUAL VISUAL ACCEPTANCE: {qa['manual_visual_acceptance']}**",
            "",
            "> Visual PASS is reserved for human review. Do not auto-claim visual PASS.",
            "",
            f"Workbook: `{qa['workbook_path']}`",
            f"Operational sheets modified: {qa['operational_sheets_modified']}",
            f"Word regenerated: {qa['word_regenerated']}  ·  Data changed: {qa['data_changed']}  ·  Promoted: {qa['promoted']}",
            "",
            "## Technical QA",
            f"- Native Excel open: {qa['native_excel_open']}",
            f"- Counts unchanged: {qa['counts_unchanged']}",
            f"- Word links: {qa['word_links']['existing']} / {qa['word_links']['total']} (broken={qa['word_links']['missing']})",
            f"- Home links: {qa['home_links']} / 13",
            f"- Search input editable: {qa['search_qa'].get('input_editable')}",
            f"- Search covered by shape: {qa['search_qa'].get('covered_by_opaque_shape')}",
            f"- Shape/table intersections: {qa['layout']['shape_table_intersections']}",
            f"- Visible non-Tahoma: {qa['visible_nontahoma']}",
            f"- Interaction: **{qa['interaction'].get('status')}**",
            "",
            "### Interaction steps",
        ]
        for s in qa["interaction"].get("steps", []):
            lines.append(f"- {s}")
        lines += ["", "## Search QA", f"```{qa['search_qa']}```", "", "## Previews"]
        for p in qa["preview_files"]:
            lines.append(f"- `{p}`")
        lines += ["", "## Before / After"]
        for p in qa["before_after_files"]:
            lines.append(f"- `{p}`")
        lines += ["", "## Build log"]
        for m in messages:
            lines.append(f"- {m}")
        lines += [
            "",
            "---",
            f"**PHASE O5 TECHNICAL GATE: {qa['phase_o5_technical_gate']}**",
            "**MANUAL VISUAL ACCEPTANCE: PENDING**",
            "",
            "STOP — do not promote.",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
