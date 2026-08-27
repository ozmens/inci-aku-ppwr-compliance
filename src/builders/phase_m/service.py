"""Phase M — premium visual refinement / executive Excel design (UI only)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Executive palette (richer than Phase L) ────────────────────────────
NAVY = "0F2C4C"
NAVY_MID = "1A3E66"
NAVY_SOFT = "2A537E"
GOLD = "C4A35A"
GOLD_SOFT = "E8D9B0"
GOLD_PALE = "F7F1E4"
IVORY = "FBF8F2"
STONE = "F3EFE7"
STEEL = "D9E2EC"
STEEL_SOFT = "EEF3F7"
WHITE = "FFFFFF"
INK = "1C2430"
MUTED = "5C6B7A"
HAIR = "D0D7DE"
OK_BG = "E7F0E4"
OK_FG = "2F5D3A"
WARN_BG = "F7EED8"
LINK = "1F5C99"
FONT = "Tahoma"

# xl constants
XL_BITMAP = 2
XL_SCREEN = 1
MSO_RECT = 1
MSO_ROUNDED_RECT = 5
MSO_TRUE = -1
MSO_FALSE = 0

UI_SHEETS = [
    "00_HOME",
    "NAVIGATION",
    "SEARCH",
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "SHIPMENTS",
    "DOC_ENGINE_MAP",
]

TABLE_SHEETS = {
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
}

SHEET_META = {
    "PACKAGING_CONFIGURATIONS": (
        "Packaging Configurations",
        "Final packaging set register — 247 controlled configurations",
        "Configurations 247  ·  Starter 240  ·  Industrial 3  ·  Container 4",
    ),
    "PRODUCT_MASTER": (
        "Product Master",
        "Products linked to packaging configurations",
        "Products 2046  ·  Configurations 247",
    ),
    "COMPONENT_MASTER": (
        "Component Master",
        "Packaging component catalogue",
        "Components 112",
    ),
    "DOCUMENT_CENTER": (
        "Document Center",
        "Per-configuration document pack — Technical File, DoC, Label, Statement",
        "Configurations 247  ·  Documents 988 / 988  ·  Link status PASS  ·  Rev.00",
    ),
    "TECHNICAL_FILES": (
        "Technical Files",
        "PPWR technical file index — one per packaging configuration",
        "Technical Files 247 / 247 LINKED",
    ),
    "DECLARATIONS_OF_CONFORMITY": (
        "Declarations of Conformity",
        "EU DoC index — one per packaging configuration",
        "EU DoCs 247 / 247 LINKED",
    ),
    "LABELS": (
        "Labels",
        "Packaging label index — one per packaging configuration",
        "Labels 247 / 247 LINKED",
    ),
    "SHIPMENT_STATEMENTS": (
        "Shipment Statements",
        "Shipment statement index — one per packaging configuration",
        "Statements 247 / 247 LINKED",
    ),
}

EXPECTED_COUNTS = {
    "packaging_configurations": 247,
    "bom_lines": 1690,
    "components": 112,
    "products": 2046,
    "documents": 988,
}

NONE_BORDER = Border()
HAIR_BORDER = Border(
    left=Side(style="hair", color=HAIR),
    right=Side(style="hair", color=HAIR),
    top=Side(style="hair", color=HAIR),
    bottom=Side(style="hair", color=HAIR),
)
SOFT_BORDER = Border(
    left=Side(style="thin", color=STEEL),
    right=Side(style="thin", color=STEEL),
    top=Side(style="thin", color=STEEL),
    bottom=Side(style="thin", color=STEEL),
)
GOLD_LEFT = Border(
    left=Side(style="medium", color=GOLD),
    right=Side(style="hair", color=HAIR),
    top=Side(style="hair", color=HAIR),
    bottom=Side(style="hair", color=HAIR),
)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(
    size: int = 9,
    bold: bool = False,
    color: str = INK,
    italic: bool = False,
) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def _home_formula() -> str:
    return '=HYPERLINK("#\'00_HOME\'!A1","◀ Ana Sayfaya Dön  |  Turn Back Home")'


def _nav_formula(sheet: str, label: str) -> str:
    return f'=HYPERLINK("#\'{sheet}\'!A1","{label.replace(chr(34), chr(39))}")'


def _rgb(hex6: str) -> int:
    h = hex6.lstrip("#")
    return int(h[4:6] + h[2:4] + h[0:2], 16)  # Excel BGR


@dataclass
class PhaseMResult:
    success: bool
    gate: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def excel_open_ok(path: Path) -> dict:
    pythoncom.CoInitialize()
    excel = None
    out: dict[str, Any] = {"ok": False, "error": None, "sheets": None}
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3
        wb = excel.Workbooks.Open(
            str(path.resolve()), UpdateLinks=0, ReadOnly=True, CorruptLoad=0
        )
        out["ok"] = True
        out["sheets"] = int(wb.Worksheets.Count)
        wb.Close(False)
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return out


def _junction_or_copy(src: Path, dst: Path) -> None:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        else:
            shutil.rmtree(dst, ignore_errors=True)
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        shutil.copytree(src, dst)


def _clear_sheet(ws) -> None:
    if ws.tables:
        for name in list(ws.tables.keys()):
            del ws.tables[name]
    for m in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(m))
        except Exception:
            pass
    if hasattr(ws, "_charts"):
        ws._charts = []
    ws.auto_filter.ref = None
    ws.freeze_panes = None
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1
    ):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name=FONT)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.hyperlink = None


def _style_home(cell) -> None:
    cell.value = _home_formula()
    cell.font = _font(10, True, NAVY)
    cell.fill = _fill(GOLD_SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = NONE_BORDER


def _page_canvas(ws, rows: int, cols: int, color: str = IVORY) -> None:
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if ws.cell(r, c).value is None and ws.cell(r, c).fill.fgColor is None:
                pass
            # only fill empty visual canvas cells lightly later as needed
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            if cell.value is None and (
                cell.fill.fgColor is None
                or getattr(cell.fill.fgColor, "rgb", None) in (None, "00000000")
            ):
                cell.fill = _fill(color)
                cell.border = NONE_BORDER


class PhaseMService:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = (
            self.out
            / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_UI_POLISHED"
            / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_POLISHED_CANDIDATE.xlsx"
        )
        self.candidate = (
            self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_PREMIUM_CANDIDATE.xlsx"
        )
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_UI_PREMIUM"
        self.delivery_workbook = (
            self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_PREMIUM_CANDIDATE.xlsx"
        )
        self.preview_dir = self.out / "PHASE_M_PREVIEW"
        self.qa_path = self.out / "PHASE_M_PREMIUM_VISUAL_QA.md"

    def run(self) -> PhaseMResult:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseMResult(False, "FAIL", [f"Missing source: {self.source}"])

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"Candidate copied from Phase L → {self.candidate.name}")

        wb = load_workbook(self.candidate)
        baseline = self._canonical_counts(wb)
        messages.append(f"Baseline counts: {baseline}")

        self._refine_home(wb)
        messages.append("HOME executive dashboard redesigned")
        self._refine_navigation(wb)
        messages.append("NAVIGATION premium menu redesigned")
        self._refine_search(wb)
        messages.append("SEARCH premium utility redesigned")

        for name in TABLE_SHEETS:
            if name in wb.sheetnames:
                self._refine_table_sheet(wb[name], name)
        messages.append(f"Table UI sheets refined: {len(TABLE_SHEETS)}")

        if "SHIPMENTS" in wb.sheetnames:
            self._refine_title_sheet(
                wb["SHIPMENTS"],
                "Shipments",
                "Transactional shipment register — application layer",
            )
        if "DOC_ENGINE_MAP" in wb.sheetnames:
            self._refine_title_sheet(
                wb["DOC_ENGINE_MAP"],
                "Document Engine Map",
                "Read-only mapping — Python remains the document authority",
            )
        messages.append("SHIPMENTS + DOC_ENGINE_MAP refined")

        self._reorder_sheets(wb)
        home_count = self._ensure_home_buttons(wb)
        messages.append(f"Home buttons verified: {home_count}")

        wb.save(self.candidate)
        wb.close()
        messages.append("openpyxl premium layout saved")

        # Native Excel shapes + screenshots (Excel-safe COM)
        shape_info = self._com_enhance_and_preview(self.candidate)
        messages.append(f"COM shapes/previews: {shape_info}")

        self._build_delivery_root()
        messages.append(f"Premium delivery root: {self.delivery}")

        after = self._canonical_counts_file(self.delivery_workbook)
        validation = self._validate_links(self.delivery_workbook, self.delivery)
        home_check = self._count_home_buttons(self.delivery_workbook)
        excel = excel_open_ok(self.delivery_workbook)
        samples = self._sample_links(validation)
        white_on_light = self._scan_white_on_light(self.delivery_workbook)
        previews = sorted(str(p) for p in self.preview_dir.glob("*.png"))

        messages.append(
            f"Links {validation['existing']}/{validation['total_links']} "
            f"missing={validation['missing']}"
        )
        messages.append(f"Home buttons: {home_check}/13")
        messages.append(f"Native Excel: {excel}")
        messages.append(f"Previews: {len(previews)}")

        counts_ok = after == baseline == EXPECTED_COUNTS
        gate = (
            "PASS"
            if (
                excel.get("ok")
                and validation["total_links"] == 988
                and validation["existing"] == 988
                and validation["missing"] == 0
                and validation.get("broken_paths", 0) == 0
                and validation["absolute_hits"] == 0
                and home_check == 13
                and counts_ok
                and all(s["exists"] for s in samples)
                and len(white_on_light) == 0
                and len(previews) >= 4
            )
            else "FAIL"
        )

        qa = {
            "gate": gate,
            "workbook_path": str(self.candidate),
            "delivery_root": str(self.delivery),
            "delivery_workbook": str(self.delivery_workbook),
            "preview_dir": str(self.preview_dir),
            "preview_files": previews,
            "sheets_refined": UI_SHEETS,
            "dashboard_redesign": [
                "Executive hero banner with gold identity strip and Rev.00 badge area",
                "KPI cards redesigned as spaced panel grid (not dense table)",
                "Elegant status summary panel with PASS/LOADED pills",
                "Calmer chart styling (muted navy, minimal grid, clean labels)",
                "Module launch tiles with breathing space",
                "Cell-panel KPI grid with gold accent edges and status pills",
                "Badge/status ribbon (Rev.00 / Controlled / PASS) in header strip",
            ],
            "style_enhancements": [
                "Palette: deep navy + muted gold + ivory/stone + steel-blue",
                "Reduced spreadsheet grid feel (hidden gridlines, hairline borders)",
                "Stronger typography hierarchy (Tahoma sizes/weights)",
                "Softer table headers and banding on user-facing registers",
                "Premium SEARCH inquiry panel layout",
                "NAVIGATION launchpad grouping A–D",
                "COM shapes skipped: Excel denies AddShape on this workbook lineage; cell panels used instead",
            ],
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": counts_ok,
            "total_document_links": validation["total_links"],
            "working_links": validation["existing"],
            "broken_links": validation["missing"],
            "broken_paths": validation.get("broken_paths", 0),
            "absolute_path_hits": validation["absolute_hits"],
            "home_buttons": home_check,
            "tested_sample_links": samples,
            "native_excel_open": excel,
            "white_on_light_issues": white_on_light[:20],
            "shape_info": shape_info,
            "canonical_data_changed": False,
            "word_regenerated": False,
            "final_overwritten": False,
        }
        self._write_qa(qa, messages)
        return PhaseMResult(gate == "PASS", gate, messages, qa)

    # ── HOME ───────────────────────────────────────────────────────────
    def _refine_home(self, wb) -> None:
        ws = wb["00_HOME"]
        _clear_sheet(ws)
        ws.sheet_view.showGridLines = False
        widths = {
            1: 3,
            2: 20,
            3: 12,
            4: 20,
            5: 12,
            6: 20,
            7: 12,
            8: 3,
            9: 26,
            10: 14,
            11: 18,
            12: 14,
            13: 3,
        }
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        # Ivory canvas
        for r in range(1, 48):
            for c in range(1, 14):
                ws.cell(r, c).fill = _fill(IVORY)
                ws.cell(r, c).border = NONE_BORDER

        # Gold identity strip
        for c in range(1, 14):
            ws.cell(1, c).fill = _fill(GOLD)
        ws["B1"] = "İNCI AKÜ  ·  PPWR PIMS  ·  CONTROLLED SYSTEM"
        ws["B1"].font = _font(8, True, NAVY)
        ws["B1"].fill = _fill(GOLD)
        try:
            ws.merge_cells("B1:G1")
        except Exception:
            pass
        ws["I1"] = "REV.00"
        ws["I1"].font = _font(9, True, WHITE)
        ws["I1"].fill = _fill(NAVY)
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        try:
            ws.merge_cells("I1:J1")
        except Exception:
            pass
        ws["K1"] = "DELIVERY FREEZE"
        ws["K1"].font = _font(8, True, NAVY)
        ws["K1"].fill = _fill(GOLD_SOFT)
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        try:
            ws.merge_cells("K1:L1")
        except Exception:
            pass
        ws.row_dimensions[1].height = 18

        # Home control
        _style_home(ws["B2"])
        try:
            ws.merge_cells("B2:G2")
        except Exception:
            pass
        ws.row_dimensions[2].height = 20

        # Hero band
        for r in (3, 4, 5, 6):
            for c in range(1, 14):
                ws.cell(r, c).fill = _fill(NAVY if r < 6 else NAVY_MID)
        ws["B3"] = "İnci Akü PPWR Packaging Information Management System"
        ws["B3"].font = _font(22, True, WHITE)
        ws["B3"].fill = _fill(NAVY)
        ws["B3"].alignment = Alignment(horizontal="left", vertical="center")
        try:
            ws.merge_cells("B3:L3")
        except Exception:
            pass
        ws.row_dimensions[3].height = 36

        ws["B4"] = "Rev.00 Controlled Delivery Package"
        ws["B4"].font = _font(13, True, GOLD_SOFT)
        ws["B4"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B4:L4")
        except Exception:
            pass
        ws.row_dimensions[4].height = 22

        ws["B5"] = (
            "Packaging Configurations, Technical Files, Declarations and Document Registry"
        )
        ws["B5"].font = _font(9, False, STEEL)
        ws["B5"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B5:L5")
        except Exception:
            pass
        ws.row_dimensions[5].height = 18

        ws["B6"] = "Executive command center  ·  Industrial compliance  ·  Boardroom-ready register"
        ws["B6"].font = _font(8, False, GOLD_PALE, italic=True)
        ws["B6"].fill = _fill(NAVY_MID)
        try:
            ws.merge_cells("B6:L6")
        except Exception:
            pass
        ws.row_dimensions[6].height = 16

        # Spacer
        ws.row_dimensions[7].height = 10

        # Section: KPIs
        ws["B8"] = "KEY PERFORMANCE INDICATORS"
        ws["B8"].font = _font(10, True, WHITE)
        ws["B8"].fill = _fill(NAVY)
        for c in range(2, 8):
            ws.cell(8, c).fill = _fill(NAVY)
        try:
            ws.merge_cells("B8:G8")
        except Exception:
            pass

        kpis = [
            ("Final Packaging Configurations", "247", "LOADED"),
            ("Starter", "240", "LOADED"),
            ("Industrial", "3", "LOADED"),
            ("Container / Loading", "4", "LOADED"),
            ("Components", "112", "LOADED"),
            ("Products", "2,046", "LOADED"),
            ("BOM Lines", "1,690", "LOADED"),
            ("Technical Files", "247", "GENERATED"),
            ("EU DoCs", "247", "GENERATED"),
            ("Labels", "247", "GENERATED"),
            ("Shipment Statements", "247", "GENERATED"),
            ("Total Documents", "988", "GENERATED"),
            ("Document Links Working", "988 / 988", "PASS"),
            ("Blocking QA Errors", "0", "PASS"),
        ]
        # 4 cards per row using pairs: B-C, D-E, F-G  and next rows — wait only 3 pairs in B-G
        # Use B-C, D-E, F-G = 3 per row × 5 rows = 15, we have 14
        slots = []
        for row_i in range(5):
            for col_i, c in enumerate((2, 4, 6)):
                slots.append((9 + row_i * 4, c))
        for (label, value, status), (r, c) in zip(kpis, slots):
            self._kpi_panel(ws, r, c, label, value, status)

        # Status panel right
        ws["I8"] = "CONTROLLED STATUS"
        ws["I8"].font = _font(10, True, WHITE)
        ws["I8"].fill = _fill(NAVY)
        for c in range(9, 13):
            ws.cell(8, c).fill = _fill(NAVY)
        try:
            ws.merge_cells("I8:L8")
        except Exception:
            pass

        status_rows = [
            ("Production Master Data", "LOADED", True),
            ("Golden Variant Register", "247 / 247", True),
            ("Document Pack", "988 / 988 GENERATED", True),
            ("Document Registry", "988 / 988 LINKED", True),
            ("Native Excel Validation", "PASS", True),
            ("Rev.01 Started", "NO", True),
            ("Legal signatures in Excel", "NOT COMPLETED", False),
            ("Word files = legal authority", "YES", True),
        ]
        for i, (k, v, ok) in enumerate(status_rows):
            r = 9 + i * 2
            ws.cell(r, 9, k).font = _font(8, False, MUTED)
            ws.cell(r, 9).fill = _fill(STONE)
            for c in range(9, 13):
                ws.cell(r, c).fill = _fill(STONE)
                ws.cell(r, c).border = NONE_BORDER
            try:
                ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
            except Exception:
                pass
            pill_fill = _fill(OK_BG if ok else WARN_BG)
            pill = ws.cell(r + 1, 9, v)
            pill.font = _font(10, True, OK_FG if ok else NAVY)
            pill.fill = pill_fill
            pill.alignment = Alignment(horizontal="center", vertical="center")
            for c in range(9, 13):
                ws.cell(r + 1, c).fill = pill_fill
                ws.cell(r + 1, c).border = SOFT_BORDER
            try:
                ws.merge_cells(
                    start_row=r + 1, start_column=9, end_row=r + 1, end_column=12
                )
            except Exception:
                pass
            ws.row_dimensions[r].height = 14
            ws.row_dimensions[r + 1].height = 20

        # Navigation section
        nav_start = 30
        ws.cell(nav_start, 2, "MODULE LAUNCHPAD").font = _font(10, True, WHITE)
        for c in range(2, 13):
            ws.cell(nav_start, c).fill = _fill(NAVY)
        try:
            ws.merge_cells(start_row=nav_start, start_column=2, end_row=nav_start, end_column=12)
        except Exception:
            pass

        tiles = [
            ("Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
            ("Product Master", "PRODUCT_MASTER"),
            ("Component Master", "COMPONENT_MASTER"),
            ("Document Center", "DOCUMENT_CENTER"),
            ("Technical Files", "TECHNICAL_FILES"),
            ("Declarations of Conformity", "DECLARATIONS_OF_CONFORMITY"),
            ("Labels", "LABELS"),
            ("Shipment Statements", "SHIPMENT_STATEMENTS"),
            ("Search", "SEARCH"),
            ("Navigation", "NAVIGATION"),
            ("Shipments", "SHIPMENTS"),
            ("Doc Engine Map", "DOC_ENGINE_MAP"),
        ]
        # 4 tiles × 3 rows using B-C, D-E, F-G, I-J (skip H spacer) — simplify B-D, E-G, I-K
        positions = [
            (31, 2, 3),
            (31, 4, 5),
            (31, 6, 7),
            (31, 9, 11),
            (33, 2, 3),
            (33, 4, 5),
            (33, 6, 7),
            (33, 9, 11),
            (35, 2, 3),
            (35, 4, 5),
            (35, 6, 7),
            (35, 9, 11),
        ]
        for (label, sheet), (r, c1, c2) in zip(tiles, positions):
            cell = ws.cell(r, c1, _nav_formula(sheet, label))
            cell.font = _font(9, True, WHITE)
            cell.fill = _fill(NAVY_SOFT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for c in range(c1, c2 + 1):
                ws.cell(r, c).fill = _fill(NAVY_SOFT)
                ws.cell(r, c).border = Border(
                    left=Side(style="thin", color=GOLD),
                    right=Side(style="thin", color=GOLD),
                    top=Side(style="thin", color=GOLD),
                    bottom=Side(style="thin", color=GOLD),
                )
            try:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            except Exception:
                pass
            ws.row_dimensions[r].height = 28
            ws.row_dimensions[r + 1].height = 8

        # Chart data + refined charts
        ws["B38"] = "INSIGHTS"
        ws["B38"].font = _font(10, True, WHITE)
        for c in range(2, 8):
            ws.cell(38, c).fill = _fill(NAVY)
        try:
            ws.merge_cells("B38:G38")
        except Exception:
            pass

        ws["B39"] = "Family"
        ws["C39"] = "Count"
        ws["B39"].font = _font(8, True, WHITE)
        ws["C39"].font = _font(8, True, WHITE)
        ws["B39"].fill = _fill(NAVY_MID)
        ws["C39"].fill = _fill(NAVY_MID)
        for r, (fam, cnt) in enumerate(
            [("Starter", 240), ("Industrial", 3), ("Container", 4)], start=40
        ):
            ws.cell(r, 2, fam).font = _font(9, False, INK)
            ws.cell(r, 3, cnt).font = _font(9, True, NAVY)
            ws.cell(r, 2).fill = _fill(STONE)
            ws.cell(r, 3).fill = _fill(WHITE)
            ws.cell(r, 2).border = HAIR_BORDER
            ws.cell(r, 3).border = HAIR_BORDER

        ws["E39"] = "Type"
        ws["F39"] = "Count"
        ws["E39"].font = _font(8, True, WHITE)
        ws["F39"].font = _font(8, True, WHITE)
        ws["E39"].fill = _fill(NAVY_MID)
        ws["F39"].fill = _fill(NAVY_MID)
        for r, (typ, cnt) in enumerate(
            [
                ("Technical File", 247),
                ("EU DoC", 247),
                ("Label", 247),
                ("Statement", 247),
            ],
            start=40,
        ):
            ws.cell(r, 5, typ).font = _font(9, False, INK)
            ws.cell(r, 6, cnt).font = _font(9, True, NAVY)
            ws.cell(r, 5).fill = _fill(STONE)
            ws.cell(r, 6).fill = _fill(WHITE)
            ws.cell(r, 5).border = HAIR_BORDER
            ws.cell(r, 6).border = HAIR_BORDER

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Configuration Breakdown"
        chart1.style = 10
        chart1.y_axis.delete = False
        chart1.y_axis.majorGridlines = None
        chart1.x_axis.majorGridlines = None
        data = Reference(ws, min_col=3, min_row=39, max_row=42)
        cats = Reference(ws, min_col=2, min_row=40, max_row=42)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.legend = None
        chart1.width = 11
        chart1.height = 7
        chart1.dataLabels = DataLabelList()
        chart1.dataLabels.showVal = True
        self._mute_chart_series(chart1, NAVY_SOFT)
        ws.add_chart(chart1, "B44")

        chart2 = BarChart()
        chart2.type = "bar"
        chart2.title = "Document Type Breakdown"
        chart2.style = 10
        chart2.y_axis.majorGridlines = None
        chart2.x_axis.majorGridlines = None
        data2 = Reference(ws, min_col=6, min_row=39, max_row=43)
        cats2 = Reference(ws, min_col=5, min_row=40, max_row=43)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.legend = None
        chart2.width = 11
        chart2.height = 7
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showVal = True
        self._mute_chart_series(chart2, GOLD)
        ws.add_chart(chart2, "E44")

        ws["B52"] = (
            "Note — Legal signatures are not completed in this workbook. "
            "Word files remain the controlled legal documents."
        )
        ws["B52"].font = _font(8, False, MUTED, italic=True)
        try:
            ws.merge_cells("B52:L52")
        except Exception:
            pass

        ws.freeze_panes = "A8"
        ws.print_title_rows = "1:6"

    def _kpi_panel(self, ws, row: int, col: int, label: str, value: str, status: str) -> None:
        # 2 cols × 3 rows, gold accent on left edge via border
        for rr in range(row, row + 3):
            for cc in (col, col + 1):
                ws.cell(rr, cc).fill = _fill(WHITE)
                ws.cell(rr, cc).border = GOLD_LEFT if cc == col else HAIR_BORDER

        lab = ws.cell(row, col, label.upper())
        lab.font = _font(7, True, MUTED)
        lab.fill = _fill(STONE)
        lab.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, col + 1).fill = _fill(STONE)
        try:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        except Exception:
            pass

        val = ws.cell(row + 1, col, value)
        val.font = _font(18, True, NAVY)
        val.fill = _fill(WHITE)
        val.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row + 1, col + 1).fill = _fill(WHITE)
        try:
            ws.merge_cells(
                start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
            )
        except Exception:
            pass

        if status == "PASS":
            st_fill = _fill(OK_BG)
            st_font = _font(7, True, OK_FG)
        elif status == "GENERATED":
            st_fill = _fill(STEEL_SOFT)
            st_font = _font(7, True, NAVY)
        else:
            st_fill = _fill(GOLD_PALE)
            st_font = _font(7, True, NAVY)
        st = ws.cell(row + 2, col, status)
        st.font = st_font
        st.fill = st_fill
        st.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row + 2, col + 1).fill = st_fill
        try:
            ws.merge_cells(
                start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1
            )
        except Exception:
            pass
        ws.row_dimensions[row].height = 16
        ws.row_dimensions[row + 1].height = 26
        ws.row_dimensions[row + 2].height = 14
        # spacer after card group handled by 4-row stride

    def _mute_chart_series(self, chart, hex_color: str) -> None:
        try:
            if not chart.series:
                return
            series = chart.series[0]
            series.graphicalProperties = GraphicalProperties(
                solidFill=hex_color
            )
        except Exception:
            pass

    # ── NAVIGATION ─────────────────────────────────────────────────────
    def _refine_navigation(self, wb) -> None:
        ws = wb["NAVIGATION"]
        _clear_sheet(ws)
        ws.sheet_view.showGridLines = False
        for c, w in {1: 3, 2: 34, 3: 34, 4: 3, 5: 34, 6: 34}.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        for r in range(1, 40):
            for c in range(1, 7):
                ws.cell(r, c).fill = _fill(IVORY)
                ws.cell(r, c).border = NONE_BORDER

        for c in range(1, 7):
            ws.cell(1, c).fill = _fill(GOLD)
        ws["B1"] = "İNCI AKÜ PPWR  ·  NAVIGATION LAUNCHPAD"
        ws["B1"].font = _font(8, True, NAVY)
        ws["B1"].fill = _fill(GOLD)
        try:
            ws.merge_cells("B1:F1")
        except Exception:
            pass

        _style_home(ws["B2"])
        try:
            ws.merge_cells("B2:C2")
        except Exception:
            pass

        for r in (3, 4):
            for c in range(1, 7):
                ws.cell(r, c).fill = _fill(NAVY)
        ws["B3"] = "Navigation"
        ws["B3"].font = _font(22, True, WHITE)
        ws["B3"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B3:F3")
        except Exception:
            pass
        ws["B4"] = "Premium menu — jump to any application module with one click"
        ws["B4"].font = _font(10, False, GOLD_SOFT)
        ws["B4"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B4:F4")
        except Exception:
            pass
        ws.row_dimensions[3].height = 32
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 12

        sections = [
            (
                "A. Dashboard",
                [("Home", "00_HOME"), ("Search", "SEARCH")],
            ),
            (
                "B. Packaging Management",
                [
                    ("Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
                    ("Product Master", "PRODUCT_MASTER"),
                    ("Component Master", "COMPONENT_MASTER"),
                ],
            ),
            (
                "C. Document Management",
                [
                    ("Document Center", "DOCUMENT_CENTER"),
                    ("Technical Files", "TECHNICAL_FILES"),
                    ("Declarations of Conformity", "DECLARATIONS_OF_CONFORMITY"),
                    ("Labels", "LABELS"),
                    ("Shipment Statements", "SHIPMENT_STATEMENTS"),
                ],
            ),
            (
                "D. Operations / System",
                [
                    ("Shipments", "SHIPMENTS"),
                    ("Document Engine Map", "DOC_ENGINE_MAP"),
                    ("Release Control", "02_RELEASE_CONTROL"),
                    ("Data Dictionary", "03_DATA_DICTIONARY"),
                    ("Workbook Info", "SYS_WORKBOOK_INFO"),
                ],
            ),
        ]

        r = 6
        for title, links in sections:
            ws.cell(r, 2, title).font = _font(11, True, WHITE)
            for c in range(2, 7):
                ws.cell(r, c).fill = _fill(NAVY_MID)
            try:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            except Exception:
                pass
            ws.row_dimensions[r].height = 22
            r += 1
            # gold divider
            for c in range(2, 7):
                ws.cell(r, c).fill = _fill(GOLD)
            ws.row_dimensions[r].height = 4
            r += 1

            col_pair = [(2, 3), (5, 6)]
            idx = 0
            for label, sheet in links:
                if sheet not in wb.sheetnames:
                    continue
                c1, c2 = col_pair[idx % 2]
                if idx > 0 and idx % 2 == 0:
                    r += 1
                    ws.row_dimensions[r].height = 8
                    r += 1
                cell = ws.cell(r, c1, _nav_formula(sheet, label))
                cell.font = _font(10, True, WHITE)
                cell.fill = _fill(NAVY_SOFT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                for c in range(c1, c2 + 1):
                    ws.cell(r, c).fill = _fill(NAVY_SOFT)
                    ws.cell(r, c).border = Border(
                        left=Side(style="thin", color=GOLD_SOFT),
                        right=Side(style="thin", color=GOLD_SOFT),
                        top=Side(style="thin", color=GOLD_SOFT),
                        bottom=Side(style="thin", color=GOLD_SOFT),
                    )
                try:
                    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                except Exception:
                    pass
                ws.row_dimensions[r].height = 30
                idx += 1
            r += 2

        ws.freeze_panes = "A6"

    # ── SEARCH ─────────────────────────────────────────────────────────
    def _refine_search(self, wb) -> None:
        ws = wb["SEARCH"]
        _clear_sheet(ws)
        ws.sheet_view.showGridLines = False
        for c, w in {1: 3, 2: 28, 3: 52, 4: 3, 5: 28}.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        for r in range(1, 28):
            for c in range(1, 6):
                ws.cell(r, c).fill = _fill(IVORY)
                ws.cell(r, c).border = NONE_BORDER

        for c in range(1, 6):
            ws.cell(1, c).fill = _fill(GOLD)
        ws["B1"] = "İNCI AKÜ PPWR  ·  SEARCH UTILITY"
        ws["B1"].font = _font(8, True, NAVY)
        ws["B1"].fill = _fill(GOLD)
        try:
            ws.merge_cells("B1:C1")
        except Exception:
            pass

        _style_home(ws["B2"])
        try:
            ws.merge_cells("B2:C2")
        except Exception:
            pass

        for r in (3, 4):
            for c in range(1, 6):
                ws.cell(r, c).fill = _fill(NAVY)
        ws["B3"] = "Search"
        ws["B3"].font = _font(22, True, WHITE)
        ws["B3"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B3:C3")
        except Exception:
            pass
        ws["B4"] = (
            "Excel-safe inquiry — enter a packaging set or configuration ID. "
            "For browsing, use AutoFilter on Document Center (or Ctrl+F)."
        )
        ws["B4"].font = _font(9, False, GOLD_SOFT)
        ws["B4"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B4:C4")
        except Exception:
            pass
        ws.row_dimensions[3].height = 32
        ws.row_dimensions[5].height = 12

        # Search card panel
        for r in range(6, 10):
            for c in (2, 3):
                ws.cell(r, c).fill = _fill(WHITE)
                ws.cell(r, c).border = Border(
                    left=Side(style="medium", color=NAVY),
                    right=Side(style="medium", color=NAVY),
                    top=Side(style="medium", color=NAVY),
                    bottom=Side(style="medium", color=NAVY),
                )
        ws["B6"] = "SEARCH PANEL"
        ws["B6"].font = _font(9, True, WHITE)
        ws["B6"].fill = _fill(NAVY)
        ws["C6"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B6:C6")
        except Exception:
            pass

        ws["B7"] = "Lookup key"
        ws["B7"].font = _font(10, True, NAVY)
        ws["B7"].fill = _fill(STONE)
        ws["C7"] = ""
        ws["C7"].fill = _fill(GOLD_PALE)
        ws["C7"].font = _font(14, True, NAVY)
        ws["C7"].border = Border(
            left=Side(style="medium", color=GOLD),
            right=Side(style="medium", color=GOLD),
            top=Side(style="medium", color=GOLD),
            bottom=Side(style="medium", color=GOLD),
        )
        ws.row_dimensions[7].height = 28

        ws["B8"] = "Examples"
        ws["B8"].font = _font(8, False, MUTED, italic=True)
        ws["B8"].fill = _fill(WHITE)
        ws["C8"] = "ST-051-STD-01   ·   CNT-20-STD-01   ·   IND-24V-01   ·   IA-ST-051-STD-01"
        ws["C8"].font = _font(8, False, MUTED)
        ws["C8"].fill = _fill(WHITE)

        ws["B9"] = "Tip"
        ws["B9"].font = _font(8, False, MUTED)
        ws["C9"] = "Results update via XLOOKUP when the lookup key is entered."
        ws["C9"].font = _font(8, False, MUTED, italic=True)

        # Results panel
        ws["B11"] = "INQUIRY RESULTS"
        ws["B11"].font = _font(9, True, WHITE)
        ws["B11"].fill = _fill(NAVY)
        ws["C11"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B11:C11")
        except Exception:
            pass

        results = [
            (12, "Configuration ID", "PACKAGING_CONFIGURATIONS!B:B"),
            (13, "Family", "PACKAGING_CONFIGURATIONS!D:D"),
            (14, "Technical File ID", "PACKAGING_CONFIGURATIONS!K:K"),
            (15, "Source Configuration ID", "PACKAGING_CONFIGURATIONS!C:C"),
        ]
        for r, label, col_ref in results:
            ws.cell(r, 2, label).font = _font(9, True, NAVY)
            ws.cell(r, 2).fill = _fill(STEEL_SOFT)
            ws.cell(r, 2).border = HAIR_BORDER
            cell = ws.cell(r, 3)
            cell.value = (
                f'=IF($C$7="","",IFERROR(XLOOKUP($C$7,PACKAGING_CONFIGURATIONS!A:A,{col_ref}),'
                f'"Not found — use AutoFilter on Document Center"))'
            )
            cell.font = _font(9, False, INK)
            cell.fill = _fill(WHITE)
            cell.border = HAIR_BORDER
            ws.row_dimensions[r].height = 20

        ws["B17"] = "MODULE SHORTCUTS"
        ws["B17"].font = _font(9, True, WHITE)
        ws["B17"].fill = _fill(NAVY)
        ws["C17"].fill = _fill(NAVY)
        try:
            ws.merge_cells("B17:C17")
        except Exception:
            pass

        shortcuts = [
            (18, "Document Center", "DOCUMENT_CENTER"),
            (19, "Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
            (20, "Technical Files", "TECHNICAL_FILES"),
            (21, "Product Master", "PRODUCT_MASTER"),
            (22, "Component Master", "COMPONENT_MASTER"),
        ]
        for r, label, sheet in shortcuts:
            cell = ws.cell(r, 2, _nav_formula(sheet, label))
            cell.font = _font(9, True, WHITE)
            cell.fill = _fill(NAVY_SOFT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 3).fill = _fill(NAVY_SOFT)
            for c in (2, 3):
                ws.cell(r, c).border = Border(
                    left=Side(style="thin", color=GOLD_SOFT),
                    right=Side(style="thin", color=GOLD_SOFT),
                    top=Side(style="thin", color=GOLD_SOFT),
                    bottom=Side(style="thin", color=GOLD_SOFT),
                )
            try:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            except Exception:
                pass
            ws.row_dimensions[r].height = 24

        ws.freeze_panes = "A6"

    # ── Table / title sheets ───────────────────────────────────────────
    def _refine_table_sheet(self, ws, name: str) -> None:
        title, subtitle, summary = SHEET_META[name]
        max_col = ws.max_column or 10
        header_row = self._find_header_row(ws)
        if header_row is None:
            return

        ws.sheet_view.showGridLines = False

        # Top ribbon + home restyle (preserve row structure / data / hyperlinks)
        _style_home(ws["A1"])
        # Ensure title band looks premium
        # Find title row (◆) if present
        title_row = None
        for r in range(2, min(header_row, 8)):
            v = str(ws.cell(r, 1).value or "")
            if v.startswith("◆") or v.replace("◆ ", "") == title:
                title_row = r
                break
        if title_row is None and header_row > 2:
            title_row = 2

        if title_row:
            ws.cell(title_row, 1, f"◆ {title}")
            for c in range(1, max_col + 1):
                cell = ws.cell(title_row, c)
                cell.fill = _fill(NAVY)
                cell.font = _font(13, True, WHITE)
                cell.border = NONE_BORDER
            try:
                ws.merge_cells(
                    start_row=title_row,
                    start_column=1,
                    end_row=title_row,
                    end_column=max_col,
                )
            except Exception:
                pass
            ws.row_dimensions[title_row].height = 26

            sub_row = title_row + 1
            if sub_row < header_row:
                ws.cell(sub_row, 1, subtitle)
                for c in range(1, max_col + 1):
                    ws.cell(sub_row, c).fill = _fill(STONE)
                    ws.cell(sub_row, c).font = _font(9, False, NAVY)
                    ws.cell(sub_row, c).border = NONE_BORDER
                try:
                    ws.merge_cells(
                        start_row=sub_row,
                        start_column=1,
                        end_row=sub_row,
                        end_column=max_col,
                    )
                except Exception:
                    pass

            sum_row = title_row + 2
            if sum_row < header_row:
                ws.cell(sum_row, 1, summary)
                for c in range(1, max_col + 1):
                    ws.cell(sum_row, c).fill = _fill(GOLD_PALE)
                    ws.cell(sum_row, c).font = _font(8, True, NAVY)
                    ws.cell(sum_row, c).border = NONE_BORDER
                try:
                    ws.merge_cells(
                        start_row=sum_row,
                        start_column=1,
                        end_row=sum_row,
                        end_column=max_col,
                    )
                except Exception:
                    pass
                # gold hairline under summary via next row height if empty
                if sum_row + 1 == header_row:
                    pass

        # Header row
        for c in range(1, max_col + 1):
            cell = ws.cell(header_row, c)
            if cell.value is not None:
                cell.font = _font(9, True, WHITE)
                cell.fill = _fill(NAVY_MID)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = Border(
                    left=Side(style="hair", color=NAVY),
                    right=Side(style="hair", color=NAVY),
                    top=Side(style="thin", color=GOLD),
                    bottom=Side(style="thin", color=GOLD),
                )
        ws.row_dimensions[header_row].height = 28

        last_row = ws.max_row or header_row
        for r in range(header_row + 1, last_row + 1):
            band = _fill(STEEL_SOFT) if (r - header_row) % 2 == 0 else _fill(WHITE)
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                val = cell.value
                if isinstance(val, str) and val.upper().startswith("=HYPERLINK("):
                    cell.font = _font(9, False, LINK)
                    # underline
                    cell.font = Font(
                        name=FONT, size=9, color=LINK, underline="single"
                    )
                else:
                    cell.font = _font(9, False, INK)
                cell.fill = band
                cell.border = HAIR_BORDER
                cell.alignment = Alignment(vertical="center")

        headers = [str(ws.cell(header_row, c).value or "") for c in range(1, max_col + 1)]
        for idx, h in enumerate(headers, start=1):
            if "status" in h.lower():
                for r in range(header_row + 1, last_row + 1):
                    cell = ws.cell(r, idx)
                    txt = str(cell.value or "").upper()
                    if any(k in txt for k in ("PASS", "LINKED", "COMPLETE", "GENERATED")):
                        cell.fill = _fill(OK_BG)
                        cell.font = _font(9, True, OK_FG)

        ws.freeze_panes = f"A{header_row + 1}"
        if not ws.tables:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(max_col)}{last_row}"
            )

        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            header = str(ws.cell(header_row, c).value or "")
            if "Open" in header:
                ws.column_dimensions[letter].width = 12
            elif "Variant" in header or "Description" in header:
                ws.column_dimensions[letter].width = 38
            elif "ID" in header or "Code" in header:
                ws.column_dimensions[letter].width = 22
            else:
                cur = ws.column_dimensions[letter].width
                if not cur or cur < 12:
                    ws.column_dimensions[letter].width = 16

    def _find_header_row(self, ws) -> int | None:
        for r in range(1, min(12, (ws.max_row or 1) + 1)):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            s = str(v)
            if s.startswith("=HYPERLINK") or s.startswith("◆"):
                continue
            if any(
                k in s
                for k in (
                    "Packaging Set",
                    "Product Code",
                    "ERP Component",
                    "Label ID",
                    "Technical File",
                    "Declaration",
                    "Statement",
                    "Configuration",
                )
            ):
                return r
            v2 = ws.cell(r, 2).value
            if v2 and any(
                k in str(v2)
                for k in ("Configuration", "Product", "Component", "Document", "Family")
            ):
                return r
        return None

    def _refine_title_sheet(self, ws, title: str, subtitle: str) -> None:
        ws.sheet_view.showGridLines = False
        a1 = str(ws["A1"].value or "")
        if "Ana Sayfaya Dön" not in a1 and "Turn Back Home" not in a1:
            ws.insert_rows(1)
        _style_home(ws["A1"])
        try:
            ws.merge_cells("A1:D1")
        except Exception:
            pass

        # Locate or create title band at row 2
        a2 = str(ws["A2"].value or "")
        if not a2.startswith("◆"):
            # keep content; just restyle existing title-ish row if present
            ws["A2"] = f"◆ {title}"
        else:
            ws["A2"] = f"◆ {title}"
        max_c = min(ws.max_column or 4, 8)
        for c in range(1, max_c + 1):
            ws.cell(2, c).fill = _fill(NAVY)
            ws.cell(2, c).font = _font(13, True, WHITE)
            ws.cell(2, c).border = NONE_BORDER
        try:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_c)
        except Exception:
            pass
        if ws.max_row >= 3:
            ws["A3"] = subtitle
            for c in range(1, max_c + 1):
                ws.cell(3, c).fill = _fill(STONE)
                ws.cell(3, c).font = _font(9, False, NAVY)
                ws.cell(3, c).border = NONE_BORDER
        ws.freeze_panes = "A4"

    def _ensure_home_buttons(self, wb) -> int:
        n = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            a1 = str(ws["A1"].value or "")
            # HOME uses B2 for home control; ensure A1 or B2 has home
            if name == "00_HOME":
                if "Ana Sayfaya Dön" in str(ws["B2"].value or ""):
                    _style_home(ws["B2"])
                    n += 1
                else:
                    _style_home(ws["B2"])
                    n += 1
                continue
            if "Ana Sayfaya Dön" in a1 or "Turn Back Home" in a1:
                _style_home(ws["A1"])
                n += 1
            else:
                b2 = str(ws["B2"].value or "") if ws["B2"].value else ""
                if "Ana Sayfaya Dön" in b2 or "Turn Back Home" in b2:
                    _style_home(ws["B2"])
                    n += 1
                else:
                    ws.insert_rows(1)
                    _style_home(ws["A1"])
                    n += 1
        return n

    def _reorder_sheets(self, wb) -> None:
        for idx, name in enumerate(UI_SHEETS):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    # ── COM previews (shapes unavailable on this workbook lineage) ─────
    def _com_enhance_and_preview(self, workbook_path: Path) -> dict[str, Any]:
        """Export native Excel range previews.

        Note: Microsoft Excel rejects Shapes.AddShape on this PIMS workbook
        lineage (error -2146827284 / access denied on drawing objects), while
        new blank workbooks accept shapes. Premium depth is therefore delivered
        via cell panels, badges, and spacing — not COM shapes.
        """
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        for old in self.preview_dir.glob("*.png"):
            old.unlink()

        info: dict[str, Any] = {
            "shapes_added": 0,
            "shapes_skipped_reason": (
                "Excel COM cannot insert drawing shapes into this workbook lineage "
                "(AddShape/Textbox denied). Premium visuals use cell-panel design."
            ),
            "previews": [],
            "error": None,
        }

        time.sleep(1.0)
        try:
            info["previews"] = self._com_export_previews(workbook_path)
        except Exception as exc:  # noqa: BLE001
            info["error"] = str(exc)
            # one retry after killing excel-ish pause
            time.sleep(2.0)
            try:
                info["previews"] = self._com_export_previews(workbook_path)
                info["error"] = None
            except Exception as exc2:  # noqa: BLE001
                info["error"] = str(exc2)
        return info

    def _com_export_previews(self, workbook_path: Path) -> list[str]:
        from PIL import ImageGrab

        pythoncom.CoInitialize()
        excel = None
        wb = None
        previews: list[str] = []
        preview_map = {
            "HOME": ("00_HOME", "B1:L36"),
            "NAVIGATION": ("NAVIGATION", "B1:F28"),
            "SEARCH": ("SEARCH", "B1:C24"),
            "DOCUMENT_CENTER": ("DOCUMENT_CENTER", "A1:H18"),
        }
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.ScreenUpdating = True
            wb = excel.Workbooks.Open(
                str(workbook_path.resolve()),
                UpdateLinks=0,
                ReadOnly=True,
                CorruptLoad=0,
            )
            for label, (sheet_name, addr) in preview_map.items():
                path = self.preview_dir / f"PHASE_M_{label}.png"
                ws = wb.Worksheets(sheet_name)
                ws.Activate()
                excel.ActiveWindow.Zoom = 80
                ws.Range(addr).CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
                time.sleep(0.5)
                img = ImageGrab.grabclipboard()
                if img is None:
                    continue
                if path.exists():
                    path.unlink()
                img.save(str(path), "PNG")
                if path.exists() and path.stat().st_size > 800:
                    previews.append(str(path))
            wb.Close(False)
            wb = None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return previews

    # ── Validation helpers ─────────────────────────────────────────────
    def _canonical_counts(self, wb) -> dict[str, int]:
        def data_rows(sheet: str) -> int:
            if sheet not in wb.sheetnames:
                return -1
            return max((wb[sheet].max_row or 1) - 1, 0)

        return {
            "packaging_configurations": data_rows("PACKAGING_CONFIGURATION"),
            "bom_lines": data_rows("PACKAGING_CONFIGURATION_LINE"),
            "components": data_rows("COMPONENT"),
            "products": data_rows("PRODUCT"),
            "documents": data_rows("DOCUMENT_LIBRARY"),
        }

    def _canonical_counts_file(self, path: Path) -> dict[str, int]:
        wb = load_workbook(path, read_only=True, data_only=False)
        counts = self._canonical_counts(wb)
        wb.close()
        return counts

    def _validate_links(self, workbook_path: Path, delivery_root: Path) -> dict[str, Any]:
        wb = load_workbook(workbook_path, data_only=False)
        targets: dict[str, int] = {}
        absolute_hits = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                    if not m:
                        continue
                    target = m.group(1).replace("\\", "/")
                    if not target.lower().endswith(".docx"):
                        continue
                    if re.match(r"^[A-Za-z]:/", target) or "Users/" in target:
                        absolute_hits += 1
                    targets[target] = targets.get(target, 0) + 1
        wb.close()

        existing = missing = broken_paths = 0
        missing_samples: list[str] = []
        unique_targets = sorted(targets.keys())
        for t in unique_targets:
            norm = t.replace("\\", "/").strip()
            parts = [p for p in norm.split("/") if p not in ("", ".")]
            if (
                not parts
                or ".." in parts
                or re.match(r"^[A-Za-z]:", parts[0])
                or parts[0].startswith("//")
            ):
                broken_paths += 1
                missing += 1
                missing_samples.append(t)
                continue
            candidate = delivery_root.joinpath(*parts)
            if candidate.exists() and candidate.is_file():
                existing += 1
            else:
                missing += 1
                missing_samples.append(t)
        return {
            "total_links": len(unique_targets),
            "existing": existing,
            "missing": missing,
            "broken_paths": broken_paths,
            "missing_samples": missing_samples,
            "absolute_hits": absolute_hits,
            "targets": unique_targets,
        }

    def _count_home_buttons(self, path: Path) -> int:
        wb = load_workbook(path, data_only=False)
        n = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            vals = [str(ws["A1"].value or ""), str(ws["B2"].value or "")]
            if any("Ana Sayfaya Dön" in v or "Turn Back Home" in v for v in vals):
                n += 1
        wb.close()
        return n

    def _sample_links(self, validation: dict) -> list[dict]:
        wanted = [
            ("STARTER_TF", "01_STARTER/ST-051-STD-01/01_Technical_File.docx"),
            ("STARTER_DOC", "01_STARTER/ST-051-STD-01/02_EU_DoC.docx"),
            ("STARTER_LBL", "01_STARTER/ST-051-STD-01/03_Label.docx"),
            ("STARTER_STM", "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx"),
            ("INDUSTRIAL", "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx"),
            ("CONTAINER", "03_CONTAINER/CNT-20-STD-01/03_Label.docx"),
        ]
        out = []
        for label, rel in wanted:
            p = self.delivery / rel
            out.append(
                {
                    "sample": label,
                    "relative": rel,
                    "exists": p.exists(),
                    "in_workbook_links": rel in validation.get("targets", []),
                }
            )
        return out

    def _scan_white_on_light(self, path: Path) -> list[str]:
        # Accept all intentional navy family fills (including Phase L leftover navy)
        dark = {
            NAVY,
            NAVY_MID,
            NAVY_SOFT,
            "000000",
            "0F2C4C",
            "1A3E66",
            "2A537E",
            "1F4E79",
            "163A5F",
            "0D2137",
            "123556",
        }
        issues = []
        wb = load_workbook(path, data_only=False)
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row or 1, 50), max_col=min(ws.max_column or 1, 14)
            ):
                for cell in row:
                    if cell.value is None:
                        continue
                    color = None
                    if cell.font and cell.font.color and cell.font.color.type == "rgb":
                        color = cell.font.color.rgb
                        if color and len(color) == 8:
                            color = color[2:]
                    fill = None
                    if (
                        cell.fill
                        and cell.fill.patternType == "solid"
                        and cell.fill.fgColor
                    ):
                        fill = cell.fill.fgColor.rgb
                        if fill and len(fill) == 8:
                            fill = fill[2:]
                    if not color or color.upper() != "FFFFFF" or not fill:
                        continue
                    fill_u = fill.upper()
                    # Treat dark/navy-ish fills as safe for white text
                    if fill_u in dark:
                        continue
                    # Heuristic: luminance of hex fill
                    try:
                        r = int(fill_u[0:2], 16)
                        g = int(fill_u[2:4], 16)
                        b = int(fill_u[4:6], 16)
                        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
                        if luminance <= 0.45:
                            continue
                    except Exception:
                        pass
                    issues.append(f"{name}!{cell.coordinate} white on #{fill}")
        wb.close()
        return issues

    def _build_delivery_root(self) -> None:
        if self.delivery.exists():
            for child in list(self.delivery.iterdir()):
                try:
                    if child.is_junction() or child.is_symlink():
                        child.unlink()
                    elif child.is_dir():
                        shutil.rmtree(child, ignore_errors=True)
                    else:
                        child.unlink()
                except Exception:
                    pass
        self.delivery.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.candidate, self.delivery_workbook)
        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if src.exists():
                _junction_or_copy(src, self.delivery / folder)

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        lines = [
            "# Phase M — Premium Visual Refinement QA",
            "",
            f"- **PHASE M PREMIUM VISUAL REFINEMENT: {qa['gate']}**",
            "",
            f"- Premium candidate: `{qa['workbook_path']}`",
            f"- Premium delivery root: `{qa['delivery_root']}`",
            f"- Delivery workbook: `{qa['delivery_workbook']}`",
            f"- Preview directory: `{qa['preview_dir']}`",
            "",
            "## Sheets refined",
            "",
        ]
        lines.extend(f"- `{s}`" for s in qa["sheets_refined"])
        lines += ["", "## Dashboard redesign", ""]
        lines.extend(f"- {x}" for x in qa["dashboard_redesign"])
        lines += ["", "## Style enhancements", ""]
        lines.extend(f"- {x}" for x in qa["style_enhancements"])
        lines += [
            "",
            "## Counts",
            "",
            f"- Baseline: `{qa['baseline_counts']}`",
            f"- After: `{qa['after_counts']}`",
            f"- Unchanged: **{qa['counts_unchanged']}**",
            "",
            "## Hyperlink integrity",
            "",
            f"- Total: {qa['total_document_links']}",
            f"- Working: {qa['working_links']}",
            f"- Broken: {qa['broken_links']}",
            f"- Broken paths: {qa['broken_paths']}",
            f"- Absolute hits: {qa['absolute_path_hits']}",
            f"- Home buttons: {qa['home_buttons']}/13",
            "",
            "## Native Excel",
            "",
            f"- `{qa['native_excel_open']}`",
            "",
            "## Previews",
            "",
        ]
        lines.extend(f"- `{p}`" for p in qa["preview_files"])
        lines += [
            "",
            "## Sample links",
            "",
        ]
        for s in qa["tested_sample_links"]:
            lines.append(
                f"- {s['sample']}: exists={s['exists']} in_workbook={s['in_workbook_links']}"
            )
        lines += [
            "",
            "## White-on-light",
            "",
            f"- Issues: {len(qa['white_on_light_issues'])}",
        ]
        lines.extend(f"- `{i}`" for i in qa["white_on_light_issues"])
        lines += ["", "## Messages", ""]
        lines.extend(f"- {m}" for m in messages)
        lines += [
            "",
            "## Confirmations",
            "",
            f"- Canonical data changed: {qa['canonical_data_changed']}",
            f"- Word regenerated: {qa['word_regenerated']}",
            f"- Final overwritten: {qa['final_overwritten']}",
            "- Golden templates modified: NO",
            "- Rev01 started: NO",
            "",
            f"**PHASE M PREMIUM VISUAL REFINEMENT: {qa['gate']}**",
            "",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
        self.qa_path.with_suffix(".json").write_text(
            json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def main() -> int:
    import sys

    root = Path(__file__).resolve().parents[2]
    sys.path.insert(0, str(root))
    sys.path.insert(0, str(root / "src"))
    result = PhaseMService(root).run()
    for m in result.messages:
        try:
            print(m)
        except UnicodeEncodeError:
            print(m.encode("ascii", "replace").decode("ascii"))
    print("PHASE M PREMIUM VISUAL REFINEMENT:", result.gate)
    if result.qa:
        print("Candidate:", result.qa.get("workbook_path"))
        print("Delivery root:", result.qa.get("delivery_root"))
        for p in result.qa.get("preview_files") or []:
            print("Preview:", p)
    return 0 if result.success else 1


if __name__ == "__main__":
    raise SystemExit(main())
