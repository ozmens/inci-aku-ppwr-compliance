"""
İnci Akü corporate dark-blue Excel UI theme (native openpyxl styles).

No VBA. Visual language only.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.table import TableStyleInfo

# ---------------------------------------------------------------------------
# Corporate palette
# ---------------------------------------------------------------------------

COLOR_PRIMARY = "0A2540"       # deep corporate navy
COLOR_PRIMARY_MID = "1B4F72"   # mid blue
COLOR_ACCENT = "2874A6"        # accent blue
COLOR_WHITE = "FFFFFF"
COLOR_INPUT = "FFF4CC"         # user input cells (Yellow)
COLOR_CALC = "D6EAF8"          # calculation cells (Blue)
COLOR_OUTPUT = "C8E6C9"        # output / results (Green)
COLOR_LOCKED = "E0E0E0"        # protected / read-only (Gray)
COLOR_NAV_BTN = "0A2540"
COLOR_NAV_BTN_FONT = "FFFFFF"
COLOR_DASH_BG = "F4F7FA"
COLOR_OK = "C6EFCE"
COLOR_OK_FONT = "006100"
COLOR_WARN = "FFEB9C"
COLOR_WARN_FONT = "9C5700"
COLOR_ERROR = "FFC7CE"
COLOR_ERROR_FONT = "9C0006"
COLOR_PENDING = "D9D9D9"
COLOR_PENDING_FONT = "595959"
COLOR_HIGH = "F5B7B1"
COLOR_MEDIUM = "F9E79F"
COLOR_LOW = "D5F5E3"

THIN = Border(
    left=Side(style="thin", color="BFCFD9"),
    right=Side(style="thin", color="BFCFD9"),
    top=Side(style="thin", color="BFCFD9"),
    bottom=Side(style="thin", color="BFCFD9"),
)

FILL_PRIMARY = PatternFill("solid", fgColor=COLOR_PRIMARY)
FILL_PRIMARY_MID = PatternFill("solid", fgColor=COLOR_PRIMARY_MID)
FILL_INPUT = PatternFill("solid", fgColor=COLOR_INPUT)
FILL_CALC = PatternFill("solid", fgColor=COLOR_CALC)
FILL_OUTPUT = PatternFill("solid", fgColor=COLOR_OUTPUT)
FILL_LOCKED = PatternFill("solid", fgColor=COLOR_LOCKED)
FILL_DASH = PatternFill("solid", fgColor=COLOR_DASH_BG)
FILL_NAV = PatternFill("solid", fgColor=COLOR_NAV_BTN)

FONT_HEADER = Font(name="Calibri", bold=True, color=COLOR_WHITE, size=11)
FONT_TITLE = Font(name="Calibri", bold=True, color=COLOR_PRIMARY, size=20)
FONT_SUBTITLE = Font(name="Calibri", bold=True, color=COLOR_PRIMARY_MID, size=14)
FONT_BODY = Font(name="Calibri", color=COLOR_PRIMARY, size=11)
FONT_NAV = Font(name="Calibri", bold=True, color=COLOR_NAV_BTN_FONT, size=11)
FONT_LINK = Font(name="Calibri", bold=True, color="0563C1", size=11, underline="single")
FONT_KPI = Font(name="Calibri", bold=True, color=COLOR_PRIMARY, size=16)
FONT_MUTED = Font(name="Calibri", color="5D6D7E", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

SHEET_PASSWORD_UI = "PIMS_UI"


def apply_header_row(ws, max_col: int, row: int = 1) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_PRIMARY
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN


def style_input_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.fill = FILL_INPUT
            cell.font = FONT_BODY
            cell.border = THIN
            cell.alignment = ALIGN_LEFT


def style_calc_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.fill = FILL_CALC
            cell.font = FONT_BODY
            cell.border = THIN
            cell.alignment = ALIGN_LEFT


def style_output_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.fill = FILL_OUTPUT
            cell.font = FONT_BODY
            cell.border = THIN
            cell.alignment = ALIGN_LEFT


def style_nav_button_cell(cell) -> None:
    cell.fill = FILL_NAV
    cell.font = FONT_NAV
    cell.alignment = ALIGN_CENTER
    cell.border = THIN
