"""
Native Excel UI layer (no VBA).

- Corporate dark-blue theme
- Dashboard + NAV hyperlink buttons
- SEARCH (input + FILTER formulas)
- Input / calculation / protected cell styling
- Table styles + AutoFilter
- Conditional formatting
- Sheet protection with unlocked input cells
"""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from models.registry import SchemaRegistry

from .engine_builder import all_engines
from .relationships import RELATIONSHIP_SHEET
from .sheet_builder import excel_col_letter
from .ui_theme import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    COLOR_ERROR,
    COLOR_ERROR_FONT,
    COLOR_HIGH,
    COLOR_LOW,
    COLOR_MEDIUM,
    COLOR_OK,
    COLOR_OK_FONT,
    COLOR_PENDING,
    COLOR_PENDING_FONT,
    COLOR_WARN,
    COLOR_WARN_FONT,
    FILL_DASH,
    FILL_INPUT,
    FILL_LOCKED,
    FILL_PRIMARY,
    FILL_PRIMARY_MID,
    FONT_BODY,
    FONT_HEADER,
    FONT_KPI,
    FONT_LINK,
    FONT_MUTED,
    FONT_SUBTITLE,
    FONT_TITLE,
    SHEET_PASSWORD_UI,
    TABLE_STYLE,
    THIN,
    apply_header_row,
    style_calc_range,
    style_input_range,
    style_nav_button_cell,
)

UI_SHEETS = ("DASHBOARD", "NAV", "SEARCH")

# Master/fact sheets users edit (input styling + unlocked body)
INPUT_SHEET_GROUPS = frozenset(
    {"LOOKUP", "ORGANIZATION", "MASTER", "CONFIGURATION", "COMMERCIAL", "LOGISTICS", "COMPLIANCE"}
)

# System sheets stay locked (not user input)
NON_INPUT_ENTITY = frozenset({"SYS_WORKBOOK_INFO", "SYS_PARAMETER"})


class UIBuilder:
    def __init__(self, workbook, registry: SchemaRegistry, settings) -> None:
        self.wb = workbook
        self.registry = registry
        self.settings = settings
        self.password = getattr(settings, "UI_SHEET_PASSWORD", SHEET_PASSWORD_UI)

    def apply(self) -> None:
        self._create_nav_sheet()
        self._create_search_sheet()
        self._create_dashboard_sheet()
        self._reorder_ui_sheets()
        self._style_entity_sheets()
        self._style_engine_sheets()
        self._style_relationship_sheet()
        self._apply_table_styles_and_filters()
        self._apply_conditional_formatting()
        self._protect_sheets()
        self._finalize_dashboard_kpis()

    # ------------------------------------------------------------------
    # UI sheets
    # ------------------------------------------------------------------

    def _create_dashboard_sheet(self) -> None:
        if "DASHBOARD" in self.wb.sheetnames:
            ws = self.wb["DASHBOARD"]
            # clear rebuild
            self.wb.remove(ws)
        ws = self.wb.create_sheet("DASHBOARD", 0)

        ws.sheet_view.showGridLines = False
        for r in range(1, 40):
            for c in range(1, 12):
                ws.cell(row=r, column=c).fill = FILL_DASH

        ws.merge_cells("B2:H2")
        title = ws["B2"]
        title.value = "İNCI AKÜ  |  PPWR Packaging Information Management System"
        title.font = FONT_TITLE
        title.alignment = ALIGN_LEFT

        ws.merge_cells("B3:H3")
        sub = ws["B3"]
        sub.value = (
            f"Workbook {self.settings.DEFAULT_OUTPUT_FILENAME}  ·  "
            f"Schema {self.settings.SCHEMA_VERSION}  ·  Native Excel UI (No VBA)"
        )
        sub.font = FONT_MUTED

        # Legend
        ws["B5"] = "CELL LEGEND"
        ws["B5"].font = FONT_SUBTITLE
        ws["B6"] = "User Input"
        ws["B6"].fill = FILL_INPUT
        ws["B6"].font = FONT_BODY
        ws["C6"] = "Editable master / transactional data"
        ws["C6"].font = FONT_MUTED
        ws["B7"] = "Calculation"
        from .ui_theme import FILL_CALC

        ws["B7"].fill = FILL_CALC
        ws["B7"].font = FONT_BODY
        ws["C7"] = "Engine formulas (locked)"
        ws["C7"].font = FONT_MUTED
        ws["B8"] = "Protected"
        ws["B8"].fill = FILL_LOCKED
        ws["B8"].font = FONT_BODY
        ws["C8"] = "Headers / system / technical"
        ws["C8"].font = FONT_MUTED

        # KPI headers
        ws["B10"] = "OPERATIONS SNAPSHOT"
        ws["B10"].font = FONT_SUBTITLE
        kpis = [
            ("B", 11, "Components", "=COUNTA(T_COMPONENT[COMPONENT_ID])"),
            ("D", 11, "Products", "=COUNTA(T_PRODUCT[PRODUCT_ID])"),
            ("F", 11, "Scenarios", "=COUNTA(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID])"),
            ("H", 11, "Shipments", "=COUNTA(T_SHIPMENT[SHIPMENT_ID])"),
            ("B", 14, "Pkg Configs", "=COUNTA(T_PACKAGING_CONFIGURATION[PACKAGING_CONFIGURATION_ID])"),
            ("D", 14, "Transport Configs", "=COUNTA(T_TRANSPORT_CONFIGURATION[TRANSPORT_CONFIGURATION_ID])"),
            ("F", 14, "Statements", "=COUNTA(T_STATEMENT[STATEMENT_ID])"),
            ("H", 14, "DoCs", "=COUNTA(T_DECLARATION_OF_CONFORMITY[DECLARATION_OF_CONFORMITY_ID])"),
        ]
        for col, row, label, formula in kpis:
            label_cell = ws[f"{col}{row}"]
            label_cell.value = label
            label_cell.fill = FILL_PRIMARY_MID
            label_cell.font = FONT_HEADER
            label_cell.alignment = ALIGN_CENTER
            label_cell.border = THIN
            val = ws[f"{col}{row + 1}"]
            val.value = formula
            val.font = FONT_KPI
            val.alignment = ALIGN_CENTER
            val.fill = FILL_LOCKED
            val.border = THIN
            val.protection = Protection(locked=True)

        # Validation summary
        ws["B17"] = "VALIDATION ENGINE"
        ws["B17"].font = FONT_SUBTITLE
        ws["B18"] = "OK"
        ws["C18"] = '=COUNTIF(T_ENG_VALIDATION[RESULT],"OK")'
        ws["B19"] = "ERROR"
        ws["C19"] = '=COUNTIF(T_ENG_VALIDATION[RESULT],"ERROR")'
        ws["B20"] = "WARN"
        ws["C20"] = '=COUNTIF(T_ENG_VALIDATION[RESULT],"WARN")'
        ws["B21"] = "PENDING"
        ws["C21"] = '=COUNTIF(T_ENG_VALIDATION[RESULT],"PENDING")'
        for r in range(18, 22):
            ws[f"B{r}"].font = FONT_BODY
            ws[f"B{r}"].fill = FILL_LOCKED
            ws[f"C{r}"].font = FONT_KPI
            ws[f"C{r}"].fill = FILL_LOCKED
            ws[f"C{r}"].alignment = ALIGN_CENTER
            ws[f"B{r}"].border = THIN
            ws[f"C{r}"].border = THIN

        # Material summaries
        ws["E17"] = "MATERIAL SUMMARIES (kg)"
        ws["E17"].font = FONT_SUBTITLE
        ws["E18"] = "Plastic"
        ws["F18"] = "=IFERROR(INDEX(T_ENG_PLASTIC_SUMMARY[TOTAL_MATERIAL_WEIGHT_KG],1),0)"
        ws["E19"] = "Paper"
        ws["F19"] = "=IFERROR(INDEX(T_ENG_PAPER_SUMMARY[TOTAL_MATERIAL_WEIGHT_KG],1),0)"
        ws["E20"] = "Wood"
        ws["F20"] = "=IFERROR(INDEX(T_ENG_WOOD_SUMMARY[TOTAL_MATERIAL_WEIGHT_KG],1),0)"
        for r in range(18, 21):
            ws[f"E{r}"].fill = FILL_LOCKED
            ws[f"E{r}"].font = FONT_BODY
            ws[f"F{r}"].fill = FILL_LOCKED
            ws[f"F{r}"].font = FONT_KPI
            ws[f"E{r}"].border = THIN
            ws[f"F{r}"].border = THIN

        # Quick navigation buttons on dashboard
        ws["B23"] = "QUICK NAVIGATION"
        ws["B23"].font = FONT_SUBTITLE
        quick = [
            ("B24", "Navigation Hub", "#NAV!A1"),
            ("D24", "Search", "#SEARCH!A1"),
            ("F24", "Validation", "#ENG_VALIDATION!A1"),
            ("H24", "Impact", "#ENG_IMPACT_ANALYSIS!A1"),
            ("B26", "Components", "#COMPONENT!A1"),
            ("D26", "Shipments", "#SHIPMENT!A1"),
            ("F26", "Statements", "#STATEMENT!A1"),
            ("H26", "Declarations", "#DECLARATION_OF_CONFORMITY!A1"),
        ]
        for addr, label, link in quick:
            cell = ws[addr]
            cell.value = label
            style_nav_button_cell(cell)
            cell.hyperlink = link

        ws["B28"] = (
            "Tips: Use table column filters on every data sheet · "
            "Search sheet for FILTER-based lookup · "
            "Yellow = input · Blue = calculation · Engines are protected"
        )
        ws["B28"].font = FONT_MUTED
        ws.merge_cells("B28:H28")

        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[12].height = 22
        ws.row_dimensions[15].height = 22

    def _create_nav_sheet(self) -> None:
        if "NAV" in self.wb.sheetnames:
            self.wb.remove(self.wb["NAV"])
        ws = self.wb.create_sheet("NAV")
        ws.sheet_view.showGridLines = False

        ws["A1"] = "PIMS NAVIGATION"
        ws["A1"].font = FONT_TITLE
        ws["A2"] = "Click a button to jump (native hyperlinks — no VBA)"
        ws["A2"].font = FONT_MUTED

        ws["A4"] = "HOME"
        ws["A4"].font = FONT_SUBTITLE
        self._nav_btn(ws, "A5", "Dashboard", "#DASHBOARD!A1")
        self._nav_btn(ws, "B5", "Search", "#SEARCH!A1")

        groups: dict[str, list[str]] = {
            "MASTERS": [],
            "CONFIGURATIONS": [],
            "OPERATIONS": [],
            "COMPLIANCE": [],
            "ENGINES": [],
            "LOOKUPS": [],
            "ORGANIZATION": [],
        }
        for table in self.registry.tables:
            if table.name in NON_INPUT_ENTITY:
                continue
            if table.group == "MASTER":
                groups["MASTERS"].append(table.name)
            elif table.group == "CONFIGURATION":
                groups["CONFIGURATIONS"].append(table.name)
            elif table.group in {"COMMERCIAL", "LOGISTICS"}:
                groups["OPERATIONS"].append(table.name)
            elif table.group == "COMPLIANCE":
                groups["COMPLIANCE"].append(table.name)
            elif table.group == "LOOKUP":
                groups["LOOKUPS"].append(table.name)
            elif table.group == "ORGANIZATION":
                groups["ORGANIZATION"].append(table.name)

        groups["ENGINES"] = [e.sheet_name for e in all_engines()]

        row = 7
        for group_name, sheets in groups.items():
            ws.cell(row=row, column=1, value=group_name).font = FONT_SUBTITLE
            row += 1
            col = 1
            for name in sheets:
                addr = f"{get_column_letter(col)}{row}"
                self._nav_btn(ws, addr, name, f"#{name}!A1")
                col += 1
                if col > 4:
                    col = 1
                    row += 1
            row += 2

        for c in range(1, 5):
            ws.column_dimensions[get_column_letter(c)].width = 28

    def _nav_btn(self, ws, addr: str, label: str, link: str) -> None:
        cell = ws[addr]
        cell.value = label
        style_nav_button_cell(cell)
        cell.hyperlink = link

    def _create_search_sheet(self) -> None:
        if "SEARCH" in self.wb.sheetnames:
            self.wb.remove(self.wb["SEARCH"])
        ws = self.wb.create_sheet("SEARCH")
        ws.sheet_view.showGridLines = False

        ws["A1"] = "PIMS SEARCH"
        ws["A1"].font = FONT_TITLE
        ws["A2"] = (
            "Type a search term in the yellow cell. Results use native FILTER "
            "(Microsoft 365). Table AutoFilters remain available on every data sheet."
        )
        ws["A2"].font = FONT_MUTED
        ws.merge_cells("A2:F2")

        ws["A4"] = "SEARCH_TERM"
        ws["A4"].fill = FILL_PRIMARY
        ws["A4"].font = FONT_HEADER
        ws["B4"] = ""
        ws["B4"].fill = FILL_INPUT
        ws["B4"].font = FONT_BODY
        ws["B4"].border = THIN
        ws["B4"].protection = Protection(locked=False)
        ws["C4"] = "← user input"
        ws["C4"].font = FONT_MUTED

        # Named range for search term
        if "NR_SEARCH_TERM" in self.wb.defined_names:
            del self.wb.defined_names["NR_SEARCH_TERM"]
        self.wb.defined_names.add(
            DefinedName(name="NR_SEARCH_TERM", attr_text="'SEARCH'!$B$4")
        )

        ws["A6"] = "COMPONENT MATCHES"
        ws["A6"].font = FONT_SUBTITLE
        ws["A7"] = (
            '=IF(NR_SEARCH_TERM="","(enter search term)",'
            'IFERROR(FILTER(HSTACK(T_COMPONENT[COMPONENT_ID],T_COMPONENT[COMPONENT_CODE],T_COMPONENT[COMPONENT_NAME]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMPONENT[COMPONENT_CODE]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMPONENT[COMPONENT_NAME]&"")))),"(no matches)"))'
        )
        ws["A7"].font = FONT_BODY

        ws["A9"] = "PRODUCT MATCHES"
        ws["A9"].font = FONT_SUBTITLE
        ws["A10"] = (
            '=IF(NR_SEARCH_TERM="","(enter search term)",'
            'IFERROR(FILTER(HSTACK(T_PRODUCT[PRODUCT_ID],T_PRODUCT[PRODUCT_CODE],T_PRODUCT[PRODUCT_NAME]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_PRODUCT[PRODUCT_CODE]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_PRODUCT[PRODUCT_NAME]&"")))),"(no matches)"))'
        )

        ws["A12"] = "SHIPMENT MATCHES"
        ws["A12"].font = FONT_SUBTITLE
        ws["A13"] = (
            '=IF(NR_SEARCH_TERM="","(enter search term)",'
            'IFERROR(FILTER(HSTACK(T_SHIPMENT[SHIPMENT_ID],T_SHIPMENT[SHIPMENT_NUMBER]),'
            'ISNUMBER(SEARCH(NR_SEARCH_TERM,T_SHIPMENT[SHIPMENT_NUMBER]&""))),"(no matches)"))'
        )

        ws["A15"] = "SCENARIO MATCHES"
        ws["A15"].font = FONT_SUBTITLE
        ws["A16"] = (
            '=IF(NR_SEARCH_TERM="","(enter search term)",'
            'IFERROR(FILTER(HSTACK(T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_ID],'
            'T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_CODE],T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_NAME]),'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_CODE]&"")))+'
            '(ISNUMBER(SEARCH(NR_SEARCH_TERM,T_COMMERCIAL_SCENARIO[COMMERCIAL_SCENARIO_NAME]&"")))),"(no matches)"))'
        )

        ws["A18"] = "ALTERNATE: Use dropdown filters on Excel Tables (filter arrows in header row)."
        ws["A18"].font = FONT_MUTED

        self._nav_btn(ws, "A20", "Dashboard", "#DASHBOARD!A1")
        self._nav_btn(ws, "B20", "Navigation", "#NAV!A1")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 40

    def _reorder_ui_sheets(self) -> None:
        # Desired front order: DASHBOARD, NAV, SEARCH
        for name in reversed(UI_SHEETS):
            if name in self.wb.sheetnames:
                current = self.wb.sheetnames.index(name)
                self.wb.move_sheet(name, offset=-current)

    # ------------------------------------------------------------------
    # Styling
    # ------------------------------------------------------------------

    def _style_entity_sheets(self) -> None:
        for table in self.registry.tables:
            ws = self.wb[table.name]
            max_col = len(table.columns)
            max_row = max(ws.max_row, 2)
            apply_header_row(ws, max_col, 1)

            if table.name in NON_INPUT_ENTITY:
                style_calc_range(ws, 2, max_row, 1, max_col)  # system = non-input look
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.protection = Protection(locked=True)
            elif table.group in INPUT_SHEET_GROUPS:
                style_input_range(ws, 2, max_row, 1, max_col)
                for cell in ws[1]:
                    cell.protection = Protection(locked=True)
                for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.protection = Protection(locked=False)
            else:
                for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.fill = FILL_LOCKED
                        cell.protection = Protection(locked=True)

            self._autosize(ws, max_col)
            # Home link
            ws.cell(row=1, column=max_col + 2).value = "Dashboard"
            ws.cell(row=1, column=max_col + 2).font = FONT_LINK
            ws.cell(row=1, column=max_col + 2).hyperlink = "#DASHBOARD!A1"

    def _style_engine_sheets(self) -> None:
        from .ui_theme import FILL_CALC

        for engine in all_engines():
            if engine.sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[engine.sheet_name]
            max_col = len(engine.headers)
            max_row = ws.max_row
            apply_header_row(ws, max_col, 1)
            style_calc_range(ws, 2, max_row, 1, max_col)
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.protection = Protection(locked=True)
            # ROW_INDEX column remains calc/locked (driver values)
            self._autosize(ws, max_col)
            ws.cell(row=1, column=max_col + 2).value = "Dashboard"
            ws.cell(row=1, column=max_col + 2).font = FONT_LINK
            ws.cell(row=1, column=max_col + 2).hyperlink = "#DASHBOARD!A1"

    def _style_relationship_sheet(self) -> None:
        if RELATIONSHIP_SHEET not in self.wb.sheetnames:
            return
        ws = self.wb[RELATIONSHIP_SHEET]
        max_col = ws.max_column
        max_row = ws.max_row
        apply_header_row(ws, max_col, 1)
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = FILL_LOCKED
                cell.font = FONT_BODY
                cell.border = THIN
                cell.protection = Protection(locked=True)

    def _apply_table_styles_and_filters(self) -> None:
        for ws in self.wb.worksheets:
            if ws.title in UI_SHEETS:
                continue
            for table in ws.tables.values():
                table.tableStyleInfo = TABLE_STYLE
                # Excel Tables include AutoFilter by default; keep ref aligned
                try:
                    from openpyxl.worksheet.filters import AutoFilter

                    table.autoFilter = AutoFilter(ref=table.ref)
                except Exception:
                    pass

    def _autosize(self, ws, max_col: int) -> None:
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            header = ws.cell(row=1, column=col).value
            width = min(max(len(str(header or "")) + 4, 12), 34)
            ws.column_dimensions[letter].width = width

    # ------------------------------------------------------------------
    # Conditional formatting
    # ------------------------------------------------------------------

    def _apply_conditional_formatting(self) -> None:
        fill_ok = PatternFill("solid", fgColor=COLOR_OK)
        font_ok = Font(color=COLOR_OK_FONT, bold=True)
        fill_err = PatternFill("solid", fgColor=COLOR_ERROR)
        font_err = Font(color=COLOR_ERROR_FONT, bold=True)
        fill_warn = PatternFill("solid", fgColor=COLOR_WARN)
        font_warn = Font(color=COLOR_WARN_FONT, bold=True)
        fill_pend = PatternFill("solid", fgColor=COLOR_PENDING)
        font_pend = Font(color=COLOR_PENDING_FONT, bold=True)

        # Validation RESULT column D
        if "ENG_VALIDATION" in self.wb.sheetnames:
            ws = self.wb["ENG_VALIDATION"]
            rng = f"D2:D{ws.max_row}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"OK"'], fill=fill_ok, font=font_ok)
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="equal", formula=['"ERROR"'], fill=fill_err, font=font_err),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="equal", formula=['"WARN"'], fill=fill_warn, font=font_warn),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal", formula=['"PENDING"'], fill=fill_pend, font=font_pend
                ),
            )

        # Impact level
        if "ENG_IMPACT_ANALYSIS" in self.wb.sheetnames:
            ws = self.wb["ENG_IMPACT_ANALYSIS"]
            # IMPACT_LEVEL is last header
            headers = [c.value for c in ws[1]]
            if "IMPACT_LEVEL" in headers:
                col = headers.index("IMPACT_LEVEL") + 1
                letter = get_column_letter(col)
                rng = f"{letter}2:{letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="equal",
                        formula=['"HIGH"'],
                        fill=PatternFill("solid", fgColor=COLOR_HIGH),
                    ),
                )
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="equal",
                        formula=['"MEDIUM"'],
                        fill=PatternFill("solid", fgColor=COLOR_MEDIUM),
                    ),
                )
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="equal",
                        formula=['"LOW"'],
                        fill=PatternFill("solid", fgColor=COLOR_LOW),
                    ),
                )

        for sheet_name, status_col in (
            ("ENG_TECHNICAL_FILE", "ENGINE_STATUS"),
            ("ENG_DECLARATION", "ENGINE_STATUS"),
            ("ENG_STATEMENT", "RECONCILE_STATUS"),
        ):
            if sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet_name]
            headers = [c.value for c in ws[1]]
            if status_col not in headers:
                continue
            col = headers.index(status_col) + 1
            letter = get_column_letter(col)
            rng = f"{letter}2:{letter}{ws.max_row}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"OK"'], fill=fill_ok, font=font_ok)
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal", formula=['"COMPLETE"'], fill=fill_ok, font=font_ok
                ),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal", formula=['"INCOMPLETE"'], fill=fill_warn, font=font_warn
                ),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal", formula=['"VARIANCE"'], fill=fill_err, font=font_err
                ),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal", formula=['"PENDING"'], fill=fill_pend, font=font_pend
                ),
            )

        # Dashboard validation counts highlight
        if "DASHBOARD" in self.wb.sheetnames:
            ws = self.wb["DASHBOARD"]
            ws.conditional_formatting.add(
                "C19",
                CellIsRule(
                    operator="greaterThan", formula=["0"], fill=fill_err, font=font_err
                ),
            )
            ws.conditional_formatting.add(
                "C18",
                CellIsRule(
                    operator="greaterThan", formula=["0"], fill=fill_ok, font=font_ok
                ),
            )

    def _finalize_dashboard_kpis(self) -> None:
        # Ensure dashboard calc cells locked
        if "DASHBOARD" not in self.wb.sheetnames:
            return
        ws = self.wb["DASHBOARD"]
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
            for cell in row:
                if cell.hyperlink:
                    cell.protection = Protection(locked=True)
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.protection = Protection(locked=True)
                else:
                    # keep decorative locked
                    if cell.protection is None or cell.coordinate != "B4":
                        cell.protection = Protection(locked=True)

    # ------------------------------------------------------------------
    # Protection
    # ------------------------------------------------------------------

    def _protect_sheets(self) -> None:
        # UI sheets
        for name in UI_SHEETS:
            if name not in self.wb.sheetnames:
                continue
            ws = self.wb[name]
            if name == "SEARCH":
                ws["B4"].protection = Protection(locked=False)
            ws.protection.sheet = True
            ws.protection.password = self.password
            ws.protection.autoFilter = True
            ws.protection.sort = True

        # Entity input sheets — allow filter while protecting structure
        for table in self.registry.tables:
            ws = self.wb[table.name]
            ws.protection.sheet = True
            ws.protection.password = self.password
            ws.protection.autoFilter = True
            ws.protection.sort = True

        # Engines fully locked
        for engine in all_engines():
            if engine.sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[engine.sheet_name]
            ws.protection.sheet = True
            ws.protection.password = self.password

        # Relationship remains hidden technical; ensure protected
        if RELATIONSHIP_SHEET in self.wb.sheetnames:
            ws = self.wb[RELATIONSHIP_SHEET]
            ws.protection.sheet = True
            if not ws.protection.password:
                ws.protection.password = getattr(
                    self.settings, "TECH_SHEET_PASSWORD", self.password
                )
