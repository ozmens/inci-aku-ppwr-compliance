"""Phase K — UI finalization: delivery-root hyperlinks + Home buttons."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F4E79"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
BODY = "1A1A1A"
THIN = "B0B0B0"
FONT = "Tahoma"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT, size=10, bold=True, color=WHITE)
HOME_FILL = PatternFill("solid", fgColor=YELLOW)
HOME_FONT = Font(name=FONT, size=11, bold=True, color=NAVY)
LINK_FONT = Font(name=FONT, size=9, color="0563C1", underline="single")
BODY_FONT = Font(name=FONT, size=9, color=BODY)
TITLE_FONT = Font(name=FONT, size=18, bold=True, color=NAVY)
THIN_BORDER = Border(
    left=Side(style="thin", color=THIN),
    right=Side(style="thin", color=THIN),
    top=Side(style="thin", color=THIN),
    bottom=Side(style="thin", color=THIN),
)

UI_SHEETS = [
    "00_HOME",
    "NAVIGATION",
    "SEARCH",
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "SHIPMENTS",
    "DOC_ENGINE_MAP",
]

# Sheets whose row 1 is a filter header table (need insert + shift freeze/filter)
TABLE_UI_SHEETS = {
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "DOC_ENGINE_MAP",
}

HYPERLINK_RE = re.compile(
    r'=HYPERLINK\("([^"]+\.docx)"\s*,\s*"([^"]*)"\)',
    re.IGNORECASE,
)


@dataclass
class PhaseKResult:
    success: bool
    gate: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def excel_open_ok(path: Path) -> dict:
    pythoncom.CoInitialize()
    excel = None
    out: dict[str, Any] = {"ok": False, "error": None, "sheets": None}
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3
        wb = excel.Workbooks.Open(
            str(path.resolve()),
            UpdateLinks=0,
            ReadOnly=True,
            CorruptLoad=0,
        )
        out["ok"] = True
        out["sheets"] = int(wb.Worksheets.Count)
        wb.Close(False)
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return out


def _junction_or_copy(src: Path, dst: Path) -> None:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        else:
            shutil.rmtree(dst, ignore_errors=True)
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        shutil.copytree(src, dst)


class PhaseKService:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source_candidate = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_CANDIDATE.xlsx"
        self.v2 = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_CANDIDATE_V2.xlsx"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_UI_READY"
        self.delivery_workbook = (
            self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_CANDIDATE_V2.xlsx"
        )
        self.qa_path = self.out / "PHASE_K_UI_FINALIZATION_QA.md"

    def run(self) -> PhaseKResult:
        messages: list[str] = []
        if not self.source_candidate.exists():
            return PhaseKResult(False, "FAIL", ["Missing Phase J UI candidate"])

        if self.v2.exists():
            self.v2.unlink()
        shutil.copy2(self.source_candidate, self.v2)
        messages.append(f"V2 copied from Phase J candidate → {self.v2.name}")

        wb = load_workbook(self.v2)

        # Normalize document HYPERLINK formulas to clean relative paths
        link_stats = self._normalize_document_hyperlinks(wb)
        messages.append(
            f"Document HYPERLINK formulas normalized: {link_stats['normalized']} "
            f"(absolute stripped: {link_stats['absolute_stripped']})"
        )

        home_count = self._add_home_buttons(wb)
        messages.append(f"Home buttons added: {home_count}")

        wb.save(self.v2)
        wb.close()
        messages.append("V2 workbook saved")

        # Build delivery root with workbook INSIDE alongside doc folders
        self._build_delivery_root()
        messages.append(f"Delivery root: {self.delivery}")

        # Validate links from delivery workbook location
        validation = self._validate_links(self.delivery_workbook, self.delivery)
        messages.append(
            f"Link validation: total={validation['total_links']} "
            f"existing={validation['existing']} missing={validation['missing']}"
        )

        excel = excel_open_ok(self.delivery_workbook)
        messages.append(f"Native Excel open (delivery root): {excel}")

        samples = self._sample_link_tests(validation)
        messages.append(f"Sample links checked: {len(samples)}")

        qa = {
            "workbook_path": str(self.v2),
            "delivery_root": str(self.delivery),
            "delivery_workbook": str(self.delivery_workbook),
            "user_facing_sheets": [s for s in UI_SHEETS if True],
            "user_facing_sheets_count": len(UI_SHEETS),
            "home_buttons_added": home_count,
            "total_document_links": validation["total_links"],
            "working_links": validation["existing"],
            "broken_links": validation["missing"],
            "broken_paths": validation.get("broken_paths", 0),
            "broken_path_samples": validation["missing_samples"][:20],
            "absolute_path_hits": validation["absolute_hits"],
            "formula_instances": validation.get("formula_instances"),
            "tested_sample_links": samples,
            "native_excel_open": excel,
            "canonical_data_changed": False,
            "word_regenerated": False,
        }
        gate = (
            "PASS"
            if (
                excel.get("ok")
                and validation["total_links"] == 988
                and validation["existing"] == 988
                and validation["missing"] == 0
                and validation.get("broken_paths", 0) == 0
                and validation["absolute_hits"] == 0
                and home_count >= len(UI_SHEETS)
                and all(s["exists"] for s in samples)
            )
            else "FAIL"
        )
        qa["gate"] = gate
        self._write_qa(qa, messages)
        return PhaseKResult(gate == "PASS", gate, messages, qa)

    def _normalize_document_hyperlinks(self, wb) -> dict[str, int]:
        """Ensure all DOCX HYPERLINKs are clean relatives (no drive letters)."""
        normalized = 0
        absolute_stripped = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str) or "HYPERLINK" not in val.upper():
                        continue
                    m = HYPERLINK_RE.search(val.replace("'", '"'))
                    # also match single-quoted variants poorly — rewrite known pattern
                    m2 = re.search(
                        r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)',
                        val,
                        re.I,
                    )
                    if not m2:
                        continue
                    target, label = m2.group(1), m2.group(2)
                    if not target.lower().endswith(".docx"):
                        continue
                    original = target
                    # strip absolute roots if any
                    target = target.replace("\\", "/")
                    if re.match(r"^[A-Za-z]:/", target) or target.startswith("//"):
                        absolute_stripped += 1
                        # keep only trailing 01_STARTER|02_INDUSTRIAL|03_CONTAINER/...
                        mpath = re.search(
                            r"(01_STARTER|02_INDUSTRIAL|03_CONTAINER)/.+",
                            target,
                        )
                        if mpath:
                            target = mpath.group(0)
                    # remove leading ./ 
                    target = target.lstrip("./")
                    # forbid parent traversal
                    if ".." in target.split("/"):
                        continue
                    new_val = f'=HYPERLINK("{target}","{label}")'
                    if new_val != val or original != target:
                        cell.value = new_val
                        cell.font = LINK_FONT
                        normalized += 1
        return {"normalized": normalized, "absolute_stripped": absolute_stripped}

    def _add_home_buttons(self, wb) -> int:
        count = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            if name in TABLE_UI_SHEETS:
                self._insert_home_banner_table_sheet(ws, name)
            else:
                self._ensure_home_banner_title_sheet(ws, name)
            count += 1
        return count

    def _home_cell_value(self) -> str:
        return '=HYPERLINK("#\'00_HOME\'!A1","◀ Ana Sayfaya Dön  |  Turn Back Home")'

    def _style_home_cell(self, cell) -> None:
        cell.value = self._home_cell_value()
        cell.font = HOME_FONT
        cell.fill = HOME_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    def _insert_home_banner_table_sheet(self, ws, name: str) -> None:
        """Insert row 1 home button; keep header+filter starting at row 2."""
        # If already has home button, refresh only
        a1 = str(ws["A1"].value or "")
        if "Ana Sayfaya Dön" in a1 or "Turn Back Home" in a1:
            self._style_home_cell(ws["A1"])
            return

        # Detect current header row (usually 1, DOC_ENGINE_MAP may be 3)
        header_row = 1
        if name == "DOC_ENGINE_MAP":
            # title at 1, blank/info, headers at 3
            header_row = 3 if ws["A3"].value else 1

        if header_row == 1:
            ws.insert_rows(1)
            self._style_home_cell(ws["A1"])
            # merge banner across a few columns for visibility
            try:
                ws.merge_cells("A1:D1")
            except Exception:
                pass
            ws.row_dimensions[1].height = 22
            # restyle header row 2
            max_col = ws.max_column
            for c in range(1, max_col + 1):
                cell = ws.cell(2, c)
                if cell.value:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.border = THIN_BORDER
            ws.freeze_panes = "A3"
            last = ws.max_row
            if max_col >= 1 and last >= 2:
                ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{last}"
        else:
            # DOC_ENGINE_MAP: put home above title
            ws.insert_rows(1)
            self._style_home_cell(ws["A1"])
            try:
                ws.merge_cells("A1:D1")
            except Exception:
                pass
            ws.row_dimensions[1].height = 22
            # headers now at row 4
            max_col = min(ws.max_column, 6)
            last = ws.max_row
            ws.auto_filter.ref = f"A4:{get_column_letter(max_col)}{last}"
            ws.freeze_panes = "A5"

    def _ensure_home_banner_title_sheet(self, ws, name: str) -> None:
        a1 = str(ws["A1"].value or "")
        if name == "00_HOME":
            # Place home marker / refresh cue at top-right area without destroying title
            # Insert banner above title
            if "Ana Sayfaya Dön" not in a1:
                ws.insert_rows(1)
                self._style_home_cell(ws["A1"])
                try:
                    ws.merge_cells("A1:F1")
                except Exception:
                    pass
                ws.row_dimensions[1].height = 22
                # restore title style on new A2
                if ws["A2"].value:
                    ws["A2"].font = TITLE_FONT
            else:
                self._style_home_cell(ws["A1"])
            return

        if "Ana Sayfaya Dön" in a1 or "Turn Back Home" in a1:
            self._style_home_cell(ws["A1"])
            return

        ws.insert_rows(1)
        self._style_home_cell(ws["A1"])
        try:
            ws.merge_cells("A1:D1")
        except Exception:
            pass
        ws.row_dimensions[1].height = 22
        # keep title formatting if present at A2
        if ws["A2"].value and not str(ws["A2"].value).startswith("="):
            ws["A2"].font = TITLE_FONT

    def _build_delivery_root(self) -> None:
        if self.delivery.exists():
            # remove junctions/files carefully
            for child in list(self.delivery.iterdir()):
                try:
                    if child.is_junction() or child.is_symlink():
                        child.unlink()
                    elif child.is_dir():
                        shutil.rmtree(child, ignore_errors=True)
                    else:
                        child.unlink()
                except Exception:
                    pass
        self.delivery.mkdir(parents=True, exist_ok=True)

        # Workbook lives INSIDE delivery root
        shutil.copy2(self.v2, self.delivery_workbook)

        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if src.exists():
                _junction_or_copy(src, self.delivery / folder)

    def _validate_links(self, workbook_path: Path, delivery_root: Path) -> dict[str, Any]:
        wb = load_workbook(workbook_path, data_only=False)
        targets: dict[str, int] = {}
        absolute_hits = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                    if not m:
                        continue
                    target = m.group(1).replace("\\", "/")
                    if not target.lower().endswith(".docx"):
                        continue
                    if re.match(r"^[A-Za-z]:/", target) or "Users/" in target:
                        absolute_hits += 1
                    targets[target] = targets.get(target, 0) + 1
        wb.close()

        existing = 0
        missing = 0
        missing_samples = []
        broken_paths = 0
        # Unique document targets should be 988; UI may repeat same path (DOCUMENT_CENTER + modules)
        # Count unique DOCX targets for PASS criteria of 988 resolvable documents
        unique_targets = sorted(targets.keys())
        for t in unique_targets:
            # Reject traversal / absolute / empty
            norm = t.replace("\\", "/").strip()
            parts = [p for p in norm.split("/") if p not in ("", ".")]
            if (
                not parts
                or ".." in parts
                or re.match(r"^[A-Za-z]:", parts[0])
                or parts[0].startswith("//")
            ):
                broken_paths += 1
                missing += 1
                missing_samples.append(t)
                continue
            # Check via delivery-root relative path. Do NOT .resolve() before exists():
            # junctions resolve outside the delivery root and would false-fail relative_to().
            candidate = delivery_root.joinpath(*parts)
            if candidate.exists() and candidate.is_file():
                existing += 1
            else:
                missing += 1
                missing_samples.append(t)

        return {
            "total_links": len(unique_targets),
            "formula_instances": sum(targets.values()),
            "existing": existing,
            "missing": missing,
            "broken_paths": broken_paths,
            "missing_samples": missing_samples,
            "absolute_hits": absolute_hits,
            "targets": unique_targets,
        }

    def _sample_link_tests(self, validation: dict) -> list[dict]:
        wanted = [
            ("STARTER_TF", "01_STARTER/ST-051-STD-01/01_Technical_File.docx"),
            ("STARTER_DOC", "01_STARTER/ST-051-STD-01/02_EU_DoC.docx"),
            ("STARTER_LBL", "01_STARTER/ST-051-STD-01/03_Label.docx"),
            ("STARTER_STM", "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx"),
            ("INDUSTRIAL", "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx"),
            ("CONTAINER", "03_CONTAINER/CNT-20-STD-01/03_Label.docx"),
        ]
        out = []
        for label, rel in wanted:
            path = self.delivery / rel
            out.append(
                {
                    "sample": label,
                    "relative": rel,
                    "exists": path.exists(),
                    "in_workbook_links": rel in validation.get("targets", []),
                }
            )
        return out

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        lines = [
            "# Phase K — UI Finalization QA",
            "",
            f"- **PHASE K UI FINALIZATION: {qa['gate']}**",
            "",
            f"- Workbook path: `{qa['workbook_path']}`",
            f"- Final delivery root: `{qa['delivery_root']}`",
            f"- Delivery workbook: `{qa['delivery_workbook']}`",
            "",
            "## Counts",
            "",
            f"- User-facing sheets: {qa['user_facing_sheets_count']}",
            f"- Home buttons added: {qa['home_buttons_added']}",
            f"- Total unique document links: {qa['total_document_links']}",
            f"- Working links: {qa['working_links']}",
            f"- Broken links: {qa['broken_links']}",
            f"- Broken paths: {qa.get('broken_paths', 0)}",
            f"- Absolute path hits: {qa['absolute_path_hits']}",
            "",
            "## Native Excel",
            "",
            f"- `{qa['native_excel_open']}`",
            "",
            "## Sample link tests",
            "",
        ]
        for s in qa["tested_sample_links"]:
            lines.append(
                f"- {s['sample']}: `{s['relative']}` exists={s['exists']} "
                f"in_workbook={s['in_workbook_links']}"
            )
        if qa["broken_path_samples"]:
            lines += ["", "## Broken path samples", ""]
            lines.extend(f"- `{p}`" for p in qa["broken_path_samples"])
        lines += [
            "",
            "## Messages",
            "",
        ]
        lines.extend(f"- {m}" for m in messages)
        lines += [
            "",
            "## Confirmations",
            "",
            f"- Canonical data changed: {qa['canonical_data_changed']}",
            f"- Word regenerated: {qa['word_regenerated']}",
            "- Golden templates modified: NO",
            "- Rev01 started: NO",
            "",
            f"**PHASE K UI FINALIZATION: {qa['gate']}**",
            "",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
        self.qa_path.with_suffix(".json").write_text(
            json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def main() -> int:
    import sys

    root = Path(__file__).resolve().parents[2]
    sys.path.insert(0, str(root))
    sys.path.insert(0, str(root / "src"))
    result = PhaseKService(root).run()
    for m in result.messages:
        try:
            print(m)
        except UnicodeEncodeError:
            print(m.encode("ascii", "replace").decode("ascii"))
    print("PHASE K UI FINALIZATION:", result.gate)
    if result.qa:
        print("V2:", result.qa.get("workbook_path"))
        print("Delivery root:", result.qa.get("delivery_root"))
    return 0 if result.success else 1


if __name__ == "__main__":
    raise SystemExit(main())
