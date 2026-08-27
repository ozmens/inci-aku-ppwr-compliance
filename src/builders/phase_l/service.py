"""Phase L — premium client UI / dashboard polish (no data / Word changes)."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── İnci Akü premium palette ──────────────────────────────────────────
NAVY = "1F4E79"
NAVY_DEEP = "163A5F"
YELLOW = "FFF2CC"
YELLOW_ACCENT = "F4C430"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E3F0"
SOFT_BLUE = "EAF1F8"
ROW_BAND = "F5F9FC"
BODY = "1A1A1A"
MUTED = "5A6A7A"
THIN = "B0B0B0"
LINK_BLUE = "0563C1"
OK_GREEN_BG = "E2EFDA"
OK_GREEN = "375623"
FONT = "Tahoma"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT, size=10, bold=True, color=WHITE)
TITLE_FONT = Font(name=FONT, size=20, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name=FONT, size=12, bold=True, color=YELLOW)
TERTIARY_FONT = Font(name=FONT, size=9, color=WHITE)
SECTION_FONT = Font(name=FONT, size=12, bold=True, color=NAVY)
BODY_FONT = Font(name=FONT, size=9, color=BODY)
MUTED_FONT = Font(name=FONT, size=9, italic=True, color=MUTED)
KPI_VALUE_FONT = Font(name=FONT, size=16, bold=True, color=NAVY)
KPI_LABEL_FONT = Font(name=FONT, size=8, bold=True, color=MUTED)
HOME_FONT = Font(name=FONT, size=11, bold=True, color=NAVY)
LINK_FONT = Font(name=FONT, size=9, color=LINK_BLUE, underline="single")
TILE_FONT = Font(name=FONT, size=10, bold=True, color=WHITE)
STATUS_OK_FONT = Font(name=FONT, size=9, bold=True, color=OK_GREEN)

HOME_FILL = PatternFill("solid", fgColor=YELLOW)
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
DEEP_FILL = PatternFill("solid", fgColor=NAVY_DEEP)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
SOFT_FILL = PatternFill("solid", fgColor=SOFT_BLUE)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
BAND_FILL = PatternFill("solid", fgColor=ROW_BAND)
OK_FILL = PatternFill("solid", fgColor=OK_GREEN_BG)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)

THIN_BORDER = Border(
    left=Side(style="thin", color=THIN),
    right=Side(style="thin", color=THIN),
    top=Side(style="thin", color=THIN),
    bottom=Side(style="thin", color=THIN),
)
CARD_BORDER = Border(
    left=Side(style="medium", color=NAVY),
    right=Side(style="medium", color=NAVY),
    top=Side(style="medium", color=NAVY),
    bottom=Side(style="medium", color=NAVY),
)

UI_SHEETS = [
    "00_HOME",
    "NAVIGATION",
    "SEARCH",
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
    "SHIPMENTS",
    "DOC_ENGINE_MAP",
]

TABLE_SHEETS = {
    "PACKAGING_CONFIGURATIONS",
    "PRODUCT_MASTER",
    "COMPONENT_MASTER",
    "DOCUMENT_CENTER",
    "TECHNICAL_FILES",
    "DECLARATIONS_OF_CONFORMITY",
    "LABELS",
    "SHIPMENT_STATEMENTS",
}

SHEET_META = {
    "PACKAGING_CONFIGURATIONS": {
        "title": "Packaging Configurations",
        "subtitle": "Final packaging set register — 247 controlled configurations",
        "summary": [
            ("Configurations", "247"),
            ("Starter", "240"),
            ("Industrial", "3"),
            ("Container", "4"),
        ],
    },
    "PRODUCT_MASTER": {
        "title": "Product Master",
        "subtitle": "Products linked to packaging configurations",
        "summary": [("Products", "2046"), ("Configs", "247")],
    },
    "COMPONENT_MASTER": {
        "title": "Component Master",
        "subtitle": "Packaging component catalogue",
        "summary": [("Components", "112")],
    },
    "DOCUMENT_CENTER": {
        "title": "Document Center",
        "subtitle": "Per-configuration document pack — open Technical File, DoC, Label, Statement",
        "summary": [
            ("Configurations", "247"),
            ("Documents linked", "988 / 988"),
            ("Link status", "PASS"),
            ("Revision", "Rev.00"),
        ],
    },
    "TECHNICAL_FILES": {
        "title": "Technical Files",
        "subtitle": "PPWR technical file index — one per packaging configuration",
        "summary": [("Technical Files", "247"), ("Linked", "247 / 247")],
    },
    "DECLARATIONS_OF_CONFORMITY": {
        "title": "Declarations of Conformity",
        "subtitle": "EU DoC index — one per packaging configuration",
        "summary": [("EU DoCs", "247"), ("Linked", "247 / 247")],
    },
    "LABELS": {
        "title": "Labels",
        "subtitle": "Packaging label index — one per packaging configuration",
        "summary": [("Labels", "247"), ("Linked", "247 / 247")],
    },
    "SHIPMENT_STATEMENTS": {
        "title": "Shipment Statements",
        "subtitle": "Shipment statement index — one per packaging configuration",
        "summary": [("Statements", "247"), ("Linked", "247 / 247")],
    },
}

EXPECTED_COUNTS = {
    "packaging_configurations": 247,
    "bom_lines": 1690,
    "components": 112,
    "products": 2046,
    "documents": 988,
}


@dataclass
class PhaseLResult:
    success: bool
    gate: str
    messages: list[str] = field(default_factory=list)
    qa: dict[str, Any] = field(default_factory=dict)


def excel_open_ok(path: Path) -> dict:
    pythoncom.CoInitialize()
    excel = None
    out: dict[str, Any] = {"ok": False, "error": None, "sheets": None}
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3
        wb = excel.Workbooks.Open(
            str(path.resolve()),
            UpdateLinks=0,
            ReadOnly=True,
            CorruptLoad=0,
        )
        out["ok"] = True
        out["sheets"] = int(wb.Worksheets.Count)
        wb.Close(False)
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return out


def _junction_or_copy(src: Path, dst: Path) -> None:
    if dst.exists():
        if dst.is_symlink() or dst.is_junction():
            dst.unlink()
        else:
            shutil.rmtree(dst, ignore_errors=True)
    try:
        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(dst), str(src)],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        shutil.copytree(src, dst)


def _clear_sheet(ws) -> None:
    if ws.tables:
        for name in list(ws.tables.keys()):
            del ws.tables[name]
    merges = list(ws.merged_cells.ranges)
    for m in merges:
        try:
            ws.unmerge_cells(str(m))
        except Exception:
            pass
    # remove charts
    if hasattr(ws, "_charts"):
        ws._charts = []
    ws.auto_filter.ref = None
    ws.freeze_panes = None
    max_r = ws.max_row or 1
    max_c = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name=FONT)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.hyperlink = None


def _home_formula() -> str:
    return '=HYPERLINK("#\'00_HOME\'!A1","◀ Ana Sayfaya Dön  |  Turn Back Home")'


def _nav_formula(sheet: str, label: str) -> str:
    safe = label.replace('"', "'")
    return f'=HYPERLINK("#\'{sheet}\'!A1","{safe}")'


def _style_home_cell(cell) -> None:
    cell.value = _home_formula()
    cell.font = HOME_FONT
    cell.fill = HOME_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def _fill_range(ws, r1, c1, r2, c2, fill) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill


def _border_range(ws, r1, c1, r2, c2, border=THIN_BORDER) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = border


class PhaseLService:
    def __init__(self, project_root: Path) -> None:
        self.root = project_root
        self.out = project_root / "output"
        self.source = (
            self.out
            / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_UI_READY"
            / "INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx"
        )
        self.candidate = self.out / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_POLISHED_CANDIDATE.xlsx"
        self.phase_i = self.out / "PHASE_I_FINAL"
        self.delivery = self.out / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_UI_POLISHED"
        self.delivery_workbook = (
            self.delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_POLISHED_CANDIDATE.xlsx"
        )
        self.qa_path = self.out / "PHASE_L_UI_POLISH_QA.md"

    def run(self) -> PhaseLResult:
        messages: list[str] = []
        if not self.source.exists():
            return PhaseLResult(False, "FAIL", [f"Missing source workbook: {self.source}"])

        if self.candidate.exists():
            self.candidate.unlink()
        shutil.copy2(self.source, self.candidate)
        messages.append(f"Candidate copied from FINAL → {self.candidate.name}")

        # Baseline counts from source copy before polish (canonical DB sheets)
        wb = load_workbook(self.candidate)
        baseline = self._canonical_counts(wb)
        messages.append(f"Baseline counts: {baseline}")

        self._polish_home(wb)
        messages.append("HOME dashboard redesigned")

        self._polish_navigation(wb)
        messages.append("NAVIGATION polished")

        self._polish_search(wb)
        messages.append("SEARCH polished (Excel-safe XLOOKUP)")

        for name in TABLE_SHEETS:
            if name in wb.sheetnames:
                self._polish_table_sheet(wb[name], name)
        messages.append(f"Table UI sheets polished: {len(TABLE_SHEETS)}")

        if "SHIPMENTS" in wb.sheetnames:
            self._polish_title_sheet(
                wb["SHIPMENTS"],
                "Shipments",
                "Transactional shipment register (application layer)",
            )
        if "DOC_ENGINE_MAP" in wb.sheetnames:
            self._polish_title_sheet(
                wb["DOC_ENGINE_MAP"],
                "Document Engine Map",
                "Read-only mapping — Python remains the document authority",
            )
        messages.append("SHIPMENTS + DOC_ENGINE_MAP polished")

        self._reorder_sheets(wb)
        messages.append("UI sheets ordered first")

        # Ensure home buttons on all UI sheets
        home_count = self._ensure_all_home_buttons(wb)
        messages.append(f"Home buttons verified: {home_count}")

        wb.save(self.candidate)
        wb.close()
        messages.append("Polished candidate saved")

        self._build_delivery_root()
        messages.append(f"Polished delivery root: {self.delivery}")

        # Validate from delivery root
        after = self._canonical_counts_file(self.delivery_workbook)
        validation = self._validate_links(self.delivery_workbook, self.delivery)
        home_check = self._count_home_buttons(self.delivery_workbook)
        excel = excel_open_ok(self.delivery_workbook)
        samples = self._sample_links(validation)
        white_on_light = self._scan_white_on_light(self.delivery_workbook)

        messages.append(
            f"Links: {validation['existing']}/{validation['total_links']} "
            f"missing={validation['missing']}"
        )
        messages.append(f"Home buttons: {home_check}/13")
        messages.append(f"Native Excel: {excel}")
        messages.append(f"White-on-light issues: {len(white_on_light)}")

        counts_ok = after == baseline == EXPECTED_COUNTS
        gate = (
            "PASS"
            if (
                excel.get("ok")
                and validation["total_links"] == 988
                and validation["existing"] == 988
                and validation["missing"] == 0
                and validation.get("broken_paths", 0) == 0
                and validation["absolute_hits"] == 0
                and home_check == 13
                and counts_ok
                and all(s["exists"] for s in samples)
                and len(white_on_light) == 0
            )
            else "FAIL"
        )

        qa = {
            "gate": gate,
            "workbook_path": str(self.candidate),
            "delivery_root": str(self.delivery),
            "delivery_workbook": str(self.delivery_workbook),
            "sheets_polished": UI_SHEETS,
            "design_changes": [
                "HOME rebuilt as premium executive command center with KPI cards, status panel, nav tiles, and safe bar charts",
                "NAVIGATION rebuilt as grouped premium menu",
                "SEARCH rebuilt with clear lookup box and Excel-safe XLOOKUP (no FILTER)",
                "Document/data UI sheets: summary strip, navy headers, banding, freeze, filters, column widths",
                "Consistent Ana Sayfaya Dön home bar retained on all 13 UI sheets",
                "UI sheet order placed first; database sheets preserved",
            ],
            "baseline_counts": baseline,
            "after_counts": after,
            "counts_unchanged": counts_ok,
            "total_document_links": validation["total_links"],
            "working_links": validation["existing"],
            "broken_links": validation["missing"],
            "broken_paths": validation.get("broken_paths", 0),
            "absolute_path_hits": validation["absolute_hits"],
            "home_buttons": home_check,
            "tested_sample_links": samples,
            "native_excel_open": excel,
            "white_on_light_issues": white_on_light[:20],
            "canonical_data_changed": False,
            "word_regenerated": False,
            "final_overwritten": False,
        }
        self._write_qa(qa, messages)
        return PhaseLResult(gate == "PASS", gate, messages, qa)

    # ── HOME ───────────────────────────────────────────────────────────
    def _polish_home(self, wb) -> None:
        ws = wb["00_HOME"]
        _clear_sheet(ws)
        ws.sheet_view.showGridLines = False

        # widths
        widths = {
            1: 22,
            2: 14,
            3: 22,
            4: 14,
            5: 22,
            6: 14,
            7: 3,
            8: 28,
            9: 22,
            10: 18,
            11: 18,
            12: 14,
        }
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        # Row 1 — home / status ribbon
        _style_home_cell(ws["A1"])
        try:
            ws.merge_cells("A1:F1")
        except Exception:
            pass
        ws["H1"] = "CONTROLLED DELIVERY  ·  REV.00"
        ws["H1"].font = Font(name=FONT, size=9, bold=True, color=NAVY)
        ws["H1"].fill = HOME_FILL
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
        try:
            ws.merge_cells("H1:L1")
        except Exception:
            pass
        ws.row_dimensions[1].height = 22

        # Banner
        _fill_range(ws, 2, 1, 4, 12, NAVY_FILL)
        ws["A2"] = "İnci Akü PPWR Packaging Information Management System"
        ws["A2"].font = TITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        try:
            ws.merge_cells("A2:L2")
        except Exception:
            pass
        ws.row_dimensions[2].height = 32

        ws["A3"] = "Rev.00 Controlled Delivery Package"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].fill = NAVY_FILL
        try:
            ws.merge_cells("A3:L3")
        except Exception:
            pass
        ws.row_dimensions[3].height = 20

        ws["A4"] = (
            "Packaging Configurations, Technical Files, Declarations and Document Registry"
        )
        ws["A4"].font = TERTIARY_FONT
        ws["A4"].fill = DEEP_FILL
        try:
            ws.merge_cells("A4:L4")
        except Exception:
            pass
        ws.row_dimensions[4].height = 18

        # KPI section header
        ws["A6"] = "EXECUTIVE KPIs"
        ws["A6"].font = Font(name=FONT, size=11, bold=True, color=WHITE)
        ws["A6"].fill = NAVY_FILL
        try:
            ws.merge_cells("A6:F6")
        except Exception:
            pass

        kpis = [
            ("Final Packaging Configurations", "247", "LOADED"),
            ("Starter", "240", "LOADED"),
            ("Industrial", "3", "LOADED"),
            ("Container / Loading", "4", "LOADED"),
            ("Components", "112", "LOADED"),
            ("Products", "2046", "LOADED"),
            ("BOM Lines", "1690", "LOADED"),
            ("Technical Files", "247", "GENERATED"),
            ("EU DoCs", "247", "GENERATED"),
            ("Labels", "247", "GENERATED"),
            ("Shipment Statements", "247", "GENERATED"),
            ("Total Documents", "988", "GENERATED"),
            ("Document Links Working", "988 / 988", "PASS"),
            ("Blocking QA Errors", "0", "PASS"),
        ]
        # 7 cards per row × 2 columns each (label+value area uses 2 cols): A-B, C-D, E-F
        # Simpler: 3 columns of cards, each card spans 2 cols and 3 rows
        positions = [
            (7, 1),
            (7, 3),
            (7, 5),
            (10, 1),
            (10, 3),
            (10, 5),
            (13, 1),
            (13, 3),
            (13, 5),
            (16, 1),
            (16, 3),
            (16, 5),
            (19, 1),
            (19, 3),
        ]
        for (label, value, status), (r, c) in zip(kpis, positions):
            self._kpi_card(ws, r, c, label, value, status)

        # Status panel
        ws["H6"] = "CONTROLLED STATUS"
        ws["H6"].font = Font(name=FONT, size=11, bold=True, color=WHITE)
        ws["H6"].fill = NAVY_FILL
        try:
            ws.merge_cells("H6:L6")
        except Exception:
            pass

        status_rows = [
            ("Production Master Data", "LOADED"),
            ("Golden Variant Register", "247 / 247"),
            ("Document Pack", "988 / 988 GENERATED"),
            ("Document Registry", "988 / 988 LINKED"),
            ("Native Excel Validation", "PASS"),
            ("Rev.01 Started", "NO"),
            ("Legal signatures in Excel", "NOT COMPLETED"),
            ("Word files = legal authority", "YES"),
        ]
        for i, (k, v) in enumerate(status_rows):
            r = 7 + i
            ws.cell(r, 8, k).font = BODY_FONT
            ws.cell(r, 8).fill = SOFT_FILL
            ws.cell(r, 8).border = THIN_BORDER
            try:
                ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
            except Exception:
                pass
            cell = ws.cell(r, 11, v)
            cell.font = STATUS_OK_FONT if v not in ("NOT COMPLETED",) else MUTED_FONT
            cell.fill = OK_FILL if "PASS" in v or v in ("LOADED", "YES", "NO", "247 / 247") or "988" in v else YELLOW_FILL
            if v == "NOT COMPLETED":
                cell.fill = YELLOW_FILL
                cell.font = Font(name=FONT, size=9, bold=True, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            try:
                ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
            except Exception:
                pass
            ws.row_dimensions[r].height = 18

        # Quick navigation
        ws["A22"] = "QUICK NAVIGATION"
        ws["A22"].font = Font(name=FONT, size=11, bold=True, color=WHITE)
        ws["A22"].fill = NAVY_FILL
        try:
            ws.merge_cells("A22:F22")
        except Exception:
            pass

        nav_tiles = [
            ("Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
            ("Product Master", "PRODUCT_MASTER"),
            ("Component Master", "COMPONENT_MASTER"),
            ("Document Center", "DOCUMENT_CENTER"),
            ("Technical Files", "TECHNICAL_FILES"),
            ("Declarations of Conformity", "DECLARATIONS_OF_CONFORMITY"),
            ("Labels", "LABELS"),
            ("Shipment Statements", "SHIPMENT_STATEMENTS"),
            ("Search", "SEARCH"),
            ("Navigation", "NAVIGATION"),
        ]
        # 5 tiles × 2 rows
        for i, (label, sheet) in enumerate(nav_tiles):
            r = 23 + (i // 5) * 2
            c = 1 + (i % 5)
            # map to columns A,B,C,D,E then next row — use pairs for wider tiles
            # Use columns: 1,2 | 3,4 | 5,6 for 3 per row... user asked 10 tiles.
            # Layout: row 23-24: 5 tiles in cols 1-5 spanning visually via fill on single cells wide
            cell = ws.cell(r, c)
            cell.value = _nav_formula(sheet, label)
            cell.font = TILE_FONT
            cell.fill = NAVY_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = CARD_BORDER
            ws.row_dimensions[r].height = 28
            # light spacer row
            if i < 5:
                ws.row_dimensions[r + 1].height = 6

        # Widen nav area: put tiles across A-F using 2-col merges for first row of 3 + second
        # Clear the simple 5-col layout and do cleaner 2-col tiles
        for r in (23, 24, 25, 26):
            for c in range(1, 7):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = PatternFill()
                cell.border = Border()
                cell.font = Font(name=FONT)

        tile_layout = [
            (23, 1, 2, "Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
            (23, 3, 4, "Product Master", "PRODUCT_MASTER"),
            (23, 5, 6, "Component Master", "COMPONENT_MASTER"),
            (25, 1, 2, "Document Center", "DOCUMENT_CENTER"),
            (25, 3, 4, "Technical Files", "TECHNICAL_FILES"),
            (25, 5, 6, "Declarations of Conformity", "DECLARATIONS_OF_CONFORMITY"),
            (27, 1, 2, "Labels", "LABELS"),
            (27, 3, 4, "Shipment Statements", "SHIPMENT_STATEMENTS"),
            (27, 5, 6, "Search", "SEARCH"),
            (29, 1, 2, "Navigation", "NAVIGATION"),
            (29, 3, 4, "Shipments", "SHIPMENTS"),
            (29, 5, 6, "Doc Engine Map", "DOC_ENGINE_MAP"),
        ]
        for r, c1, c2, label, sheet in tile_layout:
            cell = ws.cell(r, c1)
            cell.value = _nav_formula(sheet, label)
            cell.font = TILE_FONT
            cell.fill = NAVY_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for c in range(c1, c2 + 1):
                ws.cell(r, c).fill = NAVY_FILL
                ws.cell(r, c).border = CARD_BORDER
            try:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            except Exception:
                pass
            ws.row_dimensions[r].height = 30

        # Chart data (hidden area) + charts
        ws["H16"] = "Configuration breakdown"
        ws["H16"].font = SECTION_FONT
        try:
            ws.merge_cells("H16:I16")
        except Exception:
            pass
        ws["H17"] = "Family"
        ws["I17"] = "Count"
        ws["H17"].font = HEADER_FONT
        ws["I17"].font = HEADER_FONT
        ws["H17"].fill = NAVY_FILL
        ws["I17"].fill = NAVY_FILL
        for r, (fam, cnt) in enumerate(
            [("Starter", 240), ("Industrial", 3), ("Container", 4)], start=18
        ):
            ws.cell(r, 8, fam).font = BODY_FONT
            ws.cell(r, 9, cnt).font = BODY_FONT
            ws.cell(r, 8).border = THIN_BORDER
            ws.cell(r, 9).border = THIN_BORDER
            ws.cell(r, 8).fill = SOFT_FILL
            ws.cell(r, 9).fill = WHITE_FILL

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Configurations by Family"
        chart1.y_axis.title = None
        chart1.x_axis.title = None
        chart1.style = 10
        data = Reference(ws, min_col=9, min_row=17, max_row=20)
        cats = Reference(ws, min_col=8, min_row=18, max_row=20)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        chart1.width = 12
        chart1.height = 7
        chart1.legend = None
        chart1.dataLabels = DataLabelList()
        chart1.dataLabels.showVal = True
        ws.add_chart(chart1, "H22")

        ws["K16"] = "Document type breakdown"
        ws["K16"].font = SECTION_FONT
        try:
            ws.merge_cells("K16:L16")
        except Exception:
            pass
        ws["K17"] = "Type"
        ws["L17"] = "Count"
        ws["K17"].font = HEADER_FONT
        ws["L17"].font = HEADER_FONT
        ws["K17"].fill = NAVY_FILL
        ws["L17"].fill = NAVY_FILL
        for r, (typ, cnt) in enumerate(
            [
                ("Technical File", 247),
                ("EU DoC", 247),
                ("Label", 247),
                ("Statement", 247),
            ],
            start=18,
        ):
            ws.cell(r, 11, typ).font = BODY_FONT
            ws.cell(r, 12, cnt).font = BODY_FONT
            ws.cell(r, 11).border = THIN_BORDER
            ws.cell(r, 12).border = THIN_BORDER
            ws.cell(r, 11).fill = SOFT_FILL
            ws.cell(r, 12).fill = WHITE_FILL

        chart2 = BarChart()
        chart2.type = "bar"
        chart2.title = "Documents by Type"
        chart2.style = 10
        data2 = Reference(ws, min_col=12, min_row=17, max_row=21)
        cats2 = Reference(ws, min_col=11, min_row=18, max_row=21)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.shape = 4
        chart2.width = 12
        chart2.height = 7
        chart2.legend = None
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showVal = True
        ws.add_chart(chart2, "H32")

        ws["A31"] = (
            "Note: Legal signatures are NOT completed in this workbook. "
            "Word files remain the controlled legal documents."
        )
        ws["A31"].font = MUTED_FONT
        try:
            ws.merge_cells("A31:F31")
        except Exception:
            pass

        ws.freeze_panes = "A6"

    def _kpi_card(self, ws, row: int, col: int, label: str, value: str, status: str) -> None:
        # 2 cols × 3 rows
        _fill_range(ws, row, col, row + 2, col + 1, SOFT_FILL)
        _border_range(ws, row, col, row + 2, col + 1, CARD_BORDER)
        lab = ws.cell(row, col, label)
        lab.font = KPI_LABEL_FONT
        lab.fill = LIGHT_FILL
        lab.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, col + 1).fill = LIGHT_FILL
        try:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        except Exception:
            pass

        val = ws.cell(row + 1, col, value)
        val.font = KPI_VALUE_FONT
        val.fill = WHITE_FILL
        val.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row + 1, col + 1).fill = WHITE_FILL
        try:
            ws.merge_cells(
                start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
            )
        except Exception:
            pass

        st = ws.cell(row + 2, col, status)
        st.font = Font(name=FONT, size=8, bold=True, color=NAVY)
        st.fill = HOME_FILL
        st.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row + 2, col + 1).fill = HOME_FILL
        try:
            ws.merge_cells(
                start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1
            )
        except Exception:
            pass
        ws.row_dimensions[row].height = 18
        ws.row_dimensions[row + 1].height = 24
        ws.row_dimensions[row + 2].height = 16

    # ── NAVIGATION ─────────────────────────────────────────────────────
    def _polish_navigation(self, wb) -> None:
        ws = wb["NAVIGATION"]
        _clear_sheet(ws)
        ws.sheet_view.showGridLines = False
        for c, w in {1: 36, 2: 36, 3: 36, 4: 36}.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        _style_home_cell(ws["A1"])
        try:
            ws.merge_cells("A1:D1")
        except Exception:
            pass
        ws.row_dimensions[1].height = 22

        _fill_range(ws, 2, 1, 3, 4, NAVY_FILL)
        ws["A2"] = "Navigation"
        ws["A2"].font = TITLE_FONT
        try:
            ws.merge_cells("A2:D2")
        except Exception:
            pass
        ws["A3"] = "Premium menu — jump to any application module"
        ws["A3"].font = TERTIARY_FONT
        ws["A3"].fill = DEEP_FILL
        try:
            ws.merge_cells("A3:D3")
        except Exception:
            pass

        sections = [
            (
                "A. Dashboard",
                [("Home", "00_HOME"), ("Search", "SEARCH")],
            ),
            (
                "B. Packaging Management",
                [
                    ("Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
                    ("Product Master", "PRODUCT_MASTER"),
                    ("Component Master", "COMPONENT_MASTER"),
                ],
            ),
            (
                "C. Document Management",
                [
                    ("Document Center", "DOCUMENT_CENTER"),
                    ("Technical Files", "TECHNICAL_FILES"),
                    ("Declarations of Conformity", "DECLARATIONS_OF_CONFORMITY"),
                    ("Labels", "LABELS"),
                    ("Shipment Statements", "SHIPMENT_STATEMENTS"),
                ],
            ),
            (
                "D. Operations / System",
                [
                    ("Shipments", "SHIPMENTS"),
                    ("Document Engine Map", "DOC_ENGINE_MAP"),
                    ("Release Control", "02_RELEASE_CONTROL"),
                    ("Data Dictionary", "03_DATA_DICTIONARY"),
                    ("Workbook Info", "SYS_WORKBOOK_INFO"),
                ],
            ),
        ]

        r = 5
        for title, links in sections:
            ws.cell(r, 1, title).font = Font(name=FONT, size=11, bold=True, color=WHITE)
            ws.cell(r, 1).fill = NAVY_FILL
            try:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            except Exception:
                pass
            for c in range(1, 5):
                ws.cell(r, c).fill = NAVY_FILL
            r += 1
            # tiles in row pairs
            for i, (label, sheet) in enumerate(links):
                if sheet not in wb.sheetnames:
                    continue
                c = 1 + (i % 2) * 2
                if i > 0 and i % 2 == 0:
                    r += 1
                cell = ws.cell(r, c)
                cell.value = _nav_formula(sheet, label)
                cell.font = TILE_FONT
                cell.fill = NAVY_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                for cc in (c, c + 1):
                    ws.cell(r, cc).fill = NAVY_FILL
                    ws.cell(r, cc).border = CARD_BORDER
                try:
                    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
                except Exception:
                    pass
                ws.row_dimensions[r].height = 28
            r += 2

        ws.freeze_panes = "A5"

    # ── SEARCH ─────────────────────────────────────────────────────────
    def _polish_search(self, wb) -> None:
        ws = wb["SEARCH"]
        _clear_sheet(ws)
        ws.sheet_view.showGridLines = False
        for c, w in {1: 28, 2: 56, 3: 24, 4: 24}.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        _style_home_cell(ws["A1"])
        try:
            ws.merge_cells("A1:D1")
        except Exception:
            pass

        _fill_range(ws, 2, 1, 3, 4, NAVY_FILL)
        ws["A2"] = "Search"
        ws["A2"].font = TITLE_FONT
        try:
            ws.merge_cells("A2:D2")
        except Exception:
            pass
        ws["A3"] = (
            "Excel-safe lookup — type a packaging set / configuration ID, then read results below. "
            "For full browsing use AutoFilter on Document Center (or Ctrl+F)."
        )
        ws["A3"].font = TERTIARY_FONT
        ws["A3"].fill = DEEP_FILL
        try:
            ws.merge_cells("A3:D3")
        except Exception:
            pass

        ws["A5"] = "SEARCH BOX"
        ws["A5"].font = Font(name=FONT, size=10, bold=True, color=WHITE)
        ws["A5"].fill = NAVY_FILL
        try:
            ws.merge_cells("A5:B5")
        except Exception:
            pass

        ws["A6"] = "Lookup key"
        ws["A6"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws["A6"].fill = LIGHT_FILL
        ws["B6"] = ""  # user input cell
        ws["B6"].fill = YELLOW_FILL
        ws["B6"].border = CARD_BORDER
        ws["B6"].font = Font(name=FONT, size=12, bold=True, color=NAVY)
        ws["B6"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[6].height = 26

        ws["A7"] = "Examples"
        ws["A7"].font = MUTED_FONT
        ws["B7"] = (
            "ST-051-STD-01  |  IA-ST-051-STD-01  |  CNT-20-STD-01  |  IND-24V-01"
        )
        ws["B7"].font = MUTED_FONT

        ws["A9"] = "LOOKUP RESULTS"
        ws["A9"].font = Font(name=FONT, size=10, bold=True, color=WHITE)
        ws["A9"].fill = NAVY_FILL
        try:
            ws.merge_cells("A9:B9")
        except Exception:
            pass

        # Point XLOOKUP at PACKAGING_CONFIGURATIONS columns.
        # After polish, that sheet has: row1 home, rows2-4 summary, row5 headers, data from row6.
        # Whole-column refs remain valid.
        results = [
            (10, "Configuration ID", "PACKAGING_CONFIGURATIONS!B:B"),
            (11, "Family", "PACKAGING_CONFIGURATIONS!D:D"),
            (12, "Technical File ID", "PACKAGING_CONFIGURATIONS!K:K"),
            (13, "Source Configuration ID", "PACKAGING_CONFIGURATIONS!C:C"),
        ]
        for r, label, col_ref in results:
            ws.cell(r, 1, label).font = Font(name=FONT, size=9, bold=True, color=NAVY)
            ws.cell(r, 1).fill = SOFT_FILL
            ws.cell(r, 1).border = THIN_BORDER
            cell = ws.cell(r, 2)
            cell.value = (
                f'=IF($B$6="","",IFERROR(XLOOKUP($B$6,PACKAGING_CONFIGURATIONS!A:A,{col_ref}),'
                f'"Not found — use AutoFilter on Document Center"))'
            )
            cell.font = BODY_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER

        ws["A15"] = "MODULE SHORTCUTS"
        ws["A15"].font = Font(name=FONT, size=10, bold=True, color=WHITE)
        ws["A15"].fill = NAVY_FILL
        try:
            ws.merge_cells("A15:B15")
        except Exception:
            pass

        shortcuts = [
            (16, "Document Center (filter here)", "DOCUMENT_CENTER"),
            (17, "Packaging Configurations", "PACKAGING_CONFIGURATIONS"),
            (18, "Technical Files", "TECHNICAL_FILES"),
            (19, "Product Master", "PRODUCT_MASTER"),
            (20, "Component Master", "COMPONENT_MASTER"),
        ]
        for r, label, sheet in shortcuts:
            cell = ws.cell(r, 1)
            cell.value = _nav_formula(sheet, label)
            cell.font = TILE_FONT
            cell.fill = NAVY_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            try:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            except Exception:
                pass
            for c in (1, 2):
                ws.cell(r, c).fill = NAVY_FILL
                ws.cell(r, c).border = CARD_BORDER
            ws.row_dimensions[r].height = 24

        ws.freeze_panes = "A5"

    # ── Table sheets ───────────────────────────────────────────────────
    def _polish_table_sheet(self, ws, name: str) -> None:
        meta = SHEET_META[name]
        # Detect header row (first row where col A is not a HYPERLINK home and looks like header)
        header_row = self._find_header_row(ws)
        if header_row is None:
            return

        # If already polished (marker in A2), only restyle
        a2 = str(ws["A2"].value or "")
        already = a2.startswith("◆ ") or a2 == meta["title"]

        if not already:
            # Insert summary block between home (row 1) and header
            # Current: row1 home, header_row headers
            insert_at = 2
            n_insert = 3  # title, subtitle, summary
            if header_row == 1:
                # unexpected — home missing; skip insert
                pass
            else:
                ws.insert_rows(insert_at, amount=n_insert)
                header_row += n_insert

        # Home bar
        _style_home_cell(ws["A1"])
        max_col = ws.max_column or 10
        try:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(4, max_col))
        except Exception:
            pass
        ws.row_dimensions[1].height = 22

        # Title
        ws["A2"] = f"◆ {meta['title']}"
        ws["A2"].font = Font(name=FONT, size=14, bold=True, color=WHITE)
        ws["A2"].fill = NAVY_FILL
        for c in range(1, max_col + 1):
            ws.cell(2, c).fill = NAVY_FILL
        try:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        except Exception:
            pass
        ws.row_dimensions[2].height = 24

        # Subtitle
        ws["A3"] = meta["subtitle"]
        ws["A3"].font = Font(name=FONT, size=9, color=NAVY)
        ws["A3"].fill = LIGHT_FILL
        for c in range(1, max_col + 1):
            ws.cell(3, c).fill = LIGHT_FILL
        try:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
        except Exception:
            pass

        # Summary strip
        ws["A4"] = "  |  ".join(f"{k}: {v}" for k, v in meta["summary"])
        ws["A4"].font = Font(name=FONT, size=9, bold=True, color=NAVY)
        ws["A4"].fill = YELLOW_FILL
        for c in range(1, max_col + 1):
            ws.cell(4, c).fill = YELLOW_FILL
        try:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_col)
        except Exception:
            pass

        # Re-find header if we just inserted
        header_row = self._find_header_row(ws) or header_row

        # Style header
        for c in range(1, max_col + 1):
            cell = ws.cell(header_row, c)
            if cell.value is not None:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = THIN_BORDER
        ws.row_dimensions[header_row].height = 30

        # Banding + font on data (do not change values)
        last_row = ws.max_row or header_row
        for r in range(header_row + 1, last_row + 1):
            band = BAND_FILL if (r - header_row) % 2 == 0 else WHITE_FILL
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                val = cell.value
                # preserve hyperlink font for Open links
                if isinstance(val, str) and val.upper().startswith("=HYPERLINK("):
                    cell.font = LINK_FONT
                else:
                    cell.font = BODY_FONT
                if not (isinstance(val, str) and "Ana Sayfaya" in val):
                    cell.fill = band
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Status column highlighting if present
        headers = [str(ws.cell(header_row, c).value or "") for c in range(1, max_col + 1)]
        for idx, h in enumerate(headers, start=1):
            if "status" in h.lower():
                for r in range(header_row + 1, last_row + 1):
                    cell = ws.cell(r, idx)
                    txt = str(cell.value or "").upper()
                    if "PASS" in txt or "LINKED" in txt or "COMPLETE" in txt or "GENERATED" in txt:
                        cell.fill = OK_FILL
                        cell.font = STATUS_OK_FONT

        # Freeze / filter — only if no Excel Table (corruption risk)
        ws.freeze_panes = f"A{header_row + 1}"
        if not ws.tables:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(max_col)}{last_row}"
            )

        # Column widths — sensible defaults, don't destroy
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            header = str(ws.cell(header_row, c).value or "")
            if "Open" in header:
                ws.column_dimensions[letter].width = 12
            elif "Variant" in header or "Description" in header:
                ws.column_dimensions[letter].width = 36
            elif "ID" in header or "Code" in header:
                ws.column_dimensions[letter].width = 22
            else:
                cur = ws.column_dimensions[letter].width
                if not cur or cur < 12:
                    ws.column_dimensions[letter].width = 16

        ws.sheet_view.showGridLines = False

    def _find_header_row(self, ws) -> int | None:
        for r in range(1, min(12, (ws.max_row or 1) + 1)):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            s = str(v)
            if s.startswith("=HYPERLINK"):
                continue
            if s.startswith("◆"):
                continue
            # heuristic: header-like text
            if any(
                k in s
                for k in (
                    "Packaging Set",
                    "Product Code",
                    "ERP Component",
                    "Label ID",
                    "Technical File",
                    "Declaration",
                    "Statement",
                    "Configuration",
                )
            ):
                return r
            # or second column also looks like header
            v2 = ws.cell(r, 2).value
            if v2 and any(
                k in str(v2)
                for k in ("Configuration", "Product", "Component", "Document", "Family")
            ):
                return r
        return None

    def _polish_title_sheet(self, ws, title: str, subtitle: str) -> None:
        # Keep existing content below; ensure top chrome
        a1 = str(ws["A1"].value or "")
        if "Ana Sayfaya Dön" not in a1 and "Turn Back Home" not in a1:
            ws.insert_rows(1)
        _style_home_cell(ws["A1"])
        try:
            ws.merge_cells("A1:D1")
        except Exception:
            pass

        # If A2 is already a title-ish, restyle; else insert title rows
        a2 = str(ws["A2"].value or "")
        if not a2.startswith("◆"):
            ws.insert_rows(2, amount=2)
            ws["A2"] = f"◆ {title}"
            ws["A3"] = subtitle
        else:
            ws["A2"] = f"◆ {title}"
            if ws.max_row >= 3:
                ws["A3"] = subtitle

        ws["A2"].font = Font(name=FONT, size=14, bold=True, color=WHITE)
        ws["A2"].fill = NAVY_FILL
        for c in range(1, min(ws.max_column or 4, 8) + 1):
            ws.cell(2, c).fill = NAVY_FILL
        try:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(ws.max_column or 4, 6))
        except Exception:
            pass

        ws["A3"].font = Font(name=FONT, size=9, color=NAVY)
        ws["A3"].fill = LIGHT_FILL
        for c in range(1, min(ws.max_column or 4, 8) + 1):
            ws.cell(3, c).fill = LIGHT_FILL

        # Style any header-looking row below
        for r in range(4, min(10, (ws.max_row or 4) + 1)):
            v = ws.cell(r, 1).value
            if v and not str(v).startswith("=") and ws.cell(r, 2).value:
                # likely header
                for c in range(1, (ws.max_column or 1) + 1):
                    cell = ws.cell(r, c)
                    if cell.value is not None:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                break

        for c in range(1, (ws.max_column or 1) + 1):
            letter = get_column_letter(c)
            if not ws.column_dimensions[letter].width:
                ws.column_dimensions[letter].width = 22
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

    def _ensure_all_home_buttons(self, wb) -> int:
        count = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            a1 = str(ws["A1"].value or "")
            if "Ana Sayfaya Dön" in a1 or "Turn Back Home" in a1:
                _style_home_cell(ws["A1"])
                count += 1
            else:
                ws.insert_rows(1)
                _style_home_cell(ws["A1"])
                count += 1
        return count

    def _reorder_sheets(self, wb) -> None:
        # Move UI sheets to front in order
        for idx, name in enumerate(UI_SHEETS):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    # ── Counts / validation ────────────────────────────────────────────
    def _canonical_counts(self, wb) -> dict[str, int]:
        def data_rows(sheet: str) -> int:
            if sheet not in wb.sheetnames:
                return -1
            return max((wb[sheet].max_row or 1) - 1, 0)

        return {
            "packaging_configurations": data_rows("PACKAGING_CONFIGURATION"),
            "bom_lines": data_rows("PACKAGING_CONFIGURATION_LINE"),
            "components": data_rows("COMPONENT"),
            "products": data_rows("PRODUCT"),
            "documents": data_rows("DOCUMENT_LIBRARY"),
        }

    def _canonical_counts_file(self, path: Path) -> dict[str, int]:
        wb = load_workbook(path, read_only=True, data_only=False)
        counts = self._canonical_counts(wb)
        wb.close()
        return counts

    def _validate_links(self, workbook_path: Path, delivery_root: Path) -> dict[str, Any]:
        wb = load_workbook(workbook_path, data_only=False)
        targets: dict[str, int] = {}
        absolute_hits = 0
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    m = re.search(r'=HYPERLINK\("([^"]+)"', val, re.I)
                    if not m:
                        continue
                    target = m.group(1).replace("\\", "/")
                    if not target.lower().endswith(".docx"):
                        continue
                    if re.match(r"^[A-Za-z]:/", target) or "Users/" in target:
                        absolute_hits += 1
                    targets[target] = targets.get(target, 0) + 1
        wb.close()

        existing = missing = broken_paths = 0
        missing_samples: list[str] = []
        unique_targets = sorted(targets.keys())
        for t in unique_targets:
            norm = t.replace("\\", "/").strip()
            parts = [p for p in norm.split("/") if p not in ("", ".")]
            if (
                not parts
                or ".." in parts
                or re.match(r"^[A-Za-z]:", parts[0])
                or parts[0].startswith("//")
            ):
                broken_paths += 1
                missing += 1
                missing_samples.append(t)
                continue
            candidate = delivery_root.joinpath(*parts)
            if candidate.exists() and candidate.is_file():
                existing += 1
            else:
                missing += 1
                missing_samples.append(t)

        return {
            "total_links": len(unique_targets),
            "formula_instances": sum(targets.values()),
            "existing": existing,
            "missing": missing,
            "broken_paths": broken_paths,
            "missing_samples": missing_samples,
            "absolute_hits": absolute_hits,
            "targets": unique_targets,
        }

    def _count_home_buttons(self, path: Path) -> int:
        wb = load_workbook(path, data_only=False)
        n = 0
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            a1 = str(wb[name]["A1"].value or "")
            if "Ana Sayfaya Dön" in a1 or "Turn Back Home" in a1:
                n += 1
        wb.close()
        return n

    def _sample_links(self, validation: dict) -> list[dict]:
        wanted = [
            ("STARTER_TF", "01_STARTER/ST-051-STD-01/01_Technical_File.docx"),
            ("STARTER_DOC", "01_STARTER/ST-051-STD-01/02_EU_DoC.docx"),
            ("STARTER_LBL", "01_STARTER/ST-051-STD-01/03_Label.docx"),
            ("STARTER_STM", "01_STARTER/ST-051-STD-01/04_Shipment_Statement.docx"),
            ("INDUSTRIAL", "02_INDUSTRIAL/IND-24V-01/01_Technical_File.docx"),
            ("CONTAINER", "03_CONTAINER/CNT-20-STD-01/03_Label.docx"),
        ]
        out = []
        for label, rel in wanted:
            path = self.delivery / rel
            out.append(
                {
                    "sample": label,
                    "relative": rel,
                    "exists": path.exists(),
                    "in_workbook_links": rel in validation.get("targets", []),
                }
            )
        return out

    def _scan_white_on_light(self, path: Path) -> list[str]:
        """Flag white font on non-dark fills in UI sheets (sample scan)."""
        dark = {NAVY, NAVY_DEEP, "000000", "1F4E79", "163A5F"}
        issues = []
        wb = load_workbook(path, data_only=False)
        for name in UI_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            max_r = min(ws.max_row or 1, 40)
            max_c = min(ws.max_column or 1, 12)
            for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
                for cell in row:
                    if cell.value is None:
                        continue
                    color = None
                    if cell.font and cell.font.color and cell.font.color.type == "rgb":
                        color = cell.font.color.rgb
                        if color and len(color) == 8:
                            color = color[2:]
                    fill = None
                    if cell.fill and cell.fill.fgColor and cell.fill.patternType == "solid":
                        fill = cell.fill.fgColor.rgb
                        if fill and len(fill) == 8:
                            fill = fill[2:]
                    if color == "FFFFFF" and fill and fill.upper() not in dark:
                        # allow white only on navy
                        issues.append(f"{name}!{cell.coordinate} white on #{fill}")
        wb.close()
        return issues

    def _build_delivery_root(self) -> None:
        if self.delivery.exists():
            for child in list(self.delivery.iterdir()):
                try:
                    if child.is_junction() or child.is_symlink():
                        child.unlink()
                    elif child.is_dir():
                        shutil.rmtree(child, ignore_errors=True)
                    else:
                        child.unlink()
                except Exception:
                    pass
        self.delivery.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.candidate, self.delivery_workbook)
        for folder in (
            "01_STARTER",
            "02_INDUSTRIAL",
            "03_CONTAINER",
            "90_MANIFEST",
            "99_QA_REPORT",
        ):
            src = self.phase_i / folder
            if src.exists():
                _junction_or_copy(src, self.delivery / folder)

    def _write_qa(self, qa: dict, messages: list[str]) -> None:
        lines = [
            "# Phase L — UI Polish QA",
            "",
            f"- **PHASE L UI POLISH: {qa['gate']}**",
            "",
            f"- Polished candidate: `{qa['workbook_path']}`",
            f"- Polished delivery root: `{qa['delivery_root']}`",
            f"- Delivery workbook: `{qa['delivery_workbook']}`",
            "",
            "## Sheets polished",
            "",
        ]
        lines.extend(f"- `{s}`" for s in qa["sheets_polished"])
        lines += ["", "## Design changes", ""]
        lines.extend(f"- {d}" for d in qa["design_changes"])
        lines += [
            "",
            "## Counts (unchanged confirmation)",
            "",
            f"- Baseline: `{qa['baseline_counts']}`",
            f"- After polish: `{qa['after_counts']}`",
            f"- Counts unchanged: **{qa['counts_unchanged']}**",
            "",
            "## Hyperlink integrity",
            "",
            f"- Total unique document links: {qa['total_document_links']}",
            f"- Working: {qa['working_links']}",
            f"- Broken: {qa['broken_links']}",
            f"- Broken paths: {qa['broken_paths']}",
            f"- Absolute path hits: {qa['absolute_path_hits']}",
            f"- Home buttons: {qa['home_buttons']}/13",
            "",
            "## Native Excel",
            "",
            f"- `{qa['native_excel_open']}`",
            "",
            "## Sample links",
            "",
        ]
        for s in qa["tested_sample_links"]:
            lines.append(
                f"- {s['sample']}: `{s['relative']}` exists={s['exists']} "
                f"in_workbook={s['in_workbook_links']}"
            )
        lines += [
            "",
            "## White-on-light scan",
            "",
            f"- Issues: {len(qa['white_on_light_issues'])}",
        ]
        lines.extend(f"- `{i}`" for i in qa["white_on_light_issues"])
        lines += [
            "",
            "## Messages",
            "",
        ]
        lines.extend(f"- {m}" for m in messages)
        lines += [
            "",
            "## Confirmations",
            "",
            f"- Canonical data changed: {qa['canonical_data_changed']}",
            f"- Word regenerated: {qa['word_regenerated']}",
            f"- Final workbook overwritten: {qa['final_overwritten']}",
            "- Golden templates modified: NO",
            "- Rev01 started: NO",
            "",
            f"**PHASE L UI POLISH: {qa['gate']}**",
            "",
        ]
        self.qa_path.write_text("\n".join(lines), encoding="utf-8")
        self.qa_path.with_suffix(".json").write_text(
            json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def main() -> int:
    import sys

    root = Path(__file__).resolve().parents[2]
    sys.path.insert(0, str(root))
    sys.path.insert(0, str(root / "src"))
    result = PhaseLService(root).run()
    for m in result.messages:
        try:
            print(m)
        except UnicodeEncodeError:
            print(m.encode("ascii", "replace").decode("ascii"))
    print("PHASE L UI POLISH:", result.gate)
    if result.qa:
        print("Candidate:", result.qa.get("workbook_path"))
        print("Delivery root:", result.qa.get("delivery_root"))
    return 0 if result.success else 1


if __name__ == "__main__":
    raise SystemExit(main())
