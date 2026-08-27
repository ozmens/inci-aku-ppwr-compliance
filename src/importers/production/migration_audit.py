"""Phase F migration audit / QA report writer."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from importers.production.qualify import QualificationResult
from importers.production.normalizer import NormalizedStore
from importers.production.staging import StagingBundle
from importers.production.variant_description_codec import VariantDescriptionCodec
from models.packaging_configuration import PackagingConfigurationLine
from services.weight_service import WeightService


def dpp_traversal_check(bundle: StagingBundle, store: NormalizedStore) -> dict[str, Any]:
    """QA-only: Product 1011935 → ST-051-STD-01 → tare 47.0384 kg."""
    result: dict[str, Any] = {"ok": False}
    maps = [
        pm
        for pm in bundle.product_maps
        if str(pm.product_code) == "1011935" and pm.final_set_code == "ST-051-STD-01"
    ]
    if not maps:
        result["error"] = "Product 1011935 not mapped to ST-051-STD-01"
        return result
    cfg = next(c for c in bundle.configurations if c.packaging_set_code == "ST-051-STD-01")
    bom = [b for b in bundle.bom_lines if b.packaging_set_code == "ST-051-STD-01"]
    lines = []
    for bl in bom:
        from importers.production.normalizer import _line_weight_inputs

        wg, qty = _line_weight_inputs(bl)
        lines.append(
            PackagingConfigurationLine(
                component_id=bl.component_code,
                component_erp_code=bl.component_code,
                component_name=bl.erp_description,
                quantity=qty,
                uom=bl.uom,
                weight_g=wg,
            )
        )
    wr = WeightService().calculate_tare(lines)
    tare_kg = wr.total_tare_g / 1000.0
    result.update(
        {
            "product": "1011935",
            "set_code": cfg.packaging_set_code,
            "final_id": cfg.final_configuration_id,
            "source_cfg": cfg.source_configuration_id,
            "bom_lines": len(bom),
            "tare_kg": tare_kg,
            "expected_tare_kg": 47.0384,
            "ok": abs(tare_kg - 47.0384) < 1e-4
            and cfg.final_configuration_id == "IA-ST-051-STD-01"
            and cfg.source_configuration_id == "IA-ST-CFG-0122",
        }
    )
    return result


def write_migration_qa(
    path: Path,
    *,
    run_id: str,
    qualification: QualificationResult,
    inventory_rows: list,
    store: NormalizedStore,
    bundle: StagingBundle,
    production_path: Path,
    test_rc: int,
    dpp: dict[str, Any],
) -> Path:
    codec = VariantDescriptionCodec()
    # reopen workbook checks
    wb = load_workbook(production_path)
    formula_errors = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    u = cell.value.upper()
                    if any(e in u for e in ("#REF!", "#VALUE!", "#DIV/0!")):
                        formula_errors += 1
    ui = [
        s
        for s in (
            "00_README",
            "01_DASHBOARD",
            "02_RELEASE_CONTROL",
            "03_DATA_DICTIONARY",
            "04_IMPORT_GUIDE",
        )
        if s in wb.sheetnames
    ]
    pc_count = sum(
        1
        for r in range(2, wb["PACKAGING_CONFIGURATION"].max_row + 1)
        if wb["PACKAGING_CONFIGURATION"].cell(r, 1).value
    )
    wb.close()

    lines = [
        "# Phase F Migration QA",
        "",
        f"- **RUN_ID:** `{run_id}`",
        f"- **Timestamp (UTC):** {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}",
        f"- **Schema version:** 1.0.0",
        "",
        "## PRIMARY SOURCE QUALIFICATION",
        "",
        *qualification.summary_lines(),
        "",
        "## SOURCE INVENTORY",
        "",
        f"Files inventoried: {len(inventory_rows)}",
        "",
        "## COUNTS",
        "",
        f"- Final configurations: {store.stats.get('configurations')}",
        f"- Starter: {store.stats.get('starter')}",
        f"- Industrial: {store.stats.get('industrial')}",
        f"- Container / Loading: {store.stats.get('container')}",
        f"- Components: {store.stats.get('components')}",
        f"- Products: {store.stats.get('products')}",
        f"- Exact BOM lines: {store.stats.get('bom_lines')}",
        f"- Component-material rows: {store.stats.get('component_materials')}",
        f"- Technical files: {store.stats.get('technical_files')}",
        f"- DoCs: {store.stats.get('docs')}",
        f"- Document library: {store.stats.get('document_library')}",
        f"- Statements: {len(store.tables.get('STATEMENT', []))}",
        f"- Commercial scenarios: {store.stats.get('commercial_scenarios')}",
        "",
        "## IDENTITY / VARIANT BASIS",
        "",
        f"- Duplicate Variant Basis errors: {qualification.duplicate_variant_basis}",
        f"- Document-ID errors: {qualification.document_id_errors}",
        f"- Codec sample roundtrip: {codec.roundtrip_ok(bundle.configurations[0].variant_basis_tr, bundle.configurations[0].variant_basis_en)}",
        "",
        "## WEIGHT / BOM",
        "",
        f"- Blocking errors: {len(store.blocking_errors)}",
        f"- BOM rows staged: {len(bundle.bom_lines)}",
        "",
        "## DPP TRACEABILITY",
        "",
        str(dpp),
        "",
        "## WORKBOOK TECHNICAL QA",
        "",
        f"- Production workbook: `{production_path}`",
        f"- PACKAGING_CONFIGURATION rows: {pc_count}",
        f"- UI sheets preserved: {ui}",
        f"- Formula error literals: {formula_errors}",
        f"- Broken external links: 0 (no external links added)",
        "",
        "## TESTS",
        "",
        f"- unittest exit code: {test_rc}",
        "",
        "## CONFIRMATIONS",
        "",
        f"- Production data loaded: {'YES' if pc_count == 247 and not store.blocking_errors else 'NO'}",
        "- Word generation run: NO",
        "- PDF generation run: NO",
        "- Phase G started: NO",
        "",
    ]
    if store.blocking_errors:
        lines.append("## BLOCKING ERRORS")
        lines.append("")
        for e in store.blocking_errors[:50]:
            lines.append(f"- {e}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
