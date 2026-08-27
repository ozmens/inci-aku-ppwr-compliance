"""Content-based Level-1 Golden Register qualification gate."""

from __future__ import annotations

import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.id_service import IdService


VARIANT_SUFFIX_RE = re.compile(r"^.+-\d{2}$")


@dataclass
class QualificationResult:
    passed: bool
    file_path: str
    total_configurations: int = 0
    starter_count: int = 0
    industrial_count: int = 0
    container_count: int = 0
    exact_bom_rows: int = 0
    product_map_rows: int = 0
    invalid_suffixes: int = 0
    duplicate_set_codes: int = 0
    duplicate_final_ids: int = 0
    duplicate_variant_basis: int = 0
    document_id_errors: int = 0
    missing_bom_configs: int = 0
    st051_01: dict[str, Any] = field(default_factory=dict)
    st051_02: dict[str, Any] = field(default_factory=dict)
    failures: list[str] = field(default_factory=list)

    def summary_lines(self) -> list[str]:
        status = "PASS" if self.passed else "FAIL"
        return [
            f"PRIMARY SOURCE QUALIFICATION: {status}",
            f"File: {self.file_path}",
            f"Total configurations: {self.total_configurations}",
            f"Starter: {self.starter_count}",
            f"Industrial: {self.industrial_count}",
            f"Container / Loading: {self.container_count}",
            f"Exact BOM rows: {self.exact_bom_rows}",
            f"Product map rows: {self.product_map_rows}",
            f"Invalid suffixes: {self.invalid_suffixes}",
            f"Duplicate set codes: {self.duplicate_set_codes}",
            f"Duplicate final IDs: {self.duplicate_final_ids}",
            f"Duplicate Variant Basis within family: {self.duplicate_variant_basis}",
            f"Document-ID errors: {self.document_id_errors}",
            f"Configs missing BOM: {self.missing_bom_configs}",
            f"ST-051-STD-01: {self.st051_01}",
            f"ST-051-STD-02: {self.st051_02}",
            *(f"FAIL: {f}" for f in self.failures),
        ]


def _norm_basis(text: str | None) -> str:
    if not text:
        return ""
    s = unicodedata.normalize("NFKC", str(text))
    return re.sub(r"\s+", " ", s).strip().casefold()


def _parent(code: str) -> str:
    return re.sub(r"-\d{2}$", "", code)


def _sheet_dicts(ws) -> tuple[list[str], list[dict]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or first == "":
            # still try packaging set col if present
            if "Packaging Set Code" in idx and not ws.cell(r, idx["Packaging Set Code"]).value:
                continue
            if "Product Code" in idx and not ws.cell(r, idx["Product Code"]).value:
                continue
            if "Packaging Set Code" not in idx and "Product Code" not in idx:
                continue
        row = {h: ws.cell(r, idx[h]).value for h in headers}
        # skip fully empty
        if all(v is None or v == "" for v in row.values()):
            continue
        rows.append(row)
    return headers, rows


def find_register_candidate(production_dir: Path) -> Path | None:
    if not production_dir.exists():
        return None
    for p in sorted(production_dir.glob("*.xlsx")):
        if "Final_Configuration_Register" in p.name:
            return p
    return None


def qualify_golden_register(path: Path) -> QualificationResult:
    result = QualificationResult(passed=False, file_path=str(path))
    if not path.exists():
        result.failures.append("File does not exist")
        return result

    wb = load_workbook(path, data_only=True)
    required = {"01_FINAL_CONFIG_MASTER", "02_PRODUCT_MAP", "03_BOM_MASTER"}
    if not required.issubset(set(wb.sheetnames)):
        result.failures.append(f"Missing sheets: {sorted(required - set(wb.sheetnames))}")
        wb.close()
        return result

    _, configs = _sheet_dicts(wb["01_FINAL_CONFIG_MASTER"])
    # Keep only rows with Packaging Set Code
    configs = [r for r in configs if r.get("Packaging Set Code")]
    _, products = _sheet_dicts(wb["02_PRODUCT_MAP"])
    products = [r for r in products if r.get("Product Code")]
    _, bom = _sheet_dicts(wb["03_BOM_MASTER"])
    bom = [r for r in bom if r.get("Packaging Set Code") and r.get("Component Code")]

    result.total_configurations = len(configs)
    fam = Counter(str(r.get("Family") or "").upper() for r in configs)
    result.starter_count = fam.get("STARTER", 0)
    result.industrial_count = fam.get("INDUSTRIAL", 0)
    result.container_count = fam.get("CONTAINER", 0)
    result.exact_bom_rows = len(bom)
    result.product_map_rows = len(products)

    if result.total_configurations != 247:
        result.failures.append(
            f"Expected 247 configs, found {result.total_configurations} (obsolete aggregated register?)"
        )
    if result.starter_count != 240:
        result.failures.append(f"Starter expected 240, found {result.starter_count}")
    if result.industrial_count != 3:
        result.failures.append(f"Industrial expected 3, found {result.industrial_count}")
    if result.container_count != 4:
        result.failures.append(f"Container expected 4, found {result.container_count}")

    set_codes = [str(r["Packaging Set Code"]) for r in configs]
    final_ids = [str(r["Final Configuration ID"]) for r in configs]
    invalid = [c for c in set_codes if not VARIANT_SUFFIX_RE.match(c)]
    result.invalid_suffixes = len(invalid)
    if invalid:
        result.failures.append(f"Invalid variant suffixes: {invalid[:10]}")

    dup_sets = [k for k, v in Counter(set_codes).items() if v > 1]
    dup_fins = [k for k, v in Counter(final_ids).items() if v > 1]
    result.duplicate_set_codes = len(dup_sets)
    result.duplicate_final_ids = len(dup_fins)
    if dup_sets:
        result.failures.append(f"Duplicate set codes: {dup_sets[:5]}")
    if dup_fins:
        result.failures.append(f"Duplicate final IDs: {dup_fins[:5]}")

    # Variant Basis columns (TR header may be localized)
    tr_key = next(
        (k for k in configs[0].keys() if k and "Variant Basis" in str(k) and "TR" in str(k)),
        None,
    )
    if tr_key is None:
        tr_key = next(
            (k for k in configs[0].keys() if k and "Ay" in str(k) and "Variant" in str(k)),
            None,
        )
    en_key = "Variant Basis (EN)"
    by_parent: dict[str, list] = defaultdict(list)
    for r in configs:
        by_parent[_parent(str(r["Packaging Set Code"]))].append(r)
    dup_vb = 0
    for parent, items in by_parent.items():
        bases = [_norm_basis(i.get(en_key) or (i.get(tr_key) if tr_key else None)) for i in items]
        for b, n in Counter(bases).items():
            if b and n > 1:
                dup_vb += 1
                result.failures.append(f"Duplicate Variant Basis in {parent}: {b}")
    result.duplicate_variant_basis = dup_vb

    # Document IDs
    ids = IdService()
    doc_errs = 0
    for r in configs:
        sc = str(r["Packaging Set Code"])
        expected = {
            "Technical File ID": ids.technical_file_id(sc),
            "EU DoC ID": ids.doc_id(sc),
            "Label ID": ids.label_id(sc),
            "Statement ID": ids.statement_id(sc),
        }
        for field, exp in expected.items():
            if str(r.get(field)) != exp:
                doc_errs += 1
                if doc_errs <= 5:
                    result.failures.append(f"Doc ID mismatch {sc} {field}: {r.get(field)} != {exp}")
        # Final configuration ID must match IdService
        if str(r.get("Final Configuration ID")) != ids.final_configuration_id(sc):
            doc_errs += 1
            result.failures.append(
                f"Final ID mismatch {sc}: {r.get('Final Configuration ID')}"
            )
    result.document_id_errors = doc_errs

    bom_sets = {str(r["Packaging Set Code"]) for r in bom}
    missing = [c for c in set_codes if c not in bom_sets]
    result.missing_bom_configs = len(missing)
    if missing:
        result.failures.append(f"Configs without BOM: {missing[:10]}")

    # ST-051 control tests
    by_code = {str(r["Packaging Set Code"]): r for r in configs}
    a = by_code.get("ST-051-STD-01")
    b = by_code.get("ST-051-STD-02")
    bom_a = [r for r in bom if str(r["Packaging Set Code"]) == "ST-051-STD-01"]
    bom_b = [r for r in bom if str(r["Packaging Set Code"]) == "ST-051-STD-02"]

    def st_check(label: str, row, expected: dict, bom_rows) -> dict:
        info: dict[str, Any] = {"present": row is not None}
        if not row:
            result.failures.append(f"{label} missing from master")
            return info
        for k, v in expected.items():
            actual = row.get(k)
            if k == "Packaging Mass kg":
                ok = abs(float(actual) - float(v)) < 1e-6
            else:
                ok = str(actual) == str(v)
            info[k] = actual
            if not ok:
                result.failures.append(f"{label} {k}: {actual} != {v}")
        comps = {str(r["Component Code"]) for r in bom_rows}
        info["bom_components"] = sorted(comps)
        info["source_from_bom"] = sorted({str(r["Source Configuration ID"]) for r in bom_rows})
        return info

    result.st051_01 = st_check(
        "ST-051-STD-01",
        a,
        {
            "Final Configuration ID": "IA-ST-051-STD-01",
            "Source Configuration ID": "IA-ST-CFG-0122",
            "Packaging Mass kg": 47.0384,
        },
        bom_a,
    )
    if "4000782" not in result.st051_01.get("bom_components", []):
        result.failures.append("ST-051-STD-01 missing distinguishing ERP 4000782")
    result.st051_02 = st_check(
        "ST-051-STD-02",
        b,
        {
            "Final Configuration ID": "IA-ST-051-STD-02",
            "Source Configuration ID": "IA-ST-CFG-0123",
        },
        bom_b,
    )
    if "4001108" not in result.st051_02.get("bom_components", []):
        result.failures.append("ST-051-STD-02 missing distinguishing ERP 4001108")
    if set(result.st051_01.get("bom_components", [])) == set(
        result.st051_02.get("bom_components", [])
    ):
        result.failures.append("ST-051-STD-01 and STD-02 BOM component sets identical")

    # Product map SKUs
    for sku in ("1011935", "1011936", "1011939"):
        hits = [
            r
            for r in products
            if str(r.get("Product Code")) == sku
            and str(r.get("Final Set Code")) == "ST-051-STD-01"
        ]
        if not hits:
            result.failures.append(f"Product map missing {sku} → ST-051-STD-01")

    # Variant basis must not be generic Variant 01 only
    if a and en_key in a:
        vb = str(a.get(en_key) or "")
        if re.fullmatch(r"Variant\s*\d+", vb.strip(), flags=re.I):
            result.failures.append("ST-051-STD-01 Variant Basis is generic")

    result.passed = len(result.failures) == 0
    wb.close()
    return result
