"""Level-2 controlled PIMS data package importer (lineage / discrepancy only)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .source_inventory import sha256_file
from .staging import StagingBundle


def import_legacy_pims(path: Path, bundle: StagingBundle) -> StagingBundle:
    """Read Level-2 metadata; never overwrite Level-1 final configuration identity."""
    if not path.exists():
        return bundle
    file_hash = sha256_file(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    bundle.level2_meta = {
        "file_name": path.name,
        "sha256": file_hash,
        "sheets": list(wb.sheetnames),
        "role": "LEVEL_2_CONTROLLED_PIMS",
    }

    # Component master enrichment (material family / weights) without changing IDs
    if "02_COMPONENT_MASTER" in wb.sheetnames:
        ws = wb["02_COMPONENT_MASTER"]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        h = {str(x): i for i, x in enumerate(headers) if x}
        enriched = 0
        for row in rows:
            code = row[h.get("source_component_code", -1)] if "source_component_code" in h else None
            if not code:
                continue
            erp = str(code)
            if erp not in bundle.components:
                continue
            comp = bundle.components[erp]
            mf = row[h["material_family"]] if "material_family" in h else None
            if mf and not comp.material_family:
                comp.material_family = str(mf)
                enriched += 1
            uw = row[h["unit_weight_kg"]] if "unit_weight_kg" in h else None
            if comp.unit_weight_kg is None and uw not in (None, ""):
                try:
                    comp.unit_weight_kg = float(uw)
                except (TypeError, ValueError):
                    pass
            desc = row[h["description"]] if "description" in h else None
            if desc and str(desc) != comp.description and str(desc) not in comp.aliases:
                # Level-1 description stays canonical; Level-2 is alias
                comp.aliases.append(str(desc))
                bundle.discrepancies.append(
                    {
                        "severity": "INFO",
                        "object_type": "COMPONENT",
                        "object_id": erp,
                        "field": "description",
                        "golden_value": comp.description,
                        "secondary_value": str(desc),
                        "source_file": path.name,
                        "source_reference": "02_COMPONENT_MASTER",
                        "resolution": "KEEP_GOLDEN_CANONICAL",
                        "resolution_basis": "Level-1 BOM description authoritative; Level-2 stored as alias",
                        "blocking": False,
                    }
                )
        bundle.level2_meta["components_enriched"] = enriched

    # Compare configuration counts if present
    if "03_CONFIGURATION_MASTER" in wb.sheetnames:
        ws = wb["03_CONFIGURATION_MASTER"]
        n = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True)) - 0
        # recount properly
        ws2 = wb["03_CONFIGURATION_MASTER"]
        count = 0
        first = True
        for row in ws2.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row and row[0]:
                count += 1
        if count and count != len(bundle.configurations):
            bundle.discrepancies.append(
                {
                    "severity": "WARN",
                    "object_type": "CONFIGURATION_COUNT",
                    "object_id": "LEVEL2_VS_GOLDEN",
                    "field": "configuration_count",
                    "golden_value": len(bundle.configurations),
                    "secondary_value": count,
                    "source_file": path.name,
                    "source_reference": "03_CONFIGURATION_MASTER",
                    "resolution": "KEEP_GOLDEN",
                    "resolution_basis": "Level-1 Golden Register is authoritative for final configuration identity",
                    "blocking": False,
                }
            )
        bundle.level2_meta["configuration_master_rows"] = count

    wb.close()
    return bundle
