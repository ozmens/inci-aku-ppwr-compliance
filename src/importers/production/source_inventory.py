"""Phase F source inventory (SHA-256, roles, sheet lists)."""

from __future__ import annotations

import csv
import hashlib
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class SourceInventoryRow:
    file_name: str
    relative_path: str
    sha256: str
    file_size: int
    modified_utc: str
    sheet_names: str
    source_role: str
    source_priority: int
    record_counts: str
    migration_status: str
    notes: str = ""


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []
    wb = load_workbook(path, read_only=True, data_only=False)
    names = list(wb.sheetnames)
    wb.close()
    return names


def classify_source(path: Path, *, qualification_pass: bool | None = None) -> tuple[str, int, str]:
    name = path.name.upper()
    if "FINAL_CONFIGURATION_REGISTER" in name.replace(" ", "_") or (
        "FINAL_CONFIGURATION_REGISTER" in name
    ):
        role = "LEVEL_1_GOLDEN_REGISTER"
        prio = 1
        note = "Content-qualified" if qualification_pass else "Candidate"
        return role, prio, note
    if "PIMS_DATA_PACKAGE" in name or "OEM_ART5" in name or "BOM_CLOSED" in name:
        return "LEVEL_2_CONTROLLED_PIMS", 2, "Controlled PIMS/BOM lineage package"
    if path.suffix.lower() == ".7z" or "TEDAR" in name:
        return "LEVEL_4_EVIDENCE_ARCHIVE", 4, "Supplier evidence archive — do not modify"
    return "LEVEL_3_OPERATIONAL_SOURCE", 3, "Operational lineage source"


def build_inventory(
    production_dir: Path,
    *,
    qualification_pass: bool | None = None,
    record_count_hints: dict[str, str] | None = None,
) -> list[SourceInventoryRow]:
    hints = record_count_hints or {}
    rows: list[SourceInventoryRow] = []
    if not production_dir.exists():
        return rows
    for path in sorted(production_dir.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        rel = path.relative_to(production_dir.parent.parent).as_posix()
        try:
            rel = path.relative_to(production_dir.parents[1]).as_posix()
        except Exception:
            rel = str(path)
        # Prefer project-relative
        try:
            from pathlib import Path as P

            # production_dir is .../input/production → project root parents[1]
            root = production_dir.parent.parent
            rel = path.relative_to(root).as_posix()
        except Exception:
            rel = path.as_posix()

        role, prio, note = classify_source(path, qualification_pass=qualification_pass)
        sheets = _sheet_names(path)
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        )
        rows.append(
            SourceInventoryRow(
                file_name=path.name,
                relative_path=rel,
                sha256=sha256_file(path),
                file_size=path.stat().st_size,
                modified_utc=mtime,
                sheet_names=" | ".join(sheets) if sheets else "(binary/archive)",
                source_role=role,
                source_priority=prio,
                record_counts=hints.get(path.name, ""),
                migration_status="INVENTORIED",
                notes=note,
            )
        )
    return rows


def write_inventory_csv(rows: list[SourceInventoryRow], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(asdict(rows[0]).keys()) if rows else [
        "file_name",
        "relative_path",
        "sha256",
        "file_size",
        "modified_utc",
        "sheet_names",
        "source_role",
        "source_priority",
        "record_counts",
        "migration_status",
        "notes",
    ]
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for row in rows:
            w.writerow(asdict(row))
    return path


def write_inventory_md(rows: list[SourceInventoryRow], path: Path, *, title: str) -> Path:
    lines = [
        f"# {title}",
        "",
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}",
        "",
        "| File | Role | Priority | SHA-256 | Size | Sheets / notes |",
        "|------|------|----------|---------|------|----------------|",
    ]
    for r in rows:
        lines.append(
            f"| `{r.file_name}` | {r.source_role} | {r.source_priority} | "
            f"`{r.sha256[:16]}…` | {r.file_size} | {r.record_counts or r.notes} |"
        )
    lines.append("")
    lines.append("## Full hashes")
    lines.append("")
    for r in rows:
        lines.append(f"- `{r.file_name}`: `{r.sha256}`")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
