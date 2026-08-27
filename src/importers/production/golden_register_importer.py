"""Level-1 Golden Final Configuration Register importer (read-only → staging)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .source_inventory import sha256_file
from .staging import (
    StageLineage,
    StagedBomLine,
    StagedComponent,
    StagedConfiguration,
    StagedProductMap,
    StagingBundle,
)


def _headers(ws) -> dict[str, int]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h}


def _variant_tr_key(headers: dict[str, int]) -> str:
    for k in headers:
        if "Variant Basis" in str(k) and "TR" in str(k):
            return k
        if "Ay" in str(k) and "Variant" in str(k):
            return k
    raise KeyError("Variant Basis TR column not found")


def import_golden_register(path: Path, bundle: StagingBundle | None = None) -> StagingBundle:
    bundle = bundle or StagingBundle()
    file_hash = sha256_file(path)
    wb = load_workbook(path, data_only=True)

    ws = wb["01_FINAL_CONFIG_MASTER"]
    idx = _headers(ws)
    tr_key = _variant_tr_key(idx)
    for r in range(2, ws.max_row + 1):
        set_code = ws.cell(r, idx["Packaging Set Code"]).value
        if not set_code:
            continue
        lineage = StageLineage(
            source_file=path.name,
            source_sheet="01_FINAL_CONFIG_MASTER",
            source_row=r,
            source_reference=f"{path.name}!01_FINAL_CONFIG_MASTER!R{r}",
            source_configuration_id=str(ws.cell(r, idx["Source Configuration ID"]).value),
            source_hash=file_hash,
        )
        bundle.configurations.append(
            StagedConfiguration(
                family=str(ws.cell(r, idx["Family"]).value).upper(),
                packaging_set_code=str(set_code),
                final_configuration_id=str(ws.cell(r, idx["Final Configuration ID"]).value),
                source_configuration_id=str(ws.cell(r, idx["Source Configuration ID"]).value),
                configuration_name=str(ws.cell(r, idx["Configuration Name"]).value or ""),
                packaging_mass_kg=float(ws.cell(r, idx["Packaging Mass kg"]).value),
                variant_basis_tr=str(ws.cell(r, idx[tr_key]).value or ""),
                variant_basis_en=str(ws.cell(r, idx["Variant Basis (EN)"]).value or ""),
                technical_file_id=str(ws.cell(r, idx["Technical File ID"]).value),
                doc_id=str(ws.cell(r, idx["EU DoC ID"]).value),
                label_id=str(ws.cell(r, idx["Label ID"]).value),
                statement_id=str(ws.cell(r, idx["Statement ID"]).value),
                revision=str(ws.cell(r, idx["Revision"]).value or "Rev.00"),
                evidence_basis=str(ws.cell(r, idx["Evidence Basis"]).value or ""),
                nominal_product_qty=_float_or_none(ws.cell(r, idx["Nominal Product Qty"]).value),
                pallet_class=str(ws.cell(r, idx.get("Pallet Class", 1)).value or "")
                if "Pallet Class" in idx
                else None,
                lineage=lineage,
            )
        )

    bom = wb["03_BOM_MASTER"]
    bidx = _headers(bom)
    for r in range(2, bom.max_row + 1):
        set_code = bom.cell(r, bidx["Packaging Set Code"]).value
        comp = bom.cell(r, bidx["Component Code"]).value
        if not set_code or not comp:
            continue
        erp = str(comp)
        desc = str(bom.cell(r, bidx["ERP Description"]).value or "")
        uw = bom.cell(r, bidx["Unit Weight kg"]).value
        uw_f = float(uw) if uw is not None and uw != "" else None
        qty = float(bom.cell(r, bidx["Quantity"]).value or 0)
        lw = float(bom.cell(r, bidx["Line Weight kg"]).value or 0)
        group = str(bom.cell(r, bidx["Component Group"]).value or "") or None
        uom = str(bom.cell(r, bidx["UOM"]).value or "")
        basis = str(bom.cell(r, bidx["Weight Basis"]).value or "")
        lineage = StageLineage(
            source_file=str(bom.cell(r, bidx["Source File"]).value or path.name),
            source_sheet="03_BOM_MASTER",
            source_row=r,
            source_reference=str(bom.cell(r, bidx["Source Reference"]).value or f"R{r}"),
            source_configuration_id=str(bom.cell(r, bidx["Source Configuration ID"]).value),
            source_hash=file_hash,
        )
        bundle.bom_lines.append(
            StagedBomLine(
                packaging_set_code=str(set_code),
                final_configuration_id=str(bom.cell(r, bidx["Final Configuration ID"]).value),
                source_configuration_id=str(bom.cell(r, bidx["Source Configuration ID"]).value),
                component_code=erp,
                component_group=group,
                erp_description=desc,
                quantity=qty,
                uom=uom,
                unit_weight_kg=uw_f,
                line_weight_kg=lw,
                weight_basis=basis,
                lineage=lineage,
            )
        )
        _upsert_component(bundle, erp, desc, group, uw_f, lineage)

    pm = wb["02_PRODUCT_MAP"]
    pidx = _headers(pm)
    for r in range(2, pm.max_row + 1):
        pc = pm.cell(r, pidx["Product Code"]).value
        if not pc:
            continue
        lineage = StageLineage(
            source_file=path.name,
            source_sheet="02_PRODUCT_MAP",
            source_row=r,
            source_reference=f"{path.name}!02_PRODUCT_MAP!R{r}",
            source_configuration_id=_str_or_none(pm.cell(r, pidx["Source Configuration ID"]).value),
            source_hash=file_hash,
        )
        bundle.product_maps.append(
            StagedProductMap(
                product_code=str(pc),
                product_description=str(pm.cell(r, pidx["Product Description"]).value or ""),
                battery_type=_str_or_none(pm.cell(r, pidx["Battery Type"]).value),
                customer_market=_str_or_none(pm.cell(r, pidx["Customer / Market"]).value),
                nominal_qty=_float_or_none(pm.cell(r, pidx["Nominal Qty"]).value),
                source_configuration_id=_str_or_none(
                    pm.cell(r, pidx["Source Configuration ID"]).value
                ),
                final_set_code=_str_or_none(pm.cell(r, pidx["Final Set Code"]).value),
                final_configuration_id=_str_or_none(
                    pm.cell(r, pidx["Final Configuration ID"]).value
                ),
                status=str(pm.cell(r, pidx["Status"]).value or ""),
                lineage=lineage,
            )
        )

    if "04_DOMESTIC_ONLY" in wb.sheetnames:
        dom = wb["04_DOMESTIC_ONLY"]
        didx = _headers(dom)
        for r in range(2, dom.max_row + 1):
            sid = dom.cell(r, didx["Source Configuration ID"]).value
            if not sid:
                continue
            bundle.domestic_source_configs.append(
                {
                    "source_configuration_id": str(sid),
                    "name": dom.cell(r, didx.get("Configuration Name", 1)).value,
                    "status": dom.cell(r, didx.get("Status", 1)).value,
                    "source_file": path.name,
                    "source_row": r,
                }
            )

    wb.close()
    return bundle


def _upsert_component(
    bundle: StagingBundle,
    erp: str,
    desc: str,
    group: str | None,
    uw: float | None,
    lineage: StageLineage,
) -> None:
    existing = bundle.components.get(erp)
    if existing is None:
        bundle.components[erp] = StagedComponent(
            erp_code=erp,
            description=desc,
            component_group=group,
            unit_weight_kg=uw,
            lineage=lineage,
        )
        return
    if desc and desc != existing.description:
        if desc not in existing.aliases:
            existing.aliases.append(desc)
    # Prefer non-null unit weight; if conflict, keep first non-null from PROCESSED basis later
    if existing.unit_weight_kg is None and uw is not None:
        existing.unit_weight_kg = uw


def _float_or_none(v) -> float | None:
    if v is None or v == "":
        return None
    return float(v)


def _str_or_none(v) -> str | None:
    if v is None or v == "":
        return None
    return str(v)
