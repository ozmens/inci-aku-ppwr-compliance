"""
Clean Excel import templates for İnci Akü daily / bulk loading.

Headers match frozen schema columns (business columns only).
No formulas. No architecture changes.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# Business-facing columns for each import sheet (exclude audit stamps)
IMPORT_SPECS: dict[str, tuple[str, ...]] = {
    "Component_Master": (
        "COMPONENT_ID",
        "COMPONENT_CODE",
        "COMPONENT_NAME",
        "COMPONENT_TYPE_ID",
        "PACKAGING_LEVEL_ID",
        "PACKAGING_FUNCTION_ID",
        "OWNERSHIP_TYPE_ID",
        "SUPPLIER_ID",
        "WEIGHT_G",
        "LENGTH_MM",
        "WIDTH_MM",
        "HEIGHT_MM",
        "RECYCLED_CONTENT_PCT",
        "RECYCLABILITY_CLASS_ID",
        "SPEC_REF",
        "STATUS_ID",
        "EFFECTIVE_FROM",
        "EFFECTIVE_TO",
        "EXTERNAL_REF",
        "NOTES",
    ),
    "Packaging_Configuration": (
        "PACKAGING_CONFIGURATION_ID",
        "CONFIG_GROUP_CODE",
        "REVISION_NO",
        "PACKAGING_CONFIGURATION_NAME",
        "DESCRIPTION",
        "SUPERSEDES_ID",
        "STATUS_ID",
        "EFFECTIVE_FROM",
        "EFFECTIVE_TO",
        "NOTES",
    ),
    "Packaging_Configuration_Lines": (
        "PACKAGING_CONFIGURATION_LINE_ID",
        "PACKAGING_CONFIGURATION_ID",
        "COMPONENT_ID",
        "QUANTITY",
        "LINE_ROLE_ID",
        "SORT_ORDER",
        "IS_OPTIONAL",
        "NOTES",
    ),
    "Product_Master": (
        "PRODUCT_ID",
        "PRODUCT_CODE",
        "PRODUCT_NAME",
        "PRODUCT_CATEGORY_ID",
        "NET_WEIGHT_G",
        "LENGTH_MM",
        "WIDTH_MM",
        "HEIGHT_MM",
        "STATUS_ID",
        "EFFECTIVE_FROM",
        "EFFECTIVE_TO",
        "EXTERNAL_REF",
        "NOTES",
    ),
    "Commercial_Scenarios": (
        "COMMERCIAL_SCENARIO_ID",
        "COMMERCIAL_SCENARIO_CODE",
        "COMMERCIAL_SCENARIO_NAME",
        "SCENARIO_TYPE_ID",
        "PRODUCT_ID",
        "TRANSPORT_CONFIGURATION_ID",
        "CUSTOMER_ID",
        "DESTINATION_COUNTRY_ID",
        "INCOTERM_ID",
        "TRANSPORT_MODE_ID",
        "STATUS_ID",
        "VALID_FROM",
        "VALID_TO",
        "NOTES",
    ),
    "Shipment_Import": (
        "SHIPMENT_ID",
        "SHIPMENT_NUMBER",
        "COMMERCIAL_SCENARIO_ID",
        "PLANT_ID",
        "SHIP_DATE",
        "QTY_PRODUCT_UNITS",
        "PACKAGING_CONFIGURATION_ID",
        "TRANSPORT_CONFIGURATION_ID",
        "DESTINATION_COUNTRY_ID",
        "TRANSPORT_MODE_ID",
        "STATUS_ID",
        "EXTERNAL_REF",
        "NOTES",
    ),
    "Supplier_Documents": (
        "DOCUMENT_ID",
        "DOCUMENT_CODE",
        "DOCUMENT_TITLE",
        "DOCUMENT_TYPE_ID",
        "FILE_URI",
        "FILE_HASH",
        "ISSUE_DATE",
        "STATUS_ID",
        "NOTES",
        "LINK_COMPONENT_ID",
        "LINK_PACKAGING_CONFIGURATION_ID",
        "LINK_TECHNICAL_FILE_ID",
    ),
}

# Short instructions per sheet (row 1 note; headers on row 2)
SHEET_NOTES: dict[str, str] = {
    "Component_Master": "Yellow columns are required for PPWR. WEIGHT_G is mandatory. Paste into Component Master table.",
    "Packaging_Configuration": "PRIMARY OBJECT. One row = one physical packaging arrangement revision.",
    "Packaging_Configuration_Lines": "Recipe lines. QUANTITY = components per packed product unit.",
    "Product_Master": "ERP reference only — not the PPWR primary object. Links via Commercial Scenarios.",
    "Commercial_Scenarios": "Incoterms + customer + country. NEVER changes Technical Files. Selects DoC variant.",
    "Shipment_Import": "EXTERNAL_REF = Lot Number. Customer/Country/Incoterms come from COMMERCIAL_SCENARIO_ID.",
    "Supplier_Documents": "Maps to DOCUMENT_LIBRARY. Link to components/TF via DOCUMENT_LINK in workbook after import.",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FFF4CC")
NOTE_FONT = Font(name="Calibri", size=10, italic=True)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(str(cell.value or "")) + 2))


def generate_import_templates(output_dir: Path) -> Path:
    """Write Import_Templates.xlsx under templates/import/."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "Import_Templates.xlsx"

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Index sheet
    idx = wb.create_sheet("README", 0)
    idx["A1"] = "İNCI AKÜ — PPWR Import Templates"
    idx["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    idx["A3"] = "Use one sheet per entity. Copy data rows under the header into the matching workbook table."
    idx["A4"] = "Do not rename columns. IDs must be unique. Lookups use numeric FK IDs from LKP_* sheets."
    idx["A5"] = "Shipment EXTERNAL_REF = Lot Number. Technical Files are NOT imported here — create per Packaging Configuration."
    idx["A7"] = "Sheets"
    idx["A7"].font = Font(bold=True)
    for i, name in enumerate(IMPORT_SPECS, start=8):
        idx.cell(row=i, column=1, value=name)
        idx.cell(row=i, column=2, value=SHEET_NOTES.get(name, ""))
    idx.column_dimensions["A"].width = 36
    idx.column_dimensions["B"].width = 80

    for sheet_name, headers in IMPORT_SPECS.items():
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = SHEET_NOTES.get(sheet_name, "")
        ws["A1"].fill = NOTE_FILL
        ws["A1"].font = NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(6, len(headers)))

        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        _style_header_row(ws, 2, len(headers))

        # Empty stub rows for typing
        for r in range(3, 13):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = THIN

        last_col = chr(64 + min(len(headers), 26)) if len(headers) <= 26 else "Z"
        # Safe table ref using openpyxl utils
        from openpyxl.utils import get_column_letter

        end = get_column_letter(len(headers))
        table = Table(displayName=f"IMP_{sheet_name}", ref=f"A2:{end}12")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(table)
        ws.freeze_panes = "A3"
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

    wb.save(path)
    wb.close()
    return path
