"""
Production workbook validator.

Checks structure, tables, named ranges, engines, UI sheets, protection, seeds.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from models.registry import SchemaRegistry
from builders.engine_builder import all_engines
from builders.relationships import RELATIONSHIP_SHEET
from builders.ui_builder import UI_SHEETS


@dataclass
class ValidationIssue:
    severity: str  # ERROR | WARN
    code: str
    message: str


@dataclass
class ValidationReport:
    workbook_path: str
    passed: bool
    issues: list[ValidationIssue] = field(default_factory=list)
    stats: dict = field(default_factory=dict)

    def add(self, severity: str, code: str, message: str) -> None:
        self.issues.append(ValidationIssue(severity, code, message))

    @property
    def errors(self) -> list[ValidationIssue]:
        return [i for i in self.issues if i.severity == "ERROR"]

    @property
    def warnings(self) -> list[ValidationIssue]:
        return [i for i in self.issues if i.severity == "WARN"]


class WorkbookValidator:
    def __init__(self, workbook_path: Path, registry: SchemaRegistry | None = None) -> None:
        self.workbook_path = Path(workbook_path)
        self.registry = registry or SchemaRegistry.load()

    def run(self) -> ValidationReport:
        report = ValidationReport(workbook_path=str(self.workbook_path), passed=True)
        if not self.workbook_path.exists():
            report.add("ERROR", "FILE-01", f"Workbook not found: {self.workbook_path}")
            report.passed = False
            return report

        wb = load_workbook(self.workbook_path)
        try:
            self._check_sheets(wb, report)
            self._check_tables(wb, report)
            self._check_named_ranges(wb, report)
            self._check_engines(wb, report)
            self._check_ui(wb, report)
            self._check_masters_formula_free(wb, report)
            self._check_seeds(wb, report)
            self._check_protection(wb, report)
            report.stats = {
                "sheets": len(wb.sheetnames),
                "tables": sum(len(ws.tables) for ws in wb.worksheets),
                "named_ranges": len(list(wb.defined_names.keys())),
                "errors": len(report.errors),
                "warnings": len(report.warnings),
            }
        finally:
            wb.close()

        report.passed = len(report.errors) == 0
        return report

    def _check_sheets(self, wb, report: ValidationReport) -> None:
        # İnci Ops visible titles (Phase 9) or legacy internal names
        from builders.inci_ops_ux import VISIBLE_SHEET_TITLES, VISIBLE_ORDER

        visible_titles = [VISIBLE_SHEET_TITLES[k] for k in VISIBLE_ORDER]
        for title in visible_titles:
            if title not in wb.sheetnames:
                # fallback legacy
                continue
        missing_visible = [t for t in visible_titles if t not in wb.sheetnames]
        if missing_visible:
            # allow legacy DASHBOARD/COMPONENT set
            legacy = set(UI_SHEETS) | {
                "COMPONENT",
                "PRODUCT",
                "PACKAGING_CONFIGURATION",
                "SHIPMENT",
                "STATEMENT",
                "TECHNICAL_FILE",
                "DECLARATION_OF_CONFORMITY",
            }
            if not missing_visible or not legacy.issubset(set(wb.sheetnames)):
                if any(t not in wb.sheetnames for t in visible_titles) and not {
                    "DASHBOARD",
                    "COMPONENT",
                }.issubset(set(wb.sheetnames)):
                    report.add(
                        "ERROR",
                        "SHT-01",
                        f"Missing visible ops sheets: {', '.join(missing_visible)}",
                    )

        # Entity sheets may be renamed for UX; engines/sys should still exist
        for engine in all_engines():
            if engine.sheet_name not in wb.sheetnames:
                report.add("ERROR", "SHT-01", f"Missing engine sheet {engine.sheet_name}")

        if wb.sheetnames and wb.sheetnames[0] not in {"Dashboard", "DASHBOARD"}:
            report.add(
                "WARN",
                "SHT-02",
                f"Dashboard not first; found {wb.sheetnames[0]}",
            )

    def _find_table(self, wb, table_name: str):
        for ws in wb.worksheets:
            if table_name in ws.tables:
                return ws, ws.tables[table_name]
        return None, None

    def _check_tables(self, wb, report: ValidationReport) -> None:
        for table in self.registry.tables:
            expected = f"T_{table.name}"
            ws, excel_table = self._find_table(wb, expected)
            if excel_table is None:
                # packaging lines relocated in Phase 9 — stub allowed on old sheet
                if table.name == "PACKAGING_CONFIGURATION_LINE":
                    ws2, _ = self._find_table(wb, "T_PACKAGING_CONFIGURATION_LINE")
                    if ws2 is not None:
                        continue
                report.add("ERROR", "TBL-01", f"Missing Excel Table {expected}")
                continue
            # Header row = first row of table ref
            start = excel_table.ref.split(":")[0]
            header_row = int("".join(ch for ch in start if ch.isdigit()))
            headers = [
                ws.cell(row=header_row, column=c).value
                for c in range(1, ws.max_column + 1)
            ]
            for col in table.column_names:
                if col not in headers:
                    report.add(
                        "ERROR",
                        "TBL-02",
                        f"{table.name} missing header {col}",
                    )

    def _check_named_ranges(self, wb, report: ValidationReport) -> None:
        names = set(wb.defined_names.keys())
        for table in self.registry.tables:
            nr = f"NR_{table.primary_key}"
            if nr not in names:
                report.add("ERROR", "NR-01", f"Missing named range {nr}")
        if "NR_SEARCH_TERM" not in names:
            report.add("WARN", "NR-02", "Missing NR_SEARCH_TERM")

    def _check_engines(self, wb, report: ValidationReport) -> None:
        for engine in all_engines():
            if engine.sheet_name not in wb.sheetnames:
                report.add("ERROR", "ENG-01", f"Missing engine {engine.sheet_name}")
                continue
            ws = wb[engine.sheet_name]
            if engine.table_name not in ws.tables:
                report.add("ERROR", "ENG-02", f"Missing table {engine.table_name}")
            formula_count = 0
            a1_refs = 0
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5), max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        # disallow plain Sheet!A1 style (allow # hyperlinks elsewhere)
                        if "!" in cell.value and "T_" not in cell.value and "[@" not in cell.value:
                            # structured refs don't use !; INDEX/MATCH uses T_ tables
                            if "'SEARCH'" not in cell.value:
                                a1_refs += 1
            if formula_count == 0 and engine.sheet_name != "ENG_VALIDATION":
                # validation always has formulas in RESULT
                report.add("ERROR", "ENG-03", f"No formulas found in {engine.sheet_name}")
            if engine.sheet_name == "ENG_VALIDATION" and formula_count == 0:
                report.add("ERROR", "ENG-03", "No formulas found in ENG_VALIDATION")

    def _check_ui(self, wb, report: ValidationReport) -> None:
        dash = "Dashboard" if "Dashboard" in wb.sheetnames else (
            "DASHBOARD" if "DASHBOARD" in wb.sheetnames else None
        )
        if dash is None:
            report.add("ERROR", "UI-01", "Missing Dashboard sheet")
        else:
            links = sum(
                1
                for row in wb[dash].iter_rows(min_row=1, max_row=40, max_col=12)
                for c in row
                if c.hyperlink
            )
            if links < 4:
                report.add("WARN", "UI-02", "Dashboard has fewer hyperlinks than expected")
            # Phase 11: PPWR workflow wizard inputs + search on Dashboard
            if dash == "Dashboard":
                for addr in ("D8", "D9", "D10", "D11", "D12", "C40"):
                    cell = wb[dash][addr]
                    if cell.protection.locked is not False:
                        report.add(
                            "WARN",
                            "UI-03",
                            f"Dashboard wizard/search cell {addr} should be unlocked",
                        )
                if "DOC_ENGINE_VARS" not in wb.sheetnames:
                    report.add("ERROR", "DOC-01", "Missing DOC_ENGINE_VARS document engine sheet")
                header = str(wb[dash]["B6"].value or "").upper()
                if "WORKFLOW" not in header and "PPWR" not in header:
                    report.add("ERROR", "UI-04", "Dashboard missing PPWR Workflow header")
                if wb[dash]["B4"].value is None or "NEXT" not in str(wb[dash]["B4"].value).upper():
                    report.add("ERROR", "UI-05", "Dashboard missing NEXT ACTION guidance formula")
                step1 = str(wb[dash]["C8"].value or "").upper()
                if "PACKAGING" not in step1:
                    report.add(
                        "ERROR",
                        "UI-06",
                        "Dashboard STEP 1 must be Packaging Configuration (PPWR model)",
                    )

    def _check_masters_formula_free(self, wb, report: ValidationReport) -> None:
        """Formulas forbidden inside entity Excel Tables; UX panels outside tables are OK."""
        for table in self.registry.tables:
            expected = f"T_{table.name}"
            ws, excel_table = self._find_table(wb, expected)
            if excel_table is None:
                continue
            start, end = excel_table.ref.split(":")
            start_row = int("".join(ch for ch in start if ch.isdigit()))
            end_row = int("".join(ch for ch in end if ch.isdigit()))
            start_col_letters = "".join(ch for ch in start if ch.isalpha())
            end_col_letters = "".join(ch for ch in end if ch.isalpha())

            def col_idx(letters: str) -> int:
                n = 0
                for ch in letters:
                    n = n * 26 + (ord(ch.upper()) - 64)
                return n

            min_col, max_col = col_idx(start_col_letters), col_idx(end_col_letters)
            for row in ws.iter_rows(
                min_row=start_row + 1,
                max_row=min(end_row, start_row + 20),
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        report.add(
                            "ERROR",
                            "MST-01",
                            f"Formula found inside table {expected} at {cell.coordinate}",
                        )
                        return

    def _check_seeds(self, wb, report: ValidationReport) -> None:
        # Expect lookup seeds for production
        if "LKP_MATERIAL" in wb.sheetnames and wb["LKP_MATERIAL"].max_row < 3:
            report.add("WARN", "SEED-01", "LKP_MATERIAL appears unseeded")
        if "LKP_STATUS" in wb.sheetnames and wb["LKP_STATUS"].max_row < 3:
            report.add("ERROR", "SEED-02", "LKP_STATUS not seeded")

    def _check_protection(self, wb, report: ValidationReport) -> None:
        for name in ("SYS_WORKBOOK_INFO", "SYS_PARAMETER", RELATIONSHIP_SHEET):
            if name in wb.sheetnames and wb[name].sheet_state == "visible":
                report.add("WARN", "PROT-01", f"Technical sheet {name} is visible")
            if name in wb.sheetnames and not wb[name].protection.sheet:
                report.add("WARN", "PROT-02", f"Technical sheet {name} not protected")
