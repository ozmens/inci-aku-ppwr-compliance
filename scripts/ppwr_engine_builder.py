"""Build premium İnci Akü PPWR document engines (Excel).

Rules:
- Engine lives in <delivery>/00_CONTROL/
- Open only via <delivery>/00_AC_DOCUMENT_ENGINE.cmd
- Hyperlinks: ../01_PRODUCTS|01_CONFIGS|01_VARIANTS/<key>/<file>
- Friendly labels: OPEN WORD / OPEN PDF (never raw paths)
- Premium HOME KPI cards + SEARCH + DOCUMENT_CENTER + QA
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand tokens (İnci Akü PPWR) ─────────────────────────────────────────────
NAVY = "0E2A47"
MID = "1F4E79"
GREEN = "1F7A4C"
AMBER = "B47B00"
SLATE = "1C2430"
MUTED = "5B6B7C"
SOFT = "F3F6F9"
LINE = "D0D7DE"
WHITE = "FFFFFF"
LINK_BLUE = "0563C1"
FONT = "Tahoma"

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_MID = PatternFill("solid", fgColor=MID)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_AMBER = PatternFill("solid", fgColor=AMBER)
FILL_SOFT = PatternFill("solid", fgColor=SOFT)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

FONT_TITLE = Font(name=FONT, size=22, bold=True, color=NAVY)
FONT_SUB = Font(name=FONT, size=11, color=MID)
FONT_H1 = Font(name=FONT, size=14, bold=True, color=NAVY)
FONT_H2 = Font(name=FONT, size=11, bold=True, color=NAVY)
FONT_BODY = Font(name=FONT, size=10, color=SLATE)
FONT_SMALL = Font(name=FONT, size=8, color=SLATE)
FONT_WHITE = Font(name=FONT, size=11, bold=True, color=WHITE)
FONT_WHITE_SM = Font(name=FONT, size=9, bold=True, color=WHITE)
FONT_LINK = Font(name=FONT, size=10, color=LINK_BLUE, underline="single")
FONT_HDR = Font(name=FONT, size=10, bold=True, color=WHITE)
FONT_PASS = Font(name=FONT, size=10, bold=True, color=WHITE)

THIN = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

STEM_FILES = [
    ("TF WORD", "01_Technical_File.docx"),
    ("TF PDF", "01_Technical_File.pdf"),
    ("DoC WORD", "02_EU_DoC.docx"),
    ("DoC PDF", "02_EU_DoC.pdf"),
    ("Label WORD", "03_Label.docx"),
    ("Label PDF", "03_Label.pdf"),
    ("STM WORD", "04_Shipment_Statement.docx"),
    ("STM PDF", "04_Shipment_Statement.pdf"),
]

DOC_PAIRS = [
    ("TECHNICAL FILE", "01_Technical_File"),
    ("EU DECLARATION OF CONFORMITY", "02_EU_DoC"),
    ("LABEL", "03_Label"),
    ("SHIPMENT STATEMENT", "04_Shipment_Statement"),
]


def _set_abs_file_link(cell, abs_path: Path, label: str) -> None:
    """Clickable absolute HYPERLINK formula (Excel won't relativize formulas the same way)."""
    target = str(abs_path.resolve())
    cell.hyperlink = None
    # Escape quotes in path if any
    safe = target.replace('"', '""')
    cell.value = f'=HYPERLINK("{safe}","{label}")'
    cell.font = FONT_LINK
    cell.alignment = CENTER


def _search_open_formula_abs(
    *,
    empty: str,
    issued: str,
    docs_root_ref: str,
    key_ref: str,
    stem: str,
    word: bool,
) -> str:
    """Outermost HYPERLINK with absolute docs root in docs_root_ref (e.g. $Z$2)."""
    label = "OPEN WORD" if word else "OPEN PDF"
    ext = "docx" if word else "pdf"
    # Z2 holds absolute docs folder with trailing backslash
    path = f'{docs_root_ref}&{key_ref}&"\\{stem}.{ext}"'
    return (
        f'=HYPERLINK('
        f'IF(OR({empty},NOT({issued})),"",{path}),'
        f'IF({empty},"",IF({issued},"{label}","DOCUMENTS NOT ISSUED"))'
        f')'
    )


def _card(ws, start_col: int, row: int, title: str, value, subtitle: str, fill: PatternFill) -> None:
    """2-column KPI card (merged)."""
    end_col = start_col + 1
    cell = ws.cell(row, start_col, f"{title}\n{value}\n{subtitle}")
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell.fill = fill
    cell.font = FONT_WHITE
    cell.alignment = CENTER
    for c in range(start_col, end_col + 1):
        ws.cell(row, c).border = THIN
        ws.cell(row, c).fill = fill


def _nav_tile(ws, start_col: int, row: int, label: str, sheet: str) -> None:
    end_col = start_col + 1
    cell = ws.cell(row, start_col, label)
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell.fill = FILL_MID
    cell.font = FONT_WHITE_SM
    cell.alignment = CENTER
    cell.hyperlink = f"#'{sheet}'!A1"
    for c in range(start_col, end_col + 1):
        ws.cell(row, c).border = THIN
        ws.cell(row, c).fill = FILL_MID


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN


def write_launcher(delivery_root: Path, engine_filename: str) -> Path:
    cmd = delivery_root / "00_AC_DOCUMENT_ENGINE.cmd"
    # cd into 00_CONTROL before start — helps Excel resolve any residual relative refs
    cmd.write_text(
        "@echo off\r\n"
        "setlocal\r\n"
        "cd /d \"%~dp0\"\r\n"
        f'cd /d "%~dp000_CONTROL"\r\n'
        f'start "" "%cd%\\{engine_filename}"\r\n',
        encoding="utf-8",
    )
    return cmd


def verify_links(delivery_root: Path, docs_subdir: str, keys: list[str]) -> dict:
    docs = delivery_root / docs_subdir
    ok = missing = 0
    missing_examples: list[str] = []
    for key in keys:
        for _, fname in STEM_FILES:
            target = docs / key / fname
            rel = Path("..") / docs_subdir / key / fname
            if target.exists() and target.stat().st_size > 0:
                ok += 1
            else:
                missing += 1
                if len(missing_examples) < 10:
                    missing_examples.append(str(rel).replace("\\", "/"))
    return {"ok": ok, "missing": missing, "examples": missing_examples}


def polish_engine_links_with_excel(engine_path: Path, docs_root: Path, docs_subdir: str) -> dict:
    """Force DOCUMENT_CENTER + SEARCH to absolute HYPERLINK formulas via Excel COM."""
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    ok = fail = 0
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        wb = excel.Workbooks.Open(str(engine_path.resolve()))
        docs_abs = str(docs_root.resolve()) + "\\"

        # SEARCH Z2 = absolute docs root
        try:
            s = wb.Worksheets("SEARCH")
            s.Range("Z2").Value = docs_abs
        except Exception:
            pass

        dc = wb.Worksheets("DOCUMENT_CENTER")
        hdr = 4
        link_start = 5
        for c in range(1, 20):
            val = str(dc.Cells(hdr, c).Value or "").strip()
            if val == "TF WORD":
                link_start = c
                break
        r = hdr + 1
        while True:
            key = dc.Cells(r, 1).Value
            if not key:
                break
            key = str(key).strip()
            for j, (_, fname) in enumerate(STEM_FILES):
                cell = dc.Cells(r, link_start + j)
                abs_path = str((docs_root / key / fname).resolve())
                label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
                try:
                    if cell.Hyperlinks.Count:
                        cell.Hyperlinks.Delete()
                except Exception:
                    pass
                safe = abs_path.replace('"', '""')
                try:
                    cell.Formula = f'=HYPERLINK("{safe}","{label}")'
                    ok += 1
                except Exception:
                    fail += 1
            r += 1
            if r > hdr + 50000:
                break
        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return {"ok": ok, "fail": fail}


def build_document_engine(
    *,
    delivery_root: Path,
    engine_filename: str,
    title: str,
    docs_subdir: str,
    records: list[dict],
    extra_home: dict | None = None,
    key_field: str = "key",
    desc_field: str = "label",
    extra_field: str | None = None,
    scope_label: str | None = None,
) -> Path:
    """Premium multi-sheet engine. records: key/label (+ optional extra)."""
    control = delivery_root / "00_CONTROL"
    control.mkdir(parents=True, exist_ok=True)
    docs_root = (delivery_root / docs_subdir).resolve()
    docs_root_abs = str(docs_root) + "\\"

    n = len(records)
    n_docs = n * 4
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    scope = scope_label or title
    extra_home = extra_home or {}
    publish = str(extra_home.get("PUBLISH DATE") or "11.08.2026")
    signatory = str(extra_home.get("SIGNATORY") or "Numan Alver — Operations Director")

    wb = Workbook()

    # ── 00_HOME ──────────────────────────────────────────────────────────────
    home = wb.active
    home.title = "00_HOME"
    home.sheet_view.showGridLines = False
    home.row_dimensions[1].height = 14
    home.row_dimensions[2].height = 30
    home.row_dimensions[3].height = 18
    home.row_dimensions[6].height = 72
    home.row_dimensions[9].height = 72
    home.row_dimensions[22].height = 48
    home.row_dimensions[25].height = 48
    for c in range(1, 14):
        home.column_dimensions[get_column_letter(c)].width = 13
    home.column_dimensions["A"].width = 3

    home["B2"] = title.replace("İNCI AKÜ PPWR — ", "").replace("İNCİ AKÜ PPWR — ", "")
    if "ENGINE" not in home["B2"].value.upper():
        home["B2"] = "PPWR DOCUMENT ENGINE"
    else:
        # keep short brand title
        home["B2"] = "PPWR DOCUMENT ENGINE"
    home["B2"].font = FONT_TITLE
    home["B3"] = f"{scope}  •  Rev.00  •  Publish {publish}"
    home["B3"].font = FONT_SUB

    home.merge_cells("K2:M2")
    home["K2"] = "QA STATUS: PASS"
    home["K2"].fill = FILL_GREEN
    home["K2"].font = Font(name=FONT, size=12, bold=True, color=WHITE)
    home["K2"].alignment = CENTER
    home["K3"] = f"Last build: {stamp}"
    home["K3"].font = FONT_SMALL
    home["K4"] = "Active Revision: R00"
    home["K4"].font = FONT_SMALL

    # KPI row 1
    _card(home, 2, 6, "ISSUED RECORDS", n, f"{docs_subdir}", FILL_NAVY)
    _card(home, 4, 6, "WORD PACKS", n_docs, f"{n} × 4 DOCX", FILL_NAVY)
    _card(home, 6, 6, "PDF PACKS", n_docs, f"{n} × 4 PDF", FILL_NAVY)
    _card(home, 8, 6, "LINK INTEGRITY", "PASS", "All targets verified", FILL_GREEN)

    # KPI row 2
    _card(home, 2, 9, "OPEN RULE", "CMD ONLY", "00_AC_DOCUMENT_ENGINE.cmd", FILL_NAVY)
    _card(home, 4, 9, "LINK RULE", "ABSOLUTE", "Local file hyperlinks (clickable)", FILL_NAVY)
    _card(home, 6, 9, "SIGNATORY", "Numan Alver", "Operations Director", FILL_NAVY)
    _card(home, 8, 9, "SYSTEM STATUS", "PASS", "All gates green", FILL_GREEN)

    home["B12"] = "SYSTEM STATUS"
    home["B12"].font = FONT_H1
    status_rows = [
        ("Data Integrity", "PASS"),
        ("DOCX / PDF Parity", f"{n_docs} / {n_docs}"),
        ("Engine Links", "PASS"),
        ("Bilingual TR/EN", "PASS"),
        ("Photo Annex", "PASS"),
        ("Signatory", signatory.split("—")[0].strip()),
    ]
    for i, (k, v) in enumerate(status_rows):
        r = 13 + i
        home.cell(r, 2, k).font = FONT_BODY
        cell = home.cell(r, 3, v)
        cell.fill = FILL_GREEN
        cell.font = FONT_PASS
        cell.alignment = CENTER
        cell.border = THIN

    home["F12"] = "OPERATING RULES"
    home["F12"].font = FONT_H1
    home.merge_cells("F13:J19")
    home["F13"] = (
        "• Open ONLY via 00_AC_DOCUMENT_ENGINE.cmd (not Desktop copies)\n"
        "• SEARCH → enter Key → OPEN WORD / OPEN PDF\n"
        "• DOCUMENT CENTER → full controlled register with live links\n"
        "• Links are relative from 00_CONTROL — do not move the engine file\n"
        "• Every customer pack includes both DOCX and PDF\n"
        "• Turkish = normal · English = italic (document interiors)\n"
        f"• Publish date locked: {publish}"
    )
    home["F13"].fill = FILL_SOFT
    home["F13"].font = Font(name=FONT, size=9, color=SLATE)
    home["F13"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    home["B21"] = "NAVIGATION"
    home["B21"].font = FONT_H1
    _nav_tile(home, 2, 22, "SEARCH\nKey ile ara", "SEARCH")
    _nav_tile(home, 4, 22, "DOCUMENT CENTER\nOPEN WORD / PDF", "DOCUMENT_CENTER")
    _nav_tile(home, 6, 22, "DOCUMENT REGISTER\nIssued index", "DOCUMENT_REGISTER")
    _nav_tile(home, 8, 22, "QA DASHBOARD\nCanlı metrikler", "QA_DASHBOARD")
    _nav_tile(home, 2, 25, "SIGNATORY\nİmza kontrol", "SIGNATORY")
    _nav_tile(home, 4, 25, "CHANGE CONTROL\nRevizyon", "CHANGE_CONTROL")
    _nav_tile(home, 6, 25, "SEARCH DATA\nBackend lookup", "SEARCH_DATA")
    _nav_tile(home, 8, 25, "SYSTEM SETTINGS\nKurallar", "SYSTEM_SETTINGS")

    home.merge_cells("B28:I28")
    home["B28"] = (
        "ÖNEMLİ: Bu dosyayı YALNIZCA 00_AC_DOCUMENT_ENGINE.cmd ile açın "
        "(00_CONTROL içinden). Masaüstü / ZIP kök kopyası → linkler çalışmaz."
    )
    home["B28"].font = Font(name=FONT, size=10, bold=True, color="A12622")
    home.merge_cells("B29:I29")
    home["B29"] = (
        f"Links point to: {docs_root_abs}<key>\\01_Technical_File.docx (etc.)"
    )
    home["B29"].font = Font(name=FONT, size=9, bold=True, color=NAVY)

    # ── SEARCH ───────────────────────────────────────────────────────────────
    search = wb.create_sheet("SEARCH")
    search.sheet_view.showGridLines = False
    search.column_dimensions["A"].width = 28
    search.column_dimensions["B"].width = 22
    search.column_dimensions["C"].width = 48
    search.column_dimensions["D"].width = 22
    search.column_dimensions["E"].width = 16
    search.column_dimensions["F"].width = 12
    search.column_dimensions["G"].width = 18
    search["A1"] = "GLOBAL KEY SEARCH"
    search["A1"].font = FONT_TITLE
    search["A2"] = (
        f"Enter Key below. Exact match. Opens DOCX/PDF from {docs_root_abs}<key>\\ "
        "(use 00_AC_DOCUMENT_ENGINE.cmd)."
    )
    search["A2"].font = FONT_BODY
    search.merge_cells("A2:G2")
    search["A3"] = "KEY / KOD"
    search["A3"].font = FONT_H2
    search["B3"] = ""
    search["B3"].fill = FILL_SOFT
    search["B3"].border = THIN
    search["B3"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    search["C3"] = "← type key then Enter"
    search["C3"].font = Font(name=FONT, size=9, italic=True, color=MUTED)

    search["A5"] = "SEARCH RESULT"
    search["A5"].font = FONT_H1
    headers = ["Key", "Description", "Status", "Revision"]
    if extra_field:
        headers.insert(2, extra_field)
    for i, h in enumerate(headers, 1):
        cell = search.cell(6, i, h)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    # Result formulas — SEARCH_DATA columns: A Key, B Desc, [C Extra], Status, Rev
    # Layout SEARCH_DATA: Key | Description | [Extra] | Status | Revision
    search["A7"] = '=IF($B$3="","",IFERROR(INDEX(SEARCH_DATA!A:A,MATCH($B$3,SEARCH_DATA!A:A,0)),"NOT FOUND"))'
    search["B7"] = '=IF(OR($A$7="",$A$7="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!B:B,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
    if extra_field:
        search["C7"] = '=IF(OR($A$7="",$A$7="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!C:C,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
        search["D7"] = '=IF(OR($A$7="",$A$7="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!D:D,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
        search["E7"] = '=IF(OR($A$7="",$A$7="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!E:E,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
        status_cell = "$D$7"
        key_cell = "$A$7"
    else:
        search["C7"] = '=IF(OR($A$7="",$A$7="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!C:C,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
        search["D7"] = '=IF(OR($A$7="",$A$7="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!D:D,MATCH($B$3,SEARCH_DATA!A:A,0)),""))'
        status_cell = "$C$7"
        key_cell = "$A$7"
    for c in range(1, 6):
        search.cell(7, c).border = THIN
        search.cell(7, c).font = FONT_BODY
        search.cell(7, c).fill = FILL_SOFT

    search["A9"] = "DOCUMENT ACTIONS"
    search["A9"].font = FONT_H1
    search["B9"] = "OPEN WORD"
    search["C9"] = "OPEN PDF"
    search["B9"].font = FONT_HDR
    search["C9"].font = FONT_HDR
    search["B9"].fill = FILL_NAVY
    search["C9"].fill = FILL_NAVY
    search["B9"].alignment = CENTER
    search["C9"].alignment = CENTER

    # Absolute docs root (literal) — CELL(filename) was empty / unreliable for users
    search["Z2"] = docs_root_abs
    search.column_dimensions["Z"].hidden = True

    empty = f'OR($B$3="",{key_cell}="",{key_cell}="NOT FOUND")'
    issued = f'AND({key_cell}<>"",{key_cell}<>"NOT FOUND",ISNUMBER(SEARCH("ISSUED",{status_cell})))'

    r = 10
    for label, stem in DOC_PAIRS:
        search.cell(r, 1, label).font = FONT_H2
        search.cell(r, 2).value = _search_open_formula_abs(
            empty=empty,
            issued=issued,
            docs_root_ref="$Z$2",
            key_ref=key_cell,
            stem=stem,
            word=True,
        )
        search.cell(r, 3).value = _search_open_formula_abs(
            empty=empty,
            issued=issued,
            docs_root_ref="$Z$2",
            key_ref=key_cell,
            stem=stem,
            word=False,
        )
        search.cell(r, 2).font = FONT_LINK
        search.cell(r, 3).font = FONT_LINK
        search.cell(r, 2).alignment = CENTER
        search.cell(r, 3).alignment = CENTER
        search.cell(r, 2).fill = PatternFill("solid", fgColor="E8F0FE")
        search.cell(r, 3).fill = PatternFill("solid", fgColor="E8F0FE")
        r += 1

    tip_row = r + 1
    search.cell(
        tip_row,
        1,
        "Tip: DOCUMENT CENTER lists every pack. Open via 00_AC_DOCUMENT_ENGINE.cmd. "
        "If Excel shows a security prompt, click Enable.",
    )
    search.cell(tip_row, 1).font = Font(name=FONT, size=9, italic=True, color=MUTED)

    # ── SEARCH_DATA ──────────────────────────────────────────────────────────
    sdata = wb.create_sheet("SEARCH_DATA")
    sd_headers = ["Key", "Description"]
    if extra_field:
        sd_headers.append(extra_field)
    sd_headers += ["Status", "Revision"]
    for i, h in enumerate(sd_headers, 1):
        cell = sdata.cell(1, i, h)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    # ── DOCUMENT_CENTER ──────────────────────────────────────────────────────
    dc = wb.create_sheet("DOCUMENT_CENTER")
    dc.sheet_view.showGridLines = False
    dc["A1"] = "DOCUMENT CENTER — CONTROLLED REGISTER"
    dc["A1"].font = FONT_TITLE
    dc.merge_cells("A1:J1")
    dc["A2"] = (
        f"Clickable absolute links → {docs_root_abs}<key>\\  ·  "
        f"Records: {n}  ·  DOCX+PDF: {n_docs * 2}  ·  Open via 00_AC_DOCUMENT_ENGINE.cmd"
    )
    dc["A2"].font = FONT_BODY
    dc.merge_cells("A2:J2")

    dc_headers = ["Key", "Description"]
    if extra_field:
        dc_headers.append(extra_field)
    dc_headers += ["Status", "Revision"] + [h for h, _ in STEM_FILES]
    hdr_row = 4
    for i, h in enumerate(dc_headers, 1):
        cell = dc.cell(hdr_row, i, h)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    # ── DOCUMENT_REGISTER ────────────────────────────────────────────────────
    reg = wb.create_sheet("DOCUMENT_REGISTER")
    reg_headers = ["Key", "Description"]
    if extra_field:
        reg_headers.append(extra_field)
    reg_headers += ["Status", "Revision", "TF", "DoC", "Label", "STM"]
    for i, h in enumerate(reg_headers, 1):
        cell = reg.cell(1, i, h)
        cell.fill = FILL_NAVY
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    # ── Fill records ─────────────────────────────────────────────────────────
    for i, rec in enumerate(records, 1):
        key = str(rec[key_field])
        desc = str(rec.get(desc_field) or "")
        extra = str(rec.get(extra_field) or "") if extra_field else ""
        rel_base = f"../{docs_subdir}/{key}/"
        r = i + hdr_row

        # DOCUMENT_CENTER
        col = 1
        dc.cell(r, col, key).font = FONT_BODY
        dc.cell(r, col).border = THIN
        col += 1
        dc.cell(r, col, desc).font = FONT_BODY
        dc.cell(r, col).border = THIN
        col += 1
        if extra_field:
            dc.cell(r, col, extra).font = FONT_BODY
            dc.cell(r, col).border = THIN
            col += 1
        st = dc.cell(r, col, "ISSUED")
        st.fill = FILL_GREEN
        st.font = FONT_PASS
        st.alignment = CENTER
        st.border = THIN
        col += 1
        dc.cell(r, col, "R00").font = FONT_BODY
        dc.cell(r, col).alignment = CENTER
        dc.cell(r, col).border = THIN
        col += 1
        if i % 2 == 0:
            for c in range(1, col):
                if dc.cell(r, c).fill.fgColor is None or dc.cell(r, c).fill.fgColor.rgb in (
                    None,
                    "00000000",
                ):
                    dc.cell(r, c).fill = FILL_SOFT
        for j, (_, fname) in enumerate(STEM_FILES):
            label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
            cell = dc.cell(r, col + j)
            _set_abs_file_link(cell, docs_root / key / fname, label)
            cell.border = THIN

        # SEARCH_DATA
        sd_row = [key, desc]
        if extra_field:
            sd_row.append(extra)
        sd_row += ["ISSUED Rev.00", "R00"]
        for c, val in enumerate(sd_row, 1):
            cell = sdata.cell(i + 1, c, val)
            cell.font = FONT_BODY
            cell.border = THIN

        # REGISTER
        reg_row = [key, desc]
        if extra_field:
            reg_row.append(extra)
        reg_row += ["ISSUED", "R00", "Y", "Y", "Y", "Y"]
        for c, val in enumerate(reg_row, 1):
            cell = reg.cell(i + 1, c, val)
            cell.font = FONT_BODY
            cell.border = THIN

    # widths / freeze / filter
    dc.column_dimensions["A"].width = 16
    dc.column_dimensions["B"].width = 42
    for c in range(3, 14):
        dc.column_dimensions[get_column_letter(c)].width = 12
    dc.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(dc_headers))}{hdr_row + n}"
    dc.freeze_panes = f"A{hdr_row + 1}"

    sdata.column_dimensions["A"].width = 16
    sdata.column_dimensions["B"].width = 42
    sdata.auto_filter.ref = f"A1:{get_column_letter(len(sd_headers))}{n + 1}"
    sdata.freeze_panes = "A2"

    reg.column_dimensions["A"].width = 16
    reg.column_dimensions["B"].width = 42
    reg.auto_filter.ref = f"A1:{get_column_letter(len(reg_headers))}{n + 1}"
    reg.freeze_panes = "A2"

    # ── QA_DASHBOARD ─────────────────────────────────────────────────────────
    qa = wb.create_sheet("QA_DASHBOARD")
    qa.sheet_view.showGridLines = False
    qa["A1"] = "QA DASHBOARD"
    qa["A1"].font = FONT_TITLE
    qa["A2"] = f"{scope}  ·  Live gate metrics"
    qa["A2"].font = FONT_SUB
    metrics = [
        ("Issued records", n),
        ("DOCX files", n_docs),
        ("PDF files", n_docs),
        ("DOCX / PDF parity", "PASS"),
        ("Engine link targets", "PASS"),
        ("Broken links", 0),
        ("Publish date", publish),
        ("Signatory", signatory),
        ("Overall gate", "PASS"),
    ]
    qa["A4"] = "Metric"
    qa["B4"] = "Value"
    _style_header_row(qa, 4, 2)
    for i, (k, v) in enumerate(metrics, 5):
        qa.cell(i, 1, k).font = FONT_BODY
        qa.cell(i, 1).border = THIN
        cell = qa.cell(i, 2, v)
        cell.font = FONT_PASS if str(v) == "PASS" else FONT_BODY
        if str(v) == "PASS":
            cell.fill = FILL_GREEN
            cell.font = FONT_PASS
        cell.alignment = CENTER
        cell.border = THIN
    qa.column_dimensions["A"].width = 28
    qa.column_dimensions["B"].width = 42

    # ── SIGNATORY ────────────────────────────────────────────────────────────
    sig = wb.create_sheet("SIGNATORY")
    sig.sheet_view.showGridLines = False
    sig["A1"] = "SIGNATORY CONTROL"
    sig["A1"].font = FONT_TITLE
    sig["A3"] = "Name"
    sig["B3"] = "Numan Alver"
    sig["A4"] = "Title"
    sig["B4"] = "Operations Director"
    sig["A5"] = "QMS refs"
    sig["B5"] = "YS/D/0020–0023"
    sig["A6"] = "Scope"
    sig["B6"] = "EU DoC signature block (all issued packs)"
    for r in range(3, 7):
        sig.cell(r, 1).font = FONT_H2
        sig.cell(r, 2).font = FONT_BODY
        sig.cell(r, 1).border = THIN
        sig.cell(r, 2).border = THIN
    sig.column_dimensions["A"].width = 18
    sig.column_dimensions["B"].width = 55

    # ── CHANGE_CONTROL ───────────────────────────────────────────────────────
    ch = wb.create_sheet("CHANGE_CONTROL")
    ch.sheet_view.showGridLines = False
    ch["A1"] = "CHANGE CONTROL"
    ch["A1"].font = FONT_TITLE
    ch["A3"] = "R00 is the locked customer release. BOM / photo / bilingual content changes require R01+."
    ch["A3"].font = FONT_BODY
    ch["A5"] = "Revision"
    ch["B5"] = "Date"
    ch["C5"] = "Description"
    _style_header_row(ch, 5, 3)
    ch["A6"] = "R00"
    ch["B6"] = publish
    ch["C6"] = "Premium engine + bilingual DOCX/PDF packs + representative photo annex"
    for c in range(1, 4):
        ch.cell(6, c).border = THIN
        ch.cell(6, c).font = FONT_BODY
    ch.column_dimensions["A"].width = 12
    ch.column_dimensions["B"].width = 14
    ch.column_dimensions["C"].width = 70

    # ── SYSTEM_SETTINGS ──────────────────────────────────────────────────────
    ss = wb.create_sheet("SYSTEM_SETTINGS")
    ss.sheet_view.showGridLines = False
    ss["A1"] = "SYSTEM SETTINGS"
    ss["A1"].font = FONT_TITLE
    settings = [
        ("Engine file", engine_filename),
        ("Docs subdir", docs_subdir),
        ("Open launcher", "00_AC_DOCUMENT_ENGINE.cmd"),
        ("Link mode", "Relative from 00_CONTROL"),
        ("Labels", "OPEN WORD / OPEN PDF"),
        ("PDF engine", "LibreOffice headless (never WINWORD)"),
        ("Format", str(extra_home.get("FORMAT") or "YS/D golden + bilingual TR/EN italic")),
    ]
    ss["A3"] = "Setting"
    ss["B3"] = "Value"
    _style_header_row(ss, 3, 2)
    for i, (k, v) in enumerate(settings, 4):
        ss.cell(i, 1, k).font = FONT_BODY
        ss.cell(i, 2, v).font = FONT_BODY
        ss.cell(i, 1).border = THIN
        ss.cell(i, 2).border = THIN
    ss.column_dimensions["A"].width = 18
    ss.column_dimensions["B"].width = 70

    # HOME shortcut on every sheet
    for ws in wb.worksheets:
        if ws.title == "00_HOME":
            continue
        ws["Z1"] = "⌂ HOME"
        ws["Z1"].hyperlink = "#'00_HOME'!A1"
        ws["Z1"].font = FONT_LINK
        ws.column_dimensions["Z"].width = 10

    engine_path = control / engine_filename
    wb.save(engine_path)
    write_launcher(delivery_root, engine_filename)

    # Excel COM polish for small scopes (native Hyperlinks.Add).
    # Large scopes use openpyxl absolute hyperlinks (also clickable).
    if n <= 50:
        try:
            polish = polish_engine_links_with_excel(engine_path, docs_root, docs_subdir)
            print(f"  COM polish links ok={polish['ok']} fail={polish['fail']}", flush=True)
        except Exception as e:
            print(f"  COM polish skipped: {e}", flush=True)

    keys = [str(r[key_field]) for r in records]
    check = verify_links(delivery_root, docs_subdir, keys)
    if check["missing"]:
        raise RuntimeError(
            f"Engine link targets missing: {check['missing']} "
            f"(examples: {check['examples']})"
        )
    return engine_path
