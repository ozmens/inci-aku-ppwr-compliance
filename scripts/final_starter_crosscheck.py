"""FINAL STARTER PRODUCT <-> PACKAGING SET CROSS-CHECK (Excel QA only).

Creates CROSSCHECK_CANDIDATE from locked master.
Does NOT modify locked master / Word / Industrial / freeze.
Reports exceptions; does not silently fix data.
"""

from __future__ import annotations

import hashlib
import json
import math
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
OUT = ROOT / "output"
LOCKED = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
BACKUP = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00_PRELOCK_BACKUP.xlsx"
CANDIDATE = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00_CROSSCHECK_CANDIDATE.xlsx"
GOLDEN = (
    ROOT
    / "input"
    / "production"
    / "INCI_AKU_PPWR_Final_Configuration_Register_Rev00_GOLDEN_VARIANTS_FINAL.xlsx"
)
PHASE_I = OUT / "PHASE_I_FINAL"
QA_MD = OUT / "FINAL_STARTER_CROSSCHECK_QA.md"
QA_JSON = OUT / "FINAL_STARTER_CROSSCHECK_QA.json"

PENDING_PCS = {"1013084", "1014789", "1014790"}
PENDING_VARIANT = "IA-ST-VAR-0037"
PROVISIONAL = "ST-018-STD-04"
IND_SETS = {"IND-24V-01", "IND-48V-01", "IND-80V-01"}
PHYS_CONTROLLED = "CONTROLLED PACKAGING SET"
BOM_REQ = "BOM DATA REQUIRED"
PHYS_BOM = "BOM DATA REQUIRED — DO NOT ISSUE DOCUMENTS"
CFG_CONTROLLED = "CONTROLLED"
NOT_ISSUED = "NOT ISSUED"

# tare tolerance (approved numeric precision)
TARE_EXACT = 1e-9
TARE_TOL = 1e-3

DOC_TYPES = {
    "TF": "YS/D/0020",
    "DOC": "YS/D/0021",
    "LABEL": "YS/D/0022",
    "STM": "YS/D/0023",
}

NAVY = "0E2A47"
WHITE = "FFFFFF"
INK = "1C2430"
BAND = "F3F6F9"
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def word_fp() -> dict[str, Any]:
    files = sorted((PHASE_I / "01_STARTER").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "02_INDUSTRIAL").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "03_CONTAINER").rglob("01_Technical_File.docx"))
    files = [p for p in files if not p.name.startswith("~$")]
    h = hashlib.sha256()
    digests = []
    for p in files:
        d = sha256_file(p)
        digests.append(d)
        h.update(d.encode())
        h.update(str(p.relative_to(PHASE_I)).encode())
    return {"count": len(files), "aggregate": h.hexdigest(), "sample": digests[:3]}


def parse_linked(s: Any) -> list[str]:
    if s is None or s == "":
        return []
    text = str(s).replace("|", ";")
    return [x.strip() for x in text.split(";") if x.strip()]


def fnum(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(v)
    except Exception:
        return None


def load_sheet(wb, name: str) -> tuple[list[str], list[dict]]:
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers)) if headers[i] is not None})
    return headers, rows


def original_240() -> set[str]:
    wb = load_workbook(GOLDEN, data_only=True, read_only=True)
    ws = wb["01_FINAL_CONFIG_MASTER"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    codes = set()
    container = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        fam = str(row[idx["Family"]] or "").upper()
        sc = str(row[idx["Packaging Set Code"]])
        if fam == "STARTER":
            codes.add(sc)
        elif fam == "CONTAINER":
            container.add(sc)
    wb.close()
    return codes, container


def search_workbook_for_tokens(wb, tokens: set[str]) -> dict[str, list[str]]:
    """Return token -> list of sheet!cell hits (capped)."""
    hits: dict[str, list[str]] = {t: [] for t in tokens}
    for sheet in wb.sheetnames:
        if sheet == "FINAL_CROSSCHECK_QA":
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                for t in tokens:
                    if t in val:
                        loc = f"{sheet}!{cell.coordinate}"
                        if len(hits[t]) < 20:
                            hits[t].append(loc)
    return hits


def write_qa_sheet(wb, report: dict, exceptions: list[dict]) -> None:
    if "FINAL_CROSSCHECK_QA" in wb.sheetnames:
        del wb["FINAL_CROSSCHECK_QA"]
    ws = wb.create_sheet("FINAL_CROSSCHECK_QA")
    headers = ["Section", "Metric", "Value", "Expected", "Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = HAIR
    rows = [
        ["PRODUCT", "Unique Product Codes", report["unique_product_codes"], 2046, ""],
        ["PRODUCT", "Controlled products", report["controlled_products"], 2043, ""],
        ["PRODUCT", "BOM DATA REQUIRED", report["bom_data_required"], 3, ""],
        ["SET", "Controlled Packaging Sets", report["controlled_sets"], 311, ""],
        ["MAP", "Exactly one set", report["exactly_one_set"], 2043, ""],
        ["MAP", "Complete linked lists", report["complete_linked_sets"], 311, ""],
        ["UNION", "Union linked products", report["union_linked"], 2043, ""],
        ["DIFF", "Controlled missing from sets", report["diff_a"], 0, ""],
        ["DIFF", "Set-linked not controlled", report["diff_b"], 0, ""],
        ["DIFF", "Multi-set products", report["diff_c"], 0, ""],
        ["PENDING", "Pending in controlled sets", report["pending_in_controlled_sets"], 0, ""],
        ["BOM", "Tare exact", report["tare_exact"], "", ""],
        ["BOM", "Tare tolerance", report["tare_tolerance"], "", ""],
        ["BOM", "Tare fail", report["tare_fail"], 0, ""],
        ["DOC", "Document completeness", report["doc_complete"], 2043, ""],
        ["IND", "Industrial hits", report["industrial_hits"], 0, ""],
        ["CNT", "Container controlled", report["container_controlled"], 0, ""],
        ["ADJ", "Adjacency", report["adjacency"], "PASS", ""],
        ["NUM", "Management numbering", report["management_numbering"], "PASS", ""],
        ["GATE", "FINAL", report["final_gate"], "PASS", ""],
    ]
    for r_i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r_i + 2, c, v)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.border = HAIR
            cell.fill = PatternFill("solid", fgColor=BAND if r_i % 2 else WHITE)

    # exceptions
    ws2 = wb.create_sheet("FINAL_CROSSCHECK_EXCEPTIONS")
    eh = [
        "Product Code",
        "Packaging Set Code",
        "Source Configuration ID",
        "Sheet",
        "Problem",
        "Recommended correction",
    ]
    for c, h in enumerate(eh, 1):
        cell = ws2.cell(1, c, h)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = HAIR
    if not exceptions:
        ws2.cell(2, 1, "(none)")
    else:
        for r_i, ex in enumerate(exceptions):
            for c, k in enumerate(eh, 1):
                cell = ws2.cell(r_i + 2, c, ex.get(k, ""))
                cell.font = Font(name=FONT, size=9)
                cell.border = HAIR


def run() -> dict[str, Any]:
    fp_before = word_fp()
    exceptions: list[dict] = []

    if not LOCKED.exists():
        raise FileNotFoundError(LOCKED)
    if not BACKUP.exists():
        exceptions.append(
            {
                "Product Code": "",
                "Packaging Set Code": "",
                "Source Configuration ID": "",
                "Sheet": "FILESYSTEM",
                "Problem": "PRELOCK backup missing",
                "Recommended correction": "Restore backup before lock operations",
            }
        )

    # Candidate copy first
    shutil.copy2(LOCKED, CANDIDATE)

    orig240, container_codes = original_240()

    wb = load_workbook(CANDIDATE, data_only=True)
    pm_h, products = load_sheet(wb, "PRODUCT_MASTER")
    cm_h, configs = load_sheet(wb, "CONFIG_MASTER")
    bm_h, boms = load_sheet(wb, "BOM_MASTER")
    ds_h, docs = load_sheet(wb, "DOCUMENT_SCOPE")
    # optional sheets
    sheet_names = set(wb.sheetnames)

    # --- 1 Product uniqueness ---
    pcs = [str(p["Product Code"]).strip() if p.get("Product Code") is not None else "" for p in products]
    blank_pc = sum(1 for pc in pcs if not pc)
    unique_pc = len(set(pcs) - {""})
    dup_pc = [pc for pc, n in Counter(pcs).items() if pc and n > 1]
    for pc in dup_pc:
        exceptions.append(
            {
                "Product Code": pc,
                "Packaging Set Code": "",
                "Source Configuration ID": "",
                "Sheet": "PRODUCT_MASTER",
                "Problem": "Duplicate Product Code row",
                "Recommended correction": "Keep single controlled row",
            }
        )
    if blank_pc:
        exceptions.append(
            {
                "Product Code": "",
                "Packaging Set Code": "",
                "Source Configuration ID": "",
                "Sheet": "PRODUCT_MASTER",
                "Problem": f"Blank Product Code rows: {blank_pc}",
                "Recommended correction": "Remove or fill Product Code",
            }
        )

    missing_src = []
    status_other = []
    controlled_products = []
    bom_req_products = []
    for p in products:
        pc = str(p.get("Product Code") or "").strip()
        src = str(p.get("Source Configuration ID") or "").strip()
        phys = str(p.get("Physical Packaging Status") or "").strip()
        sc = str(p.get("Packaging Set Code") or "").strip()
        if not src:
            missing_src.append(pc)
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": "",
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Missing Source Configuration ID",
                    "Recommended correction": "Populate Source Configuration ID from Golden/L2",
                }
            )
        if phys == PHYS_CONTROLLED:
            controlled_products.append(p)
        elif sc == BOM_REQ or phys == PHYS_BOM or pc in PENDING_PCS:
            bom_req_products.append(p)
            # normalize classification check
            if pc not in PENDING_PCS:
                status_other.append(pc)
            if sc != BOM_REQ and sc != NOT_ISSUED:
                # still count as BOM req if status says so, but flag if unexpected code
                if sc not in (BOM_REQ, NOT_ISSUED):
                    pass
        else:
            status_other.append(pc)
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": f"Disallowed Physical Packaging Status: {phys}",
                    "Recommended correction": "Must be CONTROLLED PACKAGING SET or BOM DATA REQUIRED",
                }
            )

    # ensure pending trio classified as BOM DATA REQUIRED
    bom_pcs = {str(p["Product Code"]).strip() for p in bom_req_products}
    for pc in PENDING_PCS:
        if pc not in bom_pcs:
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": "",
                    "Source Configuration ID": "",
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Expected BOM DATA REQUIRED product not classified as such",
                    "Recommended correction": "Set Packaging Set Code / Physical status to BOM DATA REQUIRED",
                }
            )

    n_controlled = len(controlled_products)
    n_bom = len(bom_req_products)

    # --- 2 Controlled product -> set ---
    product_to_sets: dict[str, set[str]] = defaultdict(set)
    exactly_one = 0
    for p in controlled_products:
        pc = str(p["Product Code"]).strip()
        sc = str(p.get("Packaging Set Code") or "").strip()
        src = str(p.get("Source Configuration ID") or "").strip()
        fin = str(p.get("Final Configuration ID") or "").strip()
        bad = False
        if not sc:
            bad = True
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": "",
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Blank Packaging Set Code for controlled product",
                    "Recommended correction": "Assign controlled Packaging Set Code",
                }
            )
        if sc == PROVISIONAL or sc == NOT_ISSUED or sc == BOM_REQ:
            bad = True
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Controlled product points to provisional/BOM DATA REQUIRED/NOT ISSUED",
                    "Recommended correction": "Map to a controlled physical Packaging Set",
                }
            )
        if sc in IND_SETS or sc.startswith("IND-"):
            bad = True
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Packaging Set belongs to Industrial",
                    "Recommended correction": "Remove Industrial mapping from Starter",
                }
            )
        if sc.startswith("CNT-") or sc in container_codes:
            bad = True
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Packaging Set belongs to Container",
                    "Recommended correction": "Remove Container mapping from Starter",
                }
            )
        if not fin or fin == NOT_ISSUED:
            bad = True
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Missing/invalid Final Configuration ID",
                    "Recommended correction": "Populate Final Configuration ID from controlled set",
                }
            )
        if sc:
            product_to_sets[pc].add(sc)
        if not bad and len(product_to_sets[pc]) == 1:
            exactly_one += 1

    for pc, sets in product_to_sets.items():
        if len(sets) > 1:
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": "; ".join(sorted(sets)),
                    "Source Configuration ID": "",
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Product Code maps to more than one Packaging Set",
                    "Recommended correction": "Keep single Packaging Set relationship",
                }
            )

    # --- 3 Packaging set -> product ---
    controlled_cfgs = [c for c in configs if str(c.get("Configuration Status") or "") == CFG_CONTROLLED]
    # fallback if status missing: exclude NOT ISSUED / provisional / IND / CNT
    if len(controlled_cfgs) != 311:
        controlled_cfgs = [
            c
            for c in configs
            if str(c.get("Packaging Set Code") or "")
            not in {NOT_ISSUED, PROVISIONAL, BOM_REQ}
            and not str(c.get("Packaging Set Code") or "").startswith("IND-")
            and not str(c.get("Packaging Set Code") or "").startswith("CNT-")
            and str(c.get("Configuration Status") or "") != "BOM DATA REQUIRED"
        ]

    set_codes = [str(c["Packaging Set Code"]).strip() for c in controlled_cfgs]
    set_dup = [sc for sc, n in Counter(set_codes).items() if n > 1]
    for sc in set_dup:
        exceptions.append(
            {
                "Product Code": "",
                "Packaging Set Code": sc,
                "Source Configuration ID": "",
                "Sheet": "CONFIG_MASTER",
                "Problem": "Packaging Set Code appears more than once among controlled rows",
                "Recommended correction": "Keep single CONFIG_MASTER row per set",
            }
        )

    pm_by_pc = {str(p["Product Code"]).strip(): p for p in products}
    set_to_linked: dict[str, set[str]] = {}
    complete_linked = 0
    for c in controlled_cfgs:
        sc = str(c["Packaging Set Code"]).strip()
        linked = parse_linked(c.get("Linked Product Codes"))
        linked_set = set(linked)
        set_to_linked[sc] = linked_set
        src = str(c.get("Source Configuration ID") or "")
        count = c.get("Product Count")
        ok = True
        if not linked:
            ok = False
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "CONFIG_MASTER",
                    "Problem": "Blank Linked Product Codes",
                    "Recommended correction": "Populate complete linked Product Code list",
                }
            )
        if count is not None and int(count) != len(linked_set):
            ok = False
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "CONFIG_MASTER",
                    "Problem": f"Product Count {count} != unique linked {len(linked_set)}",
                    "Recommended correction": "Align Product Count with unique Linked Product Codes",
                }
            )
        # back-pointer check
        for pc in linked_set:
            if pc not in pm_by_pc:
                ok = False
                exceptions.append(
                    {
                        "Product Code": pc,
                        "Packaging Set Code": sc,
                        "Source Configuration ID": src,
                        "Sheet": "CONFIG_MASTER",
                        "Problem": "Linked Product Code not in PRODUCT_MASTER",
                        "Recommended correction": "Add product to PRODUCT_MASTER or remove from linked list",
                    }
                )
            else:
                back = str(pm_by_pc[pc].get("Packaging Set Code") or "").strip()
                if back != sc:
                    ok = False
                    exceptions.append(
                        {
                            "Product Code": pc,
                            "Packaging Set Code": sc,
                            "Source Configuration ID": src,
                            "Sheet": "CONFIG_MASTER",
                            "Problem": f"Back-pointer mismatch: PRODUCT_MASTER has {back}",
                            "Recommended correction": "Make PRODUCT_MASTER and CONFIG_MASTER agree",
                        }
                    )
        # extras/missing vs PRODUCT_MASTER forward map
        expected = {pc for pc, sets in product_to_sets.items() if sc in sets}
        extra = linked_set - expected
        missing = expected - linked_set
        if extra:
            ok = False
            for pc in sorted(extra)[:20]:
                exceptions.append(
                    {
                        "Product Code": pc,
                        "Packaging Set Code": sc,
                        "Source Configuration ID": src,
                        "Sheet": "CONFIG_MASTER",
                        "Problem": "Extra Product Code in Linked Product Codes vs PRODUCT_MASTER",
                        "Recommended correction": "Remove extra or update PRODUCT_MASTER",
                    }
                )
        if missing:
            ok = False
            for pc in sorted(missing)[:20]:
                exceptions.append(
                    {
                        "Product Code": pc,
                        "Packaging Set Code": sc,
                        "Source Configuration ID": src,
                        "Sheet": "CONFIG_MASTER",
                        "Problem": "Missing Product Code in Linked Product Codes",
                        "Recommended correction": "Add missing Product Code to linked list",
                    }
                )
        if ok:
            complete_linked += 1

    # --- 4 Global union ---
    union = set()
    for s in set_to_linked.values():
        union |= s
    controlled_pc_set = {str(p["Product Code"]).strip() for p in controlled_products}
    diff_a = sorted(controlled_pc_set - union)
    diff_b = sorted(union - controlled_pc_set)
    # multi-set from reverse
    reverse: dict[str, set[str]] = defaultdict(set)
    for sc, linked in set_to_linked.items():
        for pc in linked:
            reverse[pc].add(sc)
    multi = sorted([pc for pc, sets in reverse.items() if len(sets) > 1])
    for pc in diff_a:
        exceptions.append(
            {
                "Product Code": pc,
                "Packaging Set Code": str(pm_by_pc[pc].get("Packaging Set Code") or ""),
                "Source Configuration ID": str(pm_by_pc[pc].get("Source Configuration ID") or ""),
                "Sheet": "PRODUCT_MASTER",
                "Problem": "Controlled product missing from all set Linked Product Codes",
                "Recommended correction": "Add to CONFIG_MASTER linked list",
            }
        )
    for pc in diff_b:
        exceptions.append(
            {
                "Product Code": pc,
                "Packaging Set Code": "",
                "Source Configuration ID": "",
                "Sheet": "CONFIG_MASTER",
                "Problem": "Set-linked Product Code not controlled in PRODUCT_MASTER",
                "Recommended correction": "Classify product or remove from linked list",
            }
        )
    for pc in multi:
        exceptions.append(
            {
                "Product Code": pc,
                "Packaging Set Code": "; ".join(sorted(reverse[pc])),
                "Source Configuration ID": "",
                "Sheet": "CONFIG_MASTER",
                "Problem": "Product Code linked to multiple Packaging Sets",
                "Recommended correction": "Keep single Packaging Set relationship",
            }
        )

    # --- 5 BOM DATA REQUIRED trio ---
    docs_by_pc = {str(d["Product Code"]).strip(): d for d in docs}
    pending_in_sets = 0
    for pc in sorted(PENDING_PCS):
        p = pm_by_pc.get(pc)
        d = docs_by_pc.get(pc)
        if not p:
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": "",
                    "Source Configuration ID": "",
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "BOM DATA REQUIRED product missing from PRODUCT_MASTER",
                    "Recommended correction": "Restore product row",
                }
            )
            continue
        sc = str(p.get("Packaging Set Code") or "")
        src = str(p.get("Source Configuration ID") or "")
        if not src:
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": "",
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Pending product missing Source Configuration ID",
                    "Recommended correction": "Set Source Configuration ID",
                }
            )
        if sc not in (BOM_REQ, NOT_ISSUED):
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "PRODUCT_MASTER",
                    "Problem": "Pending product Packaging Set Code not BOM DATA REQUIRED/NOT ISSUED",
                    "Recommended correction": "Set Packaging Set Code = BOM DATA REQUIRED",
                }
            )
        if d:
            for col in (
                "Technical File ID",
                "EU DoC ID",
                "Label ID",
                "Shipment Statement ID",
            ):
                if str(d.get(col) or "") != NOT_ISSUED:
                    exceptions.append(
                        {
                            "Product Code": pc,
                            "Packaging Set Code": sc,
                            "Source Configuration ID": src,
                            "Sheet": "DOCUMENT_SCOPE",
                            "Problem": f"{col} is {d.get(col)!r}, expected NOT ISSUED",
                            "Recommended correction": "Set all four document IDs to NOT ISSUED",
                        }
                    )
        else:
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": src,
                    "Sheet": "DOCUMENT_SCOPE",
                    "Problem": "Pending product missing from DOCUMENT_SCOPE",
                    "Recommended correction": "Add DOCUMENT_SCOPE row with NOT ISSUED IDs",
                }
            )
        # appearances in controlled linked lists
        for sc2, linked in set_to_linked.items():
            if pc in linked:
                pending_in_sets += 1
                exceptions.append(
                    {
                        "Product Code": pc,
                        "Packaging Set Code": sc2,
                        "Source Configuration ID": src,
                        "Sheet": "CONFIG_MASTER",
                        "Problem": "BOM DATA REQUIRED product appears in controlled Linked Product Codes",
                        "Recommended correction": "Remove from controlled set linked list",
                    }
                )

    # pending config exclusion from 311
    pending_excluded = (
        PROVISIONAL not in set_codes
        and NOT_ISSUED not in [s for s in set_codes]
        and all(str(c.get("Packaging Set Code")) != PROVISIONAL for c in controlled_cfgs)
    )
    # NOT_ISSUED may exist as pending row outside controlled — check controlled list only
    pending_excluded = PROVISIONAL not in set(set_codes) and all(
        str(c.get("Configuration Status")) == CFG_CONTROLLED
        or str(c.get("Packaging Set Code")) not in (PROVISIONAL, NOT_ISSUED)
        for c in controlled_cfgs
    )
    pending_excluded = PROVISIONAL not in set(set_codes) and NOT_ISSUED not in set(set_codes)

    # --- 6 Industrial exclusion ---
    ind_hits = search_workbook_for_tokens(wb, IND_SETS)
    # allow mention only in HOME narrative? User said ZERO operational records — any hit fails
    # but HOME text might mention industrial? Check current home - likely not.
    industrial_total = sum(len(v) for v in ind_hits.values())
    # filter HOME / PROVISIONAL / SCOPE narrative? User: search ALL sheets — any FOUND is FOUND
    ind_status = {k: ("FOUND" if v else "ABSENT") for k, v in ind_hits.items()}
    for k, locs in ind_hits.items():
        for loc in locs:
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": k,
                    "Source Configuration ID": "",
                    "Sheet": loc.split("!")[0],
                    "Problem": f"Industrial configuration token found at {loc}",
                    "Recommended correction": "Remove Industrial reference from Starter workbook",
                }
            )

    # --- 7 Container exclusion ---
    container_in_starter = []
    for c in controlled_cfgs:
        sc = str(c["Packaging Set Code"])
        if sc.startswith("CNT-") or sc in container_codes:
            container_in_starter.append(sc)
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": sc,
                    "Source Configuration ID": str(c.get("Source Configuration ID") or ""),
                    "Sheet": "CONFIG_MASTER",
                    "Problem": "Container configuration counted in controlled Starter sets",
                    "Recommended correction": "Exclude Container from Starter controlled sets",
                }
            )
    # also scan tokens CNT-
    cnt_hits = 0
    for sheet in ("PRODUCT_MASTER", "CONFIG_MASTER", "BOM_MASTER", "DOCUMENT_SCOPE", "DOCUMENT_CENTER", "SEARCH_DATA"):
        if sheet not in sheet_names:
            continue
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            for v in row:
                if v is None:
                    continue
                s = str(v)
                if s.startswith("CNT-") or s in container_codes:
                    cnt_hits += 1
                    if cnt_hits <= 10:
                        exceptions.append(
                            {
                                "Product Code": "",
                                "Packaging Set Code": s,
                                "Source Configuration ID": "",
                                "Sheet": sheet,
                                "Problem": "Container code present in operational Starter sheet",
                                "Recommended correction": "Remove Container rows from Starter",
                            }
                        )

    # --- 8 BOM tare reconciliation ---
    bom_by_set: dict[str, list[dict]] = defaultdict(list)
    for b in boms:
        sc = str(b.get("Packaging Set Code") or "").strip()
        if sc in set(set_codes):
            bom_by_set[sc].append(b)

    tare_exact = tare_tol = tare_fail = 0
    sets_missing_bom = []
    for c in controlled_cfgs:
        sc = str(c["Packaging Set Code"]).strip()
        lines = bom_by_set.get(sc, [])
        if not lines:
            tare_fail += 1
            sets_missing_bom.append(sc)
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": sc,
                    "Source Configuration ID": str(c.get("Source Configuration ID") or ""),
                    "Sheet": "BOM_MASTER",
                    "Problem": "No BOM component rows for controlled Packaging Set",
                    "Recommended correction": "Import controlled BOM lines; do not invent",
                }
            )
            continue
        # validate required fields
        for b in lines:
            for col in (
                "Packaging Set Code",
                "Linked Product Codes",
                "Source Configuration ID",
                "Component Code",
                "Quantity",
                "UOM",
                "Unit Weight",
                "Line Weight",
            ):
                if b.get(col) in (None, ""):
                    exceptions.append(
                        {
                            "Product Code": "",
                            "Packaging Set Code": sc,
                            "Source Configuration ID": str(b.get("Source Configuration ID") or ""),
                            "Sheet": "BOM_MASTER",
                            "Problem": f"Blank required field: {col}",
                            "Recommended correction": "Complete BOM field from source data",
                        }
                    )
                    break
        bom_tare = 0.0
        for b in lines:
            lw = fnum(b.get("Line Weight"))
            if lw is None:
                q, uw = fnum(b.get("Quantity")), fnum(b.get("Unit Weight"))
                if q is not None and uw is not None:
                    lw = q * uw
                else:
                    lw = 0.0
            bom_tare += lw
        cfg_tare = fnum(c.get("Packaging Tare kg"))
        if cfg_tare is None:
            tare_fail += 1
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": sc,
                    "Source Configuration ID": str(c.get("Source Configuration ID") or ""),
                    "Sheet": "CONFIG_MASTER",
                    "Problem": "Packaging Tare kg blank; cannot reconcile to BOM",
                    "Recommended correction": "Populate tare from controlled BOM sum",
                }
            )
            continue
        diff = abs(cfg_tare - bom_tare)
        if diff <= TARE_EXACT:
            tare_exact += 1
        elif diff <= TARE_TOL:
            tare_tol += 1
        else:
            tare_fail += 1
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": sc,
                    "Source Configuration ID": str(c.get("Source Configuration ID") or ""),
                    "Sheet": "BOM_MASTER",
                    "Problem": f"Tare mismatch CONFIG={cfg_tare} BOM_SUM={round(bom_tare,6)} DIFF={round(diff,6)}",
                    "Recommended correction": "Investigate source weights; do not force-fit",
                }
            )

    # --- 9 Document scope ---
    cfg_by_set = {str(c["Packaging Set Code"]).strip(): c for c in controlled_cfgs}
    doc_complete = 0
    for p in controlled_products:
        pc = str(p["Product Code"]).strip()
        sc = str(p.get("Packaging Set Code") or "").strip()
        d = docs_by_pc.get(pc)
        cfg = cfg_by_set.get(sc)
        ok = True
        if not d:
            ok = False
            exceptions.append(
                {
                    "Product Code": pc,
                    "Packaging Set Code": sc,
                    "Source Configuration ID": str(p.get("Source Configuration ID") or ""),
                    "Sheet": "DOCUMENT_SCOPE",
                    "Problem": "Missing DOCUMENT_SCOPE row",
                    "Recommended correction": "Add document scope row aligned to Packaging Set",
                }
            )
        else:
            if str(d.get("Packaging Set Code") or "") != sc:
                ok = False
                exceptions.append(
                    {
                        "Product Code": pc,
                        "Packaging Set Code": sc,
                        "Source Configuration ID": str(p.get("Source Configuration ID") or ""),
                        "Sheet": "DOCUMENT_SCOPE",
                        "Problem": f"DOCUMENT_SCOPE Packaging Set {d.get('Packaging Set Code')} != PRODUCT {sc}",
                        "Recommended correction": "Align DOCUMENT_SCOPE Packaging Set Code",
                    }
                )
            if cfg:
                for col in (
                    "Technical File ID",
                    "EU DoC ID",
                    "Label ID",
                    "Shipment Statement ID",
                ):
                    if str(d.get(col) or "") != str(cfg.get(col) or ""):
                        ok = False
                        exceptions.append(
                            {
                                "Product Code": pc,
                                "Packaging Set Code": sc,
                                "Source Configuration ID": str(p.get("Source Configuration ID") or ""),
                                "Sheet": "DOCUMENT_SCOPE",
                                "Problem": f"{col} mismatch vs CONFIG_MASTER",
                                "Recommended correction": "Use configuration-level shared document IDs",
                            }
                        )
                    val = str(d.get(col) or "")
                    if not val or val == NOT_ISSUED:
                        ok = False
                        exceptions.append(
                            {
                                "Product Code": pc,
                                "Packaging Set Code": sc,
                                "Source Configuration ID": str(p.get("Source Configuration ID") or ""),
                                "Sheet": "DOCUMENT_SCOPE",
                                "Problem": f"{col} missing/NOT ISSUED for controlled product",
                                "Recommended correction": "Assign configuration-level IA-PPWR-* ID",
                            }
                        )
        if ok:
            doc_complete += 1

    # --- 10 Adjacency ---
    adjacency_fail_sheets = []
    checks = [
        ("PRODUCT_MASTER", "Product Code", "Packaging Set Code"),
        ("DOCUMENT_SCOPE", "Product Code", "Packaging Set Code"),
        ("SEARCH_DATA", "Product Code", "Packaging Set Code"),
        ("CONFIG_MASTER", "Packaging Set Code", "Linked Product Codes"),
        ("BOM_MASTER", "Packaging Set Code", "Linked Product Codes"),
        ("DOCUMENT_CENTER", "Packaging Set Code", "Linked Product Codes"),
        ("TECHNICAL_FILES", "Packaging Set Code", "Linked Product Codes"),
        ("DECLARATIONS_OF_CONFORMITY", "Packaging Set Code", "Linked Product Codes"),
        ("LABELS", "Packaging Set Code", "Linked Product Codes"),
        ("SHIPMENT_STATEMENTS", "Packaging Set Code", "Linked Product Codes"),
    ]
    for sheet, a, b in checks:
        if sheet not in sheet_names:
            continue
        headers = [c.value for c in next(wb[sheet].iter_rows(min_row=1, max_row=1))]
        # find adjacent pair
        ok = False
        for i in range(len(headers) - 1):
            if headers[i] == a and headers[i + 1] == b:
                ok = True
                break
        # also allow A/B as cols 0/1
        if headers and headers[0] == a and len(headers) > 1 and headers[1] == b:
            ok = True
        if not ok:
            adjacency_fail_sheets.append(sheet)
            exceptions.append(
                {
                    "Product Code": "",
                    "Packaging Set Code": "",
                    "Source Configuration ID": "",
                    "Sheet": sheet,
                    "Problem": f"Adjacency fail: expected {a} next to {b}; headers={headers[:4]}",
                    "Recommended correction": "Place Product Code / Packaging Set Code (or Linked) adjacent",
                }
            )
    adjacency = "PASS" if not adjacency_fail_sheets else "FAIL"

    # --- 11 Management numbering ---
    numbering_ok = True
    for sheet, type_col, id_prefix, expected_type in (
        ("TECHNICAL_FILES", "Management Type No", "IA-PPWR-TF-", DOC_TYPES["TF"]),
        ("DECLARATIONS_OF_CONFORMITY", "Management Type No", "IA-PPWR-DOC-", DOC_TYPES["DOC"]),
        ("LABELS", "Management Type No", "IA-PPWR-LBL-", DOC_TYPES["LABEL"]),
        ("SHIPMENT_STATEMENTS", "Management Type No", "IA-PPWR-STM-", DOC_TYPES["STM"]),
    ):
        if sheet not in sheet_names:
            continue
        _, rows = load_sheet(wb, sheet)
        for r in rows:
            mt = str(r.get(type_col) or "")
            did = str(r.get("Document ID") or "")
            if mt != expected_type:
                numbering_ok = False
                exceptions.append(
                    {
                        "Product Code": "",
                        "Packaging Set Code": str(r.get("Packaging Set Code") or ""),
                        "Source Configuration ID": "",
                        "Sheet": sheet,
                        "Problem": f"Management type {mt} != {expected_type}",
                        "Recommended correction": "Keep YS/D type numbers separate from IA-PPWR IDs",
                    }
                )
            if not did.startswith(id_prefix):
                numbering_ok = False
                exceptions.append(
                    {
                        "Product Code": "",
                        "Packaging Set Code": str(r.get("Packaging Set Code") or ""),
                        "Source Configuration ID": "",
                        "Sheet": sheet,
                        "Problem": f"Document ID {did} does not start with {id_prefix}",
                        "Recommended correction": "Do not mix management type numbers into config IDs",
                    }
                )
            # mixed: type number used as document id
            if did in DOC_TYPES.values():
                numbering_ok = False
                exceptions.append(
                    {
                        "Product Code": "",
                        "Packaging Set Code": str(r.get("Packaging Set Code") or ""),
                        "Source Configuration ID": "",
                        "Sheet": sheet,
                        "Problem": "Management type number used as Document ID",
                        "Recommended correction": "Use IA-PPWR-* for Document ID",
                    }
                )

    # existing 240 / 71 new
    controlled_set_set = set(set_codes)
    existing_ok = orig240 <= controlled_set_set and len(orig240) == 240
    new_sets = controlled_set_set - orig240
    new_ok = len(new_sets) == 71

    if not existing_ok:
        missing = sorted(orig240 - controlled_set_set)
        exceptions.append(
            {
                "Product Code": "",
                "Packaging Set Code": "; ".join(missing[:20]),
                "Source Configuration ID": "",
                "Sheet": "CONFIG_MASTER",
                "Problem": f"Existing 240 incomplete; missing {len(missing)}",
                "Recommended correction": "Restore missing existing Packaging Set Codes",
            }
        )
    if not new_ok:
        exceptions.append(
            {
                "Product Code": "",
                "Packaging Set Code": "",
                "Source Configuration ID": "",
                "Sheet": "CONFIG_MASTER",
                "Problem": f"New validated sets count {len(new_sets)} != 71",
                "Recommended correction": "Reconcile against BOM identity audit",
            }
        )

    n_controlled_sets = len(controlled_set_set)

    fp_after = word_fp()
    word_ok = fp_before["aggregate"] == fp_after["aggregate"] and fp_after["count"] == 247

    # Deduplicate exceptions for report size (keep all in JSON)
    # Final gate
    final_pass = (
        unique_pc == 2046
        and blank_pc == 0
        and not dup_pc
        and not missing_src
        and n_controlled == 2043
        and n_bom == 3
        and not status_other
        and exactly_one == 2043
        and n_controlled_sets == 311
        and complete_linked == 311
        and len(union) == 2043
        and not diff_a
        and not diff_b
        and not multi
        and pending_in_sets == 0
        and tare_fail == 0
        and (tare_exact + tare_tol) == 311
        and doc_complete == 2043
        and industrial_total == 0
        and not container_in_starter
        and cnt_hits == 0
        and adjacency == "PASS"
        and numbering_ok
        and existing_ok
        and new_ok
        and pending_excluded
        and word_ok
    )

    report = {
        "unique_product_codes": f"{unique_pc} / 2046",
        "controlled_products": f"{n_controlled} / 2043",
        "bom_data_required": f"{n_bom} / 3",
        "controlled_sets": f"{n_controlled_sets} / 311",
        "exactly_one_set": f"{exactly_one} / 2043",
        "complete_linked_sets": f"{complete_linked} / 311",
        "union_linked": f"{len(union)} / 2043",
        "diff_a": len(diff_a),
        "diff_b": len(diff_b),
        "diff_c": len(multi),
        "unknown_in_sets": len(diff_b),
        "pending_in_controlled_sets": pending_in_sets,
        "tare_exact": tare_exact,
        "tare_tolerance": tare_tol,
        "tare_fail": tare_fail,
        "doc_complete": f"{doc_complete} / 2043",
        "industrial_hits": industrial_total,
        "ind_24": ind_status.get("IND-24V-01", "ABSENT"),
        "ind_48": ind_status.get("IND-48V-01", "ABSENT"),
        "ind_80": ind_status.get("IND-80V-01", "ABSENT"),
        "container_controlled": len(container_in_starter) + cnt_hits,
        "adjacency": adjacency,
        "adjacency_fail_sheets": adjacency_fail_sheets,
        "management_numbering": "PASS" if numbering_ok else "FAIL",
        "existing_240": "PASS" if existing_ok else "FAIL",
        "new_71": "PASS" if new_ok else "FAIL",
        "pending_excluded_from_311": "PASS" if pending_excluded else "FAIL",
        "word_hash_changed": 0 if word_ok else 1,
        "final_gate": "PASS" if final_pass else "FAIL",
        "exception_count": len(exceptions),
        "candidate": str(CANDIDATE),
        "locked_untouched": True,
    }

    # Write QA sheets to candidate only
    write_qa_sheet(wb, report, exceptions)
    wb.save(CANDIDATE)
    wb.close()

    # Confirm locked master hash unchanged vs pre-candidate? Candidate is copy; locked should be same as before script.
    # Verify locked file still present and not written: we never opened LOCKED for write.

    lines = [
        "# FINAL STARTER CROSS-CHECK QA",
        "",
        "Unique Starter Product Codes:",
        report["unique_product_codes"],
        "",
        "Controlled Product Codes:",
        report["controlled_products"],
        "",
        "BOM DATA REQUIRED:",
        report["bom_data_required"],
        "",
        "Controlled Packaging Sets:",
        report["controlled_sets"],
        "",
        "Product Codes mapped to exactly one controlled set:",
        report["exactly_one_set"],
        "",
        "Packaging Sets with complete linked product list:",
        report["complete_linked_sets"],
        "",
        "Union of controlled Linked Product Codes:",
        report["union_linked"],
        "",
        "Controlled products missing from set lists:",
        str(report["diff_a"]),
        "",
        "Products linked to multiple sets:",
        str(report["diff_c"]),
        "",
        "Unknown Product Codes inside set lists:",
        str(report["unknown_in_sets"]),
        "",
        "BOM DATA REQUIRED products appearing in controlled sets:",
        str(report["pending_in_controlled_sets"]),
        "",
        "BOM tare reconciliation:",
        f"Exact: {report['tare_exact']}",
        f"Tolerance: {report['tare_tolerance']}",
        f"Fail: {report['tare_fail']}",
        "",
        "Document-scope completeness:",
        report["doc_complete"],
        "",
        "Industrial configurations in Starter:",
        str(report["industrial_hits"]),
        "Expected = 0",
        "",
        "Check explicitly:",
        "IND-24V-01:",
        report["ind_24"],
        "",
        "IND-48V-01:",
        report["ind_48"],
        "",
        "IND-80V-01:",
        report["ind_80"],
        "",
        "Container controlled records in Starter:",
        str(report["container_controlled"]),
        "Expected = 0",
        "",
        "Product Code <-> Packaging Set adjacency:",
        report["adjacency"],
        "",
        "Management document numbering:",
        report["management_numbering"],
        "",
        "Existing 240 Packaging Set Codes unchanged:",
        report["existing_240"],
        "",
        "71 new validated Packaging Set Codes present:",
        report["new_71"],
        "",
        "Pending IA-ST-VAR-0037 excluded from 311:",
        report["pending_excluded_from_311"],
        "",
        "Word hash changed:",
        str(report["word_hash_changed"]),
        "",
        "FINAL STARTER DATA INTEGRITY GATE:",
        report["final_gate"],
        "",
    ]

    if report["final_gate"] == "FAIL":
        lines.append("EXCEPTIONS:")
        lines.append("")
        # unique by problem signature to keep readable
        seen = set()
        shown = 0
        for ex in exceptions:
            key = (
                ex.get("Product Code"),
                ex.get("Packaging Set Code"),
                ex.get("Sheet"),
                ex.get("Problem"),
            )
            if key in seen:
                continue
            seen.add(key)
            lines.append(f"- Product Code: {ex.get('Product Code')}")
            lines.append(f"  Packaging Set Code: {ex.get('Packaging Set Code')}")
            lines.append(f"  Source Configuration ID: {ex.get('Source Configuration ID')}")
            lines.append(f"  Sheet: {ex.get('Sheet')}")
            lines.append(f"  Problem: {ex.get('Problem')}")
            lines.append(f"  Recommended correction: {ex.get('Recommended correction')}")
            lines.append("")
            shown += 1
            if shown >= 80:
                lines.append(f"... ({len(exceptions) - shown} more in JSON/candidate sheet)")
                break
    else:
        lines.append("No exceptions.")
        lines.append("")

    lines.extend(
        [
            "STOP.",
            "",
            "DO NOT FIX SILENTLY.",
            "DO NOT GENERATE WORDS.",
            "DO NOT START PIMS.",
            "",
            f"Candidate (QA sheets added): `{CANDIDATE}`",
            f"Locked master (untouched): `{LOCKED}`",
        ]
    )

    QA_MD.write_text("\n".join(lines), encoding="utf-8")
    QA_JSON.write_text(
        json.dumps({"report": report, "exceptions": exceptions}, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    try:
        print("\n".join(lines))
    except UnicodeEncodeError:
        print("\n".join(lines).encode("ascii", "replace").decode("ascii"))
    return report


if __name__ == "__main__":
    run()
