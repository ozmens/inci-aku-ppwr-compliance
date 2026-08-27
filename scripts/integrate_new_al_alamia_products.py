"""
Integrate 20 Al Alamia products into STARTER_MASTER.

- Exact BOM match → reuse Packaging Set (none expected for this batch)
- Unique new BOM → create new Packaging Set (cluster siblings)
- Do NOT alter existing 287 set BOMs / mappings
"""

from __future__ import annotations

import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "output" / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
NEW_XLSX = ROOT / "input" / "new_products_al_alamia_2026-08.xlsx"
BACKUP = (
    ROOT
    / "output"
    / f"INCI_AKU_PPWR_STARTER_MASTER_Rev00_BACKUP_BEFORE_AL_ALAMIA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)


def norm_uom(u: str) -> str:
    u = (u or "").upper().strip()
    return "ADT" if u in {"ADT", "PCS", "PC", "ADET"} else u


def bom_sig(lines: list[tuple[str, float, str]]) -> tuple:
    return tuple(sorted((c, round(q, 6), norm_uom(u)) for c, q, u in lines if q))


def parse_new_products() -> list[dict]:
    wb = load_workbook(NEW_XLSX, data_only=True)
    ws = wb["NEW_PRODUCTS"]
    out = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        pc = str(int(row[0])) if isinstance(row[0], float) else str(row[0]).strip()
        comps = []
        i = 5
        while i + 3 < len(row):
            code, desc, qty, uom = row[i : i + 4]
            i += 4
            if not code or float(qty or 0) == 0:
                continue
            code_s = str(int(code)) if isinstance(code, float) else str(code).strip()
            comps.append(
                {
                    "code": code_s,
                    "desc": str(desc or "").strip(),
                    "qty": float(qty),
                    "uom": norm_uom(str(uom or "")),
                }
            )
        out.append(
            {
                "product_code": pc,
                "description": str(row[1] or "").strip(),
                "battery_type": str(row[2] or "").strip(),
                "market": str(row[3] or "").strip(),
                "nominal_qty": float(row[4] or 0),
                "components": comps,
            }
        )
    wb.close()
    return out


def build_weight_lookup(wb) -> dict[str, tuple[object, float]]:
    """component_code -> (unit_weight, typical_line_weight_per_qty_unit)."""
    bm = wb["BOM_MASTER"]
    headers = [c.value for c in next(bm.iter_rows(min_row=1, max_row=1))]
    hi = {h: i for i, h in enumerate(headers)}
    # prefer first non-null unit weight seen
    lookup: dict[str, tuple[object, float]] = {}
    for row in bm.iter_rows(min_row=2, values_only=True):
        code = str(row[hi["Component Code"]] or "").strip()
        if not code or code in lookup:
            continue
        uw = row[hi["Unit Weight"]]
        lw = row[hi["Line Weight"]]
        qty = float(row[hi["Quantity"]] or 0) or 1.0
        lookup[code] = (uw, float(lw or 0) / qty if lw is not None else 0.0)
    return lookup


def next_set_code(variants: dict[str, int], nominal_qty: float) -> str:
    fam = f"ST-{int(nominal_qty):03d}-STD"
    n = variants.get(fam, 0) + 1
    variants[fam] = n
    return f"{fam}-{n:02d}"


def compute_tare(comps: list[dict], weights: dict[str, tuple[object, float]]) -> float:
    total = 0.0
    for c in comps:
        code = c["code"]
        qty = c["qty"]
        uom = c["uom"]
        uw, per = weights.get(code, (None, 0.0))
        if uom == "KG" or (isinstance(uw, str) and "MASS" in str(uw).upper()):
            total += qty  # qty already kg
        elif isinstance(uw, (int, float)):
            total += float(uw) * qty
        else:
            total += per * qty
    return round(total, 4)


def variant_basis(comps: list[dict], battery_type: str) -> tuple[str, str]:
    # Pick distinctive components for basis text
    by_code = {c["code"]: c for c in comps}
    bits_tr = [battery_type]
    bits_en = [battery_type]
    if "4000037" in by_code:
        bits_tr.append("Palet (4000037): 1 adet")
        bits_en.append("Pallet (4000037): 1 pc")
    sep = by_code.get("4000130")
    if sep:
        bits_tr.append(f"Separatör (4000130): {sep['qty']:g} adet")
        bits_en.append(f"Separator (4000130): {sep['qty']:g} pcs")
    # shrink-like
    for code in ("4000555", "4000556", "4000557", "4000558", "4000608"):
        if code in by_code:
            q = by_code[code]["qty"]
            bits_tr.append(f"Shrink film ({code}): {q:g} kg")
            bits_en.append(f"Shrink film ({code}): {q:g} kg")
            break
    return " | ".join(bits_tr), " | ".join(bits_en)


def main() -> None:
    products = parse_new_products()
    assert len(products) == 20, len(products)

    shutil.copy2(MASTER, BACKUP)
    print("BACKUP", BACKUP)

    wb = load_workbook(MASTER)
    weights = build_weight_lookup(wb)

    cm = wb["CONFIG_MASTER"]
    pm = wb["PRODUCT_MASTER"]
    bm = wb["BOM_MASTER"]

    ch = [c.value for c in cm[1]]
    ci = {h: i + 1 for i, h in enumerate(ch)}
    ph = [c.value for c in pm[1]]
    pi = {h: i + 1 for i, h in enumerate(ph)}
    bh = [c.value for c in bm[1]]
    bi = {h: i + 1 for i, h in enumerate(bh)}

    # existing controlled BOM signatures
    controlled_sets = set()
    variants: dict[str, int] = {}
    max_cfg = 0
    for r in range(2, cm.max_row + 1):
        sc = str(cm.cell(r, ci["Packaging Set Code"]).value or "").strip()
        st = str(cm.cell(r, ci["Configuration Status"]).value or "").strip()
        src = str(cm.cell(r, ci["Source Configuration ID"]).value or "")
        m = re.search(r"IA-ST-CFG-(\d+)", src)
        if m:
            max_cfg = max(max_cfg, int(m.group(1)))
        m2 = re.match(r"(ST-\d{3}-(?:STD|EUR))-(\d{2})$", sc)
        if m2:
            variants[m2.group(1)] = max(variants.get(m2.group(1), 0), int(m2.group(2)))
        if sc and st == "CONTROLLED":
            controlled_sets.add(sc)

    existing_bom: dict[str, list[tuple[str, float, str]]] = defaultdict(list)
    for r in range(2, bm.max_row + 1):
        sc = str(bm.cell(r, bi["Packaging Set Code"]).value or "").strip()
        code = str(bm.cell(r, bi["Component Code"]).value or "").strip()
        if not sc or not code or sc not in controlled_sets:
            continue
        qty = float(bm.cell(r, bi["Quantity"]).value or 0)
        uom = str(bm.cell(r, bi["UOM"]).value or "")
        existing_bom[sc].append((code, qty, uom))
    sig_to_set = {bom_sig(v): sc for sc, v in existing_bom.items()}

    # cluster new products
    clusters: dict[tuple, list[dict]] = defaultdict(list)
    for p in products:
        lines = [(c["code"], c["qty"], c["uom"]) for c in p["components"]]
        clusters[bom_sig(lines)].append(p)

    created_sets = []
    product_links: list[tuple[str, str, dict]] = []  # pc, set, product

    for sig, group in clusters.items():
        hit = sig_to_set.get(sig)
        if hit:
            set_code = hit
            # append products to existing linked list
            for r in range(2, cm.max_row + 1):
                if str(cm.cell(r, ci["Packaging Set Code"]).value) == set_code:
                    linked = str(cm.cell(r, ci["Linked Product Codes"]).value or "")
                    codes = [x.strip() for x in linked.split(";") if x.strip()]
                    for p in group:
                        if p["product_code"] not in codes:
                            codes.append(p["product_code"])
                    cm.cell(r, ci["Linked Product Codes"]).value = "; ".join(codes)
                    cm.cell(r, ci["Product Count"]).value = len(codes)
                    break
        else:
            sample = group[0]
            set_code = next_set_code(variants, sample["nominal_qty"])
            max_cfg += 1
            src_id = f"IA-ST-CFG-{max_cfg:04d}"
            final_id = f"IA-{set_code}"
            tare = compute_tare(sample["components"], weights)
            vb_tr, vb_en = variant_basis(sample["components"], sample["battery_type"])
            linked = "; ".join(p["product_code"] for p in group)
            desc = (
                f"Starter {int(sample['nominal_qty'])}-unit | "
                f"Standard Pallet 1030×1110×152 mm | Variant {set_code[-2:]}"
            )
            row = [None] * len(ch)
            # map by header
            values = {
                "Packaging Set Code": set_code,
                "Linked Product Codes": linked,
                "Source Configuration ID": src_id,
                "Final Configuration ID": final_id,
                "Product Count": len(group),
                "Nominal Qty": sample["nominal_qty"],
                "Packaging Description": desc,
                "Packaging Tare kg": tare,
                "BOM Line Count": len(sample["components"]),
                "Configuration Status": "CONTROLLED",
                "Origin": "NEW_AL_ALAMIA_2026_08",
                "Variant Basis TR": vb_tr,
                "Variant Basis EN": vb_en,
                "Technical File ID": f"IA-PPWR-TF-{set_code}-R00",
                "EU DoC ID": f"IA-PPWR-DOC-{set_code}-R00",
                "Label ID": f"IA-PPWR-LBL-{set_code}-R00",
                "Shipment Statement ID": f"IA-PPWR-STM-{set_code}-R00",
                "TF Type No": "YS/D/0020",
                "DoC Type No": "YS/D/0021",
                "Label Type No": "YS/D/0022",
                "STM Type No": "YS/D/0023",
                "Notes": "New physical packaging family from Al Alamia intake",
            }
            cm.append([values.get(h) for h in ch])

            # BOM rows
            for c in sample["components"]:
                uw, per = weights.get(c["code"], (None, 0.0))
                if c["uom"] == "KG" or (isinstance(uw, str) and "MASS" in str(uw).upper()):
                    unit_w = "MASS-BASED / N/A"
                    line_w = c["qty"]
                elif isinstance(uw, (int, float)):
                    unit_w = float(uw)
                    line_w = round(float(uw) * c["qty"], 4)
                else:
                    unit_w = per if per else None
                    line_w = round(per * c["qty"], 4) if per else None
                bm.append(
                    [
                        set_code if h == "Packaging Set Code" else
                        linked if h == "Linked Product Codes" else
                        src_id if h == "Source Configuration ID" else
                        c["code"] if h == "Component Code" else
                        c["desc"] if h == "Component Description" else
                        c["qty"] if h == "Quantity" else
                        c["uom"] if h == "UOM" else
                        unit_w if h == "Unit Weight" else
                        line_w if h == "Line Weight" else
                        None
                        for h in bh
                    ]
                )
            created_sets.append(set_code)
            print(f"NEW SET {set_code} products={[p['product_code'] for p in group]} tare={tare}")

        for p in group:
            product_links.append((p["product_code"], set_code, p))

    # PRODUCT_MASTER rows
    existing_pc = set()
    for r in range(2, pm.max_row + 1):
        existing_pc.add(str(pm.cell(r, pi["Product Code"]).value or "").strip())

    # lookup set meta for product rows
    set_meta = {}
    for r in range(2, cm.max_row + 1):
        sc = str(cm.cell(r, ci["Packaging Set Code"]).value or "").strip()
        if not sc:
            continue
        set_meta[sc] = {
            "src": cm.cell(r, ci["Source Configuration ID"]).value,
            "final": cm.cell(r, ci["Final Configuration ID"]).value,
            "tare": cm.cell(r, ci["Packaging Tare kg"]).value,
            "nom": cm.cell(r, ci["Nominal Qty"]).value,
        }

    added = 0
    for pc, set_code, p in product_links:
        if pc in existing_pc:
            print("SKIP existing product", pc)
            continue
        meta = set_meta[set_code]
        values = {
            "Product Code": pc,
            "Packaging Set Code": set_code,
            "Technical Description": p["description"],
            "Customer / Market": p["market"],
            "Source Configuration ID": meta["src"],
            "Final Configuration ID": meta["final"],
            "Packaging Tare kg": meta["tare"],
            "Scope Status": "EXPORT-READY STARTER SCOPE",
            "Physical Packaging Status": "CONTROLLED PACKAGING SET",
            "Battery Type": p["battery_type"],
            "Nominal Qty": p["nominal_qty"],
            "Legacy Market Status": "NEW / AL ALAMIA",
        }
        pm.append([values.get(h) for h in ph])
        added += 1

    wb.save(MASTER)
    wb.close()
    print("CREATED_SETS", len(created_sets), created_sets)
    print("PRODUCTS_ADDED", added)
    print("MASTER", MASTER)


if __name__ == "__main__":
    main()
