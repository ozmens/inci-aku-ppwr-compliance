"""Lock Starter master to source-validated 287 + freeze source package."""

from __future__ import annotations

import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
MASTER = ROOT / "output" / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
BACKUP = ROOT / "output" / "INCI_AKU_PPWR_STARTER_MASTER_Rev00_PRE_287_LOCK_BACKUP.xlsx"
VALIDATION = ROOT / "output" / "STARTER_71_NEW_SET_SOURCE_VALIDATION_Rev00.xlsx"
RECOVERY_QA = ROOT / "output" / "STARTER_SOURCE_BOM_RECOVERY_QA.md"
ENGINE = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
FREEZE = ROOT / "output" / "INCI_AKU_PPWR_STARTER_SOURCE_LOCK_287_Rev00"
FREEZE_ZIP = ROOT / "output" / "INCI_AKU_PPWR_STARTER_SOURCE_LOCK_287_Rev00.zip"
FREEZE_SHA = ROOT / "output" / "INCI_AKU_PPWR_STARTER_SOURCE_LOCK_287_Rev00_SHA256.txt"

PHYS_CTRL = "CONTROLLED PACKAGING SET"
PHYS_DATA = "DATA REQUIRED — COMPLETE PHYSICAL PACKAGING BOM REQUIRED"
SCOPE_IN = "IN PPWR SCOPE"
NAVY, WHITE, INK, BAND = "0E2A47", "FFFFFF", "1C2430", "F3F6F9"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)


def sha256_file(p: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_table(ws, headers, rows):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name="Tahoma", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = HAIR
    for r_i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r_i + 2, c, v)
            cell.font = Font(name="Tahoma", size=9, color=INK)
            cell.border = HAIR
            cell.fill = PatternFill("solid", fgColor=BAND if r_i % 2 else WHITE)
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(str(h)) + 2, 12), 40)


def lock_master() -> dict:
    shutil.copy2(MASTER, BACKUP)
    wb = load_workbook(MASTER)
    ws = wb["PRODUCT_MASTER"]
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ctrl = data = 0
    for r in range(2, ws.max_row + 1):
        phys = str(ws.cell(r, h.index("Physical Packaging Status") + 1).value or "")
        sc = str(ws.cell(r, h.index("Packaging Set Code") + 1).value or "")
        is_data = (
            "DATA REQUIRED" in phys
            or "BOM DATA REQUIRED" in phys
            or sc in {"BOM DATA REQUIRED", "NOT ISSUED", "NOT ISSUED / DATA REQUIRED"}
            or "DATA REQUIRED" in sc
        )
        if is_data:
            ws.cell(r, h.index("Packaging Set Code") + 1).value = "NOT ISSUED / DATA REQUIRED"
            ws.cell(r, h.index("Physical Packaging Status") + 1).value = PHYS_DATA
            ws.cell(r, h.index("Scope Status") + 1).value = SCOPE_IN
            ws.cell(r, h.index("Final Configuration ID") + 1).value = "NOT ISSUED"
            data += 1
        else:
            ws.cell(r, h.index("Physical Packaging Status") + 1).value = PHYS_CTRL
            ctrl += 1

    for sheet in ("DOCUMENT_SCOPE", "SEARCH_DATA", "DOCUMENT_CENTER", "TECHNICAL_FILES", "DECLARATIONS_OF_CONFORMITY", "LABELS", "SHIPMENT_STATEMENTS"):
        if sheet not in wb.sheetnames:
            continue
        dws = wb[sheet]
        dh = [c.value for c in next(dws.iter_rows(min_row=1, max_row=1))]
        if not dh or "Packaging Set Code" not in dh:
            continue
        for r in range(2, dws.max_row + 1):
            sc = str(dws.cell(r, dh.index("Packaging Set Code") + 1).value or "")
            if "DATA REQUIRED" in sc or sc in {"BOM DATA REQUIRED", "NOT ISSUED"}:
                dws.cell(r, dh.index("Packaging Set Code") + 1).value = "NOT ISSUED / DATA REQUIRED"
                for col in ("Technical File ID", "EU DoC ID", "Label ID", "Shipment Statement ID", "Controlled ID", "Document ID"):
                    if col in dh:
                        dws.cell(r, dh.index(col) + 1).value = "NOT ISSUED"
                if "Physical Packaging Status" in dh:
                    dws.cell(r, dh.index("Physical Packaging Status") + 1).value = PHYS_DATA
                if "Scope Status" in dh:
                    dws.cell(r, dh.index("Scope Status") + 1).value = SCOPE_IN
                if "Status" in dh:
                    dws.cell(r, dh.index("Status") + 1).value = "NOT ISSUED"

    cws = wb["CONFIG_MASTER"]
    ch = [c.value for c in next(cws.iter_rows(min_row=1, max_row=1))]
    controlled = sum(
        1
        for r in range(2, cws.max_row + 1)
        if str(cws.cell(r, ch.index("Configuration Status") + 1).value) == "CONTROLLED"
    )
    if "00_HOME" in wb.sheetnames:
        home = wb["00_HOME"]
        home["A1"] = "İNCI AKÜ PPWR — STARTER MASTER Rev.00 — FINAL SOURCE LOCK 287"
        home["A3"] = "Controlled Packaging Sets: 287 | Controlled Products: 2004 | DATA REQUIRED: 42"
        home["A4"] = "Active control count is 287. Historical 311 is archive/quarantine only."
    wb.save(MASTER)
    wb.close()
    assert ctrl == 2004 and data == 42 and controlled == 287, (ctrl, data, controlled)
    return {"controlled_products": ctrl, "data_required": data, "controlled_sets": controlled}


def freeze() -> str:
    if FREEZE.exists():
        shutil.rmtree(FREEZE)
    FREEZE.mkdir(parents=True)

    # reconciliation + config register physical workbook
    recon = FREEZE / "PRODUCT_PACKAGING_SET_RECONCILIATION.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "RECONCILIATION"
    sw = load_workbook(MASTER, data_only=True, read_only=True)
    ph = [c.value for c in next(sw["PRODUCT_MASTER"].iter_rows(min_row=1, max_row=1))]
    write_table(
        ws,
        ["Product Code", "Packaging Set Code", "Physical Packaging Status", "Source Configuration ID", "Scope Status", "Final Configuration ID", "Packaging Tare kg"],
        [
            [
                r[ph.index("Product Code")],
                r[ph.index("Packaging Set Code")],
                r[ph.index("Physical Packaging Status")],
                r[ph.index("Source Configuration ID")],
                r[ph.index("Scope Status")],
                r[ph.index("Final Configuration ID")],
                r[ph.index("Packaging Tare kg")],
            ]
            for r in sw["PRODUCT_MASTER"].iter_rows(min_row=2, values_only=True)
        ],
    )
    ws2 = wb.create_sheet("CONFIG_REGISTER")
    ch = [c.value for c in next(sw["CONFIG_MASTER"].iter_rows(min_row=1, max_row=1))]
    write_table(ws2, ch, [list(r) for r in sw["CONFIG_MASTER"].iter_rows(min_row=2, values_only=True)])
    ws3 = wb.create_sheet("BOM_QA")
    bh = [c.value for c in next(sw["BOM_MASTER"].iter_rows(min_row=1, max_row=1))]
    write_table(ws3, bh, [list(r) for r in sw["BOM_MASTER"].iter_rows(min_row=2, values_only=True)])
    sw.close()
    wb.save(recon)
    wb.close()

    copies = {
        "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx": MASTER,
        "INCI_AKU_PPWR_STARTER_MASTER_Rev00_PRE_287_LOCK_BACKUP.xlsx": BACKUP,
        "STARTER_71_NEW_SET_SOURCE_VALIDATION_Rev00.xlsx": VALIDATION,
        "STARTER_SOURCE_BOM_RECOVERY_QA.md": RECOVERY_QA,
        "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_PRE_GEN.xlsx": ENGINE if ENGINE.exists() else None,
    }
    for name, src in copies.items():
        if src and src.exists():
            shutil.copy2(src, FREEZE / name)

    (FREEZE / "SOURCE_LOCK_README.txt").write_text(
        "INCI AKU PPWR STARTER SOURCE LOCK 287 Rev.00\n"
        "Controlled Packaging Sets = 287 (240 trusted + 47 new source-proven)\n"
        "Controlled Products = 2004\n"
        "DATA REQUIRED = 42\n"
        f"Locked UTC: {datetime.now(timezone.utc).isoformat()}\n"
        "Physical copies only — no junctions.\n"
        "Do NOT promote quarantine delivery.\n",
        encoding="utf-8",
    )

    man = []
    for p in sorted(FREEZE.rglob("*")):
        if p.is_file():
            man.append(f"{sha256_file(p)}  {p.relative_to(FREEZE).as_posix()}")
    (FREEZE / "SHA256_MANIFEST.txt").write_text("\n".join(man) + "\n", encoding="utf-8")

    if FREEZE_ZIP.exists():
        FREEZE_ZIP.unlink()
    with zipfile.ZipFile(FREEZE_ZIP, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in FREEZE.rglob("*"):
            if p.is_file():
                zf.write(p, p.relative_to(FREEZE.parent).as_posix())
    digest = sha256_file(FREEZE_ZIP)
    FREEZE_SHA.write_text(digest + "\n", encoding="utf-8")
    return digest


def main():
    stats = lock_master()
    digest = freeze()
    print({"lock": stats, "freeze_zip": str(FREEZE_ZIP), "sha256": digest})


if __name__ == "__main__":
    main()
