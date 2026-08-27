"""Close 42 domestic DATA GAP starter products by assigning existing packaging sets.

User directive: same pallet/BOM family as controlled peers — take from other products.
Match by battery form factor → most common controlled Packaging Set Code.
"""

from __future__ import annotations

import json
import shutil
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from bilingual_format import translate_component, translate_product  # noqa: E402
from premium_pack_from_runtime import build_premium_pack  # noqa: E402
from convert_pdfs_libreoffice import convert_batch_via_temp, find_soffice  # noqa: E402
from ppwr_engine_builder import build_document_engine, verify_links  # noqa: E402
from fix_all_premium_bilingual import zip_delivery, DATE  # noqa: E402

MASTER = ROOT / "output" / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
MASTER_BAK = ROOT / "output" / f"INCI_AKU_PPWR_STARTER_MASTER_Rev00_BACKUP_BEFORE_GAP_CLOSE_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
STARTER = ROOT / "output" / "01_STARTER_INDIVIDUAL_DELIVERY_REV00"
PRODUCTS = STARTER / "01_PRODUCTS"
CONTROL = STARTER / "00_CONTROL"
REPORT = ROOT / "output" / "_DOMESTIC_GAP_CLOSE_QA.json"

FORM_TOKENS = [
    "TRACTORL",
    "TRACTOR",
    "C-IGYASMF",
    "C-IGYSMF",
    "C-IGYA",
    "B-IGYSMF",
    "B-IGYA",
    "C-FLAT",
    "B-FLAT",
    "A-IGYA",
    "A-FLAT",
    "D31",
    "D26",
    "D23",
    "L6",
    "L5",
    "L3",
    "L2",
    "L1",
    "AGM",
]
FORM_ALIAS = {
    "C-IGYSMF": "C-IGYA",
    "C-IGYASMF": "C-IGYASMF",
}


def _norm(s: str) -> str:
    return (
        (s or "")
        .upper()
        .replace("İ", "I")
        .replace("Ş", "S")
        .replace("Ğ", "G")
        .replace("Ü", "U")
        .replace("Ö", "O")
        .replace("Ç", "C")
    )


def form_key(desc: str) -> str:
    d = _norm(desc)
    for tok in FORM_TOKENS:
        if tok in d:
            return FORM_ALIAS.get(tok, tok)
    return "OTHER"


def load_set_bom(wb, set_code: str) -> tuple[list[dict], float | None, dict]:
    bm = wb["BOM_MASTER"]
    bh = [c.value for c in next(bm.iter_rows(min_row=1, max_row=1))]
    bhi = {h: i for i, h in enumerate(bh)}
    lines = []
    for row in bm.iter_rows(min_row=2, values_only=True):
        if str(row[bhi["Packaging Set Code"]] or "").strip() != set_code:
            continue
        desc = str(row[bhi["Component Description"]] or "").strip()
        lines.append(
            {
                "component_code": str(row[bhi["Component Code"]] or "").strip(),
                "description": desc,
                "name_en": translate_component(desc),
                "qty": row[bhi["Quantity"]],
                "uom": str(row[bhi["UOM"]] or "ADT").split("/")[0].strip() or "ADT",
                "unit_weight": row[bhi["Unit Weight"]] if "Unit Weight" in bhi else None,
                "line_weight": row[bhi["Line Weight"]] if "Line Weight" in bhi else None,
            }
        )
    cm = wb["CONFIG_MASTER"]
    ch = [c.value for c in next(cm.iter_rows(min_row=1, max_row=1))]
    chi = {h: i for i, h in enumerate(ch)}
    meta = {}
    tare = None
    for row in cm.iter_rows(min_row=2, values_only=True):
        if str(row[chi["Packaging Set Code"]] or "").strip() != set_code:
            continue
        tare = row[chi["Packaging Tare kg"]]
        meta = {h: row[chi[h]] for h in ch}
        break
    return lines, float(tare) if tare is not None else None, meta


def main() -> int:
    print("DOMESTIC GAP CLOSE start", flush=True)
    shutil.copy2(MASTER, MASTER_BAK)
    print("backup", MASTER_BAK.name, flush=True)

    wb = load_workbook(MASTER)
    pm = wb["PRODUCT_MASTER"]
    headers = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
    hi = {h: i + 1 for i, h in enumerate(headers)}  # 1-based for cell()

    # Build form → set frequency from controlled rows
    form_sets: dict[str, Counter] = defaultdict(Counter)
    for row in pm.iter_rows(min_row=2, values_only=True):
        st = str(row[headers.index("Physical Packaging Status")] or "")
        if "CONTROLLED" not in st:
            continue
        sc = str(row[headers.index("Packaging Set Code")] or "").strip()
        if not sc or "NOT" in sc:
            continue
        fk = form_key(str(row[headers.index("Technical Description")] or ""))
        form_sets[fk][sc] += 1

    default_set = max(
        ((sum(c.values()), c.most_common(1)[0][0]) for c in form_sets.values() if c),
        default=(0, "ST-072-STD-13"),
    )[1]

    assignments = []
    for r in range(2, pm.max_row + 1):
        st = str(pm.cell(r, hi["Physical Packaging Status"]).value or "")
        if "DATA REQUIRED" not in st:
            continue
        pc = str(pm.cell(r, hi["Product Code"]).value or "").strip()
        desc = str(pm.cell(r, hi["Technical Description"]).value or "")
        fk = form_key(desc)
        tops = form_sets.get(fk) or form_sets.get("OTHER")
        set_code = tops.most_common(1)[0][0] if tops else default_set
        bom, tare, meta = load_set_bom(wb, set_code)
        if not bom:
            raise RuntimeError(f"No BOM for set {set_code} (product {pc})")
        final_id = str(meta.get("Final Configuration ID") or f"IA-{set_code}")
        src_id = str(meta.get("Source Configuration ID") or "")

        pm.cell(r, hi["Packaging Set Code"]).value = set_code
        pm.cell(r, hi["Final Configuration ID"]).value = final_id
        if src_id:
            pm.cell(r, hi["Source Configuration ID"]).value = src_id
        pm.cell(r, hi["Packaging Tare kg"]).value = tare
        pm.cell(r, hi["Physical Packaging Status"]).value = "CONTROLLED PACKAGING SET"
        # keep Scope Status if present
        if "Scope Status" in hi:
            cur = str(pm.cell(r, hi["Scope Status"]).value or "")
            if "GAP" in cur.upper() or "NOT" in cur.upper() or not cur:
                pm.cell(r, hi["Scope Status"]).value = "IN PPWR SCOPE — ISSUED"

        assignments.append(
            {
                "product_code": pc,
                "description": desc,
                "form": fk,
                "set_code": set_code,
                "final_id": final_id,
                "tare": tare,
                "bom_lines": len(bom),
                "row": r,
            }
        )
        print(f"ASSIGN {pc} {fk} -> {set_code} tare={tare} bom={len(bom)}", flush=True)

    # Update CONFIG_MASTER linked product codes + counts
    cm = wb["CONFIG_MASTER"]
    ch = [c.value for c in next(cm.iter_rows(min_row=1, max_row=1))]
    chi = {h: i + 1 for i, h in enumerate(ch)}
    by_set = defaultdict(list)
    for a in assignments:
        by_set[a["set_code"]].append(a["product_code"])

    for r in range(2, cm.max_row + 1):
        sc = str(cm.cell(r, chi["Packaging Set Code"]).value or "").strip()
        if sc not in by_set:
            continue
        linked = str(cm.cell(r, chi["Linked Product Codes"]).value or "")
        codes = [c.strip() for c in linked.replace(",", ";").split(";") if c.strip()]
        for pc in by_set[sc]:
            if pc not in codes:
                codes.append(pc)
        codes = sorted(set(codes), key=lambda x: (len(x), x))
        cm.cell(r, chi["Linked Product Codes"]).value = "; ".join(codes)
        if "Product Count" in chi:
            cm.cell(r, chi["Product Count"]).value = len(codes)
        note = str(cm.cell(r, chi["Notes"]).value or "") if "Notes" in chi else ""
        add = f"GAP-CLOSE {datetime.now():%Y-%m-%d}: +{len(by_set[sc])} domestic products assigned"
        if "Notes" in chi:
            cm.cell(r, chi["Notes"]).value = (note + " | " + add).strip(" |")

    # DOCUMENT_SCOPE / SEARCH_DATA refresh lightly if sheets exist
    for sheet_name in ("DOCUMENT_SCOPE", "SEARCH_DATA", "PROVISIONAL_REGISTER"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # skip heavy rewrite — product packs are source of truth for delivery

    wb.save(MASTER)
    print(f"master updated assignments={len(assignments)}", flush=True)

    # Build packs
    pdfs = []
    pack_info = []
    wb_data = load_workbook(MASTER, data_only=True)
    for a in assignments:
        bom, tare, meta = load_set_bom(wb_data, a["set_code"])
        dest = PRODUCTS / a["product_code"]
        info = build_premium_pack(
            dest,
            key=a["product_code"],
            description_tr=a["description"],
            description_en=translate_product(a["description"]),
            set_code=a["set_code"],
            bom_lines=bom,
            scope="STARTER",
            total_tare_kg=tare,
            config_id=a["final_id"],
        )
        for stem in ["01_Technical_File", "02_EU_DoC", "03_Label", "04_Shipment_Statement"]:
            pdfs.append(dest / f"{stem}.docx")
        pack_info.append({**a, "photos": info.get("photos"), "pack": str(dest)})
        print(
            f"PACK {a['product_code']} photos={info.get('photos')} bom={info.get('bom')}",
            flush=True,
        )
    wb_data.close()

    print(f"PDF jobs={len(pdfs)}", flush=True)
    ok, fail = convert_batch_via_temp(
        find_soffice(), pdfs, ROOT / "output" / "_lo_profile_gap_close", chunk=40
    )
    print(f"PDF ok={ok} fail={fail}", flush=True)

    # Clear / rewrite DOMESTIC_DATA_GAP_LIST
    gap_xlsx = CONTROL / "DOMESTIC_DATA_GAP_LIST.xlsx"
    gwb = Workbook()
    gws = gwb.active
    gws.title = "DATA_REQUIRED"
    gws["A1"] = "Product Code"
    gws["B1"] = "Note"
    gws["A1"].font = Font(bold=True, color="FFFFFF")
    gws["B1"].font = Font(bold=True, color="FFFFFF")
    gws["A1"].fill = PatternFill("solid", fgColor="0E2A47")
    gws["B1"].fill = PatternFill("solid", fgColor="0E2A47")
    gws["A2"] = "(none)"
    gws["B2"] = (
        f"CLOSED {datetime.now():%Y-%m-%d %H:%M} — {len(assignments)} products assigned "
        "controlled packaging sets (shared pallet/BOM family). Remaining gap count: 0"
    )
    gwb.save(gap_xlsx)
    # also desktop copy
    desk = Path.home() / "Desktop" / "DOMESTIC_DATA_GAP_LIST.xlsx"
    try:
        shutil.copy2(gap_xlsx, desk)
    except Exception:
        pass

    # Assignment audit on desktop
    audit = Path.home() / "Desktop" / "DOMESTIC_GAP_CLOSE_ASSIGNMENTS.xlsx"
    awb = Workbook()
    aws = awb.active
    aws.title = "ASSIGNMENTS"
    headers_a = [
        "#",
        "Product Code",
        "Form",
        "Packaging Set",
        "Final Config ID",
        "Tare kg",
        "BOM lines",
        "Photos",
        "Description",
    ]
    for i, h in enumerate(headers_a, 1):
        c = aws.cell(1, i, h)
        c.fill = PatternFill("solid", fgColor="0E2A47")
        c.font = Font(color="FFFFFF", bold=True)
    for i, a in enumerate(pack_info, 1):
        aws.append(
            [
                i,
                a["product_code"],
                a["form"],
                a["set_code"],
                a["final_id"],
                a["tare"],
                a["bom_lines"],
                a.get("photos"),
                a["description"],
            ]
        )
    awb.save(audit)

    # Rebuild starter engine for all products
    keys = sorted(p.name for p in PRODUCTS.iterdir() if p.is_dir())
    print(f"ENGINE rebuild folders={len(keys)}", flush=True)
    build_document_engine(
        delivery_root=STARTER,
        engine_filename="INCI_AKU_PPWR_STARTER_INDIVIDUAL_ENGINE_Rev00.xlsx",
        title="İNCI AKÜ PPWR — STARTER INDIVIDUAL ENGINE Rev00",
        docs_subdir="01_PRODUCTS",
        records=[{"key": k, "label": k} for k in keys],
        scope_label="STARTER PACKAGING COMPLIANCE",
        extra_home={
            "PUBLISH DATE": DATE,
            "SIGNATORY": "Numan Alver — Operations Director",
            "FORMAT": "Premium UI + absolute HYPERLINK + gap-close 42",
            "DATA GAP": "0 (closed)",
        },
    )
    v = verify_links(STARTER, "01_PRODUCTS", keys)
    print("verify", v, flush=True)
    digest = zip_delivery(STARTER)
    print("ZIP", digest, flush=True)

    report = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "assignments": len(assignments),
        "pdf_ok": ok,
        "pdf_fail": fail,
        "starter_folders": len(keys),
        "verify": v,
        "sha256": digest,
        "backup_master": str(MASTER_BAK),
        "audit_xlsx": str(audit),
        "items": pack_info,
    }
    REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({k: report[k] for k in report if k != "items"}, indent=2), flush=True)
    print("OVERALL", "PASS" if fail == 0 and v.get("missing", 1) == 0 else "FAIL", flush=True)
    return 0 if fail == 0 and v.get("missing", 1) == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
