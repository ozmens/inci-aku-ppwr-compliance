"""FINAL REV00 delivery freeze — rename already done; validate + ZIP + SHA-256."""

from __future__ import annotations

import hashlib
import json
import sys
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from builders.phase_k.service import PhaseKService, UI_SHEETS, excel_open_ok  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> int:
    delivery = ROOT / "output" / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00_UI_READY"
    wb_path = delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx"
    zip_path = ROOT / "output" / "INCI_AKU_PPWR_FINAL_DELIVERY_REV00.zip"
    old = delivery / "INCI_AKU_PPWR_PIMS_Rev00_FINAL_UI_CANDIDATE_V2.xlsx"

    if not wb_path.exists():
        print("FAIL: renamed FINAL workbook missing:", wb_path)
        return 1
    if old.exists():
        print("FAIL: old V2 name still present:", old)
        return 1

    svc = PhaseKService(ROOT)
    svc.delivery = delivery
    svc.delivery_workbook = wb_path

    excel = excel_open_ok(wb_path)
    print("EXCEL", excel)

    validation = svc._validate_links(wb_path, delivery)
    print(
        "LINKS",
        {
            k: validation[k]
            for k in (
                "total_links",
                "existing",
                "missing",
                "broken_paths",
                "absolute_hits",
                "formula_instances",
            )
        },
    )

    wb = load_workbook(wb_path, data_only=False)
    home_ok = 0
    home_missing: list[str] = []
    for name in UI_SHEETS:
        if name not in wb.sheetnames:
            home_missing.append(f"{name}:MISSING_SHEET")
            continue
        a1 = str(wb[name]["A1"].value or "")
        if "Ana Sayfaya Dön" in a1 or "Turn Back Home" in a1:
            home_ok += 1
        else:
            home_missing.append(name)
    wb.close()
    print("HOME", home_ok, "/", len(UI_SHEETS), "missing", home_missing)

    samples = svc._sample_link_tests(validation)
    print("SAMPLES", samples)

    if zip_path.exists():
        zip_path.unlink()

    arc_root_name = "INCI_AKU_PPWR_FINAL_DELIVERY_REV00"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for path in sorted(delivery.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(delivery)
            arcname = f"{arc_root_name}/{rel.as_posix()}"
            zf.write(path, arcname)

    wb_sha = sha256(wb_path)
    zip_sha = sha256(zip_path)

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        docx = [n for n in names if n.lower().endswith(".docx")]
        xlsx = [n for n in names if n.lower().endswith(".xlsx")]

    print("ZIP", zip_path, "size", zip_path.stat().st_size)
    print("ZIP_MEMBERS", len(names), "DOCX", len(docx), "XLSX", xlsx)
    print("SHA_WORKBOOK", wb_sha)
    print("SHA_ZIP", zip_sha)

    gate = (
        bool(excel.get("ok"))
        and validation["total_links"] == 988
        and validation["existing"] == 988
        and validation["missing"] == 0
        and validation.get("broken_paths", 0) == 0
        and validation["absolute_hits"] == 0
        and home_ok == 13
        and all(s["exists"] for s in samples)
        and len(docx) == 988
        and any(n.endswith("INCI_AKU_PPWR_PIMS_Rev00_FINAL.xlsx") for n in xlsx)
        and not home_missing
    )
    status = "PASS" if gate else "FAIL"
    print("GATE", status)

    report = {
        "gate": status,
        "delivery_root": str(delivery),
        "workbook": str(wb_path),
        "zip": str(zip_path),
        "sha256_workbook": wb_sha,
        "sha256_zip": zip_sha,
        "excel": excel,
        "links": {
            "total": validation["total_links"],
            "existing": validation["existing"],
            "missing": validation["missing"],
            "broken_paths": validation.get("broken_paths", 0),
            "absolute_hits": validation["absolute_hits"],
        },
        "home_buttons": f"{home_ok}/13",
        "home_missing": home_missing,
        "samples": samples,
        "zip_docx_count": len(docx),
        "zip_member_count": len(names),
    }
    out = ROOT / "output"
    (out / "FINAL_REV00_DELIVERY_FREEZE_QA.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    md = "\n".join(
        [
            "# FINAL REV00 DELIVERY FREEZE",
            "",
            f"- **FINAL REV00 DELIVERY FREEZE: {status}**",
            "",
            f"- Delivery root: `{delivery}`",
            f"- Final workbook: `{wb_path}`",
            f"- Final ZIP: `{zip_path}`",
            "",
            "## SHA-256",
            "",
            f"- Workbook: `{wb_sha}`",
            f"- ZIP: `{zip_sha}`",
            "",
            "## Re-test",
            "",
            f"- Native Excel open: `{excel}`",
            f"- Document links: {validation['total_links']} total / "
            f"{validation['existing']} working / {validation['missing']} missing",
            f"- Broken paths: {validation.get('broken_paths', 0)}",
            f"- Absolute path hits: {validation['absolute_hits']}",
            f"- Home buttons: {home_ok}/13",
            f"- ZIP DOCX count: {len(docx)}",
            "",
            "## Confirmations",
            "",
            "- Canonical data changed: NO",
            "- Relative document links changed: NO",
            "- Word regenerated: NO",
            "- Golden templates modified: NO",
            "- Rev01 started: NO",
            "",
            f"**FINAL REV00 DELIVERY FREEZE: {status}**",
            "",
        ]
    )
    (out / "FINAL_REV00_DELIVERY_FREEZE_QA.md").write_text(md, encoding="utf-8")
    print("FINAL REV00 DELIVERY FREEZE:", status)
    return 0 if gate else 1


if __name__ == "__main__":
    raise SystemExit(main())
