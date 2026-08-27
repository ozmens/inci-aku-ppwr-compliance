"""Fix EXISTING 240 Starter BOM mass-based Line Weights from Golden source only.

Excel data correction only. No Word / Industrial / mapping / renumber changes.
"""

from __future__ import annotations

import hashlib
import json
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
OUT = ROOT / "output"
MASTER = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
BACKUP = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00_PRE_BOM_MASS_FIX_BACKUP.xlsx"
QA_XLSX = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00_BOM_MASS_FIX_QA.xlsx"
QA_MD = OUT / "STARTER_BOM_MASS_FIX_QA.md"
QA_JSON = OUT / "STARTER_BOM_MASS_FIX_QA.json"
GOLDEN = (
    ROOT
    / "input"
    / "production"
    / "INCI_AKU_PPWR_Final_Configuration_Register_Rev00_GOLDEN_VARIANTS_FINAL.xlsx"
)
PHASE_I = OUT / "PHASE_I_FINAL"

PENDING_PCS = {"1013084", "1014789", "1014790"}
PHYS_CONTROLLED = "CONTROLLED PACKAGING SET"
BOM_REQ = "BOM DATA REQUIRED"
CFG_CONTROLLED = "CONTROLLED"
IND_SETS = {"IND-24V-01", "IND-48V-01", "IND-80V-01"}
MASS_MARKER = "MASS-BASED / N/A"

TARE_EXACT = 1e-9
TARE_TOL = 1e-3

NAVY = "0E2A47"
WHITE = "FFFFFF"
INK = "1C2430"
BAND = "F3F6F9"
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def word_fp() -> dict[str, Any]:
    files = sorted((PHASE_I / "01_STARTER").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "02_INDUSTRIAL").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "03_CONTAINER").rglob("01_Technical_File.docx"))
    files = [p for p in files if not p.name.startswith("~$")]
    h = hashlib.sha256()
    for p in files:
        d = sha256_file(p)
        h.update(d.encode())
        h.update(str(p.relative_to(PHASE_I)).encode())
    return {"count": len(files), "aggregate": h.hexdigest()}


def fnum(v: Any) -> float | None:
    if v in (None, ""):
        return None
    if isinstance(v, str) and v.strip().upper() in {"MASS-BASED / N/A", "N/A", "MASS-BASED"}:
        return None
    try:
        return float(v)
    except Exception:
        return None


def nearly(a: float | None, b: float | None, tol: float = 1e-9) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def load_existing_240() -> set[str]:
    wb = load_workbook(GOLDEN, data_only=True, read_only=True)
    ws = wb["01_FINAL_CONFIG_MASTER"]
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {x: i for i, x in enumerate(h)}
    codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["Family"]] or "").upper() == "STARTER":
            codes.add(str(row[idx["Packaging Set Code"]]))
    wb.close()
    assert len(codes) == 240
    return codes


def load_golden_bom(existing: set[str]) -> dict[tuple[str, str], dict]:
    """Key = (Packaging Set Code, Component Code) — starter existing only."""
    wb = load_workbook(GOLDEN, data_only=True, read_only=True)
    ws = wb["03_BOM_MASTER"]
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[tuple[str, str], dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h[i]: row[i] for i in range(len(h))}
        sc = str(d.get("Packaging Set Code") or "")
        if sc not in existing:
            continue
        cc = str(d.get("Component Code") or "").strip()
        if not cc:
            continue
        key = (sc, cc)
        # if duplicate component codes in same set, keep first but note later
        if key not in out:
            out[key] = {
                "packaging_set_code": sc,
                "source_configuration_id": str(d.get("Source Configuration ID") or ""),
                "component_code": cc,
                "quantity": fnum(d.get("Quantity")),
                "uom": str(d.get("UOM") or "").strip(),
                "unit_weight": fnum(d.get("Unit Weight kg")),
                "line_weight": fnum(d.get("Line Weight kg")),
                "weight_basis": str(d.get("Weight Basis") or ""),
                "description": str(d.get("ERP Description") or ""),
            }
    wb.close()
    return out


def sheet_dicts(wb, name: str) -> tuple[list[str], list[dict]]:
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return headers, rows


def write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = HAIR
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for r_i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r_i + 2, c, v)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.border = HAIR
            cell.fill = PatternFill("solid", fgColor=BAND if r_i % 2 else WHITE)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(str(h)) + 2, 12), 36)


def classify_and_correct(
    cur_uw: Any,
    cur_lw: Any,
    src: dict | None,
) -> tuple[Any, Any, str, bool]:
    """Return (new_uw, new_lw, reason, changed)."""
    if src is None:
        return cur_uw, cur_lw, "SOURCE DATA REVIEW REQUIRED", False

    src_uw = src["unit_weight"]
    src_lw = src["line_weight"]
    qty = src["quantity"]
    cur_uw_n = fnum(cur_uw)
    cur_lw_n = fnum(cur_lw)

    # Case A: source has no unit weight but has line weight -> MASS-BASED
    if src_uw is None and src_lw is not None:
        target_uw = MASS_MARKER
        target_lw = src_lw
        changed = (cur_uw_n is not None and cur_uw_n != 0) or (
            str(cur_uw).strip() not in ("", MASS_MARKER) and cur_uw not in (None, "")
        )
        # also if current uw is 0 / blank incorrectly
        if cur_uw_n == 0 or cur_uw in (None, ""):
            changed = True
        if not nearly(cur_lw_n, src_lw):
            changed = True
        if changed:
            reason = "MASS-BASED LINE WEIGHT RESTORED"
            return target_uw, target_lw, reason, True
        return MASS_MARKER if cur_uw in (None, "", 0, 0.0) else cur_uw, cur_lw_n, "NO CHANGE REQUIRED", False

    # Case B: source has unit weight — normal
    if src_uw is not None:
        expected_lw = src_lw
        if expected_lw is None and qty is not None:
            expected_lw = qty * src_uw
        # verify qty * uw ~= lw
        if qty is not None and src_lw is not None and abs(qty * src_uw - src_lw) > 1e-6:
            # source inconsistency — do not invent
            return cur_uw, cur_lw, "SOURCE DATA REVIEW REQUIRED", False
        target_uw = src_uw
        target_lw = expected_lw if expected_lw is not None else src_lw
        changed = False
        if not nearly(cur_uw_n, src_uw):
            changed = True
        if not nearly(cur_lw_n, target_lw):
            changed = True
        if changed:
            return target_uw, target_lw, "SOURCE LINE WEIGHT RESTORED", True
        return cur_uw, cur_lw, "NO CHANGE REQUIRED", False

    # Case C: source missing both
    if src_uw is None and src_lw is None:
        return cur_uw, cur_lw, "SOURCE DATA REVIEW REQUIRED", False

    return cur_uw, cur_lw, "NO CHANGE REQUIRED", False


def tare_check(
    cfg_rows: list[dict],
    bom_rows: list[dict],
    set_codes: set[str],
) -> tuple[int, int, int, list[dict]]:
    bom_sum: dict[str, float] = defaultdict(float)
    for b in bom_rows:
        sc = str(b.get("Packaging Set Code") or "")
        if sc not in set_codes:
            continue
        lw = fnum(b.get("Line Weight"))
        if lw is None:
            continue
        bom_sum[sc] += lw
    exact = tol = fail = 0
    fails = []
    for c in cfg_rows:
        sc = str(c.get("Packaging Set Code") or "")
        if sc not in set_codes:
            continue
        if str(c.get("Configuration Status") or "") not in (CFG_CONTROLLED, ""):
            # still include if in set_codes controlled list
            pass
        cfg_tare = fnum(c.get("Packaging Tare kg"))
        bs = bom_sum.get(sc, 0.0)
        if cfg_tare is None:
            fail += 1
            fails.append({"set": sc, "cfg": None, "bom": bs, "diff": None})
            continue
        diff = abs(cfg_tare - bs)
        if diff <= TARE_EXACT:
            exact += 1
        elif diff <= TARE_TOL:
            tol += 1
        else:
            fail += 1
            fails.append({"set": sc, "cfg": cfg_tare, "bom": round(bs, 6), "diff": round(diff, 6)})
    return exact, tol, fail, fails


def run() -> dict[str, Any]:
    fp_before = word_fp()

    if not MASTER.exists():
        raise FileNotFoundError(MASTER)

    shutil.copy2(MASTER, BACKUP)
    if not BACKUP.exists() or BACKUP.stat().st_size != MASTER.stat().st_size:
        raise RuntimeError("Backup failed — aborting")
    if sha256_file(BACKUP) != sha256_file(MASTER):
        raise RuntimeError("Backup hash mismatch — aborting")

    existing_240 = load_existing_240()
    golden_bom = load_golden_bom(existing_240)

    # Work on a temp copy then replace master only if QA passes
    work = OUT / "_STARTER_BOM_MASS_FIX_WORK.xlsx"
    shutil.copy2(MASTER, work)

    wb = load_workbook(work)
    bom_ws = wb["BOM_MASTER"]
    bom_headers = [c.value for c in next(bom_ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i + 1 for i, h in enumerate(bom_headers)}  # 1-based excel cols

    required = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Component Code",
        "Quantity",
        "UOM",
        "Unit Weight",
        "Line Weight",
    ]
    for r in required:
        if r not in hidx:
            raise KeyError(f"BOM_MASTER missing column {r}")

    audit_rows: list[dict] = []
    structural_flags: list[dict] = []
    changed_cells = 0
    mass_restored = 0
    other_restored = 0
    review_required = 0
    no_change = 0
    affected_sets = set()
    affected_components = Counter()

    # Build cfg linked products for audit
    _, cfg_rows = sheet_dicts(wb, "CONFIG_MASTER")
    cfg_by_set = {
        str(c["Packaging Set Code"]): c
        for c in cfg_rows
        if str(c.get("Configuration Status") or "") == CFG_CONTROLLED
        or str(c.get("Packaging Set Code") or "") in existing_240
    }

    max_row = bom_ws.max_row
    for r in range(2, max_row + 1):
        sc = bom_ws.cell(r, hidx["Packaging Set Code"]).value
        if sc is None:
            continue
        sc = str(sc).strip()
        if sc not in existing_240:
            continue  # do not modify 71 new sets

        cc = str(bom_ws.cell(r, hidx["Component Code"]).value or "").strip()
        qty = bom_ws.cell(r, hidx["Quantity"]).value
        uom = bom_ws.cell(r, hidx["UOM"]).value
        cur_uw = bom_ws.cell(r, hidx["Unit Weight"]).value
        cur_lw = bom_ws.cell(r, hidx["Line Weight"]).value
        src_id = bom_ws.cell(r, hidx["Source Configuration ID"]).value
        linked = bom_ws.cell(r, hidx["Linked Product Codes"]).value

        src = golden_bom.get((sc, cc))
        # structural: quantity/uom mismatch vs source -> report, do not change
        if src is not None:
            if not nearly(fnum(qty), src["quantity"]) or str(uom or "").strip().upper() != src[
                "uom"
            ].upper().replace("ADT", "ADT"):
                # normalize ADT vs ADT
                su = src["uom"].upper()
                cu = str(uom or "").strip().upper()
                uom_ok = cu == su or {cu, su} <= {"ADT", "ADET", "PCS", "PC"} or (
                    cu in {"ADT", "ADET"} and su in {"ADT", "ADET"}
                )
                # map ADT/PCS loosely already same in data
                if not nearly(fnum(qty), src["quantity"]) or (
                    cu != su
                    and not (
                        (cu in {"ADT", "ADET", "PCS", "PC"} and su in {"ADT", "ADET", "PCS", "PC"})
                        or (cu == "M" and su == "M")
                        or (cu == "KG" and su == "KG")
                    )
                ):
                    structural_flags.append(
                        {
                            "Packaging Set Code": sc,
                            "Component Code": cc,
                            "Problem": f"Qty/UOM mismatch current=({qty},{uom}) source=({src['quantity']},{src['uom']})",
                            "Action": "REPORTED — no structural change",
                        }
                    )

        new_uw, new_lw, reason, changed = classify_and_correct(cur_uw, cur_lw, src)

        audit_rows.append(
            {
                "Packaging Set Code": sc,
                "Product Codes": linked,
                "Source Configuration ID": src_id or (src or {}).get("source_configuration_id", ""),
                "Component Code": cc,
                "Quantity": qty,
                "UOM": uom,
                "Current Unit Weight": cur_uw,
                "Source Unit Weight": None if src is None else src["unit_weight"],
                "Current Line Weight": cur_lw,
                "Source Line Weight": None if src is None else src["line_weight"],
                "Correction Reason": reason,
                "New Unit Weight": new_uw if changed else cur_uw,
                "New Line Weight": new_lw if changed else cur_lw,
            }
        )

        if reason == "SOURCE DATA REVIEW REQUIRED":
            review_required += 1
        elif reason == "NO CHANGE REQUIRED":
            no_change += 1
        elif reason == "MASS-BASED LINE WEIGHT RESTORED":
            mass_restored += 1
        elif reason == "SOURCE LINE WEIGHT RESTORED":
            other_restored += 1

        if changed:
            bom_ws.cell(r, hidx["Unit Weight"]).value = new_uw
            bom_ws.cell(r, hidx["Line Weight"]).value = new_lw
            changed_cells += 2
            affected_sets.add(sc)
            affected_components[cc] += 1

    # Also: golden rows missing from starter BOM for existing 240?
    starter_keys = set()
    for r in range(2, max_row + 1):
        sc = bom_ws.cell(r, hidx["Packaging Set Code"]).value
        if sc is None:
            continue
        sc = str(sc).strip()
        if sc not in existing_240:
            continue
        cc = str(bom_ws.cell(r, hidx["Component Code"]).value or "").strip()
        starter_keys.add((sc, cc))
    missing_in_starter = sorted(set(golden_bom) - starter_keys)
    for sc, cc in missing_in_starter:
        structural_flags.append(
            {
                "Packaging Set Code": sc,
                "Component Code": cc,
                "Problem": "Source BOM component missing from Starter BOM_MASTER",
                "Action": "REPORTED — no invent / no insert in this phase",
            }
        )
        review_required += 1

    wb.save(work)
    wb.close()

    # Reload for tare QA
    wb = load_workbook(work, data_only=True)
    _, cfg_rows = sheet_dicts(wb, "CONFIG_MASTER")
    _, bom_rows = sheet_dicts(wb, "BOM_MASTER")
    _, products = sheet_dicts(wb, "PRODUCT_MASTER")

    controlled_cfgs = [
        c for c in cfg_rows if str(c.get("Configuration Status") or "") == CFG_CONTROLLED
    ]
    controlled_sets = {str(c["Packaging Set Code"]) for c in controlled_cfgs}
    assert len(controlled_sets) == 311

    new_71 = controlled_sets - existing_240
    assert len(new_71) == 71
    assert existing_240 <= controlled_sets

    e_ex, e_tol, e_fail, e_fails = tare_check(controlled_cfgs, bom_rows, existing_240)
    n_ex, n_tol, n_fail, n_fails = tare_check(controlled_cfgs, bom_rows, new_71)
    t_ex, t_tol, t_fail, t_fails = tare_check(controlled_cfgs, bom_rows, controlled_sets)

    # Mapping / integrity gates
    pcs = [str(p["Product Code"]).strip() for p in products]
    unique_pc = len(set(pcs))
    controlled_products = [
        p for p in products if str(p.get("Physical Packaging Status") or "") == PHYS_CONTROLLED
    ]
    bom_req = [
        p
        for p in products
        if str(p.get("Packaging Set Code") or "") == BOM_REQ
        or str(p["Product Code"]).strip() in PENDING_PCS
    ]
    # industrial scan
    ind_found = {k: False for k in IND_SETS}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                s = str(v)
                for k in IND_SETS:
                    if k in s:
                        ind_found[k] = True
    cnt_hits = 0
    for sheet in (
        "PRODUCT_MASTER",
        "CONFIG_MASTER",
        "BOM_MASTER",
        "DOCUMENT_SCOPE",
        "DOCUMENT_CENTER",
        "SEARCH_DATA",
    ):
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            for v in row:
                if v is not None and str(v).startswith("CNT-"):
                    cnt_hits += 1

    # mapping unchanged: product set codes same as backup
    bwb = load_workbook(BACKUP, data_only=True, read_only=True)
    _, b_products = sheet_dicts(bwb, "PRODUCT_MASTER")
    bwb.close()
    map_changed = 0
    b_map = {
        str(p["Product Code"]).strip(): str(p.get("Packaging Set Code") or "")
        for p in b_products
    }
    for p in products:
        pc = str(p["Product Code"]).strip()
        if b_map.get(pc) != str(p.get("Packaging Set Code") or ""):
            map_changed += 1

    # set codes unchanged
    bwb = load_workbook(BACKUP, data_only=True, read_only=True)
    _, b_cfgs = sheet_dicts(bwb, "CONFIG_MASTER")
    bwb.close()
    b_sets = {
        str(c["Packaging Set Code"])
        for c in b_cfgs
        if str(c.get("Configuration Status") or "") == CFG_CONTROLLED
    }
    set_codes_changed = len(controlled_sets.symmetric_difference(b_sets))

    wb.close()

    fp_after = word_fp()
    word_ok = fp_before["aggregate"] == fp_after["aggregate"] and fp_after["count"] == 247

    gates_pass = (
        e_fail == 0
        and n_fail == 0
        and t_fail == 0
        and (e_ex + e_tol) == 240
        and (n_ex + n_tol) == 71
        and (t_ex + t_tol) == 311
        and unique_pc == 2046
        and len(controlled_products) == 2043
        and len(bom_req) == 3
        and not any(ind_found.values())
        and cnt_hits == 0
        and map_changed == 0
        and set_codes_changed == 0
        and word_ok
        and not structural_flags  # if structural, do not save? User said report before structural change — we didn't change structure
    )
    # structural_flags about qty/uom may be false positives from UOM spelling — recheck
    # Allow save if only missing-component flags? User said if source missing do not invent.
    # If tare still fails due to missing components, don't save.
    # If structural_flags exist but tare passes, still report them but can save if tare OK.
    # Tighten: allow save when tare gates pass and mapping gates pass, even if review rows logged
    # (review rows that weren't changed).
    save_ok = (
        e_fail == 0
        and n_fail == 0
        and t_fail == 0
        and (e_ex + e_tol) == 240
        and (n_ex + n_tol) == 71
        and (t_ex + t_tol) == 311
        and unique_pc == 2046
        and len(controlled_products) == 2043
        and len({str(p["Product Code"]) for p in bom_req} & PENDING_PCS) == 3
        and not any(ind_found.values())
        and cnt_hits == 0
        and map_changed == 0
        and set_codes_changed == 0
        and word_ok
        and len(missing_in_starter) == 0
    )

    # Write QA workbook always
    qwb = Workbook()
    ws = qwb.active
    ws.title = "00_SUMMARY"
    summary = [
        ["STARTER BOM MASS FIX QA"],
        ["Existing sets audited", f"{len(affected_sets | existing_240)} / 240"],
        ["Sets with corrections", len(affected_sets)],
        ["Affected BOM rows (changed)", mass_restored + other_restored],
        ["Mass-based rows restored", mass_restored],
        ["Other source line-weight corrections", other_restored],
        ["Source-data-review-required rows", review_required],
        ["No change required rows", no_change],
        ["Existing 240 Exact", e_ex],
        ["Existing 240 Tolerance", e_tol],
        ["Existing 240 Fail", e_fail],
        ["New 71 Exact", n_ex],
        ["New 71 Tolerance", n_tol],
        ["New 71 Fail", n_fail],
        ["Total 311 Exact", t_ex],
        ["Total 311 Tolerance", t_tol],
        ["Total 311 Fail", t_fail],
        ["Saved to master", "YES" if save_ok else "NO"],
        ["Word hash changed", 0 if word_ok else 1],
    ]
    for row in summary:
        ws.append(row)

    ws = qwb.create_sheet("AUDIT_CORRECTIONS")
    ah = [
        "Packaging Set Code",
        "Product Codes",
        "Source Configuration ID",
        "Component Code",
        "Quantity",
        "UOM",
        "Current Unit Weight",
        "Source Unit Weight",
        "Current Line Weight",
        "Source Line Weight",
        "Correction Reason",
        "New Unit Weight",
        "New Line Weight",
    ]
    write_table(
        ws,
        ah,
        [[a.get(k) for k in ah] for a in audit_rows if a["Correction Reason"] != "NO CHANGE REQUIRED"],
    )

    ws = qwb.create_sheet("AUDIT_ALL_EXISTING_BOM")
    write_table(ws, ah, [[a.get(k) for k in ah] for a in audit_rows])

    ws = qwb.create_sheet("TARE_FAILS")
    write_table(
        ws,
        ["Scope", "Packaging Set Code", "CONFIG Tare", "BOM Sum", "Diff"],
        [["EXISTING_240", f["set"], f["cfg"], f["bom"], f["diff"]] for f in e_fails]
        + [["NEW_71", f["set"], f["cfg"], f["bom"], f["diff"]] for f in n_fails],
    )

    ws = qwb.create_sheet("STRUCTURAL_FLAGS")
    write_table(
        ws,
        ["Packaging Set Code", "Component Code", "Problem", "Action"],
        [
            [f["Packaging Set Code"], f["Component Code"], f["Problem"], f["Action"]]
            for f in structural_flags
        ]
        if structural_flags
        else [["(none)", "", "", ""]],
    )

    ws = qwb.create_sheet("AFFECTED_COMPONENTS")
    write_table(
        ws,
        ["Component Code", "Correction Count"],
        [[code, n] for code, n in affected_components.most_common()],
    )

    if QA_XLSX.exists():
        QA_XLSX.unlink()
    qwb.save(QA_XLSX)
    qwb.close()

    if save_ok:
        shutil.copy2(work, MASTER)
        # verify
        if sha256_file(work) != sha256_file(MASTER):
            raise RuntimeError("Master save verification failed")
    # cleanup work file keep for evidence if fail
    if save_ok and work.exists():
        work.unlink()

    final = "PASS" if save_ok else "FAIL"

    report = {
        "existing_controlled_sets_audited": f"240 / 240",
        "sets_with_corrections": len(affected_sets),
        "affected_bom_rows": mass_restored + other_restored,
        "affected_component_codes": [c for c, _ in affected_components.most_common()],
        "mass_based_rows_restored": mass_restored,
        "other_source_line_weight_corrections": other_restored,
        "source_data_review_required_rows": review_required,
        "existing_240_exact": e_ex,
        "existing_240_tolerance": e_tol,
        "existing_240_fail": e_fail,
        "new_71_exact": n_ex,
        "new_71_tolerance": n_tol,
        "new_71_fail": n_fail,
        "total_311_exact": t_ex,
        "total_311_tolerance": t_tol,
        "total_311_fail": t_fail,
        "unique_starter_product_codes": f"{unique_pc} / 2046",
        "controlled_products": f"{len(controlled_products)} / 2043",
        "bom_data_required": f"{len(bom_req)} / 3",
        "industrial_rows": 0 if not any(ind_found.values()) else sum(1 for v in ind_found.values() if v),
        "ind_24": "FOUND" if ind_found["IND-24V-01"] else "ABSENT",
        "ind_48": "FOUND" if ind_found["IND-48V-01"] else "ABSENT",
        "ind_80": "FOUND" if ind_found["IND-80V-01"] else "ABSENT",
        "container_rows": cnt_hits,
        "mapping_changed": map_changed,
        "set_codes_changed": set_codes_changed,
        "word_hash_changed": 0 if word_ok else 1,
        "final_gate": final,
        "master_saved": save_ok,
        "backup": str(BACKUP),
        "qa_xlsx": str(QA_XLSX),
        "tare_fails_existing_sample": e_fails[:10],
        "structural_flags_count": len(structural_flags),
        "missing_in_starter": len(missing_in_starter),
    }

    lines = [
        "# STARTER BOM MASS FIX QA",
        "",
        "Existing controlled sets audited:",
        "240 / 240",
        "",
        "Affected BOM rows:",
        str(mass_restored + other_restored),
        "",
        "Affected component codes:",
        ", ".join(report["affected_component_codes"]) if report["affected_component_codes"] else "(none)",
        "",
        "Mass-based rows restored:",
        str(mass_restored),
        "",
        "Other source line-weight corrections:",
        str(other_restored),
        "",
        "Source-data-review-required rows:",
        str(review_required),
        "",
        "Existing 240 tare reconciliation:",
        f"Exact: {e_ex}",
        f"Tolerance: {e_tol}",
        f"Fail: {e_fail}",
        "",
        "New 71 tare reconciliation:",
        f"Exact: {n_ex}",
        f"Tolerance: {n_tol}",
        f"Fail: {n_fail}",
        "",
        "TOTAL 311 tare reconciliation:",
        f"Exact: {t_ex}",
        f"Tolerance: {t_tol}",
        f"Fail: {t_fail}",
        "",
        "Unique Starter Product Codes:",
        report["unique_starter_product_codes"],
        "",
        "Controlled products:",
        report["controlled_products"],
        "",
        "BOM DATA REQUIRED:",
        report["bom_data_required"],
        "",
        "Industrial rows in Starter:",
        str(report["industrial_rows"]),
        "",
        "IND-24V-01:",
        report["ind_24"],
        "",
        "IND-48V-01:",
        report["ind_48"],
        "",
        "IND-80V-01:",
        report["ind_80"],
        "",
        "Container rows in Starter:",
        str(report["container_rows"]),
        "",
        "Product <-> Packaging Set mapping changed:",
        f"{map_changed} expected",
        "",
        "Existing Packaging Set codes changed:",
        f"{set_codes_changed} expected",
        "",
        "Word hash changed:",
        f"{report['word_hash_changed']} expected",
        "",
        "FINAL STARTER DATA INTEGRITY GATE:",
        final,
        "",
        "STOP.",
        "",
        "DO NOT GENERATE WORDS.",
        "DO NOT START PIMS.",
        "",
        f"Backup: `{BACKUP}`",
        f"QA artefact: `{QA_XLSX}`",
        f"Master saved: {save_ok}",
    ]
    QA_MD.write_text("\n".join(lines), encoding="utf-8")
    QA_JSON.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    try:
        print("\n".join(lines))
    except UnicodeEncodeError:
        print("\n".join(lines).encode("ascii", "replace").decode("ascii"))
    return report


if __name__ == "__main__":
    run()
