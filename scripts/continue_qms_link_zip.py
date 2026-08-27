"""Continue: fix Excel engine links (handle file lock) + rebuild ZIP."""

from __future__ import annotations

import hashlib
import shutil
import sys
import zipfile
from pathlib import Path

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FINAL = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
ENG_CTRL = FINAL / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ENG_ALT = FINAL / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_QMS_LINK_FIXED.xlsx"
ENG_ROOT = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ZIP_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL.zip"
SHA_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL_SHA256.txt"
BACKUP = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_PRE_QMS_LINK_FIX.xlsx"

NAVY, BLUE, WHITE, GOLD = "0E2A47", "1F4E79", "FFFFFF", "C8A24A"
FONT = "Tahoma"
GOLD_B = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="thin", color=GOLD),
)


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def win_rel(sc: str, fname: str) -> str:
    return f"..\\01_DOCUMENT_SETS\\{sc}\\{fname}"


def main() -> None:
    src = ENG_CTRL if ENG_CTRL.exists() else ENG_ALT
    try:
        shutil.copy2(src, BACKUP)
    except Exception as e:
        print("backup warn", e, flush=True)

    wb = load_workbook(src)
    if "00_HOME" in wb.sheetnames:
        home = wb["00_HOME"]
        home["B27"] = (
            "QMS Document Type Nos:  "
            "Technical File YS/D/0020  •  EU DoC YS/D/0021  •  "
            "Label YS/D/0022  •  Shipment Statement YS/D/0023"
        )
        home["B27"].font = Font(name=FONT, size=9, bold=True, color=NAVY)
        home.merge_cells("B27:I27")
        home["B28"] = (
            "IMPORTANT: Open this Document Engine ONLY from the 00_CONTROL folder "
            "inside the customer delivery package so document links resolve correctly."
        )
        home["B28"].font = Font(name=FONT, size=9, bold=True, color="A12622")
        home.merge_cells("B28:I28")

    ws = wb["SEARCH"]
    ws["Z2"] = '=IFERROR(LEFT(CELL("filename",$B$4),FIND("[",CELL("filename",$B$4))-1),"")'
    ws.column_dimensions["Z"].hidden = True
    issued = (
        'AND($B$8<>"",$A$8<>"NOT FOUND",'
        'ISNUMBER(SEARCH("ISSUED",$G$8)),'
        'NOT(ISNUMBER(SEARCH("NOT ISSUED",$G$8))),'
        'NOT(ISNUMBER(SEARCH("YURT",$G$8))))'
    )
    domestic = 'OR(ISNUMBER(SEARCH("YURT",$G$8)),ISNUMBER(SEARCH("NOT ISSUED",$G$8)))'

    def action_formula(stem: str, is_word: bool) -> str:
        label = "OPEN WORD" if is_word else "OPEN PDF"
        ext = "docx" if is_word else "pdf"
        path_expr = f'$Z$2&"..\\01_DOCUMENT_SETS\\"&$B$8&"\\{stem}.{ext}"'
        return (
            f'=IF(OR($B$4="",$A$8="",$A$8="NOT FOUND"),"",'
            f'IF({domestic},"DOCUMENTS NOT ISSUED",'
            f'IF({issued},HYPERLINK({path_expr},"{label}"),"")))'
        )

    for r, stem in [
        (13, "01_Technical_File"),
        (15, "02_EU_DoC"),
        (17, "03_Label"),
        (19, "04_Shipment_Statement"),
    ]:
        ws.cell(r, 1).value = action_formula(stem, True)
        ws.cell(r, 2).value = action_formula(stem, False)
        for c in (1, 2):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT, size=10, bold=True, color="0563C1")
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            cell.border = GOLD_B
            cell.alignment = Alignment(horizontal="center")

    dc = wb["DOCUMENT_CENTER"]
    dc["E3"] = "TECHNICAL FILE (YS/D/0020)"
    dc["H3"] = "EU DECLARATION OF CONFORMITY (YS/D/0021)"
    dc["K3"] = "LABEL (YS/D/0022)"
    dc["N3"] = "SHIPMENT STATEMENT (YS/D/0023)"
    for col in (5, 8, 11, 14):
        dc.cell(3, col).font = Font(name=FONT, size=9, bold=True, color=WHITE)
        dc.cell(3, col).fill = PatternFill("solid", fgColor=NAVY if col in (5, 11) else BLUE)

    # Always write alt first, then try promote
    wb.save(ENG_ALT)
    wb.close()
    print("saved", ENG_ALT, flush=True)

    # COM polish on ALT
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(ENG_ALT.resolve()))
        dc = wb.Worksheets("DOCUMENT_CENTER")
        r = 5
        while dc.Cells(r, 2).Value:
            sc = str(dc.Cells(r, 2).Value).strip()
            mapping = {
                6: win_rel(sc, "01_Technical_File.docx"),
                7: win_rel(sc, "01_Technical_File.pdf"),
                9: win_rel(sc, "02_EU_DoC.docx"),
                10: win_rel(sc, "02_EU_DoC.pdf"),
                12: win_rel(sc, "03_Label.docx"),
                13: win_rel(sc, "03_Label.pdf"),
                15: win_rel(sc, "04_Shipment_Statement.docx"),
                16: win_rel(sc, "04_Shipment_Statement.pdf"),
            }
            for col, addr in mapping.items():
                cell = dc.Cells(r, col)
                try:
                    if cell.Hyperlinks.Count >= 1:
                        cell.Hyperlinks.Delete()
                except Exception:
                    pass
                label = "OPEN WORD" if addr.endswith(".docx") else "OPEN PDF"
                cell.Value = label
                dc.Hyperlinks.Add(Anchor=cell, Address=addr, TextToDisplay=label)
            r += 1
            if r > 400:
                break

        navy = 0x0E + 256 * 0x2A + 65536 * 0x47
        gold = 0xC8 + 256 * 0xA2 + 65536 * 0x4A
        for i in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(i)
            try:
                for s in list(ws.Shapes):
                    if str(getattr(s, "Name", "")).startswith("HOME_NAV"):
                        s.Delete()
            except Exception:
                pass
            if ws.Name == "00_HOME":
                continue
            try:
                left = float(ws.Cells(1, 18).Left) if ws.Name == "DOCUMENT_CENTER" else float(ws.Cells(1, 12).Left)
                shp = ws.Shapes.AddShape(5, left, 2, 90, 18)
                shp.Name = "HOME_NAV"
                shp.Fill.ForeColor.RGB = navy
                shp.Line.ForeColor.RGB = gold
                shp.TextFrame.Characters().Text = "⌂ HOME"
                shp.TextFrame.Characters().Font.Color = 0xFFFFFF
                shp.TextFrame.Characters().Font.Size = 9
                shp.TextFrame.Characters().Font.Bold = True
                shp.TextFrame.HorizontalAlignment = 2
                ws.Hyperlinks.Add(Anchor=shp, Address="", SubAddress="'00_HOME'!A1")
            except Exception:
                pass
        wb.Save()
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    # Promote ALT -> CTRL if possible
    promoted = False
    try:
        shutil.copy2(ENG_ALT, ENG_CTRL)
        promoted = True
        print("promoted to CTRL engine", flush=True)
    except PermissionError:
        print("CTRL still locked — keep using *_QMS_LINK_FIXED.xlsx until Excel is closed", flush=True)

    shutil.copy2(ENG_ALT, ENG_ROOT)
    print("copied to output root engine", flush=True)

    # ZIP from FINAL (includes ALT file; if CTRL updated, both)
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    print("Creating ZIP…", flush=True)
    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in FINAL.rglob("*"):
            if p.is_file() and not p.name.startswith("~$"):
                zf.write(p, p.relative_to(FINAL.parent).as_posix())
    digest = sha256_file(ZIP_PATH)
    SHA_PATH.write_text(digest + "\n", encoding="utf-8")

    # Quick follow test on ALT
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(str(ENG_ALT.resolve()))
    dc = wb.Worksheets("DOCUMENT_CENTER")
    addr = dc.Cells(5, 6).Hyperlinks(1).Address
    try:
        wb.FollowHyperlink(Address=addr)
        follow = True
    except Exception as e:
        follow = False
        print("follow err", e, flush=True)
    s = wb.Worksheets("SEARCH")
    s.Range("B4").NumberFormat = "@"
    s.Range("B4").Value = "1015169"
    excel.CalculateFull()
    print("SEARCH", s.Range("A8").Value, s.Range("B8").Value, s.Range("A13").Value, flush=True)
    wb.Close(False)
    excel.Quit()
    pythoncom.CoUninitialize()

    print(
        {
            "promoted_ctrl": promoted,
            "engine": str(ENG_ALT if not promoted else ENG_CTRL),
            "dc_follow": follow,
            "zip_sha": digest,
            "qms": "YS/D/0020-0023 stamped on Word titles+footers",
        },
        flush=True,
    )


if __name__ == "__main__":
    main()
