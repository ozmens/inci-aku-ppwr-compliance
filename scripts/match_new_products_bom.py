"""Match new Desktop Excel products to existing Packaging Sets; report gaps."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "output" / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
NEW_XLSX = Path(r"c:\Users\burcu\Desktop\Yeni Microsoft Excel Çalışma Sayfası.xlsx")


def bom_sig(lines: list[tuple[str, float, str]]) -> tuple:
    return tuple(sorted((c, round(q, 6), u) for c, q, u in lines if q))


def parse_new_row(row) -> dict:
    pc = str(row[0]).strip()
    desc = str(row[1] or "").strip()
    btype = str(row[2] or "").strip()
    market = str(row[3] or "").strip()
    nom = float(row[4] or 0)
    comps: list[tuple[str, float, str, str]] = []
    i = 5
    while i + 3 < len(row):
        code, desc_c, qty, uom = row[i], row[i + 1], row[i + 2], row[i + 3]
        i += 4
        if code is None or str(code).strip() in ("", "None"):
            continue
        q = float(qty or 0)
        if q == 0:
            continue
        if isinstance(code, float):
            code_s = str(int(code))
        else:
            code_s = str(code).strip()
        comps.append(
            (
                code_s,
                q,
                str(uom or "").strip().upper(),
                str(desc_c or "").strip(),
            )
        )
    return {
        "product_code": pc,
        "description": desc,
        "battery_type": btype,
        "market": market,
        "nominal_qty": nom,
        "components": comps,
    }


def main() -> None:
    wb = load_workbook(MASTER, data_only=True, read_only=True)
    bm = wb["BOM_MASTER"]
    headers = [c.value for c in next(bm.iter_rows(min_row=1, max_row=1))]
    hi = {h: i for i, h in enumerate(headers)}
    bom_by_set: dict[str, list[tuple[str, float, str]]] = defaultdict(list)
    for row in bm.iter_rows(min_row=2, values_only=True):
        sc = str(row[hi["Packaging Set Code"]] or "").strip()
        code = str(row[hi["Component Code"]] or "").strip()
        if not sc or not code:
            continue
        qty = float(row[hi["Quantity"]] or 0)
        uom = str(row[hi["UOM"]] or "").strip().upper()
        bom_by_set[sc].append((code, qty, uom))

    cm = wb["CONFIG_MASTER"]
    ch = [c.value for c in next(cm.iter_rows(min_row=1, max_row=1))]
    ci = {h: i for i, h in enumerate(ch)}
    controlled = set()
    for row in cm.iter_rows(min_row=2, values_only=True):
        sc = str(row[ci["Packaging Set Code"]] or "").strip()
        st = str(row[ci["Configuration Status"]] or "").strip()
        if sc and st == "CONTROLLED":
            controlled.add(sc)

    sig_to_sets: dict[tuple, list[str]] = defaultdict(list)
    for sc, lines in bom_by_set.items():
        if sc not in controlled:
            continue
        sig_to_sets[bom_sig(lines)].append(sc)

    # also match by (component, qty, uom) ignoring order — already sorted in sig
    nb = load_workbook(NEW_XLSX, data_only=True)["Sayfa1"]
    report = {"exact": [], "new_family": []}
    for row in nb.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row or not row[0]:
            continue
        rec = parse_new_row(row)
        lines = [(c, q, u) for c, q, u, _d in rec["components"]]
        hits = sig_to_sets.get(bom_sig(lines), [])
        item = {
            "product_code": rec["product_code"],
            "description": rec["description"],
            "battery_type": rec["battery_type"],
            "market": rec["market"],
            "nominal_qty": rec["nominal_qty"],
            "component_count": len(lines),
            "matched_sets": hits,
            "components": [
                {"code": c, "qty": q, "uom": u, "desc": d}
                for c, q, u, d in rec["components"]
            ],
        }
        if hits:
            report["exact"].append(item)
        else:
            report["new_family"].append(item)
        print(
            f"{rec['product_code']} nom={rec['nominal_qty']} comps={len(lines)} "
            f"match={hits[:2]} n={len(hits)}"
        )

    out = ROOT / "output" / "_new_products_bom_match.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print("EXACT", len(report["exact"]), "NEW_FAMILY", len(report["new_family"]))
    print("wrote", out)
    wb.close()


if __name__ == "__main__":
    main()
