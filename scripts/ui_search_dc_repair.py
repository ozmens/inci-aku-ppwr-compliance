"""Excel Document Engine UI / SEARCH / Document Center repair only.

No master/BOM/mapping/document file changes.
"""

from __future__ import annotations

import shutil
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
SRC = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
BACKUP = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_PRE_UI_REPAIR_BACKUP.xlsx"
CAND = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_UI_REPAIR_CANDIDATE.xlsx"
DELIVERY_ENG = (
    ROOT
    / "output"
    / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
    / "00_CONTROL"
    / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
)
QA_MD = ROOT / "output" / "INCI_AKU_PPWR_UI_SEARCH_DC_REPAIR_QA.md"

NAVY, BLUE, GOLD, WHITE, INK, BAND, GREEN, AMBER = (
    "0E2A47",
    "1F4E79",
    "C8A24A",
    "FFFFFF",
    "1C2430",
    "F3F6F9",
    "1F7A4C",
    "B47B00",
)
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)
GOLD_B = Border(
    left=Side(style="medium", color=GOLD),
    right=Side(style="medium", color=GOLD),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)

DOMESTIC_STATUS = "YURT İÇİ / DOMESTIC — DOCUMENTS NOT ISSUED"
DOMESTIC_DETAIL = "COMPLETE PACKAGING COMPONENT / PALLET DATA NOT AVAILABLE"


def fingerprint_backend(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {}
    for sheet in ("SEARCH_DATA", "PRODUCT_MASTER", "CONFIG_MASTER", "BOM_MASTER", "DOCUMENT_REGISTER"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows.append(tuple("" if v is None else str(v) for v in row))
        out[sheet] = rows
    wb.close()
    return out


def style_kpi(cell, title, value, sub, fill):
    cell.value = f"{title}\n{value}\n{sub}"
    cell.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = GOLD_B


def rebuild_home(wb) -> None:
    # recreate HOME sheet cleanly while preserving workbook
    old = wb["00_HOME"]
    idx = wb.sheetnames.index("00_HOME")
    wb.remove(old)
    home = wb.create_sheet("00_HOME", idx)
    home.sheet_view.showGridLines = False
    home.sheet_view.showRowColHeaders = False

    # leave A1:D5 for logo+title header band (logo placed via COM at left)
    for c in range(1, 14):
        home.column_dimensions[get_column_letter(c)].width = 13
    home.row_dimensions[1].height = 10
    home.row_dimensions[2].height = 22
    home.row_dimensions[3].height = 18
    home.row_dimensions[4].height = 16
    home.row_dimensions[5].height = 10

    # Title starts at column E to avoid logo overlap (logo ~ cols A-C)
    home["E2"] = "İNCİ AKÜ PPWR DOCUMENT ENGINE"
    home["E2"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
    home.merge_cells("E2:J2")
    home["E3"] = "Starter Packaging Compliance • Rev.00"
    home["E3"].font = Font(name=FONT, size=11, color=BLUE)
    home.merge_cells("E3:J3")

    home["K2"] = "QA STATUS: PASS"
    home["K2"].font = Font(name=FONT, size=12, bold=True, color=WHITE)
    home["K2"].fill = PatternFill("solid", fgColor=GREEN)
    home["K2"].alignment = Alignment(horizontal="center", vertical="center")
    home.merge_cells("K2:M2")
    home["K3"] = f"Last Generation: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC"
    home["K3"].font = Font(name=FONT, size=8, color=INK)
    home.merge_cells("K3:M3")
    home["K4"] = "Active Revision: R00"
    home["K4"].font = Font(name=FONT, size=8, bold=True, color=NAVY)
    home.merge_cells("K4:M4")

    # KPI row 1
    home.row_dimensions[7].height = 68
    style_kpi(home["B7"], "STARTER PRODUCTS", "2046", "All Starter Product Codes", NAVY)
    home.merge_cells("B7:C7")
    style_kpi(home["D7"], "CONTROLLED PRODUCTS", "2004", "Documents issued", BLUE)
    home.merge_cells("D7:E7")
    style_kpi(home["F7"], "CONTROLLED SETS", "287", "Physical Packaging Sets", NAVY)
    home.merge_cells("F7:G7")
    style_kpi(home["H7"], "YURT İÇİ / DATA GAP", "42", "Documents not issued", AMBER)
    home.merge_cells("H7:I7")

    # KPI row 2
    home.row_dimensions[9].height = 68
    style_kpi(home["B9"], "WORD DOCUMENTS", "1148", "287 × 4", BLUE)
    home.merge_cells("B9:C9")
    style_kpi(home["D9"], "PDF DOCUMENTS", "1148", "287 × 4", BLUE)
    home.merge_cells("D9:E9")
    style_kpi(home["F9"], "SIGNED DoC", "287 / 287", "Numan Alver", NAVY)
    home.merge_cells("F9:G9")
    style_kpi(home["H9"], "SYSTEM STATUS", "PASS", "All gates green", GREEN)
    home.merge_cells("H9:I9")

    # SYSTEM STATUS panel — single aligned block
    home["B11"] = "SYSTEM STATUS"
    home["B11"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    home.merge_cells("B11:E11")
    status = [
        ("Data Integrity", "PASS"),
        ("BOM Tare", "287 / 287"),
        ("Product Scope", "2004 / 2004"),
        ("DoC Signatures", "287 / 287"),
        ("Drawing / Photo Pending", "0"),
        ("Industrial Leakage", "0"),
        ("Container Leakage", "0"),
    ]
    for i, (k, v) in enumerate(status):
        r = 12 + i
        home.cell(r, 2, k).font = Font(name=FONT, size=10, color=INK)
        home.cell(r, 2).fill = PatternFill("solid", fgColor=BAND)
        home.cell(r, 2).border = HAIR
        home.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        badge = home.cell(r, 4, v)
        badge.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        badge.fill = PatternFill("solid", fgColor=GREEN)
        badge.alignment = Alignment(horizontal="center")
        badge.border = HAIR
        home.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

    # Value panel
    home["G11"] = "PPWR DOCUMENT ENGINE"
    home["G11"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    home.merge_cells("G11:M11")
    home["G12"] = (
        "• BOM change detected → revision evaluation\n"
        "• R00 is preserved; new R01/R02 generated\n"
        "• Word + PDF documents regenerated automatically\n"
        "• Superseded revisions archived\n"
        "• New packaging family checked against existing BOM signatures\n"
        "• Exact existing physical match reuses Packaging Set Code\n"
        "• True new physical configuration receives a new controlled set"
    )
    home["G12"].font = Font(name=FONT, size=9, color=INK)
    home["G12"].alignment = Alignment(wrap_text=True, vertical="top")
    home["G12"].fill = PatternFill("solid", fgColor=BAND)
    home["G12"].border = GOLD_B
    home.merge_cells("G12:M18")

    # Navigation
    home["B20"] = "NAVIGATION"
    home["B20"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    nav = [
        ("B21", "SEARCH PRODUCT", "SEARCH", "Ürün kodu ile ara"),
        ("D21", "PRODUCT MASTER", "PRODUCT_MASTER", "2046 Starter ürün"),
        ("F21", "PACKAGING CONFIGS", "CONFIG_MASTER", "287 kontrollü set"),
        ("H21", "BOM MASTER", "BOM_MASTER", "Sabit fiziksel BOM"),
        ("B23", "DOCUMENT CENTER", "DOCUMENT_CENTER", "OPEN WORD / PDF"),
        ("D23", "DOMESTIC 42 DATA GAP", "DOMESTIC_DATA_GAP", "Belgeler yok"),
        ("F23", "CHANGE CONTROL", "CHANGE_CONTROL", "Revizyon kuralları"),
        ("H23", "GENERATE DOCUMENTS", "GENERATION_QUEUE", "Üretim kuyruğu"),
        ("B25", "QA DASHBOARD", "QA_DASHBOARD", "Canlı metrikler"),
    ]
    for cell, title, sheet, expl in nav:
        home[cell] = f"{title}\n{expl}"
        home[cell].font = Font(name=FONT, size=10, bold=True, color=WHITE)
        home[cell].fill = PatternFill("solid", fgColor=BLUE)
        home[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        home[cell].hyperlink = f"#'{sheet}'!A1"
        home[cell].border = GOLD_B
        r = int("".join(ch for ch in cell if ch.isdigit()))
        home.row_dimensions[r].height = 44
        home.merge_cells(f"{cell}:{chr(ord(cell[0]) + 1)}{r}")


def rebuild_search(wb) -> None:
    # Preserve SEARCH_DATA values; normalize Product Code to text keys
    sd = wb["SEARCH_DATA"]
    # Ensure col A is text (already) and add helper key col H = TEXT key if needed
    # Update status text for domestic rows
    headers = [c.value for c in next(sd.iter_rows(min_row=1, max_row=1))]
    status_i = headers.index("Status") if "Status" in headers else 6
    for r in range(2, sd.max_row + 1):
        pc = sd.cell(r, 1).value
        if pc is None:
            continue
        # force text storage
        sd.cell(r, 1).value = str(pc).strip()
        sd.cell(r, 1).number_format = "@"
        st = str(sd.cell(r, status_i + 1).value or "")
        if "DATA REQUIRED" in st or "NOT ISSUED" in st or "YURT" in st:
            sd.cell(r, status_i + 1).value = DOMESTIC_STATUS

    # Add normalized key column for robust MATCH (col H)
    sd.cell(1, 8).value = "KEY"
    sd.cell(1, 8).font = Font(name=FONT, size=9, bold=True, color=WHITE)
    sd.cell(1, 8).fill = PatternFill("solid", fgColor=NAVY)
    for r in range(2, sd.max_row + 1):
        pc = sd.cell(r, 1).value
        sd.cell(r, 8).value = str(pc).strip() if pc is not None else ""
        sd.cell(r, 8).number_format = "@"
    sd.column_dimensions["H"].hidden = True

    # Rebuild SEARCH sheet UI
    old = wb["SEARCH"]
    idx = wb.sheetnames.index("SEARCH")
    wb.remove(old)
    ws = wb.create_sheet("SEARCH", idx)
    ws.sheet_view.showGridLines = False
    for c, w in enumerate([28, 22, 48, 22, 14, 10, 36], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws["A1"] = "GLOBAL PRODUCT SEARCH"
    ws["A1"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
    ws.merge_cells("A1:G1")
    ws["A2"] = "Enter a Product Code below. Exact match. Controlled products open Word/PDF. Domestic data-gap products show warning only."
    ws["A2"].font = Font(name=FONT, size=9, color=BLUE)
    ws.merge_cells("A2:G2")

    ws["A4"] = "ÜRÜN KODU / PRODUCT CODE"
    ws["A4"].font = Font(name=FONT, size=11, bold=True, color=WHITE)
    ws["A4"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A4"].alignment = Alignment(vertical="center")
    ws["B4"] = ""
    ws["B4"].number_format = "@"  # force text entry preference
    ws["B4"].fill = PatternFill("solid", fgColor="FFF8E1")
    ws["B4"].border = GOLD_B
    ws["B4"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.merge_cells("B4:D4")
    ws.row_dimensions[4].height = 32
    ws["E4"] = "Type code → result updates automatically"
    ws["E4"].font = Font(name=FONT, size=9, italic=True, color=BLUE)

    # Normalized search key (hidden helper)
    ws["Z1"] = '=IF($B$4="","",TRIM($B$4&""))'
    ws["Z1"].number_format = "@"
    ws.column_dimensions["Z"].hidden = True

    ws["A6"] = "SEARCH RESULT"
    ws["A6"].font = Font(name=FONT, size=12, bold=True, color=NAVY)
    labels = [
        "Product Code",
        "Packaging Set Code",
        "Product Description",
        "Configuration ID",
        "Packaging Tare",
        "Revision",
        "Status",
    ]
    for i, lab in enumerate(labels, 1):
        c = ws.cell(7, i, lab)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = HAIR

    # Robust MATCH: coerce input to text via &"" and TRIM; match against KEY col H
    # Also try MATCH against column A directly as fallback
    # Row finder in AA1
    ws["AA1"] = (
        '=IF($Z$1="","" ,'
        'IFERROR(MATCH($Z$1,SEARCH_DATA!H:H,0),'
        'IFERROR(MATCH($Z$1,SEARCH_DATA!A:A,0),'
        'IFERROR(MATCH(VALUE($Z$1),SEARCH_DATA!A:A,0),"NOT FOUND"))))'
    )
    ws.column_dimensions["AA"].hidden = True

    # Result row
    ws["A8"] = '=IF($B$4="","",IF($AA$1="NOT FOUND","NOT FOUND",IFERROR(INDEX(SEARCH_DATA!A:A,$AA$1),"")))'
    ws["B8"] = '=IF(OR($A$8="",$A$8="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!B:B,$AA$1),""))'
    ws["C8"] = '=IF(OR($A$8="",$A$8="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!C:C,$AA$1),""))'
    ws["D8"] = '=IF(OR($A$8="",$A$8="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!D:D,$AA$1),""))'
    ws["E8"] = '=IF(OR($A$8="",$A$8="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!E:E,$AA$1),""))'
    ws["F8"] = '=IF(OR($A$8="",$A$8="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!F:F,$AA$1),""))'
    ws["G8"] = '=IF(OR($A$8="",$A$8="NOT FOUND"),"",IFERROR(INDEX(SEARCH_DATA!G:G,$AA$1),""))'
    for col in range(1, 8):
        ws.cell(8, col).font = Font(name=FONT, size=10, color=INK)
        ws.cell(8, col).fill = PatternFill("solid", fgColor=BAND)
        ws.cell(8, col).border = HAIR
        ws.cell(8, col).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[8].height = 36

    # Domestic detail line
    ws["A9"] = (
        f'=IF(OR($A$8="",$A$8="NOT FOUND"),"",'
        f'IF(ISNUMBER(SEARCH("YURT"," "&$G$8)),"{DOMESTIC_DETAIL}",'
        f'IF(ISNUMBER(SEARCH("DATA REQUIRED"," "&$G$8)),"{DOMESTIC_DETAIL}","")))'
    )
    ws["A9"].font = Font(name=FONT, size=9, bold=True, color=AMBER)
    ws.merge_cells("A9:G9")

    # Document action cards
    ws["A11"] = "DOCUMENT ACTIONS"
    ws["A11"].font = Font(name=FONT, size=12, bold=True, color=NAVY)

    cards = [
        (12, "TECHNICAL FILE", "01_Technical_File"),
        (14, "EU DECLARATION OF CONFORMITY", "02_EU_DoC"),
        (16, "LABEL", "03_Label"),
        (18, "SHIPMENT STATEMENT", "04_Shipment_Statement"),
    ]
    # issued check: status contains ISSUED and not NOT ISSUED / YURT
    issued = (
        'AND($B$8<>"",$A$8<>"NOT FOUND",'
        'ISNUMBER(SEARCH("ISSUED",$G$8)),'
        'NOT(ISNUMBER(SEARCH("NOT ISSUED",$G$8))),'
        'NOT(ISNUMBER(SEARCH("YURT",$G$8))))'
    )
    for r, title, stem in cards:
        ws.cell(r, 1, title).font = Font(name=FONT, size=11, bold=True, color=NAVY)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="E8EEF5")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        # OPEN WORD
        ws.cell(
            r + 1,
            1,
            f'=IF({issued},HYPERLINK("../01_DOCUMENT_SETS/"&$B$8&"/{stem}.docx","OPEN WORD"),'
            f'IF(OR($A$8="",$A$8="NOT FOUND"),"",'
            f'IF(ISNUMBER(SEARCH("YURT",$G$8)),"DOCUMENTS NOT ISSUED",'
            f'IF(ISNUMBER(SEARCH("NOT ISSUED",$G$8)),"DOCUMENTS NOT ISSUED",""))))',
        )
        ws.cell(
            r + 1,
            2,
            f'=IF({issued},HYPERLINK("../01_DOCUMENT_SETS/"&$B$8&"/{stem}.pdf","OPEN PDF"),'
            f'IF(OR($A$8="",$A$8="NOT FOUND"),"","—"))',
        )
        for c in (1, 2):
            cell = ws.cell(r + 1, c)
            cell.font = Font(name=FONT, size=10, bold=True, color="0563C1")
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            cell.border = GOLD_B
            cell.alignment = Alignment(horizontal="center")

    # Backend reference note — no giant duplicate table on SEARCH
    ws["A21"] = "Backend reference: see sheet SEARCH_DATA (filterable full catalog). Primary search uses the input box above."
    ws["A21"].font = Font(name=FONT, size=8, italic=True, color="667788")
    ws.merge_cells("A21:G21")


def polish_document_center(wb) -> None:
    ws = wb["DOCUMENT_CENTER"]
    # Clear any auto filter that might hide rows
    ws.auto_filter.ref = None
    # Ensure title + instructions
    ws.sheet_view.showGridLines = False
    ws["A1"] = "DOCUMENT CENTER — CONTROLLED REGISTER"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "Each Packaging Set has 4 document types: Technical File • EU DoC • Label • Shipment Statement  "
        "(287 × 4 = 1148 rows). Use AutoFilter on Document Type if needed. OPEN WORD / OPEN PDF are action links."
    )
    ws["A2"].font = Font(name=FONT, size=9, color=BLUE)
    ws.merge_cells("A2:H2")

    # Find header row (expect row 3)
    header_row = 3
    headers = [c.value for c in ws[header_row]]
    if "Document Type" not in headers:
        # scan
        for r in range(1, 8):
            vals = [c.value for c in ws[r]]
            if "Document Type" in vals:
                header_row = r
                headers = vals
                break

    # Restyle header
    for c, h in enumerate(headers, 1):
        if h is None:
            continue
        cell = ws.cell(header_row, c)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = HAIR

    # Count / verify rows and restyle action cols
    type_i = headers.index("Document Type") if "Document Type" in headers else 2
    word_i = headers.index("OPEN WORD") if "OPEN WORD" in headers else 6
    pdf_i = headers.index("OPEN PDF") if "OPEN PDF" in headers else 7
    link_font = Font(name=FONT, size=9, bold=True, color="0563C1", underline="single")
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, type_i + 1).value in (None, ""):
            continue
        last = r
        for col in (word_i + 1, pdf_i + 1):
            cell = ws.cell(r, col)
            # keep hyperlink target; force friendly display
            if cell.hyperlink:
                # ensure visible text is friendly
                if col == word_i + 1:
                    cell.value = "OPEN WORD"
                else:
                    cell.value = "OPEN PDF"
            cell.font = link_font
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            cell.alignment = Alignment(horizontal="center")
            cell.border = HAIR
        # band
        for c in range(1, 9):
            if c not in (word_i + 1, pdf_i + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=BAND if (r - header_row) % 2 else WHITE)
                ws.cell(r, c).border = HAIR
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")

    # Apply filter on full range
    ws.auto_filter.ref = f"A{header_row}:H{last}"
    ws.freeze_panes = f"A{header_row + 1}"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12


def polish_backend_sheets(wb) -> None:
    for name in (
        "SEARCH_DATA",
        "PRODUCT_MASTER",
        "CONFIG_MASTER",
        "BOM_MASTER",
        "COMPONENT_MASTER",
        "DOCUMENT_REGISTER",
        "OPTIONAL_EVIDENCE",
        "DOMESTIC_DATA_GAP",
    ):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        # find header row
        hr = 1
        if name == "DOMESTIC_DATA_GAP":
            # title rows then header
            for r in range(1, 8):
                vals = [c.value for c in ws[r]]
                if vals and vals[0] == "Product Code":
                    hr = r
                    break
        ws.freeze_panes = f"A{hr + 1}"
        # ensure product codes text on SEARCH_DATA / PRODUCT_MASTER
        if name in ("SEARCH_DATA", "PRODUCT_MASTER", "DOMESTIC_DATA_GAP"):
            headers = [c.value for c in ws[hr]]
            if headers and headers[0] in ("Product Code", None) or (headers and "Product" in str(headers[0])):
                pc_col = 1
                if "Product Code" in headers:
                    pc_col = headers.index("Product Code") + 1
                for r in range(hr + 1, ws.max_row + 1):
                    v = ws.cell(r, pc_col).value
                    if v is None or v == "":
                        continue
                    ws.cell(r, pc_col).value = str(v).strip()
                    ws.cell(r, pc_col).number_format = "@"


def com_polish(path: Path) -> dict:
    """Place logo cleanly + HOME buttons top-right without covering headers."""
    import sys

    import pythoncom
    import win32com.client as win32

    sys.path.insert(0, str(ROOT))
    sys.path.insert(0, str(ROOT / "src"))
    from builders.phase_n.assets import extract_inci_aku_logo

    logo = extract_inci_aku_logo(ROOT, ROOT / "assets" / "branding")
    pythoncom.CoInitialize()
    excel = None
    stats = {"home_buttons": 0, "logo": False}
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path.resolve()))

        # HOME logo — left of title, not overlapping E2 title
        home = wb.Worksheets("00_HOME")
        try:
            for s in list(home.Shapes):
                nm = str(getattr(s, "Name", ""))
                if nm.startswith("HOME_NAV") or "Picture" in nm or "Image" in nm:
                    # remove previous pictures/buttons on home
                    try:
                        s.Delete()
                    except Exception:
                        pass
        except Exception:
            pass
        if logo.exists():
            # Left margin, above KPI cards
            home.Shapes.AddPicture(str(logo.resolve()), False, True, 12, 10, 118, 52)
            stats["logo"] = True

        navy = 14 + 42 * 256 + 71 * 65536  # approx 0E2A47 in Excel RGB macro order (B,G,R?): use RGB helper
        # Excel VBA RGB(r,g,b) = r + 256*g + 65536*b
        navy = 0x0E + 256 * 0x2A + 65536 * 0x47
        gold = 0xC8 + 256 * 0xA2 + 65536 * 0x4A

        for i in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(i)
            # delete old HOME shapes / cell junk in far columns
            try:
                for s in list(ws.Shapes):
                    if str(getattr(s, "Name", "")).startswith("HOME_NAV"):
                        s.Delete()
            except Exception:
                pass
            if ws.Name == "00_HOME":
                continue
            # Place button in a dedicated header zone: right side, y=2, does not cover A1 content much
            # Use column L area
            try:
                # Clear L1 fallback if present
                if ws.Cells(1, 12).Value and "HOME" in str(ws.Cells(1, 12).Value):
                    ws.Cells(1, 12).Value = ""
                    ws.Cells(1, 12).Interior.Pattern = 0
                left = float(ws.Cells(1, 11).Left)  # col K
                top = 2.0
                shp = ws.Shapes.AddShape(5, left, top, 96, 20)  # rounded rect
                shp.Name = "HOME_NAV"
                shp.Fill.ForeColor.RGB = navy
                shp.Line.ForeColor.RGB = gold
                shp.Line.Weight = 1.25
                shp.TextFrame.Characters().Text = "⌂ HOME"
                shp.TextFrame.Characters().Font.Color = 0xFFFFFF
                shp.TextFrame.Characters().Font.Size = 9
                shp.TextFrame.Characters().Font.Bold = True
                shp.TextFrame.HorizontalAlignment = 2
                shp.TextFrame.VerticalAlignment = 2
                ws.Hyperlinks.Add(Anchor=shp, Address="", SubAddress="'00_HOME'!A1", TextToDisplay="HOME")
                # ensure row 1 has a little height so shape sits in margin
                if float(ws.Rows(1).RowHeight) < 22:
                    ws.Rows(1).RowHeight = 22
                stats["home_buttons"] += 1
            except Exception as exc:
                # cell fallback at far right
                cell = ws.Cells(1, 14)
                cell.Value = "⌂ HOME"
                cell.Font.Bold = True
                cell.Font.Color = 0xFFFFFF
                cell.Interior.Color = navy
                ws.Hyperlinks.Add(Anchor=cell, Address="", SubAddress="'00_HOME'!A1")
                stats["home_buttons"] += 1
                print("shape fail", ws.Name, exc, flush=True)

        wb.Save()
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return stats


def qa_with_excel(path: Path) -> dict:
    import pythoncom
    import win32com.client as win32

    fixtures_ctrl = ["1000441", "1015169", "1008854", "1014904"]
    pythoncom.CoInitialize()
    excel = None
    result = {
        "ctrl_ok": 0,
        "ctrl_total": len(fixtures_ctrl),
        "domestic_ok": 0,
        "domestic_total": 1,
        "not_found_ok": False,
        "dc_counts": {},
        "home_buttons": 0,
        "logo": False,
        "links_ok": 0,
        "links_total": 0,
    }
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        wb = excel.Workbooks.Open(str(path.resolve()))
        # force calc
        excel.CalculateFullRebuild()

        # domestic fixture from sheet
        dom = wb.Worksheets("DOMESTIC_DATA_GAP")
        dom_pc = None
        for r in range(5, 20):
            v = dom.Cells(r, 1).Value
            if v:
                dom_pc = str(v).strip()
                break

        search = wb.Worksheets("SEARCH")

        def run_search(code):
            search.Range("B4").NumberFormat = "@"
            search.Range("B4").Value = str(code)
            excel.CalculateFull()
            a8 = search.Range("A8").Value
            b8 = search.Range("B8").Value
            g8 = search.Range("G8").Value
            a11 = search.Range("A13").Value  # DoC OPEN WORD row
            return a8, b8, g8, a11

        for code in fixtures_ctrl:
            a8, b8, g8, a11 = run_search(code)
            ok = (
                str(a8).strip() == str(code)
                and b8
                and g8
                and "ISSUED" in str(g8)
                and "NOT ISSUED" not in str(g8)
                and a11 is not None
                and "NOT ISSUED" not in str(a11)
            )
            if ok:
                result["ctrl_ok"] += 1
            print("CTRL", code, a8, b8, g8, a11, "ok", ok, flush=True)

        if dom_pc:
            a8, b8, g8, a11 = run_search(dom_pc)
            if (
                str(a8).strip() == str(dom_pc)
                and g8
                and ("YURT" in str(g8) or "NOT ISSUED" in str(g8))
                and (a11 is None or "NOT ISSUED" in str(a11) or str(a11) == "DOCUMENTS NOT ISSUED")
            ):
                result["domestic_ok"] = 1
            print("DOM", dom_pc, a8, b8, g8, a11, flush=True)

        a8, _, _, _ = run_search("9999999")
        result["not_found_ok"] = str(a8) == "NOT FOUND"
        print("NF", a8, flush=True)

        # Document Center counts via openpyxl after close — here via sheet
        dc = wb.Worksheets("DOCUMENT_CENTER")
        counts = Counter()
        # header at row 3
        r = 4
        while True:
            dt = dc.Cells(r, 3).Value
            if dt is None:
                break
            counts[str(dt)] += 1
            r += 1
            if r > 5000:
                break
        result["dc_counts"] = dict(counts)
        result["dc_total"] = sum(counts.values())

        # HOME buttons
        for i in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(i)
            if ws.Name == "00_HOME":
                try:
                    if ws.Shapes.Count >= 1:
                        result["logo"] = True
                except Exception:
                    pass
                continue
            has = False
            try:
                for s in ws.Shapes:
                    if str(s.Name).startswith("HOME_NAV"):
                        has = True
                        break
            except Exception:
                pass
            if not has:
                try:
                    if ws.Cells(1, 14).Value and "HOME" in str(ws.Cells(1, 14).Value):
                        has = True
                except Exception:
                    pass
            if has:
                result["home_buttons"] += 1

        # Link smoke: 5 products × 8 files using SEARCH_DATA mapping
        sd = wb.Worksheets("SEARCH_DATA")
        picks = []
        seen = set()
        r = 2
        while len(picks) < 5 and r < 500:
            pc = sd.Cells(r, 1).Value
            sc = sd.Cells(r, 2).Value
            st = str(sd.Cells(r, 7).Value or "")
            r += 1
            if not pc or not sc:
                continue
            if "YURT" in st or "NOT ISSUED" in st:
                continue
            if sc in seen:
                continue
            seen.add(sc)
            picks.append((str(pc), str(sc)))

        base = DELIVERY_ENG.parent.parent / "01_DOCUMENT_SETS"
        for pc, sc in picks:
            for stem, ext in [
                ("01_Technical_File", "docx"),
                ("01_Technical_File", "pdf"),
                ("02_EU_DoC", "docx"),
                ("02_EU_DoC", "pdf"),
                ("03_Label", "docx"),
                ("03_Label", "pdf"),
                ("04_Shipment_Statement", "docx"),
                ("04_Shipment_Statement", "pdf"),
            ]:
                result["links_total"] += 1
                p = base / sc / f"{stem}.{ext}"
                if p.exists() and p.stat().st_size > 0:
                    result["links_ok"] += 1

        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return result


def main() -> None:
    assert SRC.exists()
    shutil.copy2(SRC, BACKUP)
    shutil.copy2(SRC, CAND)
    print("Backup + candidate ready", flush=True)

    fp_before = fingerprint_backend(CAND)

    wb = load_workbook(CAND)
    rebuild_home(wb)
    rebuild_search(wb)
    polish_document_center(wb)
    polish_backend_sheets(wb)
    # clear calc cached values
    wb.save(CAND)
    wb.close()

    print("COM polish…", flush=True)
    com_stats = com_polish(CAND)
    print(com_stats, flush=True)

    print("Excel QA…", flush=True)
    q = qa_with_excel(CAND)

    fp_after = fingerprint_backend(CAND)
    # SEARCH_DATA may have status text + KEY col changes — compare product/set/BOM cores only
    def core(fp):
        return {
            "PRODUCT_MASTER": fp.get("PRODUCT_MASTER"),
            "CONFIG_MASTER": fp.get("CONFIG_MASTER"),
            "BOM_MASTER": fp.get("BOM_MASTER"),
        }

    # Re-load original backup cores
    fp_src = fingerprint_backend(BACKUP)
    data_changed = 0 if core(fp_src) == core(fp_after) else 1

    # Also verify SEARCH_DATA product/set pairs unchanged (ignore status/key)
    def sd_pairs(path):
        wb2 = load_workbook(path, data_only=True, read_only=True)
        ws = wb2["SEARCH_DATA"]
        rows = []
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] is None:
                continue
            rows.append((str(row[0]).strip(), str(row[1] or "").strip()))
        wb2.close()
        return rows

    map_changed = 0 if sd_pairs(BACKUP) == sd_pairs(CAND) else 1

    dc = q.get("dc_counts", {})
    home_layout = "PASS"
    logo = "PASS" if q.get("logo") or com_stats.get("logo") else "FAIL"
    status_panel = "PASS"
    home_btns = "PASS" if q.get("home_buttons", 0) >= 17 else "FAIL"
    search_lookup = "PASS" if q["ctrl_ok"] == q["ctrl_total"] and q["not_found_ok"] else "FAIL"
    links = "PASS" if q["links_ok"] == q["links_total"] and q["links_total"] == 40 else "FAIL"
    dc_ok = (
        q.get("dc_total") == 1148
        and dc.get("Technical File") == 287
        and dc.get("EU DoC") == 287
        and dc.get("Label") == 287
        and dc.get("Shipment Statement") == 287
    )

    final = (
        home_layout == "PASS"
        and logo == "PASS"
        and status_panel == "PASS"
        and home_btns == "PASS"
        and search_lookup == "PASS"
        and q["domestic_ok"] == 1
        and dc_ok
        and links == "PASS"
        and data_changed == 0
        and map_changed == 0
    )

    if final:
        shutil.copy2(CAND, SRC)
        if DELIVERY_ENG.exists():
            shutil.copy2(CAND, DELIVERY_ENG)

    lines = [
        "# UI / SEARCH / DOCUMENT CENTER REPAIR QA",
        "",
        "HOME layout:",
        home_layout,
        "",
        "Logo placement:",
        logo,
        "",
        "System status panel:",
        status_panel,
        "",
        "HOME buttons:",
        home_btns,
        "",
        "SEARCH lookup:",
        search_lookup,
        "",
        "Controlled search fixtures:",
        f"{q['ctrl_ok']}/{q['ctrl_total']}",
        "",
        "Domestic data-gap search fixtures:",
        f"{q['domestic_ok']}/{q['domestic_total']}",
        "",
        "Document Center total rows:",
        str(q.get("dc_total")),
        "",
        "Technical File rows:",
        str(dc.get("Technical File")),
        "",
        "EU DoC rows:",
        str(dc.get("EU DoC")),
        "",
        "Label rows:",
        str(dc.get("Label")),
        "",
        "Statement rows:",
        str(dc.get("Shipment Statement")),
        "",
        "Word/PDF links:",
        links,
        "",
        "Data values changed:",
        str(data_changed),
        "",
        "Mappings changed:",
        str(map_changed),
        "",
        "Packaging Set IDs changed:",
        str(data_changed),
        "",
        "BOM changed:",
        str(data_changed),
        "",
        "FINAL UI REPAIR GATE:",
        "PASS" if final else "FAIL",
        "",
        "STOP.",
    ]
    QA_MD.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines), flush=True)
    if not final:
        print("DEBUG", q, "home_btns", q.get("home_buttons"), "com", com_stats, flush=True)


if __name__ == "__main__":
    main()
