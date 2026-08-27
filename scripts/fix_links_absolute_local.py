"""Make Document Center / SEARCH links open reliably on this machine.

Uses absolute paths to FINAL\\01_DOCUMENT_SETS\\...
Removes the trap engine copy in output\\ root.
"""

from __future__ import annotations

import hashlib
import shutil
import zipfile
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
FINAL = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
ENG = FINAL / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
DOC = FINAL / "01_DOCUMENT_SETS"
ROOT_ENG = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
TRAP = ROOT / "output" / "!!!_ACMA_BU_ENGINE_LINKLERI_KIRIK.xlsx"
ZIP_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL.zip"
SHA_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL_SHA256.txt"

COLS = [
    (6, "01_Technical_File.docx"),
    (7, "01_Technical_File.pdf"),
    (9, "02_EU_DoC.docx"),
    (10, "02_EU_DoC.pdf"),
    (12, "03_Label.docx"),
    (13, "03_Label.pdf"),
    (15, "04_Shipment_Statement.docx"),
    (16, "04_Shipment_Statement.pdf"),
]


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def write_trap_stub() -> None:
    """Replace confusing root engine with a stub that opens the real one."""
    for p in (ROOT_ENG, TRAP):
        if p.exists():
            try:
                p.unlink()
            except Exception:
                pass
    wb = Workbook()
    ws = wb.active
    ws.title = "ACMA"
    ws["A1"] = "YANLIŞ DOSYA — LINKLER BURADA ÇALIŞMAZ"
    ws["A1"].font = Font(name="Tahoma", size=18, bold=True, color="A12622")
    ws.merge_cells("A1:F1")
    ws["A3"] = "Doğru dosyayı açın:"
    ws["A3"].font = Font(name="Tahoma", size=12, bold=True)
    ws["A4"] = str(ENG.resolve())
    ws["A4"].font = Font(name="Tahoma", size=11, color="0563C1", underline="single")
    ws["A4"].hyperlink = str(ENG.resolve())
    ws["A6"] = "veya çalıştırın:"
    ws["A7"] = str((FINAL / "00_AC_DOCUMENT_ENGINE.cmd").resolve())
    ws.column_dimensions["A"].width = 120
    wb.save(TRAP)
    wb.close()
    # also save under old name so habit-open shows warning
    shutil.copy2(TRAP, ROOT_ENG)


def main() -> None:
    assert ENG.exists()
    assert DOC.exists()

    (FINAL / "00_AC_DOCUMENT_ENGINE.cmd").write_text(
        "@echo off\r\n"
        "cd /d \"%~dp0\"\r\n"
        "start \"\" \"%~dp000_CONTROL\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx\"\r\n",
        encoding="utf-8",
    )
    write_trap_stub()

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(ENG.resolve()))

        try:
            home = wb.Worksheets("00_HOME")
            home.Range("B28").Value = (
                "OPEN VIA: 00_CONTROL\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx "
                "or 00_AC_DOCUMENT_ENGINE.cmd — never open output\\ root engine copy."
            )
        except Exception:
            pass

        dc = wb.Worksheets("DOCUMENT_CENTER")
        r = 5
        n = 0
        while dc.Cells(r, 2).Value:
            sc = str(dc.Cells(r, 2).Value).strip()
            for col, fname in COLS:
                cell = dc.Cells(r, col)
                abs_path = str((DOC / sc / fname).resolve())
                label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
                try:
                    if cell.Hyperlinks.Count:
                        cell.Hyperlinks.Delete()
                except Exception:
                    pass
                cell.Value = label
                dc.Hyperlinks.Add(Anchor=cell, Address=abs_path, TextToDisplay=label)
                n += 1
            r += 1
            if r > 400:
                break
        print("DC absolute links:", n)

        # SEARCH — absolute HYPERLINK formulas (outermost)
        search = wb.Worksheets("SEARCH")
        base = str(FINAL.resolve()) + "\\"
        search.Range("Z1").Value = base
        search.Columns("Z").Hidden = True

        issued = (
            'AND($B$8<>"",$A$8<>"NOT FOUND",ISNUMBER(SEARCH("ISSUED",$G$8)),'
            'NOT(ISNUMBER(SEARCH("NOT ISSUED",$G$8))),NOT(ISNUMBER(SEARCH("YURT",$G$8))))'
        )
        domestic = 'OR(ISNUMBER(SEARCH("YURT",$G$8)),ISNUMBER(SEARCH("NOT ISSUED",$G$8)))'
        empty = 'OR($B$4="",$A$8="",$A$8="NOT FOUND")'

        def af(stem: str, word: bool) -> str:
            label = "OPEN WORD" if word else "OPEN PDF"
            ext = "docx" if word else "pdf"
            path = f'$Z$1&"01_DOCUMENT_SETS\\"&$B$8&"\\{stem}.{ext}"'
            return (
                f'=IF({empty},"",'
                f'IF({domestic},"DOCUMENTS NOT ISSUED",'
                f'HYPERLINK(IF({issued},{path},""),IF({issued},"{label}",""))))'
            )

        for row, stem in [
            (13, "01_Technical_File"),
            (15, "02_EU_DoC"),
            (17, "03_Label"),
            (19, "04_Shipment_Statement"),
        ]:
            search.Cells(row, 1).Formula = af(stem, True)
            search.Cells(row, 2).Formula = af(stem, False)

        wb.Save()
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    # Verify 8 opens from DC
    pythoncom.CoInitialize()
    excel = None
    ok = 0
    fail = []
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(ENG.resolve()), ReadOnly=True)
        dc = wb.Worksheets("DOCUMENT_CENTER")
        for col, fname in COLS:
            cell = dc.Cells(5, col)
            addr = cell.Hyperlinks(1).Address
            try:
                wb.FollowHyperlink(Address=addr)
                ok += 1
            except Exception as e:
                fail.append(f"{fname}:{e}")
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    print("verify", ok, "/8", fail)

    # ZIP rebuild
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    with zipfile.ZipFile(ZIP_PATH, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for p in FINAL.rglob("*"):
            if p.is_file() and not p.name.startswith("~$"):
                zf.write(p, p.relative_to(FINAL).as_posix())
    digest = sha256_file(ZIP_PATH)
    SHA_PATH.write_text(digest + "\n", encoding="utf-8")
    print("SHA256", digest)
    print("OPEN:", ENG)
    if fail or ok != 8:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
