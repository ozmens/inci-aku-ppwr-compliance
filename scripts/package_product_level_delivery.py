"""Package product-level delivery like Component Packaging (no Word/Excel UI)."""

from __future__ import annotations

import hashlib
import json
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DELIVERY = (
    ROOT
    / "output"
    / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_CUSTOMER_DELIVERY_REV00_CANDIDATE"
)
CONTROL = DELIVERY / "00_CONTROL"
PRODUCT_SETS = DELIVERY / "01_PRODUCT_DOCUMENT_SETS"
ENGINE_CTRL = CONTROL / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ENGINE_ROOT = ROOT / "output" / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_ENGINE_Rev00.xlsx"
ENGINE_DESKTOP = Path.home() / "Desktop" / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_ENGINE_Rev00.xlsx"
ZIP_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_CUSTOMER_DELIVERY_REV00.zip"
SHA_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_CUSTOMER_DELIVERY_REV00.sha256"
STEMS = [
    "01_Technical_File",
    "02_EU_DoC",
    "03_Label",
    "04_Shipment_Statement",
]


def count_docs() -> dict:
    folders = [p for p in PRODUCT_SETS.iterdir() if p.is_dir()]
    word = pdf = missing = 0
    for folder in folders:
        for stem in STEMS:
            d = folder / f"{stem}.docx"
            p = folder / f"{stem}.pdf"
            if d.exists() and not d.name.startswith("~$"):
                word += 1
            if p.exists() and p.stat().st_size > 0:
                pdf += 1
            else:
                missing += 1
    return {
        "product_folders": len(folders),
        "word": word,
        "pdf": pdf,
        "missing_pdf": missing,
        "expected": len(folders) * 4,
    }


def link_sample_test(sample_rows: int = 25) -> dict:
    wb = load_workbook(ENGINE_CTRL, data_only=False)
    dc = wb["DOCUMENT_CENTER"]
    ok = fail = checked = 0
    fails: list[str] = []
    # data starts row 5; 8 link cols: 7,8,10,11,13,14,16,17
    cols = (7, 8, 10, 11, 13, 14, 16, 17)
    max_row = 4 + sample_rows
    for row in range(5, max_row + 1):
        pc = dc.cell(row, 1).value
        if not pc:
            break
        for col in cols:
            cell = dc.cell(row, col)
            target = cell.hyperlink.target if cell.hyperlink else None
            checked += 1
            if not target:
                fail += 1
                fails.append(f"{pc} col{col}: no hyperlink")
                continue
            resolved = (ENGINE_CTRL.parent / target).resolve()
            if resolved.exists() and resolved.stat().st_size > 0:
                ok += 1
            else:
                fail += 1
                fails.append(f"{pc} -> {target}")
    wb.close()
    return {"checked": checked, "ok": ok, "fail": fail, "fails_sample": fails[:20]}


def make_zip() -> str:
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        for p in DELIVERY.rglob("*"):
            if not p.is_file():
                continue
            if p.name.startswith("~$"):
                continue
            zf.write(p, p.relative_to(DELIVERY.parent).as_posix())
    h = hashlib.sha256()
    with ZIP_PATH.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    digest = h.hexdigest()
    SHA_PATH.write_text(digest + "\n", encoding="utf-8")
    return digest


def main() -> None:
    assert ENGINE_CTRL.exists(), f"missing engine: {ENGINE_CTRL}"
    counts = count_docs()
    print("counts", counts, flush=True)
    links = link_sample_test(40)
    print("links", links, flush=True)

    shutil.copy2(ENGINE_CTRL, ENGINE_ROOT)
    try:
        shutil.copy2(ENGINE_CTRL, ENGINE_DESKTOP)
        desktop_ok = True
    except Exception as exc:  # noqa: BLE001
        desktop_ok = False
        print("desktop_copy_fail", exc, flush=True)

    # Replace dangerous stub pointer note beside legacy tiny engine if present
    stub = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
    note = ROOT / "output" / "!!!_ACMA_BU_ENGINE_LINKLERI_KIRIK.xlsx"
    if stub.exists() and stub.stat().st_size < 50_000:
        # leave stub but ensure warning sibling exists
        if not note.exists():
            note.write_bytes(stub.read_bytes())

    qa = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "delivery": str(DELIVERY),
        "counts": counts,
        "link_sample": links,
        "engine_control": str(ENGINE_CTRL),
        "engine_root": str(ENGINE_ROOT),
        "engine_desktop": str(ENGINE_DESKTOP) if desktop_ok else None,
        "gate": "PASS"
        if counts["missing_pdf"] == 0
        and counts["word"] == counts["expected"]
        and counts["pdf"] == counts["expected"]
        and links["fail"] == 0
        else "FAIL",
    }
    (CONTROL / "PRODUCT_LEVEL_FINAL_QA.json").write_text(
        json.dumps(qa, indent=2), encoding="utf-8"
    )
    txt = (
        "İNCI AKÜ PPWR — PRODUCT-LEVEL CUSTOMER DELIVERY QA\n"
        f"Generated: {qa['generated_at']}\n"
        f"Products: {counts['product_folders']}\n"
        f"Word: {counts['word']} / {counts['expected']}\n"
        f"PDF:  {counts['pdf']} / {counts['expected']}\n"
        f"Missing PDF: {counts['missing_pdf']}\n"
        f"Link sample: {links['ok']}/{links['checked']} OK (fail={links['fail']})\n"
        f"Engine (CONTROL): {ENGINE_CTRL.name}\n"
        f"Engine (output root): {ENGINE_ROOT.name}\n"
        f"Gate: {qa['gate']}\n"
        "Open path: 00_AC_DOCUMENT_ENGINE.cmd\n"
    )
    (CONTROL / "PRODUCT_LEVEL_FINAL_QA.txt").write_text(txt, encoding="utf-8")
    print(txt, flush=True)

    print("Building ZIP (large)…", flush=True)
    digest = make_zip()
    qa["zip"] = str(ZIP_PATH)
    qa["sha256"] = digest
    qa["zip_bytes"] = ZIP_PATH.stat().st_size
    (CONTROL / "PRODUCT_LEVEL_FINAL_QA.json").write_text(
        json.dumps(qa, indent=2), encoding="utf-8"
    )
    with (CONTROL / "PRODUCT_LEVEL_FINAL_QA.txt").open("a", encoding="utf-8") as f:
        f.write(f"ZIP: {ZIP_PATH.name}\nSHA256: {digest}\n")
    print("ZIP", ZIP_PATH, "bytes", ZIP_PATH.stat().st_size, flush=True)
    print("SHA256", digest, flush=True)
    print("DONE", qa["gate"], flush=True)


if __name__ == "__main__":
    main()
