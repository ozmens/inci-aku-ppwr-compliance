"""STARTER MASTER FINAL DATA LOCK — Excel only.

NO Word / PHASE_I / freeze / PIMS changes.
"""

from __future__ import annotations

import hashlib
import json
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
OUT = ROOT / "output"
STARTER = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
BACKUP = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00_PRELOCK_BACKUP.xlsx"
GOLDEN = (
    ROOT
    / "input"
    / "production"
    / "INCI_AKU_PPWR_Final_Configuration_Register_Rev00_GOLDEN_VARIANTS_FINAL.xlsx"
)
PHASE_I = OUT / "PHASE_I_FINAL"
QA_MD = OUT / "STARTER_MASTER_FINAL_LOCK_QA.md"
QA_JSON = OUT / "STARTER_MASTER_FINAL_LOCK_QA.json"

PENDING_PCS = {"1013084", "1014789", "1014790"}
PENDING_VARIANT = "IA-ST-VAR-0037"
PENDING_SOURCE = "IA-ST-CFG-0018"
PROVISIONAL_SET = "ST-018-STD-04"

SCOPE_STATUS = "EXPORT-READY STARTER SCOPE"
PHYS_CONTROLLED = "CONTROLLED PACKAGING SET"
PHYS_BOM_REQUIRED = "BOM DATA REQUIRED — DO NOT ISSUE DOCUMENTS"
CFG_CONTROLLED = "CONTROLLED"
CFG_BOM_REQUIRED = "BOM DATA REQUIRED"
NOT_ISSUED = "NOT ISSUED"
BOM_DATA_REQUIRED = "BOM DATA REQUIRED"
PROVISIONAL_NOTE = "PROVISIONAL — NOT ISSUED / BOM DATA REQUIRED"

DOC_TYPES = {
    "TF": "YS/D/0020",
    "DOC": "YS/D/0021",
    "LABEL": "YS/D/0022",
    "STM": "YS/D/0023",
}

NAVY = "0E2A47"
WHITE = "FFFFFF"
INK = "1C2430"
BAND = "F3F6F9"
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)


def _font(size=9, bold=False, color=INK) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color)


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def style_header(ws, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = _font(9, True, WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = HAIR
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    style_header(ws, headers)
    for r_i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(r_i + 2, c, v)
            cell.font = _font(9)
            cell.border = HAIR
            cell.fill = _fill(BAND if r_i % 2 else WHITE)
            cell.alignment = Alignment(vertical="center")
    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(str(h)) + 2, 14), 42)


def linked_str(codes: list[str], sep: str = "; ") -> str:
    return sep.join(sorted(set(codes), key=lambda x: (len(x), x)))


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def word_tf_fingerprint() -> dict[str, Any]:
    files = sorted((PHASE_I / "01_STARTER").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "02_INDUSTRIAL").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "03_CONTAINER").rglob("01_Technical_File.docx"))
    files = [p for p in files if not p.name.startswith("~$")]
    digests = []
    h = hashlib.sha256()
    for p in files:
        d = sha256_file(p)
        digests.append(d)
        h.update(d.encode())
        h.update(str(p.relative_to(PHASE_I)).encode())
    return {"count": len(files), "aggregate": h.hexdigest(), "sample": digests[:3]}


def sheet_rows(path: Path, name: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    wb.close()
    return headers, rows


def original_240_codes() -> set[str]:
    wb = load_workbook(GOLDEN, data_only=True, read_only=True)
    ws = wb["01_FINAL_CONFIG_MASTER"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["Family"]] or "").upper() == "STARTER":
            codes.add(str(row[idx["Packaging Set Code"]]))
    wb.close()
    return codes


def run() -> dict[str, Any]:
    fp_before = word_tf_fingerprint()

    if not STARTER.exists():
        raise FileNotFoundError(STARTER)

    # Backup first (do not overwrite an existing backup if already locked once —
    # user asked create backup; refresh from current pre-lock state)
    shutil.copy2(STARTER, BACKUP)

    _, products = sheet_rows(STARTER, "PRODUCT_MASTER")
    _, configs = sheet_rows(STARTER, "CONFIG_MASTER")
    _, boms = sheet_rows(STARTER, "BOM_MASTER")
    _, scope_recon = sheet_rows(STARTER, "SCOPE_RECONCILIATION")

    assert len(products) == 2046

    controlled_cfgs = [c for c in configs if str(c["Packaging Set Code"]) != PROVISIONAL_SET]
    provisional_cfg = next(
        (c for c in configs if str(c["Packaging Set Code"]) == PROVISIONAL_SET), None
    )
    assert len(controlled_cfgs) == 311, f"expected 311 controlled, got {len(controlled_cfgs)}"
    assert provisional_cfg is not None, "missing provisional ST-018-STD-04 row"

    cfg_by_set = {str(c["Packaging Set Code"]): c for c in controlled_cfgs}
    orig_240 = original_240_codes()
    assert orig_240 <= set(cfg_by_set), "existing 240 missing from controlled sets"
    new_71 = set(cfg_by_set) - orig_240
    assert len(new_71) == 71, f"expected 71 new, got {len(new_71)}"

    # Build workbook
    wb = Workbook()

    # HOME
    ws = wb.active
    ws.title = "00_HOME"
    ws["A1"] = "İNCI AKÜ PPWR — STARTER MASTER Rev.00 — FINAL DATA LOCK"
    ws["A1"].font = _font(14, True, NAVY)
    ws["A3"] = "Scope coverage: 2046 / 2046 Product Codes"
    ws["A4"] = "Controlled physical Packaging Sets: 311 (240 existing + 71 new)"
    ws["A5"] = "Products with proven Packaging Set: 2043 / 2046"
    ws["A6"] = "Products BOM DATA REQUIRED: 3 / 2046 (1013084 | 1014789 | 1014790)"
    ws["A7"] = "Pending variant: IA-ST-VAR-0037 — NOT ISSUED / no documents"
    ws["A8"] = "Provisional code ST-018-STD-04: PROVISIONAL — NOT ISSUED / BOM DATA REQUIRED (not in 311)"
    ws["A10"] = "Document type numbers: TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023"
    ws["A11"] = "NO Word regeneration in this lock phase."

    # PRODUCT_MASTER
    ws = wb.create_sheet("PRODUCT_MASTER")
    pm_headers = [
        "Product Code",
        "Packaging Set Code",
        "Technical Description",
        "Customer / Market",
        "Source Configuration ID",
        "Final Configuration ID",
        "Packaging Tare kg",
        "Scope Status",
        "Physical Packaging Status",
        "Battery Type",
        "Nominal Qty",
        "Legacy Market Status",
    ]
    pm_rows = []
    for p in sorted(products, key=lambda x: str(x["Product Code"])):
        pc = str(p["Product Code"])
        if pc in PENDING_PCS:
            pm_rows.append(
                [
                    pc,
                    BOM_DATA_REQUIRED,
                    p.get("Technical Description"),
                    p.get("Customer / Market"),
                    p.get("Source Configuration ID") or PENDING_SOURCE,
                    NOT_ISSUED,
                    "",
                    SCOPE_STATUS,
                    PHYS_BOM_REQUIRED,
                    p.get("Battery Type"),
                    p.get("Nominal Qty"),
                    p.get("Legacy Market Status"),
                ]
            )
        else:
            sc = str(p["Packaging Set Code"])
            # if was pointing at provisional, should not happen for other products
            assert sc != PROVISIONAL_SET, f"{pc} still on provisional set"
            cfg = cfg_by_set[sc]
            pm_rows.append(
                [
                    pc,
                    sc,
                    p.get("Technical Description"),
                    p.get("Customer / Market"),
                    p.get("Source Configuration ID"),
                    cfg.get("Final Configuration ID"),
                    cfg.get("Packaging Tare kg"),
                    SCOPE_STATUS,
                    PHYS_CONTROLLED,
                    p.get("Battery Type"),
                    p.get("Nominal Qty"),
                    p.get("Legacy Market Status"),
                ]
            )
    write_table(ws, pm_headers, pm_rows)

    # CONFIG_MASTER — 311 controlled + 1 pending
    ws = wb.create_sheet("CONFIG_MASTER")
    cm_headers = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Final Configuration ID",
        "Product Count",
        "Nominal Qty",
        "Packaging Description",
        "Packaging Tare kg",
        "BOM Line Count",
        "Configuration Status",
        "Origin",
        "Variant Basis TR",
        "Variant Basis EN",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "TF Type No",
        "DoC Type No",
        "Label Type No",
        "STM Type No",
        "Notes",
    ]
    cm_rows = []
    for sc in sorted(cfg_by_set):
        c = cfg_by_set[sc]
        origin = str(c.get("Origin") or "")
        if origin.startswith("EXISTING"):
            origin_out = "EXISTING_240"
        else:
            origin_out = "NEW_VALIDATED_PHYSICAL"
        cm_rows.append(
            [
                sc,
                c.get("Linked Product Codes"),
                c.get("Source Configuration ID"),
                c.get("Final Configuration ID"),
                c.get("Product Count"),
                c.get("Nominal Qty"),
                c.get("Pallet / Packaging Description"),
                c.get("Packaging Tare kg"),
                c.get("BOM Line Count"),
                CFG_CONTROLLED,
                origin_out,
                c.get("Variant Basis TR"),
                c.get("Variant Basis EN"),
                c.get("Technical File ID"),
                c.get("EU DoC ID"),
                c.get("Label ID"),
                c.get("Shipment Statement ID"),
                DOC_TYPES["TF"],
                DOC_TYPES["DOC"],
                DOC_TYPES["LABEL"],
                DOC_TYPES["STM"],
                "",
            ]
        )
    # pending row
    cm_rows.append(
        [
            NOT_ISSUED,
            " | ".join(sorted(PENDING_PCS)),
            PENDING_SOURCE,
            NOT_ISSUED,
            3,
            provisional_cfg.get("Nominal Qty"),
            provisional_cfg.get("Pallet / Packaging Description"),
            "",
            0,
            CFG_BOM_REQUIRED,
            "PENDING_BOM_DATA",
            provisional_cfg.get("Variant Basis TR"),
            f"{PENDING_VARIANT}",
            NOT_ISSUED,
            NOT_ISSUED,
            NOT_ISSUED,
            NOT_ISSUED,
            DOC_TYPES["TF"],
            DOC_TYPES["DOC"],
            DOC_TYPES["LABEL"],
            DOC_TYPES["STM"],
            f"{PROVISIONAL_NOTE}; former provisional code {PROVISIONAL_SET}; variant {PENDING_VARIANT}",
        ]
    )
    write_table(ws, cm_headers, cm_rows)

    # BOM_MASTER — controlled only
    ws = wb.create_sheet("BOM_MASTER")
    bm_headers = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Component Code",
        "Component Description",
        "Quantity",
        "UOM",
        "Unit Weight",
        "Line Weight",
    ]
    bm_rows = []
    for b in boms:
        sc = str(b["Packaging Set Code"])
        if sc == PROVISIONAL_SET or sc not in cfg_by_set:
            continue
        bm_rows.append(
            [
                sc,
                b.get("Linked Product Codes"),
                b.get("Source Configuration ID"),
                b.get("Component Code"),
                b.get("Component Description"),
                b.get("Quantity"),
                b.get("UOM"),
                b.get("Unit Weight"),
                b.get("Line Weight"),
            ]
        )
    write_table(ws, bm_headers, bm_rows)

    # DOCUMENT_SCOPE
    ws = wb.create_sheet("DOCUMENT_SCOPE")
    ds_headers = [
        "Product Code",
        "Packaging Set Code",
        "Source Configuration ID",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "TF Type No",
        "DoC Type No",
        "Label Type No",
        "STM Type No",
        "Word Status",
        "Physical Packaging Status",
    ]
    ds_rows = []
    for p in sorted(products, key=lambda x: str(x["Product Code"])):
        pc = str(p["Product Code"])
        if pc in PENDING_PCS:
            ds_rows.append(
                [
                    pc,
                    BOM_DATA_REQUIRED,
                    p.get("Source Configuration ID") or PENDING_SOURCE,
                    NOT_ISSUED,
                    NOT_ISSUED,
                    NOT_ISSUED,
                    NOT_ISSUED,
                    DOC_TYPES["TF"],
                    DOC_TYPES["DOC"],
                    DOC_TYPES["LABEL"],
                    DOC_TYPES["STM"],
                    "DO_NOT_ISSUE",
                    PHYS_BOM_REQUIRED,
                ]
            )
        else:
            sc = str(p["Packaging Set Code"])
            cfg = cfg_by_set[sc]
            origin = str(cfg.get("Origin") or "")
            word_status = (
                "EXISTING_LINKED" if origin.startswith("EXISTING") else "PENDING_GENERATION"
            )
            ds_rows.append(
                [
                    pc,
                    sc,
                    p.get("Source Configuration ID"),
                    cfg.get("Technical File ID"),
                    cfg.get("EU DoC ID"),
                    cfg.get("Label ID"),
                    cfg.get("Shipment Statement ID"),
                    DOC_TYPES["TF"],
                    DOC_TYPES["DOC"],
                    DOC_TYPES["LABEL"],
                    DOC_TYPES["STM"],
                    word_status,
                    PHYS_CONTROLLED,
                ]
            )
    write_table(ws, ds_headers, ds_rows)

    # DOCUMENT_CENTER — 311 controlled + pending NOT ISSUED
    ws = wb.create_sheet("DOCUMENT_CENTER")
    dc_headers = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Final Configuration ID",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "Configuration Status",
        "Word Status",
        "TF Type No",
        "DoC Type No",
        "Label Type No",
        "STM Type No",
    ]
    dc_rows = []
    for sc in sorted(cfg_by_set):
        c = cfg_by_set[sc]
        origin = str(c.get("Origin") or "")
        dc_rows.append(
            [
                sc,
                c.get("Linked Product Codes"),
                c.get("Source Configuration ID"),
                c.get("Final Configuration ID"),
                c.get("Technical File ID"),
                c.get("EU DoC ID"),
                c.get("Label ID"),
                c.get("Shipment Statement ID"),
                CFG_CONTROLLED,
                "EXISTING_LINKED" if origin.startswith("EXISTING") else "PENDING_GENERATION",
                DOC_TYPES["TF"],
                DOC_TYPES["DOC"],
                DOC_TYPES["LABEL"],
                DOC_TYPES["STM"],
            ]
        )
    dc_rows.append(
        [
            NOT_ISSUED,
            " | ".join(sorted(PENDING_PCS)),
            PENDING_SOURCE,
            NOT_ISSUED,
            NOT_ISSUED,
            NOT_ISSUED,
            NOT_ISSUED,
            NOT_ISSUED,
            CFG_BOM_REQUIRED,
            "DO_NOT_ISSUE",
            DOC_TYPES["TF"],
            DOC_TYPES["DOC"],
            DOC_TYPES["LABEL"],
            DOC_TYPES["STM"],
        ]
    )
    write_table(ws, dc_headers, dc_rows)

    # Document indexes — controlled only (no provisional docs)
    for sheet_name, id_key in (
        ("TECHNICAL_FILES", "Technical File ID"),
        ("DECLARATIONS_OF_CONFORMITY", "EU DoC ID"),
        ("LABELS", "Label ID"),
        ("SHIPMENT_STATEMENTS", "Shipment Statement ID"),
    ):
        ws = wb.create_sheet(sheet_name)
        headers = [
            "Packaging Set Code",
            "Linked Product Codes",
            "Document ID",
            "Management Type No",
            "Source Configuration ID",
            "Configuration Status",
            "Word Status",
        ]
        type_key = {
            "TECHNICAL_FILES": "TF",
            "DECLARATIONS_OF_CONFORMITY": "DOC",
            "LABELS": "LABEL",
            "SHIPMENT_STATEMENTS": "STM",
        }[sheet_name]
        rows = []
        for sc in sorted(cfg_by_set):
            c = cfg_by_set[sc]
            origin = str(c.get("Origin") or "")
            rows.append(
                [
                    sc,
                    c.get("Linked Product Codes"),
                    c.get(id_key),
                    DOC_TYPES[type_key],
                    c.get("Source Configuration ID"),
                    CFG_CONTROLLED,
                    "EXISTING_LINKED" if origin.startswith("EXISTING") else "PENDING_GENERATION",
                ]
            )
        write_table(ws, headers, rows)

    # SEARCH_DATA
    ws = wb.create_sheet("SEARCH_DATA")
    sd_headers = [
        "Product Code",
        "Packaging Set Code",
        "Technical Description",
        "Source Configuration ID",
        "Final Configuration ID",
        "Customer / Market",
        "Packaging Tare kg",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "Scope Status",
        "Physical Packaging Status",
    ]
    sd_rows = []
    for p in sorted(products, key=lambda x: str(x["Product Code"])):
        pc = str(p["Product Code"])
        if pc in PENDING_PCS:
            sd_rows.append(
                [
                    pc,
                    BOM_DATA_REQUIRED,
                    p.get("Technical Description"),
                    p.get("Source Configuration ID") or PENDING_SOURCE,
                    NOT_ISSUED,
                    p.get("Customer / Market"),
                    "",
                    NOT_ISSUED,
                    NOT_ISSUED,
                    NOT_ISSUED,
                    NOT_ISSUED,
                    SCOPE_STATUS,
                    PHYS_BOM_REQUIRED,
                ]
            )
        else:
            sc = str(p["Packaging Set Code"])
            cfg = cfg_by_set[sc]
            sd_rows.append(
                [
                    pc,
                    sc,
                    p.get("Technical Description"),
                    p.get("Source Configuration ID"),
                    cfg.get("Final Configuration ID"),
                    p.get("Customer / Market"),
                    cfg.get("Packaging Tare kg"),
                    cfg.get("Technical File ID"),
                    cfg.get("EU DoC ID"),
                    cfg.get("Label ID"),
                    cfg.get("Shipment Statement ID"),
                    SCOPE_STATUS,
                    PHYS_CONTROLLED,
                ]
            )
    write_table(ws, sd_headers, sd_rows)

    # SCOPE_RECONCILIATION — summary + detail
    ws = wb.create_sheet("SCOPE_RECONCILIATION")
    sr_headers = ["Metric", "Value", "Notes"]
    sr_rows = [
        ["STARTER PRODUCT CODES", 2046, ""],
        ["SCOPE COVERAGE", "2046 / 2046", "All Product Codes in export-ready scope"],
        ["CONTROLLED PHYSICAL PACKAGING SETS", 311, "240 existing + 71 new validated"],
        ["PRODUCTS WITH PROVEN PACKAGING SET", "2043 / 2046", PHYS_CONTROLLED],
        ["PRODUCTS REQUIRING BOM DATA", "3 / 2046", PHYS_BOM_REQUIRED],
        ["PENDING SOURCE / VARIANT", PENDING_VARIANT, PENDING_SOURCE],
        ["PENDING PRODUCT CODES", " | ".join(sorted(PENDING_PCS)), ""],
        ["Existing controlled sets", 240, "unchanged codes"],
        ["New validated physical sets", 71, ""],
        ["Final controlled physical sets", 311, "excludes provisional"],
        [
            "PROVISIONAL CODE (NOT CONTROLLED)",
            PROVISIONAL_SET,
            PROVISIONAL_NOTE,
        ],
        ["Pending configuration Packaging Set Code", NOT_ISSUED, "do not issue documents"],
    ]
    # append variant-level reconciliation (controlled only + pending)
    sr_rows.append(["---", "---", "Variant reconciliation detail below"])
    for r in scope_recon:
        vid = str(r.get("Variant ID") or "")
        sc = str(r.get("Packaging Set Code") or "")
        if sc == PROVISIONAL_SET or vid == PENDING_VARIANT:
            sr_rows.append(
                [
                    f"VARIANT {PENDING_VARIANT}",
                    NOT_ISSUED,
                    f"{PROVISIONAL_NOTE}; former {PROVISIONAL_SET}; products "
                    + " | ".join(sorted(PENDING_PCS)),
                ]
            )
        else:
            sr_rows.append(
                [
                    f"VARIANT {vid}",
                    sc,
                    f"Source {r.get('Source Configuration ID')}; action {r.get('Action')}; "
                    f"products={r.get('Product Count')}",
                ]
            )
    write_table(ws, sr_headers, sr_rows)

    # PROVISIONAL_REGISTER (clarity sheet)
    ws = wb.create_sheet("PROVISIONAL_REGISTER")
    write_table(
        ws,
        [
            "Provisional Code",
            "Status",
            "Variant ID",
            "Source Configuration ID",
            "Linked Product Codes",
            "Counted in 311",
            "Documents",
        ],
        [
            [
                PROVISIONAL_SET,
                PROVISIONAL_NOTE,
                PENDING_VARIANT,
                PENDING_SOURCE,
                " | ".join(sorted(PENDING_PCS)),
                "NO",
                "NOT ISSUED — DO NOT GENERATE",
            ]
        ],
    )

    if STARTER.exists():
        STARTER.unlink()
    wb.save(STARTER)
    wb.close()

    # QA
    fp_after = word_tf_fingerprint()
    sw = load_workbook(STARTER, read_only=True, data_only=True)

    pm = sw["PRODUCT_MASTER"]
    ph = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
    assert ph[0] == "Product Code" and ph[1] == "Packaging Set Code"
    pm_rows_qa = list(pm.iter_rows(min_row=2, values_only=True))
    assert len(pm_rows_qa) == 2046
    pci, psi = ph.index("Product Code"), ph.index("Packaging Set Code")
    physi = ph.index("Physical Packaging Status")
    pcs = [str(r[pci]) for r in pm_rows_qa]
    assert len(set(pcs)) == 2046
    controlled_products = [
        r for r in pm_rows_qa if str(r[physi]) == PHYS_CONTROLLED
    ]
    bom_req_products = [
        r for r in pm_rows_qa if str(r[psi]) == BOM_DATA_REQUIRED
    ]
    assert len(controlled_products) == 2043
    assert len(bom_req_products) == 3
    assert {str(r[pci]) for r in bom_req_products} == PENDING_PCS

    cm = sw["CONFIG_MASTER"]
    ch = [c.value for c in next(cm.iter_rows(min_row=1, max_row=1))]
    assert ch[0] == "Packaging Set Code" and ch[1] == "Linked Product Codes"
    cm_rows_qa = list(cm.iter_rows(min_row=2, values_only=True))
    sci = ch.index("Packaging Set Code")
    lpi = ch.index("Linked Product Codes")
    sti = ch.index("Configuration Status")
    pci_count = ch.index("Product Count")
    controlled_cfg_rows = [r for r in cm_rows_qa if str(r[sti]) == CFG_CONTROLLED]
    pending_cfg_rows = [r for r in cm_rows_qa if str(r[sti]) == CFG_BOM_REQUIRED]
    assert len(controlled_cfg_rows) == 311
    assert len(pending_cfg_rows) == 1
    assert str(pending_cfg_rows[0][sci]) == NOT_ISSUED
    assert PROVISIONAL_SET not in {str(r[sci]) for r in controlled_cfg_rows}

    # every controlled product maps to exactly one set
    set_of_product = {}
    one_to_one_ok = True
    for r in controlled_products:
        pc, sc = str(r[pci]), str(r[psi])
        if pc in set_of_product and set_of_product[pc] != sc:
            one_to_one_ok = False
        set_of_product[pc] = sc
        if sc not in cfg_by_set:
            one_to_one_ok = False

    # linked products complete
    linked_ok = True
    set_to_pm = defaultdict(set)
    for pc, sc in set_of_product.items():
        set_to_pm[sc].add(pc)
    for r in controlled_cfg_rows:
        sc = str(r[sci])
        linked = {x.strip() for x in str(r[lpi] or "").replace("|", ";").split(";") if x.strip()}
        expected = set_to_pm.get(sc, set())
        if linked != expected:
            linked_ok = False
        if int(r[pci_count] or 0) != len(linked):
            linked_ok = False

    adjacency_ok = True
    for sheet, a, b in (
        ("PRODUCT_MASTER", "Product Code", "Packaging Set Code"),
        ("CONFIG_MASTER", "Packaging Set Code", "Linked Product Codes"),
        ("BOM_MASTER", "Packaging Set Code", "Linked Product Codes"),
        ("DOCUMENT_SCOPE", "Product Code", "Packaging Set Code"),
        ("DOCUMENT_CENTER", "Packaging Set Code", "Linked Product Codes"),
        ("SEARCH_DATA", "Product Code", "Packaging Set Code"),
        ("TECHNICAL_FILES", "Packaging Set Code", "Linked Product Codes"),
    ):
        headers = [c.value for c in next(sw[sheet].iter_rows(min_row=1, max_row=1))]
        if headers[0] != a or headers[1] != b:
            adjacency_ok = False

    # doc scope pending
    dsh = [c.value for c in next(sw["DOCUMENT_SCOPE"].iter_rows(min_row=1, max_row=1))]
    for row in sw["DOCUMENT_SCOPE"].iter_rows(min_row=2, values_only=True):
        if str(row[dsh.index("Product Code")]) in PENDING_PCS:
            assert str(row[dsh.index("Packaging Set Code")]) == BOM_DATA_REQUIRED
            assert str(row[dsh.index("Technical File ID")]) == NOT_ISSUED
            assert str(row[dsh.index("EU DoC ID")]) == NOT_ISSUED
            assert str(row[dsh.index("Label ID")]) == NOT_ISSUED
            assert str(row[dsh.index("Shipment Statement ID")]) == NOT_ISSUED

    # TF index count = 311
    tf_n = sum(1 for _ in sw["TECHNICAL_FILES"].iter_rows(min_row=2))
    assert tf_n == 311

    sw.close()

    provisional_not_in_311 = PROVISIONAL_SET not in set(cfg_by_set)
    existing_unchanged = orig_240 <= set(cfg_by_set) and len(orig_240) == 240
    word_ok = fp_before["aggregate"] == fp_after["aggregate"] and fp_after["count"] == 247

    final = (
        len(pcs) == 2046
        and len(controlled_products) == 2043
        and len(bom_req_products) == 3
        and len(controlled_cfg_rows) == 311
        and len(pending_cfg_rows) == 1
        and one_to_one_ok
        and linked_ok
        and adjacency_ok
        and provisional_not_in_311
        and existing_unchanged
        and word_ok
        and BACKUP.exists()
    )

    qa = {
        "unique_starter_product_codes": 2046,
        "scope_coverage": "2046 / 2046",
        "products_with_controlled_packaging_set": "2043 / 2046",
        "products_bom_data_required": "3 / 2046",
        "controlled_existing_packaging_sets": 240,
        "controlled_new_packaging_sets": 71,
        "final_controlled_packaging_sets": 311,
        "pending_configurations": 1,
        "pending_configuration": PENDING_VARIANT,
        "pending_product_codes": sorted(PENDING_PCS),
        "every_controlled_product_one_set": "PASS" if one_to_one_ok else "FAIL",
        "every_set_complete_linked_products": "PASS" if linked_ok else "FAIL",
        "product_code_packaging_set_adjacency": "PASS" if adjacency_ok else "FAIL",
        "no_provisional_in_311": "PASS" if provisional_not_in_311 else "FAIL",
        "existing_240_unchanged": "PASS" if existing_unchanged else "FAIL",
        "word_hash_changed": 0 if word_ok else 1,
        "final_master_data_lock": "PASS" if final else "FAIL",
        "starter_workbook": str(STARTER),
        "backup_workbook": str(BACKUP),
        "provisional_code": PROVISIONAL_SET,
        "word_fingerprint": fp_after,
    }

    lines = [
        "# STARTER MASTER FINAL LOCK QA",
        "",
        "Unique Starter Product Codes:",
        "2046",
        "",
        "Scope coverage:",
        "2046 / 2046",
        "",
        "Products with controlled Packaging Set:",
        "2043 / 2046",
        "",
        "Products BOM DATA REQUIRED:",
        "3 / 2046",
        "",
        "Controlled existing Packaging Sets:",
        "240",
        "",
        "Controlled new Packaging Sets:",
        "71",
        "",
        "FINAL CONTROLLED PACKAGING SETS:",
        "311",
        "",
        "Pending configurations:",
        "1",
        "",
        "Pending configuration:",
        PENDING_VARIANT,
        "",
        "Pending Product Codes:",
        "1013084",
        "1014789",
        "1014790",
        "",
        "Every controlled Product Code maps to exactly one Packaging Set:",
        qa["every_controlled_product_one_set"],
        "",
        "Every Packaging Set shows complete Linked Product Codes:",
        qa["every_set_complete_linked_products"],
        "",
        "Product Code <-> Packaging Set adjacency:",
        qa["product_code_packaging_set_adjacency"],
        "",
        "No provisional configuration counted inside 311:",
        qa["no_provisional_in_311"],
        "",
        "Existing 240 set codes unchanged:",
        qa["existing_240_unchanged"],
        "",
        "Word hash changed:",
        "0" if word_ok else "1",
        "",
        "FINAL MASTER DATA LOCK:",
        qa["final_master_data_lock"],
        "",
        "STOP.",
        "",
        "DO NOT GENERATE WORDS.",
        "DO NOT START PIMS.",
        "",
        f"Backup: `{BACKUP}`",
        f"Locked master: `{STARTER}`",
    ]
    QA_MD.write_text("\n".join(lines), encoding="utf-8")
    QA_JSON.write_text(json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8")
    try:
        print("\n".join(lines))
    except UnicodeEncodeError:
        print("\n".join(lines).encode("ascii", "replace").decode("ascii"))
    return qa


if __name__ == "__main__":
    run()
