"""Add BULK_SEARCH sheet + customer pack export tool to delivery package."""

from __future__ import annotations

import hashlib
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
FINAL = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
ENG = FINAL / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
DOC = FINAL / "01_DOCUMENT_SETS"
PACKS = FINAL / "05_CUSTOMER_PACKS"
ZIP = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL.zip"
SHA = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL_SHA256.txt"
TOOLS = FINAL / "00_TOOLS"

NAVY, GOLD, WHITE, INK, BAND = "0E2A47", "C8A24A", "FFFFFF", "1C2430", "F3F6F9"
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)

FILES = [
    "01_Technical_File.docx",
    "01_Technical_File.pdf",
    "02_EU_DoC.docx",
    "02_EU_DoC.pdf",
    "03_Label.docx",
    "03_Label.pdf",
    "04_Shipment_Statement.docx",
    "04_Shipment_Statement.pdf",
]


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def add_bulk_search_sheet() -> None:
    wb = load_workbook(ENG)
    if "BULK_SEARCH" in wb.sheetnames:
        del wb["BULK_SEARCH"]
    # insert after SEARCH
    idx = wb.sheetnames.index("SEARCH") + 1 if "SEARCH" in wb.sheetnames else 1
    ws = wb.create_sheet("BULK_SEARCH", idx)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "TOPLU ÜRÜN KODU / MÜŞTERİ PAKETİ"
    ws["A1"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
    ws.merge_cells("A1:H1")

    ws["A2"] = (
        "A müşterisine verilen tüm ürün kodlarını aşağıya yapıştırın (her satıra 1 kod). "
        "Sonra 00_TOOLS\\EXPORT_CUSTOMER_PACK.cmd çalıştırın — sistem ilgili Packaging Set "
        "dokümanlarını tek klasöre kopyalar."
    )
    ws["A2"].font = Font(name=FONT, size=10, color=INK)
    ws.merge_cells("A2:H2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 36

    ws["A4"] = "MÜŞTERİ ADI / CUSTOMER NAME"
    ws["A4"].font = Font(name=FONT, size=10, bold=True, color=WHITE)
    ws["A4"].fill = PatternFill("solid", fgColor=NAVY)
    ws["B4"] = "Musteri_A"
    ws["B4"].font = Font(name=FONT, size=12, bold=True, color=NAVY)
    ws["B4"].fill = PatternFill("solid", fgColor="FFF8E7")
    ws["B4"].border = Border(
        left=Side(style="medium", color=GOLD),
        right=Side(style="medium", color=GOLD),
        top=Side(style="medium", color=GOLD),
        bottom=Side(style="medium", color=GOLD),
    )

    ws["A5"] = "NOT"
    ws["B5"] = (
        "Kodları A8:A107 aralığına yapıştırın. Virgüllü satır da kabul edilir. "
        "Export: paket kökünde 00_TOOLS\\EXPORT_CUSTOMER_PACK.cmd"
    )
    ws["B5"].font = Font(name=FONT, size=9, color="A12622")
    ws.merge_cells("B5:H5")

    ws["A7"] = "ÜRÜN KODLARI (1 satır = 1 kod)"
    ws["A7"].font = Font(name=FONT, size=10, bold=True, color=WHITE)
    ws["A7"].fill = PatternFill("solid", fgColor=NAVY)
    ws["B7"] = "Örnek"
    ws["B7"].font = Font(name=FONT, size=9, bold=True, color=WHITE)
    ws["B7"].fill = PatternFill("solid", fgColor="1F4E79")

    # input area A8:A107
    for i in range(8, 108):
        cell = ws.cell(i, 1, "")
        cell.border = HAIR
        cell.fill = PatternFill("solid", fgColor="FFFDF5" if i % 2 == 0 else WHITE)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.number_format = "@"

    # examples in B
    for i, ex in enumerate(["1002103", "1015169", "1000069"], start=8):
        ws.cell(i, 2, ex).font = Font(name=FONT, size=9, italic=True, color="6B7280")

    # Results header area (filled by export tool)
    ws["D7"] = "SON EXPORT ÖZETİ (script doldurur)"
    ws["D7"].font = Font(name=FONT, size=10, bold=True, color=WHITE)
    ws["D7"].fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells("D7:H7")
    headers = ["Product Code", "Packaging Set", "Status", "Docs Copied", "Notes"]
    for c, h in enumerate(headers, 4):
        cell = ws.cell(8, c, h)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = HAIR

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 3
    for c, w in enumerate([16, 18, 18, 12, 40], 4):
        ws.column_dimensions[get_column_letter(c)].width = w

    # HOME nav note
    if "00_HOME" in wb.sheetnames:
        home = wb["00_HOME"]
        home["B30"] = (
            "TOPLU MÜŞTERİ: BULK_SEARCH sayfasına ürün kodlarını yapıştırın → "
            "00_TOOLS\\EXPORT_CUSTOMER_PACK.cmd ile klasör oluşturun."
        )
        home["B30"].font = Font(name=FONT, size=9, bold=True, color=NAVY)
        home.merge_cells("B30:I30")

    # hyperlink from HOME if possible — openpyxl internal
    # Add sheet order already set

    tmp = ENG.with_name("_eng_bulk_tmp.xlsx")
    wb.save(tmp)
    wb.close()
    ENG.unlink()
    shutil.move(str(tmp), str(ENG))

    # COM: HOME button to BULK_SEARCH + keep existing links
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        xwb = excel.Workbooks.Open(str(ENG.resolve()))
        # ensure BULK exists
        names = [xwb.Worksheets(i).Name for i in range(1, xwb.Worksheets.Count + 1)]
        if "BULK_SEARCH" not in names:
            raise RuntimeError("BULK_SEARCH missing after save")
        home = xwb.Worksheets("00_HOME")
        # delete old bulk nav shapes
        try:
            for shp in list(home.Shapes):
                if str(getattr(shp, "Name", "")).startswith("NAV_BULK"):
                    shp.Delete()
        except Exception:
            pass
        try:
            shp = home.Shapes.AddShape(5, 400, 120, 160, 28)  # msoShapeRoundedRectangle
            shp.Name = "NAV_BULK"
            shp.Fill.ForeColor.RGB = 0x0E + 256 * 0x2A + 65536 * 0x47
            shp.Line.ForeColor.RGB = 0xC8 + 256 * 0xA2 + 65536 * 0x4A
            shp.TextFrame.Characters().Text = "TOPLU MÜŞTERİ"
            shp.TextFrame.Characters().Font.Color = 0xFFFFFF
            shp.TextFrame.Characters().Font.Size = 11
            shp.TextFrame.Characters().Font.Bold = True
            shp.TextFrame.HorizontalAlignment = 2
            home.Hyperlinks.Add(Anchor=shp, Address="", SubAddress="'BULK_SEARCH'!A1")
        except Exception as e:
            print("nav shape skip", e)
        xwb.Save()
        xwb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def write_export_tool() -> None:
    TOOLS.mkdir(parents=True, exist_ok=True)
    PACKS.mkdir(parents=True, exist_ok=True)
    (PACKS / "README.txt").write_text(
        "Musteri paketleri buraya olusturulur.\r\n"
        "Kullanim: BULK_SEARCH'e kodlari yapistirin, sonra EXPORT_CUSTOMER_PACK.cmd calistirin.\r\n",
        encoding="utf-8",
    )

    script = TOOLS / "export_customer_pack.py"
    script.write_text(
        r'''# -*- coding: utf-8 -*-
"""Export customer document pack from BULK_SEARCH product codes."""
from __future__ import annotations

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

HERE = Path(__file__).resolve().parent
FINAL = HERE.parent
ENG = FINAL / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
DOC = FINAL / "01_DOCUMENT_SETS"
OUT_ROOT = FINAL / "05_CUSTOMER_PACKS"

FILES = [
    "01_Technical_File.docx",
    "01_Technical_File.pdf",
    "02_EU_DoC.docx",
    "02_EU_DoC.pdf",
    "03_Label.docx",
    "03_Label.pdf",
    "04_Shipment_Statement.docx",
    "04_Shipment_Statement.pdf",
]

def norm_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s.strip()

def split_codes(raw: str) -> list[str]:
    parts = re.split(r"[\s,;]+", raw.strip())
    return [p for p in (norm_code(x) for x in parts) if p]

def safe_name(s: str) -> str:
    s = re.sub(r'[<>:"/\\|?*]+', "_", s.strip())
    s = re.sub(r"\s+", "_", s)
    return s[:80] or "Musteri"

def main() -> int:
    if not ENG.exists():
        print("Engine bulunamadi:", ENG)
        return 1
    wb = load_workbook(ENG, data_only=False)
    if "BULK_SEARCH" not in wb.sheetnames:
        print("BULK_SEARCH sayfasi yok.")
        return 1
    if "SEARCH_DATA" not in wb.sheetnames:
        print("SEARCH_DATA yok.")
        return 1

    bulk = wb["BULK_SEARCH"]
    customer = safe_name(str(bulk["B4"].value or "Musteri"))
    codes: list[str] = []
    for r in range(8, 108):
        v = bulk.cell(r, 1).value
        if v is None or str(v).strip() == "":
            continue
        codes.extend(split_codes(str(v)))
    # unique preserve order
    seen = set()
    uniq = []
    for c in codes:
        if c not in seen:
            seen.add(c)
            uniq.append(c)
    codes = uniq
    if not codes:
        print("BULK_SEARCH A8:A107 bos. Urun kodlarini yapistirin.")
        return 1

    sd = wb["SEARCH_DATA"]
    # build lookup KEY/A -> row values
    lookup: dict[str, dict] = {}
    for r in range(2, sd.max_row + 1):
        pc = norm_code(sd.cell(r, 1).value)
        key = norm_code(sd.cell(r, 8).value) or pc
        row = {
            "product_code": pc,
            "set_code": str(sd.cell(r, 2).value or "").strip(),
            "desc": str(sd.cell(r, 3).value or "").strip(),
            "config_id": str(sd.cell(r, 4).value or "").strip(),
            "status": str(sd.cell(r, 7).value or "").strip(),
        }
        if pc:
            lookup[pc] = row
        if key:
            lookup[key] = row

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    dest = OUT_ROOT / f"{customer}_{stamp}"
    sets_dir = dest / "SETS"
    sets_dir.mkdir(parents=True, exist_ok=True)

    results = []
    copied_sets: set[str] = set()
    for code in codes:
        row = lookup.get(code)
        if not row:
            results.append([code, "", "NOT FOUND", 0, "Kod katalogda yok"])
            continue
        sc = row["set_code"]
        st = row["status"]
        note = ""
        ncopy = 0
        if "YURT" in st.upper() or "NOT ISSUED" in st.upper() or "DATA REQUIRED" in st.upper():
            note = "DOCUMENTS NOT ISSUED / DOMESTIC DATA GAP"
            results.append([code, sc, st, 0, note])
            continue
        if not sc:
            results.append([code, "", st, 0, "Packaging Set yok"])
            continue
        set_folder = sets_dir / sc
        set_folder.mkdir(parents=True, exist_ok=True)
        src_dir = DOC / sc
        if not src_dir.exists():
            results.append([code, sc, st, 0, "Set klasoru yok"])
            continue
        if sc not in copied_sets:
            for fname in FILES:
                src = src_dir / fname
                if src.exists() and src.stat().st_size > 0:
                    shutil.copy2(src, set_folder / fname)
                    ncopy += 1
            copied_sets.add(sc)
            note = f"Set kopyalandi ({ncopy} dosya)"
        else:
            note = "Ayni set — tekrar kopyalanmadi (paylasilan set)"
            ncopy = 8
        results.append([code, sc, st, ncopy, note])

    # INDEX workbook
    idx = Workbook()
    ws = idx.active
    ws.title = "INDEX"
    ws["A1"] = f"Musteri Paketi: {customer}"
    ws["A1"].font = Font(name="Tahoma", size=14, bold=True, color="0E2A47")
    ws["A2"] = f"Olusturma: {stamp}"
    ws["A3"] = f"Kod sayisi: {len(codes)} | Benzersiz set: {len(copied_sets)}"
    headers = ["Product Code", "Packaging Set", "Status", "Docs Copied", "Notes", "Description"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(5, c, h)
        cell.font = Font(name="Tahoma", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0E2A47")
    for i, row in enumerate(results, 6):
        code = row[0]
        desc = lookup.get(code, {}).get("desc", "")
        for c, v in enumerate(row + [desc], 1):
            ws.cell(i, c, v).font = Font(name="Tahoma", size=9)
    for c in range(1, 7):
        ws.column_dimensions[chr(64 + c)].width = 22
    idx_path = dest / "00_INDEX.xlsx"
    idx.save(idx_path)
    idx.close()

    # README
    (dest / "README.txt").write_text(
        f"Musteri: {customer}\n"
        f"Tarih: {stamp}\n"
        f"Urun kodu: {len(codes)}\n"
        f"Packaging Set klasoru: SETS\\\n"
        f"Detay: 00_INDEX.xlsx\n",
        encoding="utf-8",
    )

    # write summary back into BULK_SEARCH D9...
    # clear old summary
    for r in range(9, 120):
        for c in range(4, 9):
            bulk.cell(r, c).value = None
    for i, row in enumerate(results):
        for c, v in enumerate(row, 4):
            bulk.cell(9 + i, c, v)
    bulk["D6"] = f"Son export: {dest.name}  |  set={len(copied_sets)}  |  kod={len(codes)}"
    bulk["D6"].font = Font(name="Tahoma", size=10, bold=True, color="0E2A47")
    try:
        wb.save(ENG)
    except PermissionError:
        print("UYARI: Engine acik — ozet BULK_SEARCH'e yazilamadi. Klasor yine olusturuldu.")
    wb.close()

    print("OK")
    print("KLASOR:", dest)
    print("SET:", len(copied_sets))
    print("KOD:", len(codes))
    print("NOT_FOUND:", sum(1 for r in results if r[2] == "NOT FOUND" or r[1] == ""))
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
''',
        encoding="utf-8",
    )

    (TOOLS / "EXPORT_CUSTOMER_PACK.cmd").write_text(
        "@echo off\r\n"
        "setlocal\r\n"
        "cd /d \"%~dp0\"\r\n"
        "echo.\r\n"
        "echo === MUSTERI PAKETI EXPORT ===\r\n"
        "echo 1) Excel'de BULK_SEARCH sayfasina urun kodlarini yapistirin\r\n"
        "echo 2) Musteri adini B4'e yazin\r\n"
        "echo 3) Document Engine'i KAYDEDIP KAPATIN\r\n"
        "echo.\r\n"
        "pause\r\n"
        "python \"%~dp0export_customer_pack.py\"\r\n"
        "if errorlevel 1 (\r\n"
        "  echo HATA — Python / engine kontrol edin\r\n"
        "  pause\r\n"
        "  exit /b 1\r\n"
        ")\r\n"
        "echo.\r\n"
        "explorer \"%~dp0..\\05_CUSTOMER_PACKS\"\r\n"
        "pause\r\n",
        encoding="utf-8",
    )

    # also root shortcut cmd
    (FINAL / "00_EXPORT_CUSTOMER_PACK.cmd").write_text(
        "@echo off\r\n"
        "cd /d \"%~dp0\"\r\n"
        "call \"%~dp000_TOOLS\\EXPORT_CUSTOMER_PACK.cmd\"\r\n",
        encoding="utf-8",
    )


def rebuild_zip() -> str:
    if ZIP.exists():
        ZIP.unlink()
    with zipfile.ZipFile(ZIP, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for p in FINAL.rglob("*"):
            if p.is_file() and not p.name.startswith("~$"):
                zf.write(p, p.relative_to(FINAL).as_posix())
    digest = sha256_file(ZIP)
    SHA.write_text(digest + "\n", encoding="utf-8")
    return digest


def main() -> None:
    print("1) BULK_SEARCH sheet…", flush=True)
    add_bulk_search_sheet()
    print("2) Export tool…", flush=True)
    write_export_tool()
    print("3) ZIP…", flush=True)
    digest = rebuild_zip()
    print("ZIP", ZIP)
    print("SHA256", digest)
    print("DONE")


if __name__ == "__main__":
    main()
