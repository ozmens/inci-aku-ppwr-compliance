"""Finalize all 4 PPWR scopes: bulletproof engines, promote, ZIP, SHA256.

Never opens Microsoft Word. Assumes DOCX+PDF packs already exist.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import time
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
import sys

sys.path.insert(0, str(ROOT / "scripts"))
from ppwr_engine_builder import build_document_engine, verify_links  # noqa: E402

OUT = ROOT / "output"
DATE = "11.08.2026"
REPORT = OUT / "_MULTI_SCOPE_ALL_FINAL_QA.json"


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def zip_tree(src: Path, zip_path: Path) -> str:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        for f in src.rglob("*"):
            if f.is_file() and not f.name.startswith("~$"):
                zf.write(f, f.relative_to(src).as_posix())
    digest = sha256_file(zip_path)
    (zip_path.parent / f"{zip_path.stem}_SHA256.txt").write_text(digest + "\n", encoding="utf-8")
    # also name pattern used elsewhere: NAME_SHA256.txt where NAME is folder name
    (OUT / f"{src.name}_SHA256.txt").write_text(digest + "\n", encoding="utf-8")
    return digest


def count_packs(docs: Path) -> dict:
    folders = [p for p in docs.iterdir() if p.is_dir()] if docs.exists() else []
    docx = pdf = 0
    for folder in folders:
        for stem in [
            "01_Technical_File",
            "02_EU_DoC",
            "03_Label",
            "04_Shipment_Statement",
        ]:
            d = folder / f"{stem}.docx"
            p = folder / f"{stem}.pdf"
            if d.exists():
                docx += 1
            if p.exists() and p.stat().st_size > 500:
                pdf += 1
    return {"folders": len(folders), "docx": docx, "pdf": pdf}


def promote(candidate: Path, final: Path) -> None:
    if final.exists():
        bak = OUT / f"_BACKUP_{final.name}_{time.strftime('%Y%m%d_%H%M%S')}"
        final.rename(bak)
        print(f"backed up {final} -> {bak}", flush=True)
    print(f"promoting {candidate} -> {final}", flush=True)
    shutil.copytree(candidate, final)


def finalize_starter() -> dict:
    cand = OUT / "01_STARTER_INDIVIDUAL_DELIVERY_REV00_CANDIDATE"
    final = OUT / "01_STARTER_INDIVIDUAL_DELIVERY_REV00"
    docs = cand / "01_PRODUCTS"
    assert docs.exists(), "starter candidate missing"
    counts = count_packs(docs)
    print("STARTER counts", counts, flush=True)
    assert counts["folders"] >= 2000
    assert counts["docx"] == counts["folders"] * 4
    assert counts["pdf"] == counts["folders"] * 4

    # records from folders + master descriptions if available
    master = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
    desc_map: dict[str, str] = {}
    if master.exists():
        wb = load_workbook(master, data_only=True)
        ws = wb["PRODUCT_MASTER"]
        h = [c.value for c in next(ws.iter_rows(1, 1))]
        hi = {x: i for i, x in enumerate(h)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[hi["Product Code"]] or "").strip()
            if not code:
                continue
            desc = str(row[hi.get("Technical Description", 1)] or "").strip()
            desc_map[code] = desc
        wb.close()

    records = []
    for folder in sorted(docs.iterdir(), key=lambda p: p.name):
        if not folder.is_dir():
            continue
        code = folder.name
        records.append({"key": code, "label": desc_map.get(code, f"Starter product {code}")})

    print(f"STARTER engine build records={len(records)}", flush=True)
    eng = build_document_engine(
        delivery_root=cand,
        engine_filename="INCI_AKU_PPWR_STARTER_INDIVIDUAL_ENGINE_Rev00.xlsx",
        title="İNCI AKÜ PPWR — STARTER INDIVIDUAL ENGINE Rev00",
        docs_subdir="01_PRODUCTS",
        records=records,
        extra_home={
            "PUBLISH DATE": DATE,
            "SIGNATORY": "Numan Alver — Operations Director",
            "QMS": "TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023",
            "MODEL": "1 Product Code = 1 set (DOCX+PDF)",
            "DATA GAP": "See 00_CONTROL/DOMESTIC_DATA_GAP_LIST (not issued)",
        },
    )
    link = verify_links(cand, "01_PRODUCTS", [r["key"] for r in records])
    print("STARTER link verify", link, flush=True)
    assert link["missing"] == 0

    promote(cand, final)
    # ensure engine+cmd on final (copytree already has them)
    print("STARTER zipping…", flush=True)
    digest = zip_tree(final, OUT / f"{final.name}.zip")
    return {
        "GATE": "PASS",
        "folders": counts["folders"],
        "docx": counts["docx"],
        "pdf": counts["pdf"],
        "links_ok": link["ok"],
        "engine": str(final / "00_CONTROL" / eng.name),
        "final": str(final),
        "zip": str(OUT / f"{final.name}.zip"),
        "sha256": digest,
    }


def finalize_container() -> dict:
    final = OUT / "03_CONTAINER_DELIVERY_REV00"
    docs = final / "01_CONFIGS"
    assert docs.exists()
    counts = count_packs(docs)
    print("CONTAINER counts", counts, flush=True)
    assert counts["folders"] >= 4
    assert counts["docx"] == counts["folders"] * 4
    assert counts["pdf"] == counts["folders"] * 4

    labels = {
        "CNT-20-EUR-01": "20ft Euro pallet container loading",
        "CNT-20-STD-01": "20ft standard container loading",
        "CNT-40-IND-01": "40ft industrial container loading",
        "CNT-DE-HAM-01": "DE-HAM container loading config",
    }
    records = []
    for folder in sorted(docs.iterdir(), key=lambda p: p.name):
        if folder.is_dir():
            records.append(
                {"key": folder.name, "label": labels.get(folder.name, folder.name)}
            )

    eng = build_document_engine(
        delivery_root=final,
        engine_filename="INCI_AKU_PPWR_CONTAINER_ENGINE_Rev00.xlsx",
        title="İNCI AKÜ PPWR — CONTAINER LOADING ENGINE Rev00",
        docs_subdir="01_CONFIGS",
        records=records,
        extra_home={
            "PUBLISH DATE": DATE,
            "SIGNATORY": "Numan Alver — Operations Director",
            "QMS": "TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023",
            "MODEL": "1 Config = 1 set (DOCX+PDF)",
        },
    )
    link = verify_links(final, "01_CONFIGS", [r["key"] for r in records])
    print("CONTAINER link verify", link, flush=True)
    assert link["missing"] == 0

    print("CONTAINER zipping…", flush=True)
    digest = zip_tree(final, OUT / f"{final.name}.zip")
    return {
        "GATE": "PASS",
        "folders": counts["folders"],
        "docx": counts["docx"],
        "pdf": counts["pdf"],
        "links_ok": link["ok"],
        "engine": str(eng),
        "final": str(final),
        "zip": str(OUT / f"{final.name}.zip"),
        "sha256": digest,
    }


def verify_existing(name: str, final: Path, docs_subdir: str, expected_folders: int) -> dict:
    docs = final / docs_subdir
    counts = count_packs(docs)
    keys = [p.name for p in sorted(docs.iterdir()) if p.is_dir()]
    # rebuild engine from existing folders to guarantee link quality
    records = [{"key": k, "label": k} for k in keys]
    # try enrich industrial/component labels lightly
    if name == "INDUSTRIAL":
        master = final / "00_CONTROL" / "INCI_AKU_PPWR_INDUSTRIAL_MASTER_FROM_EXCEL_Rev00.xlsx"
        if not master.exists():
            master = OUT / "INCI_AKU_PPWR_INDUSTRIAL_MASTER_FROM_EXCEL_Rev00.xlsx"
        if master.exists():
            wb = load_workbook(master, data_only=True)
            ws = wb["PRODUCT_MASTER"]
            h = [c.value for c in next(ws.iter_rows(1, 1))]
            hi = {x: i for i, x in enumerate(h)}
            m = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[hi["Product Code"]] or "").strip()
                m[code] = {
                    "label": str(row[hi["Technical Description"]] or ""),
                    "voltage": str(row[hi["Voltage"]] or ""),
                }
            wb.close()
            records = [
                {
                    "key": k,
                    "label": m.get(k, {}).get("label") or k,
                    "voltage": m.get(k, {}).get("voltage") or "",
                }
                for k in keys
            ]
            eng = build_document_engine(
                delivery_root=final,
                engine_filename="INCI_AKU_PPWR_INDUSTRIAL_ENGINE_Rev00.xlsx",
                title="İNCI AKÜ PPWR — INDUSTRIAL INDIVIDUAL ENGINE Rev00",
                docs_subdir=docs_subdir,
                records=records,
                extra_home={
                    "PUBLISH DATE": DATE,
                    "SIGNATORY": "Numan Alver — Operations Director",
                    "QMS": "TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023",
                },
                extra_field="voltage",
            )
        else:
            eng = build_document_engine(
                delivery_root=final,
                engine_filename="INCI_AKU_PPWR_INDUSTRIAL_ENGINE_Rev00.xlsx",
                title="İNCI AKÜ PPWR — INDUSTRIAL INDIVIDUAL ENGINE Rev00",
                docs_subdir=docs_subdir,
                records=records,
            )
    elif name == "COMPONENT":
        labels = {
            "CMP-TEK-SIRA": "Komponent Ambalaj — Tek Sıra",
            "CMP-CIFT-SIRA": "Komponent Ambalaj — Çift Sıra",
        }
        records = [{"key": k, "label": labels.get(k, k)} for k in keys]
        eng = build_document_engine(
            delivery_root=final,
            engine_filename="INCI_AKU_PPWR_COMPONENT_SPARE_ENGINE_Rev00.xlsx",
            title="İNCI AKÜ PPWR — COMPONENT / SPARE ENGINE Rev00",
            docs_subdir=docs_subdir,
            records=records,
            extra_home={
                "PUBLISH DATE": DATE,
                "SIGNATORY": "Numan Alver — Operations Director",
                "QMS": "TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023",
            },
        )
    else:
        eng = Path("?")

    link = verify_links(final, docs_subdir, keys)
    assert counts["folders"] == expected_folders
    assert counts["docx"] == expected_folders * 4
    assert counts["pdf"] == expected_folders * 4
    assert link["missing"] == 0

    zip_path = OUT / f"{final.name}.zip"
    sha_path = OUT / f"{final.name}_SHA256.txt"
    if not zip_path.exists() or not sha_path.exists():
        print(f"{name} re-zip…", flush=True)
        digest = zip_tree(final, zip_path)
    else:
        # refresh zip after engine rebuild
        print(f"{name} refresh zip after engine…", flush=True)
        digest = zip_tree(final, zip_path)

    return {
        "GATE": "PASS",
        "folders": counts["folders"],
        "docx": counts["docx"],
        "pdf": counts["pdf"],
        "links_ok": link["ok"],
        "engine": str(eng),
        "final": str(final),
        "zip": str(zip_path),
        "sha256": digest,
    }


def main() -> int:
    report: dict = {}
    print("=== 1) STARTER FINAL ===", flush=True)
    report["STARTER"] = finalize_starter()

    print("=== 2) CONTAINER FINAL ===", flush=True)
    report["CONTAINER"] = finalize_container()

    print("=== 3) INDUSTRIAL VERIFY/REFRESH ===", flush=True)
    report["INDUSTRIAL"] = verify_existing(
        "INDUSTRIAL", OUT / "02_INDUSTRIAL_DELIVERY_REV00", "01_PRODUCTS", 2736
    )

    print("=== 4) COMPONENT VERIFY/REFRESH ===", flush=True)
    report["COMPONENT"] = verify_existing(
        "COMPONENT", OUT / "04_COMPONENT_SPARE_DELIVERY_REV00", "01_VARIANTS", 2
    )

    overall = all(report[k]["GATE"] == "PASS" for k in report)
    report["OVERALL"] = "PASS" if overall else "FAIL"
    REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"OVERALL": report["OVERALL"]}, indent=2), flush=True)
    for k in ("STARTER", "INDUSTRIAL", "CONTAINER", "COMPONENT"):
        r = report[k]
        print(
            f"{k}: folders={r['folders']} docx={r['docx']} pdf={r['pdf']} "
            f"links={r['links_ok']} sha={r['sha256'][:16]}…",
            flush=True,
        )
    print("QA", REPORT, flush=True)
    return 0 if overall else 1


if __name__ == "__main__":
    raise SystemExit(main())
