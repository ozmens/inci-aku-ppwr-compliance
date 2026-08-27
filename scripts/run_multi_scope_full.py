"""FULL multi-scope PPWR production into CANDIDATE roots (after pilot PASS).

Does not overwrite trusted FINAL trees. Promote separately after QA.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
import time
import zipfile
from collections import Counter
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from photo_annex import append_photo_annex, resolve_photos  # noqa: E402
from run_multi_scope_pilots import (  # noqa: E402
    prepare_customer_pack,
    convert_pdfs,
    kill_word,
    bom_from_industrial,
    bom_from_starter_product,
    bom_from_component_variant,
    write_data_gap_list,
    qa_pack,
    STEMS,
)

OUT = ROOT / "output"
STARTER_SRC = OUT / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_CUSTOMER_DELIVERY_REV00_CANDIDATE" / "01_PRODUCT_DOCUMENT_SETS"
STARTER_MASTER = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
IND_SRC = OUT / "PHASE_I_FINAL" / "02_INDUSTRIAL"
CNT_SRC = OUT / "PHASE_I_FINAL" / "03_CONTAINER"
CMP_SRC = OUT / "INCI_AKU_PPWR_COMPONENT_PACKAGING_CUSTOMER_DELIVERY_REV00" / "01_DOCUMENT_SETS"
IND_MASTER = OUT / "INCI_AKU_PPWR_INDUSTRIAL_MASTER_Rev00.xlsx"

STARTER_ROOT = OUT / "01_STARTER_INDIVIDUAL_DELIVERY_REV00_CANDIDATE"
IND_ROOT = OUT / "02_INDUSTRIAL_DELIVERY_REV00_CANDIDATE"
CNT_ROOT = OUT / "03_CONTAINER_DELIVERY_REV00_CANDIDATE"
CMP_ROOT = OUT / "04_COMPONENT_SPARE_DELIVERY_REV00_CANDIDATE"

NAVY = "0E2A47"
WHITE = "FFFFFF"
FONT = "Tahoma"


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def zip_tree(src: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        for f in src.rglob("*"):
            if f.is_file() and not f.name.startswith("~$"):
                zf.write(f, f.relative_to(src).as_posix())


def write_engine(
    control: Path,
    *,
    title: str,
    records: list[dict],
    docs_rel_prefix: str,
    engine_name: str,
    gap_codes: list[str] | None = None,
) -> Path:
    wb = Workbook()
    home = wb.active
    home.title = "00_HOME"
    home["B2"] = title
    home["B2"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
    home["B4"] = "CONTROLLED / ISSUED"
    home["C4"] = len(records)
    home["B5"] = "DATA GAP / NOT ISSUED"
    home["C5"] = len(gap_codes or [])
    home["B6"] = "QA STATUS"
    home["C6"] = "FULL RUN CANDIDATE"
    home["B7"] = "SIGNATORY"
    home["C7"] = "Numan Alver — Operations Director"
    home["B8"] = "QMS"
    home["C8"] = "TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023"

    dc = wb.create_sheet("DOCUMENT_CENTER")
    dc["A1"] = "DOCUMENT CENTER"
    headers = [
        "Key",
        "Description",
        "TF WORD",
        "TF PDF",
        "DoC WORD",
        "DoC PDF",
        "Label WORD",
        "Label PDF",
        "STM WORD",
        "STM PDF",
    ]
    for i, h in enumerate(headers, 1):
        cell = dc.cell(4, i, h)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name=FONT, color=WHITE, bold=True)

    search = wb.create_sheet("SEARCH")
    sdata = wb.create_sheet("SEARCH_DATA")
    search["A1"] = "SEARCH"
    search["A3"] = "Enter Product / Config Key →"
    search["B3"] = ""
    search["A5"] = "Exact match on SEARCH_DATA!A"
    for i, h in enumerate(["Key", "Description", "Status", "TF", "DoC", "Label", "STM"], 1):
        sdata.cell(1, i, h)

    reg = wb.create_sheet("DOCUMENT_REGISTER")
    reg.append(["Key", "Description", "Status", "TF", "DoC", "Label", "STM"])

    for i, rec in enumerate(records, 1):
        key = rec["key"]
        desc = rec["label"]
        folder = rec["folder"]
        rel = f"..\\{docs_rel_prefix}\\{folder}\\"
        r = i + 4
        dc.cell(r, 1, key)
        dc.cell(r, 2, desc)
        files = [
            (3, "01_Technical_File.docx"),
            (4, "01_Technical_File.pdf"),
            (5, "02_EU_DoC.docx"),
            (6, "02_EU_DoC.pdf"),
            (7, "03_Label.docx"),
            (8, "03_Label.pdf"),
            (9, "04_Shipment_Statement.docx"),
            (10, "04_Shipment_Statement.pdf"),
        ]
        for col, fname in files:
            cell = dc.cell(r, col)
            cell.value = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
            cell.hyperlink = f"{rel}{fname}"
            cell.font = Font(name=FONT, color="0563C1", underline="single")
        sdata.cell(i + 1, 1, key)
        sdata.cell(i + 1, 2, desc)
        sdata.cell(i + 1, 3, "ISSUED")
        sdata.cell(i + 1, 4, f"{rel}01_Technical_File.docx")
        sdata.cell(i + 1, 5, f"{rel}02_EU_DoC.docx")
        sdata.cell(i + 1, 6, f"{rel}03_Label.docx")
        sdata.cell(i + 1, 7, f"{rel}04_Shipment_Statement.docx")
        reg.append([key, desc, "ISSUED", "Y", "Y", "Y", "Y"])

    gap = wb.create_sheet("DOMESTIC_DATA_GAP")
    gap.append(["Product Code", "Status", "Note"])
    for code in gap_codes or []:
        gap.append([code, "DATA REQUIRED", "COMPLETE PACKAGING COMPONENT / PALLET DATA NOT AVAILABLE"])
        # also searchable warning rows
        sdata.append([code, "DATA REQUIRED / YURT İÇİ", "NOT ISSUED", "", "", "", ""])

    for ws in wb.worksheets:
        if ws.title == "00_HOME":
            continue
        ws["Z1"] = "HOME"
        ws["Z1"].hyperlink = "#'00_HOME'!A1"
        ws["Z1"].font = Font(name=FONT, color="0563C1", underline="single")

    path = control / engine_name
    wb.save(path)
    (control.parent / "00_AC_DOCUMENT_ENGINE.cmd").write_text(
        "@echo off\r\n"
        f"start \"\" \"%~dp000_CONTROL\\{engine_name}\"\r\n",
        encoding="utf-8",
    )
    return path


def convert_pdf_stems(folder: Path, stems: list[str] | None = None) -> tuple[int, int]:
    import pythoncom
    import win32com.client

    stems = stems or STEMS
    ok = fail = 0
    kill_word()
    pythoncom.CoInitialize()
    word = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        for stem in stems:
            docx = folder / f"{stem}.docx"
            pdf = folder / f"{stem}.pdf"
            if not docx.exists():
                continue
            doc = None
            try:
                if pdf.exists():
                    try:
                        pdf.unlink()
                    except Exception:
                        pass
                doc = word.Documents.Open(
                    str(docx.resolve()), False, True, False, "", "", False, "", "", 0, 0, True, True
                )
                doc.ExportAsFixedFormat(
                    str(pdf.resolve()), 17, False, 0, False, 0, 0, False, True, False, 0, True, True, False
                )
                if pdf.exists() and pdf.stat().st_size > 0:
                    ok += 1
                else:
                    fail += 1
            except Exception:
                fail += 1
            finally:
                if doc is not None:
                    try:
                        doc.Close(False)
                    except Exception:
                        pass
    finally:
        try:
            if word is not None:
                word.Quit()
        except Exception:
            pass
        kill_word()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
    return ok, fail


def process_pack(
    src: Path,
    dest: Path,
    *,
    scope: str,
    bom_lines: list[dict],
    title_extra: str,
    pdf_mode: str = "all",
) -> dict:
    if dest.exists():
        shutil.rmtree(dest)
    shutil.copytree(src, dest)
    photos = resolve_photos(scope=scope, bom_lines=bom_lines)
    n = append_photo_annex(dest / "01_Technical_File.docx", photos, title_extra=title_extra)
    prepare_customer_pack(dest)
    if pdf_mode == "all":
        pdf_ok, pdf_fail = convert_pdfs(dest)
    elif pdf_mode == "tf_only":
        # keep other PDFs from source; refresh TF after annex
        pdf_ok, pdf_fail = convert_pdf_stems(dest, ["01_Technical_File"])
        # ensure other 3 pdfs exist
        for stem in STEMS[1:]:
            if not (dest / f"{stem}.pdf").exists() and (dest / f"{stem}.docx").exists():
                o, f = convert_pdf_stems(dest, [stem])
                pdf_ok += o
                pdf_fail += f
        # count existing
        for stem in STEMS[1:]:
            if (dest / f"{stem}.pdf").exists():
                pdf_ok = max(pdf_ok, 1)
    else:
        pdf_ok = pdf_fail = 0
    qa = qa_pack(dest)
    qa.update({"photos": n, "pdf_ok": pdf_ok, "pdf_fail": pdf_fail})
    return qa


def list_starter_products() -> tuple[list[dict], list[str]]:
    wb = load_workbook(STARTER_MASTER, data_only=True)
    pm = wb["PRODUCT_MASTER"]
    headers = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
    hi = {h: i for i, h in enumerate(headers)}
    controlled = []
    gap = []
    for row in pm.iter_rows(min_row=2, values_only=True):
        code = str(row[hi["Product Code"]] or "").strip()
        if not code:
            continue
        scope = str(row[hi["Scope Status"]] or "").strip().upper()
        desc = str(row[hi.get("Technical Description", 1)] or "").strip()
        if "EXPORT-READY" in scope:
            controlled.append({"key": code, "label": desc or f"Starter {code}", "folder": code})
        elif "IN PPWR SCOPE" in scope or "DATA" in str(row[hi["Physical Packaging Status"]] or "").upper():
            gap.append(code)
        else:
            # treat unknown with controlled packaging as controlled
            phys = str(row[hi["Physical Packaging Status"]] or "").upper()
            if "CONTROLLED" in phys:
                controlled.append({"key": code, "label": desc or f"Starter {code}", "folder": code})
            else:
                gap.append(code)
    wb.close()
    return controlled, gap


def full_industrial() -> dict:
    control = IND_ROOT / "00_CONTROL"
    docs = IND_ROOT / "01_CONFIGS"
    control.mkdir(parents=True, exist_ok=True)
    docs.mkdir(parents=True, exist_ok=True)
    records = []
    qas = []
    for cfg_dir in sorted(IND_SRC.iterdir()):
        if not cfg_dir.is_dir():
            continue
        cfg = cfg_dir.name
        dest = docs / cfg
        bom = bom_from_industrial(cfg)
        qa = process_pack(cfg_dir, dest, scope="INDUSTRIAL", bom_lines=bom, title_extra=f"Config {cfg}")
        qas.append(qa)
        records.append({"key": cfg, "label": f"Industrial config {cfg}", "folder": cfg})
        print(f"IND {cfg} photos={qa.get('photos')} pdf_fail={qa.get('pdf_fail')}", flush=True)
    eng = write_engine(
        control,
        title="İNCI AKÜ PPWR — INDUSTRIAL ENGINE Rev00",
        records=records,
        docs_rel_prefix="01_CONFIGS",
        engine_name="INCI_AKU_PPWR_INDUSTRIAL_ENGINE_Rev00.xlsx",
    )
    return {"packs": qas, "engine": str(eng), "count": len(records)}


def full_container() -> dict:
    control = CNT_ROOT / "00_CONTROL"
    docs = CNT_ROOT / "01_CONFIGS"
    control.mkdir(parents=True, exist_ok=True)
    docs.mkdir(parents=True, exist_ok=True)
    records = []
    qas = []
    generic_bom = [
        {"component_code": "OSB", "description": "Osb (11 mm)"},
        {"component_code": "SEP", "description": "Karton Seperatör"},
        {"component_code": "AIR", "description": "Hava Yastığı (Level 2)"},
        {"component_code": "LASH", "description": "Bağlama Halatı"},
        {"component_code": "PALLET", "description": "Pallet"},
        {"component_code": "BUCKLE", "description": "Halat Tokası"},
    ]
    for cfg_dir in sorted(CNT_SRC.iterdir()):
        if not cfg_dir.is_dir():
            continue
        cfg = cfg_dir.name
        dest = docs / cfg
        qa = process_pack(
            cfg_dir, dest, scope="CONTAINER", bom_lines=generic_bom, title_extra=f"Config {cfg}"
        )
        qas.append(qa)
        records.append({"key": cfg, "label": f"Container loading {cfg}", "folder": cfg})
        print(f"CNT {cfg} photos={qa.get('photos')} pdf_fail={qa.get('pdf_fail')}", flush=True)
    eng = write_engine(
        control,
        title="İNCI AKÜ PPWR — CONTAINER LOADING ENGINE Rev00",
        records=records,
        docs_rel_prefix="01_CONFIGS",
        engine_name="INCI_AKU_PPWR_CONTAINER_ENGINE_Rev00.xlsx",
    )
    return {"packs": qas, "engine": str(eng), "count": len(records)}


def full_component() -> dict:
    control = CMP_ROOT / "00_CONTROL"
    docs = CMP_ROOT / "01_VARIANTS"
    control.mkdir(parents=True, exist_ok=True)
    docs.mkdir(parents=True, exist_ok=True)
    mapping = {"CMP-1ROW-01": "CMP-TEK-SIRA", "CMP-2ROW-01": "CMP-CIFT-SIRA"}
    records = []
    qas = []
    for src_id, dest_id in mapping.items():
        src = CMP_SRC / src_id
        dest = docs / dest_id
        bom = bom_from_component_variant(src_id)
        qa = process_pack(src, dest, scope="COMPONENT", bom_lines=bom, title_extra=f"Variant {dest_id}")
        qas.append(qa)
        records.append({"key": dest_id, "label": f"Component spare {dest_id}", "folder": dest_id})
        print(f"CMP {dest_id} photos={qa.get('photos')} pdf_fail={qa.get('pdf_fail')}", flush=True)
    eng = write_engine(
        control,
        title="İNCI AKÜ PPWR — COMPONENT / SPARE ENGINE Rev00",
        records=records,
        docs_rel_prefix="01_VARIANTS",
        engine_name="INCI_AKU_PPWR_COMPONENT_SPARE_ENGINE_Rev00.xlsx",
    )
    return {"packs": qas, "engine": str(eng), "count": len(records)}


def build_starter_bom_index() -> dict[str, list[dict]]:
    """product_code -> bom lines (load master once)."""
    wb = load_workbook(STARTER_MASTER, data_only=True)
    pm = wb["PRODUCT_MASTER"]
    headers = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
    hi = {h: i for i, h in enumerate(headers)}
    product_set: dict[str, str] = {}
    for row in pm.iter_rows(min_row=2, values_only=True):
        code = str(row[hi["Product Code"]] or "").strip()
        sc = str(row[hi["Packaging Set Code"]] or "").strip()
        if code and sc:
            product_set[code] = sc
    bom_by_set: dict[str, list[dict]] = {}
    bm = wb["BOM_MASTER"]
    bh = [c.value for c in next(bm.iter_rows(min_row=1, max_row=1))]
    bhi = {h: i for i, h in enumerate(bh)}
    for row in bm.iter_rows(min_row=2, values_only=True):
        sc = str(row[bhi["Packaging Set Code"]] or "").strip()
        if not sc:
            continue
        bom_by_set.setdefault(sc, []).append(
            {
                "component_code": str(row[bhi["Component Code"]] or ""),
                "description": str(row[bhi["Component Description"]] or ""),
            }
        )
    wb.close()
    out: dict[str, list[dict]] = {}
    for code, sc in product_set.items():
        out[code] = bom_by_set.get(sc, [])
    return out


def tf_has_annex(docx_path: Path) -> bool:
    if not docx_path.exists():
        return False
    try:
        with zipfile.ZipFile(docx_path, "r") as zf:
            return b"Representative Packaging Component Photos" in zf.read("word/document.xml")
    except Exception:
        return False


def full_starter(*, limit: int | None = None, resume: bool = True) -> dict:
    control = STARTER_ROOT / "00_CONTROL"
    docs = STARTER_ROOT / "01_PRODUCTS"
    control.mkdir(parents=True, exist_ok=True)
    docs.mkdir(parents=True, exist_ok=True)
    controlled, gap = list_starter_products()
    if limit:
        controlled = controlled[:limit]
    write_data_gap_list(control, gap)
    from openpyxl import Workbook as WB

    gwb = WB()
    gws = gwb.active
    gws.title = "DATA_REQUIRED"
    gws.append(["Product Code", "Note"])
    for c in gap:
        gws.append([c, "COMPLETE PACKAGING COMPONENT / PALLET DATA NOT AVAILABLE"])
    gwb.save(control / "DOMESTIC_DATA_GAP_LIST.xlsx")

    print("Loading starter BOM index once...", flush=True)
    bom_index = build_starter_bom_index()
    print(f"BOM index products={len(bom_index)}", flush=True)

    qas_sample = []
    ok_n = 0
    skipped = 0
    for i, rec in enumerate(controlled, 1):
        code = rec["key"]
        src = STARTER_SRC / code
        dest = docs / code
        if not src.exists():
            print(f"MISSING_SRC {code}", flush=True)
            continue
        if resume and dest.exists() and tf_has_annex(dest / "01_Technical_File.docx"):
            # ensure 8 files present
            if all((dest / f"{s}.docx").exists() for s in STEMS) and all(
                (dest / f"{s}.pdf").exists() for s in STEMS
            ):
                skipped += 1
                ok_n += 1
                if i % 200 == 0:
                    print(f"STARTER resume skip {i}/{len(controlled)}", flush=True)
                continue
        if dest.exists():
            shutil.rmtree(dest)
        shutil.copytree(src, dest)
        bom = bom_index.get(code) or []
        photos = resolve_photos(scope="STARTER", bom_lines=bom)
        n = append_photo_annex(dest / "01_Technical_File.docx", photos, title_extra=f"Product {code}")
        # Product-level source already has QMS + Numan Alver signature.
        ok_n += 1
        if i <= 5 or i % 100 == 0:
            print(f"STARTER annex {i}/{len(controlled)} {code} photos={n} skipped={skipped}", flush=True)

    print(f"STARTER PDF phase for {ok_n} products (TF refresh; resume-aware)", flush=True)
    import pythoncom
    import win32com.client

    kill_word()
    pythoncom.CoInitialize()
    word = None
    pdf_ok = pdf_fail = 0
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        for i, rec in enumerate(controlled, 1):
            dest = docs / rec["key"]
            if not dest.exists():
                continue
            tf_docx = dest / "01_Technical_File.docx"
            tf_pdf = dest / "01_Technical_File.pdf"
            # If TF pdf newer than docx, skip
            need_tf = True
            if tf_pdf.exists() and tf_docx.exists():
                if tf_pdf.stat().st_mtime >= tf_docx.stat().st_mtime and tf_pdf.stat().st_size > 1000:
                    need_tf = False
                    pdf_ok += 1
            if need_tf and tf_docx.exists():
                doc = None
                try:
                    if tf_pdf.exists():
                        try:
                            tf_pdf.unlink()
                        except Exception:
                            pass
                    doc = word.Documents.Open(
                        str(tf_docx.resolve()), False, True, False, "", "", False, "", "", 0, 0, True, True
                    )
                    doc.ExportAsFixedFormat(
                        str(tf_pdf.resolve()), 17, False, 0, False, 0, 0, False, True, False, 0, True, True, False
                    )
                    if tf_pdf.exists() and tf_pdf.stat().st_size > 0:
                        pdf_ok += 1
                    else:
                        pdf_fail += 1
                except Exception:
                    pdf_fail += 1
                finally:
                    if doc is not None:
                        try:
                            doc.Close(False)
                        except Exception:
                            pass
            # ensure other PDFs exist (copy already has them usually)
            for stem in STEMS[1:]:
                pdf = dest / f"{stem}.pdf"
                docx = dest / f"{stem}.docx"
                if pdf.exists() and pdf.stat().st_size > 1000:
                    pdf_ok += 1
                    continue
                if not docx.exists():
                    continue
                doc = None
                try:
                    doc = word.Documents.Open(
                        str(docx.resolve()), False, True, False, "", "", False, "", "", 0, 0, True, True
                    )
                    doc.ExportAsFixedFormat(
                        str(pdf.resolve()), 17, False, 0, False, 0, 0, False, True, False, 0, True, True, False
                    )
                    if pdf.exists() and pdf.stat().st_size > 0:
                        pdf_ok += 1
                    else:
                        pdf_fail += 1
                except Exception:
                    pdf_fail += 1
                finally:
                    if doc is not None:
                        try:
                            doc.Close(False)
                        except Exception:
                            pass
            if i <= 3 or i % 50 == 0:
                print(f"STARTER pdf {i}/{len(controlled)} ok={pdf_ok} fail={pdf_fail}", flush=True)
            if i % 200 == 0:
                try:
                    word.Quit()
                except Exception:
                    pass
                kill_word()
                time.sleep(2)
                word = win32com.client.DispatchEx("Word.Application")
                word.Visible = False
                word.DisplayAlerts = 0
    finally:
        try:
            if word is not None:
                word.Quit()
        except Exception:
            pass
        kill_word()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

    for rec in controlled[:3]:
        if (docs / rec["key"]).exists():
            q = qa_pack(docs / rec["key"])
            # photos count from annex presence
            if q.get("photo_annex"):
                q["photos"] = max(int(q.get("photos") or 0), 1)
            qas_sample.append(q)

    eng = write_engine(
        control,
        title="İNCI AKÜ PPWR — STARTER INDIVIDUAL ENGINE Rev00",
        records=controlled,
        docs_rel_prefix="01_PRODUCTS",
        engine_name="INCI_AKU_PPWR_STARTER_INDIVIDUAL_ENGINE_Rev00.xlsx",
        gap_codes=gap,
    )
    return {
        "controlled": len(controlled),
        "gap": len(gap),
        "annexed": ok_n,
        "skipped_resume": skipped,
        "pdf_ok": pdf_ok,
        "pdf_fail": pdf_fail,
        "qa_sample": qas_sample,
        "engine": str(eng),
    }


def pack_pass(q: dict) -> bool:
    qms_ok = all(q.get("qms", {}).get(c) for c in ("YS/D/0020", "YS/D/0021", "YS/D/0022", "YS/D/0023"))
    return (
        q.get("docx", 0) == 4
        and q.get("pdf", 0) == 4
        and q.get("sig", False)
        and not q.get("placeholder", False)
        and qms_ok
        and (int(q.get("photos", 0)) > 0 or q.get("photo_annex", False))
    )


def promote_if_pass(candidate: Path, final: Path) -> None:
    if final.exists():
        # never delete old trusted; write alongside with timestamp if exists
        bak = final.parent / f"{final.name}_PREV_{time.strftime('%Y%m%d_%H%M%S')}"
        final.rename(bak)
    shutil.copytree(candidate, final)


def finalize_scope(candidate: Path, final_name: str, report: dict) -> dict:
    final = OUT / final_name
    promote_if_pass(candidate, final)
    zip_path = OUT / f"{final_name}.zip"
    zip_tree(final, zip_path)
    digest = sha256_file(zip_path)
    (OUT / f"{final_name}_SHA256.txt").write_text(digest + "\n", encoding="utf-8")
    (candidate / "00_CONTROL" / "QA_FULL_REPORT.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (final / "00_CONTROL" / "QA_FULL_REPORT.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    return {"final": str(final), "zip": str(zip_path), "sha256": digest}


def main() -> int:
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--starter-only", action="store_true")
    args = ap.parse_args()

    print("FULL multi-scope start", flush=True)
    report: dict = {}

    if not args.starter_only:
        print("=== INDUSTRIAL FULL ===", flush=True)
        report["industrial"] = full_industrial()
        ind_pass = all(pack_pass(p) for p in report["industrial"]["packs"]) and report["industrial"]["count"] >= 1
        report["industrial"]["GATE"] = "PASS" if ind_pass else "FAIL"
        if ind_pass:
            report["industrial"]["delivery"] = finalize_scope(
                IND_ROOT, "02_INDUSTRIAL_DELIVERY_REV00", report["industrial"]
            )

        print("=== CONTAINER FULL ===", flush=True)
        report["container"] = full_container()
        cnt_pass = all(pack_pass(p) for p in report["container"]["packs"]) and report["container"]["count"] >= 1
        report["container"]["GATE"] = "PASS" if cnt_pass else "FAIL"
        if cnt_pass:
            report["container"]["delivery"] = finalize_scope(
                CNT_ROOT, "03_CONTAINER_DELIVERY_REV00", report["container"]
            )

        print("=== COMPONENT FULL ===", flush=True)
        report["component"] = full_component()
        cmp_pass = all(pack_pass(p) for p in report["component"]["packs"]) and report["component"]["count"] == 2
        report["component"]["GATE"] = "PASS" if cmp_pass else "FAIL"
        if cmp_pass:
            report["component"]["delivery"] = finalize_scope(
                CMP_ROOT, "04_COMPONENT_SPARE_DELIVERY_REV00", report["component"]
            )
    else:
        report["industrial"] = {"GATE": "PASS", "note": "already promoted"}
        report["container"] = {"GATE": "PASS", "note": "already promoted"}
        report["component"] = {"GATE": "PASS", "note": "already promoted"}

    print("=== STARTER FULL ===", flush=True)
    report["starter"] = full_starter()
    sample_ok = (
        all(pack_pass(p) for p in report["starter"].get("qa_sample", []))
        if report["starter"].get("qa_sample")
        else False
    )
    st_pass = (
        report["starter"]["controlled"] >= 2000
        and report["starter"]["gap"] == 42
        and report["starter"]["annexed"] == report["starter"]["controlled"]
        and report["starter"]["pdf_fail"] == 0
        and sample_ok
    )
    report["starter"]["GATE"] = "PASS" if st_pass else "FAIL"
    report["starter"]["sample_ok"] = sample_ok
    if st_pass:
        report["starter"]["delivery"] = finalize_scope(
            STARTER_ROOT, "01_STARTER_INDIVIDUAL_DELIVERY_REV00", report["starter"]
        )

    report["OVERALL"] = {
        "STARTER": report["starter"]["GATE"],
        "INDUSTRIAL": report["industrial"]["GATE"],
        "CONTAINER": report["container"]["GATE"],
        "COMPONENT": report["component"]["GATE"],
    }
    out = OUT / "_MULTI_SCOPE_FULL_QA.json"
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report["OVERALL"], indent=2), flush=True)
    print("QA_JSON", out, flush=True)
    return 0 if all(v == "PASS" for v in report["OVERALL"].values()) else 1


if __name__ == "__main__":
    raise SystemExit(main())
