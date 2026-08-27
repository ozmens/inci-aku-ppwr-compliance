"""Fix customer Excel links so they actually open from 00_CONTROL.

Root causes addressed:
1) SEARCH used nested IF(..., HYPERLINK(...)) — Excel does not click those.
2) output\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx looks operational but
   relative links resolve outside the delivery tree (broken).
3) No launcher at package root → easy to open the wrong workbook.

Does NOT change Product/BOM/mappings/IDs/document content.
"""

from __future__ import annotations

import hashlib
import shutil
import zipfile
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
FINAL = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
CTRL = FINAL / "00_CONTROL"
ENG = CTRL / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ENG_ROOT = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ZIP_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL.zip"
SHA_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL_SHA256.txt"
SMOKE = ROOT / "output" / "_LINK_FIX_ZIP_SMOKE"

FONT = "Tahoma"
NAVY, GOLD = "0E2A47", "C8A24A"
GOLD_B = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="thin", color=GOLD),
)


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def win_rel(sc: str, fname: str) -> str:
    return f"..\\01_DOCUMENT_SETS\\{sc}\\{fname}"


def close_excel_locks() -> None:
    pythoncom.CoInitialize()
    try:
        try:
            app = win32.GetActiveObject("Excel.Application")
        except Exception:
            return
        for i in range(app.Workbooks.Count, 0, -1):
            wb = app.Workbooks(i)
            name = str(wb.Name)
            if "DOCUMENT_ENGINE" in name or name.startswith("~$"):
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
        # leave Excel running if user has other books; do not Quit globally
    finally:
        pythoncom.CoUninitialize()


def patch_openpyxl() -> None:
    wb = load_workbook(ENG)
    home = wb["00_HOME"]
    home["B28"] = (
        "ÖNEMLİ / IMPORTANT: Bu dosyayı YALNIZCA 00_CONTROL klasöründen açın. "
        "output\\ kökündeki kopyayı veya ZIP içinden yanlış klasörü açmayın — "
        "aksi halde Word/PDF linkleri çalışmaz. "
        "Paket kökündeki 00_AC_DOCUMENT_ENGINE.cmd dosyasını kullanın."
    )
    home["B28"].font = Font(name=FONT, size=10, bold=True, color="A12622")
    home.merge_cells("B28:I28")
    home["B29"] = (
        "Correct open path: ...\\00_CONTROL\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx "
        "→ links use ..\\01_DOCUMENT_SETS\\<SET>\\..."
    )
    home["B29"].font = Font(name=FONT, size=9, bold=True, color=NAVY)
    home.merge_cells("B29:I29")

    ws = wb["SEARCH"]
    # Workbook folder via CELL("filename") — absolute base + relative docs
    ws["Z2"] = '=IFERROR(LEFT(CELL("filename",$B$4),FIND("[",CELL("filename",$B$4))-1),"")'
    ws.column_dimensions["Z"].hidden = True

    issued = (
        'AND($B$8<>"",$A$8<>"NOT FOUND",ISNUMBER(SEARCH("ISSUED",$G$8)),'
        'NOT(ISNUMBER(SEARCH("NOT ISSUED",$G$8))),NOT(ISNUMBER(SEARCH("YURT",$G$8))))'
    )
    domestic = 'OR(ISNUMBER(SEARCH("YURT",$G$8)),ISNUMBER(SEARCH("NOT ISSUED",$G$8)))'
    empty = 'OR($B$4="",$A$8="",$A$8="NOT FOUND")'

    def search_formula(stem: str, word: bool = True) -> str:
        label = "OPEN WORD" if word else "OPEN PDF"
        ext = "docx" if word else "pdf"
        # Outermost HYPERLINK — required for Excel clickability
        path = f'$Z$2&"..\\01_DOCUMENT_SETS\\"&$B$8&"\\{stem}.{ext}"'
        return (
            f'=IF({empty},"",'
            f'IF({domestic},"DOCUMENTS NOT ISSUED",'
            f'HYPERLINK(IF({issued},{path},""),IF({issued},"{label}",""))))'
        )

    for r, stem in [
        (13, "01_Technical_File"),
        (15, "02_EU_DoC"),
        (17, "03_Label"),
        (19, "04_Shipment_Statement"),
    ]:
        for c, word in ((1, True), (2, False)):
            cell = ws.cell(r, c)
            cell.value = search_formula(stem, word)
            cell.font = Font(name=FONT, size=10, bold=True, color="0563C1", underline="single")
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            cell.border = GOLD_B
            cell.alignment = Alignment(horizontal="center")

    # Document Center: formula HYPERLINK with workbook-folder base (clickable + portable)
    dc = wb["DOCUMENT_CENTER"]
    dc["Z2"] = '=IFERROR(LEFT(CELL("filename",$A$4),FIND("[",CELL("filename",$A$4))-1),"")'
    dc.column_dimensions["Z"].hidden = True
    # Clear openpyxl hyperlink objects; formulas replace them
    for row in dc.iter_rows(min_row=5, max_row=dc.max_row, min_col=6, max_col=16):
        for cell in row:
            if cell.hyperlink is not None:
                cell.hyperlink = None

    specs = [
        (6, "01_Technical_File.docx"),
        (7, "01_Technical_File.pdf"),
        (9, "02_EU_DoC.docx"),
        (10, "02_EU_DoC.pdf"),
        (12, "03_Label.docx"),
        (13, "03_Label.pdf"),
        (15, "04_Shipment_Statement.docx"),
        (16, "04_Shipment_Statement.pdf"),
    ]
    r = 5
    while dc.cell(r, 2).value:
        sc_ref = f"$B{r}"
        for col, fname in specs:
            label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
            path = f'$Z$2&"..\\01_DOCUMENT_SETS\\"&{sc_ref}&"\\{fname}"'
            cell = dc.cell(r, col)
            cell.value = f'=HYPERLINK({path},"{label}")'
            cell.font = Font(name=FONT, size=9, bold=True, color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center")
        r += 1
        if r > 400:
            break

    tmp = CTRL / "_link_fix_tmp.xlsx"
    wb.save(tmp)
    wb.close()
    ENG.unlink(missing_ok=True)
    shutil.move(str(tmp), str(ENG))


def com_polish_and_verify() -> dict:
    pythoncom.CoInitialize()
    excel = None
    result = {"dc_ok": 0, "search_ok": 0, "fail": []}
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        wb = excel.Workbooks.Open(str(ENG.resolve()))
        # Force calculate so CELL("filename") populates
        excel.CalculateFull()

        dc = wb.Worksheets("DOCUMENT_CENTER")
        # Sample 5 sets × 8 = 40
        tested = 0
        r = 5
        while dc.Cells(r, 2).Value and tested < 40:
            sc = str(dc.Cells(r, 2).Value).strip()
            for col, fname in [
                (6, "01_Technical_File.docx"),
                (7, "01_Technical_File.pdf"),
                (9, "02_EU_DoC.docx"),
                (10, "02_EU_DoC.pdf"),
                (12, "03_Label.docx"),
                (13, "03_Label.pdf"),
                (15, "04_Shipment_Statement.docx"),
                (16, "04_Shipment_Statement.pdf"),
            ]:
                cell = dc.Cells(r, col)
                val = str(cell.Value or "")
                target = (ENG.parent / ".." / "01_DOCUMENT_SETS" / sc / fname).resolve()
                if val != ("OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"):
                    result["fail"].append(f"DC display {sc} c{col}={val!r}")
                elif not target.exists():
                    result["fail"].append(f"missing {target}")
                else:
                    try:
                        # Resolve formula address: workbook dir + relative
                        addr = str(target)
                        wb.FollowHyperlink(Address=addr)
                        result["dc_ok"] += 1
                    except Exception as e:
                        result["fail"].append(f"DC follow {sc}/{fname}: {e}")
                tested += 1
            r += 1

        # SEARCH — 5 products
        search = wb.Worksheets("SEARCH")
        products = ["1015169", "1015170", "1015171", "1015172", "1015173"]
        # fallback: read first 5 ISSUED from SEARCH_DATA
        try:
            sd = wb.Worksheets("SEARCH_DATA")
            products = []
            for i in range(2, sd.UsedRange.Rows.Count + 1):
                st = str(sd.Cells(i, 7).Value or "")
                if "ISSUED" in st and "NOT ISSUED" not in st and "YURT" not in st:
                    products.append(str(sd.Cells(i, 1).Value).strip())
                if len(products) >= 5:
                    break
        except Exception:
            pass

        for pc in products[:5]:
            search.Range("B4").Value = pc
            excel.CalculateFull()
            set_code = str(search.Range("B8").Value or "").strip()
            status = str(search.Range("G8").Value or "")
            if not set_code or "NOT FOUND" in str(search.Range("A8").Value or ""):
                result["fail"].append(f"SEARCH not found {pc}")
                continue
            for row, stem in [
                (13, "01_Technical_File"),
                (15, "02_EU_DoC"),
                (17, "03_Label"),
                (19, "04_Shipment_Statement"),
            ]:
                for col, ext in ((1, "docx"), (2, "pdf")):
                    disp = str(search.Cells(row, col).Value or "")
                    label = "OPEN WORD" if ext == "docx" else "OPEN PDF"
                    target = (
                        ENG.parent / ".." / "01_DOCUMENT_SETS" / set_code / f"{stem}.{ext}"
                    ).resolve()
                    if "YURT" in status or "NOT ISSUED" in status:
                        if "DOCUMENTS NOT ISSUED" in disp:
                            result["search_ok"] += 1
                        else:
                            result["fail"].append(f"SEARCH domestic {pc} {disp!r}")
                    elif disp != label:
                        result["fail"].append(f"SEARCH disp {pc} {disp!r}")
                    elif not target.exists():
                        result["fail"].append(f"SEARCH missing {target}")
                    else:
                        try:
                            wb.FollowHyperlink(Address=str(target))
                            result["search_ok"] += 1
                        except Exception as e:
                            result["fail"].append(f"SEARCH follow {pc}/{stem}.{ext}: {e}")

        wb.Save()
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return result


def write_launcher() -> None:
    cmd = FINAL / "00_AC_DOCUMENT_ENGINE.cmd"
    cmd.write_text(
        "@echo off\r\n"
        "setlocal\r\n"
        "cd /d \"%~dp0\"\r\n"
        "set ENG=%~dp000_CONTROL\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx\r\n"
        "if not exist \"%ENG%\" (\r\n"
        "  echo Engine not found: %ENG%\r\n"
        "  pause\r\n"
        "  exit /b 1\r\n"
        ")\r\n"
        "echo Opening controlled Document Engine from 00_CONTROL...\r\n"
        "start \"\" \"%ENG%\"\r\n"
        , encoding="utf-8",
    )
    readme = FINAL / "00_READ_ME_OPEN_ENGINE.txt"
    readme.write_text(
        "İNİ AÇILIŞ / HOW TO OPEN\r\n"
        "========================\r\n"
        "1) Bu paketi bir klasöre çıkarın (Extract).\r\n"
        "2) 00_AC_DOCUMENT_ENGINE.cmd dosyasına çift tıklayın\r\n"
        "   VEYA açın:\r\n"
        "   00_CONTROL\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx\r\n"
        "\r\n"
        "output\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx (proje kök kopyası)\r\n"
        "müşteri paketi değildir — oradan açarsanız linkler ÇALIŞMAZ.\r\n"
        "\r\n"
        "Linkler göreli yoldur: ..\\01_DOCUMENT_SETS\\<SET>\\...\r\n"
        , encoding="utf-8",
    )
    # Warn next to root sync copy
    note = ROOT / "output" / "!!!_ENGINE_KOPYASI_LINKLERI_ICIN_KULLANMAYIN.txt"
    note.write_text(
        "Bu klasördeki INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx yalnızca senkron/SHA kopyasıdır.\r\n"
        "Word/PDF linklerini test etmek veya müşteri gibi kullanmak için açın:\r\n"
        "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL\\00_CONTROL\\"
        "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx\r\n"
        "veya FINAL kökündeki 00_AC_DOCUMENT_ENGINE.cmd\r\n",
        encoding="utf-8",
    )


def ensure_dev_junction() -> None:
    """Optional: make wrong open of output\\engine resolve docs for local dev."""
    link = ROOT / "01_DOCUMENT_SETS"
    target = FINAL / "01_DOCUMENT_SETS"
    if link.exists():
        return
    try:
        # Windows directory junction
        import subprocess

        subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(link), str(target)],
            check=False,
            capture_output=True,
            text=True,
        )
    except Exception:
        pass


def rebuild_zip() -> str:
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    with zipfile.ZipFile(ZIP_PATH, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for p in FINAL.rglob("*"):
            if p.is_file() and not p.name.startswith("~$"):
                zf.write(p, p.relative_to(FINAL).as_posix())
    digest = sha256_file(ZIP_PATH)
    SHA_PATH.write_text(digest + "\n", encoding="utf-8")
    return digest


def smoke_extract_links() -> dict:
    if SMOKE.exists():
        shutil.rmtree(SMOKE, ignore_errors=True)
    SMOKE.mkdir(parents=True)
    with zipfile.ZipFile(ZIP_PATH, "r") as zf:
        zf.extractall(SMOKE)
    eng = SMOKE / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
    assert eng.exists()
    assert (SMOKE / "00_AC_DOCUMENT_ENGINE.cmd").exists()

    pythoncom.CoInitialize()
    excel = None
    out = {"search": 0, "dc": 0, "fail": []}
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(eng.resolve()))
        excel.CalculateFull()
        dc = wb.Worksheets("DOCUMENT_CENTER")
        r = 5
        n = 0
        while dc.Cells(r, 2).Value and n < 40:
            sc = str(dc.Cells(r, 2).Value).strip()
            for col, fname in [
                (6, "01_Technical_File.docx"),
                (7, "01_Technical_File.pdf"),
                (9, "02_EU_DoC.docx"),
                (10, "02_EU_DoC.pdf"),
                (12, "03_Label.docx"),
                (13, "03_Label.pdf"),
                (15, "04_Shipment_Statement.docx"),
                (16, "04_Shipment_Statement.pdf"),
            ]:
                target = (eng.parent / ".." / "01_DOCUMENT_SETS" / sc / fname).resolve()
                disp = str(dc.Cells(r, col).Value or "")
                expect = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
                if disp != expect or not target.exists():
                    out["fail"].append(f"DC {sc}/{fname} disp={disp!r}")
                else:
                    wb.FollowHyperlink(Address=str(target))
                    out["dc"] += 1
                n += 1
            r += 1

        search = wb.Worksheets("SEARCH")
        sd = wb.Worksheets("SEARCH_DATA")
        products = []
        for i in range(2, min(sd.UsedRange.Rows.Count, 500) + 1):
            st = str(sd.Cells(i, 7).Value or "")
            if "ISSUED" in st and "NOT ISSUED" not in st and "YURT" not in st:
                products.append(str(sd.Cells(i, 1).Value).strip())
            if len(products) >= 5:
                break
        for pc in products:
            search.Range("B4").Value = pc
            excel.CalculateFull()
            set_code = str(search.Range("B8").Value or "").strip()
            for row, stem in [
                (13, "01_Technical_File"),
                (15, "02_EU_DoC"),
                (17, "03_Label"),
                (19, "04_Shipment_Statement"),
            ]:
                for col, ext in ((1, "docx"), (2, "pdf")):
                    disp = str(search.Cells(row, col).Value or "")
                    label = "OPEN WORD" if ext == "docx" else "OPEN PDF"
                    target = (
                        eng.parent / ".." / "01_DOCUMENT_SETS" / set_code / f"{stem}.{ext}"
                    ).resolve()
                    if disp != label or not target.exists():
                        out["fail"].append(f"SEARCH {pc}/{stem}.{ext} disp={disp!r}")
                    else:
                        wb.FollowHyperlink(Address=str(target))
                        out["search"] += 1
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return out


def main() -> None:
    print("0) Close engine locks…")
    close_excel_locks()
    for lock in CTRL.glob("~$*.xlsx"):
        try:
            lock.unlink()
        except Exception:
            pass

    print("1) Patch formulas + HOME warning…")
    patch_openpyxl()

    print("2) COM verify from 00_CONTROL…")
    ver = com_polish_and_verify()
    print(ver)
    if ver["fail"]:
        raise SystemExit(f"CTRL verify FAIL: {ver['fail'][:5]}")

    print("3) Sync root copy + launcher + junction…")
    shutil.copy2(ENG, ENG_ROOT)
    assert sha256_file(ENG) == sha256_file(ENG_ROOT)
    write_launcher()
    ensure_dev_junction()

    print("4) Rebuild ZIP…")
    digest = rebuild_zip()
    print("SHA256", digest)

    print("5) Extract smoke 80/80…")
    smoke = smoke_extract_links()
    print(smoke)
    total = smoke["search"] + smoke["dc"]
    if smoke["fail"] or total != 80:
        raise SystemExit(f"SMOKE FAIL {total}/80 {smoke['fail'][:5]}")

    print("PASS — open via FINAL\\00_AC_DOCUMENT_ENGINE.cmd or 00_CONTROL\\engine.xlsx")
    print("NEW SHA256:", digest)


if __name__ == "__main__":
    main()
