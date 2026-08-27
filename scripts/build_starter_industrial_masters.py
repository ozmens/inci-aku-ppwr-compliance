"""Build STARTER + INDUSTRIAL master workbooks (no Word regeneration).

Reconciles former domestic Source Configurations against existing 240
Starter Packaging Sets by physical BOM identity.
"""

from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
OUT = ROOT / "output"
GOLDEN = (
    ROOT
    / "input"
    / "production"
    / "INCI_AKU_PPWR_Final_Configuration_Register_Rev00_GOLDEN_VARIANTS_FINAL.xlsx"
)
L2 = (
    ROOT
    / "input"
    / "production"
    / "INCI_AKU_PPWR_PIMS_Data_Package_Rev00_OEM_ART5_FIXED.xlsx"
)
PHASE_I = OUT / "PHASE_I_FINAL"
STARTER_XLSX = OUT / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
INDUSTRIAL_XLSX = OUT / "INCI_AKU_PPWR_INDUSTRIAL_MASTER_Rev00.xlsx"
QA_MD = OUT / "STARTER_INDUSTRIAL_SPLIT_QA.md"
QA_JSON = OUT / "STARTER_INDUSTRIAL_SPLIT_QA.json"

NAVY = "0E2A47"
GOLD = "C8A24A"
WHITE = "FFFFFF"
INK = "1C2430"
BAND = "F3F6F9"
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)
SCOPE_STATUS = "EXPORT-READY STARTER SCOPE"
DOC_TYPE_NUMBERS = {
    "TF": "YS/D/0020",
    "DOC": "YS/D/0021",
    "LABEL": "YS/D/0022",
    "STM": "YS/D/0023",
}


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def _font(size=9, bold=False, color=INK) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color)


def _headers(ws) -> dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(h): i for i, h in enumerate(row) if h is not None}


def _f(v) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(v)
    except Exception:
        return None


def _round(v: float | None, n: int = 6) -> float:
    if v is None:
        return 0.0
    return round(float(v), n)


@dataclass(frozen=True)
class BomLine:
    component_code: str
    quantity: float
    uom: str
    unit_weight_kg: float

    def key(self) -> tuple:
        return (
            self.component_code.strip().upper(),
            _round(self.quantity, 6),
            (self.uom or "").strip().upper(),
            _round(self.unit_weight_kg, 6),
        )


def bom_signature(nominal_qty: float | None, lines: list[BomLine]) -> tuple:
    tare = _round(sum(_round(l.quantity * l.unit_weight_kg, 8) for l in lines), 6)
    return (
        _round(nominal_qty, 6) if nominal_qty is not None else None,
        tare,
        tuple(sorted(l.key() for l in lines)),
    )


@dataclass
class ConfigRec:
    packaging_set_code: str
    final_configuration_id: str
    source_configuration_id: str
    family: str
    nominal_qty: float | None
    packaging_tare_kg: float | None
    description: str
    variant_basis_tr: str
    variant_basis_en: str
    tf_id: str
    doc_id: str
    label_id: str
    stm_id: str
    bom: list[BomLine] = field(default_factory=list)
    origin: str = "EXISTING_240"  # EXISTING_240 | REUSED_FOR_DOMESTIC | NEW_FROM_DOMESTIC
    status: str = SCOPE_STATUS
    signature: tuple | None = None


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def word_tf_fingerprint() -> dict[str, Any]:
    """Hash inventory of PHASE_I TF files — must remain unchanged."""
    files = sorted((PHASE_I / "01_STARTER").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "02_INDUSTRIAL").rglob("01_Technical_File.docx"))
    files += sorted((PHASE_I / "03_CONTAINER").rglob("01_Technical_File.docx"))
    files = [p for p in files if not p.name.startswith("~$")]
    digests = []
    h = hashlib.sha256()
    for p in files:
        d = sha256_file(p)
        digests.append(d)
        h.update(d.encode())
        h.update(str(p.relative_to(PHASE_I)).encode())
    return {"count": len(files), "aggregate": h.hexdigest(), "sample": digests[:3]}


def load_golden_starter() -> tuple[dict[str, ConfigRec], dict[str, list[dict]], list[dict]]:
    wb = load_workbook(GOLDEN, data_only=True, read_only=True)
    # configs
    ws = wb["01_FINAL_CONFIG_MASTER"]
    idx = _headers(ws)
    configs: dict[str, ConfigRec] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fam = str(row[idx["Family"]] or "").upper()
        if fam != "STARTER":
            continue
        set_code = str(row[idx["Packaging Set Code"]])
        configs[set_code] = ConfigRec(
            packaging_set_code=set_code,
            final_configuration_id=str(row[idx["Final Configuration ID"]]),
            source_configuration_id=str(row[idx["Source Configuration ID"]]),
            family="STARTER",
            nominal_qty=_f(row[idx["Nominal Product Qty"]]),
            packaging_tare_kg=_f(row[idx["Packaging Mass kg"]]),
            description=str(row[idx["Configuration Name"]] or ""),
            variant_basis_tr="",
            variant_basis_en=str(row[idx["Variant Basis (EN)"]] or ""),
            tf_id=str(row[idx["Technical File ID"]] or ""),
            doc_id=str(row[idx["EU DoC ID"]] or ""),
            label_id=str(row[idx["Label ID"]] or ""),
            stm_id=str(row[idx["Statement ID"]] or ""),
            origin="EXISTING_240",
        )

    # TR variant basis column
    ws = wb["01_FINAL_CONFIG_MASTER"]
    idx = _headers(ws)
    tr_key = next(
        (
            k
            for k in idx
            if ("Variant Basis" in k and "TR" in k)
            or ("Ay" in k and "Variant" in k)
        ),
        None,
    )
    if tr_key:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[idx["Family"]] or "").upper() != "STARTER":
                continue
            set_code = str(row[idx["Packaging Set Code"]])
            if set_code in configs:
                configs[set_code].variant_basis_tr = str(row[idx[tr_key]] or "")

    # BOM
    ws = wb["03_BOM_MASTER"]
    idx = _headers(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        set_code = str(row[idx["Packaging Set Code"]] or "")
        if set_code not in configs:
            continue
        code = str(row[idx["Component Code"]] or "").strip()
        if not code:
            continue
        configs[set_code].bom.append(
            BomLine(
                component_code=code,
                quantity=float(row[idx["Quantity"]] or 0),
                uom=str(row[idx["UOM"]] or ""),
                unit_weight_kg=float(row[idx["Unit Weight kg"]] or 0),
            )
        )
    for cfg in configs.values():
        cfg.signature = bom_signature(cfg.nominal_qty, cfg.bom)

    # products (all starter product map rows — all 2046 are starter)
    ws = wb["02_PRODUCT_MAP"]
    idx = _headers(ws)
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pc = row[idx["Product Code"]]
        if pc in (None, ""):
            continue
        products.append(
            {
                "product_code": str(pc).strip(),
                "product_description": str(row[idx["Product Description"]] or ""),
                "battery_type": str(row[idx["Battery Type"]] or ""),
                "customer_market": str(row[idx["Customer / Market"]] or ""),
                "nominal_qty": _f(row[idx["Nominal Qty"]]),
                "source_configuration_id": str(row[idx["Source Configuration ID"]] or ""),
                "final_set_code": (
                    str(row[idx["Final Set Code"]]).strip()
                    if row[idx["Final Set Code"]] not in (None, "")
                    else None
                ),
                "final_configuration_id": (
                    str(row[idx["Final Configuration ID"]]).strip()
                    if row[idx["Final Configuration ID"]] not in (None, "")
                    else None
                ),
                "legacy_status": str(row[idx["Status"]] or ""),
            }
        )

    # domestic source list
    ws = wb["04_DOMESTIC_ONLY"]
    idx = _headers(ws)
    domestic_meta = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[idx["Source Configuration ID"]] or "")
        if not sid:
            continue
        domestic_meta[sid] = {
            "configuration_name": str(row[idx["Configuration Name"]] or ""),
            "nominal_qty": _f(row[idx["Nominal Qty"]]),
            "product_count": row[idx["Product Count"]],
            "markets": str(row[idx["Markets"]] or ""),
            "legacy_status": str(row[idx["Status"]] or ""),
        }
    wb.close()
    return configs, domestic_meta, products


def load_l2_domestic_variant_boms(
    domestic_ids: set[str],
) -> tuple[dict[str, list[BomLine]], dict[str, dict], dict[str, list[str]], dict[str, str]]:
    """Return base BOM by config, variant meta, products by variant, variant->config."""
    wb = load_workbook(L2, data_only=True, read_only=True)

    # base components
    ws = wb["04_CONFIG_COMPONENTS"]
    idx = _headers(ws)
    base: dict[str, list[BomLine]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[idx["configuration_id"]] or "")
        if cid not in domestic_ids:
            continue
        code = str(row[idx["component_code"]] or "").strip()
        if not code:
            continue
        base[cid].append(
            BomLine(
                component_code=code,
                quantity=float(row[idx["quantity"]] or 0),
                uom=str(row[idx["uom"]] or ""),
                unit_weight_kg=float(row[idx["unit_weight_kg"]] or 0),
            )
        )

    # variants
    ws = wb["05_VARIANT_MASTER"]
    idx = _headers(ws)
    variants: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[idx["configuration_id"]] or "")
        if cid not in domestic_ids:
            continue
        vid = str(row[idx["variant_id"]])
        variants[vid] = {
            "configuration_id": cid,
            "variant_name": str(row[idx["variant_name"]] or ""),
            "product_count": row[idx["product_count"]],
        }

    # variant components
    ws = wb["06_VARIANT_COMPONENTS"]
    idx = _headers(ws)
    var_lines: dict[str, list[BomLine]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid = str(row[idx["variant_id"]] or "")
        if vid not in variants:
            continue
        code = str(row[idx["component_code"]] or "").strip()
        if not code:
            continue
        var_lines[vid].append(
            BomLine(
                component_code=code,
                quantity=float(row[idx["quantity"]] or 0),
                uom=str(row[idx["uom"]] or ""),
                unit_weight_kg=float(row[idx["unit_weight_kg"]] or 0),
            )
        )

    # product -> variant
    ws = wb["07_PRODUCT_MAP"]
    idx = _headers(ws)
    products_by_variant: dict[str, list[str]] = defaultdict(list)
    product_to_variant: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[idx["configuration_id"]] or "")
        if cid not in domestic_ids:
            continue
        pc = str(row[idx["product_code"]] or "").strip()
        vid = str(row[idx["variant_id"]] or "").strip()
        if not pc or not vid:
            continue
        products_by_variant[vid].append(pc)
        product_to_variant[pc] = vid

    # config master nominal
    ws = wb["03_CONFIGURATION_MASTER"]
    idx = _headers(ws)
    cfg_nom: dict[str, float | None] = {}
    cfg_name: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[idx["configuration_id"]] or "")
        if cid not in domestic_ids:
            continue
        cfg_nom[cid] = _f(row[idx["nominal_product_quantity"]])
        cfg_name[cid] = str(row[idx["configuration_name"]] or "")

    wb.close()

    # merge base+variant per variant_id
    merged: dict[str, list[BomLine]] = {}
    for vid, meta in variants.items():
        cid = meta["configuration_id"]
        by_code: dict[str, BomLine] = {}
        for line in base.get(cid, []):
            by_code[line.component_code.upper()] = line
        for line in var_lines.get(vid, []):
            by_code[line.component_code.upper()] = line  # variant overrides/adds
        merged[vid] = list(by_code.values())
        meta["nominal_qty"] = cfg_nom.get(cid)
        meta["configuration_name"] = cfg_name.get(cid, "")

    return merged, variants, products_by_variant, product_to_variant


def suggest_set_prefix(configuration_name: str, markets: str) -> str:
    """Derive ST-{qty}-{TYPE} prefix from configuration name."""
    name = configuration_name or ""
    m = re.search(r"(\d+)\s*-?\s*unit", name, re.I)
    qty = int(m.group(1)) if m else 0
    market_u = (markets or name).upper()
    if "EUR" in market_u or "EURO" in market_u:
        kind = "EUR"
    elif "90" in name or "STD90" in name.upper():
        kind = "STD90"
    else:
        kind = "STD"
    if qty <= 0:
        # fallback bucket
        return f"ST-DOM-{kind}"
    return f"ST-{qty:03d}-{kind}"


def allocate_set_code(prefix: str, used: set[str], prefix_max: dict[str, int]) -> str:
    n = prefix_max.get(prefix, 0) + 1
    while True:
        code = f"{prefix}-{n:02d}"
        if code not in used:
            prefix_max[prefix] = n
            used.add(code)
            return code
        n += 1


def make_doc_ids(set_code: str) -> dict[str, str]:
    return {
        "tf_id": f"IA-PPWR-TF-{set_code}-R00",
        "doc_id": f"IA-PPWR-DOC-{set_code}-R00",
        "label_id": f"IA-PPWR-LBL-{set_code}-R00",
        "stm_id": f"IA-PPWR-STM-{set_code}-R00",
    }


def style_header(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row, c, h)
        cell.font = _font(9, True, WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = HAIR
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    style_header(ws, headers)
    for r_i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(r_i + 2, c, v)
            cell.font = _font(9)
            cell.border = HAIR
            cell.fill = _fill(BAND) if r_i % 2 else _fill(WHITE)
            cell.alignment = Alignment(vertical="center")
    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(str(h)) + 2, 14), 42)


def linked_products_str(codes: list[str]) -> str:
    return "; ".join(sorted(set(codes), key=lambda x: (len(x), x)))


def build() -> dict[str, Any]:
    fp_before = word_tf_fingerprint()

    configs, domestic_meta, products = load_golden_starter()
    domestic_ids = set(domestic_meta)
    assert len(configs) == 240
    assert len(products) == 2046
    assert len(domestic_ids) == 46

    # existing signatures
    sig_to_set: dict[tuple, str] = {}
    for set_code, cfg in configs.items():
        assert cfg.signature is not None
        # first wins if collision among existing (should be unique)
        sig_to_set.setdefault(cfg.signature, set_code)

    used_codes = set(configs.keys())
    prefix_max: dict[str, int] = defaultdict(int)
    for code in used_codes:
        parts = code.rsplit("-", 1)
        if len(parts) == 2 and parts[1].isdigit():
            prefix_max[parts[0]] = max(prefix_max[parts[0]], int(parts[1]))

    merged_boms, variants, products_by_variant, product_to_variant = load_l2_domestic_variant_boms(
        domestic_ids
    )

    reuse_count = 0
    new_count = 0
    variant_resolutions: list[dict] = []
    # Map: product_code -> packaging_set_code
    product_set: dict[str, str] = {}
    # For OEM products already mapped
    for p in products:
        if p["final_set_code"]:
            product_set[p["product_code"]] = p["final_set_code"]

    # Evaluate each domestic variant as a physical configuration
    for vid, meta in sorted(variants.items(), key=lambda x: x[0]):
        cid = meta["configuration_id"]
        lines = merged_boms.get(vid, [])
        nom = meta.get("nominal_qty")
        if nom is None:
            nom = domestic_meta.get(cid, {}).get("nominal_qty")
        sig = bom_signature(nom, lines)
        pcs = products_by_variant.get(vid, [])
        dmeta = domestic_meta.get(cid, {})

        if not lines:
            # BOM missing — still assign a new pending set code so product stays in scope
            prefix = suggest_set_prefix(
                meta.get("configuration_name") or dmeta.get("configuration_name", ""),
                dmeta.get("markets", ""),
            )
            new_code = allocate_set_code(prefix, used_codes, prefix_max)
            ids = make_doc_ids(new_code)
            configs[new_code] = ConfigRec(
                packaging_set_code=new_code,
                final_configuration_id=f"IA-{new_code}",
                source_configuration_id=cid,
                family="STARTER",
                nominal_qty=nom,
                packaging_tare_kg=None,
                description=meta.get("configuration_name") or dmeta.get("configuration_name", ""),
                variant_basis_tr=meta.get("variant_name") or "",
                variant_basis_en=meta.get("variant_name") or "",
                tf_id=ids["tf_id"],
                doc_id=ids["doc_id"],
                label_id=ids["label_id"],
                stm_id=ids["stm_id"],
                bom=[],
                origin="NEW_FROM_DOMESTIC_BOM_PENDING",
                status=SCOPE_STATUS,
                signature=sig,
            )
            new_count += 1
            for pc in pcs:
                product_set[pc] = new_code
            variant_resolutions.append(
                {
                    "source_configuration_id": cid,
                    "variant_id": vid,
                    "action": "NEW_BOM_PENDING",
                    "packaging_set_code": new_code,
                    "product_count": len(pcs),
                }
            )
            continue

        existing = sig_to_set.get(sig)
        if existing:
            reuse_count += 1
            for pc in pcs:
                product_set[pc] = existing
            # mark origin note on existing config
            if configs[existing].origin == "EXISTING_240":
                configs[existing].origin = "EXISTING_240_REUSED_FOR_DOMESTIC"
            variant_resolutions.append(
                {
                    "source_configuration_id": cid,
                    "variant_id": vid,
                    "action": "REUSE",
                    "packaging_set_code": existing,
                    "product_count": len(pcs),
                }
            )
        else:
            prefix = suggest_set_prefix(
                meta.get("configuration_name") or dmeta.get("configuration_name", ""),
                dmeta.get("markets", ""),
            )
            new_code = allocate_set_code(prefix, used_codes, prefix_max)
            ids = make_doc_ids(new_code)
            tare = _round(sum(l.quantity * l.unit_weight_kg for l in lines), 6)
            configs[new_code] = ConfigRec(
                packaging_set_code=new_code,
                final_configuration_id=f"IA-{new_code}",
                source_configuration_id=cid,
                family="STARTER",
                nominal_qty=nom,
                packaging_tare_kg=tare,
                description=meta.get("configuration_name") or dmeta.get("configuration_name", ""),
                variant_basis_tr=meta.get("variant_name") or "",
                variant_basis_en=meta.get("variant_name") or "",
                tf_id=ids["tf_id"],
                doc_id=ids["doc_id"],
                label_id=ids["label_id"],
                stm_id=ids["stm_id"],
                bom=list(lines),
                origin="NEW_FROM_DOMESTIC",
                status=SCOPE_STATUS,
                signature=sig,
            )
            sig_to_set[sig] = new_code
            new_count += 1
            for pc in pcs:
                product_set[pc] = new_code
            variant_resolutions.append(
                {
                    "source_configuration_id": cid,
                    "variant_id": vid,
                    "action": "NEW",
                    "packaging_set_code": new_code,
                    "product_count": len(pcs),
                }
            )

    # Ensure every product has a set (fallback by source if L2 miss)
    unresolved = []
    for p in products:
        pc = p["product_code"]
        if pc in product_set:
            continue
        # try map via source only if single existing set for that source among OEM
        sid = p["source_configuration_id"]
        candidates = [
            c.packaging_set_code
            for c in configs.values()
            if c.source_configuration_id == sid and c.origin.startswith("EXISTING")
        ]
        if len(candidates) == 1:
            product_set[pc] = candidates[0]
        else:
            unresolved.append(pc)

    # products by set
    set_to_products: dict[str, list[str]] = defaultdict(list)
    for pc, sc in product_set.items():
        set_to_products[sc].append(pc)

    # Build STARTER workbook
    wb = Workbook()
    # HOME
    ws = wb.active
    ws.title = "00_HOME"
    ws["A1"] = "İNCI AKÜ PPWR — STARTER MASTER Rev.00"
    ws["A1"].font = _font(16, True, NAVY)
    ws["A3"] = "Export-ready Starter scope — all Product Codes included"
    ws["A5"] = f"Product Codes: {len(products)}"
    ws["A6"] = f"Packaging Sets: {len(configs)}"
    ws["A7"] = f"Former domestic Source Configurations evaluated: {len(domestic_ids)}"
    ws["A8"] = "Document type numbers: TF YS/D/0020 · DoC YS/D/0021 · Label YS/D/0022 · STM YS/D/0023"
    ws["A10"] = "NO Word regeneration in this phase."

    # PRODUCT_MASTER
    ws = wb.create_sheet("PRODUCT_MASTER")
    headers = [
        "Product Code",
        "Packaging Set Code",
        "Technical Description",
        "Customer / Market",
        "Battery Type",
        "Nominal Qty",
        "Source Configuration ID",
        "Final Configuration ID",
        "Packaging Tare kg",
        "Scope Status",
        "Legacy Market Status",
    ]
    rows = []
    prod_by_code = {p["product_code"]: p for p in products}
    for pc in sorted(prod_by_code):
        p = prod_by_code[pc]
        sc = product_set.get(pc, "")
        cfg = configs.get(sc)
        rows.append(
            [
                pc,
                sc,
                p["product_description"],
                p["customer_market"],
                p["battery_type"],
                p["nominal_qty"],
                p["source_configuration_id"],
                cfg.final_configuration_id if cfg else (p["final_configuration_id"] or ""),
                cfg.packaging_tare_kg if cfg else "",
                SCOPE_STATUS,
                p["legacy_status"],
            ]
        )
    write_table(ws, headers, rows)

    # CONFIG_MASTER
    ws = wb.create_sheet("CONFIG_MASTER")
    headers = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Final Configuration ID",
        "Product Count",
        "Nominal Qty",
        "Pallet / Packaging Description",
        "Variant Basis TR",
        "Variant Basis EN",
        "Packaging Tare kg",
        "BOM Line Count",
        "Status",
        "Origin",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "TF Type No",
        "DoC Type No",
        "Label Type No",
        "STM Type No",
    ]
    rows = []
    for sc in sorted(configs):
        cfg = configs[sc]
        pcs = set_to_products.get(sc, [])
        rows.append(
            [
                sc,
                linked_products_str(pcs),
                cfg.source_configuration_id,
                cfg.final_configuration_id,
                len(pcs),
                cfg.nominal_qty,
                cfg.description,
                cfg.variant_basis_tr,
                cfg.variant_basis_en,
                cfg.packaging_tare_kg,
                len(cfg.bom),
                cfg.status,
                cfg.origin,
                cfg.tf_id,
                cfg.doc_id,
                cfg.label_id,
                cfg.stm_id,
                DOC_TYPE_NUMBERS["TF"],
                DOC_TYPE_NUMBERS["DOC"],
                DOC_TYPE_NUMBERS["LABEL"],
                DOC_TYPE_NUMBERS["STM"],
            ]
        )
    write_table(ws, headers, rows)

    # BOM_MASTER
    ws = wb.create_sheet("BOM_MASTER")
    headers = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Component Code",
        "Component Description",
        "Quantity",
        "UOM",
        "Unit Weight",
        "Line Weight",
    ]
    # component descriptions from golden BOM if available — reload briefly
    gwb = load_workbook(GOLDEN, data_only=True, read_only=True)
    gbom = gwb["03_BOM_MASTER"]
    gidx = _headers(gbom)
    desc_map: dict[tuple[str, str], str] = {}
    for row in gbom.iter_rows(min_row=2, values_only=True):
        sc = str(row[gidx["Packaging Set Code"]] or "")
        cc = str(row[gidx["Component Code"]] or "")
        desc_map[(sc, cc)] = str(row[gidx["ERP Description"]] or "")
    gwb.close()

    rows = []
    for sc in sorted(configs):
        cfg = configs[sc]
        pcs = linked_products_str(set_to_products.get(sc, []))
        for line in cfg.bom:
            lw = _round(line.quantity * line.unit_weight_kg, 6)
            rows.append(
                [
                    sc,
                    pcs,
                    cfg.source_configuration_id,
                    line.component_code,
                    desc_map.get((sc, line.component_code), ""),
                    line.quantity,
                    line.uom,
                    line.unit_weight_kg,
                    lw,
                ]
            )
    write_table(ws, headers, rows)

    # DOCUMENT_SCOPE — product level
    ws = wb.create_sheet("DOCUMENT_SCOPE")
    headers = [
        "Product Code",
        "Packaging Set Code",
        "Source Configuration ID",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "TF Type No",
        "DoC Type No",
        "Label Type No",
        "STM Type No",
        "Word Status",
    ]
    rows = []
    for pc in sorted(prod_by_code):
        sc = product_set.get(pc, "")
        cfg = configs.get(sc)
        word_status = (
            "EXISTING_LINKED"
            if cfg and cfg.origin.startswith("EXISTING")
            else "PENDING_GENERATION"
        )
        rows.append(
            [
                pc,
                sc,
                prod_by_code[pc]["source_configuration_id"],
                cfg.tf_id if cfg else "",
                cfg.doc_id if cfg else "",
                cfg.label_id if cfg else "",
                cfg.stm_id if cfg else "",
                DOC_TYPE_NUMBERS["TF"],
                DOC_TYPE_NUMBERS["DOC"],
                DOC_TYPE_NUMBERS["LABEL"],
                DOC_TYPE_NUMBERS["STM"],
                word_status,
            ]
        )
    write_table(ws, headers, rows)

    # DOCUMENT_CENTER — configuration level with linked products
    ws = wb.create_sheet("DOCUMENT_CENTER")
    headers = [
        "Packaging Set Code",
        "Linked Product Codes",
        "Source Configuration ID",
        "Final Configuration ID",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "Word Status",
        "TF Type No",
        "DoC Type No",
        "Label Type No",
        "STM Type No",
    ]
    rows = []
    for sc in sorted(configs):
        cfg = configs[sc]
        word_status = (
            "EXISTING_LINKED"
            if cfg.origin.startswith("EXISTING")
            else "PENDING_GENERATION"
        )
        rows.append(
            [
                sc,
                linked_products_str(set_to_products.get(sc, [])),
                cfg.source_configuration_id,
                cfg.final_configuration_id,
                cfg.tf_id,
                cfg.doc_id,
                cfg.label_id,
                cfg.stm_id,
                word_status,
                DOC_TYPE_NUMBERS["TF"],
                DOC_TYPE_NUMBERS["DOC"],
                DOC_TYPE_NUMBERS["LABEL"],
                DOC_TYPE_NUMBERS["STM"],
            ]
        )
    write_table(ws, headers, rows)

    # TECHNICAL_FILES / DoC / LABELS / STATEMENTS indexes
    for sheet_name, id_attr, type_key in (
        ("TECHNICAL_FILES", "tf_id", "TF"),
        ("DECLARATIONS_OF_CONFORMITY", "doc_id", "DOC"),
        ("LABELS", "label_id", "LABEL"),
        ("SHIPMENT_STATEMENTS", "stm_id", "STM"),
    ):
        ws = wb.create_sheet(sheet_name)
        headers = [
            "Packaging Set Code",
            "Linked Product Codes",
            "Document ID",
            "Management Type No",
            "Source Configuration ID",
            "Word Status",
        ]
        rows = []
        for sc in sorted(configs):
            cfg = configs[sc]
            rows.append(
                [
                    sc,
                    linked_products_str(set_to_products.get(sc, [])),
                    getattr(cfg, id_attr),
                    DOC_TYPE_NUMBERS[type_key],
                    cfg.source_configuration_id,
                    "EXISTING_LINKED"
                    if cfg.origin.startswith("EXISTING")
                    else "PENDING_GENERATION",
                ]
            )
        write_table(ws, headers, rows)

    # SEARCH_DATA
    ws = wb.create_sheet("SEARCH_DATA")
    headers = [
        "Product Code",
        "Packaging Set Code",
        "Technical Description",
        "Source Configuration ID",
        "Final Configuration ID",
        "Customer / Market",
        "Packaging Tare kg",
        "Technical File ID",
        "EU DoC ID",
        "Label ID",
        "Shipment Statement ID",
        "Scope Status",
    ]
    rows = []
    for pc in sorted(prod_by_code):
        p = prod_by_code[pc]
        sc = product_set.get(pc, "")
        cfg = configs.get(sc)
        rows.append(
            [
                pc,
                sc,
                p["product_description"],
                p["source_configuration_id"],
                cfg.final_configuration_id if cfg else "",
                p["customer_market"],
                cfg.packaging_tare_kg if cfg else "",
                cfg.tf_id if cfg else "",
                cfg.doc_id if cfg else "",
                cfg.label_id if cfg else "",
                cfg.stm_id if cfg else "",
                SCOPE_STATUS,
            ]
        )
    write_table(ws, headers, rows)

    # SCOPE_RECONCILIATION
    ws = wb.create_sheet("SCOPE_RECONCILIATION")
    headers = [
        "Source Configuration ID",
        "Variant ID",
        "Action",
        "Packaging Set Code",
        "Product Count",
        "Notes",
    ]
    rows = [
        [
            r["source_configuration_id"],
            r["variant_id"],
            r["action"],
            r["packaging_set_code"],
            r["product_count"],
            "",
        ]
        for r in variant_resolutions
    ]
    write_table(ws, headers, rows)

    if STARTER_XLSX.exists():
        STARTER_XLSX.unlink()
    wb.save(STARTER_XLSX)
    wb.close()

    # INDUSTRIAL workbook
    gwb = load_workbook(GOLDEN, data_only=True, read_only=True)
    cfg_ws = gwb["01_FINAL_CONFIG_MASTER"]
    idx = _headers(cfg_ws)
    ind_cfgs = []
    for row in cfg_ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["Family"]] or "").upper() != "INDUSTRIAL":
            continue
        ind_cfgs.append(row)
    bom_ws = gwb["03_BOM_MASTER"]
    bidx = _headers(bom_ws)
    ind_bom = defaultdict(list)
    for row in bom_ws.iter_rows(min_row=2, values_only=True):
        sc = str(row[bidx["Packaging Set Code"]] or "")
        if sc.startswith("IND-"):
            ind_bom[sc].append(row)
    gwb.close()

    iwb = Workbook()
    ws = iwb.active
    ws.title = "00_HOME"
    ws["A1"] = "İNCI AKÜ PPWR — INDUSTRIAL MASTER Rev.00"
    ws["A1"].font = _font(16, True, NAVY)
    ws["A3"] = "Baseline industrial packaging configurations (24V / 48V / 80V)"
    ws["A5"] = f"Configuration count: {len(ind_cfgs)}"
    ws["A6"] = "Product Code mapping: PRODUCT MAP PENDING (do not invent codes)"

    ws = iwb.create_sheet("PRODUCT_MASTER")
    write_table(
        ws,
        [
            "Product Code",
            "Packaging Set Code",
            "Technical Description",
            "Customer / Market",
            "Source Configuration ID",
            "Final Configuration ID",
            "Status",
        ],
        [
            [
                "PRODUCT MAP PENDING",
                str(r[idx["Packaging Set Code"]]),
                str(r[idx["Configuration Name"]] or ""),
                "",
                str(r[idx["Source Configuration ID"]]),
                str(r[idx["Final Configuration ID"]]),
                "PRODUCT MAP PENDING",
            ]
            for r in ind_cfgs
        ],
    )

    ws = iwb.create_sheet("CONFIG_MASTER")
    write_table(
        ws,
        [
            "Packaging Set Code",
            "Linked Product Codes",
            "Source Configuration ID",
            "Final Configuration ID",
            "Product Count",
            "Nominal Qty",
            "Pallet / Packaging Description",
            "Packaging Tare kg",
            "BOM Line Count",
            "Status",
            "Technical File ID",
            "EU DoC ID",
            "Label ID",
            "Shipment Statement ID",
        ],
        [
            [
                str(r[idx["Packaging Set Code"]]),
                "PRODUCT MAP PENDING",
                str(r[idx["Source Configuration ID"]]),
                str(r[idx["Final Configuration ID"]]),
                0,
                r[idx["Nominal Product Qty"]],
                str(r[idx["Configuration Name"]] or ""),
                r[idx["Packaging Mass kg"]],
                len(ind_bom[str(r[idx["Packaging Set Code"]])]),
                "CONTROLLED INDUSTRIAL SET",
                str(r[idx["Technical File ID"]]),
                str(r[idx["EU DoC ID"]]),
                str(r[idx["Label ID"]]),
                str(r[idx["Statement ID"]]),
            ]
            for r in ind_cfgs
        ],
    )

    ws = iwb.create_sheet("BOM_MASTER")
    brows = []
    for sc, lines in sorted(ind_bom.items()):
        for row in lines:
            brows.append(
                [
                    sc,
                    "PRODUCT MAP PENDING",
                    str(row[bidx["Source Configuration ID"]]),
                    str(row[bidx["Component Code"]]),
                    str(row[bidx["ERP Description"]] or ""),
                    row[bidx["Quantity"]],
                    row[bidx["UOM"]],
                    row[bidx["Unit Weight kg"]],
                    row[bidx["Line Weight kg"]],
                ]
            )
    write_table(
        ws,
        [
            "Packaging Set Code",
            "Linked Product Codes",
            "Source Configuration ID",
            "Component Code",
            "Component Description",
            "Quantity",
            "UOM",
            "Unit Weight",
            "Line Weight",
        ],
        brows,
    )

    for sheet_name, id_col in (
        ("DOCUMENT_SCOPE", "Technical File ID"),
        ("DOCUMENT_CENTER", "Technical File ID"),
    ):
        ws = iwb.create_sheet(sheet_name)
        write_table(
            ws,
            [
                "Product Code",
                "Packaging Set Code",
                "Source Configuration ID",
                "Technical File ID",
                "EU DoC ID",
                "Label ID",
                "Shipment Statement ID",
                "Status",
            ],
            [
                [
                    "PRODUCT MAP PENDING",
                    str(r[idx["Packaging Set Code"]]),
                    str(r[idx["Source Configuration ID"]]),
                    str(r[idx["Technical File ID"]]),
                    str(r[idx["EU DoC ID"]]),
                    str(r[idx["Label ID"]]),
                    str(r[idx["Statement ID"]]),
                    "PRODUCT MAP PENDING",
                ]
                for r in ind_cfgs
            ],
        )

    if INDUSTRIAL_XLSX.exists():
        INDUSTRIAL_XLSX.unlink()
    iwb.save(INDUSTRIAL_XLSX)
    iwb.close()

    # QA
    fp_after = word_tf_fingerprint()
    gwb = load_workbook(GOLDEN, data_only=True, read_only=True)
    gidx = _headers(gwb["01_FINAL_CONFIG_MASTER"])
    original_240 = set()
    container_count = 0
    for row in gwb["01_FINAL_CONFIG_MASTER"].iter_rows(min_row=2, values_only=True):
        fam = str(row[gidx["Family"]]).upper()
        if fam == "STARTER":
            original_240.add(str(row[gidx["Packaging Set Code"]]))
        elif fam == "CONTAINER":
            container_count += 1
    gwb.close()

    starter_wb = load_workbook(STARTER_XLSX, read_only=True)
    # industrial/container leakage
    cfg_sheet = starter_wb["CONFIG_MASTER"]
    ch = [c.value for c in next(cfg_sheet.iter_rows(min_row=1, max_row=1))]
    sc_i = ch.index("Packaging Set Code")
    starter_sets = []
    for row in cfg_sheet.iter_rows(min_row=2, values_only=True):
        starter_sets.append(str(row[sc_i]))
    has_ind = any(s.startswith("IND-") for s in starter_sets)
    has_cnt = any(s.startswith("CNT-") for s in starter_sets)
    # adjacency checks
    adjacency_ok = True
    for sheet, a_name, b_name in (
        ("PRODUCT_MASTER", "Product Code", "Packaging Set Code"),
        ("CONFIG_MASTER", "Packaging Set Code", "Linked Product Codes"),
        ("BOM_MASTER", "Packaging Set Code", "Linked Product Codes"),
        ("DOCUMENT_SCOPE", "Product Code", "Packaging Set Code"),
        ("DOCUMENT_CENTER", "Packaging Set Code", "Linked Product Codes"),
        ("SEARCH_DATA", "Product Code", "Packaging Set Code"),
        ("TECHNICAL_FILES", "Packaging Set Code", "Linked Product Codes"),
    ):
        ws = starter_wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if headers[0] != a_name or headers[1] != b_name:
            adjacency_ok = False
            print("ADJ FAIL", sheet, headers[:2])
    starter_wb.close()

    products_with_src = sum(1 for p in products if p["source_configuration_id"])
    products_with_set = sum(1 for p in products if product_set.get(p["product_code"]))
    domestic_sources_in_resolutions = {r["source_configuration_id"] for r in variant_resolutions}
    all_domestic_covered = domestic_ids <= domestic_sources_in_resolutions

    final = (
        len(products) == 2046
        and products_with_src == 2046
        and products_with_set == 2046
        and len(unresolved) == 0
        and original_240 <= set(configs.keys())
        and all_domestic_covered
        and not has_ind
        and not has_cnt
        and len(ind_cfgs) == 3
        and adjacency_ok
        and fp_before["aggregate"] == fp_after["aggregate"]
        and fp_before["count"] == fp_after["count"] == 247
    )

    qa = {
        "starter_product_codes": len(products),
        "existing_starter_sets": 240,
        "former_domestic_source_configs_evaluated": len(domestic_ids),
        "domestic_variants_evaluated": len(variant_resolutions),
        "exact_bom_duplicates_reused": reuse_count,
        "new_packaging_sets_required": new_count,
        "final_starter_packaging_set_count": len(configs),
        "products_without_source_configuration": 2046 - products_with_src,
        "products_without_packaging_set_resolution": len(unresolved),
        "starter_workbook": str(STARTER_XLSX),
        "industrial_workbook": str(INDUSTRIAL_XLSX),
        "industrial_baseline_config_count": len(ind_cfgs),
        "container_records_excluded_from_starter": container_count,
        "product_code_packaging_set_adjacency": "PASS" if adjacency_ok else "FAIL",
        "existing_240_set_codes_unchanged": "PASS"
        if original_240 <= set(configs.keys())
        else "FAIL",
        "word_hash_changed": 0
        if fp_before["aggregate"] == fp_after["aggregate"]
        else 1,
        "word_fingerprint_before": fp_before,
        "word_fingerprint_after": fp_after,
        "no_industrial_in_starter": not has_ind,
        "no_container_in_starter": not has_cnt,
        "all_domestic_sources_present": all_domestic_covered,
        "final": "PASS" if final else "FAIL",
        "variant_resolutions_sample": variant_resolutions[:10],
    }

    lines = [
        "# STARTER / INDUSTRIAL SPLIT QA",
        "",
        f"Starter Product Codes: {qa['starter_product_codes']}",
        f"Existing Starter sets: {qa['existing_starter_sets']}",
        f"Former domestic source configs evaluated: {qa['former_domestic_source_configs_evaluated']}",
        f"Exact BOM duplicates reused: {qa['exact_bom_duplicates_reused']}",
        f"New Packaging Sets required: {qa['new_packaging_sets_required']}",
        f"Final Starter Packaging Set count: {qa['final_starter_packaging_set_count']}",
        "",
        f"Products without Source Configuration: {qa['products_without_source_configuration']}",
        f"Products without Packaging Set resolution: {qa['products_without_packaging_set_resolution']}",
        "",
        f"Starter workbook: `{qa['starter_workbook']}`",
        f"Industrial workbook: `{qa['industrial_workbook']}`",
        "",
        f"Industrial baseline config count: {qa['industrial_baseline_config_count']}",
        "",
        f"Container records excluded from Starter: {qa['container_records_excluded_from_starter']}",
        "",
        f"Product Code <-> Packaging Set adjacency: {qa['product_code_packaging_set_adjacency']}",
        "",
        f"Existing 240 set codes unchanged: {qa['existing_240_set_codes_unchanged']}",
        "",
        f"Word hash changed: {qa['word_hash_changed']} expected",
        "",
        f"FINAL: {qa['final']}",
        "",
        "STOP.",
    ]
    QA_MD.write_text("\n".join(lines), encoding="utf-8")
    QA_JSON.write_text(json.dumps(qa, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    try:
        print("\n".join(lines))
    except UnicodeEncodeError:
        print("\n".join(lines).encode("ascii", "replace").decode("ascii"))
    return qa


if __name__ == "__main__":
    build()
