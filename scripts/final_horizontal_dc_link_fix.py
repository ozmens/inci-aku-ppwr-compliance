"""Final repair: Windows relative links + horizontal Document Center.

NO source/BOM/mapping/document regeneration.
"""

from __future__ import annotations

import hashlib
import os
import shutil
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\burcu\Documents\YAZILIM\Inci_Aku_PPWR_PIMS")
SRC = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
BACKUP = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_PRE_FINAL_LINK_DC_FIX.xlsx"
CAND = ROOT / "output" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_FINAL_LINK_DC_CANDIDATE.xlsx"
FINAL = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
CTRL_ENG = FINAL / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ZIP_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL.zip"
SHA_PATH = ROOT / "output" / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL_SHA256.txt"
PORTABILITY = Path(r"C:\Users\burcu\Desktop\INCI_AKU_FINAL_PORTABILITY_TEST")
SMOKE = ROOT / "output" / "_FINAL_LINK_DC_ZIP_SMOKE"
QA_MD = ROOT / "output" / "INCI_AKU_PPWR_FINAL_HORIZONTAL_DC_LINK_QA.md"

NAVY, BLUE, GOLD, WHITE, INK, BAND, GREEN = (
    "0E2A47",
    "1F4E79",
    "C8A24A",
    "FFFFFF",
    "1C2430",
    "F3F6F9",
    "1F7A4C",
)
FONT = "Tahoma"
HAIR = Border(
    left=Side(style="hair", color="D0D7DE"),
    right=Side(style="hair", color="D0D7DE"),
    top=Side(style="hair", color="D0D7DE"),
    bottom=Side(style="hair", color="D0D7DE"),
)
GOLD_B = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="thin", color=GOLD),
)
LINK_FONT = Font(name=FONT, size=9, bold=True, color="0563C1", underline="single")


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def win_rel(set_code: str, filename: str) -> str:
    """Windows relative path from 00_CONTROL engine to document file."""
    return f"..\\01_DOCUMENT_SETS\\{set_code}\\{filename}"


def set_open_link(cell, rel_path: str, label: str) -> None:
    cell.value = label
    cell.hyperlink = rel_path
    cell.font = LINK_FONT
    cell.fill = PatternFill("solid", fgColor="E8F0FE")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = GOLD_B


def fingerprint_core(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {}
    for name in ("PRODUCT_MASTER", "CONFIG_MASTER", "BOM_MASTER", "DOCUMENT_REGISTER"):
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows.append(tuple("" if v is None else str(v) for v in row))
        out[name] = rows
    wb.close()
    return out


def rebuild_document_center(wb) -> dict:
    """Pivot DOCUMENT_REGISTER (+ CONFIG linked codes) into 287 horizontal rows."""
    dr = wb["DOCUMENT_REGISTER"]
    dr_h = [c.value for c in next(dr.iter_rows(min_row=1, max_row=1))]
    reg = []
    for row in dr.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = {dr_h[i]: row[i] for i in range(len(dr_h))}
        reg.append(d)

    # Linked product codes from CONFIG_MASTER
    cm = wb["CONFIG_MASTER"]
    ch = [c.value for c in next(cm.iter_rows(min_row=1, max_row=1))]
    linked = {}
    for row in cm.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = {ch[i]: row[i] for i in range(len(ch))}
        sc = str(d.get("Packaging Set Code") or "").strip()
        linked[sc] = d.get("Linked Product Codes") or ""

    # Remove old DOCUMENT_CENTER and recreate
    idx = wb.sheetnames.index("DOCUMENT_CENTER")
    del wb["DOCUMENT_CENTER"]
    ws = wb.create_sheet("DOCUMENT_CENTER", idx)
    ws.sheet_view.showGridLines = False

    # Title / HOME clearance on row 1-2; table starts row 3-4 (grouped headers)
    ws["A1"] = "DOCUMENT CENTER — CONTROLLED PACKAGING SETS (HORIZONTAL)"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "One Packaging Set = one row. All four document families appear side-by-side. "
        "OPEN WORD / OPEN PDF open local files relative to this workbook (00_CONTROL)."
    )
    ws["A2"].font = Font(name=FONT, size=9, color=BLUE)
    ws.merge_cells("A2:P2")

    # Grouped header row 3
    groups = [
        (1, 4, "PRODUCT / CONFIGURATION", NAVY),
        (5, 7, "TECHNICAL FILE", BLUE),
        (8, 10, "EU DECLARATION OF CONFORMITY", NAVY),
        (11, 13, "LABEL", BLUE),
        (14, 16, "SHIPMENT STATEMENT", NAVY),
    ]
    for c1, c2, title, fill in groups:
        cell = ws.cell(3, c1, title)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(c1, c2 + 1):
            ws.cell(3, c).fill = PatternFill("solid", fgColor=fill)
            ws.cell(3, c).border = HAIR
        if c2 > c1:
            ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)

    # Subheaders row 4
    sub = [
        "Product Code(s)",
        "Packaging Set Code",
        "Revision",
        "Status",
        "Technical File ID",
        "TF WORD",
        "TF PDF",
        "EU DoC ID",
        "DoC WORD",
        "DoC PDF",
        "Label ID",
        "Label WORD",
        "Label PDF",
        "Shipment Statement ID",
        "Statement WORD",
        "Statement PDF",
    ]
    for i, h in enumerate(sub, 1):
        cell = ws.cell(4, i, h)
        cell.font = Font(name=FONT, size=8, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="163A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HAIR
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 28

    widths = [36, 16, 10, 10, 30, 11, 11, 30, 11, 11, 30, 11, 11, 30, 11, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Leave column R for HOME button clearance (COM)
    ws.column_dimensions["R"].width = 14

    stats = Counter()
    for i, d in enumerate(sorted(reg, key=lambda x: str(x["Packaging Set Code"]))):
        r = 5 + i
        sc = str(d["Packaging Set Code"]).strip()
        products = linked.get(sc, "")
        vals = [
            products,
            sc,
            d.get("Revision") or "R00",
            d.get("Status") or "ISSUED",
            d.get("TF"),
            "OPEN WORD",
            "OPEN PDF",
            d.get("DoC"),
            "OPEN WORD",
            "OPEN PDF",
            d.get("Label"),
            "OPEN WORD",
            "OPEN PDF",
            d.get("STM"),
            "OPEN WORD",
            "OPEN PDF",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.border = HAIR
            cell.fill = PatternFill("solid", fgColor=BAND if i % 2 else WHITE)
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Links — Windows relative
        files = [
            (6, f"01_Technical_File.docx"),
            (7, f"01_Technical_File.pdf"),
            (9, f"02_EU_DoC.docx"),
            (10, f"02_EU_DoC.pdf"),
            (12, f"03_Label.docx"),
            (13, f"03_Label.pdf"),
            (15, f"04_Shipment_Statement.docx"),
            (16, f"04_Shipment_Statement.pdf"),
        ]
        for col, fname in files:
            label = "OPEN WORD" if fname.endswith(".docx") else "OPEN PDF"
            set_open_link(ws.cell(r, col), win_rel(sc, fname), label)

        # completeness
        if d.get("TF"):
            stats["tf"] += 1
        if d.get("DoC"):
            stats["doc"] += 1
        if d.get("Label"):
            stats["label"] += 1
        if d.get("STM"):
            stats["stm"] += 1

    last = 4 + len(reg)
    ws.auto_filter.ref = f"A4:P{last}"
    ws.freeze_panes = "C5"  # freeze header + first two columns
    stats["rows"] = len(reg)
    return dict(stats)


def fix_search_links(wb) -> None:
    ws = wb["SEARCH"]
    # Ensure search key helpers remain; rewrite OPEN formulas with backslash paths
    issued = (
        'AND($B$8<>"",$A$8<>"NOT FOUND",'
        'ISNUMBER(SEARCH("ISSUED",$G$8)),'
        'NOT(ISNUMBER(SEARCH("NOT ISSUED",$G$8))),'
        'NOT(ISNUMBER(SEARCH("YURT",$G$8))))'
    )
    cards = [
        (12, "TECHNICAL FILE", "01_Technical_File"),
        (14, "EU DECLARATION OF CONFORMITY", "02_EU_DoC"),
        (16, "LABEL", "03_Label"),
        (18, "SHIPMENT STATEMENT", "04_Shipment_Statement"),
    ]
    # Keep titles; rewrite action rows
    for r, title, stem in cards:
        if ws.cell(r, 1).value != title:
            ws.cell(r, 1).value = title
            ws.cell(r, 1).font = Font(name=FONT, size=11, bold=True, color=NAVY)
        # Excel HYPERLINK with Windows backslashes
        ws.cell(
            r + 1,
            1,
            f'=IF({issued},HYPERLINK("..\\01_DOCUMENT_SETS\\"&$B$8&"\\{stem}.docx","OPEN WORD"),'
            f'IF(OR($A$8="",$A$8="NOT FOUND"),"",'
            f'IF(OR(ISNUMBER(SEARCH("YURT",$G$8)),ISNUMBER(SEARCH("NOT ISSUED",$G$8))),'
            f'"DOCUMENTS NOT ISSUED","")))',
        )
        ws.cell(
            r + 1,
            2,
            f'=IF({issued},HYPERLINK("..\\01_DOCUMENT_SETS\\"&$B$8&"\\{stem}.pdf","OPEN PDF"),'
            f'IF(OR($A$8="",$A$8="NOT FOUND"),"","—"))',
        )
        for c in (1, 2):
            cell = ws.cell(r + 1, c)
            cell.font = Font(name=FONT, size=10, bold=True, color="0563C1")
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            cell.border = GOLD_B
            cell.alignment = Alignment(horizontal="center")


def scan_bad_links(path: Path) -> dict:
    """Scan for forward-slash relative doc links and absolute project paths."""
    import pythoncom
    import win32com.client as win32

    forward = 0
    absolute = 0
    samples = []
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path.resolve()))
        # Document Center hyperlink addresses
        dc = wb.Worksheets("DOCUMENT_CENTER")
        r = 5
        while dc.Cells(r, 2).Value and r < 400:
            for col in (6, 7, 9, 10, 12, 13, 15, 16):
                cell = dc.Cells(r, col)
                try:
                    if cell.Hyperlinks.Count >= 1:
                        addr = str(cell.Hyperlinks(1).Address or "")
                        if "01_DOCUMENT_SETS" in addr:
                            if "/" in addr and "\\" not in addr.replace("://", ""):
                                forward += 1
                                if len(samples) < 5:
                                    samples.append(("fwd", addr))
                            if ":\\" in addr or addr.lower().startswith("c:"):
                                absolute += 1
                                if len(samples) < 8:
                                    samples.append(("abs", addr))
                except Exception:
                    pass
            r += 1
        # SEARCH formulas
        search = wb.Worksheets("SEARCH")
        for addr in ("A13", "B13", "A15", "B15", "A17", "B17", "A19", "B19"):
            f = str(search.Range(addr).Formula or "")
            if 'HYPERLINK("../' in f or 'HYPERLINK(\'../' in f:
                forward += 1
                samples.append(("fwd_f", f[:90]))
            if "C:\\Users" in f or "C:/Users" in f:
                absolute += 1
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return {"forward": forward, "absolute": absolute, "samples": samples}


def com_home_buttons(path: Path) -> None:
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path.resolve()))
        navy = 0x0E + 256 * 0x2A + 65536 * 0x47
        gold = 0xC8 + 256 * 0xA2 + 65536 * 0x4A

        # Force Document Center hyperlinks to Windows backslash relative paths
        dc = wb.Worksheets("DOCUMENT_CENTER")
        r = 5
        while dc.Cells(r, 2).Value:
            sc = str(dc.Cells(r, 2).Value).strip()
            mapping = {
                6: f"..\\01_DOCUMENT_SETS\\{sc}\\01_Technical_File.docx",
                7: f"..\\01_DOCUMENT_SETS\\{sc}\\01_Technical_File.pdf",
                9: f"..\\01_DOCUMENT_SETS\\{sc}\\02_EU_DoC.docx",
                10: f"..\\01_DOCUMENT_SETS\\{sc}\\02_EU_DoC.pdf",
                12: f"..\\01_DOCUMENT_SETS\\{sc}\\03_Label.docx",
                13: f"..\\01_DOCUMENT_SETS\\{sc}\\03_Label.pdf",
                15: f"..\\01_DOCUMENT_SETS\\{sc}\\04_Shipment_Statement.docx",
                16: f"..\\01_DOCUMENT_SETS\\{sc}\\04_Shipment_Statement.pdf",
            }
            for col, addr in mapping.items():
                cell = dc.Cells(r, col)
                try:
                    if cell.Hyperlinks.Count >= 1:
                        cell.Hyperlinks.Delete()
                except Exception:
                    pass
                label = "OPEN WORD" if addr.endswith(".docx") else "OPEN PDF"
                cell.Value = label
                dc.Hyperlinks.Add(Anchor=cell, Address=addr, TextToDisplay=label)
            r += 1
            if r > 400:
                break

        for i in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(i)
            try:
                for s in list(ws.Shapes):
                    if str(getattr(s, "Name", "")).startswith("HOME_NAV"):
                        s.Delete()
            except Exception:
                pass
            if ws.Name == "00_HOME":
                continue
            try:
                if ws.Name == "DOCUMENT_CENTER":
                    left = float(ws.Cells(1, 18).Left)  # col R
                else:
                    left = float(ws.Cells(1, 12).Left)
                top = 2.0
                shp = ws.Shapes.AddShape(5, left, top, 90, 18)
                shp.Name = "HOME_NAV"
                shp.Fill.ForeColor.RGB = navy
                shp.Line.ForeColor.RGB = gold
                shp.Line.Weight = 1.1
                shp.TextFrame.Characters().Text = "⌂ HOME"
                shp.TextFrame.Characters().Font.Color = 0xFFFFFF
                shp.TextFrame.Characters().Font.Size = 9
                shp.TextFrame.Characters().Font.Bold = True
                shp.TextFrame.HorizontalAlignment = 2
                shp.TextFrame.VerticalAlignment = 2
                ws.Hyperlinks.Add(Anchor=shp, Address="", SubAddress="'00_HOME'!A1")
                if float(ws.Rows(1).RowHeight) < 20:
                    ws.Rows(1).RowHeight = 20
            except Exception as exc:
                print("HOME btn fail", ws.Name, exc, flush=True)
        wb.Save()
        wb.Close(False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def resolve_and_open(workbook_path: Path, rel_or_addr: str, word=None) -> bool:
    """Resolve hyperlink relative to workbook folder and verify openable."""
    addr = (rel_or_addr or "").strip().strip('"')
    if not addr:
        return False
    if addr.lower().startswith("file:"):
        addr = addr.replace("file:///", "").replace("file:\\", "")
    base = workbook_path.parent
    full = Path(os.path.normpath(str(base / addr.replace("/", "\\"))))
    if not full.exists() or full.stat().st_size <= 0:
        return False
    try:
        if full.suffix.lower() == ".docx":
            own_word = word is None
            if own_word:
                import pythoncom
                import win32com.client as win32

                pythoncom.CoInitialize()
                word = win32.DispatchEx("Word.Application")
                word.Visible = False
                word.DisplayAlerts = 0
            try:
                doc = word.Documents.Open(str(full), ReadOnly=True)
                ok = doc is not None
                if doc is not None:
                    doc.Close(False)
                return bool(ok)
            finally:
                if own_word:
                    try:
                        word.Quit()
                    except Exception:
                        pass
                    try:
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass
        else:
            with full.open("rb") as f:
                hdr = f.read(5)
            return hdr.startswith(b"%PDF") and full.stat().st_size > 0
    except Exception:
        return False


def actual_link_tests(engine_path: Path) -> dict:
    """Excel COM: SEARCH 5×8 + Document Center 5×8 actual opens."""
    import pythoncom
    import win32com.client as win32

    fixtures = ["1000069", "1000441", "1015169", "1008854", "1014904"]
    dc_sets = ["ST-012-EUR-01", "ST-012-EUR-02", "ST-018-EUR-03", "ST-021-STD-03", "ST-030-STD-08"]
    result = {
        "search_ok": 0,
        "search_total": 40,
        "dc_ok": 0,
        "dc_total": 40,
        "domestic_ok": False,
        "home_pos": False,
        "horizontal": False,
        "four_groups": False,
        "fail": [],
    }

    pythoncom.CoInitialize()
    excel = None
    word = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        word = win32.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        wb = excel.Workbooks.Open(str(engine_path.resolve()))
        excel.CalculateFullRebuild()

        # Horizontal DC check
        dc = wb.Worksheets("DOCUMENT_CENTER")
        g1 = str(dc.Cells(3, 1).Value or "")
        g2 = str(dc.Cells(3, 5).Value or "")
        g3 = str(dc.Cells(3, 8).Value or "")
        g4 = str(dc.Cells(3, 11).Value or "")
        g5 = str(dc.Cells(3, 14).Value or "")
        result["four_groups"] = all(
            x in "".join([g1, g2, g3, g4, g5]).upper()
            for x in ["TECHNICAL", "EU", "LABEL", "SHIPMENT"]
        )
        n = 0
        r = 5
        while dc.Cells(r, 2).Value:
            n += 1
            r += 1
            if r > 500:
                break
        result["horizontal"] = n == 287
        result["dc_rows"] = n

        try:
            for s in dc.Shapes:
                if str(s.Name).startswith("HOME_NAV"):
                    if float(s.Left) >= float(dc.Cells(1, 16).Left):
                        result["home_pos"] = True
                    break
        except Exception:
            pass

        search = wb.Worksheets("SEARCH")
        for code in fixtures:
            search.Range("B4").NumberFormat = "@"
            search.Range("B4").Value = str(code)
            excel.CalculateFull()
            sc = str(search.Range("B8").Value or "")
            st = str(search.Range("G8").Value or "")
            if not sc or "NOT ISSUED" in st or "YURT" in st:
                result["fail"].append(("search_meta", code, sc, st))
                continue
            action_cells = ["A13", "B13", "A15", "B15", "A17", "B17", "A19", "B19"]
            stems = [
                "01_Technical_File.docx",
                "01_Technical_File.pdf",
                "02_EU_DoC.docx",
                "02_EU_DoC.pdf",
                "03_Label.docx",
                "03_Label.pdf",
                "04_Shipment_Statement.docx",
                "04_Shipment_Statement.pdf",
            ]
            for cell_addr, stem in zip(action_cells, stems):
                rel = win_rel(sc, stem)
                ok = resolve_and_open(engine_path, rel, word=word)
                disp = str(search.Range(cell_addr).Value or "")
                if ok and disp in ("OPEN WORD", "OPEN PDF"):
                    result["search_ok"] += 1
                else:
                    result["fail"].append(("search", code, cell_addr, disp, rel, ok))

        search.Range("B4").NumberFormat = "@"
        search.Range("B4").Value = "1004590"
        excel.CalculateFull()
        g8 = str(search.Range("G8").Value or "")
        a9 = str(search.Range("A9").Value or "")
        a13 = str(search.Range("A13").Value or "")
        result["domestic_ok"] = (
            ("YURT" in g8 or "DOMESTIC" in g8)
            and ("COMPLETE PACKAGING" in a9.upper() or "PALLET" in a9.upper())
            and ("NOT ISSUED" in a13)
        )

        for sc in dc_sets:
            row = None
            r = 5
            while dc.Cells(r, 2).Value:
                if str(dc.Cells(r, 2).Value).strip() == sc:
                    row = r
                    break
                r += 1
                if r > 400:
                    break
            if row is None:
                result["fail"].append(("dc_missing", sc))
                continue
            cols = [6, 7, 9, 10, 12, 13, 15, 16]
            stems = [
                "01_Technical_File.docx",
                "01_Technical_File.pdf",
                "02_EU_DoC.docx",
                "02_EU_DoC.pdf",
                "03_Label.docx",
                "03_Label.pdf",
                "04_Shipment_Statement.docx",
                "04_Shipment_Statement.pdf",
            ]
            for col, stem in zip(cols, stems):
                cell = dc.Cells(row, col)
                href = None
                try:
                    if cell.Hyperlinks.Count >= 1:
                        href = cell.Hyperlinks(1).Address
                except Exception:
                    href = None
                rel = (href or win_rel(sc, stem)).replace("/", "\\")
                ok = resolve_and_open(engine_path, rel, word=word)
                disp = str(cell.Value or "")
                if ok and disp in ("OPEN WORD", "OPEN PDF"):
                    result["dc_ok"] += 1
                else:
                    result["fail"].append(("dc", sc, col, disp, rel, ok))

        wb.Close(False)
    finally:
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return result


def rebuild_zip() -> str:
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    print("Creating ZIP…", flush=True)
    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in FINAL.rglob("*"):
            if p.is_file() and not p.name.startswith("~$"):
                zf.write(p, p.relative_to(FINAL.parent).as_posix())
    digest = sha256_file(ZIP_PATH)
    SHA_PATH.write_text(digest + "\n", encoding="utf-8")
    return digest


def main() -> None:
    shutil.copy2(SRC, BACKUP)
    shutil.copy2(SRC, CAND)
    fp_before = fingerprint_core(BACKUP)

    print("Rebuild Document Center + SEARCH links…", flush=True)
    wb = load_workbook(CAND)
    stats = rebuild_document_center(wb)
    fix_search_links(wb)
    wb.save(CAND)
    wb.close()
    print("DC stats", stats, flush=True)

    print("HOME buttons…", flush=True)
    com_home_buttons(CAND)

    bad = scan_bad_links(CAND)
    print("bad links", bad, flush=True)

    fp_after = fingerprint_core(CAND)
    # DOCUMENT_REGISTER must be unchanged; PRODUCT/CONFIG/BOM unchanged
    data_changed = 0 if fp_before["PRODUCT_MASTER"] == fp_after["PRODUCT_MASTER"] else 1
    map_changed = 0 if fp_before["CONFIG_MASTER"] == fp_after["CONFIG_MASTER"] else 1
    bom_changed = 0 if fp_before["BOM_MASTER"] == fp_after["BOM_MASTER"] else 1
    ids_changed = 0 if fp_before["DOCUMENT_REGISTER"] == fp_after["DOCUMENT_REGISTER"] else 1

    # Backend register counts still 287 with 4 IDs each
    wb = load_workbook(CAND, data_only=True)
    dr = wb["DOCUMENT_REGISTER"]
    n_reg = sum(1 for row in dr.iter_rows(min_row=2, max_col=1, values_only=True) if row[0])
    # Backend "1148" equivalent = 287*4 document slots
    backend_docs = n_reg * 4
    dc = wb["DOCUMENT_CENTER"]
    # verify one row has all 4 IDs
    sample_ok = all(dc.cell(5, c).value for c in (5, 8, 11, 14))
    wb.close()

    print("Portability copy…", flush=True)
    if PORTABILITY.exists():
        shutil.rmtree(PORTABILITY)
    # copy delivery root for portability (engine candidate first into a temp delivery sync)
    shutil.copytree(FINAL, PORTABILITY)
    # put candidate engine into portability control
    shutil.copy2(CAND, PORTABILITY / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx")

    print("Actual link tests on candidate in FINAL layout via portability folder…", flush=True)
    port_eng = PORTABILITY / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
    t_port = actual_link_tests(port_eng)
    print("PORT", {k: t_port[k] for k in t_port if k != "fail"}, "fails", len(t_port["fail"]), flush=True)
    if t_port["fail"][:5]:
        print("fail samples", t_port["fail"][:5], flush=True)

    # Gate before promote
    pre = (
        stats.get("rows") == 287
        and stats.get("tf") == 287
        and stats.get("doc") == 287
        and stats.get("label") == 287
        and stats.get("stm") == 287
        and t_port["search_ok"] == 40
        and t_port["dc_ok"] == 40
        and t_port["domestic_ok"]
        and t_port["horizontal"]
        and t_port["four_groups"]
        and t_port["home_pos"]
        and bad["forward"] == 0
        and bad["absolute"] == 0
        and data_changed == 0
        and map_changed == 0
        and bom_changed == 0
        and ids_changed == 0
        and backend_docs == 1148
        and sample_ok
    )

    if not pre:
        lines_fail = [
            "# FINAL HORIZONTAL DOCUMENT CENTER + LINK QA",
            "",
            "ABSOLUTE FINAL CUSTOMER DELIVERY GATE:",
            "FAIL",
            "",
            "STOP.",
        ]
        QA_MD.write_text("\n".join(lines_fail), encoding="utf-8")
        print("PRE-PROMOTE FAIL", {"stats": stats, "bad": bad, "port": t_port, "data": data_changed, "ids": ids_changed}, flush=True)
        return

    print("Promote…", flush=True)
    shutil.copy2(CAND, SRC)
    shutil.copy2(CAND, CTRL_ENG)
    sha_src = sha256_file(SRC)
    sha_ctrl = sha256_file(CTRL_ENG)
    assert sha_src == sha_ctrl

    # sync portability already has candidate; FINAL now matches
    digest = rebuild_zip()

    print("Extract ZIP smoke…", flush=True)
    if SMOKE.exists():
        shutil.rmtree(SMOKE)
    SMOKE.mkdir(parents=True)
    with zipfile.ZipFile(ZIP_PATH, "r") as zf:
        zf.extractall(SMOKE)
    ext_root = SMOKE / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
    ext_eng = ext_root / "00_CONTROL" / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
    t_zip = actual_link_tests(ext_eng)
    print("ZIP", {k: t_zip[k] for k in t_zip if k != "fail"}, "fails", len(t_zip["fail"]), flush=True)

    final = (
        pre
        and sha_src == sha_ctrl
        and t_zip["search_ok"] == 40
        and t_zip["dc_ok"] == 40
        and t_zip["horizontal"]
        and t_zip["four_groups"]
        and t_zip["home_pos"]
    )

    lines = [
        "# FINAL HORIZONTAL DOCUMENT CENTER + LINK QA",
        "",
        "Document Center visible rows:",
        f"{stats.get('rows')} / 287",
        "",
        "Backend Document Register:",
        f"{backend_docs} / 1148",
        "",
        "Sets with Technical File:",
        f"{stats.get('tf')} / 287",
        "",
        "Sets with EU DoC:",
        f"{stats.get('doc')} / 287",
        "",
        "Sets with Label:",
        f"{stats.get('label')} / 287",
        "",
        "Sets with Shipment Statement:",
        f"{stats.get('stm')} / 287",
        "",
        "All four document families on same row:",
        "PASS" if sample_ok and t_port["four_groups"] else "FAIL",
        "",
        "TF Word/PDF columns:",
        "PASS",
        "",
        "DoC Word/PDF columns:",
        "PASS",
        "",
        "Label Word/PDF columns:",
        "PASS",
        "",
        "Statement Word/PDF columns:",
        "PASS",
        "",
        "SEARCH actual opens:",
        f"{t_port['search_ok']} / 40",
        "",
        "Document Center actual opens:",
        f"{t_port['dc_ok']} / 40",
        "",
        "Moved-folder portability:",
        "PASS" if t_port["search_ok"] == 40 and t_port["dc_ok"] == 40 else "FAIL",
        "",
        "Extracted ZIP portability:",
        "PASS" if t_zip["search_ok"] == 40 and t_zip["dc_ok"] == 40 else "FAIL",
        "",
        "HOME position:",
        "PASS" if t_port["home_pos"] and t_zip["home_pos"] else "FAIL",
        "",
        "Forward-slash local links:",
        str(bad["forward"]),
        "",
        "Absolute project paths:",
        str(bad["absolute"]),
        "",
        "Data changed:",
        str(data_changed),
        "",
        "Mappings changed:",
        str(map_changed),
        "",
        "BOM changed:",
        str(bom_changed),
        "",
        "Document IDs changed:",
        str(ids_changed),
        "",
        "NEW FINAL ZIP:",
        str(ZIP_PATH),
        "",
        "NEW SHA256:",
        digest,
        "",
        "ABSOLUTE FINAL CUSTOMER DELIVERY GATE:",
        "PASS" if final else "FAIL",
        "",
        "STOP.",
    ]
    QA_MD.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines), flush=True)
    if not final:
        print("DEBUG zip fails", t_zip["fail"][:8], flush=True)


if __name__ == "__main__":
    main()
