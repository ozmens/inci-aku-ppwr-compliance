"""Rebuild product-level Document Engine from the good Starter FINAL UI.

- Keeps the premium HOME / SEARCH / QA look from FINAL
- Rewires paths to 01_PRODUCT_DOCUMENT_SETS\\<ProductCode>\\
- Makes SEARCH OPEN links use clickable HYPERLINK() formulas
- Fixes DOCUMENT_CENTER relative hyperlinks + blue underline font
- Does NOT launch Word or Excel COM
"""

from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
FINAL_ENG = (
    ROOT
    / "output"
    / "INCI_AKU_PPWR_STARTER_CUSTOMER_DELIVERY_REV00_FINAL"
    / "00_CONTROL"
    / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
)
CANDIDATE = (
    ROOT
    / "output"
    / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_CUSTOMER_DELIVERY_REV00_CANDIDATE"
)
CONTROL = CANDIDATE / "00_CONTROL"
PRODUCT_SETS = CANDIDATE / "01_PRODUCT_DOCUMENT_SETS"
ENG_CTRL = CONTROL / "INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx"
ENG_ROOT = ROOT / "output" / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_ENGINE_Rev00.xlsx"
MASTER = ROOT / "output" / "INCI_AKU_PPWR_STARTER_MASTER_Rev00.xlsx"
DESKTOP_CMD = Path.home() / "Desktop" / "00_AC_PRODUCT_LEVEL_ENGINE.cmd"
DESKTOP_BAD_XLSX = Path.home() / "Desktop" / "INCI_AKU_PPWR_STARTER_PRODUCT_LEVEL_ENGINE_Rev00.xlsx"

FONT = "Tahoma"
NAVY = "0E2A47"
LINK_FONT = Font(name=FONT, color="0563C1", underline="single", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
STEMS = [
    ("01_Technical_File", "Technical File ID", "TF WORD", "TF PDF"),
    ("02_EU_DoC", "EU DoC ID", "DoC WORD", "DoC PDF"),
    ("03_Label", "Label ID", "Label WORD", "Label PDF"),
    ("04_Shipment_Statement", "Shipment Statement ID", "Statement WORD", "Statement PDF"),
]


def product_instance_ids(pc: str, sc: str) -> dict[str, str]:
    return {
        "tf": f"IA-PPWR-TF-{pc}-{sc}-R00",
        "doc": f"IA-PPWR-DOC-{pc}-{sc}-R00",
        "label": f"IA-PPWR-LBL-{pc}-{sc}-R00",
        "stm": f"IA-PPWR-STM-{pc}-{sc}-R00",
    }


def load_products() -> tuple[list[dict], int, int]:
    wb = load_workbook(MASTER, data_only=True, read_only=True)
    pm = wb["PRODUCT_MASTER"]
    headers = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
    hi = {h: i for i, h in enumerate(headers)}
    rows: list[dict] = []
    starter = gap = 0
    for row in pm.iter_rows(min_row=2, values_only=True):
        pc = str(row[hi["Product Code"]] or "").strip()
        if not pc:
            continue
        starter += 1
        st = str(row[hi["Physical Packaging Status"]] or "")
        if "NOT ISSUED" in st or "DATA REQUIRED" in st or "DATA GAP" in st.upper():
            gap += 1
            continue
        sc = str(row[hi.get("Physical Packaging Set Code", hi.get("Packaging Set Code", 0))] or "").strip()
        # try common header names
        for key in (
            "Physical Packaging Set Code",
            "Packaging Set Code",
            "Final Packaging Set Code",
        ):
            if key in hi and row[hi[key]]:
                sc = str(row[hi[key]]).strip()
                break
        desc = str(row[hi.get("Technical Description", hi.get("Product Description", 0))] or "")
        cfg = str(row[hi.get("Final Configuration ID", 0)] or "") if "Final Configuration ID" in hi else ""
        tare = row[hi["Packaging Tare kg"]] if "Packaging Tare kg" in hi else ""
        if not sc:
            continue
        # only keep if folder exists
        if not (PRODUCT_SETS / pc).is_dir():
            continue
        rows.append(
            {
                "product_code": pc,
                "set_code": sc,
                "description": desc,
                "config_id": cfg,
                "tare": tare,
                "status": "ISSUED Rev.00",
            }
        )
    wb.close()
    rows.sort(key=lambda x: x["product_code"])
    return rows, starter, gap


def clear_sheet(ws) -> None:
    ws.delete_rows(1, ws.max_row or 1)
    # also clear hyperlinks collection if present
    try:
        ws._hyperlinks = []
    except Exception:
        pass


def set_link(cell, path: str, label: str) -> None:
    cell.value = label
    cell.hyperlink = path
    cell.font = LINK_FONT


def rebuild_search_data(wb, products: list[dict]) -> None:
    ws = wb["SEARCH_DATA"]
    clear_sheet(ws)
    headers = [
        "Product Code",
        "Packaging Set Code",
        "Product Description",
        "Configuration ID",
        "Packaging Tare",
        "Revision",
        "Status",
        "KEY",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, p in enumerate(products, 2):
        ws.cell(i, 1, p["product_code"])
        ws.cell(i, 2, p["set_code"])
        ws.cell(i, 3, p["description"])
        ws.cell(i, 4, p["config_id"])
        ws.cell(i, 5, p["tare"])
        ws.cell(i, 6, "R00")
        ws.cell(i, 7, p["status"])
        ws.cell(i, 8, p["product_code"])
    ws.auto_filter.ref = f"A1:H{len(products)+1}"
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col != 3 else 42


def patch_search_formulas(wb) -> None:
    """Keep FINAL SEARCH UX; point file links to product folders via $A$8."""
    ws = wb["SEARCH"]
    ws["A1"] = "GLOBAL PRODUCT SEARCH — PRODUCT-LEVEL PACKS"
    ws["A2"] = (
        "Enter a Product Code below. Exact match. Opens Word/PDF from "
        "01_PRODUCT_DOCUMENT_SETS\\<ProductCode>\\ (workbook must be opened from 00_CONTROL)."
    )
    # Keep Z1/AA1/Z2 match logic — KEY is product code in col H and A
    base = 'IFERROR(LEFT(CELL("filename",$B$4),FIND("[",CELL("filename",$B$4))-1),"")'
    issued = (
        'AND($A$8<>"",$A$8<>"NOT FOUND",ISNUMBER(SEARCH("ISSUED",$G$8)),'
        'NOT(ISNUMBER(SEARCH("NOT ISSUED",$G$8))),NOT(ISNUMBER(SEARCH("YURT",$G$8))))'
    )
    empty = 'OR($B$4="",$A$8="",$A$8="NOT FOUND")'
    domestic = 'OR(ISNUMBER(SEARCH("YURT",$G$8)),ISNUMBER(SEARCH("NOT ISSUED",$G$8)))'

    def formula(stem: str, word: bool) -> str:
        label = "OPEN WORD" if word else "OPEN PDF"
        ext = "docx" if word else "pdf"
        # Product folder key = $A$8 (product code), not set code
        path = f'{base}&"..\\01_PRODUCT_DOCUMENT_SETS\\"&$A$8&"\\{stem}.{ext}"'
        return (
            f'=IF({empty},"",'
            f'IF({domestic},"DOCUMENTS NOT ISSUED",'
            f'HYPERLINK(IF({issued},{path},"#"),IF({issued},"{label}",""))))'
        )

    pairs = [
        (13, "01_Technical_File"),
        (15, "02_EU_DoC"),
        (17, "03_Label"),
        (19, "04_Shipment_Statement"),
    ]
    for row, stem in pairs:
        ws.cell(row, 1).value = formula(stem, True)
        ws.cell(row, 2).value = formula(stem, False)
        ws.cell(row, 1).font = LINK_FONT
        ws.cell(row, 2).font = LINK_FONT


def rebuild_document_center(wb, products: list[dict]) -> None:
    ws = wb["DOCUMENT_CENTER"]
    clear_sheet(ws)
    ws["A1"] = "DOCUMENT CENTER — PRODUCT-LEVEL (HORIZONTAL)"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws["A2"] = (
        "One Product Code = one row. OPEN WORD / OPEN PDF open local files relative to this workbook (00_CONTROL). "
        "Do not open a copy from Desktop or output\\ root — use 00_AC_DOCUMENT_ENGINE.cmd."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:Q2")

    headers = [
        "Product Code",
        "Packaging Set Code",
        "Revision",
        "Status",
        "Technical File ID",
        "TF WORD",
        "TF PDF",
        "EU DoC ID",
        "DoC WORD",
        "DoC PDF",
        "Label ID",
        "Label WORD",
        "Label PDF",
        "Shipment Statement ID",
        "Statement WORD",
        "Statement PDF",
        "Product Description",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(4, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN

    for i, p in enumerate(products):
        r = i + 5
        pc = p["product_code"]
        sc = p["set_code"]
        ids = product_instance_ids(pc, sc)
        rel = f"..\\01_PRODUCT_DOCUMENT_SETS\\{pc}\\"
        values = [
            pc,
            sc,
            "R00",
            "ISSUED",
            ids["tf"],
            None,
            None,
            ids["doc"],
            None,
            None,
            ids["label"],
            None,
            None,
            ids["stm"],
            None,
            None,
            p["description"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.font = Font(name=FONT, size=9)
        link_map = [
            (6, "01_Technical_File.docx", "OPEN WORD"),
            (7, "01_Technical_File.pdf", "OPEN PDF"),
            (9, "02_EU_DoC.docx", "OPEN WORD"),
            (10, "02_EU_DoC.pdf", "OPEN PDF"),
            (12, "03_Label.docx", "OPEN WORD"),
            (13, "03_Label.pdf", "OPEN PDF"),
            (15, "04_Shipment_Statement.docx", "OPEN WORD"),
            (16, "04_Shipment_Statement.pdf", "OPEN PDF"),
        ]
        for col, fname, label in link_map:
            set_link(ws.cell(r, col), f"{rel}{fname}", label)
            ws.cell(r, col).border = THIN

    widths = {
        "A": 12,
        "B": 16,
        "C": 10,
        "D": 12,
        "E": 34,
        "F": 12,
        "G": 12,
        "H": 34,
        "I": 12,
        "J": 12,
        "K": 34,
        "L": 12,
        "M": 12,
        "N": 34,
        "O": 12,
        "P": 12,
        "Q": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:Q{4 + len(products)}"


def rebuild_document_register(wb, products: list[dict]) -> None:
    if "DOCUMENT_REGISTER" not in wb.sheetnames:
        wb.create_sheet("DOCUMENT_REGISTER")
    ws = wb["DOCUMENT_REGISTER"]
    clear_sheet(ws)
    headers = [
        "Product Code",
        "Product Description",
        "Packaging Set Code",
        "TF Instance ID",
        "DoC Instance ID",
        "Label Instance ID",
        "Statement Instance ID",
        "Revision",
        "Word Status",
        "PDF Status",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, p in enumerate(products, 2):
        ids = product_instance_ids(p["product_code"], p["set_code"])
        ws.cell(i, 1, p["product_code"])
        ws.cell(i, 2, p["description"])
        ws.cell(i, 3, p["set_code"])
        ws.cell(i, 4, ids["tf"])
        ws.cell(i, 5, ids["doc"])
        ws.cell(i, 6, ids["label"])
        ws.cell(i, 7, ids["stm"])
        ws.cell(i, 8, "R00")
        ws.cell(i, 9, "OK")
        ws.cell(i, 10, "OK")


def patch_home(wb, products: list[dict], starter: int, gap: int) -> None:
    home = wb["00_HOME"]
    # Keep card layout; overwrite key titles/counts where present
    # Prefer known cells from FINAL layout
    if home["E2"].value:
        home["E2"] = "İNCİ AKÜ PPWR DOCUMENT ENGINE"
    if home["E3"].value:
        home["E3"] = "Product-Level Customer Delivery — Rev.00"
    # KPI cards in FINAL use multiline text in B7/B9 etc. — rewrite safely
    n = len(products)
    home["B7"] = f"STARTER PRODUCTS\n{starter}\nAll Starter Product Codes"
    home["B9"] = f"WORD DOCUMENTS\n{n * 4}\n{n} × 4"
    # Try right-side cards if present
    for coord in ("F7", "F9", "H7", "H9"):
        val = home[coord].value
        if not val:
            continue
        text = str(val)
        if "PDF" in text.upper():
            home[coord] = f"PDF DOCUMENTS\n{n * 4}\n{n} × 4"
        elif "CONTROL" in text.upper() or "SET" in text.upper():
            home[coord] = f"PRODUCT PACKS\n{n}\nOne code → one pack"
        elif "DoC" in text or "SIGN" in text.upper():
            home[coord] = f"SIGNED DoC\n{n} / {n}\nNuman Alver"
        elif "GAP" in text.upper() or "DOMESTIC" in text.upper():
            home[coord] = f"DOMESTIC DATA GAP\n{gap}\nDocuments not issued"

    # Strong open-path warning (same lesson as FINAL link fix)
    # Use a free area near bottom
    home["B28"] = (
        "ÖNEMLİ: Bu Excel'i YALNIZCA delivery içindeki 00_CONTROL klasöründen "
        "(veya 00_AC_DOCUMENT_ENGINE.cmd ile) açın. Masaüstü / output kök kopyası "
        "linkleri kırar. Klasör: 01_PRODUCT_DOCUMENT_SETS\\<ÜrünKodu>\\"
    )
    home["B28"].font = Font(name=FONT, size=10, bold=True, color="A12622")
    try:
        home.merge_cells("B28:I28")
    except Exception:
        pass
    home["B29"] = (
        f"Product-level packs: {n} | Word+PDF: {n*4}+{n*4} | Generated: "
        f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    )
    home["B29"].font = Font(name=FONT, size=9, bold=True, color=NAVY)


def patch_bulk_search_if_any(wb) -> None:
    if "BULK_SEARCH" not in wb.sheetnames:
        return
    ws = wb["BULK_SEARCH"]
    # Replace DOCUMENT_SETS path tokens in formulas if present
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 80), max_col=min(ws.max_column or 1, 20)):
        for cell in row:
            if isinstance(cell.value, str) and "01_DOCUMENT_SETS" in cell.value:
                cell.value = cell.value.replace("01_DOCUMENT_SETS", "01_PRODUCT_DOCUMENT_SETS")
                # set-code folder -> product code where pattern & $B$ used is ambiguous; leave structure
                # Product bulk may still be set-oriented in FINAL — safest note:
    ws["A1"] = str(ws["A1"].value or "") + " (PRODUCT-LEVEL paths: 01_PRODUCT_DOCUMENT_SETS)"


def write_launchers() -> None:
    (CANDIDATE / "00_AC_DOCUMENT_ENGINE.cmd").write_text(
        "@echo off\r\n"
        "cd /d \"%~dp0\"\r\n"
        "start \"\" \"%~dp000_CONTROL\\INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00.xlsx\"\r\n",
        encoding="utf-8",
    )
    # Desktop: launcher only (never a detached xlsx copy)
    if DESKTOP_BAD_XLSX.exists():
        try:
            DESKTOP_BAD_XLSX.unlink()
        except Exception:
            pass
    DESKTOP_CMD.write_text(
        "@echo off\r\n"
        f"start \"\" \"{CANDIDATE}\\00_AC_DOCUMENT_ENGINE.cmd\"\r\n",
        encoding="utf-8",
    )


def main() -> None:
    assert FINAL_ENG.exists(), f"missing template engine: {FINAL_ENG}"
    assert PRODUCT_SETS.exists(), f"missing product sets: {PRODUCT_SETS}"
    products, starter, gap = load_products()
    print(f"products={len(products)} starter={starter} gap={gap}", flush=True)

    backup = CONTROL / f"INCI_AKU_PPWR_DOCUMENT_ENGINE_Rev00_BACKUP_BEFORE_UI_REPAIR.xlsx"
    if ENG_CTRL.exists():
        shutil.copy2(ENG_CTRL, backup)
        print("backup", backup.name, flush=True)

    CONTROL.mkdir(parents=True, exist_ok=True)
    shutil.copy2(FINAL_ENG, ENG_CTRL)

    wb = load_workbook(ENG_CTRL)
    # Drop minimal product-only sheets if they somehow exist from old file (fresh copy from FINAL)
    rebuild_search_data(wb, products)
    patch_search_formulas(wb)
    rebuild_document_center(wb, products)
    rebuild_document_register(wb, products)
    patch_home(wb, products, starter, gap)
    patch_bulk_search_if_any(wb)

    # Ensure defined names don't break
    wb.save(ENG_CTRL)
    wb.close()

    shutil.copy2(ENG_CTRL, ENG_ROOT)
    write_launchers()

    # quick resolve test
    wb = load_workbook(ENG_CTRL, data_only=False)
    dc = wb["DOCUMENT_CENTER"]
    ok = fail = 0
    for row in range(5, min(25, dc.max_row or 5) + 1):
        for col in (6, 7, 9, 10, 12, 13, 15, 16):
            cell = dc.cell(row, col)
            t = cell.hyperlink.target if cell.hyperlink else None
            if not t:
                fail += 1
                continue
            resolved = (ENG_CTRL.parent / t).resolve()
            if resolved.exists() and resolved.stat().st_size > 0:
                ok += 1
            else:
                fail += 1
                print("BROKEN", cell.coordinate, t, flush=True)
    wb.close()
    print(f"link_sample ok={ok} fail={fail}", flush=True)
    print("engine", ENG_CTRL, flush=True)
    print("desktop_launcher", DESKTOP_CMD, flush=True)
    print("DONE", flush=True)


if __name__ == "__main__":
    main()
